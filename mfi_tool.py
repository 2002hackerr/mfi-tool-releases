import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, re, threading
from datetime import datetime

# --- 2. UTILITIES ---
def extract_sid(val):
    s = str(val).strip()
    parts = re.findall(r'\d{10,}', s)
    return max(parts, key=len) if parts else s

def safe_num(val):
    try:
        if pd.isna(val): return 0.0
        return float(val)
    except:
        return 0.0

def clean(val):
    if pd.isna(val): return ""
    return str(val).strip()

def split_comma(val):
    if not val or pd.isna(val): return []
    return [s.strip() for s in str(val).split(',') if s.strip()]

def fmt_qty(val):
    n = safe_num(val)
    if n == 0: return ""
    return str(int(n)) if n == int(n) else str(n)

# --- 3. DIALOG CLASSES ---

class SIDRequestDialog(tk.Toplevel):
    def __init__(self, parent, invoice, po, asin, callback):
        super().__init__(parent)
        self.callback = callback
        self.title("SID Required — DICES Validation")
        self.geometry("520x200")
        self.resizable(True, True)
        self.configure(bg="#16213e")
        self.focus_set()

        tk.Label(self, text="⚠  SID Not Found in REBNI",
                 bg="#16213e", fg="#e94560",
                 font=("Segoe UI", 12, "bold")).pack(pady=(14, 4))

        info = f"Invoice: {invoice}   PO: {po}   ASIN: {asin}"
        tk.Label(self, text=info, bg="#16213e", fg="#e0e0e0",
                 font=("Segoe UI", 9)).pack(pady=2)

        tk.Label(self, text="Validate this invoice in DICES and enter the SID below:",
                 bg="#16213e", fg="#aaaacc", font=("Segoe UI", 9)).pack(pady=6)

        ef = tk.Frame(self, bg="#16213e")
        ef.pack()
        tk.Label(ef, text="SID from DICES:", bg="#16213e", fg="#e0e0e0",
                 font=("Segoe UI", 10)).pack(side="left", padx=6)
        self._sid_var = tk.StringVar()
        self._entry = tk.Entry(ef, textvariable=self._sid_var,
                               width=28, font=("Segoe UI", 10),
                               bg="#1e1e3a", fg="#e0e0e0",
                               insertbackground="white", relief="flat")
        self._entry.pack(side="left", padx=4)
        self._entry.focus_set()

        bf = tk.Frame(self, bg="#16213e")
        bf.pack(pady=10)
        tk.Button(bf, text="✔  Continue", command=self._ok,
                  bg="#2d6a4f", fg="white", font=("Segoe UI", 10, "bold"),
                  padx=14, pady=6, relief="flat", cursor="hand2").pack(side="left", padx=6)
        tk.Button(bf, text="✖  Skip", command=self._skip,
                  bg="#6b2737", fg="white", font=("Segoe UI", 10),
                  padx=14, pady=6, relief="flat", cursor="hand2").pack(side="left", padx=6)

        self.bind('<Return>', lambda e: self._ok())
        self.protocol("WM_DELETE_WINDOW", self._skip)

        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width() - self.winfo_width()) // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _ok(self):
        sid = extract_sid(self._sid_var.get())
        if sid:
            self.callback(sid)
            self.destroy()
        else: self._entry.config(bg="#3a1e1e")

    def _skip(self):
        self.callback(None)
        self.destroy()

class ManualLevelDialog(tk.Toplevel):
    def __init__(self, parent, matches, remaining_pqv, callback):
        super().__init__(parent)
        self.callback = callback
        self.matches = matches
        self.rem_pqv = remaining_pqv
        
        self.title("Manual Investigation — Next Step")
        self.geometry("620x420")
        self.configure(bg="#0f0f1a")
        self.resizable(True, True)
        self.focus_set()
        
        tk.Label(self, text="Select Invoice to Continue", bg="#0f0f1a", fg="#4a9eff", font=("Segoe UI", 12, "bold")).pack(pady=(15, 5))
        
        self.combo = ttk.Combobox(self, width=70, state="readonly", font=("Segoe UI", 10))
        opts = [f"Qty={m['mtc_qty']}  |  Inv={m['mtc_inv']}  |  PO={m['mtc_po']}  |  ASIN={m['mtc_asin']}" for m in matches]
        self.combo['values'] = opts
        if opts: self.combo.current(0)
        self.combo.pack(pady=5)
        
        tk.Label(self, text="IBC = PBC Validation", bg="#0f0f1a", fg="#e0e0e0", font=("Segoe UI", 11, "bold")).pack(pady=(20, 10))
        
        self.val_var = tk.StringVar(value="valid")
        rf = tk.Frame(self, bg="#0f0f1a")
        rf.pack(fill="x", padx=40)
        
        tk.Radiobutton(rf, text="✔  IBC = PBC  VALID  — Continue investigation", variable=self.val_var, value="valid", 
                       bg="#0f0f1a", fg="#2d6a4f", font=("Segoe UI", 10, "bold"), selectcolor="#16213e", command=self._toggle).pack(anchor="w", pady=5)
        tk.Radiobutton(rf, text="✗  IBC ≠ PBC  INVALID — Exclude units", variable=self.val_var, value="invalid", 
                       bg="#0f0f1a", fg="#e94560", font=("Segoe UI", 10, "bold"), selectcolor="#16213e", command=self._toggle).pack(anchor="w", pady=5)
                       
        self.dyn_frame = tk.Frame(self, bg="#1a1a2e", padx=15, pady=15, highlightbackground="#4a9eff", highlightthickness=1)
        self.dyn_frame.pack(fill="x", padx=40, pady=15)
        
        self.v_frame = tk.Frame(self.dyn_frame, bg="#1a1a2e")
        tk.Label(self.v_frame, text="SID from DICES:", bg="#1a1a2e", fg="white").grid(row=0, column=0, sticky="w", pady=5)
        self.sid_ent = tk.Entry(self.v_frame, width=25)
        self.sid_ent.grid(row=0, column=1, padx=10, pady=5)
        tk.Label(self.v_frame, text="Barcode from DICES:", bg="#1a1a2e", fg="white").grid(row=1, column=0, sticky="w", pady=5)
        self.bc_ent = tk.Entry(self.v_frame, width=25)
        self.bc_ent.grid(row=1, column=1, padx=10, pady=5)
        
        self.inv_frame = tk.Frame(self.dyn_frame, bg="#1a1a2e")
        tk.Label(self.inv_frame, text="Units matched to invalid invoice:", bg="#1a1a2e", fg="white").grid(row=0, column=0, sticky="w", pady=5)
        self.qty_ent = tk.Entry(self.inv_frame, width=15)
        self.qty_ent.grid(row=0, column=1, padx=10, pady=5)
        
        self._toggle()
        
        bf = tk.Frame(self, bg="#0f0f1a")
        bf.pack(pady=20)
        tk.Button(bf, text="▶  CONTINUE", bg="#2d6a4f", fg="white", font=("Segoe UI", 10, "bold"), width=15, command=self._continue).pack(side="left", padx=10)
        tk.Button(bf, text="⬛  STOP THIS ASIN", bg="#4a2020", fg="white", font=("Segoe UI", 10, "bold"), width=18, command=self._stop).pack(side="left", padx=10)
        
        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width() - self.winfo_width()) // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")
        
    def _toggle(self):
        if self.val_var.get() == "valid":
            self.inv_frame.pack_forget()
            self.v_frame.pack(fill="x")
        else:
            self.v_frame.pack_forget()
            self.inv_frame.pack(fill="x")
            
    def _continue(self):
        idx = self.combo.current()
        if idx < 0: return
        match = self.matches[idx]
        if self.val_var.get() == "valid":
            sid = extract_sid(self.sid_ent.get())
            bc = clean(self.bc_ent.get())
            if not sid:
                messagebox.showerror("Error", "SID field cannot be empty.")
                return
            result = {'action':'valid', 'chosen_match':match, 'sid':sid, 'barcode':bc or "[DICES]"}
        else:
            qty_str = self.qty_ent.get().strip()
            try: qty = float(qty_str)
            except:
                messagebox.showerror("Error", "Units must be a valid number.")
                return
            result = {'action':'invalid', 'chosen_match':match, 'invalid_qty':qty}
        self.callback(result)
        self.destroy()
        
    def _stop(self):
        self.callback({'action':'stop'})
        self.destroy()

class CrossPODialog(tk.Toplevel):
    CASE_DESCRIPTIONS = {
        "Case 1": (
            "Case 1 — No PO but ASIN received",
            "Rec=0 at claiming PO. Same ASIN received in different PO within same SID.\n"
            "Those units are overage under a different PO."
        ),
        "Case 2": (
            "Case 2 — PO exists but ASIN not ordered there",
            "This PO exists in the claiming SID, but the ASIN was never invoiced there.\n"
            "Inv Qty = 0 but units were received. This is a Cross PO overage."
        ),
        "Case 3": (
            "Case 3 — PO and ASIN exist but Rec > Inv",
            "Both PO and ASIN present. Invoiced qty = X, but received more than X.\n"
            "Excess units are Cross PO overage."
        ),
    }

    def __init__(self, parent, candidates, callback):
        super().__init__(parent)
        self.callback = callback
        self.candidates = candidates
        self.title("Cross PO Detected — Confirm")
        self.geometry("740x540")
        self.resizable(True, True)
        self.configure(bg="#0f0f1a")
        self.grab_set()
        
        tk.Label(self, text="Select the confirmed Cross PO", bg="#0f0f1a", fg="#4a9eff", font=("Segoe UI", 12, "bold")).pack(pady=10)
        
        frame = tk.Frame(self, bg="#0f0f1a")
        frame.pack(fill="both", expand=True, padx=20)
        
        cols = ('PO', 'ASIN', 'Rec Qty', 'Type')
        self.tree = ttk.Treeview(frame, columns=cols, show='headings', height=6)
        for c in cols: self.tree.heading(c, text=c); self.tree.column(c, width=120)
        self.tree.pack(fill="x")
        
        for i, c in enumerate(candidates):
            self.tree.insert('', 'end', iid=str(i), values=(c['po'], c['asin'], fmt_qty(c['rec_qty']), c['cross_type']))
            
        self.combo = ttk.Combobox(self, state="readonly", width=50)
        self.combo['values'] = [f"{c['po']} ({c['cross_type']})" for c in candidates]
        if candidates: self.combo.current(0)
        self.combo.pack(pady=15)
        
        self.case_frame = tk.LabelFrame(self, text="Confirm Cross PO Case", bg="#0f0f1a", fg="white", padx=10, pady=5)
        self.case_frame.pack(fill="x", padx=20, pady=5)
        
        self._case_var = tk.StringVar(value="Case 1")
        for k in ["Case 1", "Case 2", "Case 3"]:
            tk.Radiobutton(self.case_frame, text=k, variable=self._case_var, value=k, bg="#0f0f1a", fg="white", selectcolor="#16213e", command=self._update_desc).pack(side="left", padx=20)
            
        self.desc_lbl = tk.Label(self, text="", bg="#0f0f1a", fg="#aaaacc", font=("Segoe UI", 9), justify="left")
        self.desc_lbl.pack(pady=5, padx=25, fill="x")
        self._update_desc()
        
        bf = tk.Frame(self, bg="#0f0f1a")
        bf.pack(pady=10)
        tk.Button(bf, text="✔  Confirm & Investigate", bg="#2d6a4f", fg="white", font=("Segoe UI", 10, "bold"), padx=15, command=self._ok).pack(side="left", padx=10)
        tk.Button(bf, text="✖  Skip", bg="#4a2020", fg="white", font=("Segoe UI", 10), padx=15, command=self._skip).pack(side="left", padx=10)
        self.protocol("WM_DELETE_WINDOW", self._skip)

    def _update_desc(self):
        k = self._case_var.get()
        title, desc = self.CASE_DESCRIPTIONS.get(k, ("",""))
        self.desc_lbl.config(text=f"{title}\n{desc}")

    def _ok(self):
        idx = self.combo.current()
        if idx >= 0:
            self.callback({
                'action': 'confirmed',
                'candidate': self.candidates[idx],
                'case': self._case_var.get()
            })
        self.destroy()

    def _skip(self):
        self.callback({'action':'skip'})
        self.destroy()

class PreviewPanel(tk.Toplevel):
    COLS = ['Barcode','Inv no','SID','PO','ASIN','Inv Qty','Rec Qty','Mtc Qty','Mtc Inv','Remarks','Date']
    COL_W_PX = [130,160,130,90,110,60,60,60,160,220,150]

    def __init__(self, parent):
        super().__init__(parent)
        self.title("Investigation Preview — Manual Mode (editable)")
        self.geometry("1200x500")
        self.configure(bg="#0f0f1a")
        self.resizable(True, True)

        tk.Label(self, text="  Live Investigation Preview — double-click any cell to edit",
                 bg="#16213e", fg="#4a9eff", font=("Segoe UI", 10, "bold"), height=2).pack(fill="x")

        frame = tk.Frame(self, bg="#0f0f1a")
        frame.pack(fill="both", expand=True, padx=8, pady=6)
        vsb = ttk.Scrollbar(frame, orient="vertical")
        hsb = ttk.Scrollbar(frame, orient="horizontal")
        self.tree = ttk.Treeview(frame, columns=self.COLS, show='headings', yscrollcommand=vsb.set, xscrollcommand=hsb.set, height=20)
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        for col, w in zip(self.COLS, self.COL_W_PX):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, minwidth=40, anchor='w')
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.grid_rowconfigure(0, weight=1); frame.grid_columnconfigure(0, weight=1)
        self.tree.bind('<Double-1>', self._on_double_click)
        self._row_data = {}

        bb = tk.Frame(self, bg="#0f0f1a")
        bb.pack(fill="x", padx=8, pady=4)
        tk.Label(bb, text="Double-click any cell to edit", bg="#0f0f1a", fg="#8888aa", font=("Segoe UI", 8)).pack(side="left")
        tk.Button(bb, text="Clear All", command=self.clear_all, bg="#2d2d5e", fg="white", font=("Segoe UI", 9), relief="flat", padx=10, pady=4).pack(side="right")

        style = ttk.Style()
        style.configure("Treeview", font=("Calibri",10), rowheight=22, background="#1e1e3a", foreground="#e0e0e0", fieldbackground="#1e1e3a")
        style.configure("Treeview.Heading", font=("Calibri",10,"bold"), background="#203864", foreground="white")
        self.tree.tag_configure('header', background='#203864', foreground='white')
        self.tree.tag_configure('depth0', background='#1e1e3a')
        self.tree.tag_configure('depth1', background='#1e3a28')
        self.tree.tag_configure('sub', background='#1a1a35')
        self.tree.tag_configure('root', background='#3a1e1e', foreground='#ff8888')
        self.tree.tag_configure('sr', background='#3a1a1a', foreground='#ff6666')
        self.tree.tag_configure('dices', background='#3a3010')
        self.tree.tag_configure('invalid', background='#3a1e1e', foreground='#ff8888')
        self.tree.tag_configure('crosspo', background='#2a1a00', foreground='#f0c060')

    def add_header_row(self, asin_label=""):
        vals = list(self.COLS)
        if asin_label: vals[4] = f"── {asin_label} ──"
        iid = self.tree.insert('', 'end', values=vals, tags=('header',))
        self._row_data[iid] = dict(zip(self.COLS, vals))

    def add_row(self, rd):
        vals = [rd.get('barcode',''), rd.get('invoice',''), rd.get('sid',''), rd.get('po',''), rd.get('asin',''), 
                rd.get('inv_qty',''), rd.get('rec_qty',''), rd.get('mtc_qty',''), rd.get('mtc_inv',''), rd.get('remarks',''), rd.get('date','')]
        remarks = rd.get('remarks', '').lower()
        tag = 'invalid' if 'invalid invoice' in remarks else \
              'crosspo' if rd.get('type') == 'crosspo' else \
              'sub' if rd.get('type') == 'subrow' else \
              'root' if 'found' in remarks and 'short' in remarks else \
              'sr' if rd.get('remarks','').startswith('SR') else \
              'dices' if '[dices]' in str(rd.get('barcode','')).lower() else \
              f"depth{min(rd.get('depth',0),1)}"
        iid = self.tree.insert('', 'end', values=vals, tags=(tag,))
        self._row_data[iid] = dict(zip(self.COLS, vals))
        self._row_data[iid]['_rd'] = rd
        self.tree.see(iid)

    def get_all_rows(self):
        KEY_MAP = {'Barcode':'barcode','Inv no':'invoice','SID':'sid','PO':'po','ASIN':'asin','Inv Qty':'inv_qty',
                   'Rec Qty':'rec_qty','Mtc Qty':'mtc_qty','Mtc Inv':'mtc_inv','Remarks':'remarks','Date':'date'}
        rows = []
        for iid in self.tree.get_children():
            data = self._row_data.get(iid, {})
            if data.get(self.COLS[0]) == self.COLS[0]: continue
            rd = data.get('_rd', {}).copy()
            for col in self.COLS: rd[KEY_MAP[col]] = data.get(col, '')
            rows.append(rd)
        return rows

    def clear_all(self):
        for iid in self.tree.get_children(): self.tree.delete(iid)
        self._row_data.clear()

    def _on_double_click(self, event):
        if self.tree.identify_region(event.x, event.y) != 'cell': return
        iid = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if not iid or not col: return
        col_idx = int(col.replace('#','')) - 1
        col_name = self.COLS[col_idx]
        x, y, w, h = self.tree.bbox(iid, col)
        current = self._row_data.get(iid, {}).get(col_name, '')
        entry_var = tk.StringVar(value=str(current))
        entry = tk.Entry(self.tree, textvariable=entry_var, font=("Calibri",10), bg="#2d2d5e", fg="white", insertbackground="white", relief="flat", bd=1)
        entry.place(x=x, y=y, width=w, height=h); entry.focus_set(); entry.select_range(0, 'end')
        def save(ev=None):
            new_val = entry_var.get(); self._row_data[iid][col_name] = new_val
            vals = list(self.tree.item(iid, 'values')); vals[col_idx] = new_val
            self.tree.item(iid, values=vals); entry.destroy()
        entry.bind('<Return>', save); entry.bind('<Tab>', save); entry.bind('<Escape>', lambda e: entry.destroy()); entry.bind('<FocusOut>', save)

# --- 4. DATA LOADERS ---
def load_claims(path): return pd.read_excel(path, header=0, dtype=str)
def detect_claim_cols(df):
    cols = {c.lower().strip(): c for c in df.columns}
    m = {}
    for k in ['barcode', 'upc', 'ean']:
        if k in cols: m['Barcode'] = cols[k]; break
    for k in ['invoice', 'inv no', 'inv_no', 'invoice number']:
        if any(x in k for x in ['invoice', 'inv']):
            if cols.get(k): m['Invoice'] = cols[k]; break
    if 'Invoice' not in m:
        for c in df.columns:
            if 'inv' in c.lower(): m['Invoice'] = c; break
    for k in ['shipment id', 'sid', 'shipment_id']:
        if k in cols: m['SID'] = cols[k]; break
    if 'SID' not in m:
        for c in df.columns:
            if 'sid' in c.lower() or 'shipment' in c.lower(): m['SID'] = c; break
    for k in ['po', 'purchase order', 'purchase_order']:
        if k in cols: m['PO'] = cols[k]; break
    for k in ['asin', 'amazon product id']:
        if k in cols: m['ASIN'] = cols[k]; break
    for k in ['invoice qty', 'inv qty', 'inv_qty', 'quantity_invoiced']:
        if k in cols: m['InvQty'] = cols[k]; break
    for k in ['pqv', 'missing qty', 'shortage', 'missing_qty', 'short']:
        if k in cols: m['PQV'] = cols[k]; break
    return m

def load_rebni(path):
    df = pd.read_excel(path, header=0, dtype=str)
    df.columns = ['vendor_code', 'po', 'asin', 'shipment_id', 'received_datetime', 'warehouse_id', 'item_cost', 'quantity_unpacked', 
                  'quantity_adjusted', 'qty_received_postadj', 'quantity_matched', 'rebni_available', 'cnt_invoice_matched', 'matched_invoice_numbers']
    return df

def load_invoice_search(path):
    df = pd.read_excel(path, header=0, dtype=str)
    df.columns = ['vendor_code', 'purchase_order_id', 'asin', 'invoice_number', 'invoice_date', 'invoice_item_status', 'quantity_invoiced', 
                  'quantity_matched_total', 'no_of_shipments', 'shipment_id', 'shipmentwise_matched_qty', 'matched_po', 'matched_asin']
    return df

# --- 5. INDEX BUILDERS ---
def build_rebni_index(df):
    p, s, fb = {}, {}, {}
    for _, row in df.iterrows():
        sid = extract_sid(row['shipment_id']); po = clean(row['po']); asin = clean(row['asin'])
        p.setdefault((sid, po, asin), []).append(row.to_dict())
        s.setdefault((po, asin), []).append(row.to_dict())
        for inv in split_comma(row['matched_invoice_numbers']):
            if inv: fb.setdefault((sid, po, inv), []).append(row.to_dict())
    return p, s, fb

def build_invoice_index(df):
    idx, fb = {}, {}
    for _, row in df.iterrows():
        sids = split_comma(row['shipment_id']); pos = split_comma(row['matched_po']); asins = split_comma(row['matched_asin']); qtys = split_comma(row['shipmentwise_matched_qty'])
        for i in range(max(len(sids), len(pos), len(asins), len(qtys))):
            s_frag = extract_sid(sids[i] if i < len(sids) else "")
            p_val = pos[i] if i < len(pos) else ""
            a_val = asins[i] if i < len(asins) else ""
            q_val = safe_num(qtys[i] if i < len(qtys) else "0")
            inv_no = clean(row['invoice_number']); mtc_po = clean(row['purchase_order_id']); mtc_asin = clean(row['asin'])
            entry = {'mtc_inv': inv_no, 'mtc_po': mtc_po, 'mtc_asin': mtc_asin, 'inv_qty': safe_num(row['quantity_invoiced']), 'mtc_qty': q_val, 'date': clean(row['invoice_date'])}
            idx.setdefault((s_frag, p_val, a_val), []).append(entry)
            if inv_no: fb.setdefault((s_frag, p_val, inv_no), []).append(entry)
    return idx, fb

# --- 6. INVESTIGATION ENGINE ---
class InvestigationEngine:
    def __init__(self, rp, rs, rfb, ip, ifb, sid_cb=None):
        self.rebni_p, self.rebni_s, self.rebni_fb = rp, rs, rfb
        self.inv_p, self.inv_fb = ip, ifb
        self.sid_cb = sid_cb
        self.max_depth = 10
        self.stop_requested = False
        self.cache_sid = {}
        self.cache_bc = {}

    def _rebni_lookup(self, sid, po, asin, inv_no=None):
        rows = self.rebni_p.get((sid, po, asin), [])
        if not rows and inv_no: rows = self.rebni_fb.get((sid, po, inv_no), [])
        return rows

    def _inv_lookup(self, sid, po, asin, inv_no=None):
        m = self.inv_p.get((sid, po, asin), [])
        if not m and inv_no: m = self.inv_fb.get((sid, po, inv_no), [])
        return m

    def detect_cross_po(self, sid, current_po, asin):
        candidates = []
        seen_po = set()
        current_rows = self.rebni_p.get((sid, current_po, asin), [])
        rec_at_current = safe_num(current_rows[0].get('quantity_unpacked', 0)) if current_rows else 0.0
        for (s, p, a), rebni_rows in self.rebni_p.items():
            if s != sid or a != asin or p == current_po or p in seen_po: continue
            for r in rebni_rows:
                rec = safe_num(r.get('quantity_unpacked', 0))
                if rec <= 0: continue
                seen_po.add(p)
                inv_matches = self.inv_p.get((sid, p, asin), [])
                inv_qty_cross = safe_num(inv_matches[0].get('inv_qty', 0)) if inv_matches else 0.0
                if rec_at_current == 0 and inv_qty_cross == 0:
                    cross_type = "Case 2 — ASIN not invoiced at this PO, but received"
                elif rec > inv_qty_cross and inv_qty_cross > 0:
                    cross_type = "Case 3 — Rec qty > Inv qty (overage in cross PO)"
                else:
                    cross_type = "Case 1 — Rec=0 at current PO, units received here"
                candidates.append({'po': p, 'asin': asin, 'sid': sid, 'inv_qty': fmt_qty(inv_qty_cross), 'rec_qty': rec, 'cross_type': cross_type, 'date': clean(r.get('received_datetime', ''))})
        return candidates

    def run_cross_po_investigation(self, candidate, case_type, budget,
                                  depth=0, visited=None):
        """
        Investigate a confirmed Cross PO overage.
        budget = candidate['rec_qty'] — the overage units to explain.

        Uses same investigation logic as run_auto but:
          - Starts from Cross PO's SID + PO + ASIN
          - Stops when total shortage found >= budget
          - budget replaces rem_pqv for this subtree

        Returns (rows, total_found).
        """
        if visited is None: visited = set()

        cross_sid  = candidate['sid']
        cross_po   = candidate['po']
        cross_asin = candidate['asin']
        cross_iqty = candidate.get('inv_qty', budget)

        # Lookup Invoice Search matched invoices for this cross PO
        raw    = self._inv_lookup(cross_sid, cross_po, cross_asin)
        seen   = set()
        unique = []
        for m in raw:
            combo = (m['mtc_inv'], m['mtc_po'], m['mtc_asin'])
            if combo not in seen:
                seen.add(combo); unique.append(m)
        unique.sort(key=lambda x: safe_num(x['mtc_qty']), reverse=True)

        # REBNI data for cross PO node
        rebni_rows  = self._rebni_lookup(cross_sid, cross_po, cross_asin)
        rec_qty     = safe_num(rebni_rows[0].get('quantity_unpacked', 0)) if rebni_rows else 0.0
        rebni_avail = safe_num(rebni_rows[0].get('rebni_available', 0))   if rebni_rows else 0.0
        rec_date    = clean(rebni_rows[0].get('received_datetime', ''))   if rebni_rows else ''
        shortage    = max(0.0, safe_num(cross_iqty) - rec_qty)

        main_mtc_inv = unique[0]['mtc_inv'] if unique else "Short Received"
        main_mtc_qty = fmt_qty(unique[0]['mtc_qty']) if unique else ""

        if not unique and shortage > 0:
            remarks = (f"Found {int(min(shortage, budget))} units short "
                       f"as loop started from {int(budget)} matched qty, no remaining pqv")
        elif rebni_avail > 0:
            remarks = (f"REBNI Available = {int(rebni_avail)} units at "
                       f"matching shipment level — Suggest TSP to utilize")
        else:
            remarks = f"Cross PO {case_type} — investigating {int(budget)} overage units"

        main_row = self._make_row(
            '[CROSS PO]', '—', cross_sid, cross_po, cross_asin,
            fmt_qty(cross_iqty), rec_qty, remarks, rec_date, depth
        )
        main_row['mtc_qty'] = main_mtc_qty
        main_row['mtc_inv'] = main_mtc_inv
        rows = [main_row]

        for m in unique[1:]:
            rows.append(self._make_row("", "", "", "", "", "", "", "", "", depth, 'subrow', m))

        # Stop if REBNI found or no matches
        if rebni_avail > 0 or not unique:
            return rows, min(shortage, budget)

        # Recurse through matched invoices using budget
        total_found = min(shortage, budget)
        current_rem = budget - total_found

        for match in unique:
            if current_rem <= 0: break
            next_inv    = match['mtc_inv']
            next_po     = match['mtc_po']
            next_asin   = match['mtc_asin']
            next_iqty   = match['inv_qty']
            next_budget = safe_num(match['mtc_qty'])

            state = (cross_sid, clean(next_inv), clean(next_po), clean(next_asin))
            if state in visited: continue
            visited = visited | {state}

            next_sid = self.cache_sid.get(next_inv)
            if not next_sid:
                next_sid = self._find_sid(next_po, next_asin, next_inv)
            if not next_sid and self.sid_cb:
                next_sid = self.sid_cb(next_inv, next_po, next_asin)
                if next_sid: self.cache_sid[next_inv] = next_sid

            if not next_sid:
                rows.append(self._make_row(
                    "[DICES]", next_inv, "[ENTER SID FROM DICES]",
                    next_po, next_asin, next_iqty, 0.0, "SID not found — validate in DICES", "", depth+1))
                continue

            child_rows, found = self.run_auto(
                self.cache_bc.get(next_inv, "[DICES]"),
                next_inv, next_sid, next_po, next_asin,
                next_iqty, current_rem, depth+1, visited,
                current_rem, False, min(next_budget, current_rem))

            rows.extend(child_rows)
            contribution = min(current_rem, found)
            total_found += contribution
            current_rem -= contribution

        return rows, total_found

    def _build_cross_po_rows(self, sid, po, asin, depth):
        candidates = self.detect_cross_po(sid, po, asin)
        rows = []
        for c in candidates:
            budget = safe_num(c['rec_qty'])
            rows.append({
                'barcode': '[CROSS PO]',
                'invoice': '—',
                'sid':     c['sid'],
                'po':      c['po'],
                'asin':    c['asin'],
                'inv_qty': c.get('inv_qty', ''),
                'rec_qty': fmt_qty(c['rec_qty']),
                'mtc_qty': '',
                'mtc_inv': '',
                'remarks': (f"Cross PO — {c['cross_type']} "
                            f"| Overage = {fmt_qty(c['rec_qty'])} units — investigating chain"),
                'date':    c['date'],
                'depth':   depth,
                'type':    'crosspo',
            })
            if budget > 0:
                case_label = c['cross_type'].split("—")[0].strip()
                child_rows, found = self.run_cross_po_investigation(
                    c, case_label, budget, depth=depth+1)
                rows.extend(child_rows)
        return rows

    def run_auto(self, barcode, inv_no, sid, po, asin, inv_qty, pqv, depth=0, visited=None, rem_pqv=None, is_claiming=True, branch_budget=None):
        if self.stop_requested: return [], 0.0
        if visited is None: visited = set()
        if rem_pqv is None: rem_pqv = safe_num(pqv)
        if branch_budget is None: branch_budget = rem_pqv
        sid_frag = extract_sid(sid)
        state = (sid_frag, clean(inv_no), clean(po), clean(asin))
        if state in visited or depth >= self.max_depth: return [], 0.0
        visited.add(state)
        rows, matches, level_shortage, rem_pqv_after_level = self._build_level_logic(barcode, inv_no, sid, po, asin, inv_qty, rem_pqv, depth, is_claiming, branch_budget)
        total_found = level_shortage; cur_rem = rem_pqv_after_level; remaining_budget = branch_budget - level_shortage
        if remaining_budget > 0:
            for match in matches:
                if self.stop_requested or remaining_budget <= 0: break
                next_inv = match['mtc_inv']
                if next_inv == clean(inv_no): continue
                next_sid = self.cache_sid.get(next_inv) or self._find_sid(match['mtc_po'], match['mtc_asin'], next_inv)
                if not next_sid and self.sid_cb: next_sid = self.sid_cb(next_inv, match['mtc_po'], match['mtc_asin'])
                if not next_sid:
                    d_row = self._make_row("[DICES]", next_inv, "[DICES]", match['mtc_po'], match['mtc_asin'], match['inv_qty'], 0.0, "SID not found in REBNI — follow manually", match['date'], depth+1)
                    d_row['mtc_qty'], d_row['mtc_inv'] = match['mtc_qty'], match['mtc_inv']; rows.append(d_row); continue
                next_budget = safe_num(match['mtc_qty'])
                c_rows, found_in_child = self.run_auto(self.cache_bc.get(next_inv, "[DICES]"), next_inv, next_sid, match['mtc_po'], match['mtc_asin'], match['inv_qty'], cur_rem, depth+1, visited, cur_rem, False, next_budget)
                contribution = min(remaining_budget, found_in_child)
                total_found += contribution; remaining_budget -= contribution; cur_rem -= contribution; rows.extend(c_rows)
        return rows, total_found

    def build_one_level(self, b, i, s, p, a, iq, rem, depth=0, is_claiming=True, branch_budget=None):
        if branch_budget is None: branch_budget = rem
        rows, matches, found, new_rem = self._build_level_logic(b, i, s, p, a, iq, rem, depth, is_claiming, branch_budget)
        return rows, matches, found, new_rem

    def _build_level_logic(self, barcode, inv_no, sid, po, asin, inv_qty, rem_pqv, depth, is_claiming, branch_budget):
        sid_frag = extract_sid(sid)
        rebni_rows = self._rebni_lookup(sid_frag, clean(po), clean(asin), clean(inv_no))
        rec_qty, rebni_avail, remarks, rec_date = 0.0, 0.0, "", ""
        if rebni_rows:
            r = rebni_rows[0]; rec_qty = safe_num(r.get('quantity_unpacked', 0)); rebni_avail = safe_num(r.get('rebni_available', 0)); rec_date = clean(r.get('received_datetime',''))
            if rebni_avail > 0:
                lvl = 'claiming shipment' if is_claiming else 'matching shipment'
                remarks = f"REBNI Available = {int(rebni_avail)} units at {lvl} level — Suggest TSP to utilize"
        else: remarks = "SID not found in REBNI — check DICES manually"
        if not rebni_rows and rec_qty == 0 and depth > 0: remarks = "SR"
        shortage = safe_num(inv_qty) - rec_qty
        contribution = min(branch_budget, max(0, shortage))
        if contribution >= branch_budget and branch_budget > 0:
            suffix = "no remaining pqv" if (rem_pqv - contribution) <= 0 else "now continue for remaining pqv"
            remarks = f"Found {fmt_qty(branch_budget)} units short as loop started from {fmt_qty(branch_budget)} matched qty, {suffix}"
        new_rem = rem_pqv - contribution
        raw = self._inv_lookup(sid_frag, clean(po), clean(asin), clean(inv_no))
        seen, unique = set(), []
        for m in raw:
            combo = (m['mtc_inv'], m['mtc_po'], m['mtc_asin'])
            if combo not in seen: seen.add(combo); unique.append(m)
        sorted_m = sorted(unique, key=lambda x: x['mtc_qty'], reverse=True)
        rows = []
        main = self._make_row(barcode, inv_no, sid, po, asin, inv_qty, rec_qty, remarks, rec_date, depth)
        if not sorted_m and shortage > 0: main['mtc_inv'] = "Short Received"
        elif sorted_m:
            if all(m['mtc_inv'] == clean(inv_no) for m in sorted_m): main['mtc_inv'] = "Self Matching"; main['remarks'] = ""
            else: main['mtc_qty'], main['mtc_inv'] = sorted_m[0]['mtc_qty'], sorted_m[0]['mtc_inv']
            rows.append(main)
            for i in range(1, len(sorted_m)): rows.append(self._make_row("", "", "", "", "", "", "", "", "", depth, 'subrow', sorted_m[i]))
        else: rows.append(main)
        # Cross PO check — runs at EVERY level automatically
        rows.extend(self._build_cross_po_rows(sid_frag, clean(po), clean(asin), depth))
        return rows, sorted_m, contribution, new_rem

    def _make_row(self, b, i, s, p, a, iq, rq, rem, d, depth, rtype='dominant', match=None):
        if rtype == 'subrow' and match:
            return {'barcode': "", 'invoice': "", 'sid': "", 'po': "", 'asin': "", 'inv_qty': "", 'rec_qty': "", 'mtc_qty': match['mtc_qty'], 'mtc_inv': match['mtc_inv'], 'remarks': "", 'date': "", 'depth': depth, 'type': 'subrow'}
        return {'barcode': b, 'invoice': i, 'sid': extract_sid(s), 'po': p, 'asin': a, 'inv_qty': iq, 'rec_qty': rq, 'mtc_qty': "", 'mtc_inv': "", 'remarks': rem, 'date': d, 'depth': depth, 'type': 'root' if depth==0 else 'dominant'}

    def _find_sid(self, po, asin, inv_no):
        rows = self.rebni_s.get((clean(po), clean(asin)), [])
        for r in rows:
            if inv_no in split_comma(r.get('matched_invoice_numbers', '')): return r['shipment_id']
        return rows[0]['shipment_id'] if rows else None

# --- 7. EXCEL WRITER ---
def write_excel(all_blocks, path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Investigation"
    headers = ["Barcode", "Inv no", "SID", "PO", "ASIN", "Inv Qty", "Rec Qty", "Mtc Qty", "Mtc Inv", "Remarks", "Date"]
    h_fill = PatternFill("solid", fgColor="203864"); h_font = Font(color="FFFFFF", bold=True)
    dom_fill, sub_fill = PatternFill("solid", fgColor="E2EFDA"), PatternFill("solid", fgColor="EBF3FB")
    root_fill = PatternFill("solid", fgColor="FFE0E0"); root_font = Font(color="9C0006", bold=True)
    dices_fill = PatternFill("solid", fgColor="FFF2CC"); sr_fill = PatternFill("solid", fgColor="FFD7D7"); sr_font = Font(color="CC0000", bold=True)
    INVLD_FILL, REBNI_FILL = PatternFill("solid", fgColor="FFD0D0"), PatternFill("solid", fgColor="D0F0FF")
    CROSS_F = PatternFill("solid", fgColor="FFF0C0"); CROSS_FT = Font(bold=True, color="7a5c00", name="Calibri", size=10)
    INVLD_FONT = Font(bold=True, color="880000", name="Calibri", size=10, italic=True); REBNI_FONT = Font(bold=True, color="005580", name="Calibri", size=10)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    KM = {'Barcode':'barcode','Inv no':'invoice','SID':'sid','PO':'po','ASIN':'asin','Inv Qty':'inv_qty','Rec Qty':'rec_qty','Mtc Qty':'mtc_qty','Mtc Inv':'mtc_inv','Remarks':'remarks','Date':'date'}
    curr = 1
    for block in all_blocks:
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=curr, column=c, value=h); cell.fill, cell.font, cell.border = h_fill, h_font, border
        curr += 1
        for rd in block:
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=curr, column=c, value=rd.get(KM[h], "")); cell.border = border
                rem = str(rd.get('remarks', ''))
                if 'Cross PO' in rem: cell.fill, cell.font = CROSS_F, CROSS_FT
                elif 'invalid invoice' in rem.lower(): cell.fill, cell.font = INVLD_FILL, INVLD_FONT
                elif 'REBNI Available' in rem: cell.fill, cell.font = REBNI_FILL, REBNI_FONT
                elif rd['type'] == 'subrow': cell.fill = sub_fill
                elif rd['type'] == 'dominant' and rd['depth'] > 0: cell.fill = dom_fill
                if "found" in rem.lower() and "short" in rem.lower(): cell.fill, cell.font = root_fill, root_font
                if "SR" == rem: cell.fill, cell.font = sr_fill, sr_font
                if "[DICES]" in str(rd.get('barcode', '')):
                    if c <= 3: cell.fill = dices_fill
            curr += 1
        curr += 1
    w = [18, 22, 18, 12, 14, 9, 9, 9, 26, 36, 22]
    for i, val in enumerate(w, 1): ws.column_dimensions[get_column_letter(i)].width = val
    wb.save(path)

# --- 8. GUI ---
class MFIToolApp:
    def __init__(self):
        self.root = tk.Tk(); self.root.title("MFI Investigation Tool  v4.9  |  ROW IB")
        try: self.root.state('zoomed')
        except: self.root.attributes('-zoomed', True)
        self.root.minsize(900, 620); self.root.configure(bg="#0f0f1a")
        self.claims_path, self.rebni_path, self.inv_path, self.ticket_id, self.mode_var = tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar(value="auto")
        self.all_blocks, self.preview = [], None; self._build_ui()

    def _build_ui(self):
        t = tk.Frame(self.root, bg="#16213e", height=60); t.pack(fill="x")
        tk.Label(t, text="MFI Investigation Tool", fg="#e94560", bg="#16213e", font=("Segoe UI", 20, "bold")).pack(side="left", padx=15, pady=10)
        tk.Label(t, text="v4.9  |  ROW IB", fg="#8888aa", bg="#16213e", font=("Segoe UI", 10)).pack(side="right", padx=15)
        tk.Label(t, text="Developed by Mukesh", fg="#4a9eff", bg="#16213e", font=("Segoe UI", 10, "italic")).pack(side="right", padx=5)
        leg = tk.Frame(self.root, bg="#1a1a2e", height=28); leg.pack(fill="x")
        l = [("Claiming", "white", "#0f0f1a"), ("Dominant", "black", "#E2EFDA"), ("Sub-rows", "black", "#EBF3FB"), ("Root shortage", "#9C0006", "#FFE0E0"), ("DICES", "black", "#FFF2CC"), ("SR", "#CC0000", "#FFD7D7"), ("Cross PO", "#7a5c00", "#FFF0C0")]
        for txt, fg, bg in l: tk.Label(leg, text=txt, fg=fg, bg=bg, font=("Segoe UI", 8, "bold"), padx=10).pack(side="left", padx=5, pady=2)
        body = tk.Frame(self.root, bg="#0f0f1a", padx=20, pady=10); body.pack(fill="both", expand=True)
        inp = tk.LabelFrame(body, text="Input Files", fg="white", bg="#0f0f1a", padx=10, pady=5); inp.pack(fill="x", pady=5)
        self._f_row(inp, "Claims Sheet:", self.claims_path, 0); self._f_row(inp, "REBNI Result:", self.rebni_path, 1); self._f_row(inp, "Invoice Search:", self.inv_path, 2)
        tk.Label(body, text="Ticket ID:", fg="white", bg="#0f0f1a").pack(anchor="w", pady=(10,0)); tk.Entry(body, textvariable=self.ticket_id, width=30).pack(anchor="w")
        m = tk.LabelFrame(body, text="Investigation Mode", fg="white", bg="#0f0f1a", padx=10, pady=5); m.pack(fill="x", pady=10)
        tk.Radiobutton(m, text="AUTO  —  Tool investigates automatically. Prompts for SID when not found in REBNI.", variable=self.mode_var, value="auto", fg="white", bg="#0f0f1a", selectcolor="#16213e").pack(anchor="w", padx=10)
        tk.Radiobutton(m, text="MANUAL  —  One level at a time. Validation dialog pops up for each match.", variable=self.mode_var, value="manual", fg="white", bg="#0f0f1a", selectcolor="#16213e").pack(anchor="w", padx=10)
        self.status = tk.Label(body, text="Ready", fg="#4a9eff", bg="#0f0f1a"); self.status.pack(pady=(10,0))
        self.pb = ttk.Progressbar(body, length=740, mode='determinate'); self.pb.pack(pady=5)
        bf = tk.Frame(body, bg="#0f0f1a"); bf.pack(pady=10)
        self.run_btn = tk.Button(bf, text="▶ RUN INVESTIGATION", bg="#e94560", fg="white", font=("Segoe UI", 15, "bold"), padx=36, pady=14, command=self.start_run); self.run_btn.pack(side="left", padx=10)
        self.stop_btn = tk.Button(bf, text="🛑 STOP", bg="#4a2020", fg="white", font=("Segoe UI", 14, "bold"), padx=20, pady=14, state="disabled", command=self.request_stop); self.stop_btn.pack(side="left", padx=10)
        self.save_btn = tk.Button(bf, text="💾  SAVE", bg="#2d6a4f", fg="white", font=("Segoe UI", 13, "bold"), padx=28, pady=14, state="disabled", command=self.save_output); self.save_btn.pack(side="left", padx=10)

    def _f_row(self, p, l, v, r):
        tk.Label(p, text=l, fg="white", bg="#0f0f1a").grid(row=r, column=0, sticky="w", pady=2)
        tk.Entry(p, textvariable=v, width=65).grid(row=r, column=1, padx=5)
        tk.Button(p, text="Browse", command=lambda: v.set(filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")]))).grid(row=r, column=2)

    def _set_status(self, msg, pct=None):
        self.status.config(text=msg)
        if pct is not None: self.pb['value'] = pct
        self.root.update_idletasks()

    def start_run(self):
        if not all([self.claims_path.get(), self.rebni_path.get(), self.inv_path.get()]):
            messagebox.showerror("Error", "Please select all 3 input files."); return
        self.run_btn.config(state="disabled"); self.save_btn.config(state="disabled"); self.stop_btn.config(state="normal")
        self.all_blocks = []
        if self.mode_var.get() == "manual":
            if not self.preview or not self.preview.winfo_exists(): self.preview = PreviewPanel(self.root)
            else: self.preview.clear_all()
        threading.Thread(target=self._process, daemon=True).start()

    def request_stop(self):
        if hasattr(self, 'engine'): self.engine.stop_requested = True
        self._set_status("STOP REQUESTED — Finishing current step...")

    def _process(self):
        try:
            self._set_status("Loading Data...", 5)
            df_c = load_claims(self.claims_path.get()); map_cols = detect_claim_cols(df_c)
            df_r = load_rebni(self.rebni_path.get()); rp, rs, rfb = build_rebni_index(df_r)
            df_i = load_invoice_search(self.inv_path.get()); ip, ifb = build_invoice_index(df_i)
            self.engine = InvestigationEngine(rp, rs, rfb, ip, ifb, self._req_sid)
            tot = len(df_c)
            if self.mode_var.get() == "auto":
                for i, (_, r) in enumerate(df_c.iterrows()):
                    if self.engine.stop_requested: break
                    self._set_status(f"Processing {i+1} of {tot}...", 10 + (i/tot * 85))
                    b, _ = self.engine.run_auto(clean(r[map_cols['Barcode']]), clean(r[map_cols['Invoice']]), clean(r[map_cols['SID']]), clean(r[map_cols['PO']]), clean(r[map_cols['ASIN']]), safe_num(r[map_cols['InvQty']]), safe_num(r[map_cols['PQV']]), is_claiming=True)
                    self.all_blocks.append(b)
                self._finish()
            else:
                self.manual_q = df_c.to_dict('records'); self.map_cols = map_cols; self._next_man()
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Error", str(e))); self._finish()

    def _req_sid(self, inv, po, asin):
        if inv in self.engine.cache_sid: return self.engine.cache_sid[inv]
        res, done = [None], threading.Event()
        def on_submit(sid): 
            if sid: self.engine.cache_sid[inv] = sid
            res[0] = sid; done.set()
        def show(): SIDRequestDialog(self.root, inv, po, asin, on_submit)
        self.root.after(0, show); done.wait(); return res[0]

    def _next_man(self):
        if not self.manual_q or self.engine.stop_requested: self._finish(); return
        r = self.manual_q.pop(0)
        self.curr_m = {'b':clean(r[self.map_cols['Barcode']]), 'i':clean(r[self.map_cols['Invoice']]), 's':clean(r[self.map_cols['SID']]), 'p':clean(r[self.map_cols['PO']]), 'a':clean(r[self.map_cols['ASIN']]), 'iq':safe_num(r[self.map_cols['InvQty']]), 'pqv':safe_num(r[self.map_cols['PQV']]), 'rem':safe_num(r[self.map_cols['PQV']]), 'depth':0, 'block':[], 'processed': set(), 'rendered': False}
        self.preview.add_header_row(self.curr_m['a']); self._man_step()

    def _man_step(self):
        if self.engine.stop_requested: self._finish(); return
        m = self.curr_m; is_clm = (m['depth'] == 0)
        rows, matches, found, n_rem = self.engine.build_one_level(m['b'], m['i'], m['s'], m['p'], m['a'], m['iq'], m['rem'], m['depth'], is_clm)
        if not m['rendered']:
            m['block'].extend(rows); [self.preview.add_row(r) for r in rows]; m['rendered'] = True
        m['rem'] = n_rem
        
        # Cross PO check for manual mode
        cross_rows = [r for r in rows if r.get('type') == 'crosspo']
        if cross_rows:
            cands = self.engine.detect_cross_po(m['s'], m['p'], m['a'])
            if cands:
                self.root.after(0, lambda c=cands: CrossPODialog(self.root, c, self._handle_cross_po))

        matches = [x for x in matches if x['mtc_inv'] not in m['processed']]
        if not matches or any(x in rows[0].get('remarks', '') for x in ["no remaining pqv", "Self", "REBNI", "SR"]):
            if not cross_rows: # Only move to next if we aren't waiting for a cross PO confirmation
                self.all_blocks.append(m['block']); self._next_man()
            return
        self.root.after(0, lambda: self._show_dlg(matches))

    def _handle_cross_po(self, res):
        if res['action'] == 'skip':
            return
        cand      = res['candidate']
        case_type = res.get('case', 'Case 1')
        budget    = safe_num(cand['rec_qty'])

        def investigate():
            cross_rows, found = self.engine.run_cross_po_investigation(
                cand, case_type, budget,
                depth=self.curr_m['depth'] + 1)

            for r in cross_rows:
                self.curr_m['block'].append(r)
                if self.preview and self.preview.winfo_exists():
                    self.root.after(0, lambda row=r: self.preview.add_row(row))

            self.curr_m['rem'] = max(0.0, self.curr_m['rem'] - found)

            if self.curr_m['rem'] <= 0:
                self.root.after(100, lambda: (
                    self.all_blocks.append(self.curr_m['block']),
                    self._next_man()
                ))
            else:
                self.root.after(100, self._man_step)

        threading.Thread(target=investigate, daemon=True).start()

    def _show_dlg(self, matches):
        first_match = matches[0]; inv = first_match['mtc_inv']
        if inv in self.engine.cache_sid:
            res = {'action': 'valid', 'chosen_match': first_match, 'sid': self.engine.cache_sid[inv], 'barcode': self.engine.cache_bc.get(inv, "[DICES]")}; self._handle_res(res, matches); return
        def on_res(res): self._handle_res(res, matches)
        ManualLevelDialog(self.root, matches, self.curr_m['rem'], on_res)

    def _handle_res(self, res, matches):
        if not res or res['action'] == 'stop': self.all_blocks.append(self.curr_m['block']); self._next_man(); return
        match = res['chosen_match']; self.curr_m['processed'].add(match['mtc_inv'])
        if res['action'] == 'invalid':
            excl = res['invalid_qty']; row = {'barcode': '[INVALID]', 'invoice': match['mtc_inv'], 'sid': '—', 'po': match['mtc_po'], 'asin': match['mtc_asin'], 'inv_qty': match['inv_qty'], 'rec_qty': '', 'mtc_qty': '', 'mtc_inv': '', 'remarks': f"{int(excl)} units matched to invalid invoice {match['mtc_inv']} — excluded from PQV", 'date': match['date'], 'depth': self.curr_m['depth'], 'type': 'subrow'}
            self.curr_m['block'].append(row); self.preview.add_row(row); self.curr_m['rem'] = max(0, self.curr_m['rem'] - excl)
            if self.curr_m['rem'] <= 0: self.all_blocks.append(self.curr_m['block']); self._next_man()
            else:
                rem_m = [x for x in matches if x['mtc_inv'] != match['mtc_inv']]
                if rem_m: self.root.after(0, lambda: self._show_dlg(rem_m))
                else: self.all_blocks.append(self.curr_m['block']); self._next_man()
        else:
            self.engine.cache_sid[match['mtc_inv']] = res['sid']; self.engine.cache_bc[match['mtc_inv']] = res['barcode']
            self.curr_m.update({'b': res['barcode'], 'i': match['mtc_inv'], 's': res['sid'], 'p': match['mtc_po'], 'a': match['mtc_asin'], 'iq': match['inv_qty'], 'depth': self.curr_m['depth']+1, 'rendered': False, 'processed': set()})
            threading.Thread(target=self._man_step, daemon=True).start()

    def _finish(self):
        self._set_status("Complete. Click SAVE.", 100); self.root.after(0, lambda: self.save_btn.config(state="normal")); self.root.after(0, lambda: self.run_btn.config(state="normal")); self.root.after(0, lambda: self.stop_btn.config(state="disabled"))
        self.root.after(0, lambda: messagebox.showinfo("Done", "Finished!"))

    def save_output(self):
        t = self.ticket_id.get().strip().replace(' ', '_'); ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        p = os.path.join(os.path.dirname(self.claims_path.get()), f"MFI_{t}_{ts}.xlsx" if t else f"MFI_Investigation_{ts}.xlsx")
        try:
            b = self.preview.get_all_rows() if self.mode_var.get() == "manual" else self.all_blocks
            if self.mode_var.get() == "manual":
                 fb, cur = [], []
                 for r in b:
                     if r.get('depth') == 0 and cur: fb.append(cur); cur = []
                     cur.append(r)
                 if cur: fb.append(cur)
                 b = fb
            write_excel(b, p); messagebox.showinfo("Saved", f"Saved to:\n{p}")
        except Exception as e: messagebox.showerror("Error", str(e))

    def run(self): self.root.mainloop()

if __name__ == '__main__': MFIToolApp().run()
