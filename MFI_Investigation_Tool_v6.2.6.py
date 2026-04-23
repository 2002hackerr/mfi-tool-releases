"""
MFI Investigation Tool  v6.2.6  |  ROW IB (Secured Edition)
==========================================
ROW IB  |  Amazon
Developed by Mukesh

CHANGES IN v6.2.6:
- [TBD] Improving logic based on v6.2.5 feedback

CHANGES IN v6.2.5:
- [ISOLATION] Per-ASIN "Pending Invoices" tracking (Fixes global data leakage)
- [CROSS PO] Multi-Invoice detection for the same PO (Checks all invoice combinations)
- [AUTO] Global SID/Barcode persistence (Never asks for the same info twice)
- [AUTO] Loop Result Block auto-duplication (Copies previous results for repeated loops)
- [UI] New "RECEIVED QTY" button in Preview (Main SID/PO/ASIN status lookup)
- [UI] New "FAST FETCH" button in Search (High-speed vendor-only bulk filtering)
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext, simpledialog
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import os
import re
import sys
import threading
import subprocess
import webbrowser
import requests
from datetime import datetime

# ==========================================
# SECURED PRODUCTION VERSION
# ==========================================
ACTIVATION_URL = "https://gist.githubusercontent.com/2002hackerr/3f76afc8a819c6879e06676a36173999/raw/activation.txt"

def check_activation():
    import urllib.request
    try:
        req = urllib.request.Request(ACTIVATION_URL, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as response:
            content = response.read().decode('utf-8').strip().upper()
            return content == "ENABLED"
    except Exception as e:
        return False

#  OTA UPDATE ENGINE (MIGRATED TO GITHUB RELEASES API)
# ═══════════════════════════════════════════════════════════
APP_VERSION = "6.2.6"
# Official GitHub Releases API URL for the public repository
UPDATE_MANIFEST_URL = "https://api.github.com/repos/2002hackerr/mfi-tool-releases/releases/latest"

class UpdateDialog(tk.Toplevel):
    def __init__(self, parent, remote_ver, download_url):
        super().__init__(parent)
        self.download_url = download_url
        self.title("Software Update Available")
        self.geometry("480x280")
        self.configure(bg="#0f0f1a")
        self.resizable(False, False)
        self.lift(); self.focus_force()
        self.grab_set()

        tk.Label(self, text="🚀  New Update Available", bg="#16213e", fg="#4a9eff", 
                 font=("Segoe UI", 14, "bold"), height=2).pack(fill="x")
        
        info_frame = tk.Frame(self, bg="#0f0f1a", padx=20, pady=20)
        info_frame.pack(fill="both", expand=True)
        
        tk.Label(info_frame, text=f"A new version (v{remote_ver}) of the MFI Tool is ready.", 
                 bg="#0f0f1a", fg="white", font=("Segoe UI", 11)).pack(anchor="w")
        tk.Label(info_frame, text=f"Current version: v{APP_VERSION}", 
                 bg="#0f0f1a", fg="#888888", font=("Segoe UI", 10)).pack(anchor="w", pady=(5, 15))
        
        tk.Label(info_frame, text="Would you like to download and install it now?\nThe tool will restart automatically.", 
                 bg="#0f0f1a", fg="#cccccc", font=("Segoe UI", 10), justify="left").pack(anchor="w")

        btn_f = tk.Frame(self, bg="#0f0f1a", pady=15)
        btn_f.pack(fill="x")
        
        tk.Button(btn_f, text="Update Now", bg="#2d6a4f", fg="white", font=("Segoe UI", 10, "bold"),
                  padx=20, pady=8, relief="flat", cursor="hand2", command=self.on_update).pack(side="right", padx=20)
        tk.Button(btn_f, text="Later", bg="#333333", fg="#aaaaaa", font=("Segoe UI", 10),
                  padx=15, pady=8, relief="flat", cursor="hand2", command=self.destroy).pack(side="right")

    def on_update(self):
        self.destroy()
        DownloadProgressDialog(self.master, self.download_url)

class DownloadProgressDialog(tk.Toplevel):
    def __init__(self, parent, url):
        super().__init__(parent)
        self.url = url
        self.title("Downloading Update...")
        self.geometry("400x150")
        self.configure(bg="#0f0f1a")
        self.resizable(False, False)
        self.lift(); self.focus_force()
        self.grab_set()

        tk.Label(self, text="Downloading MFI Tool Update...", bg="#0f0f1a", fg="white", 
                 font=("Segoe UI", 11)).pack(pady=(20, 10))
        
        self.pb = ttk.Progressbar(self, length=300, mode='determinate')
        self.pb.pack(pady=10)
        
        self.status_lbl = tk.Label(self, text="Connecting...", bg="#0f0f1a", fg="#4a9eff", font=("Segoe UI", 9))
        self.status_lbl.pack()
        
        threading.Thread(target=self.start_download, daemon=True).start()

    def start_download(self):
        try:
            response = requests.get(self.url, stream=True, timeout=30)
            response.raise_for_status()  # v6.2.5 Security: Verify download success before writing
            total_size = int(response.headers.get('content-length', 0))
            
            # Use current executable path to determine update filename
            current_exe = sys.executable
            update_exe = os.path.join(os.path.dirname(current_exe), "MFI_Update_Temp.exe")
            
            downloaded = 0
            with open(update_exe, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
                        downloaded += len(chunk)
                        if total_size > 0:
                            pct = (downloaded / total_size) * 100
                            self.after(0, lambda p=pct, d=downloaded, t=total_size: self.update_status(p, d, t))
            
            self.after(0, self.finalize_update)
        except Exception as e:
            self.after(0, lambda ex=e: messagebox.showerror("Update Error", f"Failed to download update: {ex}"))
            self.after(0, self.destroy)

    def update_status(self, pct, done, total):
        self.pb['value'] = pct
        self.status_lbl.config(text=f"Downloaded {done//1024} KB / {total//1024} KB ({int(pct)}%)")

    def finalize_update(self):
        current_exe = sys.executable
        # Generate batch script for self-replacement
        # Since we use --onefile, the running EXE is what we replace.
        bat_content = f"""@echo off
timeout /t 3 /nobreak > nul
move /y "MFI_Update_Temp.exe" "{os.path.basename(current_exe)}"
start "" "{os.path.basename(current_exe)}"
del "%~f0"
"""
        bat_path = os.path.join(os.path.dirname(current_exe), "mfi_updater.bat")
        with open(bat_path, "w") as f:
            f.write(bat_content)
        
        messagebox.showinfo("Update Ready", "The update has been downloaded. The tool will now restart to apply changes.")
        subprocess.Popen([bat_path], shell=True, cwd=os.path.dirname(current_exe))
        os._exit(0)

def check_for_updates(parent):
    def _run():
        try:
            # v6.2.5: Migrated to GitHub Releases API (Robust for office networks)
            headers = {'Accept': 'application/vnd.github.v3+json'}
            resp = requests.get(UPDATE_MANIFEST_URL, headers=headers, timeout=10)
            if resp.status_code == 200:
                data = resp.json()
                # GitHub API uses "tag_name" (e.g., "v6.2.5") and "assets"
                tag_name = data.get("tag_name", "")
                
                # Strip leading 'v' if present for comparison
                remote_ver = tag_name.lstrip('v')
                
                # Find the download URL for the .exe asset
                download_url = ""
                assets = data.get("assets", [])
                for asset in assets:
                    if asset.get("name", "").endswith(".exe"):
                        download_url = asset.get("browser_download_url", "")
                        break
                
                if remote_ver and remote_ver != APP_VERSION and download_url:
                    # Show dialog on main thread
                    parent.after(0, lambda: UpdateDialog(parent, remote_ver, download_url))
        except:
            pass # Silent failure to prevent startup issues
    
    threading.Thread(target=_run, daemon=True).start()


# ═══════════════════════════════════════════════════════════
#  GLOBAL THEME ENGINE
# ═══════════════════════════════════════════════════════════

GLOBAL_THEME_NAME = "Dark Mode (Default)"

THEME_PALETTES = {
    "Dark Mode (Default)": {}, 
    "Light Mode": {
        "#0f0f1a": "#f0f2f5", "#16213e": "#dfe3ee", "#0d0d1a": "#ffffff",
        "#1e1e3a": "#ffffff", "#131320": "#ffffff", "#1a1a2e": "#e9ebee",
        "#2d2d5e": "#8b9dc3", "white": "black", "#ffffff": "black", "#cccccc": "#333333",
        "#e0e0e0": "#222222", "#4a9eff": "#0056b3", "#e94560": "#d9534f",
        "#0d1117": "#f6f8fa", "#161b22": "#ffffff", "#21262d": "#ffffff",
        "#3fb950": "#2ea043", "#f85149": "#cf222e"
    },
    "Ocean Blue": {
        "#0f0f1a": "#0a192f", "#16213e": "#112240", "#0d0d1a": "#020c1b",
        "#1e1e3a": "#233554", "#131320": "#020c1b", "#1a1a2e": "#112240",
        "#2d2d5e": "#1d4ed8", "white": "#e6f1ff", "#ffffff": "#e6f1ff", 
        "#cccccc": "#8892b0", "#e0e0e0": "#a8b2d1", "#4a9eff": "#64ffda", 
        "#e94560": "#00b4d8", "#0d1117": "#0a192f", "#161b22": "#112240", 
        "#21262d": "#233554", "#f0a500": "#48cae4", "#3fb950": "#64ffda"
    },
    "Forest Green": {
        "#0f0f1a": "#0f1a14", "#16213e": "#162b1e", "#0d0d1a": "#050f09",
        "#1e1e3a": "#1d3a2a", "#131320": "#050f09", "#1a1a2e": "#162b1e",
        "#2d2d5e": "#15803d", "white": "#e6ffe6", "#ffffff": "#e6ffe6", 
        "#cccccc": "#8fbc8f", "#e0e0e0": "#aaddaa", "#4a9eff": "#90ee90", 
        "#e94560": "#3fb950", "#0d1117": "#0f1a14", "#161b22": "#162b1e", 
        "#21262d": "#1d3a2a", "#f0a500": "#90ee90", "#3fb950": "#2ea043", "#f85149": "#cf222e"
    },
    "Sunset Orange": {
        "#0f0f1a": "#2a1610", "#16213e": "#3e1a10", "#0d0d1a": "#1a0d08",
        "#1e1e3a": "#4a2414", "#131320": "#1a0d08", "#1a1a2e": "#3e1a10",
        "#2d2d5e": "#c2410c", "white": "#fff0e6", "#ffffff": "#fff0e6", 
        "#cccccc": "#cfafa1", "#e0e0e0": "#dfcfc1", "#4a9eff": "#ffc107", 
        "#e94560": "#ff4560", "#0d1117": "#2a1610", "#161b22": "#3e1a10", 
        "#21262d": "#4a2414", "#f0a500": "#ff9e4a", "#3fb950": "#85c88a"
    },
    "Purple Midnight": {
        "#0f0f1a": "#1a0f1f", "#16213e": "#2d163e", "#0d0d1a": "#13051a",
        "#1e1e3a": "#3c1e54", "#131320": "#13051a", "#1a1a2e": "#2d163e",
        "#2d2d5e": "#7e22ce", "white": "#f8e6ff", "#ffffff": "#f8e6ff", 
        "#cccccc": "#cfa1cf", "#e0e0e0": "#dfc1df", "#4a9eff": "#d884ff", 
        "#e94560": "#b336ff", "#0d1117": "#1a0f1f", "#161b22": "#2d163e", 
        "#21262d": "#3c1e54", "#f0a500": "#e0aaff", "#3fb950": "#85c88a"
    }
}

def apply_global_theme_to_widget(w):
    active_map = THEME_PALETTES.get(GLOBAL_THEME_NAME, {})
    def _apply(w_node):
        if not hasattr(w_node, '_orig_colors'):
            w_node._orig_colors = {}
            for opt in ["bg", "fg", "insertbackground", "selectcolor"]:
                try: w_node._orig_colors[opt] = str(w_node.cget(opt))
                except: pass
        for opt, orig_val in w_node._orig_colors.items():
            key = str(orig_val)
            if key in active_map:
                try: w_node.configure(**{opt: active_map[key]}); continue
                except: pass
            try: w_node.configure(**{opt: key})
            except: pass
        for child in w_node.winfo_children(): _apply(child)
    _apply(w)

_orig_toplevel_init = tk.Toplevel.__init__
def _custom_toplevel_init(self, *args, **kw):
    _orig_toplevel_init(self, *args, **kw)
    self.after(20, lambda: apply_global_theme_to_widget(self))
    
    # v6.2.5: Append a floating Pin/Lock toggle to all dialog panels (Globally persistent state)
    def _create_pin_button():
        if not self.winfo_exists(): return
        
        # Read global lock state from the main UI root
        app_root = self.nametowidget(".")
        is_pinned = getattr(app_root, "_global_dialog_pinned", False)
        
        # Inherit the state to ensure new popups don't drop the lock
        self.attributes("-topmost", is_pinned)
        
        pin_btn = tk.Button(self, text="🔒" if is_pinned else "🔓", 
                            font=("Segoe UI", 9), 
                            bg="#0a0a1a" if is_pinned else "#1e1e3a", 
                            fg="#4a9eff" if is_pinned else "#ffffff", 
                            relief="ridge", cursor="hand2", activebackground="#2d2d5e",
                            bd=1, padx=4, pady=2)
                            
        def toggle():
            if not self.winfo_exists(): return
            
            # Invert the global state so FUTURE popups inherit the user's choice
            current_state = getattr(app_root, "_global_dialog_pinned", False)
            new_state = not current_state
            app_root._global_dialog_pinned = new_state
            
            # Apply state visually to the CURRENT window
            self.attributes("-topmost", new_state)
            pin_btn.config(text="🔒" if new_state else "🔓", 
                           fg="#4a9eff" if new_state else "#ffffff",
                           bg="#0a0a1a" if new_state else "#1e1e3a")
                           
        pin_btn.config(command=toggle)
        # Float securely in the top right, clearing vertical scrollbars and borders
        pin_btn.place(relx=1.0, rely=0.0, anchor="ne", x=-25, y=8)
        pin_btn.lift()
        
    self.after(100, _create_pin_button)
    
tk.Toplevel.__init__ = _custom_toplevel_init



# ═══════════════════════════════════════════════════════════
#  UTILITIES
# ═══════════════════════════════════════════════════════════

def extract_sid(val):
    s = str(val).strip()
    parts = re.findall(r'\d{10,}', s)
    return max(parts, key=len) if parts else s

def strip_scr(inv_no):
    """Remove trailing SCR suffix(es) from an invoice number."""
    return re.sub(r'(?:SCR)+$', '', str(inv_no).strip(), flags=re.IGNORECASE)

def safe_num(val):
    try:
        if pd.isna(val): return 0.0
    except: pass
    try:
        v_str = str(val).replace(',', '').strip()
        if not v_str: return 0.0
        return float(v_str)
    except: return 0.0

def clean(val):
    try:
        if pd.isna(val): return ""
    except: pass
    s = str(val).strip()
    if s.endswith('.0'): s = s[:-2]
    return s

def split_comma(val):
    if not val: return []
    try:
        if pd.isna(val): return []
    except: pass
    return [s.strip() for s in str(val).split(',') if s.strip()]

def fmt_qty(val):
    n = safe_num(val)
    if n == 0: return ""
    return str(int(n)) if n == int(n) else str(n)

# ── FIX 2 helper ──────────────────────────────────────────────────────────────
# Keywords whose presence in a remark means it must NEVER be overwritten.
_REMARK_PRESERVE = (
    'Phase 1', 'Direct Shortage', 'REBNI Available', 'Shipment-level REBNI',
    'SR', 'Phase 4', 'No Invoice Search', 'short received directly',
)

def _remark_overwritable(rem):
    """Return True only when none of the preserved keywords appear in rem."""
    return not any(kw in rem for kw in _REMARK_PRESERVE)


# ═══════════════════════════════════════════════════════════
#  DIALOGS
# ═══════════════════════════════════════════════════════════

class HeaderCorrectionDialog(tk.Toplevel):
    """Shown when claims file has non-standard column headers."""
    def __init__(self, parent, corrections, mapping, df_columns, callback):
        super().__init__(parent)
        self.callback    = callback
        self.corrections = corrections
        self.mapping     = mapping
        self.df_columns  = df_columns

        # v6.2.5 Fix: Updated title version
        self.title("Column Header Mismatch Detected — v6.2.5")
        self.geometry("700x480")
        self.configure(bg="#0f0f1a")
        self.resizable(True, True)
        self.lift(); self.focus_force()

        tk.Label(self,
                 text="⚠  Non-standard column headers detected in Claims file",
                 bg="#16213e", fg="#f0a500",
                 font=("Segoe UI", 12, "bold"), height=2).pack(fill="x")
        tk.Label(self,
                 text="The tool has automatically matched the columns below. "
                      "Please confirm or correct before proceeding.",
                 bg="#0f0f1a", fg="#cccccc",
                 font=("Segoe UI", 9)).pack(pady=4)

        outer = tk.Frame(self, bg="#0f0f1a"); outer.pack(fill="both", expand=True, padx=16, pady=6)

        hdrs = ["Field", "Expected", "Found in file", "Status"]
        w    = [14, 20, 28, 12]
        for ci, (h, ww) in enumerate(zip(hdrs, w)):
            tk.Label(outer, text=h, bg="#203864", fg="white",
                     font=("Calibri", 10, "bold"),
                     width=ww, anchor="w", padx=4
                     ).grid(row=0, column=ci, padx=1, pady=1, sticky="w")

        all_fields       = list(COLUMN_ALIASES.keys())
        corrected_fields = {c[0] for c in corrections}
        self._override_vars = {}

        for ri, field in enumerate(all_fields, 1):
            canonical = COLUMN_ALIASES[field][0]
            found_col = mapping.get(field, "")
            is_corrected = field in corrected_fields
            is_missing   = field not in mapping

            if is_missing:
                status_txt, status_fg, row_bg = "MISSING",   "#ff4444", "#2a0000"
            elif is_corrected:
                status_txt, status_fg, row_bg = "Auto-fixed", "#f0a500", "#1a1500"
            else:
                status_txt, status_fg, row_bg = "✔ OK",      "#44ff88", "#001a00"

            tk.Label(outer, text=field, bg=row_bg, fg="#e0e0e0",
                     font=("Calibri", 10, "bold"), width=14, anchor="w", padx=4
                     ).grid(row=ri, column=0, padx=1, pady=1, sticky="w")
            tk.Label(outer, text=canonical, bg=row_bg, fg="#aaaacc",
                     font=("Calibri", 10), width=20, anchor="w", padx=4
                     ).grid(row=ri, column=1, padx=1, pady=1, sticky="w")

            v = tk.StringVar(value=found_col or "— not found —")
            self._override_vars[field] = v
            opts = ["— not found —"] + list(df_columns)
            cb = ttk.Combobox(outer, textvariable=v, values=opts,
                              state="readonly", width=26, font=("Calibri", 9))
            cb.grid(row=ri, column=2, padx=1, pady=1, sticky="w")

            tk.Label(outer, text=status_txt, bg=row_bg, fg=status_fg,
                     font=("Calibri", 10, "bold"), width=12, anchor="w", padx=4
                     ).grid(row=ri, column=3, padx=1, pady=1, sticky="w")

        tk.Label(self,
                 text="You can change any column assignment using the dropdowns above.",
                 bg="#0f0f1a", fg="#888899", font=("Segoe UI", 8)).pack(pady=2)

        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=10)
        tk.Button(bf, text="✔  Auto-correct & Proceed",
                  command=self._proceed,
                  bg="#2d6a4f", fg="white",
                  font=("Segoe UI", 12, "bold"),
                  padx=20, pady=8, relief="flat",
                  cursor="hand2").pack(side="left", padx=10)
        tk.Button(bf, text="✖  Cancel",
                  command=self._cancel,
                  bg="#4a2020", fg="white",
                  font=("Segoe UI", 11),
                  padx=16, pady=8, relief="flat",
                  cursor="hand2").pack(side="left", padx=10)

        self.protocol("WM_DELETE_WINDOW", self._cancel)
        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _proceed(self):
        final = {}
        for field, var in self._override_vars.items():
            v = var.get()
            if v and v != "— not found —":
                final[field] = v
        self.callback({'action': 'proceed', 'mapping': final})
        self.destroy()

    def _cancel(self):
        self.callback({'action': 'cancel'})
        self.destroy()


class SIDRequestDialog(tk.Toplevel):
    """Auto mode: SID not found in REBNI → ask user."""
    def __init__(self, parent, invoice, po, asin, callback):
        super().__init__(parent)
        self.callback = callback
        self.title("SID Required — DICES Validation")
        self.geometry("540x210")
        self.resizable(True, True)
        self.configure(bg="#16213e")
        self.lift(); self.focus_force()

        tk.Label(self, text="⚠  SID Not Found in REBNI",
                 bg="#16213e", fg="#e94560",
                 font=("Segoe UI", 13, "bold")).pack(pady=(14, 4))
        tk.Label(self, text=f"Invoice: {invoice}   PO: {po}   ASIN: {asin}",
                 bg="#16213e", fg="#e0e0e0", font=("Segoe UI", 9)).pack(pady=2)
        tk.Label(self, text="Validate this invoice in DICES and enter the SID below:",
                 bg="#16213e", fg="#aaaacc", font=("Segoe UI", 9)).pack(pady=6)

        ef = tk.Frame(self, bg="#16213e"); ef.pack()
        tk.Label(ef, text="SID from DICES:", bg="#16213e", fg="#e0e0e0",
                 font=("Segoe UI", 10)).pack(side="left", padx=8)
        self._sid = tk.StringVar()
        self._entry = tk.Entry(ef, textvariable=self._sid, width=30,
                               font=("Segoe UI", 10), bg="#1e1e3a", fg="#e0e0e0",
                               insertbackground="white", relief="flat")
        self._entry.pack(side="left", padx=4)
        self._entry.focus_set()

        bf = tk.Frame(self, bg="#16213e"); bf.pack(pady=12)
        tk.Button(bf, text="✔  Continue", command=self._ok,
                  bg="#2d6a4f", fg="white", font=("Segoe UI", 11, "bold"),
                  padx=16, pady=7, relief="flat", cursor="hand2").pack(side="left", padx=8)
        tk.Button(bf, text="✖  Skip", command=self._skip,
                  bg="#6b2737", fg="white", font=("Segoe UI", 10),
                  padx=16, pady=7, relief="flat", cursor="hand2").pack(side="left", padx=8)

        self.bind('<Return>', lambda e: self._ok())
        self.protocol("WM_DELETE_WINDOW", self._skip)
        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _ok(self):
        sid = extract_sid(self._sid.get().strip())
        if sid:
            self.callback(sid); self.destroy()
        else:
            self._entry.config(bg="#3a1e1e")

    def _skip(self):
        self.callback(None); self.destroy()

class CrossPODialog(tk.Toplevel):
    # v6.1: Added automated Cross PO Analysis with engine support
    """Cross PO Confirmation Dialog — v5.3.0 logic restored for manual selection"""
    CASE_DESCRIPTIONS = {
        "Case 1": (
            "Case 1 — No PO, but ASIN received",
            "Rec=0 at claiming PO. Same ASIN received in different PO within same SID.\n"
            "Those units are overage under a different PO."
        ),
        "Case 2": (
            "Case 2 — PO exists but ASIN not ordered there",
            "This PO exists in the claiming SID, but the ASIN was never invoiced there.\n"
            "Inv Qty = 0, but units were received. This is a Cross PO overage."
        ),
        "Case 3": (
            "Case 3 — PO and ASIN exist but Rec > Inv",
            "Both PO and ASIN are present. Invoiced qty = X, but received more than X.\n"
            "Excess units are Cross PO overage."
        ),
    }

    def __init__(self, parent, candidates, current_inv, sid, callback, engine=None):
        super().__init__(parent)
        self.callback   = callback
        self.candidates = [dict(c) for c in candidates]  # v6.1: mutable copy so ANALYZE can update inv_qty/cross_type
        self.sid        = sid
        self.engine     = engine  # v6.1: engine reference for _resolve_inv_qty lookups
        self.file_path  = tk.StringVar()

        self.title("Cross PO Overage — Confirm & Investigate")
        self.geometry("740x540")
        self.resizable(True, True)
        self.configure(bg="#0f0f1a")
        self.protocol("WM_DELETE_WINDOW", self.withdraw)
        self.lift(); self.focus_force()

        tk.Label(self, text="🔄  Cross PO Overage Detected",
                 bg="#16213e", fg="#f0a500",
                 font=("Segoe UI", 13, "bold"), height=2).pack(fill="x")
        tk.Label(self,
                 text=f"SID: {sid}   |   Investigation Invoice: {current_inv}",
                 bg="#0f0f1a", fg="#cccccc",
                 font=("Segoe UI", 9)).pack(pady=2)
        tk.Label(self,
                 text="On confirming, the tool will investigate the Cross PO chain "
                      "to find equivalent shortage.",
                 bg="#0f0f1a", fg="#4a9eff",
                 font=("Segoe UI", 9)).pack(pady=2)

        self.table_frame = tk.LabelFrame(self, text="  Detected Cross PO Candidates  ",
                           bg="#0f0f1a", fg="#e0e0e0",
                           font=("Segoe UI", 9, "bold"), padx=10, pady=6)
        self.table_frame.pack(fill="x", padx=16, pady=6)
        # v6_Secured_Remedy_v6.1: Cross PO SID Details attachment UI directly in panel
        uf = tk.LabelFrame(self, text='  Cross PO SID Details attachment  ', bg='#0f0f1a', fg='#4a9eff', font=('Segoe UI', 9, 'bold'), padx=10, pady=6)
        uf.pack(fill='x', padx=16, pady=4)
        tk.Label(uf, text='Attach SID Details:', bg='#0f0f1a', fg='#cccccc', font=('Segoe UI', 9)).pack(side='left', padx=4)
        tk.Entry(uf, textvariable=self.file_path, width=50, bg='#1e1e3a', fg='white', relief='flat').pack(side='left', padx=6)
        tk.Button(uf, text='Browse', command=lambda: self.file_path.set(filedialog.askopenfilename(filetypes=[('Excel', '*.xlsx *.xls')])), bg='#2d2d5e', fg='white', relief='flat', cursor='hand2').pack(side='left', padx=4)
        tk.Button(uf, text='🔍 ANALYZE', command=self._analyze_cross_po_file, bg='#6f42c1', fg='white', font=('Segoe UI', 9, 'bold'), relief='flat', cursor='hand2', padx=12).pack(side='right', padx=4)
        
        self._render_candidates_table()

        sf = tk.Frame(self, bg="#0f0f1a"); sf.pack(fill="x", padx=16, pady=4)
        tk.Label(sf, text="Select Cross PO to investigate:",
                 bg="#0f0f1a", fg="#e0e0e0",
                 font=("Segoe UI", 10), width=30, anchor="w").pack(side="left")
        opts = [f"PO={c['po']}  Rec={fmt_qty(c['rec_qty'])}  {c['cross_type'].split(chr(8212))[0].strip()}"
                for c in candidates] + ["None — Skip"]
        self._sel_var = tk.StringVar()

        self._sel_cb  = ttk.Combobox(sf, textvariable=self._sel_var,
                                      values=opts, state="readonly", width=50,
                                      font=("Segoe UI", 9))
        self._sel_cb.current(0)
        self._sel_cb.pack(side="left", padx=6)
        self._sel_cb.bind("<<ComboboxSelected>>", self._on_candidate_change)

        cf = tk.LabelFrame(self, text="  Confirm Cross PO Case  ",
                           bg="#0f0f1a", fg="#e0e0e0",
                           font=("Segoe UI", 9, "bold"), padx=12, pady=8)
        cf.pack(fill="x", padx=16, pady=4)

        self._case_var = tk.StringVar(value="Case 1")
        self._case_desc_lbl = tk.Label(cf, text="",
                                        bg="#0f0f1a", fg="#aaaacc",
                                        font=("Segoe UI", 9), justify="left",
                                        wraplength=640, anchor="w")

        for case_key, (case_label, _) in self.CASE_DESCRIPTIONS.items():
            tk.Radiobutton(cf, text=case_label,
                           variable=self._case_var, value=case_key,
                           bg="#0f0f1a", fg="#f0c060",
                           selectcolor="#1a1500",
                           font=("Segoe UI", 10, "bold"),
                           command=self._on_case_change
                           ).pack(anchor="w", pady=2)

        self._case_desc_lbl.pack(anchor="w", pady=4, padx=8)
        self._on_case_change()

        tk.Label(self,
                 text="⚡  On confirming: tool will investigate Cross PO chain "
                      "until full Cross PO rec_qty is explained as shortage.",
                 bg="#0f0f1a", fg="#88ccff",
                 font=("Segoe UI", 9, "italic")).pack(pady=4, padx=16, anchor="w")

        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=10)
        tk.Button(bf, text="✔  Confirm & Investigate",
                  command=self._confirm,
                  bg="#2d6a4f", fg="white",
                  font=("Segoe UI", 12, "bold"),
                  padx=20, pady=9, relief="flat",
                  cursor="hand2").pack(side="left", padx=10)
        tk.Button(bf, text="✖  Skip",
                  command=self._skip,
                  bg="#4a2020", fg="white",
                  font=("Segoe UI", 11),
                  padx=16, pady=9, relief="flat",
                  cursor="hand2").pack(side="left", padx=10)

        self.protocol("WM_DELETE_WINDOW", self._skip)
        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self._sel_cb.current(0)
        self._sel_cb.pack(side="left", padx=6)
        self._sel_cb.bind("<<ComboboxSelected>>", self._on_candidate_change)

        cf = tk.LabelFrame(self, text="  Confirm Cross PO Case  ",
                           bg="#0f0f1a", fg="#e0e0e0",
                           font=("Segoe UI", 9, "bold"), padx=12, pady=8)
        cf.pack(fill="x", padx=16, pady=4)

        self._case_var = tk.StringVar(value="Case 1")
        self._case_desc_lbl = tk.Label(cf, text="",
                                        bg="#0f0f1a", fg="#aaaacc",
                                        font=("Segoe UI", 9), justify="left",
                                        wraplength=640, anchor="w")

        for case_key, (case_label, _) in self.CASE_DESCRIPTIONS.items():
            tk.Radiobutton(cf, text=case_label,
                           variable=self._case_var, value=case_key,
                           bg="#0f0f1a", fg="#f0c060",
                           selectcolor="#1a1500",
                           font=("Segoe UI", 10, "bold"),
                           command=self._on_case_change
                           ).pack(anchor="w", pady=2)

        self._case_desc_lbl.pack(anchor="w", pady=4, padx=8)
        self._on_case_change()

        tk.Label(self,
                 text="⚡  On confirming: tool will investigate Cross PO chain "
                      "until full Cross PO rec_qty is explained as shortage.",
                 bg="#0f0f1a", fg="#88ccff",
                 font=("Segoe UI", 9, "italic")).pack(pady=4, padx=16, anchor="w")

        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=10)
        tk.Button(bf, text="✔  Confirm & Investigate",
                  command=self._confirm,
                  bg="#2d6a4f", fg="white",
                  font=("Segoe UI", 12, "bold"),
                  padx=20, pady=9, relief="flat",
                  cursor="hand2").pack(side="left", padx=10)
        tk.Button(bf, text="✖  Skip",
                  command=self._skip,
                  bg="#4a2020", fg="white",
                  font=("Segoe UI", 11),
                  padx=16, pady=9, relief="flat",
                  cursor="hand2").pack(side="left", padx=10)

        self.protocol("WM_DELETE_WINDOW", self._skip)
        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")


    def _analyze_cross_po_file(self):
        # v6.2.5: Enhanced multi-invoice iteration logic (Checks all invoices for the PO)
        path = self.file_path.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showwarning('File Required', 'Please attach the Cross PO SID level Excel first.', parent=self); return
        try:
            df = pd.read_excel(path, header=0)
            df.columns = [str(c).strip().upper() for c in df.columns]
            sid_col = next((c for c in df.columns if 'SHIPMENT' in c or 'SID' in c), None)
            po_col  = next((c for c in df.columns if 'PO' in c), None)
            inv_col = next((c for c in df.columns if 'INVOICE' in c), None)
            bc_col  = next((c for c in df.columns if 'BARCODE' in c or 'FNSKU' in c or 'UPC' in c), None)
            
            if not all([sid_col, po_col, inv_col]):
                messagebox.showerror('Format Error', f'Excel must have Shipment ID, PO, and Invoice Number columns.', parent=self); return
            
            findings = []
            updated_count = 0
            target_sid = clean(self.sid)
            for c in self.candidates:
                # Find all rows matching SID and PO
                matches = df[(df[sid_col].astype(str).str.contains(target_sid)) & (df[po_col].astype(str) == c['po'])]
                if not matches.empty:
                    # v6.2.5: Iterate through all matching invoices until one with valid quantity is found
                    best_match = None
                    for _, row in matches.iterrows():
                        found_inv = clean(str(row[inv_col]))
                        if self.engine:
                            iq = self.engine._resolve_inv_qty(found_inv, c['asin'], None)
                            if iq is not None and iq > 0:
                                best_match = (found_inv, iq, clean(str(row[bc_col])) if bc_col else "")
                                break
                    
                    if best_match:
                        found_inv, iq, found_bc = best_match
                        c['found_inv'] = found_inv
                        c['found_bc']  = found_bc
                        c['inv_qty']   = iq
                        rec_n = safe_num(c['rec_qty'])
                        if iq >= rec_n: c['cross_type'] = 'Case 0 — Verified (Not Overage)'
                        findings.append(f"• PO {c['po']}: Found Inv {found_inv} (Qty {int(iq)}) | Barcode: {found_bc if found_bc else 'N/A'}")
                        updated_count += 1
                    else:
                        # Fallback to the first one if no valid quantity found, but report as such
                        first_inv = clean(str(matches.iloc[0][inv_col]))
                        c['found_inv'] = first_inv
                        c['found_bc']  = clean(str(matches.iloc[0][bc_col])) if bc_col else ""
                        findings.append(f"• PO {c['po']}: Multiple invoices found, but none matched quantity in database.")
                else:
                    findings.append(f"• PO {c['po']}: Not found in attached file")
            
            # v6.2.5: Dynamically refresh the UI grid immediately
            self._render_candidates_table()
            
            report = "\n".join(findings) if findings else "No candidates processed."
            messagebox.showinfo('Analysis Results', f'Found {updated_count} valid matches in file:\n\n{report}', parent=self)
            
            opts = [f"PO={c['po']}  Rec={fmt_qty(c['rec_qty'])}  {c['cross_type'].split('—')[0].strip()}" for c in self.candidates] + ["None — Skip"]
            self._sel_cb['values'] = opts
        except Exception as e:
            messagebox.showerror('Error', f'Analysis failed: {e}', parent=self)

    def _render_candidates_table(self):
        # v6.1.2: Clear and redraw the table dynamically
        for child in self.table_frame.winfo_children():
            child.destroy()
            
        for ci, h in enumerate(["Cross PO", "ASIN", "Inv Qty", "Rec Qty", "Overage", "Type"]):
            tk.Label(self.table_frame, text=h, bg="#203864", fg="white",
                     font=("Calibri", 10, "bold"), width=14, anchor="w", padx=3
                     ).grid(row=0, column=ci, padx=1, pady=1)
        for ri, c in enumerate(self.candidates, 1):
            inv_n = safe_num(c.get('inv_qty', 0))
            rec_n = safe_num(c['rec_qty'])
            ovg   = max(0.0, rec_n - inv_n)
            for ci, v in enumerate([c['po'], c['asin'],
                                     fmt_qty(inv_n), fmt_qty(rec_n),
                                     fmt_qty(ovg) or "—",
                                     c['cross_type'].split("—")[0].strip()]):
                tk.Label(self.table_frame, text=str(v), bg="#1e1e3a", fg="#e0e0e0",
                         font=("Calibri", 10), width=14, anchor="w", padx=3
                         ).grid(row=ri, column=ci, padx=1, pady=1)

    def _on_candidate_change(self, event=None):
        pass

    def _on_case_change(self):
        case_key = self._case_var.get()
        _, desc  = self.CASE_DESCRIPTIONS.get(case_key, ("", ""))
        self._case_desc_lbl.config(text=desc)

    def _confirm(self):
        idx = self._sel_cb.current()
        if idx >= len(self.candidates):
            self.callback({'action': 'skip'}); self.destroy(); return
        self.callback({
            'action'   : 'confirmed',
            'candidate': self.candidates[idx],
            'case'     : self._case_var.get(),
        })
        self.destroy()

    def _skip(self):
        self.callback({'action': 'skip'}); self.destroy()

class CrossPOVaultDialog(tk.Toplevel):
    """v6.2.6: Vault to manage and investigate Cross POs across multiple shipments."""
    def __init__(self, parent, engine, callback):
        super().__init__(parent)
        self.engine = engine
        self.callback = callback
        self.title("Cross PO Vault — v6.2.6")
        self.geometry("760x520")
        self.configure(bg="#0f111a")
        self.resizable(True, True)
        self.lift(); self.focus_force()

        header = tk.Frame(self, bg="#16213e", height=60)
        header.pack(fill="x")
        tk.Label(header, text="📦  Cross PO Vault", bg="#16213e", fg="#f0a500", font=("Segoe UI", 14, "bold")).pack(side="left", padx=20, pady=12)
        tk.Label(header, text="v6.2.6 | Secured Edition", bg="#16213e", fg="#4a9eff", font=("Segoe UI", 9)).pack(side="right", padx=20)

        ctrl = tk.Frame(self, bg="#0f111a", padx=20, pady=15)
        ctrl.pack(fill="x")
        tk.Label(ctrl, text="Select Shipment ID:", bg="#0f111a", fg="#cccccc", font=("Segoe UI", 10)).pack(side="left")
        
        self.sid_var = tk.StringVar()
        sids = sorted(list(self.engine.cross_po_vault.keys()))
        self.sid_combo = ttk.Combobox(ctrl, textvariable=self.sid_var, values=sids, state="readonly", width=30)
        self.sid_combo.pack(side="left", padx=10)
        self.sid_combo.bind("<<ComboboxSelected>>", lambda e: self._refresh_list())
        
        tk.Button(ctrl, text="🔄 REFRESH", command=self._refresh_list, bg="#1e1e3a", fg="white", relief="flat", padx=10).pack(side="left", padx=5)
        tk.Button(ctrl, text="🗑️ CLEAR VAULT", command=self._clear_vault, bg="#4a2020", fg="white", relief="flat", padx=10).pack(side="right")

        self.list_frame = tk.Frame(self, bg="#0d0d1a", padx=20, pady=5)
        self.list_frame.pack(fill="both", expand=True)
        
        self.canvas = tk.Canvas(self.list_frame, bg="#0d0d1a", highlightthickness=0)
        self.scroll = ttk.Scrollbar(self.list_frame, orient="vertical", command=self.canvas.yview)
        self.scroll_frame = tk.Frame(self.canvas, bg="#0d0d1a")
        
        self.canvas.create_window((0,0), window=self.scroll_frame, anchor="nw", tags="frame")
        self.canvas.configure(yscrollcommand=self.scroll.set)
        
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scroll.pack(side="right", fill="y")
        
        self.scroll_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        
        if sids:
            self.sid_var.set(sids[0])
            self._refresh_list()
        else:
            tk.Label(self.scroll_frame, text="Vault is empty. Stored Cross POs will appear here.", bg="#0d0d1a", fg="#666666", font=("Segoe UI", 10, "italic")).pack(pady=40)

    def _refresh_list(self):
        for w in self.scroll_frame.winfo_children(): w.destroy()
        sid = self.sid_var.get()
        if not sid or sid not in self.engine.cross_po_vault: return
        
        data = self.engine.cross_po_vault[sid]
        for i, item in enumerate(data):
            f = tk.Frame(self.scroll_frame, bg="#1a1a2e", padx=15, pady=10, relief="flat", bd=1)
            f.pack(fill="x", pady=4, padx=2)
            
            c = item['candidate']
            info = f"PO: {c['po']}  |  ASIN: {c['asin']}  |  Case: {item['case']}  |  Budget: {int(item['budget'])} units"
            tk.Label(f, text=info, bg="#1a1a2e", fg="#e0e0e0", font=("Segoe UI", 10, "bold")).pack(side="left")
            
            btn_f = tk.Frame(f, bg="#1a1a2e")
            btn_f.pack(side="right")
            
            tk.Button(btn_f, text="🔍 INVESTIGATE", 
                      command=lambda it=item: self._investigate(it),
                      bg="#2d6a4f", fg="white", font=("Segoe UI", 9, "bold"), relief="flat", padx=12, pady=4).pack(side="left", padx=5)
            
            tk.Button(btn_f, text="❌", 
                      command=lambda idx=i: self._remove_item(idx),
                      bg="#3d0000", fg="#ff4d4d", relief="flat", padx=8).pack(side="left")

    def _investigate(self, item):
        self.callback(item['candidate']['sid'], item['candidate'], item['case'], item['budget'])
        self.destroy()

    def _remove_item(self, idx):
        sid = self.sid_var.get()
        if sid in self.engine.cross_po_vault:
            self.engine.cross_po_vault[sid].pop(idx)
            if not self.engine.cross_po_vault[sid]: del self.engine.cross_po_vault[sid]
            self.sid_combo.config(values=sorted(list(self.engine.cross_po_vault.keys())))
            if not self.engine.cross_po_vault.get(sid): self.sid_var.set("")
            self._refresh_list()

    def _clear_vault(self):
        if messagebox.askyesno("Confirm", "Are you sure you want to clear ALL stored Cross POs?"):
            self.engine.cross_po_vault.clear()
            self.sid_combo.config(values=[])
            self.sid_var.set("")
            self._refresh_list()


class ManualLevelDialog(tk.Toplevel):
    """Manual mode: per-level dialog with IBC/PBC + Cross PO + Mismatches."""
    def __init__(self, parent, matches, remaining_pqv, branch_budget, callback,
                 pending_cb=None, engine=None):
        super().__init__(parent)
        self.callback      = callback
        self.matches       = matches
        self.rem_pqv       = remaining_pqv
        self.branch_budget = branch_budget
        self._pending_cb   = pending_cb
        self._engine_ref   = engine

        self.title("Manual Investigation — Next Step")
        self.geometry("680x560")
        self.configure(bg="#0f0f1a")
        self.resizable(True, True)
        self.protocol("WM_DELETE_WINDOW", self.withdraw)
        self.lift(); self.focus_force()

        tk.Label(self, text="  Manual Investigation — Continue",
                 bg="#16213e", fg="#4a9eff",
                 font=("Segoe UI", 12, "bold"), height=2).pack(fill="x")

        info = f"Remaining PQV: {int(remaining_pqv)}    Branch budget: {int(branch_budget)}"
        tk.Label(self, text=info, bg="#0f0f1a", fg="#cccccc",
                 font=("Segoe UI", 9)).pack(pady=2)

        inv_f = tk.LabelFrame(self, text="  Select Invoice to Continue  ",
                              font=("Segoe UI", 9, "bold"),
                              bg="#0f0f1a", fg="#e0e0e0", padx=10, pady=6)
        inv_f.pack(fill="x", padx=16, pady=4)
        opts = [f"Qty={fmt_qty(m['mtc_qty'])}  |  Inv={m['mtc_inv']}  |  PO={m['mtc_po']}  |  ASIN={m['mtc_asin']}"
                for m in matches]
        self._branch_var = tk.StringVar()
        self._branch_cb  = ttk.Combobox(inv_f, textvariable=self._branch_var,
                                         values=opts, state="readonly", width=70,
                                         font=("Segoe UI", 9))
        if opts: self._branch_cb.current(0)
        self._branch_cb.pack()

        ibc_f = tk.LabelFrame(self, text="  IBC = PBC Validation  ",
                               font=("Segoe UI", 9, "bold"),
                               bg="#0f0f1a", fg="#e0e0e0", padx=10, pady=6)
        ibc_f.pack(fill="x", padx=16, pady=4)

        self._validity = tk.StringVar(value="valid")
        rf = tk.Frame(ibc_f, bg="#0f0f1a"); rf.pack(fill="x")
        tk.Radiobutton(rf, text="✔  IBC = PBC  VALID — Continue investigation",
                       variable=self._validity, value="valid",
                       bg="#0f0f1a", fg="#90ee90", selectcolor="#1e3a28",
                       font=("Segoe UI", 10, "bold"),
                       command=self._toggle).pack(side="left", padx=6)
        tk.Radiobutton(rf, text="✗  IBC ≠ PBC  INVALID — Exclude units",
                       variable=self._validity, value="invalid",
                       bg="#0f0f1a", fg="#ff8888", selectcolor="#3a1e1e",
                       font=("Segoe UI", 10, "bold"),
                       command=self._toggle).pack(side="left", padx=14)
        self._invalid_frame = tk.Frame(ibc_f, bg="#0f0f1a")
        self._invalid_frame.pack(fill="x", pady=3)
        tk.Label(self._invalid_frame, text="Units matched to invalid invoice:",
                 bg="#0f0f1a", fg="#ff8888", font=("Segoe UI", 9)).pack(side="left", padx=4)
        self._inv_qty_var = tk.StringVar()
        tk.Entry(self._invalid_frame, textvariable=self._inv_qty_var, width=10,
                 font=("Segoe UI", 10), bg="#1e1e3a", fg="#ff8888",
                 insertbackground="white", relief="flat").pack(side="left", padx=4)

        self._dices_frame = tk.LabelFrame(self, text="  DICES Details  ",
                                           font=("Segoe UI", 9, "bold"),
                                           bg="#0f0f1a", fg="#e0e0e0", padx=10, pady=6)
        self._dices_frame.pack(fill="x", padx=16, pady=4)
        r1 = tk.Frame(self._dices_frame, bg="#0f0f1a"); r1.pack(fill="x", pady=2)
        tk.Label(r1, text="SID from DICES:", bg="#0f0f1a", fg="#e0e0e0",
                 font=("Segoe UI", 9), width=20, anchor="w").pack(side="left")
        self._sid_var = tk.StringVar()
        tk.Entry(r1, textvariable=self._sid_var, width=28, font=("Segoe UI", 9),
                 bg="#1e1e3a", fg="#e0e0e0", insertbackground="white",
                 relief="flat").pack(side="left", padx=4)

        r2 = tk.Frame(self._dices_frame, bg="#0f0f1a"); r2.pack(fill="x", pady=2)
        tk.Label(r2, text="Barcode from DICES:", bg="#0f0f1a", fg="#e0e0e0",
                 font=("Segoe UI", 9), width=20, anchor="w").pack(side="left")
        self._bc_var = tk.StringVar()
        tk.Entry(r2, textvariable=self._bc_var, width=28, font=("Segoe UI", 9),
                 bg="#1e1e3a", fg="#e0e0e0", insertbackground="white",
                 relief="flat").pack(side="left", padx=4)

        self._toggle()

        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=6)
        tk.Button(bf, text="▶  CONTINUE",
                  command=self._ok, bg="#2d6a4f", fg="white",
                  font=("Segoe UI", 12, "bold"), padx=16, pady=8,
                  relief="flat", cursor="hand2").pack(side="left", padx=6)
        tk.Button(bf, text="🔄  CROSS PO",
                  command=self._cross_po, bg="#7a5c00", fg="white",
                  font=("Segoe UI", 10, "bold"), padx=12, pady=8,
                  relief="flat", cursor="hand2").pack(side="left", padx=6)
        tk.Button(bf, text="⚖  MISMATCH / OVERAGE",
                  command=self._mismatch, bg="#2d4a7a", fg="white",
                  font=("Segoe UI", 10, "bold"), padx=12, pady=8,
                  relief="flat", cursor="hand2").pack(side="left", padx=6)
        tk.Button(bf, text="⬛  STOP THIS ASIN",
                  command=self._stop, bg="#4a2020", fg="white",
                  font=("Segoe UI", 10), padx=12, pady=8,
                  relief="flat", cursor="hand2").pack(side="left", padx=6)

        # ── Pending Invoices & Utility buttons ──────────────────────────────
        bf2 = tk.Frame(self, bg="#0f0f1a"); bf2.pack(pady=4, fill='x', padx=16)
        tk.Button(bf2, text="📋  VIEW ALL PENDING INVOICES",
                  command=self._show_pending, bg="#3a2a00", fg="#f0c060",
                  font=("Segoe UI", 10, "bold"), padx=16, pady=7,
                  relief="flat", cursor="hand2").pack(side="left", expand=True, fill="x", padx=(0,4))
        
        tk.Button(bf2, text="🔍 LOOKUP INV QTY",
                  command=self._lookup_inv_qty, bg="#1c2c42", fg="#80a0ff",
                  font=("Segoe UI", 10, "bold"), padx=16, pady=7,
                  relief="flat", cursor="hand2").pack(side="right", padx=(4,0))
        
        tk.Button(bf2, text="📦 RECEIVED QTY",
                  command=self._show_rec_qty_lookup, bg="#d4a017", fg="black",
                  font=("Segoe UI", 10, "bold"), padx=16, pady=7,
                  relief="flat", cursor="hand2").pack(side="right", padx=4)

        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _lookup_inv_qty(self):
        dlg = tk.Toplevel(self)
        dlg.title("Lookup Invoice Quantity")
        dlg.configure(bg="#0d1117")
        dlg.attributes("-topmost", True)
        
        tk.Label(dlg, text="Invoice No:", bg="#0d1117", fg="white", font=("Segoe UI", 10)).grid(row=0, column=0, padx=10, pady=(15,5), sticky='e')
        inv_ent = tk.Entry(dlg, font=("Segoe UI", 10), bg="#21262d", fg="white", insertbackground="white")
        inv_ent.grid(row=0, column=1, padx=10, pady=(15,5))
        
        tk.Label(dlg, text="ASIN:", bg="#0d1117", fg="white", font=("Segoe UI", 10)).grid(row=1, column=0, padx=10, pady=5, sticky='e')
        asin_ent = tk.Entry(dlg, font=("Segoe UI", 10), bg="#21262d", fg="white", insertbackground="white")
        asin_ent.grid(row=1, column=1, padx=10, pady=5)
        
        res_lbl = tk.Label(dlg, text="", bg="#0d1117", fg="#f0c060", font=("Segoe UI", 11, "bold"))
        res_lbl.grid(row=3, column=0, columnspan=2, pady=(10,15))
        
        def do_lookup():
            ino = inv_ent.get().strip()
            asn = asin_ent.get().strip()
            
            # Use direct engine reference passed from app
            engine = getattr(self, '_engine_ref', None)
            
            if not engine:
                res_lbl.config(text="Engine not connected.", fg="#f85149")
                return
                
            # Optimized lookup using the high-speed indexer (inv_iam)
            found_qty = engine._resolve_inv_qty(ino, asn, None)
            
            if found_qty is not None:
                res_lbl.config(text=f"Exact Inv Qty = {int(found_qty)} units", fg="#3fb950")
            else:
                res_lbl.config(text="Not found in Invoice Data.", fg="#f85149")
                
        tk.Button(dlg, text="🔍 Search Data", bg="#238636", fg="white", font=("Segoe UI", 10, "bold"),
                  command=do_lookup, cursor="hand2", padx=20).grid(row=2, column=0, columnspan=2, pady=10)
        
        dlg.update_idletasks()
        px_dlg = self.winfo_x() + (self.winfo_width() - dlg.winfo_width()) // 2
        py_dlg = self.winfo_y() + 50
        dlg.geometry(f"+{px_dlg}+{py_dlg}")
        inv_ent.focus_set()

    def _show_rec_qty_lookup(self):
        """v6.2.5: Interactive Received Qty Lookup within Manual Mode."""
        # Get current selection from combobox
        idx = self._branch_cb.current()
        if idx < 0:
            messagebox.showwarning("Selection Required", "Please select a matching invoice first.")
            return
            
        m = self.matches[idx]
        asin = clean(m.get('mtc_asin', ''))
        po   = clean(m.get('mtc_po', ''))
        
        # SID from current entry field (user might have entered it)
        sid = extract_sid(self._sid_var.get())
        
        if not sid or not po or not asin:
            messagebox.showerror("Data Missing", f"Please ensure SID, PO, and ASIN are identified.\nSID: {sid}\nPO: {po}\nASIN: {asin}")
            return
            
        dlg = tk.Toplevel(self)
        dlg.title(f"Received Qty Lookup — {asin}")
        dlg.geometry("480x420")
        dlg.configure(bg="#0d1117")
        dlg.attributes("-topmost", True)
        
        tk.Label(dlg, text="  Shipment Reconciliation Data", bg="#161b22", fg="#4a9eff", font=("Segoe UI", 12, "bold"), height=2).pack(fill="x")
        
        info_f = tk.Frame(dlg, bg="#0d1117")
        info_f.pack(pady=15, padx=20, fill='x')
        
        details = [("SID:", sid), ("PO:", po), ("ASIN:", asin)]
        for i, (l, v) in enumerate(details):
            tk.Label(info_f, text=l, bg="#0d1117", fg="#888888", font=("Segoe UI", 10)).grid(row=i, column=0, sticky='w', pady=2)
            tk.Label(info_f, text=v, bg="#0d1117", fg="white", font=("Segoe UI", 10, "bold")).grid(row=i, column=1, sticky='w', padx=10, pady=2)

        res_f = tk.LabelFrame(dlg, text="  Calculation Summary  ", bg="#0d1117", fg="#4a9eff", font=("Segoe UI", 9, "bold"), padx=15, pady=15)
        res_f.pack(fill="both", expand=True, padx=20, pady=10)

        engine = getattr(self, '_engine_ref', None)
        if not engine or not hasattr(engine, 'rebni_p'):
            tk.Label(res_f, text="Investigation Engine not initialized.", bg="#0d1117", fg="#f85149").pack()
        else:
            rows = engine.rebni_p.get((sid, po, asin), [])
            if not rows:
                tk.Label(res_f, text="No REBNI data found for this combination.", bg="#0d1117", fg="#f85149").pack()
            else:
                r0 = rows[0]
                rec_post = safe_num(r0.get('qty_received_postadj', 0))
                rec_unp  = safe_num(r0.get('quantity_unpacked', 0))
                avail    = safe_num(r0.get('rebni_available', 0))
                
                # Highlight logic
                tk.Label(res_f, text=f"Received (Post Adj):", bg="#0d1117", fg="#cccccc", font=("Segoe UI", 10)).grid(row=0, column=0, sticky='w')
                tk.Label(res_f, text=f"{int(rec_post)} units", bg="#0d1117", fg="#3fb950", font=("Segoe UI", 12, "bold")).grid(row=0, column=1, sticky='w', padx=10)
                
                tk.Label(res_f, text=f"Quantity Unpacked:", bg="#0d1117", fg="#cccccc", font=("Segoe UI", 10)).grid(row=1, column=0, sticky='w', pady=10)
                tk.Label(res_f, text=f"{int(rec_unp)} units", bg="#0d1117", fg="#58a6ff", font=("Segoe UI", 12, "bold")).grid(row=1, column=1, sticky='w', padx=10)
                
                # v6.2.5: Crucial Addition requested by Mukesh
                tk.Label(res_f, text=f"Quantity Adjusted:", bg="#0d1117", fg="#cccccc", font=("Segoe UI", 10)).grid(row=2, column=0, sticky='w', pady=5)
                tk.Label(res_f, text=f"{int(safe_num(r0.get('quantity_adjusted', 0)))} units", bg="#0d1117", fg="#ff4d4d", font=("Segoe UI", 12, "bold")).grid(row=2, column=1, sticky='w', padx=10)
                
                tk.Label(res_f, text=f"REBNI Available:", bg="#0d1117", fg="#cccccc", font=("Segoe UI", 10)).grid(row=3, column=0, sticky='w')
                tk.Label(res_f, text=f"{int(avail)} units", bg="#0d1117", fg="#f0a500", font=("Segoe UI", 12, "bold")).grid(row=3, column=1, sticky='w', padx=10)
                
                tk.Label(res_f, text="\u24D8  Verified using REBNI strictly.", bg="#0d1117", fg="#6e7681", font=("Segoe UI", 8, "italic")).grid(row=4, column=0, columnspan=2, pady=(20,0))

        tk.Button(dlg, text="CLOSE", command=dlg.destroy, bg="#333", fg="white", font=("Segoe UI", 9, "bold"), relief="flat", padx=20, pady=8).pack(pady=15)
        
        dlg.update_idletasks()
        px_dlg = self.winfo_x() + (self.winfo_width()  - dlg.winfo_width())  // 2
        py_dlg = self.winfo_y() + (self.winfo_height() - dlg.winfo_height()) // 2
        dlg.geometry(f"+{px_dlg}+{py_dlg}")

    def _toggle(self):
        if self._validity.get() == "valid":
            self._invalid_frame.pack_forget()
            self._dices_frame.pack(fill="x", padx=16, pady=4)
        else:
            self._dices_frame.pack_forget()
            self._invalid_frame.pack(fill="x", pady=3)

    def _ok(self):
        sel = self._branch_cb.current()
        if sel < 0 or sel >= len(self.matches):
            messagebox.showwarning("Select Invoice", "Please select an invoice.", parent=self)
            return
        match = self.matches[sel]
        if self._validity.get() == "valid":
            sid = extract_sid(self._sid_var.get().strip())
            if not sid:
                messagebox.showwarning("SID Required", "Please enter SID from DICES.", parent=self)
                return
            self.callback({'action': 'valid', 'chosen_match': match,
                           'sid': sid, 'barcode': self._bc_var.get().strip() or "[DICES]"})
        else:
            qty_str = self._inv_qty_var.get().strip()
            try:
                qty = float(qty_str)
            except:
                messagebox.showwarning("Qty Required", "Enter units matched to invalid invoice.", parent=self)
                return
            self.callback({'action': 'invalid', 'chosen_match': match, 'invalid_qty': qty})
        self.destroy()

    def _cross_po(self):
        self.callback({'action': 'cross_po',
                       'chosen_match': self.matches[self._branch_cb.current()] if self.matches else None})
        self.destroy()

    def _mismatch(self):
        dlg = tk.Toplevel(self)
        dlg.title("Mismatch / Overage Details")
        dlg.geometry("460x260")
        dlg.configure(bg="#0f0f1a")
        dlg.lift(); dlg.focus_force()

        fields = [("ASIN received:", "asin"), ("SID:", "sid"), ("PO:", "po"),
                  ("Inv Qty (invoiced):", "inv_qty"), ("Overage Qty received:", "ovg_qty")]
        vars_ = {}
        for i, (lbl, key) in enumerate(fields):
            tk.Label(dlg, text=lbl, bg="#0f0f1a", fg="#e0e0e0",
                     font=("Segoe UI", 10), width=22, anchor="w"
                     ).grid(row=i, column=0, padx=12, pady=5)
            v = tk.StringVar()
            tk.Entry(dlg, textvariable=v, width=26, font=("Segoe UI", 10),
                     bg="#1e1e3a", fg="#e0e0e0", insertbackground="white",
                     relief="flat").grid(row=i, column=1, padx=8, pady=5)
            vars_[key] = v

        def submit():
            data = {k: v.get().strip() for k, v in vars_.items()}
            dlg.destroy()
            self.callback({'action': 'mismatch', 'mismatch_data': data})
            self.destroy()

        tk.Button(dlg, text="✔  Confirm & Investigate", command=submit,
                  bg="#2d6a4f", fg="white", font=("Segoe UI", 11, "bold"),
                  padx=14, pady=7, relief="flat", cursor="hand2"
                  ).grid(row=len(fields), column=0, columnspan=2, pady=12)

    def refresh_from_engine(self, engine):
        """Update combobox options dynamically when Confirm Edits is clicked."""
        if not engine.user_overrides:
            return
        patched = []
        for mtch in self.matches:
            k = (clean(mtch.get('mtc_inv', '')), clean(mtch.get('mtc_asin', '')), clean(mtch.get('mtc_po', '')))
            override = engine.user_overrides.get(k, {})
            if not override: override = engine.user_overrides.get(clean(mtch.get('mtc_inv', '')), {})
            
            if override:
                mtch = dict(mtch)
                # Apply arbitrary fields edited by the user
                for key_name, override_val in override.items():
                    if override_val != '':
                        mtch[key_name] = safe_num(override_val) if 'qty' in key_name else override_val
            patched.append(mtch)
        
        self.matches = patched
        opts = [f"Qty={fmt_qty(m.get('mtc_qty',0))}  |  Inv={m.get('mtc_inv','')}  |  PO={m.get('mtc_po','')}  |  ASIN={m.get('mtc_asin','')}"
                for m in self.matches]
        self._branch_cb['values'] = opts
        idx = self._branch_cb.current()
        if opts and idx >= 0:
            self._branch_cb.current(idx)

    def _show_pending(self):
        # v6.2.5: Isolated per-ASIN matches
        matches = self._engine_ref.asin_pending_matches if self._engine_ref else []
        if not matches and self._pending_cb:
            self._pending_cb()
            return
        if not matches:
            messagebox.showinfo("Pending Invoices", "No pending invoices for this ASIN chain.", parent=self)
            return
        self._pending_cb()  # Always show if we have matches

    def _stop(self):
        self.callback({'action': 'stop'}); self.destroy()


class PendingInvoicesDialog(tk.Toplevel):
    """
    Shown before finalizing a claiming ASIN.
    Lists all matched invoices that were identified but not yet investigated.
    User can pick one to continue investigation or skip.
    """
    def __init__(self, parent, pending_invoices, asin, callback):
        super().__init__(parent)
        self.callback         = callback
        self.pending_invoices = pending_invoices

        self.title("Pending Matched Invoices — Action Required")
        self.geometry("760x480")
        self.configure(bg="#0f0f1a")
        self.resizable(True, True)
        self.protocol("WM_DELETE_WINDOW", self.withdraw)
        self.lift(); self.focus_force()

        tk.Label(self,
                 text="📋  Uninvestigated Matched Invoices Found",
                 bg="#16213e", fg="#f0a500",
                 font=("Segoe UI", 12, "bold"), height=2).pack(fill="x")

        tk.Label(self,
                 text=(f"ASIN: {asin} — The following matched invoices were identified during\n"
                       "investigation but have not yet been investigated. "
                       "Select one to continue or click 'Go to Next ASIN'."),
                 bg="#0f0f1a", fg="#cccccc",
                 font=("Segoe UI", 9), justify="left").pack(pady=6, padx=16, anchor="w")

        # Table of pending invoices
        tf = tk.LabelFrame(self, text="  Pending Matched Invoices  ",
                           bg="#0f0f1a", fg="#e0e0e0",
                           font=("Segoe UI", 9, "bold"), padx=10, pady=6)
        tf.pack(fill="both", expand=True, padx=16, pady=4)

        cols = ("Mtc Qty", "Invoice No", "PO", "ASIN", "Level")
        vsb = ttk.Scrollbar(tf, orient="vertical")
        hsb = ttk.Scrollbar(tf, orient="horizontal")
        self.tree = ttk.Treeview(tf, columns=cols, show='headings', yscrollcommand=vsb.set, xscrollcommand=hsb.set, height=10)
        vsb.config(command=self.tree.yview); hsb.config(command=self.tree.xview)
        
        col_w = {"Mtc Qty": 80, "Invoice No": 220, "PO": 140, "ASIN": 140, "Level": 60}
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=col_w[c], anchor='w', stretch=tk.YES)
            
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tf.grid_rowconfigure(0, weight=1)
        tf.grid_columnconfigure(0, weight=1)
        
        # New Feature: Styling Treeview for dark theme pending window
        s = ttk.Style()
        if "clam" in s.theme_names(): s.theme_use("clam")
        s.configure("Pending.Treeview", font=("Calibri", 10), rowheight=24, background="#1e1e3a", foreground="#e0e0e0", fieldbackground="#1e1e3a")
        s.configure("Pending.Treeview.Heading", font=("Calibri", 10, "bold"), background="#203864", foreground="white")
        s.map("Pending.Treeview", background=[('selected', '#2d4a7a')])
        self.tree.configure(style="Pending.Treeview")

        self._populate_tree()

        # Dropdown to select one
        sf = tk.Frame(self, bg="#0f0f1a"); sf.pack(fill="x", padx=16, pady=4)
        tk.Label(sf, text="Select invoice to investigate:",
                 bg="#0f0f1a", fg="#e0e0e0",
                 font=("Segoe UI", 10), width=30, anchor="w").pack(side="left")
        opts = [f"Qty={fmt_qty(inv.get('mtc_qty',''))}  |  Inv={inv.get('mtc_inv','')}  |  "
                f"PO={inv.get('mtc_po','')}  |  ASIN={inv.get('mtc_asin','')}"
                for inv in pending_invoices]
        self._sel = tk.StringVar()
        self._cb  = ttk.Combobox(sf, textvariable=self._sel,
                                   values=opts, state="readonly", width=60,
                                   font=("Segoe UI", 9))
        if opts: self._cb.current(0)
        self._cb.pack(side="left", padx=6)

        # Buttons
        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=10)
        tk.Button(bf, text="🔍  Investigate Selected Invoice",
                  command=self._investigate,
                  bg="#2d4a7a", fg="white",
                  font=("Segoe UI", 11, "bold"),
                  padx=16, pady=8, relief="flat",
                  cursor="hand2").pack(side="left", padx=8)
        tk.Button(bf, text="▶▶  Go to Next ASIN",
                  command=self._next_asin,
                  bg="#2d6a4f", fg="white",
                  font=("Segoe UI", 11, "bold"),
                  padx=16, pady=8, relief="flat",
                  cursor="hand2").pack(side="left", padx=8)

        self.protocol("WM_DELETE_WINDOW", self._next_asin)
        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _populate_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for inv in self.pending_invoices:
            vals = [
                fmt_qty(inv.get('mtc_qty', '')),
                inv.get('mtc_inv', ''),
                inv.get('mtc_po', ''),
                inv.get('mtc_asin', ''),
                str(inv.get('_depth', '?'))
            ]
            self.tree.insert('', 'end', values=vals)

    def refresh_from_engine(self, engine):
        if not engine.user_overrides: return
        patched = []
        for mtch in self.pending_invoices:
            k = (clean(mtch.get('mtc_inv', '')), clean(mtch.get('mtc_asin', '')), clean(mtch.get('mtc_po', '')))
            ov = engine.user_overrides.get(k, {})
            if not ov: ov = engine.user_overrides.get(clean(mtch.get('mtc_inv', '')), {})
            if ov:
                mtch = dict(mtch)
                for key_name, override_val in ov.items():
                    if override_val != '':
                        mtch[key_name] = safe_num(override_val) if 'qty' in key_name else override_val
            patched.append(mtch)
        
        self.pending_invoices = patched
        self._populate_tree()
        opts = [f"Qty={fmt_qty(inv.get('mtc_qty',''))}  |  Inv={inv.get('mtc_inv','')}  |  PO={inv.get('mtc_po','')}  |  ASIN={inv.get('mtc_asin','')}" for inv in self.pending_invoices]
        self._cb['values'] = opts
        if opts and self._cb.current() < 0: self._cb.current(0)

    def _investigate(self):
        idx = self._cb.current()
        if idx < 0 or idx >= len(self.pending_invoices):
            messagebox.showwarning("Select Invoice", "Please select an invoice to investigate.", parent=self)
            return
        self.callback({'action': 'investigate', 'match': self.pending_invoices[idx]})
        self.destroy()

    def _next_asin(self):
        self.callback({'action': 'next_asin'})
        self.destroy()

class CorrespondenceDialog(tk.Toplevel):
    def __init__(self, parent, all_rows, app=None):
        super().__init__(parent)
        self.all_rows = all_rows
        self.app = app
        self.title("Scenario Selection — Get Correspondence — v6.2.5")
        self.geometry("1000x850")
        self.configure(bg="#0f0f1a")
        
        # Enable Maximize/Minimize and ensure centered display
        self.resizable(True, True)
        self.focus_force()

        self.scenarios = [
            "Claiming Short",
            "REBNI",
            "Matching (Bulk)",
            "Mismatch ASIN Overages",
            "Invalid/Dummy Invoice [Dev]",
            "IBC vs PBC [Dev]"
        ]
        
        self.v_code_var = tk.StringVar(value="[Vendor Code]")
        self.fc_id_var = tk.StringVar(value="[FC ID]")
        
        self._build_ui()
        
    def _build_ui(self):
        f = tk.Frame(self, bg="#0f0f1a", padx=25, pady=25)
        f.pack(fill="both", expand=True)
        
        # --- Header with Professional Styling ---
        hdr = tk.Frame(f, bg="#0f0f1a")
        hdr.pack(fill="x", pady=(0, 20))
        tk.Label(hdr, text="Investigation Correspondence Generator",
                 fg="#4a9eff", bg="#0f0f1a",
                 font=("Segoe UI", 16, "bold")).pack(side="left")

        # --- SELECTION AREA ---
        sel_f = tk.LabelFrame(f, text=" STEP 1: SELECT SCENARIO ", fg="#4a9eff", bg="#0f0f1a", font=("Segoe UI", 9, "bold"))
        sel_f.pack(fill="x", pady=10)
        self.scenario_var = tk.StringVar()
        self.scenario_cb = ttk.Combobox(sel_f, textvariable=self.scenario_var, 
                                        values=self.scenarios, state="readonly", 
                                        font=("Segoe UI", 12))
        self.scenario_cb.pack(fill="x", padx=15, pady=15)
        self.scenario_cb.bind("<<ComboboxSelected>>", self.generate_text)
        
        # --- INPUT AREA ---
        inp_f = tk.LabelFrame(f, text=" STEP 2: HYDRATION DETAILS ", fg="#4a9eff", bg="#0f0f1a", font=("Segoe UI", 9, "bold"))
        inp_f.pack(fill="x", pady=10)
        grid_f = tk.Frame(inp_f, bg="#0f0f1a")
        grid_f.pack(fill="x", padx=15, pady=10)
        
        tk.Label(grid_f, text="Vendor Code:", bg="#0f0f1a", fg="#aaaaaa").grid(row=0, column=0, sticky="w")
        tk.Entry(grid_f, textvariable=self.v_code_var, width=25, font=("Segoe UI", 10)).grid(row=0, column=1, padx=(5, 30))
        
        tk.Label(grid_f, text="FC ID:", bg="#0f0f1a", fg="#aaaaaa").grid(row=0, column=2, sticky="w")
        tk.Entry(grid_f, textvariable=self.fc_id_var, width=25, font=("Segoe UI", 10)).grid(row=0, column=3, padx=5)

        # --- TEXT AREA ---
        tk.Label(f, text=" STEP 3: REVIEW CORRESPONDENCE ( интерпретированный ) ",
                 fg="#e0e0e0", bg="#0f0f1a", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(15, 5))
        
        # Use a higher contrast text box for legibility
        self.text_area = tk.Text(f, font=("Segoe UI", 11), 
                                 bg="#ffffff", fg="#000000", padx=15, pady=15,
                                 insertbackground="black", wrap="word", relief="flat")
        self.text_area.pack(fill="both", expand=True)

        # --- FOOTER BUTTONS ---
        btn_f = tk.Frame(f, bg="#0f0f1a")
        btn_f.pack(fill="x", pady=(25, 0))
        
        tk.Button(btn_f, text="📋  COPY TO CLIPBOARD ", 
                  command=self.copy_to_clip,
                  bg="#2d6a4f", fg="white", font=("Segoe UI", 11, "bold"),
                  relief="flat", padx=25, pady=12, cursor="hand2").pack(side="left")

        if self.app:
            tk.Button(btn_f, text="💬  AI INVESTIGATOR ", 
                      command=lambda: self.app.open_ai_chat(),
                      bg="#e94560", fg="white", font=("Segoe UI", 11, "bold"),
                      relief="flat", padx=25, pady=12, cursor="hand2").pack(side="left", padx=10)
        
        tk.Button(btn_f, text="CLOSE", command=self.destroy,
                  bg="#333333", fg="white", font=("Segoe UI", 10),
                  relief="flat", padx=20, pady=12, cursor="hand2").pack(side="right")

    def copy_to_clip(self):
        txt = self.text_area.get("1.0", tk.END).strip()
        if txt:
            self.clipboard_clear(); self.clipboard_append(txt)
            messagebox.showinfo("Success", "Professional correspondence copied!", parent=self)

    def generate_text(self, event=None):
        scenario = self.scenario_var.get()
        if not scenario: return
        
        # Gather data
        primary_row = next((r for r in self.all_rows if r.get('depth', 0) == 0), {})
        sid = str(primary_row.get('sid', '[SID missing]'))
        po  = str(primary_row.get('po', '[PO missing]'))
        inv = str(primary_row.get('invoice', '[Invoice missing]'))
        vc  = self.v_code_var.get()
        fc  = self.fc_id_var.get()
        
        # Total sums for header
        t_billed = int(sum(safe_num(r.get('inv_qty', 0)) for r in self.all_rows if r.get('depth', 0) == 0))
        t_received = int(sum(safe_num(r.get('rec_qty', 0)) for r in self.all_rows if r.get('depth', 0) == 0))
        
        # ASIN list generation
        asin_lines = []
        for r in self.all_rows:
            # Only include "real" rows with shortages
            if str(r.get('barcode','')).strip() and not str(r.get('barcode','')).startswith('['):
                iq = int(safe_num(r.get('inv_qty', 0)))
                rq = int(safe_num(r.get('rec_qty', 0)))
                if iq > rq:
                    asin = str(r.get('asin', ''))
                    missing = iq - rq
                    amt = float(safe_num(r.get('cp_status', 0)))
                    asin_lines.append(f"Asin: {asin} ---- Invoice Qty: {iq} ---- Received Qty: {rq} ---- Shortage Qty: {missing} ---- Amount: INR ---- {amt:.2f}")
        
        asin_block = "\n".join(asin_lines) if asin_lines else "[No ASIN Shortages detected in Preview]"
        
        signature = f"Regards,\nMUKESH | pathlmuk\nROW IB"
        
        if scenario == "Claiming Short":
            text = (
                f"Hello FC Team, \n\n"
                f"Note: We have already performed the all the virtual research/checks such as cross receiving, overage, REBNI, adjustments etc. "
                f"and need FC support for physical search on floor to locate missing units as per revised SOP and update at the earliest.\n\n"
                f"We are able to see that unit’s shortage received in claiming shipment.\n\n"
                f"SID#{sid} | Total Billed Qty - {t_billed} | Received Qty - {t_received}\n"
                f"Invoice#{inv} | Total Billed Qty - {t_billed} | Received Qty - {t_received}\n\n"
                f"Please locate the following ASINs that are missing from PO#{po}, If units not found physically take support from SLP Team and give update.\n\n"
                f"{asin_block}\n\n"
                f"NOTE : FC Team we are able to units are short at SID Level and No overages are Found.\n\n"
                f"**Post the investigation, please flip it back to our CTI**\n"
                f"C: TOC-India\nT: MFI\nI: PD\nGroup: NOC MFI\n\n"
                f"{signature}"
            )
        elif scenario == "REBNI":
            # Identify where REBNI matches occurred
            c_sids = set()
            m_sids = set()
            for r in self.all_rows:
                rem = str(r.get('remarks', ''))
                rsid = str(r.get('sid', ''))
                if 'REBNI' in rem or 'TSP to utilize' in rem or r.get('type') == 'rebni_shipment':
                    if r.get('depth', 0) == 0: c_sids.add(rsid)
                    else: m_sids.add(rsid)
            
            # Construct the SID line based on where REBNI was found
            sid_parts = []
            if c_sids: sid_parts.append(f"claiming SID - {', '.join(sorted(c_sids))}")
            if m_sids: sid_parts.append(f"matching SID - {', '.join(sorted(m_sids))}")
            sid_statement = " / ".join(sid_parts) if sid_parts else f"claiming SID - {sid}"

            # Build REBNI Table from collected data
            rebni_table = f"{'PO':<16} {'ASIN':<16} {'SID':<18} {'CP':<15} {'Available REBNI'}\n"
            rebni_table += "═" * 85 + "\n"
            engine = getattr(self.app, 'engine', None)
            if engine and engine.collected_rebni:
                for k, rows in engine.collected_rebni.items():
                    for r_row in rows:
                        r_po = str(r_row.get('po', ''))
                        r_asin = str(r_row.get('asin', ''))
                        r_sid = str(r_row.get('shipment_id', ''))
                        r_cp = float(safe_num(r_row.get('item_cost', 0)))
                        r_avail = int(safe_num(r_row.get('rebni_available', 0)))
                        rebni_table += f"{r_po:<16} {r_asin:<16} {r_sid:<18} {r_cp:<15.2f} {r_avail}\n"
            else:
                rebni_table += "[No REBNI matches found in investigation summary]\n"

            text = (
                f"Hello Team, \n\n"
                f"We see that vendor sent overages of mismatch ASIN in {sid_statement}. "
                f"Below suggested REBNIs are available as per the current REBNI report in line with same state FCs and CP criteria. \n\n"
                f"{rebni_table}\n\n"
                f"Please check and utilize the REBNI and update the remaining PQV units. If suggested REBNI are utilized somewhere else, "
                f"then share Invoice, ASIN and PO level details along with Invoice copy where its matched for validation. \n\n"
                f"Note: \n"
                f"1. If Suggested REBNI comes under same CP limit, the shipment ID and PO shouldn't be the factor. \n"
                f"2. Suggested REBNI comes under same shipment, them CP variance isn't the factor.\n\n"
                f"{signature}"
            )
        elif scenario == "Matching (Bulk)":
            text = (
                f"Hello FC Team, \n\n"
                f"Note: We have already performed the all the virtual research/checks such as cross receiving, overage, REBNI, adjustments etc. "
                f"and need FC support for physical search on floor to locate missing units as per revised SOP and update at the earliest.\n\n"
                f"We are able to see that units are received completely in claiming Shipment ID#{sid} and received units got matched with different invoices where we found shortage units received in matched invoices.\n\n"
                f"Please locate the following ASINs that are missing from PO#{po}, If units not found physically take support from SLP Team and give update.\n\n"
                f"{asin_block}\n\n"
                f"**Post the investigation, please flip it back to our CTI**\n"
                f"C: TOC-India\nT: MFI\nI: PD\nGroup: NOC MFI\n\n"
                f"{signature}"
            )
        elif scenario == "Mismatch ASIN Overages":
            text = (
                f"### Hello FinOps Team,  \n\n"
                f"This is regarding the suggested overage ASIN observed under the claiming Shipment ID [{sid}].  \n"
                f"It is clearly visible that:  \n\n"
                f"- Invoiced Quantity: {t_billed} units\n"
                f"- Received Quantity: {t_received} units\n"
                f"- Overage Units: {t_received - t_billed if t_received > t_billed else 0} units  \n\n"
                f"This indicates that the vendor has sent indifferent ASIN instead of the claiming ASINs.  \n\n"
                f"Request:\n"
                f"- Kindly utilize the suggested overages and resolve the PQV at the earliest.  \n"
                f"- If the suggested overages are not utilized, please obtain a credit note from the vendor code, explaining why different ASIN overages were sent in place of the claiming ASIN.  \n\n"
                f"{signature}"
            )
        else:
            text = f"Hello Team,\n\nScenario '{scenario}' template is pending final formatting logic.\n\n{signature}"

        self.text_area.delete("1.0", tk.END)
        self.text_area.insert(tk.END, text)

# ═══════════════════════════════════════════════════════════
class PreviewPanel(tk.Toplevel):
    COLS      = ['Barcode', 'Inv no', 'SID', 'PO', 'ASIN', 'Inv Qty',
                 'Rec Qty', 'Mtc Qty', 'Mtc Inv', 'Mtc ASIN', 'Mtc PO', 'FC', 'Remarks', 'Date', 'CP']
    COL_W_PX  = [130, 160, 130, 90, 110, 60, 60, 60, 160, 130, 130, 80, 240, 150, 180]

    def __init__(self, parent):
        super().__init__(parent)
        self._app = None
        self.title("Investigation Preview — v6.2.5 (editable)")
        self.geometry("1400x750")
        self.configure(bg="#0f0f1a")
        self.resizable(True, True)
        self.protocol("WM_DELETE_WINDOW", self.withdraw)
        self.protocol("WM_DESTROY", lambda *args: None) # ignore secondary destroys
        self._undo_stack = []   # v6.0.0: undo snapshots (max 10)
        self._redo_stack = []   # v6.0.0: redo snapshots
        self._clipboard_cell = None  # for right-click paste
        
        hdr_frame = tk.Frame(self, bg="#16213e")
        hdr_frame.pack(fill="x")
        tk.Label(hdr_frame, text="  Live Investigation Preview",
                 bg="#16213e", fg="#4a9eff", font=("Segoe UI", 10, "bold"), height=2).pack(side="left")
        
        # v6.0.0 Fix: Smaller buttons (font 9, less padding)
        tk.Button(hdr_frame, text="✔ CONFIRM EDITS", command=self.confirm_edits,
                  bg="#1a5a1a", fg="#90ff90", font=("Segoe UI", 9, "bold"), relief="flat", padx=10, pady=4, cursor="hand2").pack(side="right", padx=6, pady=4)
        
        tk.Button(hdr_frame, text="↪ REDO", command=self._redo,
                  bg="#5e3c8b", fg="white", font=("Segoe UI", 9, "bold"), relief="flat", padx=8, pady=4, cursor="hand2").pack(side="right", padx=2, pady=4)

        tk.Button(hdr_frame, text="↩ UNDO", command=self._undo,
                  bg="#8b5e3c", fg="white", font=("Segoe UI", 9, "bold"), relief="flat", padx=8, pady=4, cursor="hand2").pack(side="right", padx=2, pady=4)
        
        tk.Button(hdr_frame, text="✉ CORRESPONDENCE", command=self.show_correspondence,
                  bg="#3949ab", fg="white", font=("Segoe UI", 9, "bold"), relief="flat", padx=10, pady=4, cursor="hand2").pack(side="right", padx=6, pady=4)
        
        tk.Button(hdr_frame, text="🔍 MTC QTY", command=self._show_mtc_qty_lookup,
                  bg="#6f42c1", fg="white", font=("Segoe UI", 9, "bold"), relief="flat", padx=10, pady=4, cursor="hand2").pack(side="right", padx=6, pady=4)
                  
        tk.Button(hdr_frame, text="💵 CP Variance", command=self._show_cp_variance,
                  bg="#00563b", fg="#e0ffe0", font=("Segoe UI", 9, "bold"), relief="flat", padx=10, pady=4, cursor="hand2").pack(side="right", padx=6, pady=4)

        tk.Button(hdr_frame, text="📋 INV QTY", command=self._show_inv_qty_lookup,
                  bg="#1c2c42", fg="#80a0ff", font=("Segoe UI", 9, "bold"), relief="flat", padx=10, pady=4, cursor="hand2").pack(side="right", padx=6, pady=4)

        # v6.2.5: New Received Quantity lookup for main unit
        tk.Button(hdr_frame, text="📦 RECEIVED QTY", command=self._show_rec_qty_lookup,
                  bg="#d4a017", fg="black", font=("Segoe UI", 9, "bold"), relief="flat", padx=10, pady=4, cursor="hand2").pack(side="right", padx=6, pady=4)

        # Action bar moved "somewhat upper" (above the treeview)
        action_bar = tk.Frame(self, bg="#0f0f1a")
        action_bar.pack(fill="x", padx=8, pady=4)
        tk.Button(action_bar, text="💾 SAVE DIRECTLY", command=self._save_preview, bg="#005a9e", fg="white", font=("Segoe UI", 9, "bold"), relief="flat", padx=10, pady=4, cursor="hand2").pack(side="right", padx=6)
        tk.Button(action_bar, text="🗑 DELETE ROW", command=self._delete_selected_rows, bg="#5c1a1a", fg="#ff9090", font=("Segoe UI", 9, "bold"), relief="flat", padx=10, pady=4, cursor="hand2").pack(side="right", padx=6)
        tk.Button(action_bar, text="Clear All", command=self.clear_all, bg="#2d2d5e", fg="white", font=("Segoe UI", 9), relief="flat", padx=10, pady=4, cursor="hand2").pack(side="right", padx=6)

        frame = tk.Frame(self, bg="#0f0f1a")
        frame.pack(fill="both", expand=True, padx=8, pady=6)
        vsb = ttk.Scrollbar(frame, orient="vertical")
        hsb = ttk.Scrollbar(frame, orient="horizontal")
        self.tree = ttk.Treeview(frame, columns=self.COLS, show='headings', yscrollcommand=vsb.set, xscrollcommand=hsb.set, height=22)
        vsb.config(command=self.tree.yview); hsb.config(command=self.tree.xview)
        for col, w in zip(self.COLS, self.COL_W_PX):
            self.tree.heading(col, text=col); self.tree.column(col, width=w, minwidth=40, anchor='w')
        self.tree.grid(row=0, column=0, sticky="nsew"); vsb.grid(row=0, column=1, sticky="ns"); hsb.grid(row=1, column=0, sticky="ew")
        frame.grid_rowconfigure(0, weight=1); frame.grid_columnconfigure(0, weight=1)
        self.tree.bind('<Double-1>', self._on_double_click)
        self.tree.bind('<Delete>', lambda e: self._delete_selected_rows())
        self.tree.bind('<Button-3>', self._show_context_menu)
        self._row_data = {}

        # Right-click context menu
        self._ctx_menu = tk.Menu(self, tearoff=0, bg="#1e1e3a", fg="#e0e0e0", font=("Segoe UI", 9))
        self._ctx_menu.add_command(label="📋 Copy Cell", command=self._ctx_copy)
        self._ctx_menu.add_command(label="📌 Paste Cell", command=self._ctx_paste)
        self._ctx_menu.add_separator()
        self._ctx_menu.add_command(label="🗑 Clear Cell(s)", command=self._ctx_clear)
        self._ctx_menu.add_separator()
        self._ctx_menu.add_command(label="➕ Insert Blank Row", command=self._ctx_insert_row)
        self._ctx_menu.add_command(label="🗑 Delete Row(s)", command=self._delete_selected_rows)
        self._ctx_iid = None
        self._ctx_col_idx = None

        s = ttk.Style()
        s.configure("Treeview", font=("Calibri", 10), rowheight=24, background="#1e1e3a", foreground="#e0e0e0", fieldbackground="#1e1e3a")
        s.configure("Treeview.Heading", font=("Calibri", 10, "bold"), background="#203864", foreground="white")
        for tag, bg, fg in [('header','#203864','white'),('shortage_red','#ffcccc','#9c0006'),('crosspo','#2a1a00','#f0c060')]:
            self.tree.tag_configure(tag, background=bg, foreground=fg)

    def _show_mtc_qty_lookup(self):
        dlg = tk.Toplevel(self)
        dlg.title("Lookup Matched Quantity")
        dlg.configure(bg="#0d1117")
        dlg.attributes("-topmost", True)
        
        tk.Label(dlg, text="SID:", bg="#0d1117", fg="white", font=("Segoe UI", 10)).grid(row=0, column=0, padx=10, pady=(15,5), sticky='e')
        sid_ent = tk.Entry(dlg, font=("Segoe UI", 10), bg="#21262d", fg="white", insertbackground="white")
        sid_ent.grid(row=0, column=1, padx=10, pady=(15,5))

        tk.Label(dlg, text="PO:", bg="#0d1117", fg="white", font=("Segoe UI", 10)).grid(row=1, column=0, padx=10, pady=(15,5), sticky='e')
        po_ent = tk.Entry(dlg, font=("Segoe UI", 10), bg="#21262d", fg="white", insertbackground="white")
        po_ent.grid(row=1, column=1, padx=10, pady=(15,5))
        
        tk.Label(dlg, text="ASIN:", bg="#0d1117", fg="white", font=("Segoe UI", 10)).grid(row=2, column=0, padx=10, pady=(15,5), sticky='e')
        asin_ent = tk.Entry(dlg, font=("Segoe UI", 10), bg="#21262d", fg="white", insertbackground="white")
        asin_ent.grid(row=2, column=1, padx=10, pady=(15,5))
        
        # v6.0.0 Fix: Use Text widget instead of Label to show all entries
        res_frame = tk.Frame(dlg, bg="#0d1117")
        res_frame.grid(row=4, column=0, columnspan=2, pady=(10,15), padx=10)
        res_txt = tk.Text(res_frame, width=75, height=8, bg="#0d1117", fg="#3fb950", font=("Segoe UI", 10), state='disabled', bd=0, wrap="none")
        res_txt.grid(row=0, column=0, sticky="nsew")
        vsb = ttk.Scrollbar(res_frame, orient="vertical", command=res_txt.yview)
        hsb = ttk.Scrollbar(res_frame, orient="horizontal", command=res_txt.xview)
        res_txt.config(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        res_txt.tag_configure("error", foreground="#f85149")
        res_txt.tag_configure("success", foreground="#3fb950")
        
        sel = self.tree.selection()
        if sel:
            d = self._row_data.get(sel[0], {})
            sid_ent.insert(0, str(d.get('SID', '')))
            po_ent.insert(0, str(d.get('PO', '')))
            asin_ent.insert(0, str(d.get('ASIN', '')))

        def do_lookup():
            s = extract_sid(sid_ent.get())
            p = clean(po_ent.get())
            a = clean(asin_ent.get())
            
            res_txt.config(state='normal')
            res_txt.delete('1.0', 'end')
            
            engine = getattr(self._app, 'engine', None)
            if not engine:
                res_txt.insert('end', "Engine not connected.", "error")
                res_txt.config(state='disabled')
                return
            
            rows = engine.inv_p.get((s, p, a), [])
            if not rows:
                res_txt.insert('end', "No matches found.", "error")
                res_txt.config(state='disabled')
                return
            
            # v6.0.1 Fix: Deduplicate entries and show Inv, PO, ASIN, Mtc Qty
            seen = set()
            dlg.current_matches = []
            for r in rows:
                k = (r.get('mtc_inv',''), r.get('mtc_po',''), r.get('mtc_asin',''), fmt_qty(r.get('mtc_qty','')))
                if k not in seen:
                    seen.add(k)
                    dlg.current_matches.append(r)
            
            res_txt.insert('end', f"Found {len(dlg.current_matches)} unique matches:\n", "success")
            msg = "\n".join([f"Inv: {r.get('mtc_inv','')} | PO: {r.get('mtc_po','')} | ASIN: {r.get('mtc_asin','')} | Mtc Qty: {fmt_qty(r.get('mtc_qty',''))}" for r in dlg.current_matches])
            res_txt.insert('end', msg, "success")
            res_txt.config(state='disabled')
            
        def do_replace():
            sel = self.tree.selection()
            if not sel:
                messagebox.showinfo("Replace", "No destination rows selected in Preview Panel.", parent=dlg)
                return
            
            matches = getattr(dlg, 'current_matches', [])
            if not matches:
                messagebox.showinfo("Replace", "No matches found from search. Search for data first.", parent=dlg)
                return
                
            self._take_snapshot()
            replaced_count = 0
            deleted_count = 0
            
            for i, iid in enumerate(sel):
                if i >= len(matches):
                    # User requested deletion of excess selected invalid duplicate rows
                    if iid in self._row_data:
                        self._row_data.pop(iid, None)
                        try:
                            self.tree.delete(iid)
                            deleted_count += 1
                        except: pass
                    continue
                
                m = matches[i]
                if iid in self._row_data:
                    d = self._row_data[iid]
                    if 'mtc_inv' in m: d['Mtc Inv'] = str(m['mtc_inv']).strip()
                    if 'mtc_po' in m: d['Mtc PO'] = str(m['mtc_po']).strip()
                    if 'mtc_asin' in m: d['Mtc ASIN'] = str(m['mtc_asin']).strip()
                    if 'mtc_qty' in m: d['Mtc Qty'] = fmt_qty(safe_num(m['mtc_qty'])) # Enforces clean integer strings
                    
                    try:
                        vals = list(self.tree.item(iid, 'values'))
                        if 'Mtc Inv' in self.COLS: vals[self.COLS.index('Mtc Inv')] = d['Mtc Inv']
                        if 'Mtc PO' in self.COLS: vals[self.COLS.index('Mtc PO')] = d['Mtc PO']
                        if 'Mtc ASIN' in self.COLS: vals[self.COLS.index('Mtc ASIN')] = d['Mtc ASIN']
                        if 'Mtc Qty' in self.COLS: vals[self.COLS.index('Mtc Qty')] = d['Mtc Qty']
                        self.tree.item(iid, values=vals)
                        replaced_count += 1
                    except Exception: pass
            
            msg_txt = f"Replaced data correctly into {replaced_count} row(s) sequentially!"
            if deleted_count > 0:
                msg_txt += f"\nDeleted {deleted_count} excess erroneous row(s) from the selection."
            messagebox.showinfo("Success", msg_txt, parent=dlg)

        tf = tk.Frame(dlg, bg="#0d1117")
        tf.grid(row=3, column=0, columnspan=2, pady=10)
        tk.Button(tf, text="🔍 Search Data", bg="#6f42c1", fg="white", font=("Segoe UI", 10, "bold"),
                  command=do_lookup, cursor="hand2", padx=15).pack(side="left", padx=5)
        tk.Button(tf, text="🔄 Replace into Highlighted Rows", bg="#2ea043", fg="white", font=("Segoe UI", 10, "bold"),
                  command=do_replace, cursor="hand2", padx=15).pack(side="left", padx=5)
        
        dlg.update_idletasks()
        px_dlg = self.winfo_x() + (self.winfo_width() - dlg.winfo_width()) // 2
        py_dlg = self.winfo_y() + 50
        dlg.geometry(f"+{px_dlg}+{py_dlg}")
        sid_ent.focus_set()

    def _show_cp_variance(self):
        dlg = tk.Toplevel(self)
        dlg.title("CP Variance Comparison")
        dlg.configure(bg="#0d1117")
        dlg.attributes("-topmost", True)
        dlg.geometry("500x350")
        
        data = {'source': None, 'target': None, 'source_iid': None, 'target_iid': None}
        
        sf = tk.LabelFrame(dlg, text="Source ASIN (Selected Row Data)", bg="#0d1117", fg="#4a9eff", font=("Segoe UI", 10, "bold"))
        sf.pack(fill="x", padx=10, pady=5)
        s_lbl = tk.Label(sf, text="Not set", bg="#0d1117", fg="white", font=("Segoe UI", 9))
        s_lbl.pack(fill="x", pady=5)
        
        tf = tk.LabelFrame(dlg, text="Target ASIN (Selected Row Data)", bg="#0d1117", fg="#f0a500", font=("Segoe UI", 10, "bold"))
        tf.pack(fill="x", padx=10, pady=5)
        t_lbl = tk.Label(tf, text="Not set", bg="#0d1117", fg="white", font=("Segoe UI", 9))
        t_lbl.pack(fill="x", pady=5)
        
        res_lbl = tk.Label(dlg, text="", bg="#0d1117", fg="white", font=("Segoe UI", 11, "bold"))
        res_lbl.pack(fill="x", pady=10)

        def get_item_cost(sid, po, asin):
            engine = getattr(self._app, 'engine', None)
            if not engine or not hasattr(engine, 'rebni_p') or not engine.rebni_p: return None
            
            c_sid, c_po, c_asin = clean(sid), clean(po), clean(asin)
            
            # 1. Exact match (SID, PO, ASIN)
            key = (c_sid, c_po, c_asin)
            if key in engine.rebni_p:
                for r in engine.rebni_p[key]:
                    if pd.notna(r.get('item_cost')):
                        return safe_num(r.get('item_cost', 0))
                        
            # 2. Fallback: (SID, ASIN) - same shipment, cross POs / overages
            s_key = (c_sid, c_asin)
            if hasattr(engine, 'rebni_s') and s_key in engine.rebni_s:
                for r in engine.rebni_s[s_key]:
                    if pd.notna(r.get('item_cost')):
                        return safe_num(r.get('item_cost', 0))
            
            # v6.1.2 Decision: Global fallback (searching across all shipments) removed to prevent incorrect data matching.
            return None

        def eval_variance():
            if not data['source'] or not data['target']: return
            s_cp = data['source']['cp']
            t_cp = data['target']['cp']
            low, high = data['source']['low'], data['source']['high']
            if low <= t_cp <= high:
                res_lbl.config(text=f"✔ MATCH: Target CP (${t_cp:.2f})\n falls within 10% tolerance", fg="#3fb950")
            else:
                res_lbl.config(text=f"❌ MISMATCH: Target CP (${t_cp:.2f})\n violates 10% tolerance limit", fg="#f85149")
                iid = data['target_iid']
                old_rem = self.tree.set(iid, 'Remarks')
                if "Not matching within 10% cp range" not in old_rem:
                    new_rem = (old_rem + " | Not matching within 10% cp range").strip(" |")
                    self.tree.set(iid, 'Remarks', new_rem)
                    self._row_data[iid]['Remarks'] = new_rem
                    self.tree.item(iid, tags=('shortage_red',))

        def _handle_set(btn_type):
            sel = self.tree.selection()
            if not sel: return
            iid = sel[0]
            r = self._row_data.get(iid, {})
            # v6.1.2 Fix: Both Source and Target extract exactly from Main PO (Col D) and Main ASIN (Col E)
            asin = clean(r.get('ASIN', ''))
            po = clean(r.get('PO', ''))
            inv = clean(r.get('Inv no', ''))
            
            # Sub-rows match items to a parent's SID
            sid = extract_sid(r.get('SID',''))
            if getattr(self._app, 'engine', None) and hasattr(self._app.engine, 'cache_sid'):
                if inv in self._app.engine.cache_sid:
                    sid = self._app.engine.cache_sid[inv]

            cp = get_item_cost(sid, po, asin)
            if cp is None:
                # Prompt user for manual CP if it's completely missing from REBNI
                user_cp = simpledialog.askfloat("CP Not Found", 
                    f"Could not automatically find item_cost in REBNI for:\nSID: {sid}\nPO: {po}\nASIN: {asin}\n\nPlease enter the CP value manually:", 
                    parent=dlg, minvalue=0.0)
                if user_cp is not None and user_cp > 0:
                    cp = user_cp
                else:
                    return
            
            low, high = cp * 0.90, cp * 1.10
            lbl_txt = f"SID: {sid} | PO: {po} | ASIN: {asin}\nCP: ${cp:.2f} (10% Range: ${low:.2f} - ${high:.2f})"
            data[btn_type], data[f'{btn_type}_iid'] = {'cp':cp, 'low':low, 'high':high}, iid
            if btn_type == 'source': s_lbl.config(text=lbl_txt)
            else: t_lbl.config(text=lbl_txt)
            if data['source'] and data['target']: eval_variance()

        bf = tk.Frame(dlg, bg="#0d1117")
        bf.pack(fill="x", pady=10)
        tk.Button(bf, text="Set Source\nfrom Selected Row", command=lambda: _handle_set('source'), bg="#21262d", fg="#4a9eff", font=("Segoe UI", 9, "bold")).pack(side="left", expand=True, padx=5)
        tk.Button(bf, text="Set Target\nfrom Selected Row", command=lambda: _handle_set('target'), bg="#21262d", fg="#f0a500", font=("Segoe UI", 9, "bold")).pack(side="left", expand=True, padx=5)


    def _show_inv_qty_lookup(self):
        """Lookup Invoice Quantity from Invoice Search data and auto-fill the row."""
        dlg = tk.Toplevel(self)
        dlg.title("Lookup Invoice Quantity")
        dlg.configure(bg="#0d1117")
        dlg.attributes("-topmost", True)
        dlg.geometry("520x380")

        data = {'inv': '', 'asin': '', 'target_iid': None}

        # Invoice field
        tk.Label(dlg, text="Invoice No:", bg="#0d1117", fg="white",
                 font=("Segoe UI", 10)).grid(row=0, column=0, padx=10, pady=(15,5), sticky='e')
        inv_ent = tk.Entry(dlg, font=("Segoe UI", 10), bg="#21262d", fg="white",
                           insertbackground="white", width=30)
        inv_ent.grid(row=0, column=1, padx=10, pady=(15,5), sticky='w')

        # ASIN field
        tk.Label(dlg, text="ASIN:", bg="#0d1117", fg="white",
                 font=("Segoe UI", 10)).grid(row=1, column=0, padx=10, pady=5, sticky='e')
        asin_ent = tk.Entry(dlg, font=("Segoe UI", 10), bg="#21262d", fg="white",
                            insertbackground="white", width=30)
        asin_ent.grid(row=1, column=1, padx=10, pady=5, sticky='w')

        # Target row label
        tgt_lbl = tk.Label(dlg, text="Target row: (not set)", bg="#0d1117", fg="#888888",
                           font=("Segoe UI", 9), wraplength=480, justify='left')
        tgt_lbl.grid(row=2, column=0, columnspan=2, padx=10, pady=4, sticky='w')

        # Result label
        res_lbl = tk.Label(dlg, text="", bg="#0d1117", fg="#f0c060",
                           font=("Segoe UI", 12, "bold"), wraplength=480)
        res_lbl.grid(row=5, column=0, columnspan=2, pady=(10,15))

        def _set_inv_from_selection():
            sel = self.tree.selection()
            if not sel: return
            iid = sel[0]
            r = self._row_data.get(iid, {})
            inv_val = clean(r.get('Inv no', ''))  # Column B strictly
            if inv_val and not inv_val.startswith('['):
                inv_ent.delete(0, tk.END)
                inv_ent.insert(0, inv_val)
                data['target_iid'] = iid
                tgt_lbl.config(text=f"Target row: Inv={inv_val} | ASIN={clean(r.get('ASIN',''))} | PO={clean(r.get('PO',''))}", fg="#4a9eff")

        def _set_asin_from_selection():
            sel = self.tree.selection()
            if not sel: return
            iid = sel[0]
            r = self._row_data.get(iid, {})
            asin_val = clean(r.get('ASIN', ''))  # Column E strictly
            if asin_val and not asin_val.startswith('['):
                asin_ent.delete(0, tk.END)
                asin_ent.insert(0, asin_val)
                if not data['target_iid']:
                    data['target_iid'] = iid

        def do_lookup():
            ino = inv_ent.get().strip()
            asn = asin_ent.get().strip()
            if not ino or not asn:
                res_lbl.config(text="Please provide both Invoice No and ASIN.", fg="#f85149")
                return

            engine = getattr(self._app, 'engine', None)
            if not engine:
                res_lbl.config(text="Engine not connected.", fg="#f85149")
                return

            found_qty = engine._resolve_inv_qty(ino, asn, None)

            if found_qty is not None:
                qty_int = int(found_qty)
                res_lbl.config(text=f"\u2714 Inv Qty = {qty_int} units", fg="#3fb950")
                # Auto-fill Inv Qty in the target row
                iid = data.get('target_iid')
                if iid and iid in self._row_data:
                    self.tree.set(iid, 'Inv Qty', str(qty_int))
                    self._row_data[iid]['Inv Qty'] = str(qty_int)
                    res_lbl.config(text=f"\u2714 Inv Qty = {qty_int} units  (row updated!)", fg="#3fb950")
            else:
                res_lbl.config(text="Not found in Invoice Search data.", fg="#f85149")

        # Buttons row 1: Set from selection
        bf1 = tk.Frame(dlg, bg="#0d1117")
        bf1.grid(row=3, column=0, columnspan=2, pady=6)
        tk.Button(bf1, text="\u2b06 Set Invoice\nfrom Selected Row", command=_set_inv_from_selection,
                  bg="#21262d", fg="#4a9eff", font=("Segoe UI", 9, "bold"),
                  cursor="hand2", padx=10, pady=4).pack(side="left", padx=8)
        tk.Button(bf1, text="\u2b06 Set ASIN\nfrom Selected Row", command=_set_asin_from_selection,
                  bg="#21262d", fg="#f0a500", font=("Segoe UI", 9, "bold"),
                  cursor="hand2", padx=10, pady=4).pack(side="left", padx=8)

        # Button row 2: Search
        tk.Button(dlg, text="\ud83d\udd0d Search Inv Qty", bg="#238636", fg="white",
                  font=("Segoe UI", 10, "bold"), command=do_lookup,
                  cursor="hand2", padx=20).grid(row=4, column=0, columnspan=2, pady=8)

        dlg.update_idletasks()
        px_dlg = self.winfo_x() + (self.winfo_width() - dlg.winfo_width()) // 2
        py_dlg = self.winfo_y() + 50
        dlg.geometry(f"+{px_dlg}+{py_dlg}")
        inv_ent.focus_set()

    def show_correspondence(self):
        all_rows = self.get_all_rows()
        if not all_rows:
            messagebox.showinfo("No Data", "Investigation preview is empty.", parent=self)
            return
        CorrespondenceDialog(self, all_rows, app=self._app)

    def _show_rec_qty_lookup(self):
        """v6.2.5: Lookup expected Received Qty and available units for the main unit (SID/PO/ASIN)."""
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Selection Required", "Please select a row in the preview panel first.")
            return
            
        iid = sel[0]
        r = self._row_data.get(iid, {})
        asin = clean(r.get('ASIN', ''))
        po = clean(r.get('PO', ''))
        inv = clean(r.get('Inv no', ''))
        
        # Determine SID (check cache or row)
        sid = extract_sid(r.get('SID',''))
        engine = getattr(self._app, 'engine', None)
        if engine and hasattr(engine, 'cache_sid'):
            if inv in engine.cache_sid:
                sid = engine.cache_sid[inv]
        
        if not all([sid, po, asin]):
            messagebox.showerror("Data Missing", f"Selected row is missing critical data:\nSID: {sid}\nPO: {po}\nASIN: {asin}")
            return
            
        dlg = tk.Toplevel(self)
        dlg.title(f"Received Qty Lookup — {asin}")
        dlg.geometry("550x420")
        dlg.configure(bg="#0d1117")
        dlg.attributes("-topmost", True)
        
        tk.Label(dlg, text="📦  Shipment Unit Identity (REBNI Data)", bg="#16213e", fg="#f0a500", font=("Segoe UI", 11, "bold"), height=2).pack(fill="x")
        
        info_f = tk.Frame(dlg, bg="#0d1117", padx=20, pady=15)
        info_f.pack(fill="x")
        
        details = [("SID:", sid), ("PO:", po), ("ASIN:", asin)]
        for i, (l, v) in enumerate(details):
            tk.Label(info_f, text=l, bg="#0d1117", fg="#888888", font=("Segoe UI", 10)).grid(row=i, column=0, sticky='w', pady=2)
            tk.Label(info_f, text=v, bg="#0d1117", fg="white", font=("Segoe UI", 10, "bold")).grid(row=i, column=1, sticky='w', padx=10, pady=2)

        res_f = tk.LabelFrame(dlg, text="  Calculation Summary  ", bg="#0d1117", fg="#4a9eff", font=("Segoe UI", 9, "bold"), padx=15, pady=15)
        res_f.pack(fill="both", expand=True, padx=20, pady=10)

        # Execution
        if not engine or not hasattr(engine, 'rebni_p'):
            tk.Label(res_f, text="Investigation Engine not initialized.", bg="#0d1117", fg="#f85149").pack()
        else:
            rows = engine.rebni_p.get((sid, po, asin), [])
            if not rows:
                tk.Label(res_f, text="No REBNI data found for this combination.", bg="#0d1117", fg="#f85149").pack()
            else:
                r0 = rows[0]
                rec_post = safe_num(r0.get('qty_received_postadj', 0))
                rec_unp = safe_num(r0.get('quantity_unpacked', 0))
                avail = safe_num(r0.get('rebni_available', 0))
                
                # Highlight logic
                tk.Label(res_f, text=f"Received (Post Adj):", bg="#0d1117", fg="#cccccc", font=("Segoe UI", 10)).grid(row=0, column=0, sticky='w')
                tk.Label(res_f, text=f"{int(rec_post)} units", bg="#0d1117", fg="#3fb950", font=("Segoe UI", 12, "bold")).grid(row=0, column=1, sticky='w', padx=10)
                
                tk.Label(res_f, text=f"Quantity Unpacked:", bg="#0d1117", fg="#cccccc", font=("Segoe UI", 10)).grid(row=1, column=0, sticky='w', pady=10)
                tk.Label(res_f, text=f"{int(rec_unp)} units", bg="#0d1117", fg="#58a6ff", font=("Segoe UI", 12, "bold")).grid(row=1, column=1, sticky='w', padx=10)
                
                tk.Label(res_f, text=f"Quantity Adjusted:", bg="#0d1117", fg="#cccccc", font=("Segoe UI", 10)).grid(row=2, column=0, sticky='w', pady=5)
                tk.Label(res_f, text=f"{int(safe_num(r0.get('quantity_adjusted', 0)))} units", bg="#0d1117", fg="#ff4d4d", font=("Segoe UI", 12, "bold")).grid(row=2, column=1, sticky='w', padx=10)
                
                tk.Label(res_f, text=f"REBNI Available:", bg="#0d1117", fg="#cccccc", font=("Segoe UI", 10)).grid(row=3, column=0, sticky='w')
                tk.Label(res_f, text=f"{int(avail)} units", bg="#0d1117", fg="#f0a500", font=("Segoe UI", 12, "bold")).grid(row=3, column=1, sticky='w', padx=10)
                
                tk.Label(res_f, text="\u24D8  Verified using REBNI strictly.", bg="#0d1117", fg="#6e7681", font=("Segoe UI", 8, "italic")).grid(row=4, column=0, columnspan=2, pady=(20,0))

        tk.Button(dlg, text="CLOSE", command=dlg.destroy, bg="#333", fg="white", font=("Segoe UI", 9, "bold"), relief="flat", padx=20, pady=8).pack(pady=15)
        
        dlg.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - dlg.winfo_width()) // 2
        py = self.winfo_y() + (self.winfo_height() - dlg.winfo_height()) // 2
        dlg.geometry(f"+{px}+{py}")

    def add_header_row(self, label=""):
        vals = list(self.COLS); vals[4] = f"── {label} ──" if label else "Header"
        iid = self.tree.insert('', 'end', values=vals, tags=('header',))
        # v5.9.3: Explicitly tag this as a UI header so it is ignored in Excel reports
        self._row_data[iid] = dict(zip(self.COLS, vals))
        self._row_data[iid]['is_ui_header'] = True

    def add_row(self, rd):
        try:
            if not self.winfo_exists(): return
        except: return
        vals = [rd.get('barcode',''), rd.get('invoice',''), rd.get('sid',''), rd.get('po',''), rd.get('asin',''), 
                rd.get('inv_qty',''), rd.get('rec_qty',''), rd.get('mtc_qty',''), rd.get('mtc_inv',''),
                rd.get('mtc_asin',''), rd.get('mtc_po',''), rd.get('fc_id',''), rd.get('remarks',''), rd.get('date',''), rd.get('cp_status','')]
        iq, rq = safe_num(rd.get('inv_qty')), safe_num(rd.get('rec_qty'))
        tag = 'shortage_red' if (iq > 0 and rq < iq) else 'crosspo' if 'cross po' in str(rd.get('remarks','')).lower() else ''
        try:
            iid = self.tree.insert('', 'end', values=vals, tags=(tag,))
            self._row_data[iid] = dict(zip(self.COLS, vals)); self._row_data[iid]['_rd'] = rd; self.tree.see(iid)
        except Exception: pass  # Treeview destroyed

    def get_all_rows(self):
        KEY = {'Barcode':'barcode', 'Inv no':'invoice', 'SID':'sid', 'PO':'po', 'ASIN':'asin', 'Inv Qty':'inv_qty', 'Rec Qty':'rec_qty', 'Mtc Qty':'mtc_qty', 'Mtc Inv':'mtc_inv', 'Mtc ASIN':'mtc_asin', 'Mtc PO':'mtc_po', 'FC':'fc_id', 'Remarks':'remarks', 'Date':'date', 'CP':'cp_status'}
        rows = []
        for iid in self.tree.get_children():
            d = self._row_data.get(iid, {}); rd = d.get('_rd', {}).copy()
            for col in self.COLS: rd[KEY[col]] = d.get(col, '')
            if d.get('is_ui_header'): rd['is_ui_header'] = True
            rows.append(rd)
        return rows

    def clear_all(self):
        for iid in self.tree.get_children(): self.tree.delete(iid)
        self._row_data.clear()

    def _on_double_click(self, event):
        iid = self.tree.identify_row(event.y); col = self.tree.identify_column(event.x)
        if not iid or not col: return
        col_idx = int(col.replace('#',''))-1; col_name = self.COLS[col_idx]
        bbox=self.tree.bbox(iid, col); x,y,w,h = bbox
        cur = self._row_data.get(iid,{}).get(col_name,''); ev=tk.StringVar(value=str(cur))
        e = tk.Entry(self.tree, textvariable=ev); e.place(x=x,y=y,width=w,height=h); e.focus_force()
        def save(evt=None):
            nv=ev.get(); self._row_data[iid][col_name]=nv
            vals=list(self.tree.item(iid,'values')); vals[col_idx]=nv
            self.tree.item(iid, values=vals); e.destroy()
        e.bind('<Return>', save); e.bind('<FocusOut>', save)

    def _take_snapshot(self):
        """v6.0.0: Capture a full snapshot of current tree state for undo."""
        self._redo_stack.clear() # clear redo stack on new action
        snap = self._capture_state()
        self._undo_stack.append(snap)
        if len(self._undo_stack) > 10:
            self._undo_stack.pop(0)  # cap at 10

    def _capture_state(self):
        snap = []
        for iid in self.tree.get_children():
            vals = list(self.tree.item(iid, 'values'))
            tags = list(self.tree.item(iid, 'tags'))
            data_copy = {}
            if iid in self._row_data:
                data_copy = {k: v for k, v in self._row_data[iid].items() if k != '_rd'}
                if '_rd' in self._row_data[iid]:
                    data_copy['_rd'] = dict(self._row_data[iid]['_rd'])
            snap.append({'vals': vals, 'tags': tags, 'data': data_copy})
        return snap

    def _restore_state(self, snap):
        for iid in self.tree.get_children():
            self.tree.delete(iid)
        self._row_data.clear()
        for entry in snap:
            iid = self.tree.insert('', 'end', values=entry['vals'], tags=tuple(entry['tags']))
            self._row_data[iid] = entry['data']

    def _undo(self):
        """v6.0.0: Restore the last snapshot from the undo stack."""
        if not self._undo_stack:
            messagebox.showinfo("Undo", "Nothing to undo.", parent=self)
            return
        self._redo_stack.append(self._capture_state())
        snap = self._undo_stack.pop()
        self._restore_state(snap)
        messagebox.showinfo("Undo", f"Restored to previous state. ({len(self._undo_stack)} undo steps remaining)", parent=self)

    def _redo(self):
        """v6.0.0: Redo an undone action."""
        if not self._redo_stack:
            messagebox.showinfo("Redo", "Nothing to redo.", parent=self)
            return
        self._undo_stack.append(self._capture_state())
        snap = self._redo_stack.pop()
        self._restore_state(snap)
        messagebox.showinfo("Redo", f"Redo successful. ({len(self._redo_stack)} redo steps remaining)", parent=self)

    def _save_preview(self):
        """v6.0.0: Direct save from preview panel."""
        if not getattr(self._app, "ticket_id", None): return
        t, ts = self._app.ticket_id.get().strip().replace(' ','_'), datetime.now().strftime('%Y%m%d_%H%M%S')
        o = f"MFI_{t}_{ts}.xlsx" if t else f"MFI_Investigation_{ts}.xlsx"
        
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=o, parent=self)
        if not path: return
        
        try:
            ar = self.get_all_rows()
            bl, cur = [], []
            for r in ar:
                if r.get('is_ui_header'):
                    if cur: bl.append(cur)
                    cur = []
                    continue
                if str(r.get('barcode')).strip() == "Barcode": continue
                cur.append(r)
            if cur: bl.append(cur)
            
            app = self._app
            rebni_data = app.engine.collected_rebni if (app and hasattr(app, 'engine')) else None
            write_excel(bl, path, rebni_summary_data=rebni_data)
            messagebox.showinfo("Saved", f"Complete investigation report saved to:\n{path}", parent=self)
        except Exception as e:
            messagebox.showerror("Save Error", str(e), parent=self)

    def _delete_selected_rows(self):
        """v6.0.0: Delete selected rows from the preview panel."""
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Delete", "No rows selected. Click a row first.", parent=self)
            return
        self._take_snapshot()  # allow undo
        for iid in sel:
            self._row_data.pop(iid, None)
            self.tree.delete(iid)

    def _show_context_menu(self, event):
        """v6.0.0: Show right-click context menu on the Treeview."""
        iid = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if iid:
            self.tree.selection_set(iid)
            self._ctx_iid = iid
            self._ctx_col_idx = int(col.replace('#', '')) - 1 if col else None
        try:
            self._ctx_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self._ctx_menu.grab_release()

    def _ctx_copy(self):
        """v6.0.0: Copy clicked cell value to clipboard."""
        if self._ctx_iid is None or self._ctx_col_idx is None: return
        col_name = self.COLS[self._ctx_col_idx]
        val = str(self._row_data.get(self._ctx_iid, {}).get(col_name, ''))
        self.clipboard_clear()
        self.clipboard_append(val)

    def _ctx_paste(self):
        """v6.0.0: Paste clipboard content into clicked cell."""
        if self._ctx_iid is None or self._ctx_col_idx is None: return
        try:
            val = self.clipboard_get()
        except tk.TclError:
            return
        self._take_snapshot()
        col_name = self.COLS[self._ctx_col_idx]
        sel = self.tree.selection()
        if self._ctx_iid not in sel: sel = (self._ctx_iid,) # if they didn't highlight multiple, just do the clicked one
        
        for iid in sel:
            if iid in self._row_data:
                self._row_data[iid][col_name] = val
                vals = list(self.tree.item(iid, 'values'))
                vals[self._ctx_col_idx] = val
                self.tree.item(iid, values=vals)

    def _ctx_clear(self):
        """v6.2.5: Clear the targeted column across all selected rows."""
        if self._ctx_iid is None or self._ctx_col_idx is None: return
        self._take_snapshot()
        col_name = self.COLS[self._ctx_col_idx]
        sel = self.tree.selection()
        if self._ctx_iid not in sel: sel = (self._ctx_iid,)
        
        for iid in sel:
            if iid in self._row_data:
                self._row_data[iid][col_name] = ""
                vals = list(self.tree.item(iid, 'values'))
                vals[self._ctx_col_idx] = ""
                self.tree.item(iid, values=vals)

    def _ctx_insert_row(self):
        """v6.2.5: Insert an entirely blank row immediately below the selected row."""
        if self._ctx_iid is None: return
        self._take_snapshot()
        idx = self.tree.index(self._ctx_iid)
        
        vals = [''] * len(self.COLS)
        tags = self.tree.item(self._ctx_iid, 'tags')
        
        new_iid = self.tree.insert('', idx + 1, values=vals, tags=tags)
        self._row_data[new_iid] = dict(zip(self.COLS, vals))
        self._row_data[new_iid]['_rd'] = {'_blank': True} # Safe fallback flag

    def confirm_edits(self):
        self._take_snapshot()  # v6.0.0: snapshot before applying edits
        app = self._app
        if not app or not hasattr(app, 'engine'): return
        
        KEY_MAP = {'Barcode':'barcode', 'Inv no':'invoice', 'SID':'sid', 'PO':'po', 'ASIN':'asin', 'Inv Qty':'inv_qty', 'Rec Qty':'rec_qty', 'Mtc Qty':'mtc_qty', 'Mtc Inv':'mtc_inv', 'Mtc ASIN':'mtc_asin', 'Mtc PO':'mtc_po', 'FC':'fc_id', 'Remarks':'remarks', 'Date':'date', 'CP':'cp_status'}

        for iid in self.tree.get_children():
            d = self._row_data.get(iid, {}); rd = d.get('_rd', {})
            if not rd or d.get('is_ui_header'): continue
            
            mtc_inv = str(d.get('Mtc Inv', '') or rd.get('mtc_inv', '')).strip()
            mtc_po  = str(d.get('Mtc PO', '') or rd.get('mtc_po', '')).strip()
            mtc_asin= str(d.get('Mtc ASIN', '') or rd.get('mtc_asin', '')).strip()
            
            iq = safe_num(d.get('Inv Qty', '')); rq = safe_num(d.get('Rec Qty', ''))
            mq = safe_num(d.get('Mtc Qty', ''))
            
            # v6.2.5: Auto-recalculate Remarks if Inv Qty and Rec Qty create a math discrepancy
            if iq and rq and iq != rq:
                old_rem = str(d.get('Remarks', ''))
                # Only overwrite standard calculation remarks if they didn't explicitly write a custom note
                if 'short' in old_rem.lower() or 'overage' in old_rem.lower() or 'units matched' in old_rem.lower() or not old_rem.strip():
                    diff = abs(iq - rq)
                    kw = "short" if iq > rq else "overage"
                    new_rem = f"Found {int(diff)} units {kw} locally (Inv: {int(iq)}, Rec: {int(rq)})"
                    
                    if old_rem != new_rem:
                        d['Remarks'] = new_rem
                        try:
                            idx = self.COLS.index('Remarks')
                            vals = list(self.tree.item(iid, 'values'))
                            vals[idx] = new_rem
                            self.tree.item(iid, values=vals)
                        except: pass
            
            if mtc_inv: 
                k = (clean(mtc_inv), clean(mtc_asin), clean(mtc_po))
                
                # Expand overrides to support ANY cell edit syncing globally
                ext_out = {}
                for col_name in self.COLS:
                    if col_name in d: ext_out[KEY_MAP[col_name]] = d[col_name]
                
                # Make sure strict numerics remain precise
                ext_out['inv_qty'] = iq
                ext_out['mtc_qty'] = mq
                ext_out['rec_qty'] = rq
                
                app.engine.user_overrides[k] = ext_out
        
        # v6.2.5: Synchronize edits across ALL active dialogs globally
        updated_any = False
        if getattr(app, 'active_manual_dlg', None) and app.active_manual_dlg.winfo_exists():
            app.active_manual_dlg.refresh_from_engine(app.engine)
            updated_any = True
            
        if getattr(app, 'active_pending_dlg', None) and app.active_pending_dlg.winfo_exists():
            app.active_pending_dlg.refresh_from_engine(app.engine)
            updated_any = True

        if updated_any:
            messagebox.showinfo("Success", "Manual edits synchronized across all active panels.", parent=self)
        else:
            messagebox.showinfo("Success", "Manual edits saved to engine cache.", parent=self)


# ═══════════════════════════════════════════════════════════
#  DATA LOADERS
# ═══════════════════════════════════════════════════════════

def _load_file(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == '.csv':
        try:   return pd.read_csv(path, dtype=str, encoding='utf-8')
        except: return pd.read_csv(path, dtype=str, encoding='latin-1')
    else:
        try:
            return pd.read_excel(path, header=0, dtype=str, engine='calamine')
        except Exception:
            return pd.read_excel(path, header=0, dtype=str)

def load_claims(path): return _load_file(path)

COLUMN_ALIASES = {
    'Barcode': [
        'barcode', 'bar code', 'bar_code', 'upc', 'ean',
        'scan code', 'item code', 'carton barcode', 'pkg barcode',
    ],
    'Invoice': [
        'inv no', 'inv_no', 'invoice_no', 'invoice no', 'invoice number',
        'invoice_number', 'invoice', 'inv num', 'inv number',
        'inv#', 'invoice#', 'invc no', 'invc number', 'invc',
        'bill no', 'bill number',
    ],
    'SID': [
        'sid', 'shipment id', 'shipment_id', 'shipment no',
        'shipment number', 'ship id', 'fba shipment id',
        'inbound shipment', 'shipment', 'inbound sid',
    ],
    'PO': [
        'po', 'po_no', 'p.o.', 'po no', 'po number', 'po#',
        'purchase order', 'purchase_order', 'purchase order no',
        'purchase order number', 'po id', 'order no', 'order number',
        'purch order', 'header po', 'line po', 'po num',
        'invoice_lineitem_po',
    ],
    'ASIN': [
        'asin', 'po_asin', 'amazon asin', 'amazon product id',
        'product id', 'asin no', 'item asin', 'amazon id',
    ],
    'InvQty': [
        'inv qty', 'inv_qty', 'invoice qty', 'invoice quantity',
        'invoiced qty', 'invoiced quantity', 'quantity invoiced',
        'quantity_invoiced', 'billed qty', 'billed quantity',
        'total qty', 'total quantity', 'item qty',
        'po_ordered_quantity',
    ],
    'PQV': [
        'pqv', 'pqv qty', 'missing qty', 'missing quantity',
        'missing_qty', 'shortage', 'short qty', 'short quantity',
        'claim qty', 'claimed qty', 'dispute qty',
        'pending qty', 'outstanding qty', 'difference qty',
    ],
    'CP': [
        'cp', 'cost price', 'cost_price', 'unit cost', 'unit_cost',
        'item cost', 'item_cost', 'rate', 'amount', 'unit price',
    ],
}

def detect_claim_cols(df):
    actual_cols = list(df.columns)
    lower_map   = {c.lower().strip(): c for c in actual_cols}
    mapping     = {}
    corrections = []

    for field, aliases in COLUMN_ALIASES.items():
        found = None
        for alias in aliases:
            if alias in lower_map:
                found = lower_map[alias]; break
        if not found:
            for alias in aliases:
                for col in actual_cols:
                    if alias in col.lower() or col.lower() in alias:
                        found = col; break
                if found: break
        if found:
            canonical = aliases[0]
            if found.lower().strip() != canonical:
                corrections.append((field, found, canonical))
            mapping[field] = found
    return mapping, corrections

def load_rebni(path):
    df = _load_file(path)
    names = ['vendor_code', 'po', 'asin', 'shipment_id', 'received_datetime',
             'warehouse_id', 'item_cost', 'quantity_unpacked', 'quantity_adjusted',
             'qty_received_postadj', 'quantity_matched', 'rebni_available',
             'cnt_invoice_matched', 'matched_invoice_numbers']
    if len(df.columns) > len(names):
        names += [f'ext_{i}' for i in range(len(df.columns) - len(names))]
    df.columns = names[:len(df.columns)]
    return df

def load_invoice_search(path):
    df = _load_file(path)
    names = ['vendor_code', 'purchase_order_id', 'asin', 'invoice_number', 'invoice_date',
             'invoice_item_status', 'quantity_invoiced', 'quantity_matched_total',
             'no_of_shipments', 'shipment_id', 'shipmentwise_matched_qty',
             'matched_po', 'matched_asin']
    df.columns = names[:len(df.columns)]
    # v6.2.5 Deduplication: Strictly remove redundant invoice-asin-po-sid records
    df = df.drop_duplicates(subset=['purchase_order_id', 'asin', 'invoice_number', 'shipment_id']).reset_index(drop=True)
    return df

def load_shipment_master(path):
    df = _load_file(path)
    return df

# ═══════════════════════════════════════════════════════════
#  INDEX BUILDERS
# ═══════════════════════════════════════════════════════════

def build_shipment_index(df):
    """v6.2.5: Builds an index from (SID, PO) to (Barcode, Invoice) for auto-population."""
    idx = {}
    mapping, _ = detect_claim_cols(df)
    sid_col = mapping.get('SID')
    bc_col  = mapping.get('Barcode')
    po_col  = mapping.get('PO')
    inv_col = mapping.get('Invoice')
    
    for row in df.to_dict('records'):
        sid = extract_sid(clean(row.get(sid_col, ''))) if sid_col else ""
        po  = clean(row.get(po_col, '')) if po_col else ""
        bc  = clean(row.get(bc_col, '')) if bc_col else ""
        inv = clean(row.get(inv_col, '')) if inv_col else ""
        
        if sid and po:
            idx[(sid, po)] = {'bc': bc, 'inv': inv}
        if sid and bc and sid not in idx:
            idx[sid] = bc # Legacy SID -> Barcode mapping
    return idx

def build_rebni_index(df):
    p, s, fb, sid_p = {}, {}, {}, {}
    for row in df.to_dict('records'):
        sid  = extract_sid(clean(row.get('shipment_id', '')))
        po   = clean(row.get('po', ''))
        asin = clean(row.get('asin', ''))
        if not sid or not asin: continue
        p.setdefault((sid, po, asin), []).append(row)
        s.setdefault((po, asin), []).append(row)
        sid_p.setdefault(sid, []).append(row)
        for inv in split_comma(row.get('matched_invoice_numbers', '')):
            if inv: fb.setdefault((sid, po, inv), []).append(row)
    return p, s, fb, sid_p

def build_invoice_index(df):
    idx, fb, iam = {}, {}, {}
    for row in df.to_dict('records'):
        sids  = split_comma(row.get('shipment_id', ''))
        pos   = split_comma(row.get('matched_po', ''))
        asins = split_comma(row.get('matched_asin', ''))
        qtys  = split_comma(row.get('shipmentwise_matched_qty', ''))
        
        inv_no   = clean(row.get('invoice_number', ''))
        mtc_asin = clean(row.get('asin', ''))
        inv_qty  = safe_num(row.get('quantity_invoiced', '0'))
        
        # v6.1.1 Fix: Save exact target ASIN quantity; do NOT cross-pollute matching ASINs.
        if inv_no and mtc_asin:
            iam[(inv_no, mtc_asin.upper())] = inv_qty

        for i in range(max(len(sids), len(pos), len(asins), len(qtys))):
            s_frag   = extract_sid(sids[i] if i < len(sids) else "")
            p_val    = pos[i]   if i < len(pos)   else ""
            a_val    = asins[i] if i < len(asins) else ""
            q_val    = safe_num(qtys[i] if i < len(qtys) else "0")
            
            mtc_po   = clean(row.get('purchase_order_id', ''))
            if not s_frag or not p_val or not a_val: continue
            entry = {'mtc_inv':  inv_no,
                     'mtc_po':   mtc_po,
                     'mtc_asin': mtc_asin,
                     'inv_qty':  inv_qty,
                     'mtc_qty':  q_val,
                     'date':     clean(row.get('invoice_date', ''))}
            idx.setdefault((s_frag, p_val, a_val), []).append(entry)
            if inv_no: fb.setdefault((s_frag, p_val, inv_no), []).append(entry)
    return idx, fb, iam


# ═══════════════════════════════════════════════════════════
#  INVESTIGATION ENGINE
# ═══════════════════════════════════════════════════════════

class InvestigationEngine:
    MAX_DEPTH = 10

    def __init__(self, rp, rs, rfb, rsid, ip, ifb, iam, sid_cb=None, ship_idx=None):
        self.rebni_p    = rp
        self.rebni_s    = rs
        self.rebni_fb   = rfb
        self.rebni_sid  = rsid # v5.9.3: Shipment ID based REBNI index
        self.inv_p      = ip
        self.inv_fb     = ifb
        self.inv_iam    = iam
        self.sid_cb     = sid_cb
        self.shipment_index = ship_idx or {} # v6.2.5: Shipment Master index for barcodes
        self.stop_requested = False
        self.pause_requested = False
        self.ticket_type = "PDTT"
        self.cache_sid  = {}
        self.cache_bc   = {}
        self.lock_save = False
        self._cloud_busy = {"REBNI": False, "Invoice": False}
        self._cloud_progress = {"REBNI": 0, "Invoice": 0}
        self.loop_cache = {}
        self.user_overrides = {}
        self.collected_rebni = {} # (sid, po, asin) -> [rows]
        self.all_found_matches = [] # v6.1.1: Global net for pending invoices
        self.asin_pending_matches = [] # v6.2.5: Per-ASIN match isolation
        self.cross_po_vault = {} # v6.2.6: Shipment-level Cross PO repository

    def _rebni_lookup(self, sid, po, asin, inv_no=None):
        rows = self.rebni_p.get((sid, po, asin), [])
        if not rows and inv_no:
            rows = self.rebni_fb.get((sid, po, inv_no), [])
        return rows

    def _inv_lookup(self, sid, po, asin, inv_no=None):
        m = self.inv_p.get((sid, po, asin), [])
        if not m and inv_no:
            m = self.inv_fb.get((sid, po, inv_no), [])
        return m

    def _find_sid(self, po, asin, inv_no):
        rows = self.rebni_s.get((clean(po), clean(asin)), [])
        for r in rows:
            if inv_no in split_comma(r.get('matched_invoice_numbers', '')):
                return extract_sid(r['shipment_id'])
        return extract_sid(rows[0]['shipment_id']) if rows else None

    def _resolve_inv_qty(self, inv_no, asin, fallback_qty):
        qty = self.inv_iam.get((clean(inv_no), clean(asin).upper()))
        if qty is not None and qty > 0:
            return qty
        base = strip_scr(inv_no)
        if base != clean(inv_no):
            qty = self.inv_iam.get((base, clean(asin).upper()))
            if qty is not None and qty > 0:
                return qty
        return fallback_qty

    def _get_shipment_rebni(self, sid, po):
        total = 0.0
        for (s, p, a), rows in self.rebni_p.items():
            if s == sid and p == po:
                for r in rows:
                    total += safe_num(r.get('rebni_available', 0))
        return total

    def get_cp(self, sid, po, asin):
        rows = self.rebni_p.get((extract_sid(sid), clean(po), clean(asin)), [])
        for r in rows:
            cp = safe_num(r.get('item_cost', 0))
            if cp > 0: return cp
        for (s, p, a), rlist in self.rebni_p.items():
            if p == clean(po) and a == clean(asin):
                for r in rlist:
                    cp = safe_num(r.get('item_cost', 0))
                    if cp > 0: return cp
        return 0.0

    def compare_cp(self, c_sid, c_po, c_asin, m_sid, m_po, m_asin, depth):
        c_cp = self.get_cp(c_sid, c_po, c_asin)
        m_cp = self.get_cp(m_sid, m_po, m_asin)
        c_lbl = "Claiming CP" if depth == 0 else "Parent ASIN CP"
        m_lbl = "Matched CP"  if depth == 0 else "Sub-Matched CP"
        if c_cp <= 0 and m_cp <= 0: return ""
        if c_cp <= 0: return f"{m_lbl}: {m_cp:.2f} | {c_lbl}: N/A"
        if m_cp <= 0: return f"{c_lbl}: {c_cp:.2f} | {m_lbl}: N/A"
        low, high = c_cp * 0.90, c_cp * 1.10
        if low <= m_cp <= high:
            return f"Within 10% CP | {c_lbl}: {c_cp:.2f}, {m_lbl}: {m_cp:.2f} (range: {low:.2f}-{high:.2f})"
        return f"NOT within 10% CP | {c_lbl}: {c_cp:.2f}, {m_lbl}: {m_cp:.2f} (range: {low:.2f}-{high:.2f})"

    def detect_cross_po(self, sid, po, asin):
        candidates = []
        seen_po = set()
        rec_at_cur = sum(safe_num(r.get('quantity_unpacked', 0)) for r in self.rebni_p.get((sid, po, asin), []))
        for (s, p, a), rows in self.rebni_p.items():
            if s != sid or a != asin or p == po or p in seen_po: continue
            for r in rows:
                rec = safe_num(r.get('quantity_unpacked', 0))
                if rec <= 0: continue
                seen_po.add(p)
                im = self.inv_p.get((sid, p, asin), [])
                iq = safe_num(im[0].get('inv_qty', 0)) if im else 0.0
                if rec_at_cur == 0 and iq == 0:
                    tp = "Case 2 — ASIN not invoiced at this PO, but received"
                elif rec > iq and iq > 0:
                    tp = "Case 3 — Rec qty > Inv qty (overage in cross PO)"
                else:
                    tp = "Case 1 — Rec=0 at current PO, units received here"
                candidates.append({'po': p, 'asin': asin, 'sid': sid, 'inv_qty': fmt_qty(iq), 'rec_qty': rec, 'cross_type': tp, 'date': clean(r.get('received_datetime', '')), 'fc_id': clean(r.get('warehouse_id', ''))})
        return candidates

    def _make_row(self, b, i, s, p, a, iq, rq, mq, mi, rem, d, depth, rtype='dominant', cp_status='', mtc_asin='', mtc_po='', fc_id=''):
        return {
            'barcode': b, 'invoice': i, 'sid': extract_sid(s) if s else '', 'po': p, 'asin': a,
            'inv_qty': fmt_qty(iq), 'rec_qty': fmt_qty(rq), 'mtc_qty': fmt_qty(mq), 'mtc_inv': mi,
            'mtc_asin': mtc_asin, 'mtc_po': mtc_po, 'fc_id': fc_id, 'remarks': rem, 'date': d, 'depth': depth,
            'type': rtype, 'cp_status': cp_status,
        }

    def _build_level_logic(self, barcode, inv_no, sid, po, asin, iqty, rem_pqv, depth, is_claiming, is_manual=False, cross_po_indicator_only=False, initial_cp=0.0):
        while self.pause_requested and not self.stop_requested:
            import time; time.sleep(0.5)
        sid_frag = extract_sid(sid)
        
        # v6.2.5: Auto-populate barcode and invoice if missing using (SID, PO) composite index
        ship_data = self.shipment_index.get((sid_frag, clean(po)), {}) if isinstance(self.shipment_index, dict) else {}
        
        if not barcode or str(barcode).strip() == "":
            # Priority: PO-specific BC -> Legacy SID-level BC fallback
            barcode = ship_data.get('bc') or self.shipment_index.get(sid_frag, "")
            
        # REMASH TT specific: Auto-populate Invoice number from Shipment Master if missing
        if self.ticket_type == "REMASH" and (not inv_no or str(inv_no).strip() == ""):
            inv_no = ship_data.get('inv', inv_no)

        rebni_rows = self.rebni_p.get((sid_frag, clean(po), clean(asin)), [])
        fc_id = clean(rebni_rows[0].get('warehouse_id', '')) if rebni_rows else ""
        
        # v5.9.3 Shipment-Wide Collection: Capture ALL REBNI in this SID
        if sid_frag in self.rebni_sid:
            for r in self.rebni_sid[sid_frag]:
                if safe_num(r.get('rebni_available', 0)) > 0:
                    rk = (sid_frag, clean(r.get('po','')), clean(r.get('asin','')))
                    if rk not in self.collected_rebni: self.collected_rebni[rk] = []
                    if r not in self.collected_rebni[rk]: self.collected_rebni[rk].append(r)

        rec_qty = sum(safe_num(r.get('quantity_unpacked', 0)) for r in rebni_rows)
        r_avail = sum(safe_num(r.get('rebni_available', 0)) for r in rebni_rows)
        ex_adj = sum(safe_num(r.get('quantity_adjusted', 0)) for r in rebni_rows)
        rec_date = clean(rebni_rows[0].get('received_datetime', '')) if rebni_rows else ""
        cur_cp = initial_cp if initial_cp > 0 else self.get_cp(sid_frag, po, asin)
        cp_disp = f"{cur_cp:.2f}" if cur_cp > 0 else ""
        shortage = max(0.0, safe_num(iqty) - rec_qty)
        acc_at_lvl = shortage + r_avail + ex_adj
        remarks = ""
        if is_claiming or rec_qty < safe_num(iqty) or r_avail > 0 or ex_adj > 0:
            if ex_adj > 0 and shortage > 0:
                remarks = f"Found {int(ex_adj)} units of EX adjustments and {int(shortage)} units of shortage (Inv:{int(iqty)} Rec:{int(rec_qty)})"
            else:
                parts = []
                if shortage > 0: parts.append(f"Inv Qty:{int(iqty)}.Received Qty:{int(rec_qty)}- Shortage of {int(shortage)} Units")
                if r_avail > 0: parts.append(f"REBNI Available = {int(r_avail)} units at {'claiming' if is_claiming else 'matching'} level — Suggest TSP to utilize")
                if ex_adj > 0: parts.append(f"Found {int(ex_adj)} number of X adjustments")
                remarks = " | ".join(parts) if parts else "SR" if depth > 0 else ""

        if shortage >= rem_pqv > 0 and not remarks:
            rem = f"Phase 1 Direct Shortage: {int(shortage)} units short received directly"
            if acc_at_lvl > shortage: rem += f" (Total Accounted: {int(acc_at_lvl)} incl. REBNI/EX)"
            m_inv_label = "Self Matching" if is_claiming else "Short Received"
            main_row = self._make_row(barcode, inv_no, sid, po, asin, iqty, rec_qty, rec_qty, m_inv_label, rem, rec_date, depth, cp_status=cp_disp, fc_id=fc_id)
            res_rows = [main_row]
            shp_rebni = self._get_shipment_rebni(sid_frag, clean(po))
            if shp_rebni > 0:
                res_rows.append(self._make_row('[REBNI-SHP]', inv_no, sid_frag, po, asin, '', '', shp_rebni, '', f"Shipment-level REBNI = {int(shp_rebni)} units available across all ASINs in this shipment — Suggest TSP to utilize", rec_date, depth, rtype='rebni_shipment', fc_id=fc_id))
            if cross_po_indicator_only and self.ticket_type != "REMASH":
                for c in self.detect_cross_po(sid_frag, clean(po), clean(asin)):
                    res_rows.append(self._make_row('[CROSS PO?]', '—', c['sid'], c['po'], c['asin'], c['inv_qty'], c['rec_qty'], '', '', f"Phase 4: Cross PO candidate detected | {c['cross_type']} | Rec={fmt_qty(c['rec_qty'])} units | PENDING USER CONFIRMATION", c['date'], depth, rtype='crosspo', fc_id=c.get('fc_id', '')))
            else:
                # v6.2.5: For REMASH, always automate Cross PO traversal even in manual/indicator modes
                res_rows.extend(self._build_cross_po_rows(sid_frag, clean(po), clean(asin), depth))
            return res_rows, [], rec_qty, acc_at_lvl, 0.0, ex_adj

        if 'REBNI Available' in remarks or remarks == 'SR':
            rows = [self._make_row(barcode, inv_no, sid, po, asin, iqty, rec_qty, "", "", remarks, rec_date, depth, cp_status=cp_disp, fc_id=fc_id)]
            if not (cross_po_indicator_only and self.ticket_type != "REMASH"): 
                rows.extend(self._build_cross_po_rows(sid_frag, clean(po), clean(asin), depth))
            else:
                for c in self.detect_cross_po(sid_frag, clean(po), clean(asin)):
                    rows.append(self._make_row('[CROSS PO?]', '—', c['sid'], c['po'], c['asin'], c['inv_qty'], c['rec_qty'], '', '', f"Phase 4: Cross PO candidate detected | {c['cross_type']} | Rec={fmt_qty(c['rec_qty'])} units | PENDING USER CONFIRMATION", c['date'], depth, rtype='crosspo', fc_id=c.get('fc_id', '')))
            return rows, [], rec_qty, acc_at_lvl, max(0.0, rem_pqv - acc_at_lvl), ex_adj

        raw = self.inv_p.get((sid_frag, clean(po), clean(asin)), [])
        sorted_m = sorted(raw, key=lambda x: safe_num(x['mtc_qty']), reverse=True)
        if self.user_overrides:
            patched = []
            for e in sorted_m:
                ovr = self.user_overrides.get(clean(e.get('mtc_inv', '')), {})
                if ovr:
                    e = dict(e)
                    if 'inv_qty' in ovr: e['inv_qty'] = ovr['inv_qty']
                    if 'mtc_qty' in ovr: e['mtc_qty'] = ovr['mtc_qty']
                patched.append(e)
            sorted_m = sorted(patched, key=lambda x: safe_num(x['mtc_qty']), reverse=True)

        m_inv, m_qty = "", ""
        if sorted_m:
            top = sorted_m[0]
            if top['mtc_inv'] == clean(inv_no): m_inv, m_qty = "Self Matching", fmt_qty(rec_qty)
            else: m_inv, m_qty = top['mtc_inv'], fmt_qty(top['mtc_qty'])
        elif not remarks:
            if acc_at_lvl > 0:
                m_inv, m_qty = "Short Received", fmt_qty(acc_at_lvl)
                remarks = f"Accounted for {int(acc_at_lvl)} units (Shortage={int(shortage)}, REBNI={int(r_avail)}, EX={int(ex_adj)})"
            elif rec_qty > 0 and shortage == 0: remarks = "No Invoice Search matches found — Rec Qty = Inv Qty. Possible data mismatch. Verify manually in DICES."
            else: remarks = "No Invoice Search matches found — verify manually."

        cp_str, m_asn, m_po = '', '', ''
        if sorted_m and m_inv not in ('Self Matching', 'Short Received', ''):
            top = sorted_m[0]; m_asn, m_po = top.get('mtc_asin', ''), top.get('mtc_po', '')
            # v6.1.1 CP comparison uses Target ASIN (asin) instead of Matched ASIN (m_asn)
            cp_str = self.compare_cp(sid_frag, po, asin, sid_frag, m_po, asin, depth)

        rows = [self._make_row(barcode, inv_no, sid, po, asin, iqty, rec_qty, m_qty, m_inv, remarks, rec_date, depth, cp_status=cp_str, mtc_asin=m_asn, mtc_po=m_po, fc_id=fc_id)]
        if not is_manual or depth == 0:
            for m in sorted_m[(1 if (sorted_m and m_inv not in ("Self Matching", "Short Received")) else 0):]:
                # v6.1.1 CP comparison uses Target ASIN (asin)
                rows.append(self._make_row("", "", "", "", "", "", "", fmt_qty(m['mtc_qty']), m['mtc_inv'], "", "", depth, 'subrow', cp_status=self.compare_cp(sid_frag, po, asin, sid_frag, m.get('mtc_po', po), asin, depth), mtc_asin=m.get('mtc_asin',''), mtc_po=m.get('mtc_po',''), fc_id=fc_id))

        actionable = [{**m, 'inv_qty': self._resolve_inv_qty(m['mtc_inv'], m['mtc_asin'], m['inv_qty']), '_depth': depth} for m in sorted_m if m['mtc_inv'] != clean(inv_no)]
        
        # v6.2.5: Isolate found matches to the current ASIN's pending list
        self.asin_pending_matches.extend(actionable)
        self.all_found_matches.extend(actionable) 

        new_rem = max(0.0, rem_pqv - acc_at_lvl)
        if acc_at_lvl > 0:
            shp_rebni = self._get_shipment_rebni(sid_frag, clean(po))
            if shp_rebni > 0: rows.append(self._make_row('[REBNI-SHP]', inv_no, sid_frag, po, asin, '', '', shp_rebni, '', f"Shipment-level REBNI = {int(shp_rebni)} units available — Suggest TSP to utilize", rec_date, depth, rtype='rebni_shipment', fc_id=fc_id))
        if not cross_po_indicator_only: rows.extend(self._build_cross_po_rows(sid_frag, clean(po), clean(asin), depth))
        return rows, actionable, rec_qty, acc_at_lvl, new_rem, ex_adj

    def _build_cross_po_rows(self, sid, po, asin, depth):
        rows = []
        for c in self.detect_cross_po(sid, po, asin):
            rows.append(self._make_row('[CROSS PO]', '—', c['sid'], c['po'], c['asin'], c['inv_qty'], c['rec_qty'], '', '', f"Cross PO — {c['cross_type']} | Overage = {fmt_qty(c['rec_qty'])} units — investigating chain", c['date'], depth, rtype='crosspo', fc_id=c.get('fc_id', '')))
            if safe_num(c['rec_qty']) > 0:
                child_rows, _ = self.run_cross_po_investigation(c, c['cross_type'].split("\u2014")[0].strip(), safe_num(c['rec_qty']), depth=depth+1)
                rows.extend(child_rows)
        return rows

    def run_auto(self, barcode, inv_no, sid, po, asin, iqty, pqv, depth=0, visited=None, rem_pqv=None, is_claiming=True, branch_budget=None, max_depth_override=None, is_manual=False, row_callback=None, initial_cp=0.0):
        if self.stop_requested: return [], 0.0
        if visited is None: visited = set()
        if rem_pqv is None: rem_pqv = safe_num(pqv)
        if branch_budget is None: branch_budget = rem_pqv
        sid_frag = extract_sid(sid); state = (sid_frag, clean(inv_no), clean(po), clean(asin))
        eff_max = max_depth_override if max_depth_override is not None else self.MAX_DEPTH
        if state in visited:
            if self.ticket_type == "REMASH":
                loop_row = self._make_row(barcode, inv_no, sid, po, asin, iqty, 0.0, 0.0, "", "Loop repeating - Skipping duplicate investigative path for this ASIN.", "", depth, rtype='subrow')
                if row_callback: row_callback(loop_row)
                return [loop_row], 0.0
            return [], 0.0
        if depth >= eff_max: return [], 0.0
        visited = visited | {state}
        
        # v5.1.0+: Cumulative Loop Cache logic
        cached_rows, cached_acc, cached_actionable = [], 0.0, []
        if state in self.loop_cache and depth > 0:
            cached_rows, cached_acc, cached_actionable = self.loop_cache[state]
            if branch_budget <= cached_acc:
                return cached_rows, min(branch_budget, cached_acc)
        
        rows, actionable, rq, acc, n_rem, ex = self._build_level_logic(barcode, inv_no, sid, po, asin, iqty, rem_pqv, depth, is_claiming, is_manual=is_manual, initial_cp=initial_cp)
        
        # If resuming from cache, use cached matches and skip matches already explored
        if cached_actionable:
            curr_actionable = cached_actionable
            total_acc = cached_acc
            all_rows = list(cached_rows)
        else:
            curr_actionable = actionable
            total_acc = min(branch_budget, max(0.0, acc))
            all_rows = list(rows)
            if row_callback:
                for r in rows: row_callback(r)

        cur_budget = branch_budget - total_acc
        if cur_budget <= 0 or not curr_actionable or 'REBNI' in all_rows[0].get('remarks', '') or all_rows[0].get('remarks', '') == 'SR':
            if depth > 0: self.loop_cache[state] = (all_rows, total_acc, curr_actionable)
            return all_rows, total_acc

        # Resuming from the first uninvestigated match
        remaining_matches = []
        for i, match in enumerate(curr_actionable):
            if self.stop_requested: break
            if total_acc >= branch_budget:
                remaining_matches = curr_actionable[i:]
                break
            
            while self.pause_requested and not self.stop_requested:
                import time; time.sleep(0.5)

            n_inv, n_po, n_asin = match['mtc_inv'], match['mtc_po'], match['mtc_asin']
            n_budget = safe_num(match['mtc_qty']) if safe_num(match['mtc_qty']) > 0 else cur_budget
            n_iqty = self._resolve_inv_qty(n_inv, n_asin, match['inv_qty'])
            
            if hasattr(self, 'global_processed_ref'):
                mk = (clean(n_inv), clean(n_asin), clean(n_po), fmt_qty(match.get('mtc_qty',0)))
                self.global_processed_ref.add(mk)
            
            n_sid = self.cache_sid.get(n_inv) or self._find_sid(n_po, n_asin, n_inv)
            n_barcode = self.cache_bc.get(n_inv) or self.shipment_index.get(n_sid, "[DICES]")

            if self.ticket_type == "REMASH" and n_sid and n_sid != sid_frag:
                all_rows.append(self._make_row(n_barcode, n_inv, n_sid, n_po, n_asin, n_iqty, "", "", "", "units matching with different shipment Beyond shipment", "", depth + 1, fc_id=""))
                continue

            if not n_sid and self.sid_cb:
                n_sid = self.sid_cb(n_inv, n_po, n_asin)
                if n_sid: self.cache_sid[n_inv] = n_sid
            if not n_sid:
                all_rows.append(self._make_row(n_barcode, n_inv, "[ENTER SID FROM DICES]", n_po, n_asin, n_iqty, "", "", "", "Phase 2: SID not found — validate in DICES", "", depth + 1, fc_id=""))
                continue

            child_rows, child_acc = self.run_auto(n_barcode, n_inv, n_sid, n_po, n_asin, n_iqty, rem_pqv, depth+1, visited, rem_pqv-total_acc, False, n_budget, max_depth_override, is_manual, row_callback)
            all_rows.extend(child_rows)
            contribution = min(cur_budget, child_acc)
            total_acc += contribution
            cur_budget -= contribution

        if depth > 0: self.loop_cache[state] = (all_rows, total_acc, remaining_matches)
        return all_rows, total_acc

    def run_mismatch_investigation(self, data, budget, depth=0):
        # v6.2.5 Fix: Properly unpack all 6 variables and return them for consistent UI handling
        rows, match, rq, shortage, n_rem, ex = self._build_level_logic("[MISMATCH]", "", extract_sid(data.get('sid', '')), clean(data.get('po', '')), clean(data.get('asin', '')), safe_num(data.get('inv_qty', 0)), budget, depth, False)
        return rows, match, rq, shortage, n_rem, ex

    def build_one_level(self, b, i, s, p, a, iq, rem, depth=0, is_claiming=True, is_manual=False, initial_cp=0.0):
        # v6.2.5: Consistent return with 6-value internal engine return
        rows, matches, rq, acc, n_rem, ex = self._build_level_logic(b, i, s, p, a, iq, rem, depth, is_claiming, cross_po_indicator_only=is_manual, initial_cp=initial_cp)
        return rows, [m for m in matches if m['mtc_inv'] != clean(i)], rq, n_rem, ex

    def run_cross_po_investigation(self, c, case_type, budget, depth=0, visited=None):
        if visited is None: visited = set()
        c_sid, c_po, c_asin = c['sid'], c['po'], c['asin']; c_iq = safe_num(c.get('inv_qty', 0))
        state = (extract_sid(c_sid), clean(c.get('invoice_number', '')), clean(c_po), clean(c_asin))
        if state in visited:
            if self.ticket_type == "REMASH":
                return [self._make_row('[LOOP]', '—', c_sid, c_po, c_asin, c_iq, 0.0, '', '', "Loop repeating - Skipping duplicate cross-PO path.", c['date'], depth, rtype='crosspo')], 0.0
            return [], 0.0
        visited.add(state)
        raw = self.inv_p.get((extract_sid(c_sid), clean(c_po), clean(c_asin)), [])
        unique = sorted(raw, key=lambda x: safe_num(x['mtc_qty']), reverse=True)
        rebni = self.rebni_p.get((extract_sid(c_sid), clean(c_po), clean(c_asin)), [])
        rq = sum(safe_num(r.get('quantity_unpacked', 0)) for r in rebni) if rebni else 0.0
        ra = sum(safe_num(r.get('rebni_available', 0)) for r in rebni) if rebni else 0.0
        ex = sum(safe_num(r.get('quantity_adjusted', 0)) for r in rebni) if rebni else 0.0
        rd = clean(rebni[0].get('received_datetime', '')) if rebni else ''
        fc_id = clean(rebni[0].get('warehouse_id', '')) if rebni else ''
        shortage = max(0.0, c_iq - rq); acc = shortage + ra + ex
        m_inv = unique[0]['mtc_inv'] if unique else "Short Received"; m_qty = fmt_qty(unique[0]['mtc_qty']) if unique else ""
        rem = f"Phase 4 Cross PO ({case_type}): Accounted for {int(acc)} units"
        if not unique and shortage > 0: rem += " — Target met via Direct Shortage"
        elif ra > 0: rem += " — Suggest TSP to utilize REBNI"
        rows = [self._make_row('[CROSS PO]', '—', c_sid, c_po, c_asin, fmt_qty(c_iq), rq, m_qty, m_inv, rem, rd, depth, mtc_asin=unique[0].get('mtc_asin','') if unique else '', mtc_po=unique[0].get('mtc_po','') if unique else '', fc_id=fc_id)]
        for m in unique[1:]: rows.append(self._make_row("", "", "", "", "", "", "", fmt_qty(m['mtc_qty']), m['mtc_inv'], "", "", depth, 'subrow', mtc_asin=m.get('mtc_asin',''), mtc_po=m.get('mtc_po',''), fc_id=fc_id))
        total_acc = min(budget, max(0.0, acc)); cur_rem = budget - total_acc
        if ra > 0 or not unique or cur_rem <= 0: return rows, total_acc
        for match in unique:
            if match not in getattr(self, 'all_found_matches', []):
                self.all_found_matches.append(match)
                if hasattr(self, 'asin_pending_matches'):
                    self.asin_pending_matches.append(match)
            if cur_rem <= 0: break
            while self.pause_requested and not self.stop_requested:
                import time; time.sleep(0.5)

            n_inv, n_po, n_asin = match['mtc_inv'], match['mtc_po'], match['mtc_asin']
            # v6.2.6: Zero Rule - If combination not found in IAM, set Inv Qty to 0 for overage tracing.
            n_iq = self.inv_iam.get((clean(n_inv), clean(n_asin).upper()), 0)
            if n_iq == 0:
                base = strip_scr(n_inv)
                if base != clean(n_inv):
                    n_iq = self.inv_iam.get((base, clean(n_asin).upper()), 0)
            
            n_bud = safe_num(match['mtc_qty']) if safe_num(match['mtc_qty']) > 0 else cur_rem
            state = (extract_sid(c_sid), clean(n_inv), clean(n_po), clean(n_asin))
            if state in visited: continue

            # v6.1.1: Log global processing so budget-exhausted sub-items drop out of pending correctly
            if hasattr(self, 'global_processed_ref'):
                mk = (clean(n_inv), clean(n_asin), clean(n_po), fmt_qty(match.get('mtc_qty',0)))
                self.global_processed_ref.add(mk)
            
            n_sid = self.cache_sid.get(n_inv) or self._find_sid(n_po, n_asin, n_inv)
            
            # REMASH SID check
            # v6.2.5: Cross PO investigation always allows shipment transitions to trace overage chains
            pass 

            if not n_sid and self.sid_cb:
                n_sid = self.sid_cb(n_inv, n_po, n_asin)
                if n_sid: self.cache_sid[n_inv] = n_sid
            
            # Auto-populate barcode
            n_barcode = self.cache_bc.get(n_inv) or self.shipment_index.get(n_sid, "[DICES]")

            if not n_sid:
                rows.append(self._make_row(n_barcode, n_inv, "[ENTER SID]", n_po, n_asin, n_iq, "", "", "", "Phase 4: SID not found — validate in DICES", "", depth + 1))
                continue
            child_rows, child_acc = self.run_auto(n_barcode, n_inv, n_sid, n_po, n_asin, n_iq, cur_rem, depth+1, visited, cur_rem, False, min(n_bud, cur_rem))
            rows.extend(child_rows); contri = min(cur_rem, child_acc); total_acc += contri; cur_rem -= contri
        return rows, total_acc


# ═══════════════════════════════════════════════════════════
#  EXCEL WRITER
# ═══════════════════════════════════════════════════════════

def write_excel(all_blocks, path, rebni_summary_data=None):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Investigation"
    headers = ["Barcode", "Inv no", "SID", "PO", "ASIN", "Inv Qty", "Rec Qty", "Mtc Qty", "Mtc Inv", "Mtc ASIN", "Mtc PO", "FC", "Remarks", "Date", "CP"]
    H_FILL, DOM_F, SUB_F, ROOT_F, DICES_F, SR_F, INVLD_F, REBNI_F, CROSS_F, MIS_F = [PatternFill("solid", fgColor=c) for c in ["203864", "E2EFDA", "EBF3FB", "FFE0E0", "FFF2CC", "FFD7D7", "FFD0D0", "D0F0FF", "FFF0C0", "D0E8FF"]]
    H_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=10); N_FONT = Font(name="Calibri", size=10); ROOT_FT = Font(bold=True, color="9C0006", name="Calibri", size=10); SR_FT = Font(bold=True, color="CC0000", name="Calibri", size=10)
    INVLD_FT = Font(bold=True, color="880000", name="Calibri", size=10, italic=True); REBNI_FT = Font(bold=True, color="005580", name="Calibri", size=10); CROSS_FT = Font(bold=True, color="7a5c00", name="Calibri", size=10)
    SHORT_FILL = PatternFill("solid", fgColor="FFCCCC"); SHORT_FONT = Font(bold=True, color="9C0006", name="Calibri", size=10); BDR = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),  bottom=Side(style='thin'))
    KM = {'Barcode':'barcode', 'Inv no':'invoice', 'SID':'sid', 'PO':'po', 'ASIN':'asin', 'Inv Qty':'inv_qty', 'Rec Qty':'rec_qty', 'Mtc Qty':'mtc_qty', 'Mtc Inv':'mtc_inv', 'Mtc ASIN':'mtc_asin', 'Mtc PO':'mtc_po', 'FC':'fc_id', 'Remarks':'remarks', 'Date':'date', 'CP':'cp_status'}
    curr = 1
    for block in all_blocks:
        if not block: continue
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=curr, column=c, value=h); cell.fill, cell.font, cell.border = H_FILL, H_FONT, BDR
        curr += 1
        for rd in block:
            rem, rtyp, dep = str(rd.get('remarks', '')), rd.get('type', 'dominant'), rd.get('depth', 0)
            iq, rq = safe_num(rd.get('inv_qty')), safe_num(rd.get('rec_qty')); is_short = (iq > 0 and rq < iq)
            if is_short: fill, fnt = SHORT_FILL, SHORT_FONT
            elif 'invalid invoice' in rem.lower(): fill, fnt = INVLD_F, INVLD_FT
            elif 'REBNI Available' in rem or 'Shipment-level REBNI' in rem or rtyp == 'rebni_shipment': fill, fnt = REBNI_F, REBNI_FT
            elif 'Cross PO' in rem or rtyp == 'crosspo': fill, fnt = CROSS_F, CROSS_FT
            elif 'short received directly' in rem.lower() or 'Direct Shortage' in rem or 'Root cause' in rem or ('Found' in rem and 'short' in rem.lower()): fill, fnt = ROOT_F, ROOT_FT
            elif rem == 'SR': fill, fnt = SR_F, SR_FT
            elif rtyp == 'subrow': fill, fnt = SUB_F, N_FONT
            elif dep > 0: fill, fnt = DOM_F, N_FONT
            else: fill, fnt = None, N_FONT
            for c, h in enumerate(headers, 1):
                val = rd.get(KM[h], ""); final_val = val; str_val = str(val).strip()
                if val not in (None, '') and str_val.replace('.','',1).isdigit(): final_val = safe_num(str_val)
                cell = ws.cell(row=curr, column=c, value=final_val if final_val not in (None, '') else None); cell.border, cell.font = BDR, fnt
                if is_short and h == "Remarks": cell.fill, cell.font = SHORT_FILL, SHORT_FONT
                elif fill: cell.fill = fill
            curr += 1
        curr += 1
    for i, w in enumerate([18, 22, 18, 12, 14, 9, 9, 9, 26, 18, 18, 10, 42, 22, 36], 1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws2 = wb.create_sheet(title="Unique Summary")
    SH_FILL, SH_FONT = PatternFill("solid", fgColor="1F4E79"), Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    SN_FONT, SRED, SGRN = Font(name="Calibri", size=10), PatternFill("solid", fgColor="FFE2E2"), PatternFill("solid", fgColor="E2F0D9")
    CTR, LFT, SBDR = Alignment(horizontal="center", vertical="center"), Alignment(horizontal="left", vertical="center"), Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),  bottom=Side(style='thin'))
    sh = ["Barcode", "Inv no", "SID", "PO", "ASIN", "Inv Qty", "Rec Qty", "Missing QTY", "Cost Price", "Shortage Amount"]
    sw = [20, 14, 18, 12, 14, 10, 10, 14, 12, 18]
    for ci, (h, w) in enumerate(zip(sh, sw), 1):
        cell = ws2.cell(row=1, column=ci, value=h); cell.fill, cell.font, cell.border, cell.alignment = SH_FILL, SH_FONT, SBDR, CTR
        ws2.column_dimensions[get_column_letter(ci)].width = w
    seen, sr = set(), 2
    for block in all_blocks:
        for rd in block:
            b = str(rd.get('barcode', '')).strip()
            if not b or b == 'Barcode': continue
            if rd.get('type', '') == 'subrow': continue
            i, s, p, a, iq, rq = str(rd.get('invoice','')).strip(), str(rd.get('sid','')).strip(), str(rd.get('po','')).strip(), str(rd.get('asin','')).strip(), rd.get('inv_qty', 0), rd.get('rec_qty', 0)
            key = (b, i, s, p, a, str(iq), str(rq))
            if key in seen: continue
            seen.add(key); iq_n, rq_n = safe_num(iq), safe_num(rq) if str(rq).strip() != '' else 0.0
            rf = SRED if iq_n > rq_n else SGRN
            v = [b, i, s, p, a, int(iq_n) if iq_n == int(iq_n) else iq_n, int(rq_n) if rq_n == int(rq_n) else rq_n, None, safe_num(rd.get('cp_status', 0)), None]
            for ci, val in enumerate(v, 1):
                if ci == 8: cell = ws2.cell(row=sr, column=ci, value=f"=F{sr}-G{sr}")
                elif ci == 10: cell = ws2.cell(row=sr, column=ci, value=f"=H{sr}*I{sr}")
                else: cell = ws2.cell(row=sr, column=ci, value=val)
                cell.font, cell.border, cell.fill, cell.alignment = SN_FONT, SBDR, rf, (CTR if ci >= 6 else LFT)
            sr += 1
    ws2.freeze_panes = "A2"
    
    if rebni_summary_data:
        ws3 = wb.create_sheet(title="REBNI Summary")
        rh = ["PO", "ASIN", "SID", "Item_Cost", "Rebni_Available"]
        rw = [15, 15, 18, 12, 18]
        for ci, (h, w) in enumerate(zip(rh, rw), 1):
            cell = ws3.cell(row=1, column=ci, value=h); cell.fill, cell.font, cell.border, cell.alignment = SH_FILL, SH_FONT, SBDR, CTR
            ws3.column_dimensions[get_column_letter(ci)].width = w
        
        row_idx = 2
        # Data is dict: (sid, po, asin) -> [rows]
        for key, rows in rebni_summary_data.items():
            for r in rows:
                v = [r.get('po',''), r.get('asin',''), r.get('shipment_id',''), 
                     safe_num(r.get('item_cost',0)), safe_num(r.get('rebni_available',0))]
                for ci, val in enumerate(v, 1):
                    cell = ws3.cell(row=row_idx, column=ci, value=val)
                    cell.font, cell.border, cell.alignment = SN_FONT, SBDR, (CTR if ci >= 4 else LFT)
                row_idx += 1
        ws3.freeze_panes = "A2"

    wb.save(path)


# ═══════════════════════════════════════════════════════════
#  MAIN GUI
# ═══════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════
#  SEARCH CLOUD DIALOG
# ═══════════════════════════════════════════════════════════

class SearchCloudDialog(tk.Toplevel):
    def __init__(self, parent, s_type, callback):
        super().__init__(parent)
        self.callback = callback
        self.s_type = s_type
        self.title(f"Search {s_type} — v6.2.5")
        self.geometry("560x420")
        self.resizable(True, True)
        self.configure(bg="#0f0f1a")
        
        self.seller_var = tk.StringVar(value="VRP")
        self.v_code_var = tk.StringVar()
        self.source_var = tk.StringVar(value="Local")  # Local, Cloud, Custom
        self.custom_path = tk.StringVar()

        tk.Label(self, text=f"🔍 {s_type} Search Engine", 
                 bg="#16213e", fg="#4a9eff", font=("Segoe UI", 12, "bold"), height=2).pack(fill="x")
        
        # Source Selection
        sf = tk.LabelFrame(self, text="  Data Source Selection  ", bg="#0f0f1a", fg="#f0a500", font=("Segoe UI", 9, "bold"), padx=15, pady=8)
        sf.pack(fill="x", padx=20, pady=10)
        
        tk.Radiobutton(sf, text="Local (PD App Folder - Recommended)", variable=self.source_var, value="Local", 
                       bg="#0f0f1a", fg="#ffffff", selectcolor="#16213e", activebackground="#0f0f1a", font=("Segoe UI", 9)).pack(anchor="w")
        tk.Radiobutton(sf, text="Cloud Drive (W: Drive)", variable=self.source_var, value="Cloud", 
                       bg="#0f0f1a", fg="#ffffff", selectcolor="#16213e", activebackground="#0f0f1a", font=("Segoe UI", 9)).pack(anchor="w")
        tk.Radiobutton(sf, text="Custom File Selection", variable=self.source_var, value="Custom", 
                       bg="#0f0f1a", fg="#ffffff", selectcolor="#16213e", activebackground="#0f0f1a", font=("Segoe UI", 9)).pack(anchor="w")

        # Standard Fields
        f_fields = tk.Frame(self, bg="#0f0f1a"); f_fields.pack(fill="x", padx=35)
        
        f1 = tk.Frame(f_fields, bg="#0f0f1a"); f1.pack(pady=5, fill="x")
        tk.Label(f1, text="Seller Name:", bg="#0f0f1a", fg="#cccccc", width=12, anchor="w").pack(side="left")
        sellers = ["VRP", "RK World", "KayKay", "Clicktech", "Dawntech", "Etrade", "Cocoblu", "Retail EZ"]
        cb = ttk.Combobox(f1, textvariable=self.seller_var, values=sellers, state="readonly", width=25)
        cb.pack(side="left", padx=10)
        
        f2 = tk.Frame(f_fields, bg="#0f0f1a"); f2.pack(pady=5, fill="x")
        tk.Label(f2, text="Vendor Code:", bg="#0f0f1a", fg="#cccccc", width=12, anchor="w").pack(side="left")
        tk.Entry(f2, textvariable=self.v_code_var, width=28, bg="#1e1e3a", fg="white", insertbackground="white", relief="flat").pack(side="left", padx=10)

        # Custom Path Field
        bf2 = tk.Frame(self, bg="#0f0f1a"); bf2.pack(fill="x", padx=35, pady=5)
        tk.Entry(bf2, textvariable=self.custom_path, width=42, bg="#131320", fg="#4a9eff", relief="flat").pack(side="left", padx=(0,5))
        tk.Button(bf2, text="Browse", command=self._browse_custom, bg="#2d2d5e", fg="white", relief="flat").pack(side="left")

        # Action Buttons
        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=(15, 5))
        tk.Button(bf, text="CANCEL", command=self.destroy, bg="#444", fg="white", width=10, relief="flat").pack(side="left", padx=5)
        tk.Button(bf, text="SEARCH", command=self._on_search, bg="#e94560", fg="white", width=12, relief="flat", font=("Segoe UI", 9, "bold")).pack(side="left", padx=5)
        # v6.2.5: High-speed vendor-only fetch mode
        tk.Button(bf, text="FAST FETCH", command=self._on_fast_fetch, bg="#6f42c1", fg="white", width=12, relief="flat", font=("Segoe UI", 9, "bold")).pack(side="left", padx=5)
        
        if s_type == "REBNI":
            self.comp_var = tk.BooleanVar()
            tk.Checkbutton(self, text="Completely Matched & Available REBNI", variable=self.comp_var, 
                           fg="#3fb950", bg="#0f0f1a", selectcolor="#16213e", activebackground="#0f0f1a", 
                           font=("Segoe UI", 9, "bold")).pack(pady=5)
        
        apply_global_theme_to_widget(self)
        self.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{x}+{y}")

    def _browse_custom(self):
        f = filedialog.askopenfilename(filetypes=[("Text files", "*.txt"), ("CSV files", "*.csv")])
        if f: 
            self.custom_path.set(f)
            self.source_var.set("Custom")

    def _on_search(self):
        res = {
            'seller': self.seller_var.get(), 
            'vendor_code': self.v_code_var.get().strip(),
            'source': self.source_var.get(),
            'completely_matched': getattr(self, 'comp_var', None).get() if hasattr(self, 'comp_var') else False,
            'custom_path': self.custom_path.get() if self.source_var.get() == "Custom" else None,
            'mode': 'standard'
        }
        self.callback(res); self.destroy()

    def _on_fast_fetch(self):
        res = {
            'seller': self.seller_var.get(), 
            'vendor_code': self.v_code_var.get().strip(),
            'source': self.source_var.get(),
            'completely_matched': getattr(self, 'comp_var', None).get() if hasattr(self, 'comp_var') else False,
            'custom_path': self.custom_path.get() if self.source_var.get() == "Custom" else None,
            'mode': 'fast'
        }
        self.callback(res); self.destroy()

class MFIToolApp:
    def __init__(self):

        if not check_activation():
            root = tk.Tk(); root.withdraw()
            from tkinter import messagebox
            import sys
            messagebox.showerror("ACCESS DENIED", "ADMIN AUTHORIZATION REQUIRED\n\nThis version of the MFI Tool is currently inactive.\nPlease contact Mukesh (the administrator) for permission.")
            sys.exit()

        self.root = tk.Tk(); self.root.title("MFI Investigation Tool  v6.2.6  |  ROW IB (Secured Edition)")
        try: self.root.state('zoomed')
        except: self.root.attributes('-zoomed', True)
        self.root.minsize(900, 620); self.root.configure(bg="#0f0f1a")
        self.claims_path, self.rebni_path, self.inv_path, self.ship_master_path, self.ticket_id, self.mode_var, self.ticket_type_var = tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar(value="auto"), tk.StringVar(value="PDTT")
        self.ai_key = self._load_ai_key()
        self.is_light_theme = False
        self._setup_global_context_menu()
        self.all_blocks, self.preview = [], None; self._build_ui()
        # v6.1: Global ticket-wide persistence — matches and processed survive across ASINs
        self.global_matches    = []
        self.global_processed  = set()
        self.active_manual_dlg = None
        self.active_pending_dlg = None
        self.active_cross_dlg = None
        self._cloud_busy = {"REBNI": False, "Invoice": False}
        self._cloud_progress = {"REBNI": 0, "Invoice": 0}

    def _setup_global_context_menu(self):
        """v6.2.5: Provides right-click Copy/Paste/Cut context menus globally across all inputs."""
        menu = tk.Menu(self.root, tearoff=0, bg="#1e1e3a", fg="#e0e0e0", font=("Segoe UI", 9))
        menu.add_command(label="📋 Copy", command=lambda: self.root.focus_get().event_generate("<<Copy>>"))
        menu.add_command(label="📌 Paste", command=lambda: self.root.focus_get().event_generate("<<Paste>>"))
        menu.add_separator()
        menu.add_command(label="✂ Cut", command=lambda: self.root.focus_get().event_generate("<<Cut>>"))
        
        def show_context(event):
            w = event.widget
            if isinstance(w, (tk.Entry, ttk.Entry, tk.Text, ttk.Combobox)):
                try:
                    w.focus_set()
                    menu.tk_popup(event.x_root, event.y_root)
                finally:
                    menu.grab_release()

        self.root.bind_class("Entry", "<Button-3>", show_context)
        self.root.bind_class("TEntry", "<Button-3>", show_context)
        self.root.bind_class("Text", "<Button-3>", show_context)
        self.root.bind_class("TCombobox", "<Button-3>", show_context)
        
        # v6.2.5: OTA Auto-Update Check on Startup
        check_for_updates(self.root)

    def _build_ui(self):
        t = tk.Frame(self.root, bg="#16213e", height=62); t.pack(fill="x")
        tk.Label(t, text="  MFI Investigation Tool", fg="#e94560", bg="#16213e", font=("Segoe UI", 20, "bold")).pack(side="left", padx=16, pady=12)
        tk.Label(t, text="Developed by Mukesh", fg="#4a9eff", bg="#16213e", font=("Segoe UI", 10, "italic")).pack(side="right", padx=6)
        # --- Theme Palette Selector (circular color swatches) ---
        theme_frame = tk.Frame(t, bg="#16213e"); theme_frame.pack(side="right", padx=8)
        tk.Label(theme_frame, text="Theme:", fg="#cccccc", bg="#16213e", font=("Segoe UI", 8)).pack(side="left", padx=(0,4))
        self._theme_swatches = []
        swatch_colors = {
            "Dark Mode (Default)": "#0f0f1a", "Light Mode": "#f0f2f5",
            "Ocean Blue": "#0a192f", "Forest Green": "#0f1a14",
            "Sunset Orange": "#2a1610", "Purple Midnight": "#1a0f1f"
        }
        for name, color in swatch_colors.items():
            swatch = tk.Canvas(theme_frame, width=18, height=18, bg=color,
                               highlightthickness=2, highlightbackground="#555555",
                               cursor="hand2", bd=0)
            swatch.pack(side="left", padx=2)
            swatch.bind("<Button-1>", lambda e, n=name: self._apply_theme(n))
            self._theme_swatches.append((name, swatch))
        self._highlight_active_swatch()
        tk.Label(t, text="v6.2.6 | ROW IB (Secured Edition)", fg="#f0a500", bg="#16213e", font=("Segoe UI", 10, "bold")).pack(side="right", padx=10)
        leg = tk.Frame(self.root, bg="#1a1a2e", height=30); leg.pack(fill="x")
        for tx, f, b in [("Claiming","white","#0f0f1a"),("Dominant","black","#E2EFDA"),("Sub-rows","black","#EBF3FB"),("Root/Short","#9C0006","#FFE0E0"),("DICES","black","#FFF2CC"),("SR","black","#FFD7D7"),("Invalid inv","#333","#FFD0D0"),("REBNI","#333","#D0F0FF"),("Cross PO","#7a5c00","#FFF0C0"),("Mismatch","#333","#D0E8FF")]:
            tk.Label(leg, text=f"  {tx}  ", fg=f, bg=b, font=("Segoe UI", 8, "bold"), padx=8).pack(side="left", padx=3, pady=3)
        body = tk.Frame(self.root, bg="#0d0d1a", padx=24, pady=12); body.pack(fill="both", expand=True)
        inp = tk.LabelFrame(body, text="  Input Files  (Excel .xlsx or CSV .csv supported)  ", fg="#4a9eff", bg="#0d0d1a", font=("Segoe UI", 10, "bold"), padx=12, pady=8); inp.pack(fill="x", pady=6)
        self._f_row(inp, "Claims Sheet:", self.claims_path, 0); self._f_row(inp, "REBNI Result:", self.rebni_path, 1); self._f_row(inp, "Invoice Search:", self.inv_path, 2); self._f_row(inp, "Shipment Master:", self.ship_master_path, 3)
        tf = tk.Frame(body, bg="#0f0f1a"); tf.pack(anchor="w", pady=4); tk.Label(tf, text="Ticket ID:", fg="white", bg="#0f0f1a", font=("Segoe UI", 11)).pack(side="left"); tk.Entry(tf, textvariable=self.ticket_id, width=28, font=("Segoe UI", 11), bg="#1e1e3a", fg="white", insertbackground="white", relief="flat").pack(side="left", padx=8)
        m = tk.LabelFrame(body, text="Investigation Mode", fg="white", bg="#0f0f1a", padx=10, pady=5); m.pack(fill="x", pady=8); tk.Radiobutton(m, text="AUTO  —  Automatic. SID popup when not found in REBNI.", variable=self.mode_var, value="auto", fg="white", bg="#0f0f1a", selectcolor="#16213e", font=("Segoe UI", 10)).pack(anchor="w", padx=10); tk.Radiobutton(m, text="MANUAL  —  One level at a time. Live preview. Parallel interaction enabled.", variable=self.mode_var, value="manual", fg="white", bg="#0f0f1a", selectcolor="#16213e", font=("Segoe UI", 10)).pack(anchor="w", padx=10)
        tt = tk.LabelFrame(body, text="Ticket Type", fg="white", bg="#0f0f1a", padx=10, pady=5); tt.pack(fill="x", pady=6); tk.Radiobutton(tt, text="PDTT  —  Full chain investigation across all shipments (default).", variable=self.ticket_type_var, value="PDTT", fg="white", bg="#0f0f1a", selectcolor="#16213e", font=("Segoe UI", 10)).pack(anchor="w", padx=10); tk.Radiobutton(tt, text="REMASH TT  —  Claiming shipment level investigation with Overage Tracing.", variable=self.ticket_type_var, value="REMASH", fg="#f0c060", bg="#0f0f1a", selectcolor="#1a1500", font=("Segoe UI", 10)).pack(anchor="w", padx=10)
        self.status = tk.Label(body, text="Ready", fg="#4a9eff", bg="#0f0f1a", font=("Segoe UI", 10)); self.status.pack(pady=(10, 0)); self.pb = ttk.Progressbar(body, mode='determinate'); self.pb.pack(fill="x", pady=4)
        ctrl_f = tk.Frame(body, bg="#0f0f1a"); ctrl_f.pack(pady=10)
        self.run_btn = tk.Button(ctrl_f, text="▶ RUN", bg="#e94560", fg="white", font=("Segoe UI", 9, "bold"), padx=10, pady=5, relief="flat", cursor="hand2", command=self.start_run); self.run_btn.pack(side="left", padx=3)
        self.stop_inv_btn = tk.Button(ctrl_f, text="⏸ STOP", bg="#4a2020", fg="white", font=("Segoe UI", 9, "bold"), padx=8, pady=5, relief="flat", state="disabled", cursor="hand2", command=self.request_stop_investigation); self.stop_inv_btn.pack(side="left", padx=12)
        self.vault_btn = tk.Button(ctrl_f, text="📦  VIEW CROSS PO VAULT", command=self.show_vault,
                                  bg="#6a4c93", fg="white", font=("Segoe UI", 10, "bold"),
                                  padx=20, pady=8, relief="flat", cursor="hand2")
        self.vault_btn.pack(side="left", padx=12)
        bf = tk.Frame(body, bg="#0f0f1a"); bf.pack(pady=5)
        self.stop_sess_btn = tk.Button(bf, text="⏹ SESSION", bg="#3a0000", fg="white", font=("Segoe UI", 9, "bold"), padx=8, pady=5, relief="flat", state="disabled", cursor="hand2", command=self.request_stop_session); self.stop_sess_btn.pack(side="left", padx=3)
        self.save_btn = tk.Button(bf, text="💾 SAVE", bg="#2d6a4f", fg="white", font=("Segoe UI", 9, "bold"), padx=8, pady=5, relief="flat", state="normal", cursor="hand2", command=self.save_output); self.save_btn.pack(side="left", padx=3)
        self.portal_btn = tk.Button(bf, text="📑 PORTAL", bg="#1c2c42", fg="#4a9eff", font=("Segoe UI", 9, "bold"), padx=8, pady=5, relief="flat", cursor="hand2", command=self.open_summary_portal); self.portal_btn.pack(side="left", padx=3)
        self.rebni_search_btn = tk.Button(bf, text="🔍 REBNI SEARCH", bg="#ffc107", fg="#000000", font=("Segoe UI", 9, "bold"), padx=8, pady=5, relief="flat", cursor="hand2", command=lambda: self.start_cloud_search("REBNI")); self.rebni_search_btn.pack(side="left", padx=3)
        self.inv_search_btn = tk.Button(bf, text="🔍 INVOICE SEARCH", bg="#00bcd4", fg="#000000", font=("Segoe UI", 9, "bold"), padx=8, pady=5, relief="flat", cursor="hand2", command=lambda: self.start_cloud_search("Invoice")); self.inv_search_btn.pack(side="left", padx=3)
        self.restore_btn = tk.Button(bf, text="🛠️ RESTORE", bg="#333333", fg="#aaaaaa", font=("Segoe UI", 9, "bold"), padx=8, pady=5, relief="flat", cursor="hand2", command=self.restore_panels); self.restore_btn.pack(side="left", padx=3)
        self.reset_btn = tk.Button(bf, text="🔄 RESET", bg="#6c757d", fg="white", font=("Segoe UI", 9, "bold"), padx=8, pady=5, relief="flat", cursor="hand2", command=self.reset_tool); self.reset_btn.pack(side="left", padx=3)

    def _f_row(self, p, l, v, r):
        tk.Label(p, text=l, fg="#cccccc", bg="#131320", width=18, anchor="w", font=("Segoe UI", 10)).grid(row=r, column=0, sticky="w", pady=3)
        tk.Entry(p, textvariable=v, width=62, font=("Segoe UI", 10), bg="#1e1e3a", fg="white", insertbackground="white", relief="flat").grid(row=r, column=1, padx=6)
        tk.Button(p, text="Browse", command=lambda: self._browse(v), bg="#2d2d5e", fg="white", relief="flat", cursor="hand2", padx=8).grid(row=r, column=2)
        # v6.2.6: Direct Open button
        tk.Button(p, text="📂 Open", command=lambda: self._open_file(v), bg="#1a1a2e", fg="#00b4d8", font=("Segoe UI", 9, "bold"), relief="flat", cursor="hand2", padx=8).grid(row=r, column=3, padx=5)

    def _browse(self, var):
        f = filedialog.askopenfilename(filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv"), ("All Files", "*.*")])
        if f: var.set(f)

    def _open_file(self, var):
        path = var.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showwarning("Warning", "File path is empty or does not exist.")
            return
        try: os.startfile(path)
        except Exception as e: messagebox.showerror("Error", f"Could not open file: {e}")

    def show_vault(self):
        if not hasattr(self, 'engine'):
            messagebox.showinfo("Note", "Please start investigation first to initialize the engine."); return
        CrossPOVaultDialog(self.root, self.engine, self._start_vault_investigation)

    def _start_vault_investigation(self, sid, candidate, case, budget):
        if not hasattr(self, 'curr_m') or not self.curr_m:
            messagebox.showwarning("Warning", "Active investigation context required."); return
        self.curr_m.update({'depth': self.curr_m.get('depth',0)+1})
        def run_sub():
            cr, fn = self.engine.run_cross_po_investigation(candidate, case, budget, depth=self.curr_m['depth'])
            for r in cr:
                self.curr_m['block'].append(r)
                if self.preview and self.preview.winfo_exists(): self.root.after(0, lambda row=r: self.preview.add_row(row))
            self._man_step()
        threading.Thread(target=run_sub, daemon=True).start()

    def _set_status(self, msg, pct=None):
        self.root.after(0, lambda: (self.status.config(text=msg), (self.pb.__setitem__('value', pct) if pct is not None else None)))

    def start_run(self):
        if not all([self.claims_path.get(), self.rebni_path.get(), self.inv_path.get()]):
            messagebox.showerror("Error", "Please select all 3 input files."); return
        self.run_btn.config(state="disabled"); self.save_btn.config(state="disabled"); self.portal_btn.config(state="disabled"); self.stop_inv_btn.config(state="normal"); self.stop_sess_btn.config(state="normal"); self.all_blocks = []
        if self.mode_var.get() == "manual":
            if not self.preview or not self.preview.winfo_exists(): self.preview = PreviewPanel(self.root); self.preview._app = self
            else: self.preview.clear_all()
        # Initialize engine attributes
        if hasattr(self, 'engine'):
            self.engine.stop_requested = False
            self.engine.pause_requested = False
            self.engine.ticket_type = self.ticket_type_var.get()
            self.engine.collected_rebni = {} # reset
        threading.Thread(target=self._process, daemon=True).start()

    def request_stop_investigation(self):
        if not hasattr(self, 'engine'): return
        if not self.engine.pause_requested:
            self.engine.pause_requested = True
            self._set_status("Investigation paused — current results preserved. Save or resume.")
            self.root.after(0, lambda: (
                self.stop_inv_btn.config(text="▶  RESUME", bg="#2d6a4f", fg="white"),
                self.save_btn.config(state="normal"),
                self.portal_btn.config(state="normal")
            ))
        else:
            self.engine.pause_requested = False
            self._set_status("Resuming investigation...")
            self.root.after(0, lambda: (
                self.stop_inv_btn.config(text="⏸  STOP", bg="#4a2020", fg="white"),
                self.save_btn.config(state="disabled")
            ))

    def request_stop_session(self):
        if hasattr(self, 'engine'): self.engine.stop_requested = True
        self._set_status("Session ended — saving current results."); self._finish()

    def reset_tool(self):
        # v6.1: COMPLETE reset — wipe ALL memory, paths, engine state, and UI panels
        # 1) Wipe global investigation memory (fixes cross-ASIN leakage)
        self.global_matches   = []
        self.global_processed = set()
        self.all_blocks       = []
        # 2) Wipe all file paths so no old data carries to next ticket
        self.ticket_id.set("")
        self.claims_path.set("")
        self.rebni_path.set("")
        self.inv_path.set("")
        self.ship_master_path.set("") # v6.2.5 Fix: Clear Shipment Master on reset
        # 3) Wipe engine caches completely
        if hasattr(self, 'engine'):
            self.engine.user_overrides = {}
            self.engine.collected_rebni = {}
            self.engine.cache_sid = {}
            self.engine.cache_bc  = {}
            self.engine.loop_cache = {}
            self.engine.stop_requested = False
            self.engine.pause_requested = False
            self.engine.all_found_matches = []
        # 4) Destroy all open panels
        if self.preview and self.preview.winfo_exists():
            self.preview.destroy()
        self.preview = None
        if getattr(self, 'active_manual_dlg', None) and self.active_manual_dlg.winfo_exists():
            self.active_manual_dlg.destroy()
        self.active_manual_dlg = None
        if getattr(self, 'active_pending_dlg', None) and self.active_pending_dlg.winfo_exists():
            self.active_pending_dlg.destroy()
        self.active_pending_dlg = None
        # 5) Reset curr_m context if it exists
        if hasattr(self, 'curr_m'):
            del self.curr_m
        if hasattr(self, 'manual_q'):
            self.manual_q = []
        if hasattr(self, '_pending_cross_po'):
            self._pending_cross_po = []
        # 6) Reset button states
        self.rebni_search_btn.config(state="normal")
        self.inv_search_btn.config(state="normal")
        self.run_btn.config(state="normal")
        self.stop_inv_btn.config(state="disabled")
        self.stop_sess_btn.config(state="disabled")
        self._cloud_busy = {"REBNI": False, "Invoice": False}
        self._set_status("Ready — tool fully reset for next ticket.", 0)
        messagebox.showinfo("Reset", "Tool reset completely. All memory, paths, and cached data have been cleared.")

    def _update_cloud_ui_state(self):
        # Selective interlock: Only disable the active search type
        r_busy = self._cloud_busy.get("REBNI", False)
        i_busy = self._cloud_busy.get("Invoice", False)
        
        self.rebni_search_btn.config(state="disabled" if r_busy else "normal")
        self.inv_search_btn.config(state="disabled" if i_busy else "normal")
        
        # Disable RUN button if ANY search is busy to prevent investigation with partial/in-progress data
        self.run_btn.config(state="disabled" if (r_busy or i_busy) else "normal")

    def start_cloud_search(self, s_type):
        SearchCloudDialog(self.root, s_type, lambda res: self.run_cloud_query(s_type, res))

    def run_cloud_query(self, s_type, params):
        if not params: return
        self._cloud_busy[s_type] = True
        self._update_cloud_ui_state()
        
        if params.get('mode') == 'fast':
            threading.Thread(target=self._exec_fast_fetch, args=(s_type, params), daemon=True).start()
        else:
            threading.Thread(target=self._exec_cloud_query, args=(s_type, params), daemon=True).start()

    def _update_concurrent_status(self):
        msg_parts = []
        for st in ["REBNI", "Invoice"]:
            if self._cloud_busy[st]:
                msg_parts.append(f"[{st}] Scanning... {self._cloud_progress[st]:,} rows")
        
        if msg_parts:
            self._set_status(" | ".join(msg_parts))
        else:
            self._set_status("Cloud search complete.")

    def _exec_cloud_query(self, s_type, params):
        try:
            seller, v_code = params['seller'], params['vendor_code']
            comp_match = params.get('completely_matched', False)
            source = params.get('source', 'Local')
            custom_path = params.get('custom_path')
            
            f_map = {
                'Invoice': {"VRP":"vrp.txt", "RK World":"rk.txt", "KayKay":"kaykay.txt", "Clicktech":"Clicktech.txt", "Dawntech":"Dawntech.txt", "Etrade":"Etrade.txt", "Cocoblu":"Cocoblu.txt", "Retail EZ":"RetailEZ.txt"},
                'REBNI': {"VRP":"vrp.txt", "RK World":"rk.txt", "KayKay":"kk.txt", "Clicktech":"clicktech.txt", "Dawntech":"dawntech.txt", "Etrade":"etrade.txt", "Cocoblu":"cocoblu.txt", "Retail EZ":"retailEZ.txt"}
            }
            fname = f_map[s_type].get(seller)
            if comp_match and s_type == "REBNI": fname = "Available.txt"

            if source == "Custom":
                target_file = custom_path
            elif source == "Cloud":
                root_dir = r"W:\Shared With Me\PD Invoice & Rebni data"
                if not os.path.exists(root_dir):
                    self.root.after(0, lambda: messagebox.showerror("Drive Error", f"Cloud directory (W:) not found.\nPlease ensure the drive is correctly mapped.")); return
                target_file = os.path.join(root_dir, s_type, fname)
            else: # Local
                target_file = os.path.join(os.path.expanduser("~"), "Downloads", "PD App", s_type, fname)

            if not target_file or not os.path.exists(target_file):
                self.root.after(0, lambda: (
                    self._cloud_busy.update({s_type: False}),
                    self._update_cloud_ui_state(),
                    messagebox.showerror("Error", f"File not found at source ({source}):\n{target_file}")
                )); return
            
            ticket_id = self.ticket_id.get().strip()
            filtered_chunks = []
            chunk_size = 200000 # v6.2.5: Increased chunk size for bulk throughput
            total_rows_processed = 0
            
            reader = pd.read_csv(target_file, sep='\t', low_memory=False, chunksize=chunk_size, on_bad_lines='skip')
            for chunk in reader:
                total_rows_processed += len(chunk)
                self._cloud_progress[s_type] = total_rows_processed
                self.root.after(0, self._update_concurrent_status)
                
                filtered = chunk.copy()
                if ticket_id:
                    col = 'purchase_order_id' if 'purchase_order_id' in chunk.columns else 'po' if 'po' in chunk.columns else None
                    if col: filtered = filtered[filtered[col].astype(str).str.contains(ticket_id, na=False)]
                if v_code:
                    col = 'vendor_code' if 'vendor_code' in chunk.columns else None
                    if col: filtered = filtered[filtered[col].astype(str).str.contains(v_code, na=False)]
                
                if not filtered.empty:
                    filtered_chunks.append(filtered)
            
            self._finalize_cloud_results(s_type, source, v_code, filtered_chunks, ticket_id)
            
        except Exception as e:
            self.root.after(0, lambda: (
                self._cloud_busy.update({s_type: False}),
                self._update_cloud_ui_state(),
                self._update_concurrent_status(),
                messagebox.showerror("Search Error", str(e))
            ))

    def _exec_fast_fetch(self, s_type, params):
        # v6.2.5: High-speed vendor-only equality check for bulk datasets
        try:
            seller, v_code = params['seller'], params['vendor_code']
            source = params.get('source', 'Local')
            custom_path = params.get('custom_path')
            
            f_map = {
                'Invoice': {"VRP":"vrp.txt", "RK World":"rk.txt", "KayKay":"kaykay.txt", "Clicktech":"Clicktech.txt", "Dawntech":"Dawntech.txt", "Etrade":"Etrade.txt", "Cocoblu":"Cocoblu.txt", "Retail EZ":"RetailEZ.txt"},
                'REBNI': {"VRP":"vrp.txt", "RK World":"rk.txt", "KayKay":"kk.txt", "Clicktech":"clicktech.txt", "Dawntech":"dawntech.txt", "Etrade":"etrade.txt", "Cocoblu":"cocoblu.txt", "Retail EZ":"retailEZ.txt"}
            }
            fname = f_map[s_type].get(seller)
            if not fname: fname = "Available.txt" # fallback

            if source == "Custom": target_file = custom_path
            elif source == "Cloud": target_file = os.path.join(r"W:\Shared With Me\PD Invoice & Rebni data", s_type, fname)
            else: target_file = os.path.join(os.path.expanduser("~"), "Downloads", "PD App", s_type, fname)

            if not target_file or not os.path.exists(target_file):
                self.root.after(0, lambda: (
                    self._cloud_busy.update({s_type: False}),
                    self._update_cloud_ui_state(),
                    messagebox.showerror("Error", f"Fast Fetch source not found: {target_file}")
                )); return

            self._set_status(f"Fast Fetch: Connecting to {source}...")
            filtered_chunks = []
            chunk_size = 200000
            total_rows = 0
            
            # High-speed line-based filtering (avoids Pandas overhead for massive files)
            with open(target_file, 'r', encoding='utf-8', errors='ignore') as f_in:
                header = f_in.readline()
                if not header:
                    raise ValueError("Target file is empty.")
                
                cols = header.strip().split('\t')
                try:
                    v_idx = cols.index('vendor_code')
                except ValueError:
                    # Fallback to slower pandas if column not found easily
                    f_in.seek(0)
                    reader = pd.read_csv(target_file, sep='\t', low_memory=False, chunksize=chunk_size, on_bad_lines='skip')
                    for chunk in reader:
                        total_rows += len(chunk)
                        self._cloud_progress[s_type] = total_rows
                        self.root.after(0, self._update_concurrent_status)
                        if 'vendor_code' in chunk.columns:
                            filtered = chunk[chunk['vendor_code'].astype(str) == v_code]
                            if not filtered.empty: filtered_chunks.append(filtered)
                    self._finalize_cloud_results(s_type, source, v_code, filtered_chunks, "ALL")
                    return

                # Fast line-by-line matching
                matching_lines = [header]
                for line in f_in:
                    total_rows += 1
                    if total_rows % 100000 == 0:
                        self._cloud_progress[s_type] = total_rows
                        self.root.after(0, lambda tr=total_rows: (
                            self._set_status(f"Fast Fetch: Scanned {tr:,} lines..."),
                            self._update_concurrent_status()
                        ))
                    
                    parts = line.split('\t')
                    if len(parts) > v_idx and parts[v_idx].strip() == v_code:
                        matching_lines.append(line)
                
                if len(matching_lines) > 1:
                    # Convert only the matching lines to a DF for finalizing
                    from io import StringIO
                    final_df = pd.read_csv(StringIO("".join(matching_lines)), sep='\t')
                    filtered_chunks = [final_df]

            self._finalize_cloud_results(s_type, source, v_code, filtered_chunks, "ALL")

        except Exception as e:
            self.root.after(0, lambda: (
                self._cloud_busy.update({s_type: False}),
                self._update_cloud_ui_state(),
                messagebox.showerror("Fast Fetch Error", str(e))
            ))

    def _finalize_cloud_results(self, s_type, source, v_code, filtered_chunks, ticket_id):
        if not filtered_chunks:
            self.root.after(0, lambda: (
                self._cloud_busy.update({s_type: False}),
                self._update_cloud_ui_state(),
                self._update_concurrent_status(),
                messagebox.showinfo("No Results", f"No matches found in {source} for PO: {ticket_id} and Vendor: {v_code}")
            )); return
        
        final_df = pd.concat(filtered_chunks)
        res_dir = os.path.join(os.path.expanduser("~"), "Downloads")
        v_suffix = f" - {v_code}" if v_code else ""
        res_name = f"{s_type}_Search_Result{v_suffix}.csv"
        res_path = os.path.join(res_dir, res_name)
        final_df.to_csv(res_path, index=False)
        
        self.root.after(0, lambda: (
            self.rebni_path.set(res_path) if s_type == "REBNI" else self.inv_path.set(res_path),
            self._cloud_busy.update({s_type: False}),
            self._update_cloud_ui_state(),
            self._update_concurrent_status(),
            messagebox.showinfo("Success", f"{s_type} extraction ({source}) complete.\nSaved to: {res_path}\nMain path auto-populated.")
        ))
            
    def _apply_theme(self, theme_name):
        global GLOBAL_THEME_NAME
        GLOBAL_THEME_NAME = theme_name
        self.is_light_theme = (theme_name == "Light Mode")
        apply_global_theme_to_widget(self.root)
        if self.preview and self.preview.winfo_exists():
            apply_global_theme_to_widget(self.preview)
        self._highlight_active_swatch()

    def _highlight_active_swatch(self):
        for name, swatch in self._theme_swatches:
            if name == GLOBAL_THEME_NAME:
                swatch.config(highlightbackground="#ffc107", highlightthickness=3)
            else:
                swatch.config(highlightbackground="#555555", highlightthickness=2)

    def _toggle_theme(self):
        # Legacy compatibility: cycle between Dark and Light
        if GLOBAL_THEME_NAME == "Dark Mode (Default)":
            self._apply_theme("Light Mode")
        else:
            self._apply_theme("Dark Mode (Default)")

    # --- AI INTEGRATION METHODS ---
    def _load_ai_key(self):
        try:
            if os.path.exists("config.json"):
                with open("config.json", "r") as f:
                    return json.load(f).get("gemini_api_key", "")
        except: pass
        return ""

    def _save_ai_key(self, key):
        try:
            cfg = {}
            if os.path.exists("config.json"):
                with open("config.json", "r") as f: cfg = json.load(f)
            cfg["gemini_api_key"] = key
            with open("config.json", "w") as f: json.dump(cfg, f)
            self.ai_key = key
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save config: {e}")

    def _process(self):
        try:
            self._set_status("Loading Claims file…", 5); df_c = load_claims(self.claims_path.get()); mc, corr = detect_claim_cols(df_c)
            if corr or [f for f in COLUMN_ALIASES if f not in mc]:
                conf, done = [None], threading.Event(); self.root.after(0, lambda: HeaderCorrectionDialog(self.root, corr, mc, list(df_c.columns), lambda res: (conf.__setitem__(0, res['mapping']) if res['action'] == 'proceed' else None, done.set()))); done.wait()
                if conf[0] is None: self.root.after(0, lambda: (self.run_btn.config(state="normal"), self.stop_inv_btn.config(state="disabled"), self.stop_sess_btn.config(state="disabled"))); return
                mc = conf[0]
            
            # v6.2.5: Shipment Master Integration
            ship_idx = {}
            if self.ship_master_path.get().strip():
                self._set_status("Loading Shipment Master…", 8)
                ship_idx = build_shipment_index(load_shipment_master(self.ship_master_path.get()))

            self._set_status("Loading REBNI…", 12); rp, rs, rfb, rsid = build_rebni_index(load_rebni(self.rebni_path.get())); self._set_status("Loading Invoice Search…", 30); ip, ifb, iam = build_invoice_index(load_invoice_search(self.inv_path.get()))
            
            self.engine = InvestigationEngine(rp, rs, rfb, rsid, ip, ifb, iam, self._req_sid, ship_idx=ship_idx)
            self.engine.ticket_type = self.ticket_type_var.get()
            self.engine.global_processed_ref = self.global_processed
            tot = len(df_c)
            if self.mode_var.get() == "auto":
                self.preview = PreviewPanel(self.root); self.preview._app = self
                for i, (_, r) in enumerate(df_c.iterrows()):
                    if self.engine.stop_requested: break
                    self.engine.asin_pending_matches = []
                    self.global_processed.clear()
                    self._set_status(f"Auto: {i+1}/{tot}  ASIN: {clean(r.get(mc.get('ASIN',''),''))}", 60 + int((i / max(tot, 1)) * 35)); self.preview.add_header_row(f"{i+1}/{tot}: {clean(r.get(mc.get('ASIN',''),''))}")
                    rows, _ = self.engine.run_auto(clean(r.get(mc.get('Barcode', ''), '')), clean(r.get(mc.get('Invoice', ''), '')), extract_sid(clean(r.get(mc.get('SID', ''), ''))), clean(r.get(mc.get('PO', ''), '')), clean(r.get(mc.get('ASIN', ''), '')), safe_num(r.get(mc.get('InvQty', ''), 0)), safe_num(r.get(mc.get('PQV', ''), 0)), initial_cp=safe_num(r.get(mc.get('CP', ''), 0)), row_callback=lambda row: (self.preview.add_row(row) if self.preview and self.preview.winfo_exists() else None, self.root.update()))
                    self.all_blocks.append(rows)
                self._finish()
            else: self.manual_q, self.map_cols = df_c.to_dict('records'), mc; self._next_man()
        except Exception as e:
            import traceback; tb = traceback.format_exc(); self.root.after(0, lambda: messagebox.showerror("Error", f"{e}\n\n{tb}")); self._finish()

    def _req_sid(self, inv, po, asin):
        if inv in self.engine.cache_sid: return self.engine.cache_sid[inv]
        res, done = [None], threading.Event(); self.root.after(0, lambda: SIDRequestDialog(self.root, inv, po, asin, lambda s: (self.engine.cache_sid.__setitem__(inv, s) if s else None, res.__setitem__(0, s), done.set()))); done.wait(); return res[0]

    def restore_panels(self):
        options = []
        if self.preview and self.preview.winfo_exists() and self.preview.state() in ("withdrawn", "iconic"):
            options.append('Main Preview Panel')
        if getattr(self, 'active_manual_dlg', None) and self.active_manual_dlg.winfo_exists() and self.active_manual_dlg.state() in ("withdrawn", "iconic"):
            options.append('Manual Investigation Dialog')
        if getattr(self, 'active_pending_dlg', None) and self.active_pending_dlg.winfo_exists() and self.active_pending_dlg.state() in ("withdrawn", "iconic"):
            options.append('Pending Invoices Context')
        if getattr(self, 'active_cross_dlg', None) and self.active_cross_dlg.winfo_exists() and self.active_cross_dlg.state() in ("withdrawn", "iconic"):
            options.append('Cross PO Detection Dialog')
        if not options:
            if self.all_blocks:
                self.preview = PreviewPanel(self.root); self.preview._app = self
                for block in self.all_blocks:
                    self.preview.add_header_row('Restored Data')
                    for row in block: self.preview.add_row(row)
                self._set_status('Preview panel restored.')
                return
            self._set_status('No panels to restore.'); return
        if len(options) == 1: self._do_restore(options[0]); return
        dlg = tk.Toplevel(self.root)
        dlg.title('Restore Panel'); dlg.geometry('380x220'); dlg.configure(bg='#16213e'); dlg.resizable(False, False)
        tk.Label(dlg, text='Select panel to restore:', bg='#16213e', fg='white', font=('Segoe UI', 11, 'bold')).pack(pady=(15, 8))
        v = tk.StringVar(value=options[0])
        ttk.Combobox(dlg, textvariable=v, values=options, state='readonly', width=38, font=('Segoe UI', 10)).pack(pady=8)
        tk.Button(dlg, text='✔  Restore Selected', command=lambda: (self._do_restore(v.get()), dlg.destroy()), bg='#2d6a4f', fg='white', font=('Segoe UI', 10, 'bold'), padx=16, pady=8, relief='flat').pack(pady=12)
        dlg.update_idletasks()
        px = self.root.winfo_x() + (self.root.winfo_width() - dlg.winfo_width()) // 2
        py = self.root.winfo_y() + (self.root.winfo_height() - dlg.winfo_height()) // 2
        dlg.geometry(f'+{px}+{py}')

    def _do_restore(self, panel_name):
        if panel_name == 'Main Preview Panel' and getattr(self, 'preview', None) and self.preview.winfo_exists():
            self.preview.deiconify(); self.preview.lift(); self.preview.focus_force()
        elif panel_name == 'Manual Investigation Dialog' and getattr(self, 'active_manual_dlg', None) and self.active_manual_dlg.winfo_exists():
            self.active_manual_dlg.deiconify(); self.active_manual_dlg.lift(); self.active_manual_dlg.focus_force()
        elif panel_name == 'Pending Invoices Context' and getattr(self, 'active_pending_dlg', None) and self.active_pending_dlg.winfo_exists():
            self.active_pending_dlg.deiconify(); self.active_pending_dlg.lift(); self.active_pending_dlg.focus_force()
        elif panel_name == 'Cross PO Detection Dialog' and getattr(self, 'active_cross_dlg', None) and self.active_cross_dlg.winfo_exists():
            self.active_cross_dlg.deiconify(); self.active_cross_dlg.lift(); self.active_cross_dlg.focus_force()

    def _next_man(self, force_next=False):
        if hasattr(self, '_inv_proc_thread') and self._inv_proc_thread.is_alive():
            self._set_status("Waiting for background processes...")
            self.root.after(1000, lambda: self._next_man(force_next))
            return

        # v6.2.6: Prioritize Pending Cross POs (LIFO) before moving to next ASIN
        if hasattr(self, '_pending_cross_po') and self._pending_cross_po:
            p = self._pending_cross_po.pop()
            self._set_status(f"Starting Cross PO: {p['candidate']['po']}")
            def run_sub():
                cr, fn = self.engine.run_cross_po_investigation(p['candidate'], p['case'], p['budget'], depth=self.curr_m['depth']+1)
                for r in cr:
                    self.curr_m['block'].append(r)
                    if self.preview and self.preview.winfo_exists(): self.root.after(0, lambda row=r: self.preview.add_row(row))
                self._man_step()
            threading.Thread(target=run_sub, daemon=True).start(); return

        if hasattr(self, 'curr_m') and not force_next:
            if self._collect_all_pending(): self.root.after(0, self._show_pending_gateway); return
            if self.curr_m.get('block'): self.all_blocks.append(self.curr_m['block'])
        
        if not self.manual_q or self.engine.stop_requested: self._finish(); return
        
        # v6.2.5/6.2.6: Reset isolation ONLY when truly starting a NEW claiming ASIN
        self.engine.asin_pending_matches = []
        self.global_processed.clear()

        r, mc = self.manual_q.pop(0), self.map_cols
        self.curr_m = {'b':clean(r.get(mc.get('Barcode',''),'')), 'i':clean(r.get(mc.get('Invoice',''),'')), 's':extract_sid(clean(r.get(mc.get('SID',''),''))), 'p':clean(r.get(mc.get('PO',''),'')), 'a':clean(r.get(mc.get('ASIN',''),'')), 'iq':safe_num(r.get(mc.get('InvQty',''),0)), 'pqv':safe_num(r.get(mc.get('PQV',''),0)), 'rem':safe_num(r.get(mc.get('PQV',''),0)), 'budget':safe_num(r.get(mc.get('PQV',''),0)), 'depth':0, 'block':[], 'processed':self.global_processed, 'all_seen_matches':self.global_matches, 'asin_rendered_levels':set(), 'is_new_block':True, 'rendered':False, 'siblings_stack':[], 'initial_cp':safe_num(r.get(mc.get('CP',''),0))}
        # v6.1: Safety — recreate preview if it was destroyed (e.g. after reset)
        if not self.preview or not self.preview.winfo_exists():
            self.preview = PreviewPanel(self.root); self.preview._app = self
        self.preview.add_header_row(self.curr_m['a']); threading.Thread(target=self._man_step, daemon=True).start()

    def _man_step(self):
        if self.engine.stop_requested: self._finish(); return
        m = self.curr_m; rows, matches, rq, n_rem, ex = self.engine.build_one_level(m['b'], m['i'], m['s'], m['p'], m['a'], m['iq'], m['rem'], m['depth'], is_claiming=(m['depth']==0), is_manual=True, initial_cp=m.get('initial_cp',0.0))
        k = (m['depth'], clean(m['s']), clean(m['p']), clean(m['a']), clean(m['i']))
        if not m['rendered'] and k not in m['asin_rendered_levels']:
            if m.get('is_new_block') and rows: rows[0]['is_new_block']=True; m['is_new_block']=False
            m['block'].extend(rows); [self.preview.add_row(r) for r in rows]; self.root.update(); m['asin_rendered_levels'].add(k); m['rendered']=True
        elif not m['rendered']: [self.preview.add_row(r) for r in rows[1:]]; m['rendered']=True
        if not m.get('cross_po_checked'):
            m['cross_po_checked']=True; cands=self.engine.detect_cross_po(m['s'], m['p'], m['a'])
            if cands: 
                m['_awaiting_cross_po']=True; 
                def create_cross():
                    if getattr(self, 'active_cross_dlg', None) and self.active_cross_dlg.winfo_exists(): self.active_cross_dlg.destroy()
                    self.active_cross_dlg = CrossPODialog(self.root, cands, m['i'], m['s'], lambda res: self._handle_cross_po_and_finish(res), engine=self.engine)
                self.root.after(0, create_cross)
                return
        pm = []
        seen_m_keys = set()
        for mt in matches:
            aug = dict(mt); aug['_depth'] = m.get('depth',0); m.setdefault('all_seen_matches',[]).append(aug)
            
            # v6.2.6: Deduplicate by (Inv, ASIN, PO, Qty) to allow separate branches for same invoice with diff context.
            uk = (clean(mt.get('mtc_inv', '')), clean(mt.get('mtc_asin', '')), clean(mt.get('mtc_po', '')), fmt_qty(mt.get('mtc_qty',0)))
            if uk in seen_m_keys: continue
            seen_m_keys.add(uk)
            
            pm.append(mt)
        matches = pm
        
        m['rem']=n_rem; matches=[x for x in matches if self._get_loop_key(x) not in m['processed']]; rem_s=rows[0].get('remarks','') if rows else ''
        if not matches: rem_s="No unprocessed matches remaining"
        if not matches or any(kw in rem_s for kw in ["Root cause", "REBNI", "SR", "short received directly", "Direct Shortage", "Phase 1", "No Invoice Search"]):
            if m.get('siblings_stack'):
                ctx=m['siblings_stack'].pop(0); m.update({'depth':ctx['depth'], 'rem':ctx['rem'], 'budget':ctx['budget'], 'b':ctx['b'], 'i':ctx['i'], 'iq':ctx['iq'], 's':ctx['s'], 'p':ctx['p'], 'a':ctx['a'], 'processed':ctx['processed'], 'rendered':False}); self.root.after(0, lambda: self._show_dlg(ctx['siblings'])); return
            if hasattr(self, '_pending_cross_po') and self._pending_cross_po: self.root.after(0, self._next_man); return
            self.root.after(0, self._next_man); return
        self.root.after(0, lambda: self._show_dlg(matches))

    def _show_dlg(self, matches):
        m, f = self.curr_m, matches[0]
        if self.engine.user_overrides:
            pm = []
            for mt in matches:
                uk = (clean(mt.get('mtc_inv', '')), clean(mt.get('mtc_asin', '')), clean(mt.get('mtc_po', '')))
                ov = self.engine.user_overrides.get(uk, {})
                if not ov: ov = self.engine.user_overrides.get(clean(mt.get('mtc_inv', '')), {})
                if ov:
                    mt=dict(mt)
                    if 'mtc_qty' in ov: mt['mtc_qty']=ov['mtc_qty']
                    if 'inv_qty' in ov: mt['inv_qty']=ov['inv_qty']
                pm.append(mt)
            matches, f = pm, pm[0]
        if f['mtc_inv'] in self.engine.cache_sid: self._handle_res({'action':'valid','chosen_match':f,'sid':self.engine.cache_sid[f['mtc_inv']],'barcode':self.engine.cache_bc.get(f['mtc_inv'],"[DICES]")}, matches); return
        if getattr(self, 'active_manual_dlg', None) and self.active_manual_dlg.winfo_exists(): self.active_manual_dlg.destroy()
        self.active_manual_dlg = ManualLevelDialog(self.root, matches, m['rem'], m['budget'], lambda rs: self._handle_res(rs, matches), pending_cb=self._show_pending_invoices_from_dialog, engine=self.engine)

    def _handle_res(self, res, matches):
        if not res or res['action'] == 'stop': self.root.after(0, self._next_man); return
        mc = res.get('chosen_match')
        if mc: self.curr_m['processed'].add(self._get_loop_key(mc))
        if res['action'] == 'invalid':
            ex = res['invalid_qty']; ro = self.engine._make_row('[INVALID]', mc['mtc_inv'], '—', mc['mtc_po'], mc['mtc_asin'], mc['inv_qty'], '', '', '', f"{int(ex)} units matched to invalid invoice {mc['mtc_inv']} — excluded from PQV", '', self.curr_m['depth'], 'subrow')
            self.curr_m['block'].append(ro); self.preview.add_row(ro); self.curr_m['rem']=max(0, self.curr_m['rem']-ex)
            if self.curr_m['rem']<=0: self.root.after(0, self._next_man)
            else:
                rm = [x for x in matches if x['mtc_inv'] != mc['mtc_inv']]
                if rm: self.root.after(0, lambda: self._show_dlg(rm))
                else: self.root.after(0, self._next_man)
        elif res['action'] == 'cross_po':
            cands = self.engine.detect_cross_po(self.curr_m['s'], self.curr_m['p'], self.curr_m['a'])
            if cands: 
                if getattr(self, 'active_cross_dlg', None) and self.active_cross_dlg.winfo_exists(): self.active_cross_dlg.destroy()
                self.active_cross_dlg = CrossPODialog(self.root, cands, self.curr_m['i'], self.curr_m['s'], lambda r: self._handle_cross_po(r), engine=self.engine)
            else: messagebox.showinfo("No Cross PO", "No Cross PO candidates found."); self.root.after(0, lambda: self._show_dlg(matches))
        elif res['action'] == 'mismatch':
            d = res['mismatch_data']; bug = safe_num(d.get('ovg_qty', 0)) or self.curr_m['rem']
            # v6.2.5 Fix: Signature mismatch (Added 'ex' to unpack 6 returns)
            rows, sm, rq, sh, nr, ex = self.engine.run_mismatch_investigation(d, bug, depth=self.curr_m['depth'])
            for r in rows: self.curr_m['block'].append(r) or self.preview.add_row(r)
            if sm: self.root.after(0, lambda: self._show_dlg(sm))
            else:
                rm = [x for x in matches if x != res.get('chosen_match')]
                if rm: self.root.after(0, lambda: self._show_dlg(rm))
                else: self.root.after(0, self._next_man)
        else:
            self.engine.cache_sid[mc['mtc_inv']], self.engine.cache_bc[mc['mtc_inv']] = res['sid'], res['barcode']
            # v6.2.6: Filter by full loop key, not just Invoice, to allow sibling branches on same invoice.
            rem_s = [x for x in matches if self._get_loop_key(x) != self._get_loop_key(mc) and self._get_loop_key(x) not in self.curr_m['processed']]
            ps = list(self.curr_m['siblings_stack'])
            if rem_s: ps.append({'siblings':rem_s, 'depth':self.curr_m['depth'], 'rem':self.curr_m['rem'], 'budget':self.curr_m['budget'], 'b':self.curr_m.get('b',''), 'i':self.curr_m.get('i',''), 'iq':self.curr_m.get('iq',0), 's':self.curr_m['s'], 'p':self.curr_m['p'], 'a':self.curr_m['a'], 'processed':self.curr_m['processed'], 'cross_po_checked':self.curr_m.get('cross_po_checked',False)})
            bb = safe_num(mc['mtc_qty']) or self.curr_m['rem']
            self.curr_m.update({'b':res['barcode'], 'i':mc['mtc_inv'], 's':res['sid'], 'p':mc['mtc_po'], 'a':mc['mtc_asin'], 'iq':mc['inv_qty'], 'rem':bb, 'budget':bb, 'depth':self.curr_m['depth']+1, 'rendered':False, 'processed':self.curr_m['processed'], 'siblings_stack':ps, 'pending_siblings':[], 'cross_po_checked':False, '_awaiting_cross_po':False})
            threading.Thread(target=self._man_step, daemon=True).start()

    def _collect_all_pending(self):
        proc, seen, ded = self.curr_m.get('processed', set()), set(), []
        # v6.2.5: Use Per-ASIN isolation (asin_pending_matches) to prevent cross-ASIN data persistence
        src = self.engine.asin_pending_matches if hasattr(self, 'engine') and hasattr(self.engine, 'asin_pending_matches') else self.curr_m.get('all_seen_matches', [])
        for inv in src:
            k = self._get_loop_key(inv)
            if k and k not in seen and k not in proc:
                seen.add(k)
                if hasattr(self, 'engine') and self.engine.user_overrides:
                    ov_k = (clean(inv.get('mtc_inv', '')), clean(inv.get('mtc_asin', '')), clean(inv.get('mtc_po', '')))
                    ov = self.engine.user_overrides.get(ov_k, {})
                    if not ov: ov = self.engine.user_overrides.get(clean(inv.get('mtc_inv','')), {})
                    if ov:
                        inv=dict(inv)
                        if 'mtc_qty' in ov: inv['mtc_qty']=ov['mtc_qty']
                        if 'inv_qty' in ov: inv['inv_qty']=ov['inv_qty']
                ded.append(inv)
        return ded

    def _show_pending_invoices_from_dialog(self):
        ap = self._collect_all_pending()
        if not ap: messagebox.showinfo("No Pending Invoices", "All matched invoices have already been investigated or no matches were found yet.", parent=self.root); return
        if getattr(self, 'active_pending_dlg', None) and self.active_pending_dlg.winfo_exists(): self.active_pending_dlg.destroy()
        self.active_pending_dlg = PendingInvoicesDialog(self.root, ap, self.curr_m.get('a', ''), lambda res: (self.curr_m['processed'].add(self._get_loop_key(res['match'])), self.root.after(0, lambda: self._show_dlg([res['match']]))) if res['action']=='investigate' else None)

    def _show_pending_gateway(self):
        ap = self._collect_all_pending()
        if not ap: self._next_man(force_next=True); return
        if getattr(self, 'active_pending_dlg', None) and self.active_pending_dlg.winfo_exists(): self.active_pending_dlg.destroy()
        self.active_pending_dlg = PendingInvoicesDialog(self.root, ap, f"{self.curr_m.get('a', '')} (Conclusion Review)", lambda res: (self.curr_m['processed'].add(self._get_loop_key(res['match'])), self.root.after(0, lambda m=res['match']: self._show_dlg([m]))) if res['action']=='investigate' else (self.all_blocks.append(self.curr_m['block']) if self.curr_m.get('block') else None, self._next_man(force_next=True)))

    def _handle_cross_po(self, res):
        if res['action'] == 'skip': threading.Thread(target=self._man_step, daemon=True).start(); return
        sid = extract_sid(res['candidate']['sid'])
        if sid not in self.engine.cross_po_vault: self.engine.cross_po_vault[sid] = []
        
        # v6.2.6: Store in Vault categorized by Shipment ID
        self.engine.cross_po_vault[sid].append({
            'candidate': res['candidate'], 
            'case': res.get('case','Case 1'), 
            'budget': safe_num(res['candidate']['rec_qty']),
            'found_bc': res['candidate'].get('found_bc', ''),
            'found_inv': res['candidate'].get('found_inv', '')
        })
        self._set_status(f"Cross PO stored in Vault: SID {sid} | PO {res['candidate']['po']}", None)
        threading.Thread(target=self._man_step, daemon=True).start()

    def _handle_cross_po_and_finish(self, res):
        if res['action'] == 'skip': self.curr_m.pop('_awaiting_cross_po', None); threading.Thread(target=self._man_step, daemon=True).start(); return
        c, bud = res['candidate'], safe_num(res['candidate']['rec_qty'])
        # v6.1.2: Dynamically use captured Barcode, Invoice, and Inv Qty instead of empty strings
        f_bc = c.get('found_bc', '')
        f_inv = c.get('found_inv', '')
        f_iq = c.get('inv_qty', 0)
        
        self._set_status(f"Cross PO confirmed ({c['po']}) — starting manual investigation of {int(bud)} units…", None); self.curr_m.pop('_awaiting_cross_po', None)
        self.curr_m.update({'b':f_bc, 'i':f_inv, 's':c['sid'], 'p':c['po'], 'a':c['asin'], 'iq':f_iq, 'rem':bud, 'budget':bud, 'depth':self.curr_m['depth']+1, 'rendered':False, 'processed':self.curr_m['processed'], 'cross_po_checked':True, 'asin_rendered_levels':set()}); threading.Thread(target=self._man_step, daemon=True).start()

    def _finish(self):
        msg = "Investigation complete!" if not (hasattr(self, 'engine') and self.engine.stop_requested) else "Investigation stopped by user."
        self._set_status("Complete. Click SAVE.", 100); self.root.after(0, lambda: (self.run_btn.config(state="normal"), self.save_btn.config(state="normal"), self.portal_btn.config(state="normal"), self.stop_inv_btn.config(state="disabled"), self.stop_sess_btn.config(state="disabled"), messagebox.showinfo("Done", msg)))

    def _get_loop_key(self, mt): return (clean(mt.get('mtc_inv','')), clean(mt.get('mtc_asin','')), clean(mt.get('mtc_po','')), fmt_qty(mt.get('mtc_qty',0)))

    def save_output(self):
        t, ts = self.ticket_id.get().strip().replace(' ','_'), datetime.now().strftime('%Y%m%d_%H%M%S')
        o = f"MFI_{t}_{ts}.xlsx" if t else f"MFI_Investigation_{ts}.xlsx"
        p = os.path.join(os.path.dirname(self.claims_path.get()) or os.getcwd(), o)
        
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=o)
        if not path: return
        
        try:
            bl = list(self.all_blocks)
            # Include any active manual block that hasn't been committed yet
            if self.mode_var.get() == "manual" and hasattr(self, 'curr_m') and self.curr_m.get('block') and self.curr_m['block'] not in bl:
                bl.append(self.curr_m['block'])
            
            # v5.9.3 Logic: Pass collected REBNI data to Excel Writer
            rebni_data = self.engine.collected_rebni if hasattr(self, 'engine') else None
            write_excel(bl, path, rebni_summary_data=rebni_data)
            
            self._set_status(f"Project saved to: {os.path.basename(path)}")
            # v5.9.3: Enable save button for subsequent saves
            self.root.after(0, lambda: self.save_btn.config(state="normal"))
            messagebox.showinfo("Saved", f"Complete investigation report saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Save Error", str(e))

    def open_summary_portal(self):
        """v6.2.5: Launches the summary portal. Priority to Legacy HTML (Canonical), Fallback to React."""
        if hasattr(sys, '_MEIPASS'):
            base_dir = sys._MEIPASS
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            
        # Prefer legacy HTML as it contains the dashboard/AI logic the user expects
        portal_path = os.path.join(base_dir, "MFI_unique_summary_upload_export.html")
        if not os.path.exists(portal_path):
            portal_path = os.path.join(base_dir, "Web application", "dist", "index.html")
        
        if os.path.exists(portal_path):
            self._set_status("Launching Summary Portal...")
            from pathlib import Path
            webbrowser.open_new_tab(Path(portal_path).as_uri())
        else:
            messagebox.showerror("Portal Error", 
                                f"Could not find any portal files:\n{portal_path}\n\n"
                                f"Ensure the legacy HTML or 'Web application/dist' folder is present.")

    def open_generic_rec_qty_lookup(self):
        """v6.2.5: Provides a generic, standalone Received Qty lookup tool on the main screen."""
        dlg = tk.Toplevel(self.root)
        dlg.title("Generic Received Qty Lookup")
        dlg.geometry("450x300")
        dlg.configure(bg="#0d1117")
        
        tk.Label(dlg, text="  Standalone REBNI Lookup Tool", bg="#161b22", fg="#f0a500", font=("Segoe UI", 11, "bold"), height=2).pack(fill="x")
        
        form_f = tk.Frame(dlg, bg="#0d1117", padx=20, pady=15)
        form_f.pack(fill="x")
        
        tk.Label(form_f, text="Shipment ID:", bg="#0d1117", fg="#888888", font=("Segoe UI", 10)).grid(row=0, column=0, sticky='w', pady=5)
        sid_ent = tk.Entry(form_f, font=("Segoe UI", 10), bg="#21262d", fg="white", insertbackground="white", width=25)
        sid_ent.grid(row=0, column=1, padx=10, pady=5)
        
        tk.Label(form_f, text="Purchase Order:", bg="#0d1117", fg="#888888", font=("Segoe UI", 10)).grid(row=1, column=0, sticky='w', pady=5)
        po_ent = tk.Entry(form_f, font=("Segoe UI", 10), bg="#21262d", fg="white", insertbackground="white", width=25)
        po_ent.grid(row=1, column=1, padx=10, pady=5)
        
        tk.Label(form_f, text="ASIN:", bg="#0d1117", fg="#888888", font=("Segoe UI", 10)).grid(row=2, column=0, sticky='w', pady=5)
        asin_ent = tk.Entry(form_f, font=("Segoe UI", 10), bg="#21262d", fg="white", insertbackground="white", width=25)
        asin_ent.grid(row=2, column=1, padx=10, pady=5)
        
        def run_lookup():
            sid = extract_sid(sid_ent.get().strip())
            po  = clean(po_ent.get().strip())
            asin = clean(asin_ent.get().strip())
            
            if not all([sid, po, asin]):
                messagebox.showerror("Error", "All fields are required.", parent=dlg)
                return
                
            if not hasattr(self, 'engine') or not self.engine:
                messagebox.showerror("Error", "Investigation Engine not active. Run investigation first.", parent=dlg)
                return
                
            rows = self.engine.rebni_p.get((sid, po, asin), [])
            if not rows:
                messagebox.showinfo("Result", "No REBNI data found for this combination.", parent=dlg)
                return
                
            # Create the result display dialog (reusing logic or creating a small one)
            r0 = rows[0]
            res_dlg = tk.Toplevel(dlg)
            res_dlg.title(f"Result for {asin}")
            res_dlg.geometry("400x350")
            res_dlg.configure(bg="#0d1117")
            
            tk.Label(res_dlg, text=f"Reconciliation: {asin}", bg="#161b22", fg="#3fb950", font=("Segoe UI", 11, "bold"), height=2).pack(fill="x")
            
            summary_f = tk.Frame(res_dlg, bg="#0d1117", padx=20, pady=20)
            summary_f.pack(fill="both", expand=True)
            
            items = [
                ("Received (Post Adj):", r0.get('qty_received_postadj', 0), "#3fb950"),
                ("Quantity Unpacked:", r0.get('quantity_unpacked', 0), "#58a6ff"),
                ("Quantity Adjusted:", r0.get('quantity_adjusted', 0), "#ff4d4d"),
                ("REBNI Available:", r0.get('rebni_available', 0), "#f0a500")
            ]
            
            for i, (label, val, color) in enumerate(items):
                tk.Label(summary_f, text=label, bg="#0d1117", fg="#cccccc", font=("Segoe UI", 10)).grid(row=i, column=0, sticky='w', pady=10)
                tk.Label(summary_f, text=f"{int(safe_num(val))} units", bg="#0d1117", fg=color, font=("Segoe UI", 12, "bold")).grid(row=i, column=1, sticky='w', padx=15)
            
            tk.Button(res_dlg, text="CLOSE", command=res_dlg.destroy, bg="#333", fg="white").pack(pady=10)

        tk.Button(dlg, text="🔍  CHECK REBNI", command=run_lookup, bg="#d4a017", fg="black", font=("Segoe UI", 10, "bold"), padx=20, pady=8, relief="flat", cursor="hand2").pack(pady=10)

    def run(self):
        try: self.root.mainloop()
        except KeyboardInterrupt: pass

if __name__ == '__main__': MFIToolApp().run()
