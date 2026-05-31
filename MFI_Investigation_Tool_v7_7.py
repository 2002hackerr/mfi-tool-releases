"""
CHANGES IN v7.3.0 (REFINED):
  ✔ UI: Removed Investigation Path and Legend headers from main panel for cleaner look.
  ✔ Cross PO: Restored v6.2.6 logic — fixed Auto Mode chain investigation and Case 3 detection.
  ✔ Matching: Always shows Mtc Qty and Mtc Inv for Cross PO chains; labels updated to [CROSS PO?].
  ✔ Excel: Simplified "Investigation" sheet formatting — only shortages and invalid matches in Red.
  ✔ REBNI: Restored v6.2.6 availability filter (>0) to prevent zero-revenue clutter in summary.
  ✔ Bug Fix: Corrected `item_cost` mapping in independent reconciliation and UI dialogs.
  ✔ CrossPODialog: Removed duplicate UI block (double widgets stacked on top of each other)
  ✔ CrossPOVaultDialog: Removed duplicate _remove_item/_clear_vault (2nd defs caused TypeError)
  ✔ run_cross_po_investigation: Fixed UnboundLocalError cur_rem → chain_rem
  ✔ CrossPODialog._analyze_cross_po_file: Fixed self.root.after → self.after (Toplevel has no .root)
  ✔ MFIToolApp.__init__: Added self._pending_cross_po = [] initialization (prevents AttributeError)
  ✔ DownloadProgressDialog.finalize_update: messagebox now called on main thread via self.after
  ✔ finalize_update: Replaced os._exit(0) with root.destroy()+sys.exit(0) for clean shutdown
  ✔ load_invoice_search: Deduplication now keeps row with max quantity_matched_total
  ✔ config.json: Moved to ~/.mfi_tool/config.json (stable across launch directories)
  ✔ CorrespondenceDialog: Replaced Cyrillic artifact with English text

MFI Investigation Tool  v7.7  |  ROW IB
==========================================
ROW IB  |  Amazon
Developed by Mukesh

CHANGES IN v6.2.6:
- [TBD] Improving logic based on v6.2.5 feedback

CHANGES IN v6.2.5:
- [ISOLATION] Per-ASIN "Pending Invoices" tracking (Fixes global data leakage)
- [CROSS PO] Multi-Invoice detection for the same PO (Checks all invoice combinations)
- [AUTO] Global SID/Barcode persistence (Never asks for the same info twice)
- [AUTO] Loop Result Block auto-duplication (Copies previous results for repeated loops)
- [UI] New "RECEIVED QTY" button in Preview (Main SID/PO/ASIN status lookup)
- [UI] New "FAST FETCH" button in Search (High-speed vendor-only bulk filtering)
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext, simpledialog
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import os
import re
import sys
import threading
import subprocess
import webbrowser
import requests
from datetime import datetime

# ==========================================
# SECURED PRODUCTION VERSION
# ==========================================
ACTIVATION_URL = "https://gist.githubusercontent.com/2002hackerr/3f76afc8a819c6879e06676a36173999/raw/activation.txt"

def check_activation():
    import urllib.request
    try:
        req = urllib.request.Request(ACTIVATION_URL, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as response:
            content = response.read().decode('utf-8').strip().upper()
            return content == "ENABLED"
    except Exception as e:
        return False

#  OTA UPDATE ENGINE (MIGRATED TO GITHUB RELEASES API)
# ═══════════════════════════════════════════════════════════
APP_VERSION = "7.7"
# Official GitHub Releases API URL for the public repository
UPDATE_MANIFEST_URL = "https://api.github.com/repos/2002hackerr/mfi-tool-releases/releases/latest"

class UpdateDialog(tk.Toplevel):
    def __init__(self, parent, remote_ver, download_url):
        super().__init__(parent)
        self.download_url = download_url
        self.title("Software Update Available")
        self.geometry("480x280")
        self.configure(bg="#0f0f1a")
        self.resizable(False, False)
        self.lift(); self.focus_force()
        self.grab_set()

        tk.Label(self, text="🚀  New Update Available", bg="#16213e", fg="#4a9eff", 
                 font=("Segoe UI", 14, "bold"), height=2).pack(fill="x")
        
        info_frame = tk.Frame(self, bg="#0f0f1a", padx=20, pady=20)
        info_frame.pack(fill="both", expand=True)
        
        tk.Label(info_frame, text=f"A new version (v{remote_ver}) of the MFI Tool is ready.", 
                 bg="#0f0f1a", fg="white", font=("Segoe UI", 11)).pack(anchor="w")
        tk.Label(info_frame, text=f"Current version: v{APP_VERSION}", 
                 bg="#0f0f1a", fg="#888888", font=("Segoe UI", 10)).pack(anchor="w", pady=(5, 15))
        
        tk.Label(info_frame, text="Would you like to download and install it now?\nThe tool will restart automatically.", 
                 bg="#0f0f1a", fg="#cccccc", font=("Segoe UI", 10), justify="left").pack(anchor="w")

        btn_f = tk.Frame(self, bg="#0f0f1a", pady=15)
        btn_f.pack(fill="x")
        
        tk.Button(btn_f, text="Update Now", bg="#2d6a4f", fg="white", font=("Segoe UI", 10, "bold"),
                  padx=20, pady=8, relief="flat", cursor="hand2", command=self.on_update).pack(side="right", padx=20)
        tk.Button(btn_f, text="Later", bg="#333333", fg="#aaaaaa", font=("Segoe UI", 10),
                  padx=15, pady=8, relief="flat", cursor="hand2", command=self.destroy).pack(side="right")

    def on_update(self):
        self.destroy()
        DownloadProgressDialog(self.master, self.download_url)

class DownloadProgressDialog(tk.Toplevel):
    def __init__(self, parent, url):
        super().__init__(parent)
        self.url = url
        self.title("Downloading Update...")
        self.geometry("400x150")
        self.configure(bg="#0f0f1a")
        self.resizable(False, False)
        self.lift(); self.focus_force()
        self.grab_set()

        tk.Label(self, text="Downloading MFI Tool Update...", bg="#0f0f1a", fg="white", 
                 font=("Segoe UI", 11)).pack(pady=(20, 10))
        
        self.pb = ttk.Progressbar(self, length=300, mode='determinate')
        self.pb.pack(pady=10)
        
        self.status_lbl = tk.Label(self, text="Connecting...", bg="#0f0f1a", fg="#4a9eff", font=("Segoe UI", 9))
        self.status_lbl.pack()
        
        threading.Thread(target=self.start_download, daemon=True).start()

    def start_download(self):
        try:
            response = requests.get(self.url, stream=True, timeout=30)
            response.raise_for_status()  # v6.2.5 Security: Verify download success before writing
            total_size = int(response.headers.get('content-length', 0))
            
            # Use current executable path to determine update filename
            current_exe = sys.executable
            update_exe = os.path.join(os.path.dirname(current_exe), "MFI_Update_Temp.exe")
            
            downloaded = 0
            with open(update_exe, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
                        downloaded += len(chunk)
                        if total_size > 0:
                            pct = (downloaded / total_size) * 100
                            self.after(0, lambda p=pct, d=downloaded, t=total_size: self.update_status(p, d, t))
            
            self.after(0, self.finalize_update)
        except Exception as e:
            self.after(0, lambda ex=e: messagebox.showerror("Update Error", f"Failed to download update: {ex}"))
            self.after(0, self.destroy)

    def update_status(self, pct, done, total):
        self.pb['value'] = pct
        self.status_lbl.config(text=f"Downloaded {done//1024} KB / {total//1024} KB ({int(pct)}%)")

    def finalize_update(self):
        current_exe = sys.executable
        # Generate batch script for self-replacement
        # Since we use --onefile, the running EXE is what we replace.
        bat_content = f"""@echo off
timeout /t 3 /nobreak > nul
move /y "MFI_Update_Temp.exe" "{os.path.basename(current_exe)}"
start "" "{os.path.basename(current_exe)}"
del "%~f0"
"""
        bat_path = os.path.join(os.path.dirname(current_exe), "mfi_updater.bat")
        with open(bat_path, "w") as f:
            f.write(bat_content)
        
        # FIX 6+10: messagebox must be called on main thread; use self.after
        # FIX 10: os._exit(0) bypasses cleanup — use destroy+sys.exit via after
        def _apply():
            messagebox.showinfo("Update Ready", "The update has been downloaded. The tool will now restart to apply changes.")
            subprocess.Popen([bat_path], shell=True, cwd=os.path.dirname(current_exe))
            try:
                self.master.destroy()
            except Exception:
                pass
            sys.exit(0)
        self.after(0, _apply)

def check_for_updates(parent):
    def _run():
        try:
            # v6.2.5: Migrated to GitHub Releases API (Robust for office networks)
            headers = {'Accept': 'application/vnd.github.v3+json'}
            resp = requests.get(UPDATE_MANIFEST_URL, headers=headers, timeout=10)
            if resp.status_code == 200:
                data = resp.json()
                # GitHub API uses "tag_name" (e.g., "v6.2.5") and "assets"
                tag_name = data.get("tag_name", "")
                
                # Strip leading 'v' if present for comparison
                remote_ver = tag_name.lstrip('v')
                
                # Find the download URL for the .exe asset
                download_url = ""
                assets = data.get("assets", [])
                for asset in assets:
                    if asset.get("name", "").endswith(".exe"):
                        download_url = asset.get("browser_download_url", "")
                        break
                
                if remote_ver and remote_ver != APP_VERSION and download_url:
                    # Show dialog on main thread
                    parent.after(0, lambda: UpdateDialog(parent, remote_ver, download_url))
        except:
            pass # Silent failure to prevent startup issues
    
    threading.Thread(target=_run, daemon=True).start()


# ═══════════════════════════════════════════════════════════
#  GLOBAL THEME ENGINE
# ═══════════════════════════════════════════════════════════

GLOBAL_THEME_NAME = "Dark Mode (Default)"

THEME_PALETTES = {
    "Dark Mode (Default)": {}, 
    "Light Mode": {
        "#0f0f1a": "#f0f2f5", "#16213e": "#dfe3ee", "#0d0d1a": "#ffffff",
        "#1e1e3a": "#ffffff", "#131320": "#ffffff", "#1a1a2e": "#e9ebee",
        "#2d2d5e": "#8b9dc3", "white": "black", "#ffffff": "black", "#cccccc": "#444444",
        "#e0e0e0": "#222222", "#4a9eff": "#0056b3", "#e94560": "#c94a5f",
        "#0d1117": "#f6f8fa", "#161b22": "#ffffff", "#21262d": "#ffffff",
        "#3fb950": "#2ea043", "#f85149": "#cf222e"
    },
    "Ocean Blue": {
        "#0f0f1a": "#0f1a2a", "#16213e": "#1a2a4a", "#0d0d1a": "#0a1220",
        "#1e1e3a": "#2a3a5a", "#131320": "#0a1220", "#1a1a2e": "#1a2a4a",
        "#2d2d5e": "#2a5aa0", "white": "#e0f0ff", "#ffffff": "#e0f0ff", 
        "#cccccc": "#90a0c0", "#e0e0e0": "#b0c0e0", "#4a9eff": "#40c0f0", 
        "#e94560": "#30a0d0", "#0d1117": "#0f1a2a", "#161b22": "#1a2a4a", 
        "#21262d": "#2a3a5a", "#f0a500": "#40b0d0", "#3fb950": "#40c0f0"
    },
    "Forest Green": {
        "#0f0f1a": "#141a14", "#16213e": "#1e2b1e", "#0d0d1a": "#0d140d",
        "#1e1e3a": "#2a3a2a", "#131320": "#0d140d", "#1a1a2e": "#1e2b1e",
        "#2d2d5e": "#2a7a2a", "white": "#f0fff0", "#ffffff": "#f0fff0", 
        "#cccccc": "#a0c0a0", "#e0e0e0": "#c0d0c0", "#4a9eff": "#60c060", 
        "#e94560": "#40a040", "#0d1117": "#141a14", "#161b22": "#1e2b1e", 
        "#21262d": "#2a3a2a", "#f0a500": "#60c060", "#3fb950": "#2ea043", "#f85149": "#cf222e"
    },
    "Sunset Orange": {
        "#0f0f1a": "#2a1e1a", "#16213e": "#3a2a24", "#0d0d1a": "#1f1410",
        "#1e1e3a": "#4a3a34", "#131320": "#1f1410", "#1a1a2e": "#3a2a24",
        "#2d2d5e": "#a05a2a", "white": "#fff0e0", "#ffffff": "#fff0e0", 
        "#cccccc": "#c0b0a0", "#e0e0e0": "#d0c0b0", "#4a9eff": "#f0a040", 
        "#e94560": "#e06a4a", "#0d1117": "#2a1e1a", "#161b22": "#3a2a24", 
        "#21262d": "#4a3a34", "#f0a500": "#f09040", "#3fb950": "#a0c0a0"
    },
    "Purple Midnight": {
        "#0f0f1a": "#1e142a", "#16213e": "#2d1e3a", "#0d0d1a": "#140d1f",
        "#1e1e3a": "#3a2a4a", "#131320": "#140d1f", "#1a1a2e": "#2d1e3a",
        "#2d2d5e": "#7a2a9a", "white": "#f5e0ff", "#ffffff": "#f5e0ff", 
        "#cccccc": "#c0a0c0", "#e0e0e0": "#d0b0d0", "#4a9eff": "#c060f0", 
        "#e94560": "#a040d0", "#0d1117": "#1e142a", "#161b22": "#2d1e3a", 
        "#21262d": "#3a2a4a", "#f0a500": "#c090f0", "#3fb950": "#a0c0a0"
    },
    "Dark Matte (Low Intensity)": {
        "#0f0f1a": "#121212", "#16213e": "#1a1a1a", "#0d0d1a": "#121212",
        "#1e1e3a": "#1e1e1e", "#131320": "#121212", "#1a1a2e": "#1a1a1a",
        "#2d2d5e": "#2d2d2d", "white": "#b0b0b0", "#ffffff": "#b0b0b0", 
        "#cccccc": "#888888", "#e0e0e0": "#a0a0a0", "#4a9eff": "#5c7c9c", 
        "#e94560": "#8c4a54", "#f0a500": "#8c7c4a", "#3fb950": "#4a8c5c"
    },
    "Midnight Coal": {
        "#0f0f1a": "#151921", "#16213e": "#1c212b", "#0d0d1a": "#151921",
        "#1e1e3a": "#252b36", "#131320": "#151921", "#1a1a2e": "#1c212b",
        "#2d2d5e": "#363e4d", "white": "#c0c5ce", "#ffffff": "#c0c5ce", 
        "#cccccc": "#a7adba", "#e0e0e0": "#bfc1c2", "#4a9eff": "#5f81a1", 
        "#e94560": "#a15f6b", "#f0a500": "#a1925f", "#3fb950": "#5fa16b"
    },
    "Flat Charcoal": {
        "#0f0f1a": "#1c1c1c", "#16213e": "#242424", "#0d0d1a": "#1c1c1c",
        "#1e1e3a": "#2a2a2a", "#131320": "#1c1c1c", "#1a1a2e": "#242424",
        "#2d2d5e": "#333333", "white": "#d1d1d1", "#ffffff": "#d1d1d1", 
        "#cccccc": "#999999", "#e0e0e0": "#b3b3b3", "#4a9eff": "#707070", 
        "#e94560": "#806060", "#f0a500": "#808060", "#3fb950": "#608060"
    },
    "Deep Forest (Muted)": {
        "#0f0f1a": "#0d140d", "#16213e": "#141a14", "#0d0d1a": "#0d140d",
        "#1e1e3a": "#1a241a", "#131320": "#0d140d", "#1a1a2e": "#141a14",
        "#2d2d5e": "#243324", "white": "#c8d1c8", "#ffffff": "#c8d1c8", 
        "#cccccc": "#9ca89c", "#e0e0e0": "#b0bab0", "#4a9eff": "#5f8c5f", 
        "#e94560": "#8c5f5f", "#f0a500": "#8c815f", "#3fb950": "#2ea043"
    },
    "Obsidian Purple": {
        "#0f0f1a": "#140d1f", "#16213e": "#1a1224", "#0d0d1a": "#140d1f",
        "#1e1e3a": "#241a2e", "#131320": "#140d1f", "#1a1a2e": "#1a1224",
        "#2d2d5e": "#332440", "white": "#cfc8d1", "#ffffff": "#cfc8d1", 
        "#cccccc": "#a49ca8", "#e0e0e0": "#b6b0ba", "#4a9eff": "#7a5f8c", 
        "#e94560": "#8c5f7a", "#f0a500": "#8c815f", "#3fb950": "#5f8c6b"
    },
    "Soft Grey Light": {
        "#0f0f1a": "#e8e8e8", "#16213e": "#dadada", "#0d0d1a": "#e0e0e0",
        "#1e1e3a": "#ffffff", "#131320": "#e8e8e8", "#1a1a2e": "#dadada",
        "#2d2d5e": "#b0b0b0", "white": "#333333", "#ffffff": "#333333", 
        "#cccccc": "#555555", "#e0e0e0": "#444444", "#4a9eff": "#5c7c9c", 
        "#e94560": "#8c4a54", "#f0a500": "#8c7c4a", "#3fb950": "#4a8c5c"
    }
}




def apply_global_theme_to_widget(w):
    active_map = THEME_PALETTES.get(GLOBAL_THEME_NAME, {})
    def _apply(w_node):
        if not hasattr(w_node, '_orig_colors'):
            w_node._orig_colors = {}
            for opt in ["bg", "fg", "insertbackground", "selectcolor"]:
                try: w_node._orig_colors[opt] = str(w_node.cget(opt))
                except: pass
        for opt, orig_val in w_node._orig_colors.items():
            key = str(orig_val)
            if key in active_map:
                try: w_node.configure(**{opt: active_map[key]}); continue
                except: pass
            try: w_node.configure(**{opt: key})
            except: pass
        for child in w_node.winfo_children(): _apply(child)
    _apply(w)

_orig_toplevel_init = tk.Toplevel.__init__
def _custom_toplevel_init(self, *args, **kw):
    _orig_toplevel_init(self, *args, **kw)
    self.after(20, lambda: apply_global_theme_to_widget(self))
    
    # v6.2.5: Append a floating Pin/Lock toggle to all dialog panels (Globally persistent state)
    def _create_pin_button():
        if not self.winfo_exists(): return
        
        # Read global lock state from the main UI root
        app_root = self.nametowidget(".")
        is_pinned = getattr(app_root, "_global_dialog_pinned", False)
        
        # v7.1.3: Apply persistent geometry per-class if pinned
        if is_pinned:
            geoms = getattr(app_root, "_global_dialog_geoms", {})
            my_geom = geoms.get(self.__class__.__name__)
            if my_geom: self.geometry(my_geom)
            self.attributes("-topmost", True)
        
        pin_btn = tk.Button(self, text="🔒" if is_pinned else "🔓", 
                            font=("Segoe UI", 9), 
                            bg="#0a0a1a" if is_pinned else "#1e1e3a", 
                            fg="#4a9eff" if is_pinned else "#ffffff", 
                            relief="ridge", cursor="hand2", activebackground="#2d2d5e",
                            bd=1, padx=4, pady=2)
                            
        def toggle():
            if not self.winfo_exists(): return
            current_state = getattr(app_root, "_global_dialog_pinned", False)
            new_state = not current_state
            app_root._global_dialog_pinned = new_state
            
            # Apply state visually to the CURRENT window
            self.attributes("-topmost", new_state)
            pin_btn.config(text="🔒" if new_state else "🔓", 
                           fg="#4a9eff" if new_state else "#ffffff",
                           bg="#0a0a1a" if new_state else "#1e1e3a")
                           
            # Save geometry if locking
            if new_state:
                geoms = getattr(app_root, "_global_dialog_geoms", {})
                geoms[self.__class__.__name__] = self.geometry()
                app_root._global_dialog_geoms = geoms

        def _save_geom_on_config(event):
            # If pinned, track user resizing
            if getattr(app_root, "_global_dialog_pinned", False):
                if event.widget == self:
                    geoms = getattr(app_root, "_global_dialog_geoms", {})
                    geoms[self.__class__.__name__] = self.geometry()
                    app_root._global_dialog_geoms = geoms

        pin_btn.config(command=toggle)
        self.bind("<Configure>", _save_geom_on_config, add="+")
        
        pin_btn.place(relx=1.0, rely=0.0, anchor="ne", x=-25, y=8)
        pin_btn.lift()
        
    self.after(100, _create_pin_button)
    
tk.Toplevel.__init__ = _custom_toplevel_init



# ═══════════════════════════════════════════════════════════
#  UTILITIES
# ═══════════════════════════════════════════════════════════

def extract_sid(val):
    s = str(val).strip()
    parts = re.findall(r'\d{10,}', s)
    return max(parts, key=len) if parts else s

def strip_scr(inv_no):
    """Remove trailing SCR suffix(es) from an invoice number."""
    return re.sub(r'(?:SCR)+$', '', str(inv_no).strip(), flags=re.IGNORECASE)

def safe_num(val):
    try:
        if pd.isna(val): return 0.0
    except: pass
    try:
        v_str = str(val).replace(',', '').strip()
        if not v_str: return 0.0
        return float(v_str)
    except: return 0.0

def clean(val):
    try:
        if pd.isna(val): return ""
    except: pass
    s = str(val).strip()
    if s.endswith('.0'): s = s[:-2]
    return s

def split_comma(val):
    if not val: return []
    try:
        if pd.isna(val): return []
    except: pass
    return [s.strip() for s in str(val).split(',') if s.strip()]

def fmt_qty(val):
    n = safe_num(val)
    if n == 0: return ""
    return str(int(n)) if n == int(n) else str(n)

# ── FIX 2 helper ──────────────────────────────────────────────────────────────
# Keywords whose presence in a remark means it must NEVER be overwritten.
_REMARK_PRESERVE = (
    'Phase 1', 'Direct Shortage', 'REBNI Available', 'Shipment-level REBNI',
    'SR', 'Phase 4', 'No Invoice Search', 'short received directly',
)

def _remark_overwritable(rem):
    """Return True only when none of the preserved keywords appear in rem."""
    return not any(kw in rem for kw in _REMARK_PRESERVE)


# ═══════════════════════════════════════════════════════════
#  DIALOGS
# ═══════════════════════════════════════════════════════════

class HeaderCorrectionDialog(tk.Toplevel):
    """Shown when claims file has non-standard column headers."""
    def __init__(self, parent, corrections, mapping, df_columns, callback):
        super().__init__(parent)
        self.callback    = callback
        self.corrections = corrections
        self.mapping     = mapping
        self.df_columns  = df_columns

        # v6.2.5 Fix: Updated title version
        self.title("Column Header Mismatch Detected — v6.2.5")
        self.geometry("700x480")
        self.configure(bg="#0f0f1a")
        self.resizable(True, True)
        self.lift(); self.focus_force()

        tk.Label(self,
                 text="⚠  Non-standard column headers detected in Claims file",
                 bg="#16213e", fg="#f0a500",
                 font=("Segoe UI", 12, "bold"), height=2).pack(fill="x")
        tk.Label(self,
                 text="The tool has automatically matched the columns below. "
                      "Please confirm or correct before proceeding.",
                 bg="#0f0f1a", fg="#cccccc",
                 font=("Segoe UI", 9)).pack(pady=4)

        outer = tk.Frame(self, bg="#0f0f1a"); outer.pack(fill="both", expand=True, padx=16, pady=6)

        hdrs = ["Field", "Expected", "Found in file", "Status"]
        w    = [14, 20, 28, 12]
        for ci, (h, ww) in enumerate(zip(hdrs, w)):
            tk.Label(outer, text=h, bg="#203864", fg="white",
                     font=("Calibri", 10, "bold"),
                     width=ww, anchor="w", padx=4
                     ).grid(row=0, column=ci, padx=1, pady=1, sticky="w")

        all_fields       = list(COLUMN_ALIASES.keys())
        corrected_fields = {c[0] for c in corrections}
        self._override_vars = {}

        for ri, field in enumerate(all_fields, 1):
            canonical = COLUMN_ALIASES[field][0]
            found_col = mapping.get(field, "")
            is_corrected = field in corrected_fields
            is_missing   = field not in mapping

            if is_missing:
                status_txt, status_fg, row_bg = "MISSING",   "#ff4444", "#2a0000"
            elif is_corrected:
                status_txt, status_fg, row_bg = "Auto-fixed", "#f0a500", "#1a1500"
            else:
                status_txt, status_fg, row_bg = "✔ OK",      "#44ff88", "#001a00"

            tk.Label(outer, text=field, bg=row_bg, fg="#e0e0e0",
                     font=("Calibri", 10, "bold"), width=14, anchor="w", padx=4
                     ).grid(row=ri, column=0, padx=1, pady=1, sticky="w")
            tk.Label(outer, text=canonical, bg=row_bg, fg="#aaaacc",
                     font=("Calibri", 10), width=20, anchor="w", padx=4
                     ).grid(row=ri, column=1, padx=1, pady=1, sticky="w")

            v = tk.StringVar(value=found_col or "— not found —")
            self._override_vars[field] = v
            opts = ["— not found —"] + list(df_columns)
            cb = ttk.Combobox(outer, textvariable=v, values=opts,
                              state="readonly", width=26, font=("Calibri", 9))
            cb.grid(row=ri, column=2, padx=1, pady=1, sticky="w")

            tk.Label(outer, text=status_txt, bg=row_bg, fg=status_fg,
                     font=("Calibri", 10, "bold"), width=12, anchor="w", padx=4
                     ).grid(row=ri, column=3, padx=1, pady=1, sticky="w")

        tk.Label(self,
                 text="You can change any column assignment using the dropdowns above.",
                 bg="#0f0f1a", fg="#888899", font=("Segoe UI", 8)).pack(pady=2)

        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=10)
        tk.Button(bf, text="✔  Auto-correct & Proceed",
                  command=self._proceed,
                  bg="#2d6a4f", fg="white",
                  font=("Segoe UI", 12, "bold"),
                  padx=20, pady=8, relief="flat",
                  cursor="hand2").pack(side="left", padx=10)
        tk.Button(bf, text="✖  Cancel",
                  command=self._cancel,
                  bg="#4a2020", fg="white",
                  font=("Segoe UI", 11),
                  padx=16, pady=8, relief="flat",
                  cursor="hand2").pack(side="left", padx=10)

        self.protocol("WM_DELETE_WINDOW", self._cancel)
        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _proceed(self):
        final = {}
        for field, var in self._override_vars.items():
            v = var.get()
            if v and v != "— not found —":
                final[field] = v
        self.callback({'action': 'proceed', 'mapping': final})
        self.destroy()

    def _cancel(self):
        self.callback({'action': 'cancel'})
        self.destroy()


class SIDRequestDialog(tk.Toplevel):
    """Auto mode: SID not found in REBNI → ask user."""
    def __init__(self, parent, invoice, po, asin, callback):
        super().__init__(parent)
        self.callback = callback
        self.title("SID Required — DICES Validation")
        self.geometry("540x210")
        self.resizable(True, True)
        self.configure(bg="#16213e")
        self.lift(); self.focus_force()

        tk.Label(self, text="⚠  SID Not Found in REBNI",
                 bg="#16213e", fg="#e94560",
                 font=("Segoe UI", 13, "bold")).pack(pady=(14, 4))
        tk.Label(self, text=f"Invoice: {invoice}   PO: {po}   ASIN: {asin}",
                 bg="#16213e", fg="#e0e0e0", font=("Segoe UI", 9)).pack(pady=2)
        tk.Label(self, text="Validate this invoice in DICES and enter the SID below:",
                 bg="#16213e", fg="#aaaacc", font=("Segoe UI", 9)).pack(pady=6)

        ef = tk.Frame(self, bg="#16213e"); ef.pack()
        tk.Label(ef, text="SID from DICES:", bg="#16213e", fg="#e0e0e0",
                 font=("Segoe UI", 10)).pack(side="left", padx=8)
        self._sid = tk.StringVar()
        self._entry = tk.Entry(ef, textvariable=self._sid, width=30,
                               font=("Segoe UI", 10), bg="#1e1e3a", fg="#e0e0e0",
                               insertbackground="white", relief="flat")
        self._entry.pack(side="left", padx=4)
        self._entry.focus_set()

        bf = tk.Frame(self, bg="#16213e"); bf.pack(pady=12)
        tk.Button(bf, text="✔  Continue", command=self._ok,
                  bg="#2d6a4f", fg="white", font=("Segoe UI", 11, "bold"),
                  padx=16, pady=7, relief="flat", cursor="hand2").pack(side="left", padx=8)
        tk.Button(bf, text="✖  Skip", command=self._skip,
                  bg="#6b2737", fg="white", font=("Segoe UI", 10),
                  padx=16, pady=7, relief="flat", cursor="hand2").pack(side="left", padx=8)

        self.bind('<Return>', lambda e: self._ok())
        self.protocol("WM_DELETE_WINDOW", self._skip)
        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _ok(self):
        sid = extract_sid(self._sid.get().strip())
        if sid:
            self.callback(sid); self.destroy()
        else:
            self._entry.config(bg="#3a1e1e")

    def _skip(self):
        self.callback(None); self.destroy()

class CrossPODialog(tk.Toplevel):
    # v6.1: Added automated Cross PO Analysis with engine support
    """Cross PO Confirmation Dialog — v5.3.0 logic restored for manual selection"""
    CASE_DESCRIPTIONS = {
        "Case 0": (
            "Case 0 — Non-overage (Rec <= Inv)",
            "All received units are covered by the invoice at this PO.\n"
            "There is no overage to investigate. You should SKIP this."
        ),
        "Case 1": (
            "Case 1 — No PO, but ASIN received",
            "Rec=0 at claiming PO. Same ASIN received in different PO within same SID.\n"
            "Those units are overage under a different PO."
        ),
        "Case 2": (
            "Case 2 — PO exists but ASIN not ordered there",
            "This PO exists in the claiming SID, but the ASIN was never invoiced there.\n"
            "Inv Qty = 0, but units were received. This is a Cross PO overage."
        ),
        "Case 3": (
            "Case 3 — PO and ASIN exist but Rec > Inv",
            "Both PO and ASIN are present. Invoiced qty = X, but received more than X.\n"
            "Excess units are Cross PO overage."
        ),
    }

    def __init__(self, parent, candidates, current_inv, sid, callback, engine=None):
        super().__init__(parent)
        self.callback   = callback
        self.candidates = [dict(c) for c in candidates]  # v6.1: mutable copy so ANALYZE can update inv_qty/cross_type
        self.sid        = sid
        self.engine     = engine  # v6.1: engine reference for _resolve_inv_qty lookups
        self.file_path  = tk.StringVar()

        self.title("Cross PO Overage — Confirm & Investigate")
        self.geometry("740x540")
        self.resizable(True, True)
        self.configure(bg="#0f0f1a")
        self.protocol("WM_DELETE_WINDOW", self.withdraw)
        self.lift(); self.focus_force()

        tk.Label(self, text="🔄  Cross PO Overage Detected",
                 bg="#16213e", fg="#f0a500",
                 font=("Segoe UI", 13, "bold"), height=2).pack(fill="x")

        # Top Shortcut Buttons
        top_btn_frame = tk.Frame(self, bg="#0f0f1a")
        top_btn_frame.pack(pady=4)
        tk.Button(top_btn_frame, text="Confirm", bg="#2d6a4f", fg="white", font=("Segoe UI", 9, "bold"), padx=12, relief="flat", cursor="hand2", command=self._confirm).pack(side="left", padx=5)
        tk.Button(top_btn_frame, text="Skip", bg="#4a2020", fg="white", font=("Segoe UI", 9, "bold"), padx=12, relief="flat", cursor="hand2", command=self._skip).pack(side="left", padx=5)

        tk.Label(self,
                 text=f"SID: {sid}   |   Investigation Invoice: {current_inv}",
                 bg="#0f0f1a", fg="#cccccc",
                 font=("Segoe UI", 9)).pack(pady=2)
        tk.Label(self,
                 text="On confirming, the tool will investigate the Cross PO chain "
                      "to find equivalent shortage.",
                 bg="#0f0f1a", fg="#4a9eff",
                 font=("Segoe UI", 9)).pack(pady=2)

        self.table_wrapper = tk.LabelFrame(self, text="  Detected Cross PO Candidates  ",
                           bg="#0f0f1a", fg="#e0e0e0",
                           font=("Segoe UI", 9, "bold"), padx=10, pady=6)
        self.table_wrapper.pack(fill="x", padx=16, pady=6)
        
        self.table_canvas = tk.Canvas(self.table_wrapper, bg="#0f0f1a", highlightthickness=0, height=110)
        self.table_scroll = ttk.Scrollbar(self.table_wrapper, orient="vertical", command=self.table_canvas.yview)
        self.table_frame = tk.Frame(self.table_canvas, bg="#0f0f1a")
        
        self.table_frame.bind("<Configure>", lambda e: self.table_canvas.configure(scrollregion=self.table_canvas.bbox("all")))
        self.table_canvas.create_window((0, 0), window=self.table_frame, anchor="nw")
        self.table_canvas.configure(yscrollcommand=self.table_scroll.set)
        
        self.table_canvas.pack(side="left", fill="both", expand=True)
        self.table_scroll.pack(side="right", fill="y")
        # v6_Secured_Remedy_v6.1: Cross PO SID Details attachment UI directly in panel
        uf = tk.LabelFrame(self, text='  Cross PO SID Details attachment  ', bg='#0f0f1a', fg='#4a9eff', font=('Segoe UI', 9, 'bold'), padx=10, pady=6)
        uf.pack(fill='x', padx=16, pady=4)
        tk.Label(uf, text='Attach SID Details:', bg='#0f0f1a', fg='#cccccc', font=('Segoe UI', 9)).pack(side='left', padx=4)
        tk.Entry(uf, textvariable=self.file_path, width=50, bg='#1e1e3a', fg='white', relief='flat').pack(side='left', padx=6)
        tk.Button(uf, text='Browse', command=lambda: self.file_path.set(filedialog.askopenfilename(filetypes=[('Excel / CSV', '*.xlsx *.xls *.csv')])), bg='#2d2d5e', fg='white', relief='flat', cursor='hand2').pack(side='left', padx=4)
        tk.Button(uf, text='🔍 ANALYZE', command=self._analyze_cross_po_file, bg='#6f42c1', fg='white', font=('Segoe UI', 9, 'bold'), relief='flat', cursor='hand2', padx=12).pack(side='right', padx=4)
        
        self._render_candidates_table()

        sf = tk.Frame(self, bg="#0f0f1a"); sf.pack(fill="x", padx=16, pady=4)
        tk.Label(sf, text="Select Cross PO to investigate:",
                 bg="#0f0f1a", fg="#e0e0e0",
                 font=("Segoe UI", 10), width=30, anchor="w").pack(side="left")
        opts = [f"PO={c['po']}  Rec={fmt_qty(c['rec_qty'])}  {c['cross_type'].split(chr(8212))[0].strip()}"
                for c in candidates] + ["None — Skip"]
        self._sel_var = tk.StringVar()

        self._sel_cb  = ttk.Combobox(sf, textvariable=self._sel_var,
                                      values=opts, state="readonly", width=50,
                                      font=("Segoe UI", 9))
        self._sel_cb.current(0)
        self._sel_cb.pack(side="left", padx=6)
        self._sel_cb.bind("<<ComboboxSelected>>", self._on_candidate_change)

        cf = tk.LabelFrame(self, text="  Confirm Cross PO Case  ",
                           bg="#0f0f1a", fg="#e0e0e0",
                           font=("Segoe UI", 9, "bold"), padx=12, pady=8)
        cf.pack(fill="x", padx=16, pady=4)

        self._case_var = tk.StringVar(value="Case 1")
        self._case_desc_lbl = tk.Label(cf, text="",
                                        bg="#0f0f1a", fg="#aaaacc",
                                        font=("Segoe UI", 9), justify="left",
                                        wraplength=640, anchor="w")

        for case_key, (case_label, _) in self.CASE_DESCRIPTIONS.items():
            tk.Radiobutton(cf, text=case_label,
                           variable=self._case_var, value=case_key,
                           bg="#0f0f1a", fg="#f0c060",
                           selectcolor="#1a1500",
                           font=("Segoe UI", 10, "bold"),
                           command=self._on_case_change
                           ).pack(anchor="w", pady=2)

        self._case_desc_lbl.pack(anchor="w", pady=4, padx=8)
        self._on_case_change()

        tk.Label(self,
                 text="⚡  On confirming: tool will investigate Cross PO chain "
                      "until full Cross PO rec_qty is explained as shortage.",
                 bg="#0f0f1a", fg="#88ccff",
                 font=("Segoe UI", 9, "italic")).pack(pady=4, padx=16, anchor="w")

        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=10)
        tk.Button(bf, text="✔  Confirm & Investigate",
                  command=self._confirm,
                  bg="#2d6a4f", fg="white",
                  font=("Segoe UI", 12, "bold"),
                  padx=20, pady=9, relief="flat",
                  cursor="hand2").pack(side="left", padx=10)
        tk.Button(bf, text="✖  Skip",
                  command=self._skip,
                  bg="#4a2020", fg="white",
                  font=("Segoe UI", 11),
                  padx=16, pady=9, relief="flat",
                  cursor="hand2").pack(side="left", padx=10)

        self.protocol("WM_DELETE_WINDOW", self._skip)
        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")


    def _analyze_cross_po_file(self):
        """v7.1.0: Automated Cross PO analysis using internal Vendor Data first, then fallback to file."""
        findings = []
        updated_count = 0
        target_sid = clean(self.sid)

        # 1) Internal Lookup from Vendor Data (High Speed)
        # v7.2.3: Two-step match. Bridge (SID, PO) -> Lookup (Inv, ASIN)
        if self.engine and self.engine.vendor_raw:
            raw_rows = self.engine.vendor_raw.get(target_sid, [])
            if raw_rows:
                # Find column indices once
                h = self.engine.vendor_headers
                def get_idx(keys):
                    for i, col in enumerate(h):
                        if any(k.lower() in str(col).lower() for k in keys): return i
                    return -1
                
                i_po = get_idx(['po number', 'purchase order', 'po'])
                i_inv = get_idx(['invoice number', 'invoice'])
                i_bc = get_idx(['barcode', 'fnsku'])
                
                new_cands = []
                for c in self.candidates:
                    matches = []
                    seen_invoices = set()
                    for r in raw_rows:
                        r_po = clean(str(r[i_po])) if i_po != -1 else ""
                        if r_po == clean(c['po']):
                            r_inv = clean(str(r[i_inv])) if i_inv != -1 else ""
                            if r_inv and r_inv not in seen_invoices:
                                # Authoritative Quantity Lookup from Invoice Search file
                                iq = self.engine._resolve_inv_qty(r_inv, c['asin'], None)
                                if iq is not None and iq > 0:
                                    r_bc = clean(str(r[i_bc])) if i_bc != -1 else ""
                                    matches.append({'inv': r_inv, 'bc': r_bc, 'qty': iq})
                                    seen_invoices.add(r_inv)
                    
                    if matches:
                        # Update first match into existing candidate
                        c['found_inv'] = matches[0]['inv']
                        c['found_bc'] = matches[0]['bc']
                        c['inv_qty'] = matches[0]['qty']
                        if matches[0]['qty'] >= safe_num(c['rec_qty']):
                            c['cross_type'] = 'Case 0 — Verified (Not Overage)'
                        elif matches[0]['qty'] > 0:
                            c['cross_type'] = 'Case 3 — Rec qty > Inv qty (overage in cross PO)'
                        else:
                            c['cross_type'] = 'Case 1 — Rec=0 at current PO, units received here'
                        updated_count += 1
                        findings.append(f"• [VENDOR DATA] PO {c['po']}: Found {len(matches)} matching invoice(s).")
                        
                        # Add additional invoices as new candidates (if any)
                        for m in matches[1:]:
                            new_c = c.copy()
                            new_c['found_inv'] = m['inv']
                            new_c['found_bc'] = m['bc']
                            new_c['inv_qty'] = m['qty']
                            if m['qty'] >= safe_num(new_c['rec_qty']):
                                new_c['cross_type'] = 'Case 0 — Verified (Not Overage)'
                            elif m['qty'] > 0:
                                new_c['cross_type'] = 'Case 3 — Rec qty > Inv qty (overage in cross PO)'
                            else:
                                new_c['cross_type'] = 'Case 1 — Rec=0 at current PO, units received here'
                            new_cands.append(new_c)
                    else:
                        c['inv_qty'] = 0
                        c['cross_type'] = 'Case 1 — Rec=0 at current PO, units received here'
                
                if new_cands: self.candidates.extend(new_cands)

        if updated_count == len(self.candidates) and updated_count > 0:
            self.after(0, lambda: messagebox.showinfo('Analysis Complete', "All candidates automatically verified via Vendor Level Data.", parent=self))
            self._render_candidates_table()
            return

        # 2) Fallback to attached file if provided
        path = self.file_path.get().strip()
        if not path:
            self._render_candidates_table()
            if updated_count > 0:
                return
            messagebox.showwarning('File Required', 'No automated matches found at vendor level. Please attach the Cross PO SID level Excel/CSV first.', parent=self); return
            
        if not os.path.exists(path):
            messagebox.showerror('Error', 'Attached file path does not exist.', parent=self); return

        try:
            if path.lower().endswith('.csv'):
                df = pd.read_csv(path, header=0, encoding='utf-8', on_bad_lines='skip')
            else:
                df = pd.read_excel(path, header=0)
            df.columns = [str(c).strip().upper() for c in df.columns]
            sid_col = next((c for c in df.columns if 'SHIPMENT' in c or 'SID' in c), None)
            po_col  = next((c for c in df.columns if 'PO' in c), None)
            inv_col = next((c for c in df.columns if 'INVOICE' in c), None)
            bc_col  = next((c for c in df.columns if 'BARCODE' in c or 'FNSKU' in c or 'UPC' in c), None)
            
            if not all([sid_col, po_col, inv_col]):
                messagebox.showerror('Format Error', f'Excel must have Shipment ID, PO, and Invoice Number columns.', parent=self); return
            
            findings = []
            updated_count = 0
            target_sid = clean(self.sid)
            for c in self.candidates:
                # Find all rows matching SID and PO
                matches = df[(df[sid_col].astype(str).str.contains(target_sid)) & (df[po_col].astype(str) == c['po'])]
                if not matches.empty:
                    # v6.2.5: Iterate through all matching invoices until one with valid quantity is found
                    best_match = None
                    for _, row in matches.iterrows():
                        found_inv = clean(str(row[inv_col]))
                        if self.engine:
                            iq = self.engine._resolve_inv_qty(found_inv, c['asin'], None)
                            if iq is not None and iq > 0:
                                best_match = (found_inv, iq, clean(str(row[bc_col])) if bc_col else "")
                                break
                    
                    if best_match:
                        found_inv, iq, found_bc = best_match
                        c['found_inv'] = found_inv
                        c['found_bc']  = found_bc
                        c['inv_qty']   = iq
                        if iq >= rec_n: 
                            c['cross_type'] = 'Case 0 — Verified (Not Overage)'
                        elif iq > 0:
                            c['cross_type'] = 'Case 3 — Rec qty > Inv qty (overage in cross PO)'
                        else:
                            c['cross_type'] = 'Case 1 — Rec=0 at current PO, units received here'
                        findings.append(f"• PO {c['po']}: Found Inv {found_inv} (Qty {int(iq)}) | Barcode: {found_bc if found_bc else 'N/A'}")
                        updated_count += 1
                    else:
                        # Fallback to reporting mismatch, but strictly zero out inv_qty as requested
                        first_inv = clean(str(matches.iloc[0][inv_col]))
                        c['found_inv'] = first_inv
                        c['found_bc']  = clean(str(matches.iloc[0][bc_col])) if bc_col else ""
                        c['inv_qty']   = 0  # v6.2.6: Fix - Replace PO level qty with 0 if no DB match found
                        c['cross_type'] = 'Case 1 — Rec=0 at current PO, units received here' # Bugfix: update stale cross_type
                        findings.append(f"• PO {c['po']}: Multiple invoices found, but none matched quantity in database (Reset to 0).")
                else:
                    c['inv_qty'] = 0  # v6.2.6: Fix - Zero out if PO not in attached file
                    c['cross_type'] = 'Case 1 — Rec=0 at current PO, units received here' # Bugfix: update stale cross_type
                    findings.append(f"• PO {c['po']}: Not found in attached file (Reset to 0)")
            
            # v6.2.5: Dynamically refresh the UI grid immediately
            self._render_candidates_table()
            
            report = "\n".join(findings) if findings else "No candidates processed."
            messagebox.showinfo('Analysis Results', f'Found {updated_count} valid matches in file:\n\n{report}', parent=self)
            
            opts = [f"PO={c['po']}  Rec={fmt_qty(c['rec_qty'])}  {c['cross_type'].split('—')[0].strip()}" for c in self.candidates] + ["None — Skip"]
            self._sel_cb['values'] = opts
        except Exception as e:
            messagebox.showerror('Error', f'Analysis failed: {e}', parent=self)

    def _render_candidates_table(self):
        # v6.1.2: Clear and redraw the table dynamically
        for child in self.table_frame.winfo_children():
            child.destroy()
            
        for ci, h in enumerate(["Cross PO", "ASIN", "Inv Qty", "Rec Qty", "Overage", "Type"]):
            tk.Label(self.table_frame, text=h, bg="#203864", fg="white",
                     font=("Calibri", 10, "bold"), width=14, anchor="w", padx=3
                     ).grid(row=0, column=ci, padx=1, pady=1)
        for ri, c in enumerate(self.candidates, 1):
            inv_n = safe_num(c.get('inv_qty', 0))
            rec_n = safe_num(c['rec_qty'])
            ovg   = max(0.0, rec_n - inv_n)
            for ci, v in enumerate([c['po'], c['asin'],
                                     fmt_qty(inv_n), fmt_qty(rec_n),
                                     fmt_qty(ovg) or "—",
                                     c['cross_type'].split("—")[0].strip()]):
                tk.Label(self.table_frame, text=str(v), bg="#1e1e3a", fg="#e0e0e0",
                         font=("Calibri", 10), width=14, anchor="w", padx=3
                         ).grid(row=ri, column=ci, padx=1, pady=1)

    def _on_candidate_change(self, event=None):
        pass

    def _on_case_change(self):
        case_key = self._case_var.get()
        _, desc  = self.CASE_DESCRIPTIONS.get(case_key, ("", ""))
        self._case_desc_lbl.config(text=desc)

    def _confirm(self):
        if self._case_var.get() == "Case 0":
            messagebox.showwarning("Non-Overage", "Case 0 is a non-overage and cannot be investigated.\nPlease select 'Skip' or choose a valid overage case.")
            return

        idx = self._sel_cb.current()
        if idx >= len(self.candidates):
            self.callback({'action': 'skip'}); self.destroy(); return
            
        selected_po = self.candidates[idx]['po']
        
        # v6.2.6: Automated Vault Persistence - Store ALL non-selected GENUINE overages (Case 1, 2, 3)
        if self.engine and hasattr(self.engine, 'cross_po_vault'):
            if self.sid not in self.engine.cross_po_vault:
                self.engine.cross_po_vault[self.sid] = []
            
            for c in self.candidates:
                # v6.2.6.1: Only store candidates that are NOT selected and are GENUINE overages (Case 1, 2, 3)
                is_genuine = any(k in c.get('cross_type', '') for k in ["Case 1", "Case 2", "Case 3"])
                if c['po'] != selected_po and is_genuine:
                    # Check if already in vault to avoid duplicates
                    if not any(v['candidate']['po'] == c['po'] and v['candidate']['asin'] == c['asin'] for v in self.engine.cross_po_vault[self.sid]):
                        inv_n = safe_num(c.get('inv_qty', 0))
                        rec_n = safe_num(c['rec_qty'])
                        overage = max(0.0, rec_n - inv_n)
                        
                        # Store in the format expected by the Vault UI
                        self.engine.cross_po_vault[self.sid].append({
                            'candidate': dict(c),
                            'case': c['cross_type'].split(chr(8212))[0].strip(), # "Case X"
                            'budget': overage
                        })
        
        self.callback({
            'action'   : 'confirmed',
            'candidate': self.candidates[idx],
            'case'     : self._case_var.get(),
        })
        self.destroy()

    def _skip(self):
        self.callback({'action': 'skip'}); self.destroy()

class CrossPOVaultDialog(tk.Toplevel):
    """v6.2.6: Vault to manage and investigate Cross POs across multiple shipments."""
    def __init__(self, parent, engine, callback):
        super().__init__(parent)
        self.engine = engine
        self.callback = callback
        self.title("Cross PO Vault — v6.2.6")
        self.geometry("760x520")
        self.configure(bg="#0f111a")
        self.resizable(True, True)
        self.lift(); self.focus_force()

        header = tk.Frame(self, bg="#16213e", height=60)
        header.pack(fill="x")
        tk.Label(header, text="📦  Cross PO Vault", bg="#16213e", fg="#f0a500", font=("Segoe UI", 14, "bold")).pack(side="left", padx=20, pady=12)
        tk.Label(header, text="v6.2.6 | Secured Edition", bg="#16213e", fg="#4a9eff", font=("Segoe UI", 9)).pack(side="right", padx=20)

        ctrl = tk.Frame(self, bg="#0f111a", padx=20, pady=15)
        ctrl.pack(fill="x")
        tk.Label(ctrl, text="Select Shipment ID:", bg="#0f111a", fg="#cccccc", font=("Segoe UI", 10)).pack(side="left")
        
        self.sid_var = tk.StringVar()
        sids = sorted(list(self.engine.cross_po_vault.keys()))
        self.sid_combo = ttk.Combobox(ctrl, textvariable=self.sid_var, values=sids, state="readonly", width=30)
        self.sid_combo.pack(side="left", padx=10)
        self.sid_combo.bind("<<ComboboxSelected>>", lambda e: self._refresh_list())
        
        tk.Button(ctrl, text="🔄 REFRESH", command=self._refresh_list, bg="#1e1e3a", fg="white", relief="flat", padx=10).pack(side="left", padx=5)
        tk.Button(ctrl, text="🗑️ CLEAR VAULT", command=self._clear_vault, bg="#4a2020", fg="white", relief="flat", padx=10).pack(side="right")

        self.list_frame = tk.Frame(self, bg="#0d0d1a", padx=20, pady=5)
        self.list_frame.pack(fill="both", expand=True)
        
        self.canvas = tk.Canvas(self.list_frame, bg="#0d0d1a", highlightthickness=0)
        self.scroll = ttk.Scrollbar(self.list_frame, orient="vertical", command=self.canvas.yview)
        self.scroll_frame = tk.Frame(self.canvas, bg="#0d0d1a")
        
        self.canvas.create_window((0,0), window=self.scroll_frame, anchor="nw", tags="frame")
        self.canvas.configure(yscrollcommand=self.scroll.set)
        
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scroll.pack(side="right", fill="y")
        
        self.scroll_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        
        if sids:
            self.sid_var.set(sids[0])
            self._refresh_list()
        else:
            tk.Label(self.scroll_frame, text="Vault is empty. Stored Cross POs will appear here.", bg="#0d0d1a", fg="#666666", font=("Segoe UI", 10, "italic")).pack(pady=40)

    def _refresh_list(self):
        for w in self.scroll_frame.winfo_children(): w.destroy()
        sid = self.sid_var.get()
        if not sid or sid not in self.engine.cross_po_vault: return
        
        data = self.engine.cross_po_vault[sid]
        if not data:
            tk.Label(self.scroll_frame, text="No genuine overages stored for this SID.", bg="#0d0d1a", fg="#666666", font=("Segoe UI", 10, "italic")).pack(pady=20)
            return

        # Header for the list
        h_f = tk.Frame(self.scroll_frame, bg="#16213e")
        h_f.pack(fill="x", pady=(0, 5))
        headers = ["PO", "ASIN", "Overage Qty", "Case Type", "Actions"]
        widths = [15, 15, 12, 18, 25]
        for idx, text in enumerate(headers):
            tk.Label(h_f, text=text, bg="#16213e", fg="#4a9eff", font=("Segoe UI", 9, "bold"), width=widths[idx], anchor="w", padx=10).pack(side="left")

        for i, item in enumerate(data):
            f = tk.Frame(self.scroll_frame, bg="#1a1a2e", padx=5, pady=8, relief="flat", bd=1)
            f.pack(fill="x", pady=2, padx=2)
            
            c = item.get('candidate', {})
            case_name = item.get('case', 'Unknown')
            budget = item.get('budget', 0)
            
            # Data columns
            tk.Label(f, text=c.get('po', 'N/A'), bg="#1a1a2e", fg="white", font=("Consolas", 10, "bold"), width=15, anchor="w", padx=10).pack(side="left")
            tk.Label(f, text=c.get('asin', 'N/A'), bg="#1a1a2e", fg="#e0e0e0", font=("Consolas", 10), width=15, anchor="w", padx=10).pack(side="left")
            tk.Label(f, text=f"{int(budget)} units", bg="#1a1a2e", fg="#f0a500", font=("Segoe UI", 10, "bold"), width=12, anchor="w", padx=10).pack(side="left")
            tk.Label(f, text=case_name, bg="#1a1a2e", fg="#cccccc", font=("Segoe UI", 9), width=18, anchor="w", padx=10).pack(side="left")
            
            btn_f = tk.Frame(f, bg="#1a1a2e")
            btn_f.pack(side="right", padx=10)
            
            # v6.2.6.1: High contrast investigation trigger
            tk.Button(btn_f, text="🔍 INVESTIGATE FURTHER", 
                      command=lambda it=item: self._investigate(it),
                      bg="#2d6a4f", fg="white", font=("Segoe UI", 8, "bold"), 
                      relief="flat", padx=12, pady=4, cursor="hand2").pack(side="left", padx=5)
            
            tk.Button(btn_f, text="❌", 
                      command=lambda s=sid, idx=i: self._remove_item(s, idx),
                      bg="#3d0000", fg="#ff4d4d", relief="flat", padx=8, cursor="hand2").pack(side="left")

    def _remove_item(self, sid, idx):
        if sid in self.engine.cross_po_vault:
            if idx < len(self.engine.cross_po_vault[sid]):
                self.engine.cross_po_vault[sid].pop(idx)
                if not self.engine.cross_po_vault[sid]:
                    del self.engine.cross_po_vault[sid]
                self._refresh_list()

    def _clear_vault(self):
        if messagebox.askyesno("Clear Vault", "Are you sure you want to clear ALL stored Cross POs across ALL shipments?"):
            self.engine.cross_po_vault.clear()
            self.sid_combo['values'] = []
            self.sid_var.set("")
            self._refresh_list()

    def _investigate(self, item):
        # Trigger the engine investigation for this specific candidate
        # We pass the SID, candidate dict, case type, and the budget (overage)
        c = item['candidate']
        sid = c.get('sid') or self.sid_var.get()
        self.callback(sid, c, item['case'], item['budget'])
        self.destroy()




class ManualLevelDialog(tk.Toplevel):
    """Manual mode: per-level dialog with IBC/PBC + Cross PO + Mismatches."""
    def __init__(self, parent, matches, remaining_pqv, branch_budget, callback,
                 pending_cb=None, engine=None):
        super().__init__(parent)
        self.callback      = callback
        self.matches       = matches
        self.rem_pqv       = remaining_pqv
        self.branch_budget = branch_budget
        self._pending_cb   = pending_cb
        self._engine_ref   = engine

        self.title("Manual Investigation — Next Step")
        self.geometry("680x560")
        self.configure(bg="#0f0f1a")
        self.resizable(True, True)
        self.protocol("WM_DELETE_WINDOW", self.withdraw)
        self.lift(); self.focus_force()

        tk.Label(self, text="  Manual Investigation — Continue",
                 bg="#16213e", fg="#4a9eff",
                 font=("Segoe UI", 12, "bold"), height=2).pack(fill="x")

        info = f"Remaining PQV: {int(remaining_pqv)}    Branch budget: {int(branch_budget)}"
        tk.Label(self, text=info, bg="#0f0f1a", fg="#cccccc",
                 font=("Segoe UI", 9)).pack(pady=2)

        inv_f = tk.LabelFrame(self, text="  Select Invoice to Continue  ",
                              font=("Segoe UI", 9, "bold"),
                              bg="#0f0f1a", fg="#e0e0e0", padx=10, pady=6)
        inv_f.pack(fill="x", padx=16, pady=4)
        opts = [f"Qty={fmt_qty(m['mtc_qty'])}  |  Inv={m['mtc_inv']}  |  PO={m['mtc_po']}  |  ASIN={m['mtc_asin']}"
                for m in matches]
        self._branch_var = tk.StringVar()
        self._branch_cb  = ttk.Combobox(inv_f, textvariable=self._branch_var,
                                         values=opts, state="readonly", width=70,
                                         font=("Segoe UI", 9))
        if opts: self._branch_cb.current(0)
        self._branch_cb.pack()

        ibc_f = tk.LabelFrame(self, text="  IBC = PBC Validation  ",
                               font=("Segoe UI", 9, "bold"),
                               bg="#0f0f1a", fg="#e0e0e0", padx=10, pady=6)
        ibc_f.pack(fill="x", padx=16, pady=4)

        self._validity = tk.StringVar(value="valid")
        rf = tk.Frame(ibc_f, bg="#0f0f1a"); rf.pack(fill="x")
        tk.Radiobutton(rf, text="✔  IBC = PBC  VALID — Continue investigation",
                       variable=self._validity, value="valid",
                       bg="#0f0f1a", fg="#90ee90", selectcolor="#1e3a28",
                       font=("Segoe UI", 10, "bold"),
                       command=self._toggle).pack(side="left", padx=6)
        tk.Radiobutton(rf, text="✗  IBC ≠ PBC  INVALID — Exclude units",
                       variable=self._validity, value="invalid",
                       bg="#0f0f1a", fg="#ff8888", selectcolor="#3a1e1e",
                       font=("Segoe UI", 10, "bold"),
                       command=self._toggle).pack(side="left", padx=14)
        self._invalid_frame = tk.Frame(ibc_f, bg="#0f0f1a")
        self._invalid_frame.pack(fill="x", pady=3)
        tk.Label(self._invalid_frame, text="Units matched to invalid invoice:",
                 bg="#0f0f1a", fg="#ff8888", font=("Segoe UI", 9)).pack(side="left", padx=4)
        self._inv_qty_var = tk.StringVar()
        tk.Entry(self._invalid_frame, textvariable=self._inv_qty_var, width=10,
                 font=("Segoe UI", 10), bg="#1e1e3a", fg="#ff8888",
                 insertbackground="white", relief="flat").pack(side="left", padx=4)

        self._dices_frame = tk.LabelFrame(self, text="  DICES Details  ",
                                           font=("Segoe UI", 9, "bold"),
                                           bg="#0f0f1a", fg="#e0e0e0", padx=10, pady=6)
        self._dices_frame.pack(fill="x", padx=16, pady=4)
        r1 = tk.Frame(self._dices_frame, bg="#0f0f1a"); r1.pack(fill="x", pady=2)
        tk.Label(r1, text="SID from DICES:", bg="#0f0f1a", fg="#e0e0e0",
                 font=("Segoe UI", 9), width=20, anchor="w").pack(side="left")
        self._sid_var = tk.StringVar()
        tk.Entry(r1, textvariable=self._sid_var, width=28, font=("Segoe UI", 9),
                 bg="#1e1e3a", fg="#e0e0e0", insertbackground="white",
                 relief="flat").pack(side="left", padx=4)
        tk.Button(r1, text="Pull Data", command=self._pull_vendor_data,
                  bg="#2d2d5e", fg="#4a9eff", font=("Segoe UI", 8, "bold"),
                  relief="flat", cursor="hand2", padx=6).pack(side="left")
        tk.Button(r1, text="DICES ↗", command=lambda: webbrowser.open(f"https://smocentral.amazon.eu/dices/advanced-search?shipment_id={self._sid_var.get().strip()}") if self._sid_var.get().strip() and self._sid_var.get().strip() != '-' else None,
                  bg="#1f4e79", fg="white", font=("Segoe UI", 8, "bold"),
                  relief="flat", cursor="hand2", padx=6).pack(side="left", padx=4)

        r2 = tk.Frame(self._dices_frame, bg="#0f0f1a"); r2.pack(fill="x", pady=2)
        tk.Label(r2, text="Barcode from DICES:", bg="#0f0f1a", fg="#e0e0e0",
                 font=("Segoe UI", 9), width=20, anchor="w").pack(side="left")
        self._bc_var = tk.StringVar()
        tk.Entry(r2, textvariable=self._bc_var, width=28, font=("Segoe UI", 9),
                 bg="#1e1e3a", fg="#e0e0e0", insertbackground="white",
                 relief="flat").pack(side="left", padx=4)
        tk.Button(r2, text="Pull Data", command=self._pull_vendor_data,
                  bg="#2d2d5e", fg="#4a9eff", font=("Segoe UI", 8, "bold"),
                  relief="flat", cursor="hand2", padx=6).pack(side="left")
        tk.Button(r2, text="DICES ↗", command=lambda: webbrowser.open(f"https://smocentral.amazon.eu/dices/document?documentID={self._bc_var.get().strip()}") if self._bc_var.get().strip() and self._bc_var.get().strip() != '-' else None,
                  bg="#1f4e79", fg="white", font=("Segoe UI", 8, "bold"),
                  relief="flat", cursor="hand2", padx=6).pack(side="left", padx=4)

        # v7.1.0: IBC vs PBC Validation Button & Display Frame
        tk.Button(self._dices_frame, text="🔍 IBC vs PBC VALIDATION", 
                  command=self._show_ibc_validation, bg="#3a3a5e", fg="#4a9eff", 
                  font=("Segoe UI", 9, "bold"), relief="flat", cursor="hand2", padx=10).pack(pady=(8,0))
        
        self._val_frame = tk.Frame(self._dices_frame, bg="#0f0f1a")
        self._val_frame.pack(fill="x", pady=4)
        from tkinter.scrolledtext import ScrolledText
        self._val_display = ScrolledText(self._val_frame, height=8, bg="#05050a", fg="#4a9eff", font=("Consolas", 9), relief="flat")
        # Start hidden
        self._val_display.pack_forget()

        self._toggle()
        self._auto_populate()

        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=6)
        tk.Button(bf, text="▶  CONTINUE",
                  command=self._ok, bg="#2d6a4f", fg="white",
                  font=("Segoe UI", 12, "bold"), padx=16, pady=8,
                  relief="flat", cursor="hand2").pack(side="left", padx=6)
        tk.Button(bf, text="🔄  CROSS PO",
                  command=self._cross_po, bg="#7a5c00", fg="white",
                  font=("Segoe UI", 10, "bold"), padx=12, pady=8,
                  relief="flat", cursor="hand2").pack(side="left", padx=6)
        tk.Button(bf, text="⚖  MISMATCH / OVERAGE",
                  command=self._mismatch, bg="#2d4a7a", fg="white",
                  font=("Segoe UI", 10, "bold"), padx=12, pady=8,
                  relief="flat", cursor="hand2").pack(side="left", padx=6)
        tk.Button(bf, text="⬛  STOP THIS ASIN",
                  command=self._stop, bg="#4a2020", fg="white",
                  font=("Segoe UI", 10), padx=12, pady=8,
                  relief="flat", cursor="hand2").pack(side="left", padx=6)

        # ── Pending Invoices & Utility buttons ──────────────────────────────
        bf2 = tk.Frame(self, bg="#0f0f1a"); bf2.pack(pady=4, fill='x', padx=16)
        tk.Button(bf2, text="📋  VIEW ALL PENDING INVOICES",
                  command=self._show_pending, bg="#3a2a00", fg="#f0c060",
                  font=("Segoe UI", 10, "bold"), padx=16, pady=7,
                  relief="flat", cursor="hand2").pack(side="left", expand=True, fill="x", padx=(0,4))
        
        tk.Button(bf2, text="🔍 LOOKUP INV QTY",
                  command=self._lookup_inv_qty, bg="#1c2c42", fg="#80a0ff",
                  font=("Segoe UI", 10, "bold"), padx=16, pady=7,
                  relief="flat", cursor="hand2").pack(side="right", padx=(4,0))
        
        tk.Button(bf2, text="📦 RECEIVED QTY",
                  command=self._show_rec_qty_lookup, bg="#d4a017", fg="black",
                  font=("Segoe UI", 10, "bold"), padx=16, pady=7,
                  relief="flat", cursor="hand2").pack(side="right", padx=4)

        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _auto_populate(self):
        """v7.1.0: Auto-populate SID and Barcode from Vendor Data with SCR cleaning."""
        self._pull_vendor_data(show_msg=False)

    def _pull_vendor_data(self, show_msg=True):
        """v7.1.1: Pulls SID and Barcode from Vendor Data, handles SCR removal."""
        sel = self._branch_var.get()
        if not sel or not self._engine_ref: return
        try:
            parts = sel.split("|")
            inv_part = next((p for p in parts if "Inv=" in p), None)
            if not inv_part: return
            inv_no = inv_part.split("=")[1].strip()
            
            # v7.1.1: Clean Invoice (remove SCR)
            clean_inv = inv_no.upper().replace("SCR", "").strip()
            
            vd = self._engine_ref.vendor_idx.get(clean_inv)
            if not vd: 
                # Try original if cleaning didn't help
                vd = self._engine_ref.vendor_idx.get(inv_no)

            if vd:
                if vd.get('sid'): self._sid_var.set(vd['sid'])
                if vd.get('bc'): self._bc_var.set(vd['bc'])
                if show_msg: self._set_status_temp(f"Data pulled for {clean_inv}", "#90ee90")
            elif show_msg:
                messagebox.showinfo("Not Found", f"Invoice {clean_inv} not found in Vendor Level Data.")
        except Exception as e:
            if show_msg: messagebox.showerror("Error", f"Failed to pull data: {e}")

    def _set_status_temp(self, msg, color):
        # Helper to show temp status in a label if we had one, otherwise just pass
        pass

    def _lookup_inv_qty(self):
        """Restored in v7.1.1: Lookup dialog for invoice quantity."""
        dlg = tk.Toplevel(self)
        dlg.title("Lookup Invoice Quantity")
        dlg.configure(bg="#0d1117")
        dlg.attributes("-topmost", True)
        
        tk.Label(dlg, text="Invoice No:", bg="#0d1117", fg="white", font=("Segoe UI", 10)).grid(row=0, column=0, padx=10, pady=(15,5), sticky='e')
        inv_ent = tk.Entry(dlg, font=("Segoe UI", 10), bg="#21262d", fg="white", insertbackground="white")
        inv_ent.grid(row=0, column=1, padx=10, pady=(15,5))
        
        tk.Label(dlg, text="ASIN:", bg="#0d1117", fg="white", font=("Segoe UI", 10)).grid(row=1, column=0, padx=10, pady=5, sticky='e')
        asin_ent = tk.Entry(dlg, font=("Segoe UI", 10), bg="#21262d", fg="white", insertbackground="white")
        asin_ent.grid(row=1, column=1, padx=10, pady=5)
        
        res_lbl = tk.Label(dlg, text="", bg="#0d1117", fg="#f0c060", font=("Segoe UI", 11, "bold"))
        res_lbl.grid(row=3, column=0, columnspan=2, pady=(10,15))
        
        def do_lookup():
            ino = inv_ent.get().strip()
            asn = asin_ent.get().strip()
            engine = getattr(self, '_engine_ref', None)
            if not engine:
                res_lbl.config(text="Engine not connected.", fg="#f85149"); return
            found_qty = engine._resolve_inv_qty(ino, asn, None)
            if found_qty is not None:
                res_lbl.config(text=f"Exact Inv Qty = {int(found_qty)} units", fg="#3fb950")
            else:
                res_lbl.config(text="Not found in Invoice Data.", fg="#f85149")
                
        tk.Button(dlg, text="🔍 Search Data", bg="#238636", fg="white", font=("Segoe UI", 10, "bold"),
                  command=do_lookup, cursor="hand2", padx=20).grid(row=2, column=0, columnspan=2, pady=10)
        
        dlg.update_idletasks()
        px_dlg = self.winfo_x() + (self.winfo_width() - dlg.winfo_width()) // 2
        py_dlg = self.winfo_y() + 50
        dlg.geometry(f"+{px_dlg}+{py_dlg}")
        inv_ent.focus_set()

    def _show_ibc_validation(self):
        """v7.1.0: Shows the exact IBC vs PBC validation block for the current SID."""
        sid = self._sid_var.get().strip()
        if not sid:
            messagebox.showwarning("Warning", "Please enter/populate a Shipment ID first."); return
        
        if not self._engine_ref or not self._engine_ref.vendor_raw:
            messagebox.showerror("Error", "Vendor Data not loaded or index missing."); return
            
        raw_rows = self._engine_ref.vendor_raw.get(sid)
        if not raw_rows:
            messagebox.showinfo("Note", f"No Vendor Data found for SID: {sid}"); return
            
        # Toggle display
        if self._val_display.winfo_ismapped():
            self._val_display.pack_forget()
            return
            
        self._val_display.pack(fill="x", pady=4)
        self._val_display.config(state="normal")
        self._val_display.delete("1.0", tk.END)
        
        # Build Table
        headers = self._engine_ref.vendor_headers
        # We want to show: Inv, PO, ASIN, Inv Qty, IBC, PBC, Rejected, Reason
        def find_idx(patterns):
            for i, h in enumerate(headers):
                hl = str(h).lower().strip()
                if any(p in hl for p in patterns): return i
            return -1
            
        idx_inv = find_idx(['invoice number', 'invoice'])
        idx_po  = find_idx(['po number', 'purchase order', 'po'])
        idx_iq  = find_idx(['total invoiced quantity', 'invoiced qty', 'inv qty'])
        idx_ibc = find_idx(['invoice box count', 'ibc', 'invoice box'])
        idx_pbc = find_idx(['pod box quantity', 'pbc', 'physical box'])
        idx_rc  = find_idx(['number of cartons rejected', 'carton rejected', 'rejected carton'])
        idx_rr  = find_idx(['pod reject reason', 'reject reason', 'rejection reason'])
        
        # Header Row (Removed ASIN)
        row_fmt = "{:<16} | {:<12} | {:>7} | {:>4} | {:>4} | {:>4} | {}"
        header_text = row_fmt.format("Invoice", "PO", "InvQty", "IBC", "PBC", "Rej", "Reason")
        self._val_display.insert(tk.END, header_text + "\n" + ("-" * 85) + "\n")
        
        total_ibc = 0
        total_pbc = 0 
        total_inv_qty = 0
        total_rej = 0
        seen_pos = set()
        
        for r in raw_rows:
            inv = str(r[idx_inv]) if idx_inv != -1 else ""
            po  = str(r[idx_po]) if idx_po != -1 else ""
            iq  = int(safe_num(r[idx_iq])) if idx_iq != -1 else 0
            ibc = int(safe_num(r[idx_ibc])) if idx_ibc != -1 else 0
            pbc = int(safe_num(r[idx_pbc])) if idx_pbc != -1 else 0
            rej = int(safe_num(r[idx_rc])) if idx_rc != -1 else 0
            rsn = str(r[idx_rr]) if idx_rr != -1 else ""
            
            self._val_display.insert(tk.END, row_fmt.format(inv[:16], po[:12], iq, ibc, pbc, rej, rsn) + "\n")
            total_ibc += ibc
            total_inv_qty += iq
            total_rej += rej
            if po not in seen_pos:
                total_pbc += pbc
                seen_pos.add(po)

        self._val_display.insert(tk.END, ("-" * 85) + "\n")
        # v7.2.3 logic: VALID if IBC <= PBC and no rejected cartons
        is_valid = (total_ibc <= total_pbc) and (total_rej == 0)
        status = "✔ VALID" if is_valid else "❌ INVALID"
        color = "#90ee90" if is_valid else "#ff8888"
        
        summary = f"TOTAL INV QTY: {total_inv_qty}  |  TOTAL IBC: {total_ibc}  |  TOTAL PBC: {total_pbc}  |  STATUS: {status}"
        self._val_display.insert(tk.END, summary + "\n")
        
        # Color coding the status line in text
        start_idx = self._val_display.index("end-2l")
        self._val_display.tag_add("status", start_idx, tk.END)
        self._val_display.tag_config("status", foreground=color, font=("Consolas", 10, "bold"))
                
        self._val_display.config(state="disabled")
        self._val_display.see(tk.END)

    def _show_rec_qty_lookup(self):
        """v6.2.6: Fully Independent Received Qty Lookup (Asks for SID/PO/ASIN)."""
        dlg = tk.Toplevel(self)
        dlg.title("Manual Mode — Received Qty Lookup")
        dlg.geometry("450x320")
        dlg.configure(bg="#0d1117")
        dlg.attributes("-topmost", True)
        dlg.lift(); dlg.focus_force()
        
        tk.Label(dlg, text="  Independent Reconciliation Lookup", bg="#161b22", fg="#f0a500", font=("Segoe UI", 11, "bold"), height=2).pack(fill="x")
        
        f = tk.Frame(dlg, bg="#0d1117", padx=20, pady=15)
        f.pack(fill="x")
        
        # UI Fields
        tk.Label(f, text="Shipment ID:", bg="#0d1117", fg="#888888", font=("Segoe UI", 10)).grid(row=0, column=0, sticky='w', pady=5)
        sid_ent = tk.Entry(f, font=("Segoe UI", 10), bg="#21262d", fg="white", insertbackground="white", width=25)
        sid_ent.grid(row=0, column=1, padx=10, pady=5)
        # Pre-fill SID if available from the current dialog
        if hasattr(self, '_sid_var'): sid_ent.insert(0, self._sid_var.get())
        
        tk.Label(f, text="Purchase Order:", bg="#0d1117", fg="#888888", font=("Segoe UI", 10)).grid(row=1, column=0, sticky='w', pady=5)
        po_ent = tk.Entry(f, font=("Segoe UI", 10), bg="#21262d", fg="white", insertbackground="white", width=25)
        po_ent.grid(row=1, column=1, padx=10, pady=5)
        
        tk.Label(f, text="ASIN:", bg="#0d1117", fg="#888888", font=("Segoe UI", 10)).grid(row=2, column=0, sticky='w', pady=5)
        asin_ent = tk.Entry(f, font=("Segoe UI", 10), bg="#21262d", fg="white", insertbackground="white", width=25)
        asin_ent.grid(row=2, column=1, padx=10, pady=5)

        def perform_lookup():
            sid = extract_sid(sid_ent.get().strip())
            po  = clean(po_ent.get().strip())
            asin = clean(asin_ent.get().strip())
            if not all([sid, po, asin]):
                messagebox.showerror("Error", "All fields are required.", parent=dlg); return
            
            engine = getattr(self, '_engine_ref', None)
            if not engine:
                messagebox.showerror("Error", "Investigation Engine not active.", parent=dlg); return
                
            rows = engine.rebni_p.get((sid, po, asin), [])
            if not rows:
                messagebox.showinfo("Result", "No REBNI data found for this combination.", parent=dlg); return
            
            r0 = rows[0]
            # matched invoices logic
            matched_invs = []
            for (m_inv, m_po, m_asin), m_rows in engine.inv_p.items():
                if clean(m_asin) == asin and clean(m_po) == po:
                    # check if invoice belongs to SID
                    if any(clean(m_inv) == clean(i) for i in engine.cache_sid.get(sid, [])):
                        qty = sum(safe_num(r.get('quantity_invoiced', 0)) for r in m_rows)
                        matched_invs.append(f"{m_inv} (Qty:{int(qty)})")

            res_dlg = tk.Toplevel(dlg)
            res_dlg.title(f"Reconciliation Detail — {asin}")
            res_dlg.geometry("520x480")
            res_dlg.configure(bg="#0d1117")
            res_dlg.attributes("-topmost", True)
            
            tk.Label(res_dlg, text=f"📊  Full Reconciliation: {asin}", bg="#161b22", fg="#3fb950", font=("Segoe UI", 11, "bold"), height=2).pack(fill="x")
            
            sc = tk.Frame(res_dlg, bg="#0d1117", padx=25, pady=20)
            sc.pack(fill="both", expand=True)
            
            # Metrics
            metrics = [
                ("Quantity Unpacked:", r0.get('quantity_unpacked', 0), "#58a6ff"),
                ("Quantity Adjusted:", r0.get('quantity_adjusted', 0), "#ff4d4d"),
                ("REBNI Available:", r0.get('rebni_available', 0), "#f0a500"),
                ("Quantity Received (Post Adj):", r0.get('qty_received_postadj', 0), "#3fb950"),
                ("Item Cost (INR):", f"{safe_num(r0.get('item_cost', 0)):.2f}", "#abb2bf")
            ]
            
            for i, (l, v, c) in enumerate(metrics):
                tk.Label(sc, text=l, bg="#0d1117", fg="#888888", font=("Segoe UI", 10)).grid(row=i, column=0, sticky='w', pady=8)
                val_str = f"{v} units" if "Cost" not in l else v
                tk.Label(sc, text=val_str, bg="#0d1117", fg=c, font=("Segoe UI", 11, "bold")).grid(row=i, column=1, sticky='w', padx=20)

            # Matched Invoices List
            tk.Label(sc, text="Matched Invoices in SID:", bg="#0d1117", fg="#4a9eff", font=("Segoe UI", 10, "bold")).grid(row=6, column=0, sticky='w', pady=(15,5))
            inv_text = ", ".join(matched_invs) if matched_invs else "No invoices found for this ASIN in the portal."
            st = scrolledtext.ScrolledText(sc, width=45, height=4, bg="#161b22", fg="#e0e0e0", font=("Consolas", 9), relief="flat")
            st.insert("1.0", inv_text)
            st.configure(state="disabled")
            st.grid(row=7, column=0, columnspan=2, sticky='w', pady=5)

            tk.Button(res_dlg, text="CLOSE", command=res_dlg.destroy, bg="#333", fg="white", padx=20).pack(pady=15)

        tk.Button(dlg, text="🔍  SEARCH & RECONCILE", command=perform_lookup, bg="#d4a017", fg="black", font=("Segoe UI", 10, "bold"), padx=25, pady=10, relief="flat", cursor="hand2").pack(pady=10)
        
        dlg.update_idletasks()
        px_dlg = self.winfo_x() + (self.winfo_width()  - dlg.winfo_width())  // 2
        py_dlg = self.winfo_y() + (self.winfo_height() - dlg.winfo_height()) // 2
        dlg.geometry(f"+{px_dlg}+{py_dlg}")

    def _toggle(self):
        if self._validity.get() == "valid":
            self._invalid_frame.pack_forget()
            self._dices_frame.pack(fill="x", padx=16, pady=4)
        else:
            self._dices_frame.pack_forget()
            self._invalid_frame.pack(fill="x", pady=3)

    def _ok(self):
        sel = self._branch_cb.current()
        if sel < 0 or sel >= len(self.matches):
            messagebox.showwarning("Select Invoice", "Please select an invoice.", parent=self)
            return
        match = self.matches[sel]
        if self._validity.get() == "valid":
            sid = extract_sid(self._sid_var.get().strip())
            if not sid:
                messagebox.showwarning("SID Required", "Please enter SID from DICES.", parent=self)
                return
            self.callback({'action': 'valid', 'chosen_match': match,
                           'sid': sid, 'barcode': self._bc_var.get().strip() or "[DICES]"})
        else:
            qty_str = self._inv_qty_var.get().strip()
            try:
                qty = float(qty_str)
            except:
                messagebox.showwarning("Qty Required", "Enter units matched to invalid invoice.", parent=self)
                return
            self.callback({'action': 'invalid', 'chosen_match': match, 'invalid_qty': qty})
        self.destroy()

    def _cross_po(self):
        self.callback({'action': 'cross_po',
                       'chosen_match': self.matches[self._branch_cb.current()] if self.matches else None})
        self.destroy()

    def _mismatch(self):
        dlg = tk.Toplevel(self)
        dlg.title("Mismatch / Overage Details")
        dlg.geometry("460x260")
        dlg.configure(bg="#0f0f1a")
        dlg.lift(); dlg.focus_force()

        fields = [("ASIN received:", "asin"), ("SID:", "sid"), ("PO:", "po"),
                  ("Inv Qty (invoiced):", "inv_qty"), ("Overage Qty received:", "ovg_qty")]
        vars_ = {}
        for i, (lbl, key) in enumerate(fields):
            tk.Label(dlg, text=lbl, bg="#0f0f1a", fg="#e0e0e0",
                     font=("Segoe UI", 10), width=22, anchor="w"
                     ).grid(row=i, column=0, padx=12, pady=5)
            v = tk.StringVar()
            tk.Entry(dlg, textvariable=v, width=26, font=("Segoe UI", 10),
                     bg="#1e1e3a", fg="#e0e0e0", insertbackground="white",
                     relief="flat").grid(row=i, column=1, padx=8, pady=5)
            vars_[key] = v

        def submit():
            data = {k: v.get().strip() for k, v in vars_.items()}
            dlg.destroy()
            self.callback({'action': 'mismatch', 'mismatch_data': data})
            self.destroy()

        tk.Button(dlg, text="✔  Confirm & Investigate", command=submit,
                  bg="#2d6a4f", fg="white", font=("Segoe UI", 11, "bold"),
                  padx=14, pady=7, relief="flat", cursor="hand2"
                  ).grid(row=len(fields), column=0, columnspan=2, pady=12)

    def sync_with_edits(self, updated_map, new_adds):
        """Update combobox options dynamically when Confirm Edits is clicked in PreviewPanel."""
        patched = []
        for m in self.matches:
            k = (clean(m.get('mtc_inv','')), clean(m.get('mtc_asin','')), clean(m.get('mtc_po','')))
            if k in updated_map:
                m = dict(m)
                m.update(updated_map[k])
            patched.append(m)
        for n in new_adds:
            patched.append(n)
        self.matches = patched
        
        opts = [f"Qty={fmt_qty(m.get('mtc_qty',0))}  |  Inv={m.get('mtc_inv','')}  |  PO={m.get('mtc_po','')}  |  ASIN={m.get('mtc_asin','')}"
                for m in self.matches]
        self._branch_cb['values'] = opts
        idx = self._branch_cb.current()
        if opts and idx >= 0:
            self._branch_cb.current(idx)
        elif opts:
            self._branch_cb.current(0)

    def _show_pending(self):
        # v6.2.5: Isolated per-ASIN matches
        matches = self._engine_ref.asin_pending_matches if self._engine_ref else []
        if not matches and self._pending_cb:
            self._pending_cb()
            return
        if not matches:
            messagebox.showinfo("Pending Invoices", "No pending invoices for this ASIN chain.", parent=self)
            return
        self._pending_cb()  # Always show if we have matches

    def _stop(self):
        self.callback({'action': 'stop'}); self.destroy()


class PendingInvoicesDialog(tk.Toplevel):
    """
    Shown before finalizing a claiming ASIN.
    Lists all matched invoices that were identified but not yet investigated.
    User can pick one to continue investigation or skip.
    """
    def __init__(self, parent, pending_invoices, asin, callback):
        super().__init__(parent)
        self.callback         = callback
        self.pending_invoices = pending_invoices

        self.title("Pending Matched Invoices — Action Required")
        self.geometry("760x480")
        self.configure(bg="#0f0f1a")
        self.resizable(True, True)
        self.protocol("WM_DELETE_WINDOW", self.withdraw)
        self.lift(); self.focus_force()

        tk.Label(self,
                 text="📋  Uninvestigated Matched Invoices Found",
                 bg="#16213e", fg="#f0a500",
                 font=("Segoe UI", 12, "bold"), height=2).pack(fill="x")

        tk.Label(self,
                 text=(f"ASIN: {asin} — The following matched invoices were identified during\n"
                       "investigation but have not yet been investigated. "
                       "Select one to continue or click 'Go to Next ASIN'."),
                 bg="#0f0f1a", fg="#cccccc",
                 font=("Segoe UI", 9), justify="left").pack(pady=6, padx=16, anchor="w")

        # Table of pending invoices
        tf = tk.LabelFrame(self, text="  Pending Matched Invoices  ",
                           bg="#0f0f1a", fg="#e0e0e0",
                           font=("Segoe UI", 9, "bold"), padx=10, pady=6)
        tf.pack(fill="both", expand=True, padx=16, pady=4)

        cols = ("Mtc Qty", "Invoice No", "PO", "ASIN", "Level")
        vsb = ttk.Scrollbar(tf, orient="vertical")
        hsb = ttk.Scrollbar(tf, orient="horizontal")
        self.tree = ttk.Treeview(tf, columns=cols, show='headings', yscrollcommand=vsb.set, xscrollcommand=hsb.set, height=10)
        vsb.config(command=self.tree.yview); hsb.config(command=self.tree.xview)
        
        col_w = {"Mtc Qty": 80, "Invoice No": 220, "PO": 140, "ASIN": 140, "Level": 60}
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=col_w[c], anchor='w', stretch=tk.YES)
            
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tf.grid_rowconfigure(0, weight=1)
        tf.grid_columnconfigure(0, weight=1)
        
        # New Feature: Styling Treeview for dark theme pending window
        s = ttk.Style()
        if "clam" in s.theme_names(): s.theme_use("clam")
        s.configure("Pending.Treeview", font=("Calibri", 10), rowheight=24, background="#1e1e3a", foreground="#e0e0e0", fieldbackground="#1e1e3a")
        s.configure("Pending.Treeview.Heading", font=("Calibri", 10, "bold"), background="#203864", foreground="white")
        s.map("Pending.Treeview", background=[('selected', '#2d4a7a')])
        self.tree.configure(style="Pending.Treeview")

        self._populate_tree()

        # Dropdown to select one
        sf = tk.Frame(self, bg="#0f0f1a"); sf.pack(fill="x", padx=16, pady=4)
        tk.Label(sf, text="Select invoice to investigate:",
                 bg="#0f0f1a", fg="#e0e0e0",
                 font=("Segoe UI", 10), width=30, anchor="w").pack(side="left")
        opts = [f"Qty={fmt_qty(inv.get('mtc_qty',''))}  |  Inv={inv.get('mtc_inv','')}  |  "
                f"PO={inv.get('mtc_po','')}  |  ASIN={inv.get('mtc_asin','')}"
                for inv in pending_invoices]
        self._sel = tk.StringVar()
        self._cb  = ttk.Combobox(sf, textvariable=self._sel,
                                   values=opts, state="readonly", width=60,
                                   font=("Segoe UI", 9))
        if opts: self._cb.current(0)
        self._cb.pack(side="left", padx=6)

        # Buttons
        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=10)
        tk.Button(bf, text="🔍  Investigate Selected Invoice",
                  command=self._investigate,
                  bg="#2d4a7a", fg="white",
                  font=("Segoe UI", 11, "bold"),
                  padx=16, pady=8, relief="flat",
                  cursor="hand2").pack(side="left", padx=8)
        tk.Button(bf, text="▶▶  Go to Next ASIN",
                  command=self._next_asin,
                  bg="#2d6a4f", fg="white",
                  font=("Segoe UI", 11, "bold"),
                  padx=16, pady=8, relief="flat",
                  cursor="hand2").pack(side="left", padx=8)

        self.protocol("WM_DELETE_WINDOW", self._next_asin)
        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _populate_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for inv in self.pending_invoices:
            vals = [
                fmt_qty(inv.get('mtc_qty', '')),
                inv.get('mtc_inv', ''),
                inv.get('mtc_po', ''),
                inv.get('mtc_asin', ''),
                str(inv.get('_depth', '?'))
            ]
            self.tree.insert('', 'end', values=vals)

    def sync_with_edits(self, updated_map, new_adds):
        patched = []
        for m in self.pending_invoices:
            k = (clean(m.get('mtc_inv','')), clean(m.get('mtc_asin','')), clean(m.get('mtc_po','')))
            if k in updated_map:
                m = dict(m)
                m.update(updated_map[k])
            patched.append(m)
        for n in new_adds:
            patched.append(n)
        
        self.pending_invoices = patched
        self._populate_tree()
        opts = [f"Qty={fmt_qty(inv.get('mtc_qty',''))}  |  Inv={inv.get('mtc_inv','')}  |  PO={inv.get('mtc_po','')}  |  ASIN={inv.get('mtc_asin','')}" for inv in self.pending_invoices]
        self._cb['values'] = opts
        if opts and self._cb.current() < 0: self._cb.current(0)

    def _investigate(self):
        idx = self._cb.current()
        if idx < 0 or idx >= len(self.pending_invoices):
            messagebox.showwarning("Select Invoice", "Please select an invoice to investigate.", parent=self)
            return
        self.callback({'action': 'investigate', 'match': self.pending_invoices[idx]})
        self.destroy()

    def _next_asin(self):
        self.callback({'action': 'next_asin'})
        self.destroy()

class CorrespondenceDialog(tk.Toplevel):
    def __init__(self, parent, all_rows, app=None):
        super().__init__(parent)
        self.all_rows = all_rows
        self.app = app
        self.title("Scenario Selection — Get Correspondence — v7.7")
        self.geometry("1000x850")
        self.configure(bg="#0f0f1a")
        
        # Enable Maximize/Minimize and ensure centered display
        self.resizable(True, True)
        self.focus_force()

        self.scenarios = [
            "Claiming Short",
            "REBNI",
            "Matching (Bulk)",
            "Mismatch ASIN Overages",
            "Invalid/Dummy Invoice [Dev]",
            "IBC vs PBC [Dev]"
        ]
        
        self.v_code_var = tk.StringVar(value="[Vendor Code]")
        self.fc_id_var = tk.StringVar(value="[FC ID]")
        
        self._build_ui()
        
    def _build_ui(self):
        f = tk.Frame(self, bg="#0f0f1a", padx=25, pady=25)
        f.pack(fill="both", expand=True)
        
        # --- Header with Professional Styling ---
        hdr = tk.Frame(f, bg="#0f0f1a")
        hdr.pack(fill="x", pady=(0, 20))
        tk.Label(hdr, text="Investigation Correspondence Generator",
                 fg="#4a9eff", bg="#0f0f1a",
                 font=("Segoe UI", 16, "bold")).pack(side="left")

        # --- SELECTION AREA ---
        sel_f = tk.LabelFrame(f, text=" STEP 1: SELECT SCENARIO ", fg="#4a9eff", bg="#0f0f1a", font=("Segoe UI", 9, "bold"))
        sel_f.pack(fill="x", pady=10)
        self.scenario_var = tk.StringVar()
        self.scenario_cb = ttk.Combobox(sel_f, textvariable=self.scenario_var, 
                                        values=self.scenarios, state="readonly", 
                                        font=("Segoe UI", 12))
        self.scenario_cb.pack(fill="x", padx=15, pady=15)
        self.scenario_cb.bind("<<ComboboxSelected>>", self.generate_text)
        
        # --- INPUT AREA ---
        inp_f = tk.LabelFrame(f, text=" STEP 2: HYDRATION DETAILS ", fg="#4a9eff", bg="#0f0f1a", font=("Segoe UI", 9, "bold"))
        inp_f.pack(fill="x", pady=10)
        grid_f = tk.Frame(inp_f, bg="#0f0f1a")
        grid_f.pack(fill="x", padx=15, pady=10)
        
        tk.Label(grid_f, text="Vendor Code:", bg="#0f0f1a", fg="#aaaaaa").grid(row=0, column=0, sticky="w")
        tk.Entry(grid_f, textvariable=self.v_code_var, width=25, font=("Segoe UI", 10)).grid(row=0, column=1, padx=(5, 30))
        
        tk.Label(grid_f, text="FC ID:", bg="#0f0f1a", fg="#aaaaaa").grid(row=0, column=2, sticky="w")
        tk.Entry(grid_f, textvariable=self.fc_id_var, width=25, font=("Segoe UI", 10)).grid(row=0, column=3, padx=5)

        # --- TEXT AREA ---
        tk.Label(f, text=" STEP 3: REVIEW CORRESPONDENCE (Auto-generated from investigation data) ",
                 fg="#e0e0e0", bg="#0f0f1a", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(15, 5))
        
        # Use a higher contrast text box for legibility and monospaced font for perfect table alignment
        self.text_area = tk.Text(f, font=("Consolas", 10), 
                                 bg="#ffffff", fg="#000000", padx=15, pady=15,
                                 insertbackground="black", wrap="none", relief="flat")
        self.text_area.pack(fill="both", expand=True)

        # --- FOOTER BUTTONS ---
        btn_f = tk.Frame(f, bg="#0f0f1a")
        btn_f.pack(fill="x", pady=(25, 0))
        
        tk.Button(btn_f, text="📋  COPY TO CLIPBOARD ", 
                  command=self.copy_to_clip,
                  bg="#2d6a4f", fg="white", font=("Segoe UI", 11, "bold"),
                  relief="flat", padx=25, pady=12, cursor="hand2").pack(side="left")

        if self.app:
            tk.Button(btn_f, text="💬  AI INVESTIGATOR ", 
                      command=lambda: self.app.open_ai_chat(),
                      bg="#e94560", fg="white", font=("Segoe UI", 11, "bold"),
                      relief="flat", padx=25, pady=12, cursor="hand2").pack(side="left", padx=10)
        
        tk.Button(btn_f, text="CLOSE", command=self.destroy,
                  bg="#333333", fg="white", font=("Segoe UI", 10),
                  relief="flat", padx=20, pady=12, cursor="hand2").pack(side="right")

    def copy_to_clip(self):
        txt = self.text_area.get("1.0", tk.END).strip()
        if txt:
            self.clipboard_clear(); self.clipboard_append(txt)
            messagebox.showinfo("Success", "Professional correspondence copied!", parent=self)

    def generate_text(self, event=None):
        scenario = self.scenario_var.get()
        if not scenario: return
        
        # Gather data - strictly skip headers
        valid_rows = [r for r in self.all_rows if not r.get('is_ui_header') and str(r.get('sid', '')).strip() != 'SID' and str(r.get('barcode', '')).strip() != 'Barcode']
        primary_row = next((r for r in valid_rows if r.get('depth', 0) == 0), {})
        
        sid = str(primary_row.get('sid', '[SID missing]'))
        po  = str(primary_row.get('po', '[PO missing]'))
        inv = str(primary_row.get('invoice', '[Invoice missing]'))
        vc  = self.v_code_var.get()
        fc  = self.fc_id_var.get()
        
        # Total sums for header
        t_billed = int(sum(safe_num(r.get('inv_qty', 0)) for r in valid_rows if r.get('depth', 0) == 0))
        t_received = int(sum(safe_num(r.get('rec_qty', 0)) for r in valid_rows if r.get('depth', 0) == 0))
        
        # ASIN list generation
        asin_lines = []
        for r in valid_rows:
            # Only include "real" rows with shortages
            if str(r.get('barcode','')).strip() and not str(r.get('barcode','')).startswith('['):
                iq = int(safe_num(r.get('inv_qty', 0)))
                rq = int(safe_num(r.get('rec_qty', 0)))
                if iq > rq:
                    asin = str(r.get('asin', ''))
                    missing = iq - rq
                    amt = float(safe_num(r.get('cp_status', 0)))
                    asin_lines.append(f"Asin: {asin} ---- Invoice Qty: {iq} ---- Received Qty: {rq} ---- Shortage Qty: {missing} ---- Amount: INR ---- {amt:.2f}")
        
        asin_block = "\n".join(asin_lines) if asin_lines else "[No ASIN Shortages detected in Preview]"
        
        signature = f"Regards,\nMUKESH | pathlmuk\nROW IB"
        
        if scenario == "Claiming Short":
            # 1. ASIN Table Generation
            asin_hdr = f"| {'Barcode':<16} | {'Inv no':<20} | {'SID':<16} | {'PO':<12} | {'ASIN':<12} | {'Inv Qty':<9} | {'Rec Qty':<9} | {'Missing QTY':<11} |"
            asin_sep = f"| {'-'*16} | {'-'*20} | {'-'*16} | {'-'*12} | {'-'*12} | {'-'*9} | {'-'*9} | {'-'*11} |"
            asin_table_rows = [asin_hdr, asin_sep]
            for r in valid_rows:
                if str(r.get('barcode','')).strip() and not str(r.get('barcode','')).startswith('['):
                    iq = int(safe_num(r.get('inv_qty', 0)))
                    rq = int(safe_num(r.get('rec_qty', 0)))
                    if iq > rq:
                        missing = iq - rq
                        bc = clean(r.get('barcode', ''))
                        inv_val = clean(r.get('invoice', ''))
                        sid_val = clean(r.get('sid', ''))
                        po_val = clean(r.get('po', ''))
                        asin_val = clean(r.get('asin', ''))
                        asin_table_rows.append(f"| {bc:<16} | {inv_val:<20} | {sid_val:<16} | {po_val:<12} | {asin_val:<12} | {iq:<9} | {rq:<9} | {missing:<11} |")
            
            asin_block = "Example:\n" + "\n".join(asin_table_rows) if len(asin_table_rows) > 2 else "[No ASIN Shortages detected]"

            # 2. IBC/PBC Table Generation
            ibc_hdr = f"| {'Barcode':<16} | {'PO Number':<12} | {'Invoice Number':<20} | {'Invoiced Qty':<14} | {'Shipment ID':<16} | {'POD Box Qty':<12} | {'Rejected Cartons':<18} | {'POD Reject Reason':<20} | {'IBC':<5} |"
            ibc_sep = f"| {'-'*16} | {'-'*12} | {'-'*20} | {'-'*14} | {'-'*16} | {'-'*12} | {'-'*18} | {'-'*20} | {'-'*5} |"
            ibc_pbc_rows = [ibc_hdr, ibc_sep]
            engine = getattr(self.app, 'engine', None)
            
            t_billed_ibc = 0
            
            if engine and engine.vendor_raw:
                target_sid = extract_sid(sid)
                raw_data = engine.vendor_raw.get(target_sid, [])
                if raw_data:
                    h = engine.vendor_headers
                    def get_idx(keys):
                        for i, col in enumerate(h):
                            if any(k.lower() in str(col).lower() for k in keys): return i
                        return -1
                    i_bc = get_idx(['barcode', 'fnsku', 'upc'])
                    i_po = get_idx(['po number', 'purchase order', 'po'])
                    i_inv = get_idx(['invoice number', 'invoice'])
                    i_iq = get_idx(['total invoiced quantity', 'invoiced qty', 'inv qty'])
                    i_sid = get_idx(['shipment id', 'shipment_id', 'sid'])
                    i_pbc = get_idx(['pod box', 'pbc', 'physical box'])
                    i_rej = get_idx(['rejected carton', 'number of cartons rejected', 'rej'])
                    i_rej_rsn = get_idx(['pod reject reason', 'reject reason'])
                    i_ibc = get_idx(['invoiced box', 'ibc', 'invoice box count'])
                    
                    for r in raw_data:
                        iq_val = clean(str(r[i_iq])) if i_iq != -1 else ""
                        t_billed_ibc += int(safe_num(iq_val))
                        
                        rpo = clean(str(r[i_po])) if i_po != -1 else ""
                        if not rpo or rpo == po:
                            ibc_pbc_rows.append(
                                f"| {clean(str(r[i_bc])):<16} | "
                                f"{rpo:<12} | "
                                f"{clean(str(r[i_inv])):<20} | "
                                f"{iq_val:<14} | "
                                f"{clean(str(r[i_sid])):<16} | "
                                f"{clean(str(r[i_pbc])):<12} | "
                                f"{clean(str(r[i_rej])):<18} | "
                                f"{clean(str(r[i_rej_rsn]))[:20]:<20} | "
                                f"{clean(str(r[i_ibc])):<5} |"
                            )

            ibc_pbc_block = "Example:\n" + "\n".join(ibc_pbc_rows) if len(ibc_pbc_rows) > 2 else "[Please paste data from IBC and PBC validation Sheet here]"

            t_billed_display = t_billed_ibc if t_billed_ibc > 0 else t_billed

            text = (
                f"Hello FC Team/SLP Team,\n\n"
                f"We have already performed all the virtual research/checks such as cross receiving, overage, REBNI, adjustments etc. "
                f"and need FC support for physical search on floor to locate missing units as per revised SOP and update at the earliest.\n\n"
                f"SID# {sid} | Total Billed Qty - {t_billed_display} | Received Qty - [User From CIAT]\n\n"
                f"Shipment ID (to be investigated): {sid}\n"
                f"PO to be investigated: {po}\n"
                f"Invoice No. to be investigated: {inv}\n\n"
                f"ASIN/Units to be investigated:\n\n"
                f"{asin_block}\n\n\n"
                f"ROW IB Findings: ***We are able to see that unit’s shortage received in claiming shipment***\n\n"
                f"[{fc if fc != '[FC ID]' else 'FC'}] / [SLP] Investigation support required:\n\n"
                f"Note: We need FC support ONLY for PHYSICAL SEARCH on floor to locate missing units as virtual research is already completed. "
                f"If missing units not found physically and PQV value >50K then FC to take support from FC SLP team for further investigation. "
                f"TT to be flipped FC SLP queue for further investigation through CCTV footage and relevant tools.\n\n"
                f"Detailed investigation summary and findings are appended below for your reference:\n\n"
                f"Below are the shipment level IBC vs PBC Details and the same has been attached in information tab As bulk:\n\n"
                f"{ibc_pbc_block}\n\n\n"
                f"Regards\n"
                f"Investigator | login@\n"
                f"ROW IB"
            )
        elif scenario == "REBNI":
            # Identify where REBNI matches occurred
            c_sids = set()
            m_sids = set()
            for r in valid_rows:
                rem = str(r.get('remarks', ''))
                rsid = str(r.get('sid', ''))
                if 'REBNI' in rem or 'TSP to utilize' in rem or r.get('type') == 'rebni_shipment':
                    if r.get('depth', 0) == 0: c_sids.add(rsid)
                    else: m_sids.add(rsid)
            
            # Construct the SID line based on where REBNI was found
            sid_parts = []
            if c_sids: sid_parts.append(f"claiming SID - {', '.join(sorted(c_sids))}")
            if m_sids: sid_parts.append(f"matching SID - {', '.join(sorted(m_sids))}")
            sid_statement = " / ".join(sid_parts) if sid_parts else f"claiming SID - {sid}"

            # Build REBNI Table from collected data
            rebni_table = f"{'PO':<16} {'ASIN':<16} {'SID':<18} {'CP':<15} {'Available REBNI'}\n"
            rebni_table += "═" * 85 + "\n"
            engine = getattr(self.app, 'engine', None)
            if engine and engine.collected_rebni:
                for k, rows in engine.collected_rebni.items():
                    for r_row in rows:
                        r_po = str(r_row.get('po', ''))
                        r_asin = str(r_row.get('asin', ''))
                        r_sid = str(r_row.get('shipment_id', ''))
                        r_cp = float(safe_num(r_row.get('item_cost', 0)))
                        r_avail = int(safe_num(r_row.get('rebni_available', 0)))
                        rebni_table += f"{r_po:<16} {r_asin:<16} {r_sid:<18} {r_cp:<15.2f} {r_avail}\n"
            else:
                rebni_table += "[No REBNI matches found in investigation summary]\n"

            text = (
                f"Hello Team, \n\n"
                f"We see that vendor sent overages of mismatch ASIN in {sid_statement}. "
                f"Below suggested REBNIs are available as per the current REBNI report in line with same state FCs and CP criteria. \n\n"
                f"{rebni_table}\n\n"
                f"Please check and utilize the REBNI and update the remaining PQV units. If suggested REBNI are utilized somewhere else, "
                f"then share Invoice, ASIN and PO level details along with Invoice copy where its matched for validation. \n\n"
                f"Note: \n"
                f"1. If Suggested REBNI comes under same CP limit, the shipment ID and PO shouldn't be the factor. \n"
                f"2. Suggested REBNI comes under same shipment, them CP variance isn't the factor.\n\n"
                f"{signature}"
            )
        elif scenario == "Matching (Bulk)":
            text = (
                f"Hello FC Team, \n\n"
                f"Note: We have already performed the all the virtual research/checks such as cross receiving, overage, REBNI, adjustments etc. "
                f"and need FC support for physical search on floor to locate missing units as per revised SOP and update at the earliest.\n\n"
                f"We are able to see that units are received completely in claiming Shipment ID#{sid} and received units got matched with different invoices where we found shortage units received in matched invoices.\n\n"
                f"Please locate the following ASINs that are missing from PO#{po}, If units not found physically take support from SLP Team and give update.\n\n"
                f"{asin_block}\n\n"
                f"**Post the investigation, please flip it back to our CTI**\n"
                f"C: TOC-India\nT: MFI\nI: PD\nGroup: NOC MFI\n\n"
                f"{signature}"
            )
        elif scenario == "Mismatch ASIN Overages":
            text = (
                f"### Hello FinOps Team,  \n\n"
                f"This is regarding the suggested overage ASIN observed under the claiming Shipment ID [{sid}].  \n"
                f"It is clearly visible that:  \n\n"
                f"- Invoiced Quantity: {t_billed} units\n"
                f"- Received Quantity: {t_received} units\n"
                f"- Overage Units: {t_received - t_billed if t_received > t_billed else 0} units  \n\n"
                f"This indicates that the vendor has sent indifferent ASIN instead of the claiming ASINs.  \n\n"
                f"Request:\n"
                f"- Kindly utilize the suggested overages and resolve the PQV at the earliest.  \n"
                f"- If the suggested overages are not utilized, please obtain a credit note from the vendor code, explaining why different ASIN overages were sent in place of the claiming ASIN.  \n\n"
                f"{signature}"
            )
        else:
            text = f"Hello Team,\n\nScenario '{scenario}' template is pending final formatting logic.\n\n{signature}"

        self.text_area.delete("1.0", tk.END)
        self.text_area.insert(tk.END, text)

# ═══════════════════════════════════════════════════════════
class AnalysisDialog(tk.Toplevel):
    def __init__(self, parent, app, unique_sids):
        super().__init__(parent)
        self.app = app
        self.unique_sids = unique_sids
        self.title("ASIN Level Analysis (CIAT)")
        self.geometry("650x300")
        self.configure(bg="#0f0f1a")
        self.attributes("-topmost", True)
        
        tk.Label(self, text="CIAT ASIN Level Analysis", bg="#0f0f1a", fg="#d9455f", font=("Segoe UI", 14, "bold")).pack(pady=10)
        
        f1 = tk.Frame(self, bg="#0f0f1a")
        f1.pack(fill="x", padx=20, pady=10)
        tk.Label(f1, text="Select Shipment ID:", bg="#0f0f1a", fg="white", font=("Segoe UI", 10)).pack(side="left", padx=5)
        self.sid_var = tk.StringVar(value=unique_sids[0])
        cb = ttk.Combobox(f1, textvariable=self.sid_var, values=unique_sids, state="readonly", font=("Segoe UI", 10), width=30)
        cb.pack(side="left", padx=10)
        if len(unique_sids) == 1:
            cb.config(state="disabled")
            
        f2 = tk.Frame(self, bg="#0f0f1a")
        f2.pack(fill="x", padx=20, pady=10)
        tk.Label(f2, text="CIAT Receive Data:", bg="#0f0f1a", fg="white", font=("Segoe UI", 10)).pack(side="left", padx=5)
        self.ciat_var = tk.StringVar(value=self.app.ciat_path.get())
        tk.Entry(f2, textvariable=self.ciat_var, width=45, font=("Segoe UI", 10), bg="#1e1e3a", fg="white", insertbackground="white", relief="flat").pack(side="left", padx=10)
        tk.Button(f2, text="Browse", command=self._browse_ciat, bg="#2d2d5e", fg="white", relief="flat", cursor="hand2").pack(side="left", padx=5)
        
        f3 = tk.Frame(self, bg="#0f0f1a")
        f3.pack(fill="x", padx=20, pady=10)
        tk.Label(f3, text="Output Mode:", bg="#0f0f1a", fg="white", font=("Segoe UI", 10, "bold")).pack(side="left", padx=5)
        self.output_mode = tk.StringVar(value="new_file")
        tk.Radiobutton(f3, text="Generate New File (Separate Excel)", variable=self.output_mode, value="new_file", fg="white", bg="#0f0f1a", selectcolor="#16213e", font=("Segoe UI", 9)).pack(side="left", padx=10)
        tk.Radiobutton(f3, text="Append to Main Investigation Output", variable=self.output_mode, value="append", fg="white", bg="#0f0f1a", selectcolor="#16213e", font=("Segoe UI", 9)).pack(side="left", padx=10)
        
        tk.Button(self, text="▶ RUN ANALYSIS", command=self._generate, bg="#2d6a4f", fg="white", font=("Segoe UI", 11, "bold"), relief="flat", cursor="hand2", padx=20, pady=8).pack(pady=20)
        
    def _browse_ciat(self):
        f = filedialog.askopenfilename(filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv"), ("All Files", "*.*")])
        if f: self.ciat_var.set(f)
        
    def _generate(self):
        sid = self.sid_var.get().strip()
        ciat_file = self.ciat_var.get().strip()
        if not sid or not ciat_file:
            messagebox.showerror("Error", "Please provide both Shipment ID and CIAT Receive Data file.", parent=self)
            return
            
        if not os.path.exists(ciat_file):
            messagebox.showerror("Error", "CIAT file does not exist.", parent=self)
            return
            
        try:
            engine = getattr(self.app, 'engine', None)
            if not engine or not engine.vendor_raw:
                messagebox.showerror("Error", "Vendor Data is not loaded in memory.", parent=self)
                return
                
            vendor_rows = engine.vendor_raw.get(sid, [])
            if not vendor_rows:
                messagebox.showerror("Error", f"No Vendor Data found for SID {sid}.", parent=self)
                return
                
            def get_idx(keys, headers):
                for k in keys:
                    for i, h in enumerate(headers):
                        if k in str(h).lower(): return i
                return -1
                
            inv_col_idx = get_idx(['invoice number', 'invoice', 'invoice no'], engine.vendor_headers)
            if inv_col_idx == -1:
                messagebox.showerror("Error", "Could not find Invoice Number column in Vendor Data.", parent=self)
                return
                
            inv_numbers = set(str(r[inv_col_idx]).strip().upper() for r in vendor_rows if pd.notna(r[inv_col_idx]) and str(r[inv_col_idx]).strip())
            
            inv_search_df = getattr(self.app, 'inv_search_df', None)
            if inv_search_df is None or inv_search_df.empty:
                messagebox.showerror("Error", "Invoice Search Data is not loaded in memory.", parent=self)
                return
                
            inv_search_df['invoice_number_clean'] = inv_search_df['invoice_number'].astype(str).str.strip().str.upper()
            filtered_inv = inv_search_df[inv_search_df['invoice_number_clean'].isin(inv_numbers)].copy()
            
            if filtered_inv.empty:
                messagebox.showwarning("Warning", f"No matching invoices found in Invoice Search data for SID {sid}.", parent=self)
                return
                
            filtered_inv.loc[:, 'quantity_invoiced'] = pd.to_numeric(filtered_inv['quantity_invoiced'], errors='coerce').fillna(0)
            filtered_inv.loc[:, 'asin_clean'] = filtered_inv['asin'].astype(str).str.strip().str.upper()
            invoiced_grouped = filtered_inv.groupby('asin_clean')['quantity_invoiced'].sum().reset_index()
            invoiced_grouped.rename(columns={'asin_clean': 'Invoiced ASIN', 'quantity_invoiced': 'Invoiced Quantity'}, inplace=True)
            
            if ciat_file.lower().endswith('.csv'): ciat_df = pd.read_csv(ciat_file)
            else: ciat_df = pd.read_excel(ciat_file)
                
            asin_col = next((c for c in ciat_df.columns if 'asin' in str(c).lower()), None)
            qty_col = next((c for c in ciat_df.columns if 'received' in str(c).lower() and ('unit' in str(c).lower() or 'qty' in str(c).lower() or 'quantity' in str(c).lower())), None)
            
            if not asin_col and len(ciat_df.columns) > 4: asin_col = ciat_df.columns[4]
            if not qty_col and len(ciat_df.columns) > 5: qty_col = ciat_df.columns[5]
            
            if not asin_col or not qty_col:
                messagebox.showerror("Error", "Could not identify ASIN or Received Units columns in CIAT data.", parent=self)
                return
                
            ciat_df['asin_clean'] = ciat_df[asin_col].astype(str).str.strip().str.upper()
            ciat_df[qty_col] = pd.to_numeric(ciat_df[qty_col], errors='coerce').fillna(0)
            
            received_grouped = ciat_df.groupby('asin_clean')[qty_col].sum().reset_index()
            received_grouped.rename(columns={'asin_clean': 'Received ASIN', qty_col: 'Received Quantity'}, inplace=True)
            
            max_len = max(len(invoiced_grouped), len(received_grouped))
            inv_asin_list = list(invoiced_grouped['Invoiced ASIN']) + [""] * (max_len - len(invoiced_grouped))
            inv_qty_list = list(invoiced_grouped['Invoiced Quantity']) + [""] * (max_len - len(invoiced_grouped))
            rec_asin_list = list(received_grouped['Received ASIN']) + [""] * (max_len - len(received_grouped))
            rec_qty_list = list(received_grouped['Received Quantity']) + [""] * (max_len - len(received_grouped))
            
            merged = pd.merge(invoiced_grouped, received_grouped, left_on='Invoiced ASIN', right_on='Received ASIN', how='outer')
            merged['Unique ASIN'] = merged['Invoiced ASIN'].combine_first(merged['Received ASIN'])
            merged['Invoiced Quantity'] = pd.to_numeric(merged['Invoiced Quantity'], errors='coerce').fillna(0)
            merged['Received Quantity'] = pd.to_numeric(merged['Received Quantity'], errors='coerce').fillna(0)
            merged['Overage / Shortage'] = merged['Received Quantity'] - merged['Invoiced Quantity']
            
            def get_status(diff):
                if diff < 0: return "Shortage"
                if diff > 0: return "Overage"
                return "Matched"
                
            merged['Status'] = merged['Overage / Shortage'].apply(get_status)
            final_len = max(max_len, len(merged))
            final_len = max(final_len, 5)
            
            inv_asin_list += [""] * (final_len - len(inv_asin_list))
            inv_qty_list += [""] * (final_len - len(inv_qty_list))
            rec_asin_list += [""] * (final_len - len(rec_asin_list))
            rec_qty_list += [""] * (final_len - len(rec_qty_list))
            
            total_inv = merged['Invoiced Quantity'].sum()
            total_rec = merged['Received Quantity'].sum()
            total_over = merged[merged['Overage / Shortage'] > 0]['Overage / Shortage'].sum()
            total_short = abs(merged[merged['Overage / Shortage'] < 0]['Overage / Shortage'].sum())
            
            metric_names = ["", "Total Invoiced Qty", "Total Received Qty", "Overage", "Shortage"] + [""] * (final_len - 5)
            metric_vals = ["", total_inv, total_rec, total_over, total_short] + [""] * (final_len - 5)
            
            out_dict = {
                "Invoiced ASIN": inv_asin_list,
                "Invoiced Quantity": inv_qty_list,
                " ": [""] * final_len,
                "Received ASIN": rec_asin_list,
                "Received Quantity": rec_qty_list,
                "  ": [""] * final_len,
                "Unique ASIN": list(merged['Unique ASIN']) + [""] * (final_len - len(merged)),
                "Total Invoiced Qty": list(merged['Invoiced Quantity']) + [""] * (final_len - len(merged)),
                "Total Received Qty": list(merged['Received Quantity']) + [""] * (final_len - len(merged)),
                "Overage / Shortage": list(merged['Overage / Shortage']) + [""] * (final_len - len(merged)),
                "Status": list(merged['Status']) + [""] * (final_len - len(merged)),
                "   ": [""] * final_len,
                "    ": metric_names,
                "     ": metric_vals
            }
            
            final_df = pd.DataFrame(out_dict)
            
            # --- SHIPMENT LEVEL ANALYSIS ---
            po_col = next((c for c in ciat_df.columns if str(c).strip().upper() == 'PO'), None)
            if not po_col:
                po_col = next((c for c in ciat_df.columns if 'po' in str(c).lower().strip().split()), None)
                
            shipment_merges = []
            if po_col:
                ciat_df['po_clean'] = ciat_df[po_col].astype(str).str.replace(r'\.0$', '', regex=True).str.strip().str.upper()
                inv_search_df['po_clean'] = inv_search_df['purchase_order_id'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip().str.upper()
                filtered_inv_ship = inv_search_df[inv_search_df['invoice_number_clean'].isin(inv_numbers)].copy()
                filtered_inv_ship['po_clean'] = filtered_inv_ship['purchase_order_id'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip().str.upper()
                filtered_inv_ship['quantity_invoiced'] = pd.to_numeric(filtered_inv_ship['quantity_invoiced'], errors='coerce').fillna(0)
                filtered_inv_ship['asin_clean'] = filtered_inv_ship['asin'].astype(str).str.strip().str.upper()
                
                ciat_ship_grouped = ciat_df.groupby(['po_clean', 'asin_clean'])[qty_col].sum().reset_index()
                inv_ship_grouped = filtered_inv_ship.groupby(['po_clean', 'invoice_number_clean', 'asin_clean'])['quantity_invoiced'].sum().reset_index()
                
                all_pos = sorted(list(set(ciat_ship_grouped['po_clean']).union(set(inv_ship_grouped['po_clean']))))
                
                shipment_rows = []
                ship_total_inv = 0
                ship_total_rec = 0
                ship_total_over = 0
                ship_total_short = 0
                
                current_row_idx = 2 # Excel row 1 is header, data starts at row 2
                
                for po in all_pos:
                    po_ciat = ciat_ship_grouped[ciat_ship_grouped['po_clean'] == po]
                    po_inv = inv_ship_grouped[inv_ship_grouped['po_clean'] == po]
                    
                    po_asins = sorted(list(set(po_ciat['asin_clean']).union(set(po_inv['asin_clean']))))
                    po_start_row = current_row_idx
                    
                    for asin in po_asins:
                        asin_ciat = po_ciat[po_ciat['asin_clean'] == asin]
                        rec_qty = asin_ciat[qty_col].iloc[0] if not asin_ciat.empty else 0
                        
                        asin_invs = po_inv[po_inv['asin_clean'] == asin]
                        inv_list = []
                        if not asin_invs.empty:
                            for _, r in asin_invs.iterrows():
                                inv_list.append((r['invoice_number_clean'], r['quantity_invoiced']))
                        else:
                            inv_list.append(("No Invoice", 0))
                            
                        total_inv_qty = sum(x[1] for x in inv_list)
                        total_rec_qty = rec_qty
                        diff = total_rec_qty - total_inv_qty
                        status = "Matched"
                        if diff > 0: status = "Overage"
                        elif diff < 0: status = "Shortage"
                        
                        ship_total_inv += total_inv_qty
                        ship_total_rec += total_rec_qty
                        if diff > 0: ship_total_over += diff
                        elif diff < 0: ship_total_short += abs(diff)
                        
                        num_rows = len(inv_list)
                        asin_start_row = current_row_idx
                        asin_end_row = current_row_idx + num_rows - 1
                        
                        for i, (inv_no, inv_q) in enumerate(inv_list):
                            row_dict = {
                                "PO": po if i == 0 else "",
                                "Invoice Number": inv_no,
                                "Invoiced ASIN": asin if inv_q > 0 or inv_no != "No Invoice" else "",
                                "Invoiced Quantity": inv_q if inv_q > 0 or inv_no != "No Invoice" else "",
                                " ": "",
                                "Received ASIN": asin if i == 0 else "",
                                "Received Quantity": total_rec_qty if i == 0 else "",
                                "  ": "",
                                "Total Invoiced Qty": total_inv_qty if i == 0 else "",
                                "Total Received Qty": total_rec_qty if i == 0 else "",
                                "Overage / Shortage": diff if i == 0 else "",
                                "Status": status if i == 0 else "",
                                "   ": "",
                                "    ": "",
                                "     ": ""
                            }
                            shipment_rows.append(row_dict)
                            
                        if num_rows > 1:
                            from openpyxl.utils import get_column_letter
                            for col_idx in [6, 7, 9, 10, 11, 12]: # F, G, I, J, K, L
                                col_letter = get_column_letter(col_idx)
                                shipment_merges.append(f"{col_letter}{asin_start_row}:{col_letter}{asin_end_row}")
                                
                        current_row_idx += num_rows
                        
                    if current_row_idx - 1 > po_start_row:
                        shipment_merges.append(f"A{po_start_row}:A{current_row_idx - 1}")
                        
                metric_names = ["", "Total Invoiced Qty", "Total Received Qty", "Overage", "Shortage"]
                metric_vals = ["", ship_total_inv, ship_total_rec, ship_total_over, ship_total_short]
                
                for i in range(len(shipment_rows)):
                    if i < len(metric_names):
                        shipment_rows[i]["    "] = metric_names[i]
                        shipment_rows[i]["     "] = metric_vals[i]
                
                if len(shipment_rows) < len(metric_names):
                    for i in range(len(shipment_rows), len(metric_names)):
                        pad_dict = {k: "" for k in shipment_rows[0].keys()}
                        pad_dict["    "] = metric_names[i]
                        pad_dict["     "] = metric_vals[i]
                        shipment_rows.append(pad_dict)
                        
                shipment_df = pd.DataFrame(shipment_rows)
            else:
                shipment_df = pd.DataFrame({"Message": ["PO column not found in CIAT data. Shipment Level Analysis skipped."]})
                shipment_merges = []
            
            if self.output_mode.get() == "new_file":
                save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=f"ASIN_Level_Analysis_{sid}.xlsx", filetypes=[("Excel Files", "*.xlsx")])
                if save_path:
                    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                        final_df.to_excel(writer, sheet_name='ASIN Level Analysis', index=False)
                        shipment_df.to_excel(writer, sheet_name='Shipment Level Analysis', index=False)
                        
                        # Apply custom merging for Shipment Level Analysis
                        if shipment_merges:
                            ws = writer.sheets['Shipment Level Analysis']
                            for merge_range in shipment_merges:
                                ws.merge_cells(merge_range)
                                start_cell = merge_range.split(':')[0]
                                ws[start_cell].alignment = Alignment(vertical='center', horizontal='center')
                    messagebox.showinfo("Success", f"Analysis saved successfully to:\n{save_path}", parent=self)
                    self.destroy()
            else:
                # Append Mode: store dataframes in engine to be saved when write_excel is called
                engine.ciat_asin_df = final_df
                engine.ciat_shipment_df = shipment_df
                engine.ciat_shipment_merges = shipment_merges
                messagebox.showinfo("Success", "Analysis generated and stored!\n\nThe ASIN Level and Shipment Level sheets will be appended to your main investigation output when you click 'SAVE DIRECTLY'.", parent=self)
                self.destroy()
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate analysis:\n{str(e)}", parent=self)

class PreviewPanel(tk.Toplevel):
    COLS      = ['Barcode', 'Inv no', 'SID', 'PO', 'ASIN', 'Inv Qty',
                 'Rec Qty', 'Mtc Qty', 'Mtc Inv', 'Mtc ASIN', 'Mtc PO', 'FC', 'Remarks', 'Date', 'CP']
    COL_W_PX  = [130, 160, 130, 90, 110, 60, 60, 60, 160, 130, 130, 80, 240, 150, 180]

    def __init__(self, parent):
        super().__init__(parent)
        self._app = None
        self.title("Investigation Preview — v7.7 (editable)")
        self.geometry("1400x750")
        self.configure(bg="#0f0f1a")
        self.resizable(True, True)
        self.protocol("WM_DELETE_WINDOW", self.withdraw)
        self.protocol("WM_DESTROY", lambda *args: None) # ignore secondary destroys
        self._undo_stack = []   # v6.0.0: undo snapshots (max 10)
        self._redo_stack = []   # v6.0.0: redo snapshots
        self._clipboard_cell = None  # for right-click paste
        
        hdr_frame = tk.Frame(self, bg="#16213e")
        hdr_frame.pack(fill="x")
        tk.Label(hdr_frame, text="  Live Investigation Preview",
                 bg="#16213e", fg="#4a9eff", font=("Segoe UI", 10, "bold"), height=2).pack(side="left")
        
        # v6.0.0 Fix: Smaller buttons (font 9, less padding)

        tk.Button(hdr_frame, text="✔ CONFIRM EDITS", command=self.confirm_edits,
                  bg="#1a5a1a", fg="#90ff90", font=("Segoe UI", 9, "bold"), relief="flat", padx=10, pady=4, cursor="hand2").pack(side="right", padx=6, pady=4)
        
        tk.Button(hdr_frame, text="↪ REDO", command=self._redo,
                  bg="#5e3c8b", fg="white", font=("Segoe UI", 9, "bold"), relief="flat", padx=8, pady=4, cursor="hand2").pack(side="right", padx=2, pady=4)

        tk.Button(hdr_frame, text="↩ UNDO", command=self._undo,
                  bg="#8b5e3c", fg="white", font=("Segoe UI", 9, "bold"), relief="flat", padx=8, pady=4, cursor="hand2").pack(side="right", padx=2, pady=4)
        
        tk.Button(hdr_frame, text="✉ CORRESPONDENCE", command=self.show_correspondence,
                  bg="#3949ab", fg="white", font=("Segoe UI", 9, "bold"), relief="flat", padx=10, pady=4, cursor="hand2").pack(side="right", padx=6, pady=4)
        
        tk.Button(hdr_frame, text="🔍 MTC QTY", command=self._show_mtc_qty_lookup,
                  bg="#6f42c1", fg="white", font=("Segoe UI", 9, "bold"), relief="flat", padx=10, pady=4, cursor="hand2").pack(side="right", padx=6, pady=4)
                  
        tk.Button(hdr_frame, text="💵 CP Variance", command=self._show_cp_variance,
                  bg="#00563b", fg="#e0ffe0", font=("Segoe UI", 9, "bold"), relief="flat", padx=10, pady=4, cursor="hand2").pack(side="right", padx=6, pady=4)

        tk.Button(hdr_frame, text="📋 INV QTY", command=self._show_inv_qty_lookup,
                  bg="#1c2c42", fg="#80a0ff", font=("Segoe UI", 9, "bold"), relief="flat", padx=10, pady=4, cursor="hand2").pack(side="right", padx=6, pady=4)

        # v6.2.5: New Received Quantity lookup for main unit
        tk.Button(hdr_frame, text="📦 RECEIVED QTY", command=self._show_rec_qty_lookup,
                  bg="#d4a017", fg="black", font=("Segoe UI", 9, "bold"), relief="flat", padx=10, pady=4, cursor="hand2").pack(side="right", padx=6, pady=4)

        # v7.2.3: On Hold Qty lookup from Invoice Search
        tk.Button(hdr_frame, text="📌 ON HOLD QTY", command=self._show_on_hold_lookup,
                  bg="#8b0000", fg="#ffcccc", font=("Segoe UI", 9, "bold"), relief="flat", padx=10, pady=4, cursor="hand2").pack(side="right", padx=6, pady=4)

        # Action bar moved "somewhat upper" (above the treeview)
        action_bar = tk.Frame(self, bg="#0f0f1a")
        action_bar.pack(fill="x", padx=8, pady=4)
        tk.Button(action_bar, text="📊 ANALYSE", command=self._show_ciat_analysis, bg="#d9455f", fg="white", font=("Segoe UI", 9, "bold"), relief="flat", padx=15, pady=4, cursor="hand2").pack(side="left", padx=10)
        tk.Button(action_bar, text="💾 SAVE DIRECTLY", command=self._save_preview, bg="#005a9e", fg="white", font=("Segoe UI", 9, "bold"), relief="flat", padx=10, pady=4, cursor="hand2").pack(side="right", padx=6)
        tk.Button(action_bar, text="🗑 DELETE ROW", command=self._delete_selected_rows, bg="#5c1a1a", fg="#ff9090", font=("Segoe UI", 9, "bold"), relief="flat", padx=10, pady=4, cursor="hand2").pack(side="right", padx=6)
        tk.Button(action_bar, text="Clear All", command=self.clear_all, bg="#2d2d5e", fg="white", font=("Segoe UI", 9), relief="flat", padx=10, pady=4, cursor="hand2").pack(side="right", padx=6)

        frame = tk.Frame(self, bg="#0f0f1a")
        frame.pack(fill="both", expand=True, padx=8, pady=6)
        vsb = ttk.Scrollbar(frame, orient="vertical")
        hsb = ttk.Scrollbar(frame, orient="horizontal")
        self.tree = ttk.Treeview(frame, columns=self.COLS, show='headings', yscrollcommand=vsb.set, xscrollcommand=hsb.set, height=22)
        vsb.config(command=self.tree.yview); hsb.config(command=self.tree.xview)
        for col, w in zip(self.COLS, self.COL_W_PX):
            self.tree.heading(col, text=col); self.tree.column(col, width=w, minwidth=40, anchor='w', stretch=False)
        self.tree.grid(row=0, column=0, sticky="nsew"); vsb.grid(row=0, column=1, sticky="ns"); hsb.grid(row=1, column=0, sticky="ew")
        frame.grid_rowconfigure(0, weight=1); frame.grid_columnconfigure(0, weight=1)
        self.tree.bind('<Double-1>', self._on_double_click)
        self.tree.bind('<Delete>', lambda e: self._delete_selected_rows())
        self.tree.bind('<Button-3>', self._show_context_menu)
        self.tree.bind('<Double-3>', self._on_double_right_click)
        self._row_data = {}

        # Right-click context menu
        self._ctx_menu = tk.Menu(self, tearoff=0, bg="#1e1e3a", fg="#e0e0e0", font=("Segoe UI", 9))
        self._ctx_menu.add_command(label="📋 Copy Cell", command=self._ctx_copy)
        self._ctx_menu.add_command(label="📌 Paste Cell", command=self._ctx_paste)
        self._ctx_menu.add_separator()
        self._ctx_menu.add_command(label="🗑 Clear Cell(s)", command=self._ctx_clear)
        self._ctx_menu.add_separator()
        self._ctx_menu.add_command(label="➕ Insert Blank Row", command=self._ctx_insert_row)
        self._ctx_menu.add_command(label="🔍 Find (Ctrl+F)", command=self._show_find_dialog)
        self._ctx_menu.add_separator()
        self._ctx_menu.add_command(label="🗑 Delete Row(s)", command=self._delete_selected_rows)
        self._ctx_iid = None
        self._ctx_col_idx = None
        
        self._search_matches = []
        self._search_idx = -1
        self.bind("<Control-f>", lambda e: self._show_find_dialog())
        self.tree.bind("<Control-f>", lambda e: self._show_find_dialog())

        s = ttk.Style()
        s.configure("Treeview", font=("Calibri", 10), rowheight=24, background="#1e1e3a", foreground="#e0e0e0", fieldbackground="#1e1e3a")
        s.configure("Treeview.Heading", font=("Calibri", 10, "bold"), background="#203864", foreground="white")
        for tag, bg, fg in [('header','#203864','white'),('shortage_red','#ffcccc','#9c0006'),('crosspo','#2a1a00','#f0c060'),('invalid_ibc_pbc','#FFE0F0','#9c0006')]:
            self.tree.tag_configure(tag, background=bg, foreground=fg)


    def _show_ciat_analysis(self):
        unique_sids = list(set(str(d.get('SID', '')).strip() for d in self._row_data.values() if str(d.get('SID', '')).strip()))
        if not unique_sids:
            messagebox.showwarning("Warning", "No Shipment IDs found in the current preview data.", parent=self)
            return
        AnalysisDialog(self, self._app, unique_sids)

    def _show_mtc_qty_lookup(self):
        dlg = tk.Toplevel(self)
        dlg.title("Lookup Matched Quantity")
        dlg.configure(bg="#0d1117")
        dlg.attributes("-topmost", True)
        
        tk.Label(dlg, text="SID:", bg="#0d1117", fg="white", font=("Segoe UI", 10)).grid(row=0, column=0, padx=10, pady=(15,5), sticky='e')
        sid_ent = tk.Entry(dlg, font=("Segoe UI", 10), bg="#21262d", fg="white", insertbackground="white")
        sid_ent.grid(row=0, column=1, padx=10, pady=(15,5))

        tk.Label(dlg, text="PO:", bg="#0d1117", fg="white", font=("Segoe UI", 10)).grid(row=1, column=0, padx=10, pady=(15,5), sticky='e')
        po_ent = tk.Entry(dlg, font=("Segoe UI", 10), bg="#21262d", fg="white", insertbackground="white")
        po_ent.grid(row=1, column=1, padx=10, pady=(15,5))
        
        tk.Label(dlg, text="ASIN:", bg="#0d1117", fg="white", font=("Segoe UI", 10)).grid(row=2, column=0, padx=10, pady=(15,5), sticky='e')
        asin_ent = tk.Entry(dlg, font=("Segoe UI", 10), bg="#21262d", fg="white", insertbackground="white")
        asin_ent.grid(row=2, column=1, padx=10, pady=(15,5))
        
        # v6.0.0 Fix: Use Text widget instead of Label to show all entries
        res_frame = tk.Frame(dlg, bg="#0d1117")
        res_frame.grid(row=4, column=0, columnspan=2, pady=(10,15), padx=10)
        res_txt = tk.Text(res_frame, width=75, height=8, bg="#0d1117", fg="#3fb950", font=("Segoe UI", 10), state='disabled', bd=0, wrap="none")
        res_txt.grid(row=0, column=0, sticky="nsew")
        vsb = ttk.Scrollbar(res_frame, orient="vertical", command=res_txt.yview)
        hsb = ttk.Scrollbar(res_frame, orient="horizontal", command=res_txt.xview)
        res_txt.config(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        res_txt.tag_configure("error", foreground="#f85149")
        res_txt.tag_configure("success", foreground="#3fb950")
        
        sel = self.tree.selection()
        if sel:
            d = self._row_data.get(sel[0], {})
            sid_ent.insert(0, str(d.get('SID', '')))
            po_ent.insert(0, str(d.get('PO', '')))
            asin_ent.insert(0, str(d.get('ASIN', '')))

        def do_lookup():
            s = extract_sid(sid_ent.get())
            p = clean(po_ent.get())
            a = clean(asin_ent.get())
            
            res_txt.config(state='normal')
            res_txt.delete('1.0', 'end')
            
            engine = getattr(self._app, 'engine', None)
            if not engine:
                res_txt.insert('end', "Engine not connected.", "error")
                res_txt.config(state='disabled')
                return
            
            rows = engine.inv_p.get((s, p, a), [])
            if not rows:
                res_txt.insert('end', "No matches found.", "error")
                res_txt.config(state='disabled')
                return
            
            # v6.0.1 Fix: Deduplicate entries and show Inv, PO, ASIN, Mtc Qty
            seen = set()
            dlg.current_matches = []
            for r in rows:
                k = (r.get('mtc_inv',''), r.get('mtc_po',''), r.get('mtc_asin',''), fmt_qty(r.get('mtc_qty','')))
                if k not in seen:
                    seen.add(k)
                    dlg.current_matches.append(r)
            
            res_txt.insert('end', f"Found {len(dlg.current_matches)} unique matches:\n", "success")
            msg = "\n".join([f"Inv: {r.get('mtc_inv','')} | PO: {r.get('mtc_po','')} | ASIN: {r.get('mtc_asin','')} | Mtc Qty: {fmt_qty(r.get('mtc_qty',''))}" for r in dlg.current_matches])
            res_txt.insert('end', msg, "success")
            res_txt.config(state='disabled')
            
        def do_replace():
            sel = self.tree.selection()
            if not sel:
                messagebox.showinfo("Replace", "No destination rows selected in Preview Panel.", parent=dlg)
                return
            
            matches = getattr(dlg, 'current_matches', [])
            if not matches:
                messagebox.showinfo("Replace", "No matches found from search. Search for data first.", parent=dlg)
                return
                
            self._take_snapshot()
            replaced_count = 0
            deleted_count = 0
            
            for i, iid in enumerate(sel):
                if i >= len(matches):
                    # User requested deletion of excess selected invalid duplicate rows
                    if iid in self._row_data:
                        self._row_data.pop(iid, None)
                        try:
                            self.tree.delete(iid)
                            deleted_count += 1
                        except: pass
                    continue
                
                m = matches[i]
                if iid in self._row_data:
                    d = self._row_data[iid]
                    if 'mtc_inv' in m: d['Mtc Inv'] = str(m['mtc_inv']).strip()
                    if 'mtc_po' in m: d['Mtc PO'] = str(m['mtc_po']).strip()
                    if 'mtc_asin' in m: d['Mtc ASIN'] = str(m['mtc_asin']).strip()
                    if 'mtc_qty' in m: d['Mtc Qty'] = fmt_qty(safe_num(m['mtc_qty'])) # Enforces clean integer strings
                    
                    try:
                        vals = list(self.tree.item(iid, 'values'))
                        if 'Mtc Inv' in self.COLS: vals[self.COLS.index('Mtc Inv')] = d['Mtc Inv']
                        if 'Mtc PO' in self.COLS: vals[self.COLS.index('Mtc PO')] = d['Mtc PO']
                        if 'Mtc ASIN' in self.COLS: vals[self.COLS.index('Mtc ASIN')] = d['Mtc ASIN']
                        if 'Mtc Qty' in self.COLS: vals[self.COLS.index('Mtc Qty')] = d['Mtc Qty']
                        self.tree.item(iid, values=vals)
                        replaced_count += 1
                    except Exception: pass
            
            msg_txt = f"Replaced data correctly into {replaced_count} row(s) sequentially!"
            if deleted_count > 0:
                msg_txt += f"\nDeleted {deleted_count} excess erroneous row(s) from the selection."
            messagebox.showinfo("Success", msg_txt, parent=dlg)

        tf = tk.Frame(dlg, bg="#0d1117")
        tf.grid(row=3, column=0, columnspan=2, pady=10)
        tk.Button(tf, text="🔍 Search Data", bg="#6f42c1", fg="white", font=("Segoe UI", 10, "bold"),
                  command=do_lookup, cursor="hand2", padx=15).pack(side="left", padx=5)
        tk.Button(tf, text="🔄 Replace into Highlighted Rows", bg="#2ea043", fg="white", font=("Segoe UI", 10, "bold"),
                  command=do_replace, cursor="hand2", padx=15).pack(side="left", padx=5)
        
        dlg.update_idletasks()
        px_dlg = self.winfo_x() + (self.winfo_width() - dlg.winfo_width()) // 2
        py_dlg = self.winfo_y() + 50
        dlg.geometry(f"+{px_dlg}+{py_dlg}")
        sid_ent.focus_set()

    def _show_cp_variance(self):
        dlg = tk.Toplevel(self)
        dlg.title("CP Variance Comparison")
        dlg.configure(bg="#0d1117")
        dlg.attributes("-topmost", True)
        dlg.geometry("500x350")
        
        data = {'source': None, 'target': None, 'source_iid': None, 'target_iid': None}
        
        sf = tk.LabelFrame(dlg, text="Source ASIN (Selected Row Data)", bg="#0d1117", fg="#4a9eff", font=("Segoe UI", 10, "bold"))
        sf.pack(fill="x", padx=10, pady=5)
        s_lbl = tk.Label(sf, text="Not set", bg="#0d1117", fg="white", font=("Segoe UI", 9))
        s_lbl.pack(fill="x", pady=5)
        
        tf = tk.LabelFrame(dlg, text="Target ASIN (Selected Row Data)", bg="#0d1117", fg="#f0a500", font=("Segoe UI", 10, "bold"))
        tf.pack(fill="x", padx=10, pady=5)
        t_lbl = tk.Label(tf, text="Not set", bg="#0d1117", fg="white", font=("Segoe UI", 9))
        t_lbl.pack(fill="x", pady=5)
        
        res_lbl = tk.Label(dlg, text="", bg="#0d1117", fg="white", font=("Segoe UI", 11, "bold"))
        res_lbl.pack(fill="x", pady=10)

        def get_item_cost(sid, po, asin):
            engine = getattr(self._app, 'engine', None)
            if not engine or not hasattr(engine, 'rebni_p') or not engine.rebni_p: return None
            
            c_sid, c_po, c_asin = clean(sid), clean(po), clean(asin)
            
            # 1. Exact match (SID, PO, ASIN)
            key = (c_sid, c_po, c_asin)
            if key in engine.rebni_p:
                for r in engine.rebni_p[key]:
                    if pd.notna(r.get('item_cost')):
                        return safe_num(r.get('item_cost', 0))
                        
            # 2. Fallback: (SID, ASIN) - same shipment, cross POs / overages
            s_key = (c_sid, c_asin)
            if hasattr(engine, 'rebni_s') and s_key in engine.rebni_s:
                for r in engine.rebni_s[s_key]:
                    if pd.notna(r.get('item_cost')):
                        return safe_num(r.get('item_cost', 0))
            
            # v6.1.2 Decision: Global fallback (searching across all shipments) removed to prevent incorrect data matching.
            return None

        def eval_variance():
            if not data['source'] or not data['target']: return
            s_cp = data['source']['cp']
            t_cp = data['target']['cp']
            low, high = data['source']['low'], data['source']['high']
            if low <= t_cp <= high:
                res_lbl.config(text=f"✔ MATCH: Target CP (${t_cp:.2f})\n falls within 10% tolerance", fg="#3fb950")
            else:
                res_lbl.config(text=f"❌ MISMATCH: Target CP (${t_cp:.2f})\n violates 10% tolerance limit", fg="#f85149")
                iid = data['target_iid']
                old_rem = self.tree.set(iid, 'Remarks')
                if "Not matching within 10% cp range" not in old_rem:
                    new_rem = (old_rem + " | Not matching within 10% cp range").strip(" |")
                    self.tree.set(iid, 'Remarks', new_rem)
                    self._row_data[iid]['Remarks'] = new_rem
                    self.tree.item(iid, tags=('shortage_red',))

        def _handle_set(btn_type):
            sel = self.tree.selection()
            if not sel: return
            iid = sel[0]
            r = self._row_data.get(iid, {})
            # v6.1.2 Fix: Both Source and Target extract exactly from Main PO (Col D) and Main ASIN (Col E)
            asin = clean(r.get('ASIN', ''))
            po = clean(r.get('PO', ''))
            inv = clean(r.get('Inv no', ''))
            
            # Sub-rows match items to a parent's SID
            sid = extract_sid(r.get('SID',''))
            if getattr(self._app, 'engine', None) and hasattr(self._app.engine, 'cache_sid'):
                if inv in self._app.engine.cache_sid:
                    sid = self._app.engine.cache_sid[inv]

            cp = get_item_cost(sid, po, asin)
            if cp is None:
                # Prompt user for manual CP if it's completely missing from REBNI
                user_cp = simpledialog.askfloat("CP Not Found", 
                    f"Could not automatically find item_cost in REBNI for:\nSID: {sid}\nPO: {po}\nASIN: {asin}\n\nPlease enter the CP value manually:", 
                    parent=dlg, minvalue=0.0)
                if user_cp is not None and user_cp > 0:
                    cp = user_cp
                else:
                    return
            
            low, high = cp * 0.90, cp * 1.10
            lbl_txt = f"SID: {sid} | PO: {po} | ASIN: {asin}\nCP: ${cp:.2f} (10% Range: ${low:.2f} - ${high:.2f})"
            data[btn_type], data[f'{btn_type}_iid'] = {'cp':cp, 'low':low, 'high':high}, iid
            if btn_type == 'source': s_lbl.config(text=lbl_txt)
            else: t_lbl.config(text=lbl_txt)
            if data['source'] and data['target']: eval_variance()

        bf = tk.Frame(dlg, bg="#0d1117")
        bf.pack(fill="x", pady=10)
        tk.Button(bf, text="Set Source\nfrom Selected Row", command=lambda: _handle_set('source'), bg="#21262d", fg="#4a9eff", font=("Segoe UI", 9, "bold")).pack(side="left", expand=True, padx=5)
        tk.Button(bf, text="Set Target\nfrom Selected Row", command=lambda: _handle_set('target'), bg="#21262d", fg="#f0a500", font=("Segoe UI", 9, "bold")).pack(side="left", expand=True, padx=5)


    def _show_inv_qty_lookup(self):
        """Lookup Invoice Quantity from Invoice Search data and auto-fill the row."""
        dlg = tk.Toplevel(self)
        dlg.title("Lookup Invoice Quantity")
        dlg.configure(bg="#0d1117")
        dlg.attributes("-topmost", True)
        dlg.geometry("520x380")

        data = {'inv': '', 'asin': '', 'target_iid': None}

        # Invoice field
        tk.Label(dlg, text="Invoice No:", bg="#0d1117", fg="white",
                 font=("Segoe UI", 10)).grid(row=0, column=0, padx=10, pady=(15,5), sticky='e')
        inv_ent = tk.Entry(dlg, font=("Segoe UI", 10), bg="#21262d", fg="white",
                           insertbackground="white", width=30)
        inv_ent.grid(row=0, column=1, padx=10, pady=(15,5), sticky='w')

        # ASIN field
        tk.Label(dlg, text="ASIN:", bg="#0d1117", fg="white",
                 font=("Segoe UI", 10)).grid(row=1, column=0, padx=10, pady=5, sticky='e')
        asin_ent = tk.Entry(dlg, font=("Segoe UI", 10), bg="#21262d", fg="white",
                            insertbackground="white", width=30)
        asin_ent.grid(row=1, column=1, padx=10, pady=5, sticky='w')

        # Target row label
        tgt_lbl = tk.Label(dlg, text="Target row: (not set)", bg="#0d1117", fg="#888888",
                           font=("Segoe UI", 9), wraplength=480, justify='left')
        tgt_lbl.grid(row=2, column=0, columnspan=2, padx=10, pady=4, sticky='w')

        # Result label
        res_lbl = tk.Label(dlg, text="", bg="#0d1117", fg="#f0c060",
                           font=("Segoe UI", 12, "bold"), wraplength=480)
        res_lbl.grid(row=5, column=0, columnspan=2, pady=(10,15))

        def _set_inv_from_selection():
            sel = self.tree.selection()
            if not sel: return
            iid = sel[0]
            r = self._row_data.get(iid, {})
            inv_val = clean(r.get('Inv no', ''))  # Column B strictly
            if inv_val and not inv_val.startswith('['):
                inv_ent.delete(0, tk.END)
                inv_ent.insert(0, inv_val)
                data['target_iid'] = iid
                tgt_lbl.config(text=f"Target row: Inv={inv_val} | ASIN={clean(r.get('ASIN',''))} | PO={clean(r.get('PO',''))}", fg="#4a9eff")

        def _set_asin_from_selection():
            sel = self.tree.selection()
            if not sel: return
            iid = sel[0]
            r = self._row_data.get(iid, {})
            asin_val = clean(r.get('ASIN', ''))  # Column E strictly
            if asin_val and not asin_val.startswith('['):
                asin_ent.delete(0, tk.END)
                asin_ent.insert(0, asin_val)
                if not data['target_iid']:
                    data['target_iid'] = iid

        def do_lookup():
            ino = inv_ent.get().strip()
            asn = asin_ent.get().strip()
            if not ino or not asn:
                res_lbl.config(text="Please provide both Invoice No and ASIN.", fg="#f85149")
                return

            engine = getattr(self._app, 'engine', None)
            if not engine:
                res_lbl.config(text="Engine not connected.", fg="#f85149")
                return

            found_qty = engine._resolve_inv_qty(ino, asn, None)

            if found_qty is not None:
                qty_int = int(found_qty)
                res_lbl.config(text=f"\u2714 Inv Qty = {qty_int} units", fg="#3fb950")
                # Auto-fill Inv Qty in the target row
                iid = data.get('target_iid')
                if iid and iid in self._row_data:
                    self.tree.set(iid, 'Inv Qty', str(qty_int))
                    self._row_data[iid]['Inv Qty'] = str(qty_int)
                    res_lbl.config(text=f"\u2714 Inv Qty = {qty_int} units  (row updated!)", fg="#3fb950")
            else:
                res_lbl.config(text="Not found in Invoice Search data.", fg="#f85149")

        # Buttons row 1: Set from selection
        bf1 = tk.Frame(dlg, bg="#0d1117")
        bf1.grid(row=3, column=0, columnspan=2, pady=6)
        tk.Button(bf1, text="\u2b06 Set Invoice\nfrom Selected Row", command=_set_inv_from_selection,
                  bg="#21262d", fg="#4a9eff", font=("Segoe UI", 9, "bold"),
                  cursor="hand2", padx=10, pady=4).pack(side="left", padx=8)
        tk.Button(bf1, text="\u2b06 Set ASIN\nfrom Selected Row", command=_set_asin_from_selection,
                  bg="#21262d", fg="#f0a500", font=("Segoe UI", 9, "bold"),
                  cursor="hand2", padx=10, pady=4).pack(side="left", padx=8)

        # Button row 2: Search
        tk.Button(dlg, text="\ud83d\udd0d Search Inv Qty", bg="#238636", fg="white",
                  font=("Segoe UI", 10, "bold"), command=do_lookup,
                  cursor="hand2", padx=20).grid(row=4, column=0, columnspan=2, pady=8)

        dlg.update_idletasks()
        px_dlg = self.winfo_x() + (self.winfo_width() - dlg.winfo_width()) // 2
        py_dlg = self.winfo_y() + 50
        dlg.geometry(f"+{px_dlg}+{py_dlg}")
        inv_ent.focus_set()

    def show_correspondence(self):
        all_rows = self.get_all_rows()
        if not all_rows:
            messagebox.showinfo("No Data", "Investigation preview is empty.", parent=self)
            return
        CorrespondenceDialog(self, all_rows, app=self._app)

    def _show_rec_qty_lookup(self):
        """v6.2.5: Lookup expected Received Qty and available units for the main unit (SID/PO/ASIN)."""
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Selection Required", "Please select a row in the preview panel first.")
            return
            
        iid = sel[0]
        r = self._row_data.get(iid, {})
        asin = clean(r.get('ASIN', ''))
        po = clean(r.get('PO', ''))
        inv = clean(r.get('Inv no', ''))
        
        # Determine SID (check cache or row)
        sid = extract_sid(r.get('SID',''))
        engine = getattr(self._app, 'engine', None)
        if engine and hasattr(engine, 'cache_sid'):
            if inv in engine.cache_sid:
                sid = engine.cache_sid[inv]
        
        if not all([sid, po, asin]):
            messagebox.showerror("Data Missing", f"Selected row is missing critical data:\nSID: {sid}\nPO: {po}\nASIN: {asin}")
            return
            
        dlg = tk.Toplevel(self)
        dlg.title(f"Received Qty Lookup — {asin}")
        dlg.geometry("550x420")
        dlg.configure(bg="#0d1117")
        dlg.attributes("-topmost", True)
        
        tk.Label(dlg, text="📦  Shipment Unit Identity (REBNI Data)", bg="#16213e", fg="#f0a500", font=("Segoe UI", 11, "bold"), height=2).pack(fill="x")
        
        info_f = tk.Frame(dlg, bg="#0d1117", padx=20, pady=15)
        info_f.pack(fill="x")
        
        details = [("SID:", sid), ("PO:", po), ("ASIN:", asin)]
        for i, (l, v) in enumerate(details):
            tk.Label(info_f, text=l, bg="#0d1117", fg="#888888", font=("Segoe UI", 10)).grid(row=i, column=0, sticky='w', pady=2)
            tk.Label(info_f, text=v, bg="#0d1117", fg="white", font=("Segoe UI", 10, "bold")).grid(row=i, column=1, sticky='w', padx=10, pady=2)

        res_f = tk.LabelFrame(dlg, text="  Calculation Summary  ", bg="#0d1117", fg="#4a9eff", font=("Segoe UI", 9, "bold"), padx=15, pady=15)
        res_f.pack(fill="both", expand=True, padx=20, pady=10)

        # Execution
        if not engine or not hasattr(engine, 'rebni_p'):
            tk.Label(res_f, text="Investigation Engine not initialized.", bg="#0d1117", fg="#f85149").pack()
        else:
            rows = engine.rebni_p.get((sid, po, asin), [])
            if not rows:
                tk.Label(res_f, text="No REBNI data found for this combination.", bg="#0d1117", fg="#f85149").pack()
            else:
                r0 = rows[0]
                rec_post = safe_num(r0.get('qty_received_postadj', 0))
                rec_unp = safe_num(r0.get('quantity_unpacked', 0))
                avail = safe_num(r0.get('rebni_available', 0))
                
                # Highlight logic
                tk.Label(res_f, text=f"Received (Post Adj):", bg="#0d1117", fg="#cccccc", font=("Segoe UI", 10)).grid(row=0, column=0, sticky='w')
                tk.Label(res_f, text=f"{int(rec_post)} units", bg="#0d1117", fg="#3fb950", font=("Segoe UI", 12, "bold")).grid(row=0, column=1, sticky='w', padx=10)
                
                tk.Label(res_f, text=f"Quantity Unpacked:", bg="#0d1117", fg="#cccccc", font=("Segoe UI", 10)).grid(row=1, column=0, sticky='w', pady=10)
                tk.Label(res_f, text=f"{int(rec_unp)} units", bg="#0d1117", fg="#58a6ff", font=("Segoe UI", 12, "bold")).grid(row=1, column=1, sticky='w', padx=10)
                
                tk.Label(res_f, text=f"Quantity Adjusted:", bg="#0d1117", fg="#cccccc", font=("Segoe UI", 10)).grid(row=2, column=0, sticky='w', pady=5)
                tk.Label(res_f, text=f"{int(safe_num(r0.get('quantity_adjusted', 0)))} units", bg="#0d1117", fg="#ff4d4d", font=("Segoe UI", 12, "bold")).grid(row=2, column=1, sticky='w', padx=10)
                
                tk.Label(res_f, text=f"REBNI Available:", bg="#0d1117", fg="#cccccc", font=("Segoe UI", 10)).grid(row=3, column=0, sticky='w')
                tk.Label(res_f, text=f"{int(avail)} units", bg="#0d1117", fg="#f0a500", font=("Segoe UI", 12, "bold")).grid(row=3, column=1, sticky='w', padx=10)
                
                tk.Label(res_f, text="\u24D8  Verified using REBNI strictly.", bg="#0d1117", fg="#6e7681", font=("Segoe UI", 8, "italic")).grid(row=4, column=0, columnspan=2, pady=(20,0))

        tk.Button(dlg, text="CLOSE", command=dlg.destroy, bg="#333", fg="white", font=("Segoe UI", 9, "bold"), relief="flat", padx=20, pady=8).pack(pady=15)
        
        dlg.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - dlg.winfo_width()) // 2
        py = self.winfo_y() + (self.winfo_height() - dlg.winfo_height()) // 2
        dlg.geometry(f"+{px}+{py}")

    def _show_on_hold_lookup(self):
        """v7.6: On Hold Qty lookup — with live filtering and fixed dark-mode headers."""
        # Auto-fill from selected row's main invoice (Inv no column)
        auto_inv = ""
        sel = self.tree.selection()
        if sel:
            r = self._row_data.get(sel[0], {})
            auto_inv = clean(r.get('Inv no', ''))

        dlg = tk.Toplevel(self)
        dlg.title("📌 On Hold Qty — Invoice Search Lookup")
        dlg.geometry("880x560")
        dlg.configure(bg="#0d1117")
        dlg.resizable(True, True)
        dlg.attributes("-topmost", True)

        tk.Label(dlg, text="📌  On Hold Quantity Lookup", bg="#16213e", fg="#ff6b6b",
                 font=("Segoe UI", 11, "bold"), height=2).pack(fill="x")

        # v7.6: Fix header style for dark mode
        s = ttk.Style()
        s.configure("OnHold.Treeview", font=("Calibri", 10), rowheight=24,
                     background="#1e1e3a", foreground="#e0e0e0", fieldbackground="#1e1e3a")
        s.configure("OnHold.Treeview.Heading", font=("Segoe UI", 9, "bold"),
                     background="#203864", foreground="white")
        s.map("OnHold.Treeview.Heading", background=[('active', '#2a4b86')])

        # Input frame
        inp_f = tk.Frame(dlg, bg="#0d1117", padx=20, pady=10)
        inp_f.pack(fill="x")
        tk.Label(inp_f, text="Invoice Number:", bg="#0d1117", fg="#cccccc",
                 font=("Segoe UI", 10)).grid(row=0, column=0, sticky='e', padx=(0, 8))
        inv_var = tk.StringVar(value=auto_inv)
        inv_ent = tk.Entry(inp_f, textvariable=inv_var, font=("Segoe UI", 10),
                           bg="#21262d", fg="white", insertbackground="white", width=24)
        inv_ent.grid(row=0, column=1, sticky='w')
        tk.Label(inp_f, text="(auto-filled from selected row — or type any invoice number)",
                 bg="#0d1117", fg="#6e7681", font=("Segoe UI", 8, "italic")).grid(row=1, column=1, sticky='w', pady=(2, 0))

        # v7.6: Live Filter bar
        tk.Label(inp_f, text="Filter:", bg="#0d1117", fg="#ffaa44",
                 font=("Segoe UI", 10, "bold")).grid(row=0, column=3, sticky='e', padx=(20, 5))
        filter_var = tk.StringVar()
        filter_ent = tk.Entry(inp_f, textvariable=filter_var, font=("Segoe UI", 10),
                              bg="#21262d", fg="#ffaa44", insertbackground="#ffaa44", width=20)
        filter_ent.grid(row=0, column=4, sticky='w')
        tk.Label(inp_f, text="(type status, ASIN, PO — filters instantly)",
                 bg="#0d1117", fg="#6e7681", font=("Segoe UI", 8, "italic")).grid(row=1, column=4, sticky='w', pady=(2, 0))

        # Results frame
        res_f = tk.Frame(dlg, bg="#0d1117")
        res_f.pack(fill="both", expand=True, padx=16, pady=8)

        cols = ["Invoice No", "ASIN", "PO", "Status", "Qty Invoiced", "Qty Matched"]
        col_widths = [160, 130, 120, 160, 100, 100]
        tv = ttk.Treeview(res_f, columns=cols, show='headings', height=14, style="OnHold.Treeview")
        for c, w in zip(cols, col_widths):
            tv.heading(c, text=c); tv.column(c, width=w, minwidth=40, anchor='w')
        vsb = ttk.Scrollbar(res_f, orient="vertical", command=tv.yview)
        hsb = ttk.Scrollbar(res_f, orient="horizontal", command=tv.xview)
        tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tv.grid(row=0, column=0, sticky="nsew"); vsb.grid(row=0, column=1, sticky="ns"); hsb.grid(row=1, column=0, sticky="ew")
        res_f.grid_rowconfigure(0, weight=1); res_f.grid_columnconfigure(0, weight=1)
        tv.tag_configure('onhold', background='#3a1a00', foreground='#ffaa44')
        tv.tag_configure('normal', background='#1e1e3a', foreground='#e0e0e0')

        status_lbl = tk.Label(dlg, text="", bg="#0d1117", fg="#6e7681", font=("Segoe UI", 9, "italic"))
        status_lbl.pack(pady=(2, 0))

        # v7.6: In-memory store for looked-up rows (for instant filtering)
        _all_results = []

        def _apply_filter(*args):
            """Instantly filter the displayed rows based on filter_var text."""
            ftext = filter_var.get().strip().upper()
            for item in tv.get_children(): tv.delete(item)
            shown = 0
            for vals, tag in _all_results:
                if not ftext or any(ftext in str(v).upper() for v in vals):
                    tv.insert('', 'end', values=vals, tags=(tag,))
                    shown += 1
            total = len(_all_results)
            if ftext:
                status_lbl.config(text=f"Showing {shown} of {total} row(s)  |  Filter: '{filter_var.get().strip()}'")
            else:
                status_lbl.config(text=f"✔ Showing all {total} row(s)")

        def do_lookup(*args):
            inv_no = clean(inv_var.get())
            for item in tv.get_children(): tv.delete(item)
            _all_results.clear()
            filter_var.set("")
            if not inv_no:
                status_lbl.config(text="Enter an invoice number to search."); return

            app = getattr(self, '_app', None)
            inv_df = getattr(app, 'inv_search_df', None)
            if inv_df is None:
                status_lbl.config(text="⚠ Invoice Search data not loaded yet. Run investigation first."); return

            matches = inv_df[inv_df['invoice_number'].astype(str).str.strip().str.upper() == inv_no.upper()]
            if matches.empty:
                status_lbl.config(text=f"No results found for invoice: {inv_no}"); return

            for _, row in matches.iterrows():
                status_raw = str(row.get('invoice_item_status', '')).strip()
                qty_inv = safe_num(row.get('quantity_invoiced', 0))
                qty_mtc = safe_num(row.get('quantity_matched_total', 0))
                tag = 'onhold' if 'hold' in status_raw.lower() else 'normal'
                vals = [
                    clean(str(row.get('invoice_number', ''))),
                    clean(str(row.get('asin', ''))),
                    clean(str(row.get('purchase_order_id', ''))),
                    status_raw,
                    int(qty_inv) if qty_inv == int(qty_inv) else qty_inv,
                    int(qty_mtc) if qty_mtc == int(qty_mtc) else qty_mtc
                ]
                _all_results.append((vals, tag))
                tv.insert('', 'end', values=vals, tags=(tag,))
            status_lbl.config(text=f"✔ Found {len(_all_results)} row(s) for invoice {inv_no}")

        tk.Button(inp_f, text="🔍 LOOKUP", command=do_lookup,
                  bg="#8b0000", fg="#ffcccc", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=12, pady=4, cursor="hand2").grid(row=0, column=2, padx=(10, 0))
        inv_ent.bind('<Return>', do_lookup)
        filter_var.trace_add('write', _apply_filter)

        tk.Button(dlg, text="CLOSE", command=dlg.destroy, bg="#333", fg="white",
                  font=("Segoe UI", 9, "bold"), relief="flat", padx=20, pady=6).pack(pady=8)

        # Auto-trigger lookup if invoice was auto-filled
        if auto_inv:
            dlg.after(100, do_lookup)

        dlg.update_idletasks()
        px = self.winfo_x() + (self.winfo_width() - dlg.winfo_width()) // 2
        py = self.winfo_y() + (self.winfo_height() - dlg.winfo_height()) // 2
        dlg.geometry(f"+{px}+{py}")

    def add_header_row(self, label=""):
        vals = list(self.COLS); vals[4] = f"── {label} ──" if label else "Header"
        iid = self.tree.insert('', 'end', values=vals, tags=('header',))
        # v5.9.3: Explicitly tag this as a UI header so it is ignored in Excel reports
        self._row_data[iid] = dict(zip(self.COLS, vals))
        self._row_data[iid]['is_ui_header'] = True
        self.tree.see(iid)

    def add_row(self, rd):
        try:
            if not self.winfo_exists(): return
        except: return
        vals = [rd.get('barcode',''), rd.get('invoice',''), rd.get('sid',''), rd.get('po',''), rd.get('asin',''), 
                rd.get('inv_qty',''), rd.get('rec_qty',''), rd.get('mtc_qty',''), rd.get('mtc_inv',''),
                rd.get('mtc_asin',''), rd.get('mtc_po',''), rd.get('fc_id',''), rd.get('remarks',''), rd.get('date',''), rd.get('cp_status','')]
        iq, rq = safe_num(rd.get('inv_qty')), safe_num(rd.get('rec_qty'))
        tag = 'shortage_red' if (iq > 0 and rq < iq) else 'crosspo' if 'cross po' in str(rd.get('remarks','')).lower() else 'invalid_ibc_pbc' if rd.get('is_invalid_ibc_pbc') else ''
        try:
            iid = self.tree.insert('', 'end', values=vals, tags=(tag,))
            self._row_data[iid] = dict(zip(self.COLS, vals)); self._row_data[iid]['_rd'] = rd
            # v7.2.0: Store Loop Key for Sync/Undo
            self._row_data[iid]['_loop_key'] = self._app._get_loop_key(rd) if self._app and hasattr(self._app, '_get_loop_key') else None
            self.tree.see(iid)

        except Exception: pass  # Treeview destroyed

    def get_all_rows(self):
        KEY = {'Barcode':'barcode', 'Inv no':'invoice', 'SID':'sid', 'PO':'po', 'ASIN':'asin', 'Inv Qty':'inv_qty', 'Rec Qty':'rec_qty', 'Mtc Qty':'mtc_qty', 'Mtc Inv':'mtc_inv', 'Mtc ASIN':'mtc_asin', 'Mtc PO':'mtc_po', 'FC':'fc_id', 'Remarks':'remarks', 'Date':'date', 'CP':'cp_status'}
        rows = []
        for iid in self.tree.get_children():
            d = self._row_data.get(iid, {}); rd = d.get('_rd', {}).copy()
            for col in self.COLS: rd[KEY[col]] = d.get(col, '')
            if d.get('is_ui_header'): rd['is_ui_header'] = True
            rows.append(rd)
        return rows

    def clear_all(self):
        for iid in self.tree.get_children(): self.tree.delete(iid)
        self._row_data.clear()

    def _on_double_click(self, event):
        iid = self.tree.identify_row(event.y); col = self.tree.identify_column(event.x)
        if not iid or not col: return
        col_idx = int(col.replace('#',''))-1; col_name = self.COLS[col_idx]
        
        # v7.1.3: One-Click Copy — copy the cell value to clipboard
        val = str(self.tree.set(iid, col_name)).strip()
        try:
            self.clipboard_clear()
            self.clipboard_append(val)
        except: pass # fallback for headless/remote
        
        bbox=self.tree.bbox(iid, col)
        if not bbox: return
        self._on_double_right_click(event)
        x,y,w,h = bbox
        cur = self._row_data.get(iid,{}).get(col_name,''); ev=tk.StringVar(value=str(cur))
        e = tk.Entry(self.tree, textvariable=ev); e.place(x=x,y=y,width=w,height=h); e.focus_force()

        def save(evt=None):
            nv=ev.get(); self._row_data[iid][col_name]=nv
            vals=list(self.tree.item(iid,'values')); vals[col_idx]=nv
            self.tree.item(iid, values=vals); e.destroy()
        e.bind('<Return>', save); e.bind('<FocusOut>', save)


    def _take_snapshot(self):
        """v6.0.0: Capture a full snapshot of current tree state for undo."""
        self._redo_stack.clear() # clear redo stack on new action
        snap = self._capture_state()
        self._undo_stack.append(snap)
        if len(self._undo_stack) > 10:
            self._undo_stack.pop(0)  # cap at 10

    def _capture_state(self):
        snap = []
        for iid in self.tree.get_children():
            vals = list(self.tree.item(iid, 'values'))
            tags = list(self.tree.item(iid, 'tags'))
            data_copy = {}
            if iid in self._row_data:
                data_copy = {k: v for k, v in self._row_data[iid].items() if k != '_rd'}
                if '_rd' in self._row_data[iid]:
                    data_copy['_rd'] = dict(self._row_data[iid]['_rd'])
            snap.append({'vals': vals, 'tags': tags, 'data': data_copy})
        return snap

    def _restore_state(self, snap):
        for iid in self.tree.get_children():
            self.tree.delete(iid)
        self._row_data.clear()
        for entry in snap:
            iid = self.tree.insert('', 'end', values=entry['vals'], tags=tuple(entry['tags']))
            self._row_data[iid] = entry['data']

    def _undo(self):
        """v6.0.0: Restore the last snapshot from the undo stack."""
        if not self._undo_stack:
            messagebox.showinfo("Undo", "Nothing to undo.", parent=self)
            return
        self._redo_stack.append(self._capture_state())
        snap = self._undo_stack.pop()
        self._restore_state(snap)
        messagebox.showinfo("Undo", f"Restored to previous state. ({len(self._undo_stack)} undo steps remaining)", parent=self)

    def _redo(self):
        """v6.0.0: Redo an undone action."""
        if not self._redo_stack:
            messagebox.showinfo("Redo", "Nothing to redo.", parent=self)
            return
        self._undo_stack.append(self._capture_state())
        snap = self._redo_stack.pop()
        self._restore_state(snap)
        messagebox.showinfo("Redo", f"Redo successful. ({len(self._redo_stack)} redo steps remaining)", parent=self)

    def _save_preview(self):
        """v6.0.0: Direct save from preview panel."""
        if not getattr(self._app, "ticket_id", None): return
        t, ts = self._app.ticket_id.get().strip().replace(' ','_'), datetime.now().strftime('%Y%m%d_%H%M%S')
        o = f"MFI_{t}_{ts}.xlsx" if t else f"MFI_Investigation_{ts}.xlsx"
        
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=o, parent=self)
        if not path: return
        
        try:
            ar = self.get_all_rows()
            bl, cur = [], []
            for r in ar:
                if r.get('is_ui_header'):
                    if cur: bl.append(cur)
                    cur = []
                    continue
                if str(r.get('barcode')).strip() == "Barcode": continue
                cur.append(r)
            if cur: bl.append(cur)
            
            app = self._app
            rebni_data = app.engine.collected_rebni if (app and hasattr(app, 'engine')) else None
            sids = app.engine.unique_sids_found if (app and hasattr(app, 'engine')) else None
            vraw = app.engine.vendor_raw if (app and hasattr(app, 'engine')) else None
            vhdr = app.engine.vendor_headers if (app and hasattr(app, 'engine')) else None
            c_asin = getattr(app.engine, 'ciat_asin_df', None) if (app and hasattr(app, 'engine')) else None
            c_ship = getattr(app.engine, 'ciat_shipment_df', None) if (app and hasattr(app, 'engine')) else None
            write_excel(bl, path, rebni_summary_data=rebni_data, unique_sids_found=sids, vendor_raw=vraw, vendor_headers=vhdr, ciat_asin_df=c_asin, ciat_shipment_df=c_ship)
            messagebox.showinfo("Saved", f"Complete investigation report saved to:\n{path}", parent=self)
        except Exception as e:
            messagebox.showerror("Save Error", str(e), parent=self)

    # ═══════════════════════════════════════════════════════════
    #  FIND / SEARCH FEATURE
    # ═══════════════════════════════════════════════════════════
    def _show_find_dialog(self):
        """v7.3.6: Launches the non-modal Find dialog for grid searching."""
        if hasattr(self, '_find_dlg') and self._find_dlg.winfo_exists():
            self._find_dlg.lift(); return
            
        self._find_dlg = tk.Toplevel(self)
        self._find_dlg.title("Find in Preview")
        self._find_dlg.geometry("380x150")
        self._find_dlg.resizable(False, False)
        self._find_dlg.attributes("-topmost", True)
        self._find_dlg.configure(bg="#1a1a2e")
        
        tk.Label(self._find_dlg, text="Search for Invoice, PO, ASIN or Remark:", fg="#e0e0e0", bg="#1a1a2e", font=("Segoe UI", 9)).pack(pady=(15, 5))
        ent = tk.Entry(self._find_dlg, width=40, bg="#16213e", fg="white", insertbackground="white", font=("Segoe UI", 10))
        ent.pack(pady=5); ent.focus_set()
        
        btn_f = tk.Frame(self._find_dlg, bg="#1a1a2e")
        btn_f.pack(pady=10)
        
        status_lbl = tk.Label(self._find_dlg, text="", fg="#00d2ff", bg="#1a1a2e", font=("Segoe UI", 9))
        status_lbl.pack()

        def do_find(direction=1):
            query = ent.get().strip().lower()
            if not query: return
            
            # If query changed, perform new global scan for (item, col_idx) pairs
            if not hasattr(self, '_last_query') or self._last_query != query:
                self._last_query = query
                self._search_matches = []
                all_cols = list(self.tree['columns'])
                for item in self.tree.get_children(''):
                    vals = self.tree.item(item, 'values')
                    for ci, val in enumerate(vals):
                        if query in str(val).lower():
                            self._search_matches.append((item, ci))
                self._search_idx = -1
            
            if not self._search_matches:
                status_lbl.config(text="No matches found.", fg="#ff4d4d"); return
                
            self._search_idx = (self._search_idx + direction) % len(self._search_matches)
            target_item, col_idx = self._search_matches[self._search_idx]
            col_id = self.tree['columns'][col_idx]
            
            # 1. Select and scroll to row
            self.tree.selection_set(target_item)
            self.tree.see(target_item)
            self.tree.focus(target_item)
            
            # 2. Draw temporary Cell Focus Box
            try:
                bbox = self.tree.bbox(target_item, col_id)
                if bbox:
                    x, y, w, h = bbox
                    # Create a "Highlighter" frame
                    hl = tk.Frame(self.tree, bg="#00d2ff", highlightthickness=2, highlightbackground="yellow")
                    hl.place(x=x, y=y, width=w, height=h)
                    # Flash effect: blink and then destroy
                    def blink(count):
                        if count <= 0: hl.destroy(); return
                        hl.place_forget() if count % 2 == 0 else hl.place(x=x, y=y, width=w, height=h)
                        self.after(200, lambda: blink(count-1))
                    blink(6) # 3 blinks over 1.2 seconds
            except: pass
            
            status_lbl.config(text=f"Match {self._search_idx + 1} of {len(self._search_matches)} (Col: {col_id})", fg="#00d2ff")

        tk.Button(btn_f, text="  Find Previous  ", command=lambda: do_find(-1), bg="#16213e", fg="white").pack(side="left", padx=5)
        tk.Button(btn_f, text="    Find Next    ", command=lambda: do_find(1), bg="#203864", fg="white", font=("Segoe UI", 9, "bold")).pack(side="left", padx=5)
        
        ent.bind("<Return>", lambda e: do_find(1))
        ent.bind("<Escape>", lambda e: self._find_dlg.destroy())

    def _delete_selected_rows(self):
        """v6.0.0: Delete selected rows from the preview panel."""
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Delete", "No rows selected. Click a row first.", parent=self)
            return
        self._take_snapshot()  # allow undo
        for iid in sel:
            self._row_data.pop(iid, None)
            self.tree.delete(iid)

    def _on_double_right_click(self, event):
        """v7.2.0: Immediate Copy on Double Right-Click."""
        iid = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if iid and col:
            self.tree.selection_set(iid)
            self._ctx_iid = iid
            self._ctx_col_idx = int(col.replace('#','')) - 1
            self._ctx_copy()

    def _show_context_menu(self, event):
        """v6.0.0: Show right-click context menu on the Treeview."""
        iid = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if iid:
            self.tree.selection_set(iid)
            self._ctx_iid = iid
            self._ctx_col_idx = int(col.replace('#', '')) - 1 if col else None
        try:
            self._ctx_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self._ctx_menu.grab_release()

    def _ctx_copy(self):
        """v6.0.0: Copy clicked cell value to clipboard."""
        if self._ctx_iid is None or self._ctx_col_idx is None: return
        col_name = self.COLS[self._ctx_col_idx]
        val = str(self._row_data.get(self._ctx_iid, {}).get(col_name, ''))
        self.clipboard_clear()
        self.clipboard_append(val)

    def _ctx_paste(self):
        """v6.0.0: Paste clipboard content into clicked cell."""
        if self._ctx_iid is None or self._ctx_col_idx is None: return
        try:
            val = self.clipboard_get()
        except tk.TclError:
            return
        self._take_snapshot()
        col_name = self.COLS[self._ctx_col_idx]
        sel = self.tree.selection()
        if self._ctx_iid not in sel: sel = (self._ctx_iid,) # if they didn't highlight multiple, just do the clicked one
        
        for iid in sel:
            if iid in self._row_data:
                self._row_data[iid][col_name] = val
                vals = list(self.tree.item(iid, 'values'))
                vals[self._ctx_col_idx] = val
                self.tree.item(iid, values=vals)

    def _ctx_clear(self):
        """v6.2.5: Clear the targeted column across all selected rows."""
        if self._ctx_iid is None or self._ctx_col_idx is None: return
        self._take_snapshot()
        col_name = self.COLS[self._ctx_col_idx]
        sel = self.tree.selection()
        if self._ctx_iid not in sel: sel = (self._ctx_iid,)
        
        for iid in sel:
            if iid in self._row_data:
                self._row_data[iid][col_name] = ""
                vals = list(self.tree.item(iid, 'values'))
                vals[self._ctx_col_idx] = ""
                self.tree.item(iid, values=vals)

    def _ctx_insert_row(self):
        """v6.2.5: Insert an entirely blank row immediately below the selected row."""
        if self._ctx_iid is None: return
        self._take_snapshot()
        idx = self.tree.index(self._ctx_iid)
        
        vals = [''] * len(self.COLS)
        tags = self.tree.item(self._ctx_iid, 'tags')
        
        new_iid = self.tree.insert('', idx + 1, values=vals, tags=tags)
        self._row_data[new_iid] = dict(zip(self.COLS, vals))
        self._row_data[new_iid]['_rd'] = {'_blank': True} # Safe fallback flag

    def confirm_edits(self):
        self._take_snapshot()  # v6.0.0: snapshot before applying edits
        app = self._app
        if not app or not hasattr(app, 'engine'): return

        # v7.2.0: Synchronize Deletions (Undo Logic)
        # Identify which loop keys are still present in the preview panel
        remaining_loop_keys = set()
        for iid in self.tree.get_children():
            lk = self._row_data.get(iid, {}).get('_loop_key')
            if lk: remaining_loop_keys.add(lk)

        # Remove keys from 'processed' that are no longer in the preview (effectively "undoing" them)
        if hasattr(app, 'curr_m'):
            processed = app.curr_m.get('processed', set())
            to_remove = [lk for lk in processed if lk not in remaining_loop_keys]
            for lk in to_remove:
                processed.discard(lk)
                app.global_processed.discard(lk)
            if to_remove:
                self._app._set_status(f"Undo synced: {len(to_remove)} entries returned to Pending Gateway.")

        
        KEY_MAP = {'Barcode':'barcode', 'Inv no':'invoice', 'SID':'sid', 'PO':'po', 'ASIN':'asin', 'Inv Qty':'inv_qty', 'Rec Qty':'rec_qty', 'Mtc Qty':'mtc_qty', 'Mtc Inv':'mtc_inv', 'Mtc ASIN':'mtc_asin', 'Mtc PO':'mtc_po', 'FC':'fc_id', 'Remarks':'remarks', 'Date':'date', 'CP':'cp_status'}

        for iid in self.tree.get_children():
            d = self._row_data.get(iid, {}); rd = d.get('_rd', {})
            if not rd or d.get('is_ui_header'): continue
            
            mtc_inv = str(d.get('Mtc Inv', '') or rd.get('mtc_inv', '')).strip()
            mtc_po  = str(d.get('Mtc PO', '') or rd.get('mtc_po', '')).strip()
            mtc_asin= str(d.get('Mtc ASIN', '') or rd.get('mtc_asin', '')).strip()
            
            iq = safe_num(d.get('Inv Qty', '')); rq = safe_num(d.get('Rec Qty', ''))
            mq = safe_num(d.get('Mtc Qty', ''))
            
            # v6.2.5: Auto-recalculate Remarks if Inv Qty and Rec Qty create a math discrepancy
            if iq and rq and iq != rq:
                old_rem = str(d.get('Remarks', ''))
                # Only overwrite standard calculation remarks if they didn't explicitly write a custom note
                if 'short' in old_rem.lower() or 'overage' in old_rem.lower() or 'units matched' in old_rem.lower() or not old_rem.strip():
                    diff = abs(iq - rq)
                    kw = "short" if iq > rq else "overage"
                    new_rem = f"Found {int(diff)} units {kw} locally (Inv: {int(iq)}, Rec: {int(rq)})"
                    
                    if old_rem != new_rem:
                        d['Remarks'] = new_rem
                        try:
                            idx = self.COLS.index('Remarks')
                            vals = list(self.tree.item(iid, 'values'))
                            vals[idx] = new_rem
                            self.tree.item(iid, values=vals)
                        except: pass
            
            if mtc_inv: 
                k = (clean(mtc_inv), clean(mtc_asin), clean(mtc_po))
                
                # Expand overrides to support ANY cell edit syncing globally
                ext_out = {}
                for col_name in self.COLS:
                    if col_name in d: ext_out[KEY_MAP[col_name]] = d[col_name]
                
                # Make sure strict numerics remain precise
                ext_out['inv_qty'] = iq
                ext_out['mtc_qty'] = mq
                ext_out['rec_qty'] = rq
                
                app.engine.user_overrides[k] = ext_out
        
        # v6.2.5: Synchronize edits across ALL active dialogs globally
        # v7.3.0 Bug Fix: Build an exact mapping from OLD keys to NEW data, handling newly inserted rows explicitly.
        updated_map = {}
        new_adds = []
        for iid in self.tree.get_children():
            d = self._row_data.get(iid, {}); rd = d.get('_rd', {})
            if not rd or d.get('is_ui_header'): continue
            
            n_inv = str(d.get('Mtc Inv', '') or rd.get('mtc_inv', '')).strip()
            n_po  = str(d.get('Mtc PO', '') or rd.get('mtc_po', '')).strip()
            n_asin= str(d.get('Mtc ASIN', '') or rd.get('mtc_asin', '')).strip()
            n_qty = safe_num(d.get('Mtc Qty', ''))
            n_iqty= safe_num(d.get('Inv Qty', ''))
            
            if not n_inv: continue
            
            new_dict = {'mtc_inv': n_inv, 'mtc_po': n_po, 'mtc_asin': n_asin, 'mtc_qty': n_qty, 'inv_qty': n_iqty}
            if rd.get('_blank'):
                new_adds.append(new_dict)
            else:
                old_k = (clean(rd.get('mtc_inv','')), clean(rd.get('mtc_asin','')), clean(rd.get('mtc_po','')))
                updated_map[old_k] = new_dict

        updated_any = False
        if getattr(app, 'active_manual_dlg', None) and app.active_manual_dlg.winfo_exists():
            app.active_manual_dlg.sync_with_edits(updated_map, new_adds)
            updated_any = True
            
        if getattr(app, 'active_pending_dlg', None) and app.active_pending_dlg.winfo_exists():
            app.active_pending_dlg.sync_with_edits(updated_map, new_adds)
            updated_any = True

        if updated_any:
            messagebox.showinfo("Success", "Manual edits synchronized across all active panels.", parent=self)
        else:
            messagebox.showinfo("Success", "Manual edits saved to engine cache.", parent=self)


# ═══════════════════════════════════════════════════════════
#  DATA LOADERS
# ═══════════════════════════════════════════════════════════

def _load_file(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == '.csv':
        try:   return pd.read_csv(path, dtype=str, encoding='utf-8')
        except: return pd.read_csv(path, dtype=str, encoding='latin-1')
    else:
        try:
            return pd.read_excel(path, header=0, dtype=str, engine='calamine')
        except Exception:
            return pd.read_excel(path, header=0, dtype=str)

def load_claims(path): return _load_file(path)

COLUMN_ALIASES = {
    'Barcode': [
        'barcode', 'bar code', 'bar_code', 'upc', 'ean',
        'scan code', 'item code', 'carton barcode', 'pkg barcode',
    ],
    'Invoice': [
        'inv no', 'inv_no', 'invoice_no', 'invoice no', 'invoice number',
        'invoice_number', 'invoice', 'inv num', 'inv number',
        'inv#', 'invoice#', 'invc no', 'invc number', 'invc',
        'bill no', 'bill number',
    ],
    'SID': [
        'sid', 'shipment id', 'shipment_id', 'shipment no',
        'shipment number', 'ship id', 'fba shipment id',
        'inbound shipment', 'shipment', 'inbound sid',
    ],
    'PO': [
        'po', 'po_no', 'p.o.', 'po no', 'po number', 'po#',
        'purchase order', 'purchase_order', 'purchase order no',
        'purchase order number', 'po id', 'order no', 'order number',
        'purch order', 'header po', 'line po', 'po num',
        'invoice_lineitem_po',
    ],
    'ASIN': [
        'asin', 'po_asin', 'amazon asin', 'amazon product id',
        'product id', 'asin no', 'item asin', 'amazon id',
    ],
    'InvQty': [
        'inv qty', 'inv_qty', 'invoice qty', 'invoice quantity',
        'invoiced qty', 'invoiced quantity', 'quantity invoiced',
        'quantity_invoiced', 'billed qty', 'billed quantity',
        'total qty', 'total quantity', 'item qty',
        'po_ordered_quantity',
    ],
    'PQV': [
        'pqv', 'pqv qty', 'missing qty', 'missing quantity',
        'missing_qty', 'shortage', 'short qty', 'short quantity',
        'claim qty', 'claimed qty', 'dispute qty',
        'pending qty', 'outstanding qty', 'difference qty',
    ],
    'CP': [
        'cp', 'cost price', 'cost_price', 'unit cost', 'unit_cost',
        'item cost', 'item_cost', 'rate', 'amount', 'unit price',
    ],
}

def detect_claim_cols(df):
    actual_cols = list(df.columns)
    lower_map   = {c.lower().strip(): c for c in actual_cols}
    mapping     = {}
    corrections = []

    for field, aliases in COLUMN_ALIASES.items():
        found = None
        for alias in aliases:
            if alias in lower_map:
                found = lower_map[alias]; break
        if not found:
            for alias in aliases:
                for col in actual_cols:
                    if alias in col.lower() or col.lower() in alias:
                        found = col; break
                if found: break
        if found:
            canonical = aliases[0]
            if found.lower().strip() != canonical:
                corrections.append((field, found, canonical))
            mapping[field] = found
    return mapping, corrections

def load_vendor_data(path):
    df = _load_file(path)
    return df

# ─────────────────────────────────────────────────────────
#  INDEX BUILDERS
# ─────────────────────────────────────────────────────────

def build_vendor_index(df):
    """v7.0: Builds Vendor Level Data index. Groups IBC by PO."""
    idx = {}
    raw_by_sid = {}
    headers = df.columns.tolist()
    
    def get_col(*matches):
        for c in headers:
            cl = str(c).lower().strip()
            if any(m.lower() in cl for m in matches): return c
        return None

    c_inv = get_col('invoice number', 'invoice')
    c_sid = get_col('shipment id', 'shipment', 'sid')
    c_po  = get_col('po number', 'purchase order', 'po')
    c_bc  = get_col('barcode', 'fnsku')
    c_pbc = get_col('pod box quantity', 'pbc', 'physical box')
    c_ibc = get_col('invoice box count', 'ibc', 'invoice box')
    c_iq  = get_col('total invoiced quantity', 'invoiced qty', 'inv qty')
    c_rc  = get_col('number of cartons rejected', 'carton rejected', 'rejected carton', 'rejected')
    c_rr  = get_col('pod reject reason', 'reject reason', 'rejection reason')

    # Aggregation: Group by PO to calculate total IBC across all invoices for that PO
    po_totals = {}
    for r in df.to_dict('records'):
        po = clean(r.get(c_po, ''))
        if po:
            ibc = safe_num(r.get(c_ibc, 0))
            if po not in po_totals:
                po_totals[po] = {'total_ibc': 0, 'pbc': safe_num(r.get(c_pbc, 0)), 'rej_ctn': safe_num(r.get(c_rc, 0)), 'rej_rsn': clean(r.get(c_rr, ''))}
            po_totals[po]['total_ibc'] += ibc

    # Build primary index keyed by Invoice Number
    for r in df.to_dict('records'):
        inv = clean(r.get(c_inv, ''))
        sid_frag = extract_sid(clean(r.get(c_sid, ''))) if c_sid else ""
        
        # v7.0: Store raw row data for Validation Sheet extraction
        if sid_frag:
            if sid_frag not in raw_by_sid: raw_by_sid[sid_frag] = []
            raw_by_sid[sid_frag].append([r.get(h, '') for h in headers])

        if not inv: continue
        po = clean(r.get(c_po, ''))
        
        po_data = po_totals.get(po, {})
        idx[inv] = {
            'sid': sid_frag,
            'bc': clean(r.get(c_bc, '')),
            'po': po,
            'inv_qty': safe_num(r.get(c_iq, 0)),
            'ibc': po_data.get('total_ibc', safe_num(r.get(c_ibc, 0))),
            'pbc': po_data.get('pbc', safe_num(r.get(c_pbc, 0))),
            'rej_ctn': po_data.get('rej_ctn', safe_num(r.get(c_rc, 0))),
            'rej_rsn': po_data.get('rej_rsn', clean(r.get(c_rr, '')))
        }
    return idx, raw_by_sid, headers

def load_rebni(path):
    df = _load_file(path)
    names = ['vendor_code', 'po', 'asin', 'shipment_id', 'received_datetime',
             'warehouse_id', 'item_cost', 'quantity_unpacked', 'quantity_adjusted',
             'qty_received_postadj', 'quantity_matched', 'rebni_available',
             'cnt_invoice_matched', 'matched_invoice_numbers']
    if len(df.columns) > len(names):
        names += [f'ext_{i}' for i in range(len(df.columns) - len(names))]
    df.columns = names[:len(df.columns)]
    return df

def load_invoice_search(path):
    df = _load_file(path)
    names = ['vendor_code', 'purchase_order_id', 'asin', 'invoice_number', 'invoice_date',
             'invoice_item_status', 'quantity_invoiced', 'quantity_matched_total',
             'no_of_shipments', 'shipment_id', 'shipmentwise_matched_qty',
             'matched_po', 'matched_asin']
    df.columns = names[:len(df.columns)]
    # v6.2.5 Deduplication: Strictly remove redundant invoice-asin-po-sid records
    # FIX 7: Keep the row with max quantity_matched_total when deduplicating.
    # Simply dropping duplicates could discard legitimate rows with different quantities.
    if 'quantity_matched_total' in df.columns:
        df = (df.sort_values('quantity_matched_total', ascending=False)
                .drop_duplicates(subset=['purchase_order_id', 'asin', 'invoice_number', 'shipment_id'])
                .reset_index(drop=True))
    else:
        df = df.drop_duplicates(subset=['purchase_order_id', 'asin', 'invoice_number', 'shipment_id']).reset_index(drop=True)
    return df

def load_shipment_master(path):
    df = _load_file(path)
    return df

# ═══════════════════════════════════════════════════════════
#  INDEX BUILDERS
# ═══════════════════════════════════════════════════════════

def build_shipment_index(df):
    """v6.2.5: Builds an index from (SID, PO) to (Barcode, Invoice) for auto-population."""
    idx = {}
    mapping, _ = detect_claim_cols(df)
    sid_col = mapping.get('SID')
    bc_col  = mapping.get('Barcode')
    po_col  = mapping.get('PO')
    inv_col = mapping.get('Invoice')
    
    for row in df.to_dict('records'):
        sid = extract_sid(clean(row.get(sid_col, ''))) if sid_col else ""
        po  = clean(row.get(po_col, '')) if po_col else ""
        bc  = clean(row.get(bc_col, '')) if bc_col else ""
        inv = clean(row.get(inv_col, '')) if inv_col else ""
        
        if sid and po:
            idx[(sid, po)] = {'bc': bc, 'inv': inv}
        if sid and bc and sid not in idx:
            idx[sid] = bc # Legacy SID -> Barcode mapping
    return idx

def build_rebni_index(df):
    p, s, fb, sid_p = {}, {}, {}, {}
    for row in df.to_dict('records'):
        sid  = extract_sid(clean(row.get('shipment_id', '')))
        po   = clean(row.get('po', ''))
        asin = clean(row.get('asin', ''))
        if not sid or not asin: continue
        p.setdefault((sid, po, asin), []).append(row)
        s.setdefault((po, asin), []).append(row)
        sid_p.setdefault(sid, []).append(row)
        for inv in split_comma(row.get('matched_invoice_numbers', '')):
            if inv: fb.setdefault((sid, po, inv), []).append(row)
    return p, s, fb, sid_p

def build_invoice_index(df):
    idx, fb, iam = {}, {}, {}
    for row in df.to_dict('records'):
        sids  = split_comma(row.get('shipment_id', ''))
        pos   = split_comma(row.get('matched_po', ''))
        asins = split_comma(row.get('matched_asin', ''))
        qtys  = split_comma(row.get('shipmentwise_matched_qty', ''))
        
        inv_no   = clean(row.get('invoice_number', ''))
        mtc_asin = clean(row.get('asin', ''))
        inv_qty  = safe_num(row.get('quantity_invoiced', '0'))
        
        # v6.1.1 Fix: Save exact target ASIN quantity; do NOT cross-pollute matching ASINs.
        # v7.1.0: Only store/overwrite if quantity is positive to avoid adjustment rows (0 qty) wiping out real data.
        if inv_no and mtc_asin and inv_qty > 0:
            key = (inv_no, mtc_asin.upper())
            if inv_qty > iam.get(key, 0):
                iam[key] = inv_qty

        for i in range(max(len(sids), len(pos), len(asins), len(qtys))):
            s_frag   = extract_sid(sids[i] if i < len(sids) else "")
            p_val    = clean(pos[i]   if i < len(pos)   else "")
            a_val    = clean(asins[i] if i < len(asins) else "")
            q_val    = safe_num(qtys[i] if i < len(qtys) else "0")
            
            mtc_po   = clean(row.get('purchase_order_id', ''))
            if not s_frag or not p_val or not a_val: continue
            entry = {'mtc_inv':  inv_no,
                     'mtc_po':   mtc_po,
                     'mtc_asin': mtc_asin,
                     'inv_qty':  inv_qty,
                     'mtc_qty':  q_val,
                     'date':     clean(row.get('invoice_date', ''))}
            idx.setdefault((s_frag, p_val, a_val), []).append(entry)
            if inv_no: fb.setdefault((s_frag, p_val, inv_no), []).append(entry)
    return idx, fb, iam


# ═══════════════════════════════════════════════════════════
#  INVESTIGATION ENGINE
# ═══════════════════════════════════════════════════════════

class InvestigationEngine:
    MAX_DEPTH = 10

    def __init__(self, rp, rs, rfb, rsid, ip, ifb, iam, sid_cb=None, vendor_idx=None, vendor_raw=None, vendor_headers=None):
        self.rebni_p    = rp
        self.rebni_s    = rs
        self.rebni_fb   = rfb
        self.rebni_sid  = rsid # v5.9.3: Shipment ID based REBNI index
        self.inv_p      = ip
        self.inv_fb     = ifb
        self.inv_iam    = iam
        self.sid_cb     = sid_cb
        self.vendor_idx = vendor_idx or {} # v7.0: Vendor Level index
        self.vendor_raw = vendor_raw or {}
        self.vendor_headers = vendor_headers or []
        self.unique_sids_found = set()
        self.stop_requested = False
        self.pause_requested = False
        self.ticket_type = "PDTT"
        self.cache_sid  = {}
        self.cache_bc  = {}
        self.lock_save = False
        self._cloud_busy = {"REBNI": False, "Invoice": False}
        self._cloud_progress = {"REBNI": 0, "Invoice": 0}
        self._pending_cross_po = []  # FIX 5: Always initialize — prevents AttributeError in _next_man/_handle_res
        self.loop_cache = {}
        self.user_overrides = {}
        self.collected_rebni = {} # (sid, po, asin) -> [rows]
        self.all_found_matches = [] # v6.1.1: Global net for pending invoices
        self.asin_pending_matches = [] # v6.2.5: Per-ASIN match isolation
        self.seen_shp_rebni = set() # v7.3.0: Track shipment-level REBNI to avoid duplicate rows
        self.cross_po_vault = {} # v6.2.6: Shipment-level Cross PO repository

    def check_shipment_validity(self, sid):
        """v7.2.0: Automated IBC vs PBC validation following tiered logic:
        1. Rejected Cartons > 0 => INVALID
        2. SID Level IBC > PBC => INVALID
        """
        rows = self.vendor_raw.get(extract_sid(sid), [])
        if not rows: return False, "No data"
        
        h = self.vendor_headers
        def get_idx(keys):
            for i, col in enumerate(h):
                if any(k.lower() in str(col).lower() for k in keys): return i
            return -1
        
        idx_ibc = get_idx(['invoiced box', 'ibc'])
        idx_pbc = get_idx(['pod box', 'pbc', 'physical box'])
        idx_rej = get_idx(['rejected carton', 'rej'])
        idx_po  = get_idx(['po number', 'po'])

        total_ibc, total_pbc, total_rej = 0, 0, 0
        seen_pos = set()

        for r in rows:
            if idx_rej != -1: total_rej += safe_num(r[idx_rej])
            if idx_ibc != -1: total_ibc += safe_num(r[idx_ibc])
            if idx_pbc != -1:
                po_val = clean(str(r[idx_po])) if idx_po != -1 else None
                if po_val:
                    if po_val not in seen_pos:
                        total_pbc += safe_num(r[idx_pbc])
                        seen_pos.add(po_val)
                else:
                    total_pbc += safe_num(r[idx_pbc])
        
        is_invalid = (total_rej > 0) or (total_ibc > total_pbc)
        reason = ""
        if total_rej > 0: reason = f"Rejected Cartons: {int(total_rej)}"
        elif total_ibc > total_pbc: reason = f"IBC({int(total_ibc)}) > PBC({int(total_pbc)})"
        
        return is_invalid, reason

    def _rebni_lookup(self, sid, po, asin, inv_no=None):
        rows = self.rebni_p.get((sid, po, asin), [])
        if not rows and inv_no:
            rows = self.rebni_fb.get((sid, po, inv_no), [])
        return rows

    def _inv_lookup(self, sid, po, asin, inv_no=None):
        m = self.inv_p.get((sid, po, asin), [])
        if not m and inv_no:
            m = self.inv_fb.get((sid, po, inv_no), [])
        return m

    def _find_sid(self, po, asin, inv_no):
        rows = self.rebni_s.get((clean(po), clean(asin)), [])
        for r in rows:
            if inv_no in split_comma(r.get('matched_invoice_numbers', '')):
                return extract_sid(r['shipment_id'])
        return extract_sid(rows[0]['shipment_id']) if rows else None

    def _resolve_inv_qty(self, inv_no, asin, fallback_qty):
        # v7.1.0: Always prioritize 'Main' invoice (no SCR) quantity per user request
        base = strip_scr(inv_no)
        qty = self.inv_iam.get((base, clean(asin).upper()))
        if qty is not None and qty > 0:
            return qty
            
        # Fallback to literal if base failed or same
        if base != clean(inv_no):
            qty = self.inv_iam.get((clean(inv_no), clean(asin).upper()))
            if qty is not None and qty > 0:
                return qty
        return fallback_qty

    def _get_shipment_rebni(self, sid, po):
        total = 0.0
        for (s, p, a), rows in self.rebni_p.items():
            if s == sid:
                for r in rows:
                    total += safe_num(r.get('rebni_available', 0))
        return total

    def get_cp(self, sid, po, asin):
        rows = self.rebni_p.get((extract_sid(sid), clean(po), clean(asin)), [])
        for r in rows:
            cp = safe_num(r.get('item_cost', 0))
            if cp > 0: return cp
        for (s, p, a), rlist in self.rebni_p.items():
            if p == clean(po) and a == clean(asin):
                for r in rlist:
                    cp = safe_num(r.get('item_cost', 0))
                    if cp > 0: return cp
        return 0.0

    def compare_cp(self, c_sid, c_po, c_asin, m_sid, m_po, m_asin, depth):
        c_cp = self.get_cp(c_sid, c_po, c_asin)
        m_cp = self.get_cp(m_sid, m_po, m_asin)
        c_lbl = "Claiming CP" if depth == 0 else "Parent ASIN CP"
        m_lbl = "Matched CP"  if depth == 0 else "Sub-Matched CP"
        if c_cp <= 0 and m_cp <= 0: return ""
        if c_cp <= 0: return f"{m_lbl}: {m_cp:.2f} | {c_lbl}: N/A"
        if m_cp <= 0: return f"{c_lbl}: {c_cp:.2f} | {m_lbl}: N/A"
        low, high = m_cp * 0.90, m_cp * 1.10
        if low <= c_cp <= high:
            return f"Within 10% CP | {m_lbl}: {m_cp:.2f}, {c_lbl}: {c_cp:.2f} (range: {low:.2f}-{high:.2f})"
        return f"NOT within 10% CP | {m_lbl}: {m_cp:.2f}, {c_lbl}: {c_cp:.2f} (range: {low:.2f}-{high:.2f})"

    def detect_cross_po(self, sid, po, asin, exclude_pos=None):
        candidates = []
        seen_po = set()
        sid_frag = extract_sid(sid)
        
        # v7.2.0: Prioritize Vendor Level Data for Cross PO detection
        if self.vendor_raw and sid_frag in self.vendor_raw:
            h = self.vendor_headers
            def get_idx(keys):
                for i, col in enumerate(h):
                    if any(k.lower() in str(col).lower() for k in keys): return i
                return -1
            
            i_po = get_idx(['po number', 'purchase order', 'po'])
            i_asin = get_idx(['asin', 'product id'])
            i_qty = get_idx(['total received quantity', 'received qty', 'quantity unpacked', 'qty received', 'rec qty'])
            i_date = get_idx(['received date', 'date', 'datetime'])
            i_fc = get_idx(['warehouse', 'fc', 'fulfillment center'])
            i_iq = get_idx(['total invoiced quantity', 'invoiced qty', 'inv qty'])
            i_inv = get_idx(['invoice number', 'invoice', 'inv no', 'inv_no'])
            
            # Find current receiving status at this PO for the "Case" logic
            rec_at_cur = 0
            for row in self.vendor_raw[sid_frag]:
                if clean(str(row[i_po])) == clean(po) and clean(str(row[i_asin])) == clean(asin):
                    rec_at_cur += safe_num(row[i_qty])
            
            for row in self.vendor_raw[sid_frag]:
                r_po = clean(str(row[i_po])) if i_po != -1 else ""
                r_asin = clean(str(row[i_asin])) if i_asin != -1 else ""
                if r_po == clean(po) or r_asin != clean(asin) or r_po in seen_po: continue
                if exclude_pos and r_po in exclude_pos: continue  # v7.2.3: Skip already-investigated POs
                
                rec = safe_num(row[i_qty]) if i_qty != -1 else 0
                if rec <= 0: continue
                
                iq = safe_num(row[i_iq]) if i_iq != -1 else 0
                tp = ""
                if rec_at_cur == 0 and iq == 0: tp = "Case 2 — ASIN not invoiced at this PO, but received"
                elif rec > iq:
                    if iq > 0: tp = "Case 3 — Rec qty > Inv qty (overage in cross PO)"
                    else: tp = "Case 1 — Rec=0 at current PO, units received here"
                else:
                    # v7.3.0 Logic: Correctly identify non-overages as Case 0
                    tp = "Case 0 — Non-overage (Rec <= Inv)"
                
                if tp:
                    seen_po.add(r_po)
                    candidates.append({
                        'po': r_po, 'asin': asin, 'sid': sid_frag, 
                        'inv_qty': fmt_qty(iq), 'rec_qty': rec, 
                        'cross_type': tp, 
                        'date': clean(row[i_date]) if i_date != -1 else "", 
                        'fc_id': clean(row[i_fc]) if i_fc != -1 else "",
                        'found_inv': clean(str(row[i_inv])) if i_inv != -1 else ""
                    })

        # Fallback/Supplemental: Check REBNI (Shipment Level Data)
        rec_at_cur_r = sum(safe_num(r.get('quantity_unpacked', 0)) for r in self.rebni_p.get((sid_frag, clean(po), clean(asin)), []))
        for (s, p, a), rows in self.rebni_p.items():
            if s != sid_frag or a != asin or p == po or p in seen_po: continue
            if exclude_pos and p in exclude_pos: continue  # v7.2.3: Skip already-investigated POs
            for r in rows:
                rec = safe_num(r.get('quantity_unpacked', 0))
                if rec <= 0: continue
                seen_po.add(p)
                im = self.inv_p.get((sid_frag, p, asin), [])
                iq = sum(safe_num(m.get('inv_qty', 0)) for m in im) if im else 0.0
                tp = ""
                if rec_at_cur_r == 0 and iq == 0: tp = "Case 2 — ASIN not invoiced at this PO, but received"
                elif rec > iq:
                    if iq > 0: tp = "Case 3 — Rec qty > Inv qty (overage in cross PO)"
                    else: tp = "Case 1 — Rec=0 at current PO, units received here"
                else:
                    # v7.2.3: Correctly identify non-overages as Case 0
                    tp = "Case 0 — Non-overage (Rec <= Inv)"
                
                if tp:
                    candidates.append({
                        'po': p, 'asin': asin, 'sid': sid_frag, 'inv_qty': fmt_qty(iq), 
                        'rec_qty': rec, 'cross_type': tp, 
                        'date': clean(r.get('received_datetime', '')), 
                        'fc_id': clean(r.get('warehouse_id', '')),
                        'found_inv': clean(r.get('invoice_number', '')) or clean(r.get('invoice', ''))
                    })

        # v7.2.3 AUTO MODE FIX: Bridge Vendor Data to verify specific Invoice and Qty. Zero out if no match.
        final_candidates = []
        if self.vendor_raw and sid_frag in self.vendor_raw:
            raw_rows = self.vendor_raw[sid_frag]
            h = self.vendor_headers
            def get_idx(keys):
                for i, col in enumerate(h):
                    if any(k.lower() in str(col).lower() for k in keys): return i
                return -1
            
            i_po = get_idx(['po number', 'purchase order', 'po'])
            i_inv = get_idx(['invoice number', 'invoice'])
            i_bc = get_idx(['barcode', 'fnsku'])

            for c in candidates:
                matches = []
                seen_invoices = set()
                for r in raw_rows:
                    r_po = clean(str(r[i_po])) if i_po != -1 else ""
                    if r_po == clean(c['po']):
                        r_inv = clean(str(r[i_inv])) if i_inv != -1 else ""
                        if r_inv and r_inv not in seen_invoices:
                            iq = self._resolve_inv_qty(r_inv, c['asin'], None)
                            if iq is not None and iq > 0:
                                r_bc = clean(str(r[i_bc])) if i_bc != -1 else ""
                                matches.append({'inv': r_inv, 'bc': r_bc, 'qty': iq})
                                seen_invoices.add(r_inv)
                
                if matches:
                    c['found_inv'] = matches[0]['inv']
                    c['found_bc'] = matches[0]['bc']
                    c['inv_qty'] = matches[0]['qty']
                    if matches[0]['qty'] >= safe_num(c['rec_qty']): 
                        c['cross_type'] = 'Case 0 — Verified (Not Overage)'
                    elif matches[0]['qty'] > 0:
                        c['cross_type'] = 'Case 3 — Rec qty > Inv qty (overage in cross PO)'
                    else:
                        c['cross_type'] = 'Case 1 — Rec=0 at current PO, units received here'
                    final_candidates.append(c)
                    for m in matches[1:]:
                        new_c = c.copy()
                        new_c['found_inv'] = m['inv']
                        new_c['found_bc'] = m['bc']
                        new_c['inv_qty'] = m['qty']
                        if m['qty'] >= safe_num(new_c['rec_qty']): 
                            new_c['cross_type'] = 'Case 0 — Verified (Not Overage)'
                        elif m['qty'] > 0:
                            new_c['cross_type'] = 'Case 3 — Rec qty > Inv qty (overage in cross PO)'
                        else:
                            new_c['cross_type'] = 'Case 1 — Rec=0 at current PO, units received here'
                        final_candidates.append(new_c)
                else:
                    c['inv_qty'] = 0
                    c['cross_type'] = 'Case 1 — Rec=0 at current PO, units received here'
                    final_candidates.append(c)
        else:
            for c in candidates:
                c['inv_qty'] = 0
                c['cross_type'] = 'Case 1 — Rec=0 at current PO, units received here'
                final_candidates.append(c)

        return final_candidates

    def _make_row(self, b, i, s, p, a, iq, rq, mq, mi, rem, d, depth, rtype='dominant', cp_status='', mtc_asin='', mtc_po='', fc_id='', is_invalid_ibc_pbc=False):
        return {
            'barcode': b, 'invoice': i, 'sid': extract_sid(s) if s else '', 'po': p, 'asin': a,
            'inv_qty': fmt_qty(iq), 'rec_qty': fmt_qty(rq), 'mtc_qty': fmt_qty(mq), 'mtc_inv': mi,
            'mtc_asin': mtc_asin, 'mtc_po': mtc_po, 'fc_id': fc_id, 'remarks': rem, 'date': d, 'depth': depth,
            'type': rtype, 'cp_status': cp_status, 'is_invalid_ibc_pbc': is_invalid_ibc_pbc
        }

    def _build_level_logic(self, barcode, inv_no, sid, po, asin, iqty, rem_pqv, depth, is_claiming, is_manual=False, cross_po_indicator_only=False, initial_cp=0.0, exclude_pos=None):
        while self.pause_requested and not self.stop_requested:
            import time; time.sleep(0.5)
        sid_frag = extract_sid(sid)
        
        # v7.2.0: Check Shipment Validity for the CURRENT shipment (v7.6: Fixed to use SID, not invoice number)
        is_curr_invalid, curr_reason = self.check_shipment_validity(sid_frag or sid)

        # v7.0.1: Auto-populate with SCR cleaning
        v_data = {}
        if clean(inv_no):
            v_data = self.vendor_idx.get(clean(inv_no), {})
            if not v_data:
                v_data = self.vendor_idx.get(strip_scr(inv_no), {})
        # v7.1.3: SID Priority Lookup — Always prioritize SID from Vendor Level Data over REBNI/passed SID
        if v_data and v_data.get('sid'):
            v_sid = extract_sid(v_data['sid'])
            if v_sid:
                sid_frag = v_sid
                self.unique_sids_found.add(sid_frag)
        elif sid_frag:
            self.unique_sids_found.add(sid_frag)

        
        if not barcode or str(barcode).strip() == "":
            barcode = v_data.get('bc', "")
            
        # REMASH TT specific: Auto-populate Invoice number from Vendor Data if missing
        if self.ticket_type == "REMASH" and (not inv_no or str(inv_no).strip() == ""):
            pass # We already key by inv_no now, so this is unlikely unless we search by SID.

        rebni_rows = self.rebni_p.get((sid_frag, clean(po), clean(asin)), [])
        fc_id = clean(rebni_rows[0].get('warehouse_id', '')) if rebni_rows else ""
        
        # v5.9.3 Shipment-Wide Collection: Capture ALL REBNI in this SID
        if sid_frag in self.rebni_sid:
            for r in self.rebni_sid[sid_frag]:
                # v7.3.0: Only collect rows with actual availability > 0 (v6.2.6 alignment)
                if safe_num(r.get('rebni_available', 0)) > 0:
                    rk = (sid_frag, clean(r.get('po','')), clean(r.get('asin','')))
                    if rk not in self.collected_rebni: self.collected_rebni[rk] = []
                    if r not in self.collected_rebni[rk]: self.collected_rebni[rk].append(r)

        # v7.1.0: Strictly use FIRST ROW ONLY for REBNI per Row IB architectural rules
        rec_qty = sum(safe_num(r.get('quantity_unpacked', 0)) for r in rebni_rows)
        r_avail = sum(safe_num(r.get('rebni_available', 0)) for r in rebni_rows)
        ex_adj  = sum(safe_num(r.get('quantity_adjusted', 0)) for r in rebni_rows)
        rec_date = clean(rebni_rows[0].get('received_datetime', '')) if rebni_rows else ""
        cur_cp = initial_cp if initial_cp > 0 else self.get_cp(sid_frag, po, asin)
        cp_disp = f"{cur_cp:.2f}" if cur_cp > 0 else ""
        shortage = max(0.0, safe_num(iqty) - rec_qty)
        acc_at_lvl = shortage + r_avail + ex_adj
        new_rem = max(0.0, rem_pqv - acc_at_lvl)
        remarks = ""

        # v7.2.0: If current shipment is invalid, prepend warning
        if is_curr_invalid:
            remarks = f"[INVALID IBC/PBC] {curr_reason} | "

        if is_claiming or rec_qty < safe_num(iqty) or r_avail > 0 or ex_adj > 0:
            if ex_adj > 0 and shortage > 0:
                remarks += f"Found {int(ex_adj)} units of EX adjustments and {int(shortage)} units of shortage (Inv:{int(iqty)} Rec:{int(rec_qty)})"
            else:
                parts = []
                if shortage > 0: parts.append(f"Inv Qty:{int(iqty)}.Received Qty:{int(rec_qty)}- Shortage of {int(shortage)} Units")
                if r_avail > 0: parts.append(f"REBNI Available = {int(r_avail)} units at {'claiming' if is_claiming else 'matching'} level — Suggest TSP to utilize")
                if ex_adj > 0: parts.append(f"Found {int(ex_adj)} number of X adjustments")
                remarks += " | ".join(parts) if parts else "SR" if depth > 0 else ""
        
        if shortage >= rem_pqv > 0 and not remarks:
            rem = f"Phase 1 Direct Shortage: {int(shortage)} units short received directly"
            if acc_at_lvl > shortage: rem += f" (Total Accounted: {int(acc_at_lvl)} incl. REBNI/EX)"
            m_inv_label = "Self Matching" if is_claiming else "Short Received"
            main_row = self._make_row(barcode, inv_no, sid, po, asin, iqty, rec_qty, rec_qty, m_inv_label, rem, rec_date, depth, cp_status=cp_disp, fc_id=fc_id)
            res_rows = [main_row]
            shp_rebni = self._get_shipment_rebni(sid_frag, clean(po))
            if shp_rebni > 0 and sid_frag not in self.seen_shp_rebni:
                res_rows.append(self._make_row('[REBNI-SHP]', inv_no, sid_frag, po, asin, '', '', shp_rebni, '', f"Shipment-level REBNI = {int(shp_rebni)} units available across all ASINs in this shipment — Suggest TSP to utilize", rec_date, depth, rtype='rebni_shipment', fc_id=fc_id))
                self.seen_shp_rebni.add(sid_frag)
            # v7.2.3: Cross SID REBNI check — ONLY at claiming level
            if is_claiming and rem_pqv > 0:
                cross_sid_rows = self._check_cross_sid_rebni(sid_frag, clean(po), clean(asin), rem_pqv, depth)
                if cross_sid_rows:
                    res_rows.extend(cross_sid_rows)
            if shortage > 0:
                if cross_po_indicator_only and self.ticket_type != "REMASH":
                    _excl = (exclude_pos or set()) | {clean(po)}  # v7.2.3: exclude current + parent POs
                    for c in self.detect_cross_po(sid_frag, clean(po), clean(asin), exclude_pos=_excl):
                        if 'Case 0' in c.get('cross_type', ''): continue  # v7.2.3: skip Rec<=Inv (non-overage)
                        if depth == 0:
                            res_rows.append(self._make_row('[CROSS PO?]', '—', c['sid'], c['po'], c['asin'], c['inv_qty'], c['rec_qty'], '', '', f"Phase 4: Cross PO candidate detected | {c['cross_type']} | Rec={fmt_qty(c['rec_qty'])} units | PENDING USER CONFIRMATION", c['date'], depth, rtype='crosspo', fc_id=c.get('fc_id', '')))
                else:
                    # v6.2.5: For REMASH, always automate Cross PO traversal even in manual/indicator modes
                    res_rows.extend(self._build_cross_po_rows(sid_frag, clean(po), clean(asin), depth))
            return res_rows, [], rec_qty, acc_at_lvl, 0.0, ex_adj


        if 'REBNI Available' in remarks or remarks == 'SR':
            rows = [self._make_row(barcode, inv_no, sid, po, asin, iqty, rec_qty, "", "", remarks, rec_date, depth, cp_status=cp_disp, fc_id=fc_id)]
            if shortage > 0:
                if not (cross_po_indicator_only and self.ticket_type != "REMASH"): 
                    rows.extend(self._build_cross_po_rows(sid_frag, clean(po), clean(asin), depth))
                else:
                    for c in self.detect_cross_po(sid_frag, clean(po), clean(asin), exclude_pos=(exclude_pos or set()) | {clean(po)}):
                        if 'Case 0' in c.get('cross_type', ''): continue  # v7.2.3: skip Rec<=Inv (non-overage)
                        if depth == 0:
                            rows.append(self._make_row('[CROSS PO?]', '—', c['sid'], c['po'], c['asin'], c['inv_qty'], c['rec_qty'], '', '', f"Phase 4: Cross PO candidate detected | {c['cross_type']} | Rec={fmt_qty(c['rec_qty'])} units | PENDING USER CONFIRMATION", c['date'], depth, rtype='crosspo', fc_id=c.get('fc_id', '')))
            return rows, [], rec_qty, acc_at_lvl, max(0.0, rem_pqv - acc_at_lvl), ex_adj

        raw = self.inv_p.get((sid_frag, clean(po), clean(asin)), [])
        
        # v7.1.3: SCR/Main Redundancy Filter — Skip matched invoices that are identical to current claiming invoice
        # This prevents infinite loops or redundant branches when SCR and main invoices share data.
        # v7.3.6: No longer filtering filtered_raw for display. 
        # All matched units must be printed to the Preview/Excel for visibility.
        # Redundancy filtering (to prevent loops) is now moved exclusively to the 'actionable' loop.
        filtered_raw = list(raw)
        
        sorted_m = sorted(filtered_raw, key=lambda x: safe_num(x['mtc_qty']), reverse=True)

        if self.user_overrides:
            patched = []
            for e in sorted_m:
                ovr = self.user_overrides.get(clean(e.get('mtc_inv', '')), {})
                if ovr:
                    e = dict(e)
                    if 'inv_qty' in ovr: e['inv_qty'] = ovr['inv_qty']
                    if 'mtc_qty' in ovr: e['mtc_qty'] = ovr['mtc_qty']
                patched.append(e)
            sorted_m = sorted(patched, key=lambda x: safe_num(x['mtc_qty']), reverse=True)

        m_inv, m_qty = "", ""
        if sorted_m:
            top = sorted_m[0]
            m_inv, m_qty = top['mtc_inv'], fmt_qty(top['mtc_qty'])
        elif not remarks:
            if acc_at_lvl > 0:
                m_inv, m_qty = "Short Received", fmt_qty(acc_at_lvl)
                remarks = f"Accounted for {int(acc_at_lvl)} units (Shortage={int(shortage)}, REBNI={int(r_avail)}, EX={int(ex_adj)})"
            elif rec_qty > 0 and shortage == 0: remarks = "No Invoice Search matches found — Rec Qty = Inv Qty. Possible data mismatch. Verify manually in DICES."
            else: remarks = "No Invoice Search matches found — verify manually."

        cp_str, m_asn, m_po = '', '', ''
        is_self_match = sorted_m and clean(sorted_m[0].get('mtc_inv','')) == clean(inv_no)
        if sorted_m and m_inv not in ('Short Received', ''):
            top = sorted_m[0]; m_asn, m_po = top.get('mtc_asin', ''), top.get('mtc_po', '')
            if not is_self_match:
                # v6.1.1 CP comparison uses Target ASIN (asin) instead of Matched ASIN (m_asn)
                cp_str = self.compare_cp(sid_frag, po, asin, sid_frag, m_po, m_asn, depth)

        rows = [self._make_row(barcode, inv_no, sid, po, asin, iqty, rec_qty, m_qty, m_inv, remarks, rec_date, depth, cp_status=cp_str, mtc_asin=m_asn, mtc_po=m_po, fc_id=fc_id, is_invalid_ibc_pbc=(is_curr_invalid and depth > 0))]
        
        # v7.3.0 Bug Fix: Unconditionally print all matched invoices from the database so the user has full visibility.
        # The sum of these matches accounts for the Received Quantity. They should not be limited by the remaining Shortage.
        for m in sorted_m[(1 if (sorted_m and m_inv not in ("Short Received",)) else 0):]:
            # v7.6: Check Shipment Validity for subrows using MATCHED invoice's SID from Vendor Data
            m_vdata = self.vendor_idx.get(clean(m['mtc_inv']), {})
            m_sid_check = extract_sid(m_vdata.get('sid', '')) if m_vdata else ''
            is_m_invalid, m_reason = self.check_shipment_validity(m_sid_check) if m_sid_check else (False, '')
            rem_m = f"[INVALID IBC/PBC] {m_reason}" if is_m_invalid else ""
            rows.append(self._make_row("", "", "", "", "", "", "", fmt_qty(m['mtc_qty']), m['mtc_inv'], rem_m, "", depth, 'subrow', cp_status=self.compare_cp(sid_frag, po, asin, sid_frag, m.get('mtc_po', po), m.get('mtc_asin', asin), depth), mtc_asin=m.get('mtc_asin',''), mtc_po=m.get('mtc_po',''), fc_id=fc_id, is_invalid_ibc_pbc=is_m_invalid))

        actionable = []
        for m in sorted_m:
            # v7.3.0 Logic Fix: Unify Auto and Manual mode logic. Skip only true self-matches (same inv + same ASIN).
            # This allows Auto mode to properly investigate cross-ASIN mismatches.
            if (m['mtc_inv'] == clean(inv_no) or strip_scr(m['mtc_inv']) == strip_scr(inv_no)) and clean(m.get('mtc_asin', '')) == clean(asin): continue
            a_vdata = self.vendor_idx.get(clean(m['mtc_inv']), {})
            a_sid_check = extract_sid(a_vdata.get('sid', '')) if a_vdata else ''
            is_m_inv, m_reason = self.check_shipment_validity(a_sid_check) if a_sid_check else (False, '')
            actionable.append({**m, 'inv_qty': self._resolve_inv_qty(m['mtc_inv'], m['mtc_asin'], m['inv_qty']), '_depth': depth, '_rem_at_discovery': rem_pqv, '_budget': safe_num(m.get('mtc_qty',0)), 'is_invalid_ibc_pbc': is_m_inv, 'invalid_reason': m_reason})

        
        # v6.2.5: Isolate found matches to the current ASIN's pending list
        self.asin_pending_matches.extend(actionable)
        self.all_found_matches.extend(actionable) 

        # v7.0: Collect all SIDs for Vendor Data extraction
        for m in sorted_m:
            mi = clean(m.get('mtc_inv', ''))
            if mi and mi in self.vendor_idx:
                vd = self.vendor_idx[mi]
                if vd.get('sid'): self.unique_sids_found.add(vd['sid'])

        new_rem = max(0.0, rem_pqv - acc_at_lvl)
        if acc_at_lvl > 0:
            shp_rebni = self._get_shipment_rebni(sid_frag, clean(po))
            if shp_rebni > 0 and sid_frag not in self.seen_shp_rebni:
                rows.append(self._make_row('[REBNI-SHP]', inv_no, sid_frag, po, asin, '', '', shp_rebni, '', f"Shipment-level REBNI = {int(shp_rebni)} units available — Suggest TSP to utilize", rec_date, depth, rtype='rebni_shipment', fc_id=fc_id))
                self.seen_shp_rebni.add(sid_frag)
        # v7.3.6: Only trigger Cross PO rows/indicators if current PO has a shortage (iqty > rec_qty)
        if safe_num(iqty) > rec_qty:
            if not cross_po_indicator_only: 
                rows.extend(self._build_cross_po_rows(sid_frag, clean(po), clean(asin), depth, exclude_pos=exclude_pos))
            else:
                for c in self.detect_cross_po(sid_frag, clean(po), clean(asin)):
                    if depth == 0:
                        rows.append(self._make_row('[CROSS PO?]', '—', c['sid'], c['po'], c['asin'], c['inv_qty'], c['rec_qty'], '', '', f"Phase 4: Cross PO candidate detected | {c['cross_type']} | Rec={fmt_qty(c['rec_qty'])} units | PENDING USER CONFIRMATION", c['date'], depth, rtype='crosspo', fc_id=c.get('fc_id', '')))
        return rows, actionable, rec_qty, acc_at_lvl, new_rem, ex_adj

    def _check_cross_sid_rebni(self, current_sid, po, asin, pqv, depth):
        """v7.2.3: Cross SID REBNI check — only at claiming level.
        Validates cross-SID REBNI via Vendor Data + Invoice Search."""
        rows = []
        s_key = (clean(po), clean(asin))
        cross_entries = self.rebni_s.get(s_key, [])
        if not cross_entries: return rows

        for r in cross_entries:
            r_sid = extract_sid(clean(r.get('shipment_id', '')))
            if not r_sid or r_sid == current_sid: continue
            ra = safe_num(r.get('rebni_available', 0))
            if ra <= 0: continue
            rec_at_cross = safe_num(r.get('quantity_unpacked', 0))

            # Secondary validation via Vendor Data
            is_valid = False
            validation_detail = ""
            cross_inv = ""

            if self.vendor_raw and r_sid in self.vendor_raw:
                raw_rows = self.vendor_raw[r_sid]
                h = self.vendor_headers
                def get_idx(keys):
                    for i, col in enumerate(h):
                        if any(k.lower() in str(col).lower() for k in keys): return i
                    return -1

                i_po = get_idx(['po number', 'purchase order', 'po'])
                i_inv = get_idx(['invoice number', 'invoice'])

                # Find if the PO exists in vendor data for cross SID
                po_found_in_vendor = False
                for vr in raw_rows:
                    v_po = clean(str(vr[i_po])) if i_po != -1 else ""
                    if v_po == clean(po):
                        po_found_in_vendor = True
                        v_inv = clean(str(vr[i_inv])) if i_inv != -1 else ""
                        if v_inv:
                            cross_inv = v_inv
                            # Check invoice qty via Invoice Search
                            inv_qty_at_cross = self._resolve_inv_qty(v_inv, asin, None)
                            if inv_qty_at_cross is not None and inv_qty_at_cross > 0:
                                if rec_at_cross > inv_qty_at_cross:
                                    is_valid = True
                                    validation_detail = f"Rec({int(rec_at_cross)}) > Inv({int(inv_qty_at_cross)}) at Inv {v_inv}"
                                else:
                                    validation_detail = f"Rec({int(rec_at_cross)}) <= Inv({int(inv_qty_at_cross)}) — No overage"
                            else:
                                # Invoice exists but no qty found in Invoice Search — treat as valid
                                is_valid = True
                                validation_detail = f"PO invoiced at {v_inv} but no qty in IS — valid"
                        break

                if not po_found_in_vendor:
                    # PO not invoiced at cross SID → valid (orphaned units)
                    is_valid = True
                    validation_detail = "PO not invoiced at cross SID"
            else:
                # No vendor data for cross SID — cannot validate, skip
                validation_detail = "No vendor data for cross SID"

            if is_valid and ra >= pqv:
                remark = f"Cross SID — REBNI Available = {int(ra)} units at SID {r_sid} | {validation_detail}"
                rows.append(self._make_row('[CROSS SID REBNI]', cross_inv, r_sid, po, asin, '', rec_at_cross, ra, '', remark, clean(r.get('received_datetime', '')), depth, rtype='rebni_crosssid'))

        return rows

    def _build_cross_po_rows(self, sid, po, asin, depth, exclude_pos=None):
        rows = []
        for c in self.detect_cross_po(sid, po, asin, exclude_pos=exclude_pos):
            if 'Case 0' in c.get('cross_type', ''): continue
            
            if safe_num(c['rec_qty']) > 0:
                # Investigating — let run_cross_po_investigation handle the rows to avoid duplicates
                child_rows, _ = self.run_cross_po_investigation(c, c['cross_type'].split("—")[0].strip(), safe_num(c['rec_qty']), depth=depth+1, exclude_pos=exclude_pos)
                rows.extend(child_rows)
            else:
                # Not investigating — print candidate row only
                rows.append(self._make_row('[CROSS PO?]', '—', c['sid'], c['po'], c['asin'], c['inv_qty'], c['rec_qty'], '', c.get('found_inv',''), f"Cross PO — {c['cross_type']} | Overage = {fmt_qty(c['rec_qty'])} units", c['date'], depth+1, rtype='crosspo', fc_id=c.get('fc_id', '')))
        return rows

    def run_auto(self, barcode, inv_no, sid, po, asin, iqty, pqv, depth=0, visited=None, rem_pqv=None, is_claiming=True, branch_budget=None, max_depth_override=None, is_manual=False, row_callback=None, initial_cp=0.0, exclude_pos=None):
        if self.stop_requested: return [], 0.0
        if visited is None: visited = set()
        if rem_pqv is None: rem_pqv = safe_num(pqv)
        if branch_budget is None: branch_budget = rem_pqv
        sid_frag = extract_sid(sid); state = (sid_frag, clean(inv_no), clean(po), clean(asin))
        eff_max = max_depth_override if max_depth_override is not None else self.MAX_DEPTH
        
        cached_rows, cached_acc, cached_actionable = [], 0.0, None # v7.1.3: Initialize to prevent UnboundLocalError

        if state in visited:

            loop_row = self._make_row(barcode, inv_no, sid, po, asin, iqty, 0.0, 0.0, "", "Loop repeating - Skipping duplicate investigative path for this ASIN.", "", depth, rtype='subrow')
            if row_callback: row_callback(loop_row)
            return [loop_row], 0.0
        if depth >= eff_max: return [], 0.0
        visited = visited | {state}
        
        # v7.1.3: Enhanced Loop Cache logic (Always paste and highlight repeated loops)
        if state in self.loop_cache and depth > 0:
            cached_rows, cached_acc, cached_actionable = self.loop_cache[state]
            if row_callback:
                for r in cached_rows:
                    rc = dict(r)
                    if not str(rc.get('remarks','')).startswith("[REPEATED]"):
                        rc['remarks'] = f"[REPEATED] {rc.get('remarks', '')}"
                    rc['rtype'] = 'sr'
                    row_callback(rc)
            
            final_cached = []
            for r in cached_rows:
                rc = dict(r)
                if not str(rc.get('remarks','')).startswith("[REPEATED]"):
                    rc['remarks'] = f"[REPEATED] {rc.get('remarks', '')}"
                rc['rtype'] = 'sr'
                final_cached.append(rc)
            return final_cached, min(branch_budget, cached_acc)

        
        _xp = (exclude_pos or set()) | {clean(po)}
        rows, actionable, rq, acc, n_rem, ex = self._build_level_logic(barcode, inv_no, sid, po, asin, iqty, rem_pqv, depth, is_claiming, is_manual=is_manual, initial_cp=initial_cp, exclude_pos=_xp)
        
        # If resuming from cache, use cached matches and skip matches already explored
        if cached_actionable:
            curr_actionable = cached_actionable
            total_acc = cached_acc
            all_rows = list(cached_rows)
        else:
            curr_actionable = actionable
            total_acc = min(branch_budget, max(0.0, acc))
            all_rows = list(rows)
            if row_callback:
                for r in rows: row_callback(r)

        cur_budget = branch_budget - total_acc
        # v7.3.6: No longer exiting early if cur_budget <= 0. 
        # We must continue to the loop below to ensure all sibling matches are PRINTED for visibility.
        if not curr_actionable or 'REBNI' in all_rows[0].get('remarks', '') or all_rows[0].get('remarks', '') == 'SR':
            if depth > 0: self.loop_cache[state] = (all_rows, total_acc, curr_actionable)
            return all_rows, total_acc

        # Resuming from the first uninvestigated match
        remaining_matches = []
        for i, match in enumerate(curr_actionable):
            if self.stop_requested: break
            while self.pause_requested and not self.stop_requested:
                import time; time.sleep(0.5)

            n_inv, n_po, n_asin = match['mtc_inv'], match['mtc_po'], match['mtc_asin']
            n_budget = safe_num(match['mtc_qty']) if safe_num(match['mtc_qty']) > 0 else cur_budget
            n_iqty = self._resolve_inv_qty(n_inv, n_asin, match['inv_qty'])
            
            # v7.0.1: Vendor Level Data injection with SCR cleaning
            v_data = self.vendor_idx.get(clean(n_inv), {})
            if not v_data: v_data = self.vendor_idx.get(strip_scr(n_inv), {})
            n_sid = self.cache_sid.get(n_inv)
            n_barcode = self.cache_bc.get(n_inv)
            
            if v_data:
                if not n_sid: n_sid = v_data.get('sid')
                if not n_barcode: n_barcode = v_data.get('bc')
                if n_sid: self.unique_sids_found.add(n_sid)
                
            if not n_barcode: n_barcode = "[DICES]"

            if self.ticket_type == "REMASH" and n_sid and n_sid != sid_frag:
                all_rows.append(self._make_row(n_barcode, n_inv, n_sid, n_po, n_asin, n_iqty, "", "", "", "units matching with different shipment Beyond shipment", "", depth + 1, fc_id=""))
                continue

            if not n_sid and self.sid_cb:
                n_sid = self.sid_cb(n_inv, n_po, n_asin)
                if n_sid: self.cache_sid[n_inv] = n_sid
            if not n_sid:
                all_rows.append(self._make_row(n_barcode, n_inv, "[ENTER SID FROM DICES]", n_po, n_asin, n_iqty, "", "", "", "Phase 2: SID not found — validate in DICES", "", depth + 1, fc_id=""))
                continue

            if cur_budget <= 0:
                # v7.3.6: Budget exhausted. Only print the match header (level) and don't investigate deeper.
                # This ensures every sibling match found in the database is visible to the associate.
                c_rows, _, _, _, _, _ = self._build_level_logic(n_barcode, n_inv, n_sid, n_po, n_asin, n_iqty, rem_pqv, depth+1, False, initial_cp=initial_cp, exclude_pos=_xp)
                if row_callback:
                    for cr in c_rows: row_callback(cr)
                all_rows.extend(c_rows)
                continue

            child_rows, child_acc = self.run_auto(n_barcode, n_inv, n_sid, n_po, n_asin, n_iqty, rem_pqv, depth+1, visited, rem_pqv-total_acc, False, n_budget, max_depth_override, is_manual, row_callback, initial_cp=initial_cp, exclude_pos=_xp)
            all_rows.extend(child_rows)
            contribution = min(cur_budget, child_acc)
            total_acc += contribution
            cur_budget -= contribution

        if depth > 0: self.loop_cache[state] = (all_rows, total_acc, [])
        return all_rows, total_acc

    def run_mismatch_investigation(self, data, budget, depth=0):
        # v6.2.5 Fix: Properly unpack all 6 variables and return them for consistent UI handling
        rows, match, rq, shortage, n_rem, ex = self._build_level_logic("[MISMATCH]", "", extract_sid(data.get('sid', '')), clean(data.get('po', '')), clean(data.get('asin', '')), safe_num(data.get('inv_qty', 0)), budget, depth, False)
        return rows, match, rq, shortage, n_rem, ex

    def build_one_level(self, b, i, s, p, a, iq, rem, depth=0, is_claiming=True, is_manual=False, initial_cp=0.0, exclude_pos=None):
        # v6.2.5: Consistent return with 6-value internal engine return
        # v7.2.3: exclude_pos carries the ancestor PO chain to prevent redundant cross PO re-detection
        _xp = (exclude_pos or set()) | {clean(p)}
        rows, matches, rq, acc, n_rem, ex = self._build_level_logic(b, i, s, p, a, iq, rem, depth, is_claiming, is_manual=is_manual, cross_po_indicator_only=is_manual, initial_cp=initial_cp, exclude_pos=_xp)
        # Auto mode: strip all same-invoice matches (handled by run_cross_po_investigation)
        # Manual mode: only strip true self-matches; keep mismatches (different ASIN) for user investigation
        if is_manual:
            filtered = [m for m in matches if not ((m['mtc_inv'] == clean(i) or strip_scr(m['mtc_inv']) == strip_scr(i)) and clean(m.get('mtc_asin', '')) == clean(a))]
        else:
            filtered = [m for m in matches if not (m['mtc_inv'] == clean(i) or strip_scr(m['mtc_inv']) == strip_scr(i))]
        return rows, filtered, rq, n_rem, ex

    def run_cross_po_investigation(self, c, case_type, budget, depth=0, visited=None, exclude_pos=None):
        if visited is None: visited = set()
        if exclude_pos is None: exclude_pos = set()
        c_sid, c_po, c_asin = c['sid'], c['po'], c['asin']; c_iq = safe_num(c.get('inv_qty', 0))
        state = (extract_sid(c_sid), clean(c.get('invoice_number', '')), clean(c_po), clean(c_asin))
        if state in visited:
            if self.ticket_type == "REMASH":
                return [self._make_row('[LOOP]', '—', c_sid, c_po, c_asin, c_iq, 0.0, '', '', "Loop repeating - Skipping duplicate cross-PO path.", c['date'], depth, rtype='crosspo')], 0.0
            return [], 0.0
        visited.add(state)
        raw = self.inv_p.get((extract_sid(c_sid), clean(c_po), clean(c_asin)), [])
        unique = sorted(raw, key=lambda x: safe_num(x['mtc_qty']), reverse=True)
        rebni = self.rebni_p.get((extract_sid(c_sid), clean(c_po), clean(c_asin)), [])
        # v7.1.0: Strictly use FIRST ROW ONLY for REBNI per Row IB architectural rules
        rq = safe_num(rebni[0].get('quantity_unpacked', 0)) if rebni else 0.0
        ra = safe_num(rebni[0].get('rebni_available', 0)) if rebni else 0.0
        ex = safe_num(rebni[0].get('quantity_adjusted', 0)) if rebni else 0.0
        rd = clean(rebni[0].get('received_datetime', '')) if rebni else ''
        fc_id = clean(rebni[0].get('warehouse_id', '')) if rebni else ''
        shortage = max(0.0, c_iq - rq)
        overage = max(0.0, rq - c_iq)
        acc = overage + ra + ex
        m_inv = unique[0]['mtc_inv'] if unique else "Short Received"; m_qty = fmt_qty(unique[0]['mtc_qty']) if unique else ""
        rows = []
        rem = f"Phase 4 Cross PO ({case_type}): Accounted for {int(acc)} units"
        if not unique and shortage > 0: rem += " — Target met via Direct Shortage"
        elif ra > 0: rem += " — Suggest TSP to utilize REBNI"
        
        total_acc = min(budget, max(0.0, acc))
        # v7.2.3 AUTO MODE FIX: Decouple mathematical budget from chain tracing for zeroed-out Cross POs
        chain_rem = budget if c_iq == 0 else (budget - total_acc)
        
        # v6.2.6 logic: Always print the first summary row with its match (if any)
        rows = [self._make_row('[CROSS PO?]', '—', c_sid, c_po, c_asin, fmt_qty(c_iq), rq, m_qty, m_inv, rem, rd, depth, mtc_asin=unique[0].get('mtc_asin','') if unique else '', mtc_po=unique[0].get('mtc_po','') if unique else '', fc_id=fc_id)]
        
        # v7.3.6: Restored sub-row printing for all discovered matches in Cross PO.
        # This ensures that if a Cross PO has multiple matches (e.g. 14 units + 1 unit), they both appear in the output.
        if unique:
            for m in unique[1:]:
                rows.append(self._make_row("", "", "", "", "", "", "", fmt_qty(m['mtc_qty']), m['mtc_inv'], "", "", depth, 'subrow', mtc_asin=m.get('mtc_asin',''), mtc_po=m.get('mtc_po',''), fc_id=fc_id))

        if ra > 0 or not unique: return rows, total_acc
        
        for match in unique:
            if match not in getattr(self, 'all_found_matches', []):
                self.all_found_matches.append(match)
                if hasattr(self, 'asin_pending_matches'):
                    self.asin_pending_matches.append(match)
            while self.pause_requested and not self.stop_requested:
                import time; time.sleep(0.5)

            n_inv, n_po, n_asin = match['mtc_inv'], match['mtc_po'], match['mtc_asin']
            # v6.2.6: Zero Rule - If combination not found in IAM, set Inv Qty to 0 for overage tracing.
            n_iq = self.inv_iam.get((clean(n_inv), clean(n_asin).upper()), 0)
            if n_iq == 0:
                base = strip_scr(n_inv)
                if base != clean(n_inv):
                    n_iq = self.inv_iam.get((base, clean(n_asin).upper()), 0)
            
            n_bud = safe_num(match['mtc_qty']) if safe_num(match['mtc_qty']) > 0 else chain_rem
            state = (extract_sid(c_sid), clean(n_inv), clean(n_po), clean(n_asin))
            if state in visited: continue

            if chain_rem <= 0:
                # v7.3.6 (Fixed for Auto Mode): Budget exhausted. Print matches without deep investigation.
                c_rows, _, _, _, _, _ = self._build_level_logic(n_barcode, n_inv, n_sid, n_po, n_asin, n_iq, chain_rem, depth+1, False, exclude_pos=exclude_pos)
                rows.extend(c_rows)
                continue

            # v6.1.1: Log global processing so budget-exhausted sub-items drop out of pending correctly
            # v7.0.1: Vendor Level Data injection with SCR cleaning
            v_data = self.vendor_idx.get(clean(n_inv), {})
            if not v_data: v_data = self.vendor_idx.get(strip_scr(n_inv), {})
            n_sid = self.cache_sid.get(n_inv)
            n_barcode = self.cache_bc.get(n_inv)
            
            if v_data:
                if not n_sid: n_sid = v_data.get('sid')
                if not n_barcode: n_barcode = v_data.get('bc')
                if n_sid: self.unique_sids_found.add(n_sid)
                
            if not n_barcode: n_barcode = "[DICES]"

            # REMASH SID check
            # v6.2.5: Cross PO investigation always allows shipment transitions to trace overage chains
            pass 

            if not n_sid and self.sid_cb:
                n_sid = self.sid_cb(n_inv, n_po, n_asin)
                if n_sid: self.cache_sid[n_inv] = n_sid

            if not n_sid:
                rows.append(self._make_row(n_barcode, n_inv, "[ENTER SID]", n_po, n_asin, n_iq, "", "", "", "Phase 4: SID not found — validate in DICES", "", depth + 1))
                continue
            child_rows, child_acc = self.run_auto("", n_inv, n_sid, n_po, n_asin, n_iq, chain_rem, depth+1, visited, chain_rem, False, min(n_bud, chain_rem), exclude_pos=exclude_pos)
            rows.extend(child_rows); contri = min(chain_rem, child_acc); chain_rem -= contri
            if c_iq > 0: total_acc += contri
        return rows, total_acc


# ═══════════════════════════════════════════════════════════
#  EXCEL WRITER
# ═══════════════════════════════════════════════════════════

def write_excel(all_blocks, path, rebni_summary_data=None, unique_sids_found=None, vendor_raw=None, vendor_headers=None, raw_claims=None, raw_rebni=None, raw_inv=None, ciat_asin_df=None, ciat_shipment_df=None, ciat_shipment_merges=None):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Investigation"
    headers = ["Barcode", "Inv no", "SID", "PO", "ASIN", "Inv Qty", "Rec Qty", "Mtc Qty", "Mtc Inv", "Mtc ASIN", "Mtc PO", "FC", "Remarks", "Date", "CP"]
    H_FILL, DOM_F, SUB_F, ROOT_F, DICES_F, SR_F, INVLD_F, REBNI_F, CROSS_F, MIS_F, PINK_F = [PatternFill("solid", fgColor=c) for c in ["203864", "E2EFDA", "EBF3FB", "FFE0E0", "FFF2CC", "FFD7D7", "FFD0D0", "D0F0FF", "FFF0C0", "D0E8FF", "FFE0F0"]]
    H_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=10); N_FONT = Font(name="Calibri", size=10); ROOT_FT = Font(bold=True, color="9C0006", name="Calibri", size=10); SR_FT = Font(bold=True, color="CC0000", name="Calibri", size=10)
    INVLD_FT = Font(bold=True, color="880000", name="Calibri", size=10, italic=True); REBNI_FT = Font(bold=True, color="005580", name="Calibri", size=10); CROSS_FT = Font(bold=True, color="7a5c00", name="Calibri", size=10)
    SHORT_FILL = PatternFill("solid", fgColor="FFCCCC"); SHORT_FONT = Font(bold=True, color="9C0006", name="Calibri", size=10); BDR = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),  bottom=Side(style='thin'))
    INVALID_SHIPMENT_FILL = PatternFill("solid", fgColor="FFF9C4"); INVALID_SHIPMENT_FONT = Font(bold=True, color="000000", name="Calibri", size=10)
    INVALID_HIGHLIGHT_COLS = {'SID', 'Mtc Qty', 'Mtc Inv', 'Mtc ASIN', 'Mtc PO'}
    KM = {'Barcode':'barcode', 'Inv no':'invoice', 'SID':'sid', 'PO':'po', 'ASIN':'asin', 'Inv Qty':'inv_qty', 'Rec Qty':'rec_qty', 'Mtc Qty':'mtc_qty', 'Mtc Inv':'mtc_inv', 'Mtc ASIN':'mtc_asin', 'Mtc PO':'mtc_po', 'FC':'fc_id', 'Remarks':'remarks', 'Date':'date', 'CP':'cp_status'}
    curr = 1
    for block in all_blocks:
        if not block: continue
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=curr, column=c, value=h); cell.fill, cell.font, cell.border = H_FILL, H_FONT, BDR
        curr += 1
        for rd in block:
            rem, rtyp, dep = str(rd.get('remarks', '')), rd.get('type', 'dominant'), rd.get('depth', 0)
            iq, rq = safe_num(rd.get('inv_qty')), safe_num(rd.get('rec_qty')); is_short = (iq > 0 and rq < iq)
            is_invalid_row = rd.get('is_invalid_ibc_pbc')
            if is_short: fill, fnt = SHORT_FILL, SHORT_FONT
            elif 'invalid' in rem.lower(): fill, fnt = INVLD_F, INVLD_FT
            elif 'short received' in rem.lower() or 'direct shortage' in rem.lower() or 'root cause' in rem.lower() or ('found' in rem.lower() and 'short' in rem.lower()): fill, fnt = SHORT_FILL, SHORT_FONT
            else: fill, fnt = None, N_FONT
            for c, h in enumerate(headers, 1):
                val = rd.get(KM[h], ""); final_val = val; str_val = str(val).strip()
                is_hyperlink = False
                
                if val not in (None, '') and str_val and str_val != '-':
                    if h == 'Barcode' and not str_val.startswith('['):
                        final_val = f'=HYPERLINK("https://smocentral.amazon.eu/dices/document?documentID={str_val}", "{str_val}")'
                        is_hyperlink = True
                    elif h == 'SID' and not str_val.startswith('['):
                        final_val = f'=HYPERLINK("https://smocentral.amazon.eu/dices/advanced-search?shipment_id={str_val}", "{str_val}")'
                        is_hyperlink = True
                    elif str_val.replace('.','',1).isdigit(): 
                        final_val = safe_num(str_val)
                else:
                    if val not in (None, '') and str_val.replace('.','',1).isdigit(): final_val = safe_num(str_val)
                
                cell = ws.cell(row=curr, column=c, value=final_val if final_val not in (None, '') else None); cell.border, cell.font = BDR, fnt
                if is_invalid_row and h in INVALID_HIGHLIGHT_COLS: cell.fill, cell.font = INVALID_SHIPMENT_FILL, INVALID_SHIPMENT_FONT
                elif is_short and h == "Remarks": cell.fill, cell.font = SHORT_FILL, SHORT_FONT
                elif fill: cell.fill = fill
                
                if is_hyperlink:
                    cell.font = Font(name=cell.font.name, size=cell.font.size, bold=cell.font.bold, italic=cell.font.italic, color="0563C1", underline="single")
            curr += 1
        curr += 1
    for i, w in enumerate([18, 22, 18, 12, 14, 9, 9, 9, 26, 18, 18, 10, 42, 22, 36], 1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws2 = wb.create_sheet(title="Unique Summary")
    SH_FILL, SH_FONT = PatternFill("solid", fgColor="1F4E79"), Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    SN_FONT, SRED, SGRN = Font(name="Calibri", size=10), PatternFill("solid", fgColor="FFE2E2"), PatternFill("solid", fgColor="E2F0D9")
    CTR, LFT, SBDR = Alignment(horizontal="center", vertical="center"), Alignment(horizontal="left", vertical="center"), Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),  bottom=Side(style='thin'))
    sh = ["Barcode", "Inv no", "SID", "PO", "ASIN", "Inv Qty", "Rec Qty", "Missing QTY", "Cost Price", "Shortage Amount"]
    sw = [20, 14, 18, 12, 14, 10, 10, 14, 12, 18]
    for ci, (h, w) in enumerate(zip(sh, sw), 1):
        cell = ws2.cell(row=1, column=ci, value=h); cell.fill, cell.font, cell.border, cell.alignment = SH_FILL, SH_FONT, SBDR, CTR
        ws2.column_dimensions[get_column_letter(ci)].width = w
    seen, sr = set(), 2
    for block in all_blocks:
        for rd in block:
            b = str(rd.get('barcode', '')).strip()
            if not b or b == 'Barcode': continue
            if rd.get('type', '') == 'subrow': continue
            i, s, p, a, iq, rq = str(rd.get('invoice','')).strip(), str(rd.get('sid','')).strip(), str(rd.get('po','')).strip(), str(rd.get('asin','')).strip(), rd.get('inv_qty', 0), rd.get('rec_qty', 0)
            key = (b, i, s, p, a, str(iq), str(rq))
            if key in seen: continue
            seen.add(key); iq_n, rq_n = safe_num(iq), safe_num(rq) if str(rq).strip() != '' else 0.0
            rf = SRED if iq_n > rq_n else SGRN
            v = [b, i, s, p, a, int(iq_n) if iq_n == int(iq_n) else iq_n, int(rq_n) if rq_n == int(rq_n) else rq_n, None, safe_num(rd.get('cp_status', 0)), None]
            for ci, val in enumerate(v, 1):
                if ci == 8: cell = ws2.cell(row=sr, column=ci, value=f"=F{sr}-G{sr}")
                elif ci == 10: cell = ws2.cell(row=sr, column=ci, value=f"=H{sr}*I{sr}")
                else: cell = ws2.cell(row=sr, column=ci, value=val)
                cell.font, cell.border, cell.fill, cell.alignment = SN_FONT, SBDR, rf, (CTR if ci >= 6 else LFT)
            sr += 1
    ws2.freeze_panes = "A2"
    
    if rebni_summary_data:
        ws3 = wb.create_sheet(title="REBNI Summary")
        rh = ["PO", "ASIN", "SID", "Item_Cost", "Rebni_Available"]
        rw = [15, 15, 18, 12, 18]
        for ci, (h, w) in enumerate(zip(rh, rw), 1):
            cell = ws3.cell(row=1, column=ci, value=h); cell.fill, cell.font, cell.border, cell.alignment = SH_FILL, SH_FONT, SBDR, CTR
            ws3.column_dimensions[get_column_letter(ci)].width = w
        
        row_idx = 2
        # Data is dict: (sid, po, asin) -> [rows]
        for key, rows in rebni_summary_data.items():
            for r in rows:
                v = [r.get('po',''), r.get('asin',''), r.get('shipment_id',''), 
                     safe_num(r.get('item_cost',0)), safe_num(r.get('rebni_available',0))]
                for ci, val in enumerate(v, 1):
                    cell = ws3.cell(row=row_idx, column=ci, value=val)
                    cell.font, cell.border, cell.alignment = SN_FONT, SBDR, (CTR if ci >= 4 else LFT)
                row_idx += 1
        ws3.freeze_panes = "A2"

    # v7.1.3: ROOT_CAUSE_ONLY sheet (summary of all shortages found)
    ws_rc = wb.create_sheet(title="ROOT_CAUSE_ONLY")
    rch = ["Barcode", "Inv no", "SID", "PO", "ASIN", "Invoiced Quantity", "Received Quantity", "Missing Quantity", "Remarks"]
    rcw = [20, 18, 18, 12, 14, 18, 18, 18, 50]
    for ci, (h, w) in enumerate(zip(rch, rcw), 1):
        cell = ws_rc.cell(row=1, column=ci, value=h)
        cell.fill, cell.font, cell.border, cell.alignment = SH_FILL, SH_FONT, SBDR, CTR
        ws_rc.column_dimensions[get_column_letter(ci)].width = w
    
    rc_idx = 2
    seen_rc = set()
    for block in all_blocks:
        for rd in block:
            iq_n, rq_n = safe_num(rd.get('inv_qty')), safe_num(rd.get('rec_qty'))
            # User Rule: "each and every column which is in red color which are shortages"
            if iq_n > rq_n:
                miss = iq_n - rq_n
                v = [rd.get('barcode',''), rd.get('invoice',''), rd.get('sid',''), rd.get('po',''), rd.get('asin',''),
                     int(iq_n) if iq_n == int(iq_n) else iq_n,
                     int(rq_n) if rq_n == int(rq_n) else rq_n,
                     int(miss) if miss == int(miss) else miss,
                     rd.get('remarks', '')]
                
                # Global Deduplication for Root Cause Sheet
                key = tuple(str(x).strip() for x in v)
                if key in seen_rc: continue
                seen_rc.add(key)
                
                for ci, val in enumerate(v, 1):
                    is_hyperlink = False
                    if val not in (None, '') and str(val).strip() and str(val).strip() != '-' and not str(val).strip().startswith('['):
                        if ci == 1:
                            val = f'=HYPERLINK("https://smocentral.amazon.eu/dices/document?documentID={val}", "{val}")'
                            is_hyperlink = True
                        elif ci == 3:
                            val = f'=HYPERLINK("https://smocentral.amazon.eu/dices/advanced-search?shipment_id={val}", "{val}")'
                            is_hyperlink = True
                            
                    cell = ws_rc.cell(row=rc_idx, column=ci, value=val)
                    cell.font, cell.border, cell.alignment = SN_FONT, SBDR, (CTR if ci >= 6 and ci <= 8 else LFT)
                    if is_hyperlink:
                        cell.font = Font(name=SN_FONT.name, size=SN_FONT.size, color="0563C1", underline="single")
                rc_idx += 1
    ws_rc.freeze_panes = "A2"


    
    if unique_sids_found and vendor_raw and vendor_headers:

        ws_ibc = wb.create_sheet("IBC and PBC validation")
        r_idx = 1
        
        # Helper to find column indices dynamically
        def get_idx(*matches):
            for i, h in enumerate(vendor_headers):
                hl = str(h).lower().strip()
                if any(m.lower() in hl for m in matches): return i
            return -1
            
        idx_po = get_idx('po number', 'purchase order', 'po')
        idx_sid = get_idx('shipment id', 'shipment', 'sid')
        idx_bc = get_idx('barcode', 'fnsku', 'asin')
        idx_ibc = get_idx('invoice box count', 'ibc', 'invoice box')
        idx_pbc = get_idx('pod box quantity', 'pbc', 'physical box')
        idx_inv = get_idx('invoice number', 'invoice')
        idx_qty = get_idx('total invoiced quantity', 'invoiced quantity', 'qty')

        for sid in unique_sids_found:
            rows = vendor_raw.get(sid, [])
            if not rows: continue
            
            # Paste original Vendor Headers for this Shipment ID block
            for c, h in enumerate(vendor_headers, 1):
                cell = ws_ibc.cell(row=r_idx, column=c, value=h)
                cell.fill, cell.font, cell.border = H_FILL, H_FONT, BDR
            r_idx += 1
            
            seen_pos = set()
            total_ibc = 0
            total_pbc = 0
            total_inv_qty = 0
            
            # Paste exactly all original raw data rows for this Shipment ID
            for row_vals in rows:
                mod_vals = list(row_vals) # Create a copy to modify safely
                po_val = clean(str(mod_vals[idx_po])) if idx_po != -1 else None
                
                # Deduplication logic for PBC: Zero out PBC for repeated POs
                if po_val:
                    if po_val in seen_pos:
                        if idx_pbc != -1: mod_vals[idx_pbc] = 0
                    else:
                        seen_pos.add(po_val)
                
                # Accumulate Totals
                if idx_ibc != -1:
                    try: total_ibc += float(str(mod_vals[idx_ibc]).strip()) if str(mod_vals[idx_ibc]).strip() else 0
                    except: pass
                if idx_pbc != -1:
                    try: total_pbc += float(str(mod_vals[idx_pbc]).strip()) if str(mod_vals[idx_pbc]).strip() else 0
                    except: pass
                if idx_qty != -1:
                    try: total_inv_qty += float(str(mod_vals[idx_qty]).strip()) if str(mod_vals[idx_qty]).strip() else 0
                    except: pass

                for c, val in enumerate(mod_vals, 1):
                    is_hyperlink = False
                    if val not in (None, '') and str(val).strip() and str(val).strip() != '-' and not str(val).strip().startswith('['):
                        if (c - 1) == idx_bc:
                            val = f'=HYPERLINK("https://smocentral.amazon.eu/dices/document?documentID={val}", "{val}")'
                            is_hyperlink = True
                        elif (c - 1) == idx_sid:
                            val = f'=HYPERLINK("https://smocentral.amazon.eu/dices/advanced-search?shipment_id={val}", "{val}")'
                            is_hyperlink = True

                    # v7.0.1: Convert numeric strings to actual numbers to prevent "Number Stored as Text" warning
                    if not is_hyperlink and isinstance(val, str) and val.strip():
                        sval = val.strip()
                        if sval.isdigit():
                            # Retain as string if it is a barcode/number with leading zero(s) (e.g. 000123)
                            if not (sval.startswith('0') and len(sval) > 1):
                                val = int(sval)
                        else:
                            try:
                                val = float(sval)
                            except ValueError:
                                pass
                    cell = ws_ibc.cell(row=r_idx, column=c, value=val)
                    if is_hyperlink:
                        cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
                r_idx += 1
            
            # Write Total Row
            t_fill_green = PatternFill("solid", fgColor="E2EFDA")
            t_fill_red = PatternFill("solid", fgColor="FF0000")
            
            col_sid = idx_sid + 1 if idx_sid != -1 else 1
            cell_sid = ws_ibc.cell(row=r_idx, column=col_sid, value="TOTAL")
            
            # Conditional formatting: Red if IBC > PBC, otherwise light green
            if total_ibc > total_pbc:
                cell_sid.fill = t_fill_red
                cell_sid.font = Font(bold=True, color="FFFFFF")
            else:
                cell_sid.fill = t_fill_green
                cell_sid.font = Font(bold=True)
                
            # Place "TOTAL QTY" under Invoice Number column
            if idx_inv != -1:
                cell_inv = ws_ibc.cell(row=r_idx, column=idx_inv + 1, value="TOTAL QTY")
                cell_inv.font = Font(bold=True)
                cell_inv.fill = t_fill_green
                
            # Place sums
            if idx_qty != -1:
                ws_ibc.cell(row=r_idx, column=idx_qty + 1, value=int(total_inv_qty) if float(total_inv_qty).is_integer() else total_inv_qty).font = Font(bold=True)
            if idx_ibc != -1:
                ws_ibc.cell(row=r_idx, column=idx_ibc + 1, value=int(total_ibc) if float(total_ibc).is_integer() else total_ibc).font = Font(bold=True)
            if idx_pbc != -1:
                ws_ibc.cell(row=r_idx, column=idx_pbc + 1, value=int(total_pbc) if float(total_pbc).is_integer() else total_pbc).font = Font(bold=True)
            
            # Blank row between different Shipment IDs
            r_idx += 2

    # v7.6: Append raw input files as separate sheets (single-pass, in-memory)
    from openpyxl.utils.dataframe import dataframe_to_rows
    for sheet_title, df in [("ASIN Level Analysis", ciat_asin_df), ("Shipment Level Analysis", ciat_shipment_df), ("Claims Data", raw_claims), ("REBNI Data", raw_rebni), ("Invoice Search Data", raw_inv)]:
        if df is not None and not df.empty:
            ws_raw = wb.create_sheet(title=sheet_title)
            for ri, r in enumerate(dataframe_to_rows(df, index=False, header=True)):
                clean_row = []
                for v in r:
                    if ri > 0 and v is not None and not isinstance(v, (int, float)):
                        sv = str(v).strip()
                        try: v = int(sv)
                        except ValueError:
                            try: v = float(sv)
                            except ValueError: pass
                    clean_row.append(v)
                ws_raw.append(clean_row)
                
            if sheet_title == "Shipment Level Analysis" and ciat_shipment_merges:
                for merge_range in ciat_shipment_merges:
                    ws_raw.merge_cells(merge_range)
                    start_cell = merge_range.split(':')[0]
                    ws_raw[start_cell].alignment = Alignment(vertical='center', horizontal='center')

    wb.save(path)


# ═══════════════════════════════════════════════════════════
#  MAIN GUI
# ═══════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════
#  SEARCH CLOUD DIALOG
# ═══════════════════════════════════════════════════════════

class SearchCloudDialog(tk.Toplevel):
    def __init__(self, parent, s_type, callback):
        super().__init__(parent)
        self.callback = callback
        self.s_type = s_type
        self.title(f"Search {s_type} - v7.7 SECURED")
        self.geometry("560x420")
        self.resizable(True, True)
        self.configure(bg="#0f0f1a")
        
        self.seller_var = tk.StringVar(value="VRP")
        self.v_code_var = tk.StringVar()
        self.s_id_var = tk.StringVar()
        self.source_var = tk.StringVar(value="Local")  # Local, Cloud, Custom
        self.custom_path = tk.StringVar()

        tk.Label(self, text=f"🔍 {s_type} Search Engine", 
                 bg="#16213e", fg="#4a9eff", font=("Segoe UI", 12, "bold"), height=2).pack(fill="x")
        
        # Source Selection
        sf = tk.LabelFrame(self, text="  Data Source Selection  ", bg="#0f0f1a", fg="#f0a500", font=("Segoe UI", 9, "bold"), padx=15, pady=8)
        sf.pack(fill="x", padx=20, pady=10)
        
        tk.Radiobutton(sf, text="Local (PD App Folder - Recommended)", variable=self.source_var, value="Local", 
                       bg="#0f0f1a", fg="#ffffff", selectcolor="#16213e", activebackground="#0f0f1a", font=("Segoe UI", 9)).pack(anchor="w")
        tk.Radiobutton(sf, text="Cloud Drive (W: Drive)", variable=self.source_var, value="Cloud", 
                       bg="#0f0f1a", fg="#ffffff", selectcolor="#16213e", activebackground="#0f0f1a", font=("Segoe UI", 9)).pack(anchor="w")
        tk.Radiobutton(sf, text="Custom File Selection", variable=self.source_var, value="Custom", 
                       bg="#0f0f1a", fg="#ffffff", selectcolor="#16213e", activebackground="#0f0f1a", font=("Segoe UI", 9)).pack(anchor="w")

        # Standard Fields
        f_fields = tk.Frame(self, bg="#0f0f1a"); f_fields.pack(fill="x", padx=35)
        
        f1 = tk.Frame(f_fields, bg="#0f0f1a"); f1.pack(pady=5, fill="x")
        tk.Label(f1, text="Seller Name:", bg="#0f0f1a", fg="#cccccc", width=12, anchor="w").pack(side="left")
        sellers = ["VRP", "RK World", "KayKay", "Clicktech", "Dawntech", "Etrade", "Cocoblu", "Retail EZ"]
        cb = ttk.Combobox(f1, textvariable=self.seller_var, values=sellers, state="readonly", width=25)
        cb.pack(side="left", padx=10)
        
        f2 = tk.Frame(f_fields, bg="#0f0f1a"); f2.pack(pady=5, fill="x")
        tk.Label(f2, text="Vendor Code:", bg="#0f0f1a", fg="#cccccc", width=12, anchor="w").pack(side="left")
        tk.Entry(f2, textvariable=self.v_code_var, width=28, bg="#1e1e3a", fg="white", insertbackground="white", relief="flat").pack(side="left", padx=10)

        f3 = tk.Frame(f_fields, bg="#0f0f1a"); f3.pack(pady=5, fill="x")
        tk.Label(f3, text="Shipment ID:", bg="#0f0f1a", fg="#cccccc", width=12, anchor="w").pack(side="left")
        tk.Entry(f3, textvariable=self.s_id_var, width=28, bg="#1e1e3a", fg="white", insertbackground="white", relief="flat").pack(side="left", padx=10)

        # Custom Path Field
        bf2 = tk.Frame(self, bg="#0f0f1a"); bf2.pack(fill="x", padx=35, pady=5)
        tk.Entry(bf2, textvariable=self.custom_path, width=42, bg="#131320", fg="#4a9eff", relief="flat").pack(side="left", padx=(0,5))
        tk.Button(bf2, text="Browse", command=self._browse_custom, bg="#2d2d5e", fg="white", relief="flat").pack(side="left")

        # Action Buttons
        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=(15, 5))
        tk.Button(bf, text="CANCEL", command=self.destroy, bg="#444", fg="white", width=10, relief="flat").pack(side="left", padx=5)
        tk.Button(bf, text="SEARCH", command=self._on_search, bg="#e94560", fg="white", width=12, relief="flat", font=("Segoe UI", 9, "bold")).pack(side="left", padx=5)
        # v6.2.5: High-speed vendor-only fetch mode
        tk.Button(bf, text="FAST FETCH", command=self._on_fast_fetch, bg="#6f42c1", fg="white", width=12, relief="flat", font=("Segoe UI", 9, "bold")).pack(side="left", padx=5)
        
        if s_type == "REBNI":
            self.comp_var = tk.BooleanVar()
            tk.Checkbutton(self, text="Completely Matched & Available REBNI", variable=self.comp_var, 
                           fg="#3fb950", bg="#0f0f1a", selectcolor="#16213e", activebackground="#0f0f1a", 
                           font=("Segoe UI", 9, "bold")).pack(pady=5)
        
        apply_global_theme_to_widget(self)
        self.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{x}+{y}")

    def _browse_custom(self):
        f = filedialog.askopenfilename(filetypes=[("Text files", "*.txt"), ("CSV files", "*.csv")])
        if f: 
            self.custom_path.set(f)
            self.source_var.set("Custom")

    def _on_search(self):
        res = {
            'seller': self.seller_var.get(), 
            'vendor_code': self.v_code_var.get().strip(),
            'shipment_id': self.s_id_var.get().strip(),
            'source': self.source_var.get(),
            'completely_matched': getattr(self, 'comp_var', None).get() if hasattr(self, 'comp_var') else False,
            'custom_path': self.custom_path.get() if self.source_var.get() == "Custom" else None,
            'mode': 'standard'
        }
        self.callback(res); self.destroy()

    def _on_fast_fetch(self):
        res = {
            'seller': self.seller_var.get(), 
            'vendor_code': self.v_code_var.get().strip(),
            'shipment_id': self.s_id_var.get().strip(),
            'source': self.source_var.get(),
            'completely_matched': getattr(self, 'comp_var', None).get() if hasattr(self, 'comp_var') else False,
            'custom_path': self.custom_path.get() if self.source_var.get() == "Custom" else None,
            'mode': 'fast'
        }
        self.callback(res); self.destroy()

class MFIToolApp:
    def __init__(self):

        if not check_activation():
            root = tk.Tk(); root.withdraw()
            from tkinter import messagebox
            import sys
            messagebox.showerror("ACCESS DENIED", "ADMIN AUTHORIZATION REQUIRED\n\nThis version of the MFI Tool is currently inactive.\nPlease contact Mukesh (the administrator) for permission.")
            sys.exit()

        self.root = tk.Tk(); self.root.title("Row IB Investigation Tool v7.7 SECURED | ROW IB")
        try: self.root.state('zoomed')
        except: self.root.attributes('-zoomed', True)
        self.root.minsize(900, 620); self.root.configure(bg="#0f0f1a")
        self.claims_path, self.rebni_path, self.inv_path, self.vendor_level_path, self.ciat_path, self.ticket_id, self.mode_var, self.ticket_type_var = tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar(value="auto"), tk.StringVar(value="PDTT")
        self.ai_key = self._load_ai_key()
        self.is_light_theme = False
        self._setup_global_context_menu()
        self.all_blocks, self.preview = [], None; self._build_ui()
        # v6.1: Global ticket-wide persistence — matches and processed survive across ASINs
        self.global_matches    = []
        self.global_processed  = set()
        self.active_manual_dlg = None
        self.active_pending_dlg = None
        self.active_cross_dlg = None
        self._cloud_busy = {"REBNI": False, "Invoice": False}
        self._cloud_progress = {"REBNI": 0, "Invoice": 0}
        self._pending_cross_po = []  # FIX 5: Always initialize — prevents AttributeError in _next_man/_handle_res

    def _setup_global_context_menu(self):
        """v6.2.5: Provides right-click Copy/Paste/Cut context menus globally across all inputs."""
        menu = tk.Menu(self.root, tearoff=0, bg="#1e1e3a", fg="#e0e0e0", font=("Segoe UI", 9))
        menu.add_command(label="📋 Copy", command=lambda: self.root.focus_get().event_generate("<<Copy>>"))
        menu.add_command(label="📌 Paste", command=lambda: self.root.focus_get().event_generate("<<Paste>>"))
        menu.add_separator()
        menu.add_command(label="✂ Cut", command=lambda: self.root.focus_get().event_generate("<<Cut>>"))
        
        def show_context(event):
            w = event.widget
            if isinstance(w, (tk.Entry, ttk.Entry, tk.Text, ttk.Combobox)):
                try:
                    w.focus_set()
                    menu.tk_popup(event.x_root, event.y_root)
                finally:
                    menu.grab_release()

        self.root.bind_class("Entry", "<Button-3>", show_context)
        self.root.bind_class("TEntry", "<Button-3>", show_context)
        self.root.bind_class("Text", "<Button-3>", show_context)
        self.root.bind_class("TCombobox", "<Button-3>", show_context)
        
        # v6.2.5: OTA Auto-Update Check on Startup
        check_for_updates(self.root)

    def _build_ui(self):
        t = tk.Frame(self.root, bg="#16213e", height=62); t.pack(fill="x")
        tk.Label(t, text="  MFI Investigation Tool", fg="#e94560", bg="#16213e", font=("Segoe UI", 20, "bold")).pack(side="left", padx=16, pady=12)
        tk.Label(t, text="Developed by Mukesh", fg="#4a9eff", bg="#16213e", font=("Segoe UI", 10, "italic")).pack(side="right", padx=6)
        # v7.2.0: About Button
        tk.Button(t, text="ⓘ About", bg="#16213e", fg="#4a9eff", font=("Segoe UI", 10, "bold"), 
                  relief="flat", borderwidth=0, activebackground="#16213e", activeforeground="#f0a500",
                  cursor="hand2", command=self.open_about_page).pack(side="right", padx=10)
        # --- v7.1.3: Theme Dropdown (Replaces swatches for cleaner UI) ---
        theme_frame = tk.Frame(t, bg="#16213e"); theme_frame.pack(side="right", padx=12)
        tk.Label(theme_frame, text="Theme Palette:", fg="#cccccc", bg="#16213e", font=("Segoe UI", 9)).pack(side="left", padx=4)
        
        self.theme_var = tk.StringVar(value=GLOBAL_THEME_NAME)
        self.theme_cb = ttk.Combobox(theme_frame, textvariable=self.theme_var, 
                                     values=list(THEME_PALETTES.keys()), 
                                     state="readonly", width=18, font=("Segoe UI", 9))
        self.theme_cb.pack(side="left", padx=4)
        self.theme_cb.bind("<<ComboboxSelected>>", lambda e: self._apply_theme(self.theme_var.get()))

        tk.Label(t, text="v7.7 SECURED | ROW IB", fg="#f0a500", bg="#16213e", font=("Segoe UI", 10, "bold")).pack(side="right", padx=15)
        
        # --- v7.1.3: Breadcrumbs bar for path tracking ---
        bc_f = tk.Frame(self.root, bg="#1a1a2e", height=35)
        self.breadcrumb_lbl = tk.Label(bc_f, text=" Investigation Path: Ready ", fg="#88ccff", bg="#1a1a2e", font=("Segoe UI", 10, "bold"), padx=12)
        self.breadcrumb_lbl.pack(side="left")

        leg = tk.Frame(self.root, bg="#1a1a2e", height=30)
        for tx, f, b in [("Claiming","white","#0f0f1a"),("Dominant","black","#E2EFDA"),("Sub-rows","black","#EBF3FB"),("Root/Short","#9C0006","#FFE0E0"),("DICES","black","#FFF2CC"),("SR","black","#FFD7D7"),("Invalid inv","#333","#FFD0D0"),("REBNI","#333","#D0F0FF"),("Cross PO","#7a5c00","#FFF0C0"),("Mismatch","#333","#D0E8FF")]:
            tk.Label(leg, text=f"  {tx}  ", fg=f, bg=b, font=("Segoe UI", 8, "bold"), padx=8).pack(side="left", padx=3, pady=3)
        body = tk.Frame(self.root, bg="#0d0d1a", padx=24, pady=12); body.pack(fill="both", expand=True)
        inp = tk.LabelFrame(body, text="  Input Files  (Excel .xlsx or CSV .csv supported)  ", fg="#4a9eff", bg="#0d0d1a", font=("Segoe UI", 10, "bold"), padx=12, pady=8); inp.pack(fill="x", pady=6)
        self._f_row(inp, "Claims Sheet:", self.claims_path, 0); self._f_row(inp, "REBNI Result:", self.rebni_path, 1); self._f_row(inp, "Invoice Search:", self.inv_path, 2); self._f_row(inp, "Vendor Level Data:", self.vendor_level_path, 3); self._f_row(inp, "CIAT Receive Data (Optional):", self.ciat_path, 4)
        tf = tk.Frame(body, bg="#0f0f1a"); tf.pack(anchor="w", pady=4); tk.Label(tf, text="Ticket ID:", fg="white", bg="#0f0f1a", font=("Segoe UI", 11)).pack(side="left"); tk.Entry(tf, textvariable=self.ticket_id, width=28, font=("Segoe UI", 11), bg="#1e1e3a", fg="white", insertbackground="white", relief="flat").pack(side="left", padx=8)
        m = tk.LabelFrame(body, text="Investigation Mode", fg="white", bg="#0f0f1a", padx=10, pady=5); m.pack(fill="x", pady=8); tk.Radiobutton(m, text="AUTO  —  Automatic. SID popup when not found in REBNI.", variable=self.mode_var, value="auto", fg="white", bg="#0f0f1a", selectcolor="#16213e", font=("Segoe UI", 10)).pack(anchor="w", padx=10); tk.Radiobutton(m, text="MANUAL  —  One level at a time. Live preview. Parallel interaction enabled.", variable=self.mode_var, value="manual", fg="white", bg="#0f0f1a", selectcolor="#16213e", font=("Segoe UI", 10)).pack(anchor="w", padx=10)
        tt = tk.LabelFrame(body, text="Ticket Type", fg="white", bg="#0f0f1a", padx=10, pady=5); tt.pack(fill="x", pady=6); tk.Radiobutton(tt, text="PDTT  —  Full chain investigation across all shipments (default).", variable=self.ticket_type_var, value="PDTT", fg="white", bg="#0f0f1a", selectcolor="#16213e", font=("Segoe UI", 10)).pack(anchor="w", padx=10); tk.Radiobutton(tt, text="REMASH TT  —  Claiming shipment level investigation with Overage Tracing.", variable=self.ticket_type_var, value="REMASH", fg="#f0c060", bg="#0f0f1a", selectcolor="#1a1500", font=("Segoe UI", 10)).pack(anchor="w", padx=10)
        self.status = tk.Label(body, text="Ready", fg="#4a9eff", bg="#0f0f1a", font=("Segoe UI", 10)); self.status.pack(pady=(10, 0)); self.pb = ttk.Progressbar(body, mode='determinate'); self.pb.pack(fill="x", pady=4)
        ctrl_f = tk.Frame(body, bg="#0f0f1a"); ctrl_f.pack(pady=10)
        self.run_btn = tk.Button(ctrl_f, text="▶ RUN", bg="#e94560", fg="white", font=("Segoe UI", 9, "bold"), padx=10, pady=5, relief="flat", cursor="hand2", command=self.start_run); self.run_btn.pack(side="left", padx=3)
        self.stop_inv_btn = tk.Button(ctrl_f, text="⏸ STOP", bg="#4a2020", fg="white", font=("Segoe UI", 9, "bold"), padx=8, pady=5, relief="flat", state="disabled", cursor="hand2", command=self.request_stop_investigation); self.stop_inv_btn.pack(side="left", padx=12)
        self.vault_btn = tk.Button(ctrl_f, text="📦  VIEW CROSS PO VAULT", command=self.show_vault,
                                  bg="#6a4c93", fg="white", font=("Segoe UI", 10, "bold"),
                                  padx=20, pady=8, relief="flat", cursor="hand2")
        self.vault_btn.pack(side="left", padx=12)
        bf = tk.Frame(body, bg="#0f0f1a"); bf.pack(pady=5)
        self.stop_sess_btn = tk.Button(bf, text="⏹ SESSION", bg="#3a0000", fg="white", font=("Segoe UI", 9, "bold"), padx=8, pady=5, relief="flat", state="disabled", cursor="hand2", command=self.request_stop_session); self.stop_sess_btn.pack(side="left", padx=3)
        self.save_btn = tk.Button(bf, text="💾 SAVE", bg="#2d6a4f", fg="white", font=("Segoe UI", 9, "bold"), padx=8, pady=5, relief="flat", state="normal", cursor="hand2", command=self.save_output); self.save_btn.pack(side="left", padx=3)
        self.portal_btn = tk.Button(bf, text="📑 PORTAL", bg="#1c2c42", fg="#4a9eff", font=("Segoe UI", 9, "bold"), padx=8, pady=5, relief="flat", cursor="hand2", command=self.open_summary_portal); self.portal_btn.pack(side="left", padx=3)
        self.rebni_search_btn = tk.Button(bf, text="🔍 REBNI SEARCH", bg="#ffc107", fg="#000000", font=("Segoe UI", 9, "bold"), padx=8, pady=5, relief="flat", cursor="hand2", command=lambda: self.start_cloud_search("REBNI")); self.rebni_search_btn.pack(side="left", padx=3)
        self.inv_search_btn = tk.Button(bf, text="🔍 INVOICE SEARCH", bg="#00bcd4", fg="#000000", font=("Segoe UI", 9, "bold"), padx=8, pady=5, relief="flat", cursor="hand2", command=lambda: self.start_cloud_search("Invoice")); self.inv_search_btn.pack(side="left", padx=3)
        self.restore_btn = tk.Button(bf, text="🛠️ RESTORE", bg="#333333", fg="#aaaaaa", font=("Segoe UI", 9, "bold"), padx=8, pady=5, relief="flat", cursor="hand2", command=self.restore_panels); self.restore_btn.pack(side="left", padx=3)
        self.reset_btn = tk.Button(bf, text="🔄 RESET", bg="#6c757d", fg="white", font=("Segoe UI", 9, "bold"), padx=8, pady=5, relief="flat", cursor="hand2", command=self.reset_tool); self.reset_btn.pack(side="left", padx=3)

    def _f_row(self, p, l, v, r):
        tk.Label(p, text=l, fg="#cccccc", bg="#131320", width=18, anchor="w", font=("Segoe UI", 10)).grid(row=r, column=0, sticky="w", pady=3)
        tk.Entry(p, textvariable=v, width=62, font=("Segoe UI", 10), bg="#1e1e3a", fg="white", insertbackground="white", relief="flat").grid(row=r, column=1, padx=6)
        tk.Button(p, text="Browse", command=lambda: self._browse(v), bg="#2d2d5e", fg="white", relief="flat", cursor="hand2", padx=8).grid(row=r, column=2)
        # v6.2.6: Direct Open button
        tk.Button(p, text="📂 Open", command=lambda: self._open_file(v), bg="#1a1a2e", fg="#00b4d8", font=("Segoe UI", 9, "bold"), relief="flat", cursor="hand2", padx=8).grid(row=r, column=3, padx=5)

    def _browse(self, var):
        f = filedialog.askopenfilename(filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv"), ("All Files", "*.*")])
        if f: var.set(f)

    def _open_file(self, var):
        path = var.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showwarning("Warning", "File path is empty or does not exist.")
            return
        try: os.startfile(path)
        except Exception as e: messagebox.showerror("Error", f"Could not open file: {e}")

    def show_vault(self):
        if not hasattr(self, 'engine'):
            messagebox.showinfo("Note", "Please start investigation first to initialize the engine."); return
        CrossPOVaultDialog(self.root, self.engine, self._start_vault_investigation)

    def _start_vault_investigation(self, sid, candidate, case, budget):
        if not hasattr(self, 'curr_m') or not self.curr_m:
            messagebox.showwarning("Warning", "Active investigation context required."); return
            
        c, bud = candidate, safe_num(budget)
        f_bc = c.get('found_bc', '')
        f_inv = c.get('found_inv', '')
        f_iq = c.get('inv_qty', 0)
        
        self._set_status(f"Vault Cross PO confirmed ({c['po']}) - starting manual investigation of {int(bud)} units.", None)
        claiming_po = self.curr_m.get('claiming_po', self.curr_m.get('p', ''))
        
        self.curr_m.update({'b':f_bc, 'i':f_inv, 's':c['sid'], 'p':c['po'], 'a':c['asin'], 'iq':f_iq, 'rem':bud, 'budget':bud, 'depth':self.curr_m.get('depth',0)+1, 'rendered':False, 'processed':self.curr_m.get('processed', set()), 'cross_po_checked':True, 'asin_rendered_levels':set(), 'claiming_po': claiming_po})
        threading.Thread(target=self._man_step, daemon=True).start()

    def _set_status(self, msg, pct=None):
        self.root.after(0, lambda: (self.status.config(text=msg), (self.pb.__setitem__('value', pct) if pct is not None else None)))

    def start_run(self):
        if not all([self.claims_path.get(), self.rebni_path.get(), self.inv_path.get()]):
            messagebox.showerror("Error", "Please select all 3 input files."); return
        self.run_btn.config(state="disabled"); self.save_btn.config(state="disabled"); self.portal_btn.config(state="disabled"); self.stop_inv_btn.config(state="normal"); self.stop_sess_btn.config(state="normal"); self.all_blocks = []
        if self.mode_var.get() == "manual":
            if not self.preview or not self.preview.winfo_exists(): self.preview = PreviewPanel(self.root); self.preview._app = self
            else: self.preview.clear_all()
        # Initialize engine attributes
        if hasattr(self, 'engine'):
            self.engine.stop_requested = False
            self.engine.pause_requested = False
            self.engine.ticket_type = self.ticket_type_var.get()
            self.engine.collected_rebni = {} # reset
        threading.Thread(target=self._process, daemon=True).start()

    def request_stop_investigation(self):
        if not hasattr(self, 'engine'): return
        if not self.engine.pause_requested:
            self.engine.pause_requested = True
            self._set_status("Investigation paused — current results preserved. Save or resume.")
            self.root.after(0, lambda: (
                self.stop_inv_btn.config(text="▶  RESUME", bg="#2d6a4f", fg="white"),
                self.save_btn.config(state="normal"),
                self.portal_btn.config(state="normal")
            ))
        else:
            self.engine.pause_requested = False
            self._set_status("Resuming investigation...")
            self.root.after(0, lambda: (
                self.stop_inv_btn.config(text="⏸  STOP", bg="#4a2020", fg="white"),
                self.save_btn.config(state="disabled")
            ))

    def request_stop_session(self):
        if hasattr(self, 'engine'): self.engine.stop_requested = True
        self._set_status("Session ended — saving current results."); self._finish()

    def reset_tool(self):
        # v6.1: COMPLETE reset — wipe ALL memory, paths, engine state, and UI panels
        # 1) Wipe global investigation memory (fixes cross-ASIN leakage)
        self.global_matches   = []
        self.global_processed = set()
        self.all_blocks       = []
        # 2) Wipe all file paths so no old data carries to next ticket
        self.ticket_id.set("")
        self.claims_path.set("")
        self.rebni_path.set("")
        self.inv_path.set("")
        self.vendor_level_path.set("") # v7.0: Clear Vendor Level on reset
        # 3) Wipe engine caches completely
        if hasattr(self, 'engine'):
            self.engine.user_overrides = {}
            self.engine.collected_rebni = {}
            self.engine.cache_sid = {}
            self.engine.cache_bc  = {}
            self.engine.loop_cache = {}
            self.engine.stop_requested = False
            self.engine.pause_requested = False
            self.engine.all_found_matches = []
        # 4) Destroy all open panels
        if self.preview and self.preview.winfo_exists():
            self.preview.destroy()
        self.preview = None
        if getattr(self, 'active_manual_dlg', None) and self.active_manual_dlg.winfo_exists():
            self.active_manual_dlg.destroy()
        self.active_manual_dlg = None
        if getattr(self, 'active_pending_dlg', None) and self.active_pending_dlg.winfo_exists():
            self.active_pending_dlg.destroy()
        self.active_pending_dlg = None
        # 5) Reset curr_m context if it exists
        if hasattr(self, 'curr_m'):
            del self.curr_m
        if hasattr(self, 'manual_q'):
            self.manual_q = []
        if hasattr(self, '_pending_cross_po'):
            self._pending_cross_po = []
        # 6) Reset button states
        self.rebni_search_btn.config(state="normal")
        self.inv_search_btn.config(state="normal")
        self.run_btn.config(state="normal")
        self.stop_inv_btn.config(state="disabled")
        self.stop_sess_btn.config(state="disabled")
        self._cloud_busy = {"REBNI": False, "Invoice": False}
        self._set_status("Ready — tool fully reset for next ticket.", 0)
        messagebox.showinfo("Reset", "Tool reset completely. All memory, paths, and cached data have been cleared.")

    def _update_cloud_ui_state(self):
        # Selective interlock: Only disable the active search type
        r_busy = self._cloud_busy.get("REBNI", False)
        i_busy = self._cloud_busy.get("Invoice", False)
        
        self.rebni_search_btn.config(state="disabled" if r_busy else "normal")
        self.inv_search_btn.config(state="disabled" if i_busy else "normal")
        
        # Disable RUN button if ANY search is busy to prevent investigation with partial/in-progress data
        self.run_btn.config(state="disabled" if (r_busy or i_busy) else "normal")

    def start_cloud_search(self, s_type):
        SearchCloudDialog(self.root, s_type, lambda res: self.run_cloud_query(s_type, res))

    def run_cloud_query(self, s_type, params):
        if not params: return
        self._cloud_busy[s_type] = True
        self._update_cloud_ui_state()
        
        if params.get('mode') == 'fast':
            threading.Thread(target=self._exec_fast_fetch, args=(s_type, params), daemon=True).start()
        else:
            threading.Thread(target=self._exec_cloud_query, args=(s_type, params), daemon=True).start()

    def _update_concurrent_status(self):
        msg_parts = []
        for st in ["REBNI", "Invoice"]:
            if self._cloud_busy[st]:
                msg_parts.append(f"[{st}] Scanning... {self._cloud_progress[st]:,} rows")
        
        if msg_parts:
            self._set_status(" | ".join(msg_parts))
        else:
            self._set_status("Cloud search complete.")

    def _exec_cloud_query(self, s_type, params):
        try:
            seller, v_code = params['seller'], params['vendor_code']
            s_id = params.get('shipment_id', '').strip()
            v_codes_list = [v.strip() for v in v_code.replace(' ', ',').split(',') if v.strip()]
            comp_match = params.get('completely_matched', False)
            source = params.get('source', 'Local')
            custom_path = params.get('custom_path')
            
            f_map = {
                'Invoice': {"VRP":"vrp.txt", "RK World":"rk.txt", "KayKay":"kaykay.txt", "Clicktech":"Clicktech.txt", "Dawntech":"Dawntech.txt", "Etrade":"Etrade.txt", "Cocoblu":"Cocoblu.txt", "Retail EZ":"RetailEZ.txt"},
                'REBNI': {"VRP":"vrp.txt", "RK World":"rk.txt", "KayKay":"kk.txt", "Clicktech":"clicktech.txt", "Dawntech":"dawntech.txt", "Etrade":"etrade.txt", "Cocoblu":"cocoblu.txt", "Retail EZ":"retailEZ.txt"}
            }
            fname = f_map[s_type].get(seller)
            if comp_match and s_type == "REBNI": fname = "Available.txt"

            if source == "Custom":
                target_file = custom_path
            elif source == "Cloud":
                root_dir = r"W:\Shared With Me\PD Invoice & Rebni data"
                if not os.path.exists(root_dir):
                    self.root.after(0, lambda: messagebox.showerror("Drive Error", f"Cloud directory (W:) not found.\nPlease ensure the drive is correctly mapped.")); return
                target_file = os.path.join(root_dir, s_type, fname)
            else: # Local
                target_file = os.path.join(os.path.expanduser("~"), "Downloads", "PD App", s_type, fname)

            if not target_file or not os.path.exists(target_file):
                self.root.after(0, lambda: (
                    self._cloud_busy.update({s_type: False}),
                    self._update_cloud_ui_state(),
                    messagebox.showerror("Error", f"File not found at source ({source}):\n{target_file}")
                )); return
            
            ticket_id = self.ticket_id.get().strip()
            filtered_chunks = []
            chunk_size = 200000 # v6.2.5: Increased chunk size for bulk throughput
            total_rows_processed = 0
            
            reader = pd.read_csv(target_file, sep='\t', low_memory=False, chunksize=chunk_size, on_bad_lines='skip')
            for chunk in reader:
                total_rows_processed += len(chunk)
                self._cloud_progress[s_type] = total_rows_processed
                self.root.after(0, self._update_concurrent_status)
                
                filtered = chunk.copy()
                if ticket_id:
                    col = 'purchase_order_id' if 'purchase_order_id' in chunk.columns else 'po' if 'po' in chunk.columns else None
                    if col: filtered = filtered[filtered[col].astype(str).str.contains(ticket_id, na=False)]
                
                if s_id:
                    s_aliases = ['shipment_id', 'sid', 'shipment id', 'shipment no', 'shipment number', 'ship id', 'fba shipment id', 'inbound shipment', 'shipment', 'inbound sid']
                    s_col = next((c for c in chunk.columns if str(c).lower().strip() in s_aliases), None)
                    if s_col: 
                        filtered = filtered[filtered[s_col].astype(str).str.contains(s_id, na=False, case=False)]
                elif v_codes_list:
                    col = 'vendor_code' if 'vendor_code' in chunk.columns else None
                    if col: filtered = filtered[filtered[col].astype(str).str.strip().isin(v_codes_list)]
                
                if not filtered.empty:
                    filtered_chunks.append(filtered)
            
            search_term = s_id if s_id else (v_codes_list[0] if v_codes_list else "")
            self._finalize_cloud_results(s_type, source, search_term, filtered_chunks, ticket_id)
            
        except Exception as e:
            self.root.after(0, lambda: (
                self._cloud_busy.update({s_type: False}),
                self._update_cloud_ui_state(),
                self._update_concurrent_status(),
                messagebox.showerror("Search Error", str(e))
            ))

    def _exec_fast_fetch(self, s_type, params):
        # v6.2.5: High-speed vendor-only equality check for bulk datasets
        try:
            seller, v_code = params['seller'], params['vendor_code']
            s_id = params.get('shipment_id', '').strip()
            v_codes_list = [v.strip() for v in v_code.replace(' ', ',').split(',') if v.strip()]
            source = params.get('source', 'Local')
            custom_path = params.get('custom_path')
            
            f_map = {
                'Invoice': {"VRP":"vrp.txt", "RK World":"rk.txt", "KayKay":"kaykay.txt", "Clicktech":"Clicktech.txt", "Dawntech":"Dawntech.txt", "Etrade":"Etrade.txt", "Cocoblu":"Cocoblu.txt", "Retail EZ":"RetailEZ.txt"},
                'REBNI': {"VRP":"vrp.txt", "RK World":"rk.txt", "KayKay":"kk.txt", "Clicktech":"clicktech.txt", "Dawntech":"dawntech.txt", "Etrade":"etrade.txt", "Cocoblu":"cocoblu.txt", "Retail EZ":"retailEZ.txt"}
            }
            fname = f_map[s_type].get(seller)
            if not fname: fname = "Available.txt" # fallback

            if source == "Custom": target_file = custom_path
            elif source == "Cloud": target_file = os.path.join(r"W:\Shared With Me\PD Invoice & Rebni data", s_type, fname)
            else: target_file = os.path.join(os.path.expanduser("~"), "Downloads", "PD App", s_type, fname)

            if not target_file or not os.path.exists(target_file):
                self.root.after(0, lambda: (
                    self._cloud_busy.update({s_type: False}),
                    self._update_cloud_ui_state(),
                    messagebox.showerror("Error", f"Fast Fetch source not found: {target_file}")
                )); return

            self._set_status(f"Fast Fetch: Connecting to {source}...")
            filtered_chunks = []
            chunk_size = 200000
            search_term = s_id if s_id else v_code
            total_rows = 0
            
            # High-speed line-based filtering (avoids Pandas overhead for massive files)
            with open(target_file, 'r', encoding='utf-8', errors='ignore') as f_in:
                header = f_in.readline()
                if not header:
                    raise ValueError("Target file is empty.")
                
                cols = header.strip().split('\t')
                try:
                    s_aliases = ['shipment_id', 'sid', 'shipment id', 'shipment no', 'shipment number', 'ship id', 'fba shipment id', 'inbound shipment', 'shipment', 'inbound sid']
                    if s_id:
                        v_idx = next((i for i, c in enumerate(cols) if str(c).lower().strip() in s_aliases), None)
                        if v_idx is None: raise ValueError("Shipment ID column not found")
                    else:
                        v_idx = cols.index('vendor_code')
                except ValueError:
                    # Fallback to slower pandas if column not found easily
                    f_in.seek(0)
                    reader = pd.read_csv(target_file, sep='\t', low_memory=False, chunksize=chunk_size, on_bad_lines='skip')
                    for chunk in reader:
                        total_rows += len(chunk)
                        self._cloud_progress[s_type] = total_rows
                        self.root.after(0, self._update_concurrent_status)
                        if s_id:
                            s_col = next((c for c in chunk.columns if str(c).lower().strip() in s_aliases), None)
                            if s_col:
                                filtered = chunk[chunk[s_col].astype(str).str.contains(s_id, na=False, case=False)]
                                if not filtered.empty: filtered_chunks.append(filtered)
                        elif 'vendor_code' in chunk.columns:
                            filtered = chunk[chunk['vendor_code'].astype(str).str.strip().isin(v_codes_list)]
                            if not filtered.empty: filtered_chunks.append(filtered)
                    search_term = s_id if s_id else (v_codes_list[0] if v_codes_list else "")
                    self._finalize_cloud_results(s_type, source, search_term, filtered_chunks, "ALL")
                    return

                # Fast line-by-line matching
                matching_lines = [header]
                for line in f_in:
                    total_rows += 1
                    if total_rows % 100000 == 0:
                        self._cloud_progress[s_type] = total_rows
                        self.root.after(0, lambda tr=total_rows: (
                            self._set_status(f"Fast Fetch: Scanned {tr:,} lines..."),
                            self._update_concurrent_status()
                        ))
                    
                    parts = line.split('\t')
                    if len(parts) > v_idx:
                        if s_id:
                            if s_id.lower() in parts[v_idx].strip().lower():
                                matching_lines.append(line)
                        elif parts[v_idx].strip() in v_codes_list:
                            matching_lines.append(line)
                
                if len(matching_lines) > 1:
                    # Convert only the matching lines to a DF for finalizing
                    from io import StringIO
                    final_df = pd.read_csv(StringIO("".join(matching_lines)), sep='\t')
                    filtered_chunks = [final_df]

            search_term = s_id if s_id else (v_codes_list[0] if v_codes_list else "")
            self._finalize_cloud_results(s_type, source, search_term, filtered_chunks, "ALL")

        except Exception as e:
            self.root.after(0, lambda: (
                self._cloud_busy.update({s_type: False}),
                self._update_cloud_ui_state(),
                messagebox.showerror("Fast Fetch Error", str(e))
            ))

    def _finalize_cloud_results(self, s_type, source, v_code, filtered_chunks, ticket_id):
        if not filtered_chunks:
            self.root.after(0, lambda: (
                self._cloud_busy.update({s_type: False}),
                self._update_cloud_ui_state(),
                self._update_concurrent_status(),
                messagebox.showinfo("No Results", f"No matches found in {source} for PO: {ticket_id} and Vendor: {v_code}")
            )); return
        
        final_df = pd.concat(filtered_chunks)
        res_dir = os.path.join(os.path.expanduser("~"), "Downloads")
        v_suffix = f" - {v_code}" if v_code else ""
        res_name = f"{s_type}_Search_Result{v_suffix}.csv"
        res_path = os.path.join(res_dir, res_name)
        final_df.to_csv(res_path, index=False)
        
        self.root.after(0, lambda: (
            self.rebni_path.set(res_path) if s_type == "REBNI" else self.inv_path.set(res_path),
            self._cloud_busy.update({s_type: False}),
            self._update_cloud_ui_state(),
            self._update_concurrent_status(),
            messagebox.showinfo("Success", f"{s_type} extraction ({source}) complete.\nSaved to: {res_path}\nMain path auto-populated.")
        ))
            
    def _apply_theme(self, theme_name):
        global GLOBAL_THEME_NAME
        GLOBAL_THEME_NAME = theme_name
        self.is_light_theme = ("Light" in theme_name or "Pastel" in theme_name or "Cloud" in theme_name or "Sand" in theme_name)
        apply_global_theme_to_widget(self.root)
        if self.preview and self.preview.winfo_exists():
            apply_global_theme_to_widget(self.preview)


    def _toggle_theme(self):
        # Legacy compatibility: cycle between Dark and Light
        if GLOBAL_THEME_NAME == "Dark Mode (Default)":
            self._apply_theme("Light Mode")
        else:
            self._apply_theme("Dark Mode (Default)")

    # --- AI INTEGRATION METHODS ---
    def _load_ai_key(self):
        try:
            # FIX 8: Use stable user home dir — prevents key loss when CWD changes
            _cfg = os.path.join(os.path.expanduser("~"), ".mfi_tool", "config.json")
            if os.path.exists(_cfg):
                with open(_cfg, "r") as f:
                    return json.load(f).get("gemini_api_key", "")
        except: pass
        return ""

    def _save_ai_key(self, key):
        try:
            # FIX 8: Use stable user home dir for config
            _cfg_dir = os.path.join(os.path.expanduser("~"), ".mfi_tool")
            os.makedirs(_cfg_dir, exist_ok=True)
            _cfg = os.path.join(_cfg_dir, "config.json")
            cfg = {}
            if os.path.exists(_cfg):
                with open(_cfg, "r") as f: cfg = json.load(f)
            cfg["gemini_api_key"] = key
            with open(_cfg, "w") as f: json.dump(cfg, f)
            self.ai_key = key
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save config: {e}")

    def _process(self):
        try:
            self._set_status("Loading Claims file…", 5); df_c = load_claims(self.claims_path.get()); self.raw_claims_df = df_c; mc, corr = detect_claim_cols(df_c)
            if corr or [f for f in COLUMN_ALIASES if f not in mc]:
                conf, done = [None], threading.Event(); self.root.after(0, lambda: HeaderCorrectionDialog(self.root, corr, mc, list(df_c.columns), lambda res: (conf.__setitem__(0, res['mapping']) if res['action'] == 'proceed' else None, done.set()))); done.wait()
                if conf[0] is None: self.root.after(0, lambda: (self.run_btn.config(state="normal"), self.stop_inv_btn.config(state="disabled"), self.stop_sess_btn.config(state="disabled"))); return
                mc = conf[0]
            
            # v7.0: Vendor Level Data Integration
            vendor_idx, vendor_raw, vendor_headers = {}, {}, []
            if self.vendor_level_path.get().strip():
                self._set_status("Loading Vendor Level Data…", 8)
                vendor_idx, vendor_raw, vendor_headers = build_vendor_index(load_vendor_data(self.vendor_level_path.get()))

            self._set_status("Loading REBNI…", 12); _rebni_df = load_rebni(self.rebni_path.get()); self.raw_rebni_df = _rebni_df; rp, rs, rfb, rsid = build_rebni_index(_rebni_df); self._set_status("Loading Invoice Search…", 30); _inv_df = load_invoice_search(self.inv_path.get()); self.inv_search_df = _inv_df; ip, ifb, iam = build_invoice_index(_inv_df)
            
            self.engine = InvestigationEngine(rp, rs, rfb, rsid, ip, ifb, iam, self._req_sid, vendor_idx=vendor_idx, vendor_raw=vendor_raw, vendor_headers=vendor_headers)
            self.engine.ticket_type = self.ticket_type_var.get()
            tot = len(df_c)
            if self.mode_var.get() == "auto":
                self.preview = PreviewPanel(self.root); self.preview._app = self
                for i, (_, r) in enumerate(df_c.iterrows()):
                    if self.engine.stop_requested: break
                    self.engine.asin_pending_matches = []
                    self.global_processed.clear()
                    if self.engine: self.engine.loop_cache.clear()
                    self._set_status(f"Auto: {i+1}/{tot}  ASIN: {clean(r.get(mc.get('ASIN',''),''))}", 60 + int((i / max(tot, 1)) * 35)); self.preview.add_header_row(f"{i+1}/{tot}: {clean(r.get(mc.get('ASIN',''),''))}")
                    rows, _ = self.engine.run_auto(clean(r.get(mc.get('Barcode', ''), '')), clean(r.get(mc.get('Invoice', ''), '')), extract_sid(clean(r.get(mc.get('SID', ''), ''))), clean(r.get(mc.get('PO', ''), '')), clean(r.get(mc.get('ASIN', ''), '')), safe_num(r.get(mc.get('InvQty', ''), 0)), safe_num(r.get(mc.get('PQV', ''), 0)), initial_cp=safe_num(r.get(mc.get('CP', ''), 0)), row_callback=lambda row: self.root.after(0, lambda: (self.preview.add_row(row) if self.preview and self.preview.winfo_exists() else None)), visited=self.global_processed)
                    self.all_blocks.append(rows)
                self._finish()
            else: self.manual_q, self.map_cols = df_c.to_dict('records'), mc; self._next_man()
        except Exception as e:
            import traceback; tb = traceback.format_exc(); err_msg = str(e)
            self.root.after(0, lambda msg=err_msg: messagebox.showerror("Error", f"{msg}\n\n{tb}"))
            self._finish()

    def _req_sid(self, inv, po, asin):
        if inv in self.engine.cache_sid: return self.engine.cache_sid[inv]
        res, done = [None], threading.Event(); self.root.after(0, lambda: SIDRequestDialog(self.root, inv, po, asin, lambda s: (self.engine.cache_sid.__setitem__(inv, s) if s else None, res.__setitem__(0, s), done.set()))); done.wait(); return res[0]

    def restore_panels(self):
        options = []
        if self.preview and self.preview.winfo_exists() and self.preview.state() in ("withdrawn", "iconic"):
            options.append('Main Preview Panel')
        if getattr(self, 'active_manual_dlg', None) and self.active_manual_dlg.winfo_exists() and self.active_manual_dlg.state() in ("withdrawn", "iconic"):
            options.append('Manual Investigation Dialog')
        if getattr(self, 'active_pending_dlg', None) and self.active_pending_dlg.winfo_exists() and self.active_pending_dlg.state() in ("withdrawn", "iconic"):
            options.append('Pending Invoices Context')
        if getattr(self, 'active_cross_dlg', None) and self.active_cross_dlg.winfo_exists() and self.active_cross_dlg.state() in ("withdrawn", "iconic"):
            options.append('Cross PO Detection Dialog')
        if not options:
            if self.all_blocks:
                self.preview = PreviewPanel(self.root); self.preview._app = self
                for block in self.all_blocks:
                    self.preview.add_header_row('Restored Data')
                    for row in block: self.preview.add_row(row)
                self._set_status('Preview panel restored.')
                return
            self._set_status('No panels to restore.'); return
        if len(options) == 1: self._do_restore(options[0]); return
        dlg = tk.Toplevel(self.root)
        dlg.title('Restore Panel'); dlg.geometry('380x220'); dlg.configure(bg='#16213e'); dlg.resizable(False, False)
        tk.Label(dlg, text='Select panel to restore:', bg='#16213e', fg='white', font=('Segoe UI', 11, 'bold')).pack(pady=(15, 8))
        v = tk.StringVar(value=options[0])
        ttk.Combobox(dlg, textvariable=v, values=options, state='readonly', width=38, font=('Segoe UI', 10)).pack(pady=8)
        tk.Button(dlg, text='✔  Restore Selected', command=lambda: (self._do_restore(v.get()), dlg.destroy()), bg='#2d6a4f', fg='white', font=('Segoe UI', 10, 'bold'), padx=16, pady=8, relief='flat').pack(pady=12)
        dlg.update_idletasks()
        px = self.root.winfo_x() + (self.root.winfo_width() - dlg.winfo_width()) // 2
        py = self.root.winfo_y() + (self.root.winfo_height() - dlg.winfo_height()) // 2
        dlg.geometry(f'+{px}+{py}')

    def _do_restore(self, panel_name):
        if panel_name == 'Main Preview Panel' and getattr(self, 'preview', None) and self.preview.winfo_exists():
            self.preview.deiconify(); self.preview.lift(); self.preview.focus_force()
        elif panel_name == 'Manual Investigation Dialog' and getattr(self, 'active_manual_dlg', None) and self.active_manual_dlg.winfo_exists():
            self.active_manual_dlg.deiconify(); self.active_manual_dlg.lift(); self.active_manual_dlg.focus_force()
        elif panel_name == 'Pending Invoices Context' and getattr(self, 'active_pending_dlg', None) and self.active_pending_dlg.winfo_exists():
            self.active_pending_dlg.deiconify(); self.active_pending_dlg.lift(); self.active_pending_dlg.focus_force()
        elif panel_name == 'Cross PO Detection Dialog' and getattr(self, 'active_cross_dlg', None) and self.active_cross_dlg.winfo_exists():
            self.active_cross_dlg.deiconify(); self.active_cross_dlg.lift(); self.active_cross_dlg.focus_force()

    def _next_man(self, force_next=False):
        if hasattr(self, '_inv_proc_thread') and self._inv_proc_thread.is_alive():
            self._set_status("Waiting for background processes...")
            self.root.after(1000, lambda: self._next_man(force_next))
            return

        # v6.2.6: Prioritize Pending Cross POs (LIFO) before moving to next ASIN
        if hasattr(self, '_pending_cross_po') and self._pending_cross_po:
            p = self._pending_cross_po.pop()
            self._set_status(f"Starting Cross PO: {p['candidate']['po']}")
            def run_sub():
                cr, fn = self.engine.run_cross_po_investigation(p['candidate'], p['case'], p['budget'], depth=self.curr_m['depth']+1)
                for r in cr:
                    self.curr_m['block'].append(r)
                    if self.preview and self.preview.winfo_exists(): self.root.after(0, lambda row=r: self.preview.add_row(row))
                self._man_step()
            threading.Thread(target=run_sub, daemon=True).start(); return

        if hasattr(self, 'curr_m') and not force_next:
            if self._collect_all_pending(): self.root.after(0, self._show_pending_gateway); return
            if self.curr_m.get('block'): self.all_blocks.append(self.curr_m['block'])
        
        if not self.manual_q or self.engine.stop_requested: self._finish(); return
        
        # v6.2.5/6.2.6: Reset isolation ONLY when truly starting a NEW claiming ASIN
        self.engine.asin_pending_matches = []
        self.global_processed.clear()

        r, mc = self.manual_q.pop(0), self.map_cols
        self.curr_m = {'b':clean(r.get(mc.get('Barcode',''),'')), 'i':clean(r.get(mc.get('Invoice',''),'')), 's':extract_sid(clean(r.get(mc.get('SID',''),''))), 'p':clean(r.get(mc.get('PO',''),'')), 'a':clean(r.get(mc.get('ASIN',''),'')), 'iq':safe_num(r.get(mc.get('InvQty',''),0)), 'pqv':safe_num(r.get(mc.get('PQV',''),0)), 'rem':safe_num(r.get(mc.get('PQV',''),0)), 'budget':safe_num(r.get(mc.get('PQV',''),0)), 'depth':0, 'block':[], 'processed':self.global_processed, 'all_seen_matches':self.global_matches, 'asin_rendered_levels':set(), 'is_new_block':True, 'rendered':False, 'siblings_stack':[], 'initial_cp':safe_num(r.get(mc.get('CP',''),0))}
        # v6.1: Safety — recreate preview if it was destroyed (e.g. after reset)
        if not self.preview or not self.preview.winfo_exists():
            self.preview = PreviewPanel(self.root); self.preview._app = self
        self.preview.add_header_row(self.curr_m['a']); threading.Thread(target=self._man_step, daemon=True).start()

    def _man_step(self):
        if self.engine.stop_requested: self._finish(); return
        m = self.curr_m
        # v7.2.3: Build exclude_pos from claiming_po so ancestor POs are never re-detected as cross POs
        _man_excl = {clean(m['p'])}
        if m.get('claiming_po'): _man_excl.add(clean(m['claiming_po']))
        rows, matches, rq, n_rem, ex = self.engine.build_one_level(m['b'], m['i'], m['s'], m['p'], m['a'], m['iq'], m['rem'], m['depth'], is_claiming=(m['depth']==0), is_manual=True, initial_cp=m.get('initial_cp',0.0), exclude_pos=_man_excl)
        k = (m['depth'], clean(m['s']), clean(m['p']), clean(m['a']), clean(m['i']))
        if not m['rendered'] and k not in m['asin_rendered_levels']:
            if m.get('is_new_block') and rows: rows[0]['is_new_block']=True; m['is_new_block']=False
            m['block'].extend(rows); self.root.after(0, lambda: [self.preview.add_row(r) for r in rows]); m['asin_rendered_levels'].add(k); m['rendered']=True
        elif not m['rendered']: self.root.after(0, lambda: [self.preview.add_row(r) for r in rows[1:]]); m['rendered']=True
        # v7.3.5: Skip Cross PO check if current PO has no shortage (Shortage <= 0)
        shortage = safe_num(m['iq']) - safe_num(rq)
        if shortage > 0 and not m.get('cross_po_checked'):
            m['cross_po_checked'] = True
            cands = self.engine.detect_cross_po(m['s'], m['p'], m['a'], exclude_pos={clean(m['p'])})
            
            # v7.3.5: Filter out Case 0 (Non-overages) from the candidates
            if cands:
                cands = [c for c in cands if 'Case 0' not in c.get('cross_type', '')]
            
            if cands: 
                m['_awaiting_cross_po'] = True; 
                def create_cross():
                    if getattr(self, 'active_cross_dlg', None) and self.active_cross_dlg.winfo_exists(): 
                        self.active_cross_dlg.destroy()
                    self.active_cross_dlg = CrossPODialog(self.root, cands, m['i'], m['s'], lambda res: self._handle_cross_po_and_finish(res), engine=self.engine)
                self.root.after(0, create_cross)
                return
        pm = []
        seen_m_keys = set()
        for mt in matches:
            aug = dict(mt); aug['_depth'] = m.get('depth',0); m.setdefault('all_seen_matches',[]).append(aug)
            
            # v6.2.6: Deduplicate by (Inv, ASIN, PO, Qty) to allow separate branches for same invoice with diff context.
            uk = (clean(mt.get('mtc_inv', '')), clean(mt.get('mtc_asin', '')), clean(mt.get('mtc_po', '')), fmt_qty(mt.get('mtc_qty',0)))
            if uk in seen_m_keys: continue
            seen_m_keys.add(uk)
            
            pm.append(mt)
        matches = pm
        
        m['rem']=n_rem; matches=[x for x in matches if self._get_loop_key(x) not in m['processed']]; rem_s=rows[0].get('remarks','') if rows else ''
        if not matches: rem_s="No unprocessed matches remaining"
        
        # v7.1.3: Update Breadcrumbs Path Tracking
        self._update_breadcrumbs()

        if not matches or any(kw in rem_s for kw in ["Root cause", "REBNI", "SR", "short received directly", "Direct Shortage", "Phase 1", "No Invoice Search"]):
            # v7.1.5: Removed automatic siblings_stack popping and confirmation popups.
            # User now has full control via the Pending Invoices Gateway.
            if hasattr(self, '_pending_cross_po') and self._pending_cross_po: self.root.after(0, self._next_man); return
            self.root.after(0, self._next_man); return

        self.root.after(0, lambda: self._show_dlg(matches))

    def _show_dlg(self, matches):
        m, f = self.curr_m, matches[0]
        if self.engine.user_overrides:
            pm = []
            for mt in matches:
                uk = (clean(mt.get('mtc_inv', '')), clean(mt.get('mtc_asin', '')), clean(mt.get('mtc_po', '')))
                ov = self.engine.user_overrides.get(uk, {})
                if not ov: ov = self.engine.user_overrides.get(clean(mt.get('mtc_inv', '')), {})
                if ov:
                    mt=dict(mt)
                    if 'mtc_qty' in ov: mt['mtc_qty']=ov['mtc_qty']
                    if 'inv_qty' in ov: mt['inv_qty']=ov['inv_qty']
                pm.append(mt)
            matches, f = pm, pm[0]
        if f['mtc_inv'] in self.engine.cache_sid: self._handle_res({'action':'valid','chosen_match':f,'sid':self.engine.cache_sid[f['mtc_inv']],'barcode':self.engine.cache_bc.get(f['mtc_inv'],"[DICES]")}, matches); return
        if getattr(self, 'active_manual_dlg', None) and self.active_manual_dlg.winfo_exists(): self.active_manual_dlg.destroy()
        self.active_manual_dlg = ManualLevelDialog(self.root, matches, m['rem'], m['budget'], lambda rs: self._handle_res(rs, matches), pending_cb=self._show_pending_invoices_from_dialog, engine=self.engine)

    def _handle_res(self, res, matches):
        if not res or res['action'] == 'stop': self.root.after(0, self._next_man); return
        mc = res.get('chosen_match')
        if mc: self.curr_m['processed'].add(self._get_loop_key(mc))
        if res['action'] == 'invalid':
            ex = res['invalid_qty']; ro = self.engine._make_row('[INVALID]', mc['mtc_inv'], '—', mc['mtc_po'], mc['mtc_asin'], mc['inv_qty'], '', '', '', f"{int(ex)} units matched to invalid invoice {mc['mtc_inv']} — excluded from PQV", '', self.curr_m['depth'], 'subrow')
            self.curr_m['block'].append(ro); self.preview.add_row(ro); self.curr_m['rem']=max(0, self.curr_m['rem']-ex)
            if self.curr_m['rem']<=0: self.root.after(0, self._next_man)
            else:
                rm = [x for x in matches if x['mtc_inv'] != mc['mtc_inv']]
                if rm: self.root.after(0, lambda: self._show_dlg(rm))
                else: self.root.after(0, self._next_man)
        elif res['action'] == 'cross_po':
            cands = self.engine.detect_cross_po(self.curr_m['s'], self.curr_m['p'], self.curr_m['a'], exclude_pos={clean(self.curr_m['p'])})  # v7.2.3: exclude current PO
            if cands: 
                if getattr(self, 'active_cross_dlg', None) and self.active_cross_dlg.winfo_exists(): self.active_cross_dlg.destroy()
                self.active_cross_dlg = CrossPODialog(self.root, cands, self.curr_m['i'], self.curr_m['s'], lambda r: self._handle_cross_po(r), engine=self.engine)
            else: messagebox.showinfo("No Cross PO", "No Cross PO candidates found."); self.root.after(0, lambda: self._show_dlg(matches))
        elif res['action'] == 'mismatch':

            d = res['mismatch_data']; bug = safe_num(d.get('ovg_qty', 0)) or self.curr_m['rem']
            # v6.2.5 Fix: Signature mismatch (Added 'ex' to unpack 6 returns)
            rows, sm, rq, sh, nr, ex = self.engine.run_mismatch_investigation(d, bug, depth=self.curr_m['depth'])
            for r in rows: self.curr_m['block'].append(r) or self.preview.add_row(r)
            if sm: self.root.after(0, lambda: self._show_dlg(sm))
            else:
                rm = [x for x in matches if x != res.get('chosen_match')]
                if rm: self.root.after(0, lambda: self._show_dlg(rm))
                else: self.root.after(0, self._next_man)
        else:
            self.engine.cache_sid[mc['mtc_inv']], self.engine.cache_bc[mc['mtc_inv']] = res['sid'], res['barcode']
            # v6.2.6: Filter by full loop key, not just Invoice, to allow sibling branches on same invoice.
            rem_s = [x for x in matches if self._get_loop_key(x) != self._get_loop_key(mc) and self._get_loop_key(x) not in self.curr_m['processed']]
            ps = list(self.curr_m['siblings_stack'])
            if rem_s: ps.append({'siblings':rem_s, 'depth':self.curr_m['depth'], 'rem':self.curr_m['rem'], 'budget':self.curr_m['budget'], 'b':self.curr_m.get('b',''), 'i':self.curr_m.get('i',''), 'iq':self.curr_m.get('iq',0), 's':self.curr_m['s'], 'p':self.curr_m['p'], 'a':self.curr_m['a'], 'processed':self.curr_m['processed'], 'cross_po_checked':self.curr_m.get('cross_po_checked',False)})
            bb = safe_num(mc['mtc_qty']) or self.curr_m['rem']
            self.curr_m.update({'b':res['barcode'], 'i':mc['mtc_inv'], 's':res['sid'], 'p':mc['mtc_po'], 'a':mc['mtc_asin'], 'iq':mc['inv_qty'], 'rem':bb, 'budget':bb, 'depth':self.curr_m['depth']+1, 'rendered':False, 'processed':self.curr_m['processed'], 'siblings_stack':ps, 'pending_siblings':[], 'cross_po_checked':False, '_awaiting_cross_po':False})
            threading.Thread(target=self._man_step, daemon=True).start()

    def _update_breadcrumbs(self):
        """v7.1.3: Real-time path tracking label below the header."""
        if not hasattr(self, 'curr_m'): return
        m = self.curr_m
        path = f"ASIN: {m['a']} | PO: {m['p']} | Depth: {m['depth']} > Current INV: {m['i']}"
        txt = f" Investigation Path: {path} "
        self.root.after(0, lambda: self.breadcrumb_lbl.config(text=txt))



    def _collect_all_pending(self):
        proc, seen, ded = self.curr_m.get('processed', set()), set(), []
        # v6.2.5: Use Per-ASIN isolation (asin_pending_matches) to prevent cross-ASIN data persistence
        src = self.engine.asin_pending_matches if hasattr(self, 'engine') and hasattr(self.engine, 'asin_pending_matches') else self.curr_m.get('all_seen_matches', [])
        for inv in src:
            k = self._get_loop_key(inv)
            if k and k not in seen and k not in proc:
                seen.add(k)
                if hasattr(self, 'engine') and self.engine.user_overrides:
                    ov_k = (clean(inv.get('mtc_inv', '')), clean(inv.get('mtc_asin', '')), clean(inv.get('mtc_po', '')))
                    ov = self.engine.user_overrides.get(ov_k, {})
                    if not ov: ov = self.engine.user_overrides.get(clean(inv.get('mtc_inv','')), {})
                    if ov:
                        inv=dict(inv)
                        if 'mtc_qty' in ov: inv['mtc_qty']=ov['mtc_qty']
                        if 'inv_qty' in ov: inv['inv_qty']=ov['inv_qty']
                ded.append(inv)
        return ded

    def _show_pending_invoices_from_dialog(self):
        ap = self._collect_all_pending()
        if not ap: messagebox.showinfo("No Pending Invoices", "All matched invoices have already been investigated or no matches were found yet.", parent=self.root); return
        if getattr(self, 'active_pending_dlg', None) and self.active_pending_dlg.winfo_exists(): self.active_pending_dlg.destroy()
        self.active_pending_dlg = PendingInvoicesDialog(self.root, ap, self.curr_m.get('a', ''), lambda res: (self.curr_m['processed'].add(self._get_loop_key(res['match'])), self.root.after(0, lambda: self._show_dlg([res['match']]))) if res['action']=='investigate' else None)

    def _show_pending_gateway(self):
        ap = self._collect_all_pending()
        if not ap: 
            if hasattr(self, 'curr_m') and self.curr_m.get('block'):
                self.all_blocks.append(self.curr_m['block'])
            self._next_man(force_next=True); return
        
        def on_res(res):
            if res['action']=='investigate':
                m = res['match']
                self.curr_m['processed'].add(self._get_loop_key(m))
                # v7.1.5: Resume at correct depth and remaining PQV context
                self.curr_m.update({
                    'depth': m.get('_depth', 0) + 1,
                    'rem': m.get('_rem_at_discovery', self.curr_m['rem']),
                    'budget': m.get('_budget', safe_num(m.get('mtc_qty', 0))),
                    'rendered': False,
                    '_awaiting_cross_po': False
                })
                self.root.after(0, lambda: self._show_dlg([m]))
            else:
                # Go to Next ASIN
                if self.curr_m.get('block'): self.all_blocks.append(self.curr_m['block'])
                self._next_man(force_next=True)

        if getattr(self, 'active_pending_dlg', None) and self.active_pending_dlg.winfo_exists(): self.active_pending_dlg.destroy()
        self.active_pending_dlg = PendingInvoicesDialog(self.root, ap, f"{self.curr_m.get('a', '')} (Conclusion Review)", on_res)


    def _handle_cross_po(self, res):
        if res['action'] == 'skip': threading.Thread(target=self._man_step, daemon=True).start(); return
        sid = extract_sid(res['candidate']['sid'])
        if sid not in self.engine.cross_po_vault: self.engine.cross_po_vault[sid] = []
        
        # v6.2.6: Store in Vault categorized by Shipment ID
        self.engine.cross_po_vault[sid].append({
            'candidate': res['candidate'], 
            'case': res.get('case','Case 1'), 
            'budget': safe_num(res['candidate']['rec_qty']),
            'found_bc': res['candidate'].get('found_bc', ''),
            'found_inv': res['candidate'].get('found_inv', '')
        })
        self._set_status(f"Cross PO stored in Vault: SID {sid} | PO {res['candidate']['po']}", None)
        threading.Thread(target=self._man_step, daemon=True).start()

    def _handle_cross_po_and_finish(self, res):
        if res['action'] == 'skip': self.curr_m.pop('_awaiting_cross_po', None); threading.Thread(target=self._man_step, daemon=True).start(); return
        c, bud = res['candidate'], safe_num(res['candidate']['rec_qty'])
        # v6.1.2: Dynamically use captured Barcode, Invoice, and Inv Qty instead of empty strings
        f_bc = c.get('found_bc', '')
        f_inv = c.get('found_inv', '')
        f_iq = c.get('inv_qty', 0)
        
        self._set_status(f"Cross PO confirmed ({c['po']}) — starting manual investigation of {int(bud)} units…", None); self.curr_m.pop('_awaiting_cross_po', None)
        # v7.2.3: Store claiming_po so deeper levels can exclude it from cross PO re-detection
        claiming_po = self.curr_m.get('claiming_po', self.curr_m.get('p', ''))
        self.curr_m.update({'b':f_bc, 'i':f_inv, 's':c['sid'], 'p':c['po'], 'a':c['asin'], 'iq':f_iq, 'rem':bud, 'budget':bud, 'depth':self.curr_m['depth']+1, 'rendered':False, 'processed':self.curr_m['processed'], 'cross_po_checked':True, 'asin_rendered_levels':set(), 'claiming_po': claiming_po}); threading.Thread(target=self._man_step, daemon=True).start()

    def _finish(self):
        msg = "Investigation complete!" if not (hasattr(self, 'engine') and self.engine.stop_requested) else "Investigation stopped by user."
        self._set_status("Complete. Click SAVE.", 100); self.root.after(0, lambda: (self.run_btn.config(state="normal"), self.save_btn.config(state="normal"), self.portal_btn.config(state="normal"), self.stop_inv_btn.config(state="disabled"), self.stop_sess_btn.config(state="disabled"), messagebox.showinfo("Done", msg)))

    def _get_loop_key(self, mt): return (clean(mt.get('mtc_inv','')), clean(mt.get('mtc_asin','')), clean(mt.get('mtc_po','')), fmt_qty(mt.get('mtc_qty',0)))

    def save_output(self):
        t, ts = self.ticket_id.get().strip().replace(' ','_'), datetime.now().strftime('%Y%m%d_%H%M%S')
        o = f"MFI_{t}_{ts}.xlsx" if t else f"MFI_Investigation_{ts}.xlsx"
        
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=o)
        if not path: return
        
        self.save_btn.config(state="disabled")
        self._set_status("Saving output... Please wait...", 0)
        self.pb.config(mode='indeterminate')
        self.pb.start()
        
        bl = list(self.all_blocks)
        if self.mode_var.get() == "manual" and hasattr(self, 'curr_m') and self.curr_m.get('block') and self.curr_m['block'] not in bl:
            bl.append(self.curr_m['block'])
            
        rebni_data = self.engine.collected_rebni if hasattr(self, 'engine') else None
        sids = self.engine.unique_sids_found if hasattr(self, 'engine') else None
        vraw = self.engine.vendor_raw if hasattr(self, 'engine') else None
        vhdr = self.engine.vendor_headers if hasattr(self, 'engine') else None
        
        r_claims = getattr(self, 'raw_claims_df', None)
        r_rebni = getattr(self, 'raw_rebni_df', None)
        r_inv = getattr(self, 'inv_search_df', None)
        c_asin = getattr(self.engine, 'ciat_asin_df', None) if hasattr(self, 'engine') else None
        c_ship = getattr(self.engine, 'ciat_shipment_df', None) if hasattr(self, 'engine') else None
        c_merges = getattr(self.engine, 'ciat_shipment_merges', None) if hasattr(self, 'engine') else None
        
        def _save_task():
            try:
                write_excel(bl, path, rebni_summary_data=rebni_data, unique_sids_found=sids, vendor_raw=vraw, vendor_headers=vhdr, raw_claims=r_claims, raw_rebni=r_rebni, raw_inv=r_inv, ciat_asin_df=c_asin, ciat_shipment_df=c_ship, ciat_shipment_merges=c_merges)
                self.root.after(0, lambda: self._set_status(f"Project saved to: {os.path.basename(path)}"))
                self.root.after(0, lambda: self.save_btn.config(state="normal"))
                self.root.after(0, lambda: messagebox.showinfo("Saved", f"Complete investigation report saved to:\n{path}"))
            except Exception as e:
                self.root.after(0, lambda err=str(e): messagebox.showerror("Save Error", err))
                self.root.after(0, lambda: self.save_btn.config(state="normal"))
            finally:
                self.root.after(0, self.pb.stop)
                self.root.after(0, lambda: self.pb.config(mode='determinate'))
                
        threading.Thread(target=_save_task, daemon=True).start()

    def open_summary_portal(self):
        """v6.2.5: Launches the summary portal. Priority to Legacy HTML (Canonical), Fallback to React."""
        if hasattr(sys, '_MEIPASS'):
            base_dir = sys._MEIPASS
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            
        # Prefer legacy HTML as it contains the dashboard/AI logic the user expects
        portal_path = os.path.join(base_dir, "MFI_unique_summary_upload_export.html")
        if not os.path.exists(portal_path):
            portal_path = os.path.join(base_dir, "Web application", "dist", "index.html")
        
        if os.path.exists(portal_path):
            self._set_status("Launching Summary Portal...")
            from pathlib import Path
            webbrowser.open_new_tab(Path(portal_path).as_uri())
        else:
            messagebox.showerror("Portal Error", 
                                f"Could not find any portal files:\n{portal_path}\n\n"
                                f"Ensure the legacy HTML or 'Web application/dist' folder is present.")

    def open_about_page(self):
        """v7.2.0: Launches the 'About' documentation HTML page."""
        if hasattr(sys, '_MEIPASS'):
            base_dir = sys._MEIPASS
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            
        about_path = os.path.join(base_dir, "MFI_Tool.html")
        
        if os.path.exists(about_path):
            self._set_status("Opening About Page...")
            from pathlib import Path
            webbrowser.open_new_tab(Path(about_path).as_uri())
        else:
            messagebox.showerror("About Error", 
                                f"Could not find the About page:\n{about_path}\n\n"
                                f"Ensure 'MFI_Tool.html' is bundled or in the root directory.")

    def open_generic_rec_qty_lookup(self):
        """v6.2.6: Provides a fully detailed standalone Received Qty lookup tool on the main screen."""
        dlg = tk.Toplevel(self.root)
        dlg.title("Generic Received Qty Lookup")
        dlg.geometry("450x320")
        dlg.configure(bg="#0d1117")
        
        tk.Label(dlg, text="  Standalone REBNI Lookup Tool", bg="#161b22", fg="#f0a500", font=("Segoe UI", 11, "bold"), height=2).pack(fill="x")
        
        form_f = tk.Frame(dlg, bg="#0d1117", padx=20, pady=15)
        form_f.pack(fill="x")
        
        tk.Label(form_f, text="Shipment ID:", bg="#0d1117", fg="#888888", font=("Segoe UI", 10)).grid(row=0, column=0, sticky='w', pady=5)
        sid_ent = tk.Entry(form_f, font=("Segoe UI", 10), bg="#21262d", fg="white", insertbackground="white", width=25)
        sid_ent.grid(row=0, column=1, padx=10, pady=5)
        
        tk.Label(form_f, text="Purchase Order:", bg="#0d1117", fg="#888888", font=("Segoe UI", 10)).grid(row=1, column=0, sticky='w', pady=5)
        po_ent = tk.Entry(form_f, font=("Segoe UI", 10), bg="#21262d", fg="white", insertbackground="white", width=25)
        po_ent.grid(row=1, column=1, padx=10, pady=5)
        
        tk.Label(form_f, text="ASIN:", bg="#0d1117", fg="#888888", font=("Segoe UI", 10)).grid(row=2, column=0, sticky='w', pady=5)
        asin_ent = tk.Entry(form_f, font=("Segoe UI", 10), bg="#21262d", fg="white", insertbackground="white", width=25)
        asin_ent.grid(row=2, column=1, padx=10, pady=5)
        
        def run_lookup():
            sid = extract_sid(sid_ent.get().strip())
            po  = clean(po_ent.get().strip())
            asin = clean(asin_ent.get().strip())
            
            if not all([sid, po, asin]):
                messagebox.showerror("Error", "All fields are required.", parent=dlg); return
                
            if not hasattr(self, 'engine') or not self.engine:
                messagebox.showerror("Error", "Investigation Engine not active. Run investigation first.", parent=dlg); return
                
            rows = self.engine.rebni_p.get((sid, po, asin), [])
            if not rows:
                messagebox.showinfo("Result", "No REBNI data found for this combination.", parent=dlg); return
                
            r0 = rows[0]
            # matched invoices logic
            matched_invs = []
            for (m_inv, m_po, m_asin), m_rows in self.engine.inv_p.items():
                if clean(m_asin) == asin and clean(m_po) == po:
                    if any(clean(m_inv) == clean(i) for i in self.engine.cache_sid.get(sid, [])):
                        qty = sum(safe_num(r.get('quantity_invoiced', 0)) for r in m_rows)
                        matched_invs.append(f"{m_inv} (Qty:{int(qty)})")

            res_dlg = tk.Toplevel(dlg)
            res_dlg.title(f"Result for {asin}")
            res_dlg.geometry("520x480")
            res_dlg.configure(bg="#0d1117")
            
            tk.Label(res_dlg, text=f"📊  Full Reconciliation: {asin}", bg="#161b22", fg="#3fb950", font=("Segoe UI", 11, "bold"), height=2).pack(fill="x")
            
            sc = tk.Frame(res_dlg, bg="#0d1117", padx=25, pady=20)
            sc.pack(fill="both", expand=True)
            
            metrics = [
                ("Quantity Unpacked:", r0.get('quantity_unpacked', 0), "#58a6ff"),
                ("Quantity Adjusted:", r0.get('quantity_adjusted', 0), "#ff4d4d"),
                ("REBNI Available:", r0.get('rebni_available', 0), "#f0a500"),
                ("Quantity Received (Post Adj):", r0.get('qty_received_postadj', 0), "#3fb950"),
                ("Item Cost (INR):", f"{safe_num(r0.get('item_cost', 0)):.2f}", "#abb2bf")
            ]
            
            for i, (l, v, c) in enumerate(metrics):
                tk.Label(sc, text=l, bg="#0d1117", fg="#888888", font=("Segoe UI", 10)).grid(row=i, column=0, sticky='w', pady=8)
                val_str = f"{v} units" if "Cost" not in l else v
                tk.Label(sc, text=val_str, bg="#0d1117", fg=c, font=("Segoe UI", 11, "bold")).grid(row=i, column=1, sticky='w', padx=20)

            tk.Label(sc, text="Matched Invoices in SID:", bg="#0d1117", fg="#4a9eff", font=("Segoe UI", 10, "bold")).grid(row=6, column=0, sticky='w', pady=(15,5))
            inv_text = ", ".join(matched_invs) if matched_invs else "No invoices found for this ASIN in the portal."
            st = scrolledtext.ScrolledText(sc, width=45, height=4, bg="#161b22", fg="#e0e0e0", font=("Consolas", 9), relief="flat")
            st.insert("1.0", inv_text)
            st.configure(state="disabled")
            st.grid(row=7, column=0, columnspan=2, sticky='w', pady=5)

            tk.Button(res_dlg, text="CLOSE", command=res_dlg.destroy, bg="#333", fg="white", padx=20).pack(pady=15)

        tk.Button(dlg, text="🔍  CHECK REBNI", command=run_lookup, bg="#d4a017", fg="black", font=("Segoe UI", 10, "bold"), padx=25, pady=10, relief="flat", cursor="hand2").pack(pady=10)

    def run(self):
        try: self.root.mainloop()
        except KeyboardInterrupt: pass

if __name__ == '__main__': MFIToolApp().run()
