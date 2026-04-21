"""
MFI Investigation Tool  v5.9.2  |  ROW IB
==========================================
ROW IB  |  Amazon
Developed by Mukesh

CHANGES IN v5.9.2:
  ✔ [INTEGRATION] Added "📑 UNIQUE SUMMARY PORTAL" button for report aggregation.
  ✔ [INTEGRATION] Implemented auto-launch logic for MFI_unique_summary_upload_export.html.
  ✔ [UI/FIX] Standardized all branding and labels to "v5.9.2 | ROW IB".
  ✔ [SAFETY] Migrated 100% stable logic from v5.9.1 before adding features.

CHANGES IN v5.9.1:
  ✔ [UI/FIX] CorrespondenceDialog now truly non-modal, resizable, and includes maximize/minimize buttons.
  ✔ [UI/FIX] PreviewPanel.show_correspondence re-implemented to correctly pass all_rows.
  ✔ [PERF] Transitioned from iterrows() to high-speed dictionary-based indexing (to_dict('records')).
  ✔ [UI] Standardized all branding and labels to "v5.9.1 | ROW IB".
  ✔ [FIX] Corrected AttributeError in PreviewPanel.show_correspondence.
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import os
import re
import sys
import threading
import webbrowser
from datetime import datetime


# ═══════════════════════════════════════════════════════════
#  UTILITIES
# ═══════════════════════════════════════════════════════════

def extract_sid(val):
    s = str(val).strip()
    parts = re.findall(r'\d{10,}', s)
    return max(parts, key=len) if parts else s

def strip_scr(inv_no):
    """Remove trailing SCR suffix(es) from an invoice number."""
    return re.sub(r'(?:SCR)+$', '', str(inv_no).strip(), flags=re.IGNORECASE)

def safe_num(val):
    try:
        if pd.isna(val): return 0.0
    except: pass
    try:
        return float(str(val).replace(',', '').strip())
    except: return 0.0

def clean(val):
    try:
        if pd.isna(val): return ""
    except: pass
    s = str(val).strip()
    if s.endswith('.0'): s = s[:-2]
    return s

def split_comma(val):
    if not val: return []
    try:
        if pd.isna(val): return []
    except: pass
    return [s.strip() for s in str(val).split(',') if s.strip()]

def fmt_qty(val):
    n = safe_num(val)
    if n == 0: return ""
    return str(int(n)) if n == int(n) else str(n)

# ── FIX 2 helper ──────────────────────────────────────────────────────────────
# Keywords whose presence in a remark means it must NEVER be overwritten.
_REMARK_PRESERVE = (
    'Phase 1', 'Direct Shortage', 'REBNI Available', 'Shipment-level REBNI',
    'SR', 'Phase 4', 'No Invoice Search', 'short received directly',
)

def _remark_overwritable(rem):
    """Return True only when none of the preserved keywords appear in rem."""
    return not any(kw in rem for kw in _REMARK_PRESERVE)


# ═══════════════════════════════════════════════════════════
#  DIALOGS
# ═══════════════════════════════════════════════════════════

class HeaderCorrectionDialog(tk.Toplevel):
    """Shown when claims file has non-standard column headers."""
    def __init__(self, parent, corrections, mapping, df_columns, callback):
        super().__init__(parent)
        self.callback    = callback
        self.corrections = corrections
        self.mapping     = mapping
        self.df_columns  = df_columns

        # FIX 13: Updated title version
        self.title("Column Header Mismatch Detected — v5.9.2")
        self.geometry("700x480")
        self.configure(bg="#0f0f1a")
        self.resizable(True, True)
        self.lift(); self.focus_force()

        tk.Label(self,
                 text="⚠  Non-standard column headers detected in Claims file",
                 bg="#16213e", fg="#f0a500",
                 font=("Segoe UI", 12, "bold"), height=2).pack(fill="x")
        tk.Label(self,
                 text="The tool has automatically matched the columns below. "
                      "Please confirm or correct before proceeding.",
                 bg="#0f0f1a", fg="#cccccc",
                 font=("Segoe UI", 9)).pack(pady=4)

        outer = tk.Frame(self, bg="#0f0f1a"); outer.pack(fill="both", expand=True, padx=16, pady=6)

        hdrs = ["Field", "Expected", "Found in file", "Status"]
        w    = [14, 20, 28, 12]
        for ci, (h, ww) in enumerate(zip(hdrs, w)):
            tk.Label(outer, text=h, bg="#203864", fg="white",
                     font=("Calibri", 10, "bold"),
                     width=ww, anchor="w", padx=4
                     ).grid(row=0, column=ci, padx=1, pady=1, sticky="w")

        all_fields       = list(COLUMN_ALIASES.keys())
        corrected_fields = {c[0] for c in corrections}
        self._override_vars = {}

        for ri, field in enumerate(all_fields, 1):
            canonical = COLUMN_ALIASES[field][0]
            found_col = mapping.get(field, "")
            is_corrected = field in corrected_fields
            is_missing   = field not in mapping

            if is_missing:
                status_txt, status_fg, row_bg = "MISSING",   "#ff4444", "#2a0000"
            elif is_corrected:
                status_txt, status_fg, row_bg = "Auto-fixed", "#f0a500", "#1a1500"
            else:
                status_txt, status_fg, row_bg = "✔ OK",      "#44ff88", "#001a00"

            tk.Label(outer, text=field, bg=row_bg, fg="#e0e0e0",
                     font=("Calibri", 10, "bold"), width=14, anchor="w", padx=4
                     ).grid(row=ri, column=0, padx=1, pady=1, sticky="w")
            tk.Label(outer, text=canonical, bg=row_bg, fg="#aaaacc",
                     font=("Calibri", 10), width=20, anchor="w", padx=4
                     ).grid(row=ri, column=1, padx=1, pady=1, sticky="w")

            v = tk.StringVar(value=found_col or "— not found —")
            self._override_vars[field] = v
            opts = ["— not found —"] + list(df_columns)
            cb = ttk.Combobox(outer, textvariable=v, values=opts,
                              state="readonly", width=26, font=("Calibri", 9))
            cb.grid(row=ri, column=2, padx=1, pady=1, sticky="w")

            tk.Label(outer, text=status_txt, bg=row_bg, fg=status_fg,
                     font=("Calibri", 10, "bold"), width=12, anchor="w", padx=4
                     ).grid(row=ri, column=3, padx=1, pady=1, sticky="w")

        tk.Label(self,
                 text="You can change any column assignment using the dropdowns above.",
                 bg="#0f0f1a", fg="#888899", font=("Segoe UI", 8)).pack(pady=2)

        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=10)
        tk.Button(bf, text="✔  Auto-correct & Proceed",
                  command=self._proceed,
                  bg="#2d6a4f", fg="white",
                  font=("Segoe UI", 12, "bold"),
                  padx=20, pady=8, relief="flat",
                  cursor="hand2").pack(side="left", padx=10)
        tk.Button(bf, text="✖  Cancel",
                  command=self._cancel,
                  bg="#4a2020", fg="white",
                  font=("Segoe UI", 11),
                  padx=16, pady=8, relief="flat",
                  cursor="hand2").pack(side="left", padx=10)

        self.protocol("WM_DELETE_WINDOW", self._cancel)
        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _proceed(self):
        final = {}
        for field, var in self._override_vars.items():
            v = var.get()
            if v and v != "— not found —":
                final[field] = v
        self.callback({'action': 'proceed', 'mapping': final})
        self.destroy()

    def _cancel(self):
        self.callback({'action': 'cancel'})
        self.destroy()


class SIDRequestDialog(tk.Toplevel):
    """Auto mode: SID not found in REBNI → ask user."""
    def __init__(self, parent, invoice, po, asin, callback):
        super().__init__(parent)
        self.callback = callback
        self.title("SID Required — DICES Validation")
        self.geometry("540x210")
        self.resizable(True, True)
        self.configure(bg="#16213e")
        self.lift(); self.focus_force()

        tk.Label(self, text="⚠  SID Not Found in REBNI",
                 bg="#16213e", fg="#e94560",
                 font=("Segoe UI", 13, "bold")).pack(pady=(14, 4))
        tk.Label(self, text=f"Invoice: {invoice}   PO: {po}   ASIN: {asin}",
                 bg="#16213e", fg="#e0e0e0", font=("Segoe UI", 9)).pack(pady=2)
        tk.Label(self, text="Validate this invoice in DICES and enter the SID below:",
                 bg="#16213e", fg="#aaaacc", font=("Segoe UI", 9)).pack(pady=6)

        ef = tk.Frame(self, bg="#16213e"); ef.pack()
        tk.Label(ef, text="SID from DICES:", bg="#16213e", fg="#e0e0e0",
                 font=("Segoe UI", 10)).pack(side="left", padx=8)
        self._sid = tk.StringVar()
        self._entry = tk.Entry(ef, textvariable=self._sid, width=30,
                               font=("Segoe UI", 10), bg="#1e1e3a", fg="#e0e0e0",
                               insertbackground="white", relief="flat")
        self._entry.pack(side="left", padx=4)
        self._entry.focus_set()

        bf = tk.Frame(self, bg="#16213e"); bf.pack(pady=12)
        tk.Button(bf, text="✔  Continue", command=self._ok,
                  bg="#2d6a4f", fg="white", font=("Segoe UI", 11, "bold"),
                  padx=16, pady=7, relief="flat", cursor="hand2").pack(side="left", padx=8)
        tk.Button(bf, text="✖  Skip", command=self._skip,
                  bg="#6b2737", fg="white", font=("Segoe UI", 10),
                  padx=16, pady=7, relief="flat", cursor="hand2").pack(side="left", padx=8)

        self.bind('<Return>', lambda e: self._ok())
        self.protocol("WM_DELETE_WINDOW", self._skip)
        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _ok(self):
        sid = extract_sid(self._sid.get().strip())
        if sid:
            self.callback(sid); self.destroy()
        else:
            self._entry.config(bg="#3a1e1e")

    def _skip(self):
        self.callback(None); self.destroy()

class CrossPODialog(tk.Toplevel):
    """Cross PO Confirmation Dialog — v5.3.0 logic restored for manual selection"""
    CASE_DESCRIPTIONS = {
        "Case 1": (
            "Case 1 — No PO, but ASIN received",
            "Rec=0 at claiming PO. Same ASIN received in different PO within same SID.\n"
            "Those units are overage under a different PO."
        ),
        "Case 2": (
            "Case 2 — PO exists but ASIN not ordered there",
            "This PO exists in the claiming SID, but the ASIN was never invoiced there.\n"
            "Inv Qty = 0, but units were received. This is a Cross PO overage."
        ),
        "Case 3": (
            "Case 3 — PO and ASIN exist but Rec > Inv",
            "Both PO and ASIN are present. Invoiced qty = X, but received more than X.\n"
            "Excess units are Cross PO overage."
        ),
    }

    def __init__(self, parent, candidates, current_inv, sid, callback):
        super().__init__(parent)
        self.callback   = callback
        self.candidates = candidates

        self.title("Cross PO Overage — Confirm & Investigate")
        self.geometry("740x540")
        self.resizable(True, True)
        self.configure(bg="#0f0f1a")
        self.lift(); self.focus_force()

        tk.Label(self, text="🔄  Cross PO Overage Detected",
                 bg="#16213e", fg="#f0a500",
                 font=("Segoe UI", 13, "bold"), height=2).pack(fill="x")
        tk.Label(self,
                 text=f"SID: {sid}   |   Investigation Invoice: {current_inv}",
                 bg="#0f0f1a", fg="#cccccc",
                 font=("Segoe UI", 9)).pack(pady=2)
        tk.Label(self,
                 text="On confirming, the tool will investigate the Cross PO chain "
                      "to find equivalent shortage.",
                 bg="#0f0f1a", fg="#4a9eff",
                 font=("Segoe UI", 9)).pack(pady=2)

        tf = tk.LabelFrame(self, text="  Detected Cross PO Candidates  ",
                           bg="#0f0f1a", fg="#e0e0e0",
                           font=("Segoe UI", 9, "bold"), padx=10, pady=6)
        tf.pack(fill="x", padx=16, pady=6)
        for ci, h in enumerate(["Cross PO", "ASIN", "Inv Qty", "Rec Qty", "Overage", "Type"]):
            tk.Label(tf, text=h, bg="#203864", fg="white",
                     font=("Calibri", 10, "bold"), width=14, anchor="w", padx=3
                     ).grid(row=0, column=ci, padx=1, pady=1)
        for ri, c in enumerate(candidates, 1):
            inv_n = safe_num(c.get('inv_qty', 0))
            rec_n = safe_num(c['rec_qty'])
            ovg   = max(0.0, rec_n - inv_n)
            for ci, v in enumerate([c['po'], c['asin'],
                                     fmt_qty(inv_n), fmt_qty(rec_n),
                                     fmt_qty(ovg) or "—",
                                     c['cross_type'].split("—")[0].strip()]):
                tk.Label(tf, text=str(v), bg="#1e1e3a", fg="#e0e0e0",
                         font=("Calibri", 10), width=14, anchor="w", padx=3
                         ).grid(row=ri, column=ci, padx=1, pady=1)

        sf = tk.Frame(self, bg="#0f0f1a"); sf.pack(fill="x", padx=16, pady=4)
        tk.Label(sf, text="Select Cross PO to investigate:",
                 bg="#0f0f1a", fg="#e0e0e0",
                 font=("Segoe UI", 10), width=30, anchor="w").pack(side="left")
        opts = [f"PO={c['po']}  Rec={fmt_qty(c['rec_qty'])}  {c['cross_type'].split(chr(8212))[0].strip()}"
                for c in candidates] + ["None — Skip"]
        self._sel_var = tk.StringVar()
        self._sel_cb  = ttk.Combobox(sf, textvariable=self._sel_var,
                                      values=opts, state="readonly", width=50,
                                      font=("Segoe UI", 9))
        self._sel_cb.current(0)
        self._sel_cb.pack(side="left", padx=6)
        self._sel_cb.bind("<<ComboboxSelected>>", self._on_candidate_change)

        cf = tk.LabelFrame(self, text="  Confirm Cross PO Case  ",
                           bg="#0f0f1a", fg="#e0e0e0",
                           font=("Segoe UI", 9, "bold"), padx=12, pady=8)
        cf.pack(fill="x", padx=16, pady=4)

        self._case_var = tk.StringVar(value="Case 1")
        self._case_desc_lbl = tk.Label(cf, text="",
                                        bg="#0f0f1a", fg="#aaaacc",
                                        font=("Segoe UI", 9), justify="left",
                                        wraplength=640, anchor="w")

        for case_key, (case_label, _) in self.CASE_DESCRIPTIONS.items():
            tk.Radiobutton(cf, text=case_label,
                           variable=self._case_var, value=case_key,
                           bg="#0f0f1a", fg="#f0c060",
                           selectcolor="#1a1500",
                           font=("Segoe UI", 10, "bold"),
                           command=self._on_case_change
                           ).pack(anchor="w", pady=2)

        self._case_desc_lbl.pack(anchor="w", pady=4, padx=8)
        self._on_case_change()

        tk.Label(self,
                 text="⚡  On confirming: tool will investigate Cross PO chain "
                      "until full Cross PO rec_qty is explained as shortage.",
                 bg="#0f0f1a", fg="#88ccff",
                 font=("Segoe UI", 9, "italic")).pack(pady=4, padx=16, anchor="w")

        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=10)
        tk.Button(bf, text="✔  Confirm & Investigate",
                  command=self._confirm,
                  bg="#2d6a4f", fg="white",
                  font=("Segoe UI", 12, "bold"),
                  padx=20, pady=9, relief="flat",
                  cursor="hand2").pack(side="left", padx=10)
        tk.Button(bf, text="✖  Skip",
                  command=self._skip,
                  bg="#4a2020", fg="white",
                  font=("Segoe UI", 11),
                  padx=16, pady=9, relief="flat",
                  cursor="hand2").pack(side="left", padx=10)

        self.protocol("WM_DELETE_WINDOW", self._skip)
        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _on_candidate_change(self, event=None):
        pass

    def _on_case_change(self):
        case_key = self._case_var.get()
        _, desc  = self.CASE_DESCRIPTIONS.get(case_key, ("", ""))
        self._case_desc_lbl.config(text=desc)

    def _confirm(self):
        idx = self._sel_cb.current()
        if idx >= len(self.candidates):
            self.callback({'action': 'skip'}); self.destroy(); return
        self.callback({
            'action'   : 'confirmed',
            'candidate': self.candidates[idx],
            'case'     : self._case_var.get(),
        })
        self.destroy()

    def _skip(self):
        self.callback({'action': 'skip'}); self.destroy()


class ManualLevelDialog(tk.Toplevel):
    """Manual mode: per-level dialog with IBC/PBC + Cross PO + Mismatches."""
    def __init__(self, parent, matches, remaining_pqv, branch_budget, callback,
                 pending_cb=None, engine=None):
        super().__init__(parent)
        self.callback      = callback
        self.matches       = matches
        self.rem_pqv       = remaining_pqv
        self.branch_budget = branch_budget
        self._pending_cb   = pending_cb
        self._engine_ref   = engine

        self.title("Manual Investigation — Next Step")
        self.geometry("680x560")
        self.configure(bg="#0f0f1a")
        self.resizable(True, True)
        self.lift(); self.focus_force()

        tk.Label(self, text="  Manual Investigation — Continue",
                 bg="#16213e", fg="#4a9eff",
                 font=("Segoe UI", 12, "bold"), height=2).pack(fill="x")

        info = f"Remaining PQV: {int(remaining_pqv)}    Branch budget: {int(branch_budget)}"
        tk.Label(self, text=info, bg="#0f0f1a", fg="#cccccc",
                 font=("Segoe UI", 9)).pack(pady=2)

        inv_f = tk.LabelFrame(self, text="  Select Invoice to Continue  ",
                              font=("Segoe UI", 9, "bold"),
                              bg="#0f0f1a", fg="#e0e0e0", padx=10, pady=6)
        inv_f.pack(fill="x", padx=16, pady=4)
        opts = [f"Qty={fmt_qty(m['mtc_qty'])}  |  Inv={m['mtc_inv']}  |  PO={m['mtc_po']}  |  ASIN={m['mtc_asin']}"
                for m in matches]
        self._branch_var = tk.StringVar()
        self._branch_cb  = ttk.Combobox(inv_f, textvariable=self._branch_var,
                                         values=opts, state="readonly", width=70,
                                         font=("Segoe UI", 9))
        if opts: self._branch_cb.current(0)
        self._branch_cb.pack()

        ibc_f = tk.LabelFrame(self, text="  IBC = PBC Validation  ",
                               font=("Segoe UI", 9, "bold"),
                               bg="#0f0f1a", fg="#e0e0e0", padx=10, pady=6)
        ibc_f.pack(fill="x", padx=16, pady=4)

        self._validity = tk.StringVar(value="valid")
        rf = tk.Frame(ibc_f, bg="#0f0f1a"); rf.pack(fill="x")
        tk.Radiobutton(rf, text="✔  IBC = PBC  VALID — Continue investigation",
                       variable=self._validity, value="valid",
                       bg="#0f0f1a", fg="#90ee90", selectcolor="#1e3a28",
                       font=("Segoe UI", 10, "bold"),
                       command=self._toggle).pack(side="left", padx=6)
        tk.Radiobutton(rf, text="✗  IBC ≠ PBC  INVALID — Exclude units",
                       variable=self._validity, value="invalid",
                       bg="#0f0f1a", fg="#ff8888", selectcolor="#3a1e1e",
                       font=("Segoe UI", 10, "bold"),
                       command=self._toggle).pack(side="left", padx=14)

        self._invalid_frame = tk.Frame(ibc_f, bg="#0f0f1a")
        self._invalid_frame.pack(fill="x", pady=3)
        tk.Label(self._invalid_frame, text="Units matched to invalid invoice:",
                 bg="#0f0f1a", fg="#ff8888", font=("Segoe UI", 9)).pack(side="left", padx=4)
        self._inv_qty_var = tk.StringVar()
        tk.Entry(self._invalid_frame, textvariable=self._inv_qty_var, width=10,
                 font=("Segoe UI", 10), bg="#1e1e3a", fg="#ff8888",
                 insertbackground="white", relief="flat").pack(side="left", padx=4)

        self._dices_frame = tk.LabelFrame(self, text="  DICES Details  ",
                                           font=("Segoe UI", 9, "bold"),
                                           bg="#0f0f1a", fg="#e0e0e0", padx=10, pady=6)
        self._dices_frame.pack(fill="x", padx=16, pady=4)
        r1 = tk.Frame(self._dices_frame, bg="#0f0f1a"); r1.pack(fill="x", pady=2)
        tk.Label(r1, text="SID from DICES:", bg="#0f0f1a", fg="#e0e0e0",
                 font=("Segoe UI", 9), width=20, anchor="w").pack(side="left")
        self._sid_var = tk.StringVar()
        tk.Entry(r1, textvariable=self._sid_var, width=28, font=("Segoe UI", 9),
                 bg="#1e1e3a", fg="#e0e0e0", insertbackground="white",
                 relief="flat").pack(side="left", padx=4)

        r2 = tk.Frame(self._dices_frame, bg="#0f0f1a"); r2.pack(fill="x", pady=2)
        tk.Label(r2, text="Barcode from DICES:", bg="#0f0f1a", fg="#e0e0e0",
                 font=("Segoe UI", 9), width=20, anchor="w").pack(side="left")
        self._bc_var = tk.StringVar()
        tk.Entry(r2, textvariable=self._bc_var, width=28, font=("Segoe UI", 9),
                 bg="#1e1e3a", fg="#e0e0e0", insertbackground="white",
                 relief="flat").pack(side="left", padx=4)

        self._toggle()

        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=6)
        tk.Button(bf, text="▶  CONTINUE",
                  command=self._ok, bg="#2d6a4f", fg="white",
                  font=("Segoe UI", 12, "bold"), padx=16, pady=8,
                  relief="flat", cursor="hand2").pack(side="left", padx=6)
        tk.Button(bf, text="🔄  CROSS PO",
                  command=self._cross_po, bg="#7a5c00", fg="white",
                  font=("Segoe UI", 10, "bold"), padx=12, pady=8,
                  relief="flat", cursor="hand2").pack(side="left", padx=6)
        tk.Button(bf, text="⚖  MISMATCH / OVERAGE",
                  command=self._mismatch, bg="#2d4a7a", fg="white",
                  font=("Segoe UI", 10, "bold"), padx=12, pady=8,
                  relief="flat", cursor="hand2").pack(side="left", padx=6)
        tk.Button(bf, text="⬛  STOP THIS ASIN",
                  command=self._stop, bg="#4a2020", fg="white",
                  font=("Segoe UI", 10), padx=12, pady=8,
                  relief="flat", cursor="hand2").pack(side="left", padx=6)

        # ── Pending Invoices & Utility buttons ──────────────────────────────
        bf2 = tk.Frame(self, bg="#0f0f1a"); bf2.pack(pady=4, fill='x', padx=16)
        tk.Button(bf2, text="📋  VIEW ALL PENDING INVOICES",
                  command=self._show_pending, bg="#3a2a00", fg="#f0c060",
                  font=("Segoe UI", 10, "bold"), padx=16, pady=7,
                  relief="flat", cursor="hand2").pack(side="left", expand=True, fill="x", padx=(0,4))
        
        tk.Button(bf2, text="🔍 LOOKUP INV QTY",
                  command=self._lookup_inv_qty, bg="#1c2c42", fg="#80a0ff",
                  font=("Segoe UI", 10, "bold"), padx=16, pady=7,
                  relief="flat", cursor="hand2").pack(side="right")

        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _lookup_inv_qty(self):
        dlg = tk.Toplevel(self)
        dlg.title("Lookup Invoice Quantity")
        dlg.configure(bg="#0d1117")
        dlg.attributes("-topmost", True)
        
        tk.Label(dlg, text="Invoice No:", bg="#0d1117", fg="white", font=("Segoe UI", 10)).grid(row=0, column=0, padx=10, pady=(15,5), sticky='e')
        inv_ent = tk.Entry(dlg, font=("Segoe UI", 10), bg="#21262d", fg="white", insertbackground="white")
        inv_ent.grid(row=0, column=1, padx=10, pady=(15,5))
        
        tk.Label(dlg, text="ASIN:", bg="#0d1117", fg="white", font=("Segoe UI", 10)).grid(row=1, column=0, padx=10, pady=5, sticky='e')
        asin_ent = tk.Entry(dlg, font=("Segoe UI", 10), bg="#21262d", fg="white", insertbackground="white")
        asin_ent.grid(row=1, column=1, padx=10, pady=5)
        
        res_lbl = tk.Label(dlg, text="", bg="#0d1117", fg="#f0c060", font=("Segoe UI", 11, "bold"))
        res_lbl.grid(row=3, column=0, columnspan=2, pady=(10,15))
        
        def do_lookup():
            ino = inv_ent.get().strip()
            asn = asin_ent.get().strip()
            
            # Use direct engine reference passed from app
            engine = getattr(self, '_engine_ref', None)
            
            if not engine:
                res_lbl.config(text="Engine not connected.", fg="#f85149")
                return
                
            # Optimized lookup using the high-speed indexer (inv_iam)
            found_qty = engine._resolve_inv_qty(ino, asn, None)
            
            if found_qty is not None:
                res_lbl.config(text=f"Exact Inv Qty = {int(found_qty)} units", fg="#3fb950")
            else:
                res_lbl.config(text="Not found in Invoice Data.", fg="#f85149")
                
        tk.Button(dlg, text="🔍 Search Data", bg="#238636", fg="white", font=("Segoe UI", 10, "bold"),
                  command=do_lookup, cursor="hand2", padx=20).grid(row=2, column=0, columnspan=2, pady=10)
        
        dlg.update_idletasks()
        px_dlg = self.winfo_x() + (self.winfo_width() - dlg.winfo_width()) // 2
        py_dlg = self.winfo_y() + 50
        dlg.geometry(f"+{px_dlg}+{py_dlg}")
        inv_ent.focus_set()

    def _toggle(self):
        if self._validity.get() == "valid":
            self._invalid_frame.pack_forget()
            self._dices_frame.pack(fill="x", padx=16, pady=4)
        else:
            self._dices_frame.pack_forget()
            self._invalid_frame.pack(fill="x", pady=3)

    def _ok(self):
        sel = self._branch_cb.current()
        if sel < 0 or sel >= len(self.matches):
            messagebox.showwarning("Select Invoice", "Please select an invoice.", parent=self)
            return
        match = self.matches[sel]
        if self._validity.get() == "valid":
            sid = extract_sid(self._sid_var.get().strip())
            if not sid:
                messagebox.showwarning("SID Required", "Please enter SID from DICES.", parent=self)
                return
            self.callback({'action': 'valid', 'chosen_match': match,
                           'sid': sid, 'barcode': self._bc_var.get().strip() or "[DICES]"})
        else:
            qty_str = self._inv_qty_var.get().strip()
            try:
                qty = float(qty_str)
            except:
                messagebox.showwarning("Qty Required", "Enter units matched to invalid invoice.", parent=self)
                return
            self.callback({'action': 'invalid', 'chosen_match': match, 'invalid_qty': qty})
        self.destroy()

    def _cross_po(self):
        self.callback({'action': 'cross_po',
                       'chosen_match': self.matches[self._branch_cb.current()] if self.matches else None})
        self.destroy()

    def _mismatch(self):
        dlg = tk.Toplevel(self)
        dlg.title("Mismatch / Overage Details")
        dlg.geometry("460x260")
        dlg.configure(bg="#0f0f1a")
        dlg.lift(); dlg.focus_force()

        fields = [("ASIN received:", "asin"), ("SID:", "sid"), ("PO:", "po"),
                  ("Inv Qty (invoiced):", "inv_qty"), ("Overage Qty received:", "ovg_qty")]
        vars_ = {}
        for i, (lbl, key) in enumerate(fields):
            tk.Label(dlg, text=lbl, bg="#0f0f1a", fg="#e0e0e0",
                     font=("Segoe UI", 10), width=22, anchor="w"
                     ).grid(row=i, column=0, padx=12, pady=5)
            v = tk.StringVar()
            tk.Entry(dlg, textvariable=v, width=26, font=("Segoe UI", 10),
                     bg="#1e1e3a", fg="#e0e0e0", insertbackground="white",
                     relief="flat").grid(row=i, column=1, padx=8, pady=5)
            vars_[key] = v

        def submit():
            data = {k: v.get().strip() for k, v in vars_.items()}
            dlg.destroy()
            self.callback({'action': 'mismatch', 'mismatch_data': data})
            self.destroy()

        tk.Button(dlg, text="✔  Confirm & Investigate", command=submit,
                  bg="#2d6a4f", fg="white", font=("Segoe UI", 11, "bold"),
                  padx=14, pady=7, relief="flat", cursor="hand2"
                  ).grid(row=len(fields), column=0, columnspan=2, pady=12)

    def refresh_from_engine(self, engine):
        """Update combobox options dynamically when Confirm Edits is clicked."""
        if not engine.user_overrides:
            return
        patched = []
        for mtch in self.matches:
            inv_key = clean(mtch.get('mtc_inv', ''))
            override = engine.user_overrides.get(inv_key, {})
            if override:
                mtch = dict(mtch)
                if 'mtc_qty' in override: mtch['mtc_qty'] = override['mtc_qty']
                if 'inv_qty' in override: mtch['inv_qty'] = override['inv_qty']
            patched.append(mtch)
        self.matches = patched
        opts = [f"Qty={fmt_qty(m['mtc_qty'])}  |  Inv={m['mtc_inv']}  |  PO={m['mtc_po']}  |  ASIN={m['mtc_asin']}"
                for m in self.matches]
        self._branch_cb['values'] = opts
        idx = self._branch_cb.current()
        if opts and idx >= 0:
            self._branch_cb.current(idx)

    def _show_pending(self):
        if self._pending_cb:
            self._pending_cb()
        else:
            messagebox.showinfo("No Pending Invoices",
                                "No pending invoice information available yet.",
                                parent=self)

    def _stop(self):
        self.callback({'action': 'stop'}); self.destroy()


class PendingInvoicesDialog(tk.Toplevel):
    """
    Shown before finalizing a claiming ASIN.
    Lists all matched invoices that were identified but not yet investigated.
    User can pick one to continue investigation or skip.
    """
    def __init__(self, parent, pending_invoices, asin, callback):
        super().__init__(parent)
        self.callback         = callback
        self.pending_invoices = pending_invoices

        self.title("Pending Matched Invoices — Action Required")
        self.geometry("760x480")
        self.configure(bg="#0f0f1a")
        self.resizable(True, True)
        self.lift(); self.focus_force()

        tk.Label(self,
                 text="📋  Uninvestigated Matched Invoices Found",
                 bg="#16213e", fg="#f0a500",
                 font=("Segoe UI", 12, "bold"), height=2).pack(fill="x")

        tk.Label(self,
                 text=(f"ASIN: {asin} — The following matched invoices were identified during\n"
                       "investigation but have not yet been investigated. "
                       "Select one to continue or click 'Go to Next ASIN'."),
                 bg="#0f0f1a", fg="#cccccc",
                 font=("Segoe UI", 9), justify="left").pack(pady=6, padx=16, anchor="w")

        # Table of pending invoices
        tf = tk.LabelFrame(self, text="  Pending Matched Invoices  ",
                           bg="#0f0f1a", fg="#e0e0e0",
                           font=("Segoe UI", 9, "bold"), padx=10, pady=6)
        tf.pack(fill="both", expand=True, padx=16, pady=4)

        cols = ("Mtc Qty", "Invoice No", "PO", "ASIN", "Level")
        col_w = [80, 200, 120, 140, 60]
        for ci, (h, w) in enumerate(zip(cols, col_w)):
            tk.Label(tf, text=h, bg="#203864", fg="white",
                     font=("Calibri", 10, "bold"),
                     width=w // 8, anchor="w", padx=4
                     ).grid(row=0, column=ci, padx=1, pady=1, sticky="w")

        for ri, inv in enumerate(pending_invoices, 1):
            for ci, v in enumerate([
                fmt_qty(inv.get('mtc_qty', '')),
                inv.get('mtc_inv', ''),
                inv.get('mtc_po', ''),
                inv.get('mtc_asin', ''),
                str(inv.get('_depth', '?')),
            ]):
                tk.Label(tf, text=str(v), bg="#1e1e3a", fg="#e0e0e0",
                         font=("Calibri", 10),
                         width=[10, 25, 15, 18, 8][ci],
                         anchor="w", padx=4
                         ).grid(row=ri, column=ci, padx=1, pady=1, sticky="w")

        # Dropdown to select one
        sf = tk.Frame(self, bg="#0f0f1a"); sf.pack(fill="x", padx=16, pady=4)
        tk.Label(sf, text="Select invoice to investigate:",
                 bg="#0f0f1a", fg="#e0e0e0",
                 font=("Segoe UI", 10), width=30, anchor="w").pack(side="left")
        opts = [f"Qty={fmt_qty(inv.get('mtc_qty',''))}  |  Inv={inv.get('mtc_inv','')}  |  "
                f"PO={inv.get('mtc_po','')}  |  ASIN={inv.get('mtc_asin','')}"
                for inv in pending_invoices]
        self._sel = tk.StringVar()
        self._cb  = ttk.Combobox(sf, textvariable=self._sel,
                                   values=opts, state="readonly", width=60,
                                   font=("Segoe UI", 9))
        if opts: self._cb.current(0)
        self._cb.pack(side="left", padx=6)

        # Buttons
        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=10)
        tk.Button(bf, text="🔍  Investigate Selected Invoice",
                  command=self._investigate,
                  bg="#2d4a7a", fg="white",
                  font=("Segoe UI", 11, "bold"),
                  padx=16, pady=8, relief="flat",
                  cursor="hand2").pack(side="left", padx=8)
        tk.Button(bf, text="▶▶  Go to Next ASIN",
                  command=self._next_asin,
                  bg="#2d6a4f", fg="white",
                  font=("Segoe UI", 11, "bold"),
                  padx=16, pady=8, relief="flat",
                  cursor="hand2").pack(side="left", padx=8)

        self.protocol("WM_DELETE_WINDOW", self._next_asin)
        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _investigate(self):
        idx = self._cb.current()
        if idx < 0 or idx >= len(self.pending_invoices):
            messagebox.showwarning("Select Invoice", "Please select an invoice to investigate.", parent=self)
            return
        self.callback({'action': 'investigate', 'match': self.pending_invoices[idx]})
        self.destroy()

    def _next_asin(self):
        self.callback({'action': 'next_asin'})
        self.destroy()

class CorrespondenceDialog(tk.Toplevel):
    def __init__(self, parent, all_rows):
        super().__init__(parent)
        self.all_rows = all_rows
        self.title("Scenario Selection — Get Correspondence — v5.9.2")
        self.geometry("1000x850")
        self.configure(bg="#0f0f1a")
        
        # Enable Maximize/Minimize and ensure centered display
        self.resizable(True, True)
        self.focus_force()

        self.scenarios = [
            "Claiming Short",
            "REBNI",
            "Matching (Bulk)",
            "Mismatch ASIN Overages",
            "Invalid/Dummy Invoice [Dev]",
            "IBC vs PBC [Dev]"
        ]
        
        self.v_code_var = tk.StringVar(value="[Vendor Code]")
        self.fc_id_var = tk.StringVar(value="[FC ID]")
        
        self._build_ui()
        
    def _build_ui(self):
        f = tk.Frame(self, bg="#0f0f1a", padx=25, pady=25)
        f.pack(fill="both", expand=True)
        
        # --- Header with Professional Styling ---
        hdr = tk.Frame(f, bg="#0f0f1a")
        hdr.pack(fill="x", pady=(0, 20))
        tk.Label(hdr, text="Investigation Correspondence Generator",
                 fg="#4a9eff", bg="#0f0f1a",
                 font=("Segoe UI", 16, "bold")).pack(side="left")

        # --- SELECTION AREA ---
        sel_f = tk.LabelFrame(f, text=" STEP 1: SELECT SCENARIO ", fg="#4a9eff", bg="#0f0f1a", font=("Segoe UI", 9, "bold"))
        sel_f.pack(fill="x", pady=10)
        self.scenario_var = tk.StringVar()
        self.scenario_cb = ttk.Combobox(sel_f, textvariable=self.scenario_var, 
                                        values=self.scenarios, state="readonly", 
                                        font=("Segoe UI", 12))
        self.scenario_cb.pack(fill="x", padx=15, pady=15)
        self.scenario_cb.bind("<<ComboboxSelected>>", self.generate_text)
        
        # --- INPUT AREA ---
        inp_f = tk.LabelFrame(f, text=" STEP 2: HYDRATION DETAILS ", fg="#4a9eff", bg="#0f0f1a", font=("Segoe UI", 9, "bold"))
        inp_f.pack(fill="x", pady=10)
        grid_f = tk.Frame(inp_f, bg="#0f0f1a")
        grid_f.pack(fill="x", padx=15, pady=10)
        
        tk.Label(grid_f, text="Vendor Code:", bg="#0f0f1a", fg="#aaaaaa").grid(row=0, column=0, sticky="w")
        tk.Entry(grid_f, textvariable=self.v_code_var, width=25, font=("Segoe UI", 10)).grid(row=0, column=1, padx=(5, 30))
        
        tk.Label(grid_f, text="FC ID:", bg="#0f0f1a", fg="#aaaaaa").grid(row=0, column=2, sticky="w")
        tk.Entry(grid_f, textvariable=self.fc_id_var, width=25, font=("Segoe UI", 10)).grid(row=0, column=3, padx=5)

        # --- TEXT AREA ---
        tk.Label(f, text=" STEP 3: REVIEW CORRESPONDENCE ( интерпретированный ) ",
                 fg="#e0e0e0", bg="#0f0f1a", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(15, 5))
        
        # Use a higher contrast text box for legibility
        self.text_area = tk.Text(f, font=("Segoe UI", 11), 
                                 bg="#ffffff", fg="#000000", padx=15, pady=15,
                                 insertbackground="black", wrap="word", relief="flat")
        self.text_area.pack(fill="both", expand=True)

        # --- FOOTER BUTTONS ---
        btn_f = tk.Frame(f, bg="#0f0f1a")
        btn_f.pack(fill="x", pady=(25, 0))
        
        tk.Button(btn_f, text="📋  COPY TO CLIPBOARD ", 
                  command=self.copy_to_clip,
                  bg="#2d6a4f", fg="white", font=("Segoe UI", 11, "bold"),
                  relief="flat", padx=25, pady=12, cursor="hand2").pack(side="left")
        
        tk.Button(btn_f, text="CLOSE", command=self.destroy,
                  bg="#333333", fg="white", font=("Segoe UI", 10),
                  relief="flat", padx=20, pady=12, cursor="hand2").pack(side="right")

    def copy_to_clip(self):
        txt = self.text_area.get("1.0", tk.END).strip()
        if txt:
            self.clipboard_clear(); self.clipboard_append(txt)
            messagebox.showinfo("Success", "Professional correspondence copied!", parent=self)

    def generate_text(self, event=None):
        scenario = self.scenario_var.get()
        if not scenario: return
        
        # Gather data
        primary_row = next((r for r in self.all_rows if r.get('depth', 0) == 0), {})
        sid = str(primary_row.get('sid', '[SID missing]'))
        po  = str(primary_row.get('po', '[PO missing]'))
        inv = str(primary_row.get('invoice', '[Invoice missing]'))
        vc  = self.v_code_var.get()
        fc  = self.fc_id_var.get()
        
        # Total sums for header
        t_billed = int(sum(safe_num(r.get('inv_qty', 0)) for r in self.all_rows if r.get('depth', 0) == 0))
        t_received = int(sum(safe_num(r.get('rec_qty', 0)) for r in self.all_rows if r.get('depth', 0) == 0))
        
        # ASIN list generation
        asin_lines = []
        for r in self.all_rows:
            # Only include "real" rows with shortages
            if str(r.get('barcode','')).strip() and not str(r.get('barcode','')).startswith('['):
                iq = int(safe_num(r.get('inv_qty', 0)))
                rq = int(safe_num(r.get('rec_qty', 0)))
                if iq > rq:
                    asin = str(r.get('asin', ''))
                    missing = iq - rq
                    amt = float(safe_num(r.get('cp_status', 0)))
                    asin_lines.append(f"Asin: {asin} ---- Invoice Qty: {iq} ---- Received Qty: {rq} ---- Shortage Qty: {missing} ---- Amount: INR ---- {amt:.2f}")
        
        asin_block = "\n".join(asin_lines) if asin_lines else "[No ASIN Shortages detected in Preview]"
        
        signature = f"Regards,\nMUKESH | pathlmuk\nROW IB"
        
        if scenario == "Claiming Short":
            text = (
                f"Hello FC Team, \n\n"
                f"Note: We have already performed the all the virtual research/checks such as cross receiving, overage, REBNI, adjustments etc. "
                f"and need FC support for physical search on floor to locate missing units as per revised SOP and update at the earliest.\n\n"
                f"We are able to see that unit’s shortage received in claiming shipment.\n\n"
                f"SID#{sid} | Total Billed Qty - {t_billed} | Received Qty - {t_received}\n"
                f"Invoice#{inv} | Total Billed Qty - {t_billed} | Received Qty - {t_received}\n\n"
                f"Please locate the following ASINs that are missing from PO#{po}, If units not found physically take support from SLP Team and give update.\n\n"
                f"{asin_block}\n\n"
                f"NOTE : FC Team we are able to units are short at SID Level and No overages are Found.\n\n"
                f"**Post the investigation, please flip it back to our CTI**\n"
                f"C: TOC-India\nT: MFI\nI: PD\nGroup: NOC MFI\n\n"
                f"{signature}"
            )
        elif scenario == "REBNI":
            text = (
                f"Hello Team,\n\n"
                f"{asin_block}\n\n"
                f"NOTE : Requesting FinOps Team to Utilize the below suggested REBNI and RESOLVE the PQV of above ASINs and Update the overview tab accordingly.\n\n"
                f"Update on remaining ASIN will be given later as it is have X adjustments.\n\n"
                f"We see that vendor sent overages of mismatch ASIN in claiming SID#{sid} . Below suggested REBNIs are available as per the current REBNI report in line with same state FCs and CP criteria.\n\n"
                f"For reference attaching SS of REBNI.\n\n"
                f"Please check and utilize the REBNI and update the remaining PQV units. If suggested REBNI are utilized somewhere else, then share Invoice, ASIN and PO level details along with Invoice copy where its matched for validation.\n\n"
                f"Note:\n\n"
                f"1. If Suggested REBNI comes under same CP limit, the shipment ID and PO shouldn't be the factor.\n"
                f"2. Suggested REBNI comes under same shipment, them CP variance isn't the factor.\n\n"
                f"{signature}"
            )
        elif scenario == "Matching (Bulk)":
            text = (
                f"Hello FC Team, \n\n"
                f"Note: We have already performed the all the virtual research/checks such as cross receiving, overage, REBNI, adjustments etc. "
                f"and need FC support for physical search on floor to locate missing units as per revised SOP and update at the earliest.\n\n"
                f"We are able to see that units are received completely in claiming Shipment ID#{sid} and received units got matched with different invoices where we found shortage units received in matched invoices.\n\n"
                f"Please locate the following ASINs that are missing from PO#{po}, If units not found physically take support from SLP Team and give update.\n\n"
                f"{asin_block}\n\n"
                f"**Post the investigation, please flip it back to our CTI**\n"
                f"C: TOC-India\nT: MFI\nI: PD\nGroup: NOC MFI\n\n"
                f"{signature}"
            )
        elif scenario == "Mismatch ASIN Overages":
            text = (
                f"### Hello FinOps Team,  \n\n"
                f"This is regarding the suggested overage ASIN observed under the claiming Shipment ID [{sid}].  \n"
                f"It is clearly visible that:  \n\n"
                f"- Invoiced Quantity: {t_billed} units\n"
                f"- Received Quantity: {t_received} units\n"
                f"- Overage Units: {t_received - t_billed if t_received > t_billed else 0} units  \n\n"
                f"This indicates that the vendor has sent indifferent ASIN instead of the claiming ASINs.  \n\n"
                f"Request:\n"
                f"- Kindly utilize the suggested overages and resolve the PQV at the earliest.  \n"
                f"- If the suggested overages are not utilized, please obtain a credit note from the vendor code, explaining why different ASIN overages were sent in place of the claiming ASIN.  \n\n"
                f"{signature}"
            )
        else:
            text = f"Hello Team,\n\nScenario '{scenario}' template is pending final formatting logic.\n\n{signature}"

        self.text_area.delete("1.0", tk.END)
        self.text_area.insert(tk.END, text)

class PreviewPanel(tk.Toplevel):
    COLS      = ['Barcode', 'Inv no', 'SID', 'PO', 'ASIN', 'Inv Qty',
                 'Rec Qty', 'Mtc Qty', 'Mtc Inv', 'Mtc ASIN', 'Mtc PO', 'Remarks', 'Date', 'CP']
    COL_W_PX  = [130, 160, 130, 90, 110, 60, 60, 60, 160, 130, 130, 240, 150, 180]

    def __init__(self, parent):
        super().__init__(parent)
        self._app = None
        self.title("Investigation Preview — v5.9.2 (editable)")
        self.geometry("1400x750")
        self.configure(bg="#0f0f1a")
        self.resizable(True, True)
        
        hdr_frame = tk.Frame(self, bg="#16213e")
        hdr_frame.pack(fill="x")
        tk.Label(hdr_frame, text="  Live Investigation Preview — double-click any cell to edit",
                 bg="#16213e", fg="#4a9eff", font=("Segoe UI", 10, "bold"), height=2).pack(side="left")
        
        tk.Button(hdr_frame, text="  ✔ CONFIRM EDITS  ", command=self.confirm_edits,
                  bg="#1a5a1a", fg="#90ff90", font=("Segoe UI", 10, "bold"), relief="flat", padx=14, pady=6, cursor="hand2").pack(side="right", padx=12, pady=6)
        
        tk.Button(hdr_frame, text="  ✉ GET CORRESPONDENCE  ", command=self.show_correspondence,
                  bg="#3949ab", fg="white", font=("Segoe UI", 10, "bold"), relief="flat", padx=14, pady=6, cursor="hand2").pack(side="right", padx=12, pady=6)

        frame = tk.Frame(self, bg="#0f0f1a")
        frame.pack(fill="both", expand=True, padx=8, pady=6)
        vsb = ttk.Scrollbar(frame, orient="vertical")
        hsb = ttk.Scrollbar(frame, orient="horizontal")
        self.tree = ttk.Treeview(frame, columns=self.COLS, show='headings', yscrollcommand=vsb.set, xscrollcommand=hsb.set, height=22)
        vsb.config(command=self.tree.yview); hsb.config(command=self.tree.xview)
        for col, w in zip(self.COLS, self.COL_W_PX):
            self.tree.heading(col, text=col); self.tree.column(col, width=w, minwidth=40, anchor='w')
        self.tree.grid(row=0, column=0, sticky="nsew"); vsb.grid(row=0, column=1, sticky="ns"); hsb.grid(row=1, column=0, sticky="ew")
        frame.grid_rowconfigure(0, weight=1); frame.grid_columnconfigure(0, weight=1)
        self.tree.bind('<Double-1>', self._on_double_click)
        self._row_data = {}

        bb = tk.Frame(self, bg="#0f0f1a"); bb.pack(fill="x", padx=8, pady=4)
        tk.Button(bb, text="Clear All", command=self.clear_all, bg="#2d2d5e", fg="white", font=("Segoe UI", 9), relief="flat", padx=10, pady=4, cursor="hand2").pack(side="right")

        s = ttk.Style()
        s.configure("Treeview", font=("Calibri", 10), rowheight=24, background="#1e1e3a", foreground="#e0e0e0", fieldbackground="#1e1e3a")
        s.configure("Treeview.Heading", font=("Calibri", 10, "bold"), background="#203864", foreground="white")
        for tag, bg, fg in [('header','#203864','white'),('shortage_red','#ffcccc','#9c0006'),('crosspo','#2a1a00','#f0c060')]:
            self.tree.tag_configure(tag, background=bg, foreground=fg)

    def show_correspondence(self):
        all_rows = self.get_all_rows()
        if not all_rows:
            messagebox.showinfo("No Data", "Investigation preview is empty.", parent=self)
            return
        CorrespondenceDialog(self, all_rows)

    def add_header_row(self, label=""):
        vals = list(self.COLS); vals[4] = f"── {label} ──" if label else "Header"
        iid = self.tree.insert('', 'end', values=vals, tags=('header',))
        self._row_data[iid] = dict(zip(self.COLS, vals))

    def add_row(self, rd):
        vals = [rd.get('barcode',''), rd.get('invoice',''), rd.get('sid',''), rd.get('po',''), rd.get('asin',''), 
                rd.get('inv_qty',''), rd.get('rec_qty',''), rd.get('mtc_qty',''), rd.get('mtc_inv',''),
                rd.get('mtc_asin',''), rd.get('mtc_po',''), rd.get('remarks',''), rd.get('date',''), rd.get('cp_status','')]
        iq, rq = safe_num(rd.get('inv_qty')), safe_num(rd.get('rec_qty'))
        tag = 'shortage_red' if (iq > 0 and rq < iq) else 'crosspo' if 'cross po' in str(rd.get('remarks','')).lower() else ''
        iid = self.tree.insert('', 'end', values=vals, tags=(tag,))
        self._row_data[iid] = dict(zip(self.COLS, vals)); self._row_data[iid]['_rd'] = rd; self.tree.see(iid)

    def get_all_rows(self):
        KEY = {'Barcode':'barcode', 'Inv no':'invoice', 'SID':'sid', 'PO':'po', 'ASIN':'asin', 'Inv Qty':'inv_qty', 'Rec Qty':'rec_qty', 'Mtc Qty':'mtc_qty', 'Mtc Inv':'mtc_inv', 'Mtc ASIN':'mtc_asin', 'Mtc PO':'mtc_po', 'Remarks':'remarks', 'Date':'date', 'CP':'cp_status'}
        rows = []
        for iid in self.tree.get_children():
            d = self._row_data.get(iid, {}); rd = d.get('_rd', {}).copy()
            for col in self.COLS: rd[KEY[col]] = d.get(col, '')
            rows.append(rd)
        return rows

    def clear_all(self):
        for iid in self.tree.get_children(): self.tree.delete(iid)
        self._row_data.clear()

    def _on_double_click(self, event):
        iid = self.tree.identify_row(event.y); col = self.tree.identify_column(event.x)
        if not iid or not col: return
        col_idx = int(col.replace('#',''))-1; col_name = self.COLS[col_idx]
        bbox=self.tree.bbox(iid, col); x,y,w,h = bbox
        cur = self._row_data.get(iid,{}).get(col_name,''); ev=tk.StringVar(value=str(cur))
        e = tk.Entry(self.tree, textvariable=ev); e.place(x=x,y=y,width=w,height=h); e.focus_force()
        def save(evt=None):
            nv=ev.get(); self._row_data[iid][col_name]=nv
            vals=list(self.tree.item(iid,'values')); vals[col_idx]=nv
            self.tree.item(iid, values=vals); e.destroy()
        e.bind('<Return>', save); e.bind('<FocusOut>', save)

    def confirm_edits(self):
        app = self._app
        if not app or not hasattr(app, 'engine'): return
        for iid in self.tree.get_children():
            d = self._row_data.get(iid, {}); rd = d.get('_rd', {})
            if not rd: continue
            mtc_inv = str(d.get('Mtc Inv', '') or rd.get('mtc_inv', '')).strip()
            iq = safe_num(d.get('Inv Qty', '')); mq = safe_num(d.get('Mtc Qty', ''))
            if mtc_inv: app.engine.user_overrides[clean(mtc_inv)] = {'inv_qty':iq, 'mtc_qty':mq}


# ═══════════════════════════════════════════════════════════
#  DATA LOADERS
# ═══════════════════════════════════════════════════════════

def _load_file(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == '.csv':
        try:   return pd.read_csv(path, dtype=str, encoding='utf-8')
        except: return pd.read_csv(path, dtype=str, encoding='latin-1')
    else:
        try:
            return pd.read_excel(path, header=0, dtype=str, engine='calamine')
        except Exception:
            return pd.read_excel(path, header=0, dtype=str)

def load_claims(path): return _load_file(path)

COLUMN_ALIASES = {
    'Barcode': [
        'barcode', 'bar code', 'bar_code', 'upc', 'ean',
        'scan code', 'item code', 'carton barcode', 'pkg barcode',
    ],
    'Invoice': [
        'inv no', 'inv_no', 'invoice_no', 'invoice no', 'invoice number',
        'invoice_number', 'invoice', 'inv num', 'inv number',
        'inv#', 'invoice#', 'invc no', 'invc number', 'invc',
        'bill no', 'bill number',
    ],
    'SID': [
        'sid', 'shipment id', 'shipment_id', 'shipment no',
        'shipment number', 'ship id', 'fba shipment id',
        'inbound shipment', 'shipment', 'inbound sid',
    ],
    'PO': [
        'po', 'po_no', 'p.o.', 'po no', 'po number', 'po#',
        'purchase order', 'purchase_order', 'purchase order no',
        'purchase order number', 'po id', 'order no', 'order number',
        'purch order', 'header po', 'line po', 'po num',
        'invoice_lineitem_po',
    ],
    'ASIN': [
        'asin', 'po_asin', 'amazon asin', 'amazon product id',
        'product id', 'asin no', 'item asin', 'amazon id',
    ],
    'InvQty': [
        'inv qty', 'inv_qty', 'invoice qty', 'invoice quantity',
        'invoiced qty', 'invoiced quantity', 'quantity invoiced',
        'quantity_invoiced', 'billed qty', 'billed quantity',
        'total qty', 'total quantity', 'item qty',
        'po_ordered_quantity',
    ],
    'PQV': [
        'pqv', 'pqv qty', 'missing qty', 'missing quantity',
        'missing_qty', 'shortage', 'short qty', 'short quantity',
        'claim qty', 'claimed qty', 'dispute qty',
        'pending qty', 'outstanding qty', 'difference qty',
    ],
    'CP': [
        'cp', 'cost price', 'cost_price', 'unit cost', 'unit_cost',
        'item cost', 'item_cost', 'rate', 'amount', 'unit price',
    ],
}

def detect_claim_cols(df):
    actual_cols = list(df.columns)
    lower_map   = {c.lower().strip(): c for c in actual_cols}
    mapping     = {}
    corrections = []

    for field, aliases in COLUMN_ALIASES.items():
        found = None
        for alias in aliases:
            if alias in lower_map:
                found = lower_map[alias]; break
        if not found:
            for alias in aliases:
                for col in actual_cols:
                    if alias in col.lower() or col.lower() in alias:
                        found = col; break
                if found: break
        if found:
            canonical = aliases[0]
            if found.lower().strip() != canonical:
                corrections.append((field, found, canonical))
            mapping[field] = found
    return mapping, corrections

def load_rebni(path):
    df = _load_file(path)
    names = ['vendor_code', 'po', 'asin', 'shipment_id', 'received_datetime',
             'warehouse_id', 'item_cost', 'quantity_unpacked', 'quantity_adjusted',
             'qty_received_postadj', 'quantity_matched', 'rebni_available',
             'cnt_invoice_matched', 'matched_invoice_numbers']
    if len(df.columns) > len(names):
        names += [f'ext_{i}' for i in range(len(df.columns) - len(names))]
    df.columns = names[:len(df.columns)]
    return df

def load_invoice_search(path):
    df = _load_file(path)
    names = ['vendor_code', 'purchase_order_id', 'asin', 'invoice_number', 'invoice_date',
             'invoice_item_status', 'quantity_invoiced', 'quantity_matched_total',
             'no_of_shipments', 'shipment_id', 'shipmentwise_matched_qty',
             'matched_po', 'matched_asin']
    df.columns = names[:len(df.columns)]
    return df


# ═══════════════════════════════════════════════════════════
#  INDEX BUILDERS
# ═══════════════════════════════════════════════════════════

def build_rebni_index(df):
    p, s, fb = {}, {}, {}
    for row in df.to_dict('records'):
        sid  = extract_sid(clean(row.get('shipment_id', '')))
        po   = clean(row.get('po', ''))
        asin = clean(row.get('asin', ''))
        if not sid or not asin: continue
        p.setdefault((sid, po, asin), []).append(row)
        s.setdefault((po, asin), []).append(row)
        for inv in split_comma(row.get('matched_invoice_numbers', '')):
            if inv: fb.setdefault((sid, po, inv), []).append(row)
    return p, s, fb

def build_invoice_index(df):
    idx, fb, iam = {}, {}, {}
    for row in df.to_dict('records'):
        sids  = split_comma(row.get('shipment_id', ''))
        pos   = split_comma(row.get('matched_po', ''))
        asins = split_comma(row.get('matched_asin', ''))
        qtys  = split_comma(row.get('shipmentwise_matched_qty', ''))
        
        inv_no   = clean(row.get('invoice_number', ''))
        mtc_asin = clean(row.get('asin', ''))
        inv_qty  = safe_num(row.get('quantity_invoiced', '0'))
        
        if inv_no and mtc_asin:
            iam[(inv_no, mtc_asin.upper())] = inv_qty

        for i in range(max(len(sids), len(pos), len(asins), len(qtys))):
            s_frag   = extract_sid(sids[i] if i < len(sids) else "")
            p_val    = pos[i]   if i < len(pos)   else ""
            a_val    = asins[i] if i < len(asins) else ""
            q_val    = safe_num(qtys[i] if i < len(qtys) else "0")
            
            mtc_po   = clean(row.get('purchase_order_id', ''))
            if not s_frag or not p_val or not a_val: continue
            entry = {'mtc_inv':  inv_no,
                     'mtc_po':   mtc_po,
                     'mtc_asin': mtc_asin,
                     'inv_qty':  inv_qty,
                     'mtc_qty':  q_val,
                     'date':     clean(row.get('invoice_date', ''))}
            idx.setdefault((s_frag, p_val, a_val), []).append(entry)
            if inv_no and a_val:
                iam[(inv_no, clean(a_val).upper())] = inv_qty
            if inv_no: fb.setdefault((s_frag, p_val, inv_no), []).append(entry)
    return idx, fb, iam


# ═══════════════════════════════════════════════════════════
#  INVESTIGATION ENGINE
# ═══════════════════════════════════════════════════════════

class InvestigationEngine:
    MAX_DEPTH = 10

    def __init__(self, rp, rs, rfb, ip, ifb, iam, sid_cb=None):
        self.rebni_p    = rp
        self.rebni_s    = rs
        self.rebni_fb   = rfb
        self.inv_p      = ip
        self.inv_fb     = ifb
        self.inv_iam    = iam
        self.sid_cb     = sid_cb
        self.stop_requested = False
        self.cache_sid  = {}
        self.cache_bc   = {}
        self.loop_cache = {}
        self.user_overrides = {}

    def _rebni_lookup(self, sid, po, asin, inv_no=None):
        rows = self.rebni_p.get((sid, po, asin), [])
        if not rows and inv_no:
            rows = self.rebni_fb.get((sid, po, inv_no), [])
        return rows

    def _inv_lookup(self, sid, po, asin, inv_no=None):
        m = self.inv_p.get((sid, po, asin), [])
        if not m and inv_no:
            m = self.inv_fb.get((sid, po, inv_no), [])
        return m

    def _find_sid(self, po, asin, inv_no):
        rows = self.rebni_s.get((clean(po), clean(asin)), [])
        for r in rows:
            if inv_no in split_comma(r.get('matched_invoice_numbers', '')):
                return extract_sid(r['shipment_id'])
        return extract_sid(rows[0]['shipment_id']) if rows else None

    def _resolve_inv_qty(self, inv_no, asin, fallback_qty):
        qty = self.inv_iam.get((clean(inv_no), clean(asin).upper()))
        if qty is not None and qty > 0:
            return qty
        base = strip_scr(inv_no)
        if base != clean(inv_no):
            qty = self.inv_iam.get((base, clean(asin)))
            if qty is not None and qty > 0:
                return qty
        return fallback_qty

    def _get_shipment_rebni(self, sid, po):
        total = 0.0
        for (s, p, a), rows in self.rebni_p.items():
            if s == sid and p == po:
                for r in rows:
                    total += safe_num(r.get('rebni_available', 0))
        return total

    def get_cp(self, sid, po, asin):
        rows = self.rebni_p.get((extract_sid(sid), clean(po), clean(asin)), [])
        for r in rows:
            cp = safe_num(r.get('item_cost', 0))
            if cp > 0: return cp
        for (s, p, a), rlist in self.rebni_p.items():
            if p == clean(po) and a == clean(asin):
                for r in rlist:
                    cp = safe_num(r.get('item_cost', 0))
                    if cp > 0: return cp
        return 0.0

    def compare_cp(self, c_sid, c_po, c_asin, m_sid, m_po, m_asin, depth):
        c_cp = self.get_cp(c_sid, c_po, c_asin)
        m_cp = self.get_cp(m_sid, m_po, m_asin)
        c_lbl = "Claiming CP" if depth == 0 else "Parent ASIN CP"
        m_lbl = "Matched CP"  if depth == 0 else "Sub-Matched CP"
        if c_cp <= 0 and m_cp <= 0: return ""
        if c_cp <= 0: return f"{m_lbl}: {m_cp:.2f} | {c_lbl}: N/A"
        if m_cp <= 0: return f"{c_lbl}: {c_cp:.2f} | {m_lbl}: N/A"
        low, high = c_cp * 0.90, c_cp * 1.10
        if low <= m_cp <= high:
            return f"Within 10% CP | {c_lbl}: {c_cp:.2f}, {m_lbl}: {m_cp:.2f} (range: {low:.2f}-{high:.2f})"
        return f"NOT within 10% CP | {c_lbl}: {c_cp:.2f}, {m_lbl}: {m_cp:.2f} (range: {low:.2f}-{high:.2f})"

    def detect_cross_po(self, sid, po, asin):
        candidates = []
        seen_po = set()
        rec_at_cur = sum(safe_num(r.get('quantity_unpacked', 0)) for r in self.rebni_p.get((sid, po, asin), []))
        for (s, p, a), rows in self.rebni_p.items():
            if s != sid or a != asin or p == po or p in seen_po: continue
            for r in rows:
                rec = safe_num(r.get('quantity_unpacked', 0))
                if rec <= 0: continue
                seen_po.add(p)
                im = self.inv_p.get((sid, p, asin), [])
                iq = safe_num(im[0].get('inv_qty', 0)) if im else 0.0
                if rec_at_cur == 0 and iq == 0:
                    tp = "Case 2 — ASIN not invoiced at this PO, but received"
                elif rec > iq and iq > 0:
                    tp = "Case 3 — Rec qty > Inv qty (overage in cross PO)"
                else:
                    tp = "Case 1 — Rec=0 at current PO, units received here"
                candidates.append({'po': p, 'asin': asin, 'sid': sid, 'inv_qty': fmt_qty(iq), 'rec_qty': rec, 'cross_type': tp, 'date': clean(r.get('received_datetime', ''))})
        return candidates

    def _make_row(self, b, i, s, p, a, iq, rq, mq, mi, rem, d, depth, rtype='dominant', cp_status='', mtc_asin='', mtc_po=''):
        return {
            'barcode': b, 'invoice': i, 'sid': extract_sid(s) if s else '', 'po': p, 'asin': a,
            'inv_qty': fmt_qty(iq), 'rec_qty': fmt_qty(rq), 'mtc_qty': fmt_qty(mq), 'mtc_inv': mi,
            'mtc_asin': mtc_asin, 'mtc_po': mtc_po, 'remarks': rem, 'date': d, 'depth': depth,
            'type': rtype, 'cp_status': cp_status,
        }

    def _build_level_logic(self, barcode, inv_no, sid, po, asin, iqty, rem_pqv, depth, is_claiming, is_manual=False, cross_po_indicator_only=False, initial_cp=0.0):
        sid_frag = extract_sid(sid)
        rebni_rows = self.rebni_p.get((sid_frag, clean(po), clean(asin)), [])
        rec_qty = sum(safe_num(r.get('quantity_unpacked', 0)) for r in rebni_rows)
        r_avail = sum(safe_num(r.get('rebni_available', 0)) for r in rebni_rows)
        ex_adj = sum(safe_num(r.get('quantity_adjusted', 0)) for r in rebni_rows)
        rec_date = clean(rebni_rows[0].get('received_datetime', '')) if rebni_rows else ""
        cur_cp = initial_cp if initial_cp > 0 else self.get_cp(sid_frag, po, asin)
        cp_disp = f"{cur_cp:.2f}" if cur_cp > 0 else ""
        shortage = max(0.0, safe_num(iqty) - rec_qty)
        acc_at_lvl = shortage + r_avail + ex_adj
        remarks = ""
        if is_claiming or rec_qty < safe_num(iqty) or r_avail > 0 or ex_adj > 0:
            if ex_adj > 0 and shortage > 0:
                remarks = f"Found {int(ex_adj)} units of EX adjustments and {int(shortage)} units of shortage (Inv:{int(iqty)} Rec:{int(rec_qty)})"
            else:
                parts = []
                if shortage > 0: parts.append(f"Inv Qty:{int(iqty)}.Received Qty:{int(rec_qty)}- Shortage of {int(shortage)} Units")
                if r_avail > 0: parts.append(f"REBNI Available = {int(r_avail)} units at {'claiming' if is_claiming else 'matching'} level — Suggest TSP to utilize")
                if ex_adj > 0: parts.append(f"Found {int(ex_adj)} number of X adjustments")
                remarks = " | ".join(parts) if parts else "SR" if depth > 0 else ""

        if shortage >= rem_pqv > 0 and not remarks:
            rem = f"Phase 1 Direct Shortage: {int(shortage)} units short received directly"
            if acc_at_lvl > shortage: rem += f" (Total Accounted: {int(acc_at_lvl)} incl. REBNI/EX)"
            main_row = self._make_row(barcode, inv_no, sid, po, asin, iqty, rec_qty, rec_qty, "Short Received", rem, rec_date, depth, cp_status=cp_disp)
            res_rows = [main_row]
            shp_rebni = self._get_shipment_rebni(sid_frag, clean(po))
            if shp_rebni > 0:
                res_rows.append(self._make_row('[REBNI-SHP]', inv_no, sid_frag, po, asin, '', '', shp_rebni, '', f"Shipment-level REBNI = {int(shp_rebni)} units available across all ASINs in this shipment — Suggest TSP to utilize", rec_date, depth, rtype='rebni_shipment'))
            if cross_po_indicator_only:
                for c in self.detect_cross_po(sid_frag, clean(po), clean(asin)):
                    res_rows.append(self._make_row('[CROSS PO?]', '—', c['sid'], c['po'], c['asin'], c['inv_qty'], c['rec_qty'], '', '', f"Phase 4: Cross PO candidate detected | {c['cross_type']} | Rec={fmt_qty(c['rec_qty'])} units | PENDING USER CONFIRMATION", c['date'], depth, rtype='crosspo'))
            else:
                res_rows.extend(self._build_cross_po_rows(sid_frag, clean(po), clean(asin), depth))
            return res_rows, [], rec_qty, acc_at_lvl, 0.0, ex_adj

        if 'REBNI Available' in remarks or remarks == 'SR':
            rows = [self._make_row(barcode, inv_no, sid, po, asin, iqty, rec_qty, "", "", remarks, rec_date, depth, cp_status=cp_disp)]
            if not cross_po_indicator_only: rows.extend(self._build_cross_po_rows(sid_frag, clean(po), clean(asin), depth))
            else:
                for c in self.detect_cross_po(sid_frag, clean(po), clean(asin)):
                    rows.append(self._make_row('[CROSS PO?]', '—', c['sid'], c['po'], c['asin'], c['inv_qty'], c['rec_qty'], '', '', f"Phase 4: Cross PO candidate detected | {c['cross_type']} | Rec={fmt_qty(c['rec_qty'])} units | PENDING USER CONFIRMATION", c['date'], depth, rtype='crosspo'))
            return rows, [], rec_qty, acc_at_lvl, max(0.0, rem_pqv - acc_at_lvl), ex_adj

        raw = self.inv_p.get((sid_frag, clean(po), clean(asin)), [])
        sorted_m = sorted(raw, key=lambda x: safe_num(x['mtc_qty']), reverse=True)
        if self.user_overrides:
            patched = []
            for e in sorted_m:
                ovr = self.user_overrides.get(clean(e.get('mtc_inv', '')), {})
                if ovr:
                    e = dict(e)
                    if 'inv_qty' in ovr: e['inv_qty'] = ovr['inv_qty']
                    if 'mtc_qty' in ovr: e['mtc_qty'] = ovr['mtc_qty']
                patched.append(e)
            sorted_m = sorted(patched, key=lambda x: safe_num(x['mtc_qty']), reverse=True)

        m_inv, m_qty = "", ""
        if sorted_m:
            top = sorted_m[0]
            if top['mtc_inv'] == clean(inv_no): m_inv, m_qty = "Self Matching", fmt_qty(rec_qty)
            else: m_inv, m_qty = top['mtc_inv'], fmt_qty(top['mtc_qty'])
        elif not remarks:
            if acc_at_lvl > 0:
                m_inv, m_qty = "Short Received", fmt_qty(acc_at_lvl)
                remarks = f"Accounted for {int(acc_at_lvl)} units (Shortage={int(shortage)}, REBNI={int(r_avail)}, EX={int(ex_adj)})"
            elif rec_qty > 0 and shortage == 0: remarks = "No Invoice Search matches found — Rec Qty = Inv Qty. Possible data mismatch. Verify manually in DICES."
            else: remarks = "No Invoice Search matches found — verify manually."

        cp_str, m_asn, m_po = '', '', ''
        if sorted_m and m_inv not in ('Self Matching', 'Short Received', ''):
            top = sorted_m[0]; m_asn, m_po = top.get('mtc_asin', ''), top.get('mtc_po', '')
            cp_str = self.compare_cp(sid_frag, po, asin, sid_frag, m_po, m_asn, depth)

        rows = [self._make_row(barcode, inv_no, sid, po, asin, iqty, rec_qty, m_qty, m_inv, remarks, rec_date, depth, cp_status=cp_str, mtc_asin=m_asn, mtc_po=m_po)]
        if not is_manual or depth == 0:
            for m in sorted_m[(1 if (sorted_m and m_inv not in ("Self Matching", "Short Received")) else 0):]:
                rows.append(self._make_row("", "", "", "", "", "", "", fmt_qty(m['mtc_qty']), m['mtc_inv'], "", "", depth, 'subrow', cp_status=self.compare_cp(sid_frag, po, asin, sid_frag, m.get('mtc_po', po), m.get('mtc_asin', asin), depth), mtc_asin=m.get('mtc_asin',''), mtc_po=m.get('mtc_po','')))

        actionable = [{**m, 'inv_qty': self._resolve_inv_qty(m['mtc_inv'], m['mtc_asin'], m['inv_qty'])} for m in sorted_m if m['mtc_inv'] != clean(inv_no)]
        new_rem = max(0.0, rem_pqv - acc_at_lvl)
        if acc_at_lvl > 0:
            shp_rebni = self._get_shipment_rebni(sid_frag, clean(po))
            if shp_rebni > 0: rows.append(self._make_row('[REBNI-SHP]', inv_no, sid_frag, po, asin, '', '', shp_rebni, '', f"Shipment-level REBNI = {int(shp_rebni)} units available — Suggest TSP to utilize", rec_date, depth, rtype='rebni_shipment'))
        if not cross_po_indicator_only: rows.extend(self._build_cross_po_rows(sid_frag, clean(po), clean(asin), depth))
        return rows, actionable, rec_qty, acc_at_lvl, new_rem, ex_adj

    def _build_cross_po_rows(self, sid, po, asin, depth):
        rows = []
        for c in self.detect_cross_po(sid, po, asin):
            rows.append(self._make_row('[CROSS PO]', '—', c['sid'], c['po'], c['asin'], c['inv_qty'], c['rec_qty'], '', '', f"Cross PO — {c['cross_type']} | Overage = {fmt_qty(c['rec_qty'])} units — investigating chain", c['date'], depth, rtype='crosspo'))
            if safe_num(c['rec_qty']) > 0:
                child_rows, _ = self.run_cross_po_investigation(c, c['cross_type'].split("\u2014")[0].strip(), safe_num(c['rec_qty']), depth=depth+1)
                rows.extend(child_rows)
        return rows

    def run_auto(self, barcode, inv_no, sid, po, asin, iqty, pqv, depth=0, visited=None, rem_pqv=None, is_claiming=True, branch_budget=None, max_depth_override=None, is_manual=False, row_callback=None, initial_cp=0.0):
        if self.stop_requested: return [], 0.0
        if visited is None: visited = set()
        if rem_pqv is None: rem_pqv = safe_num(pqv)
        if branch_budget is None: branch_budget = rem_pqv
        sid_frag = extract_sid(sid); state = (sid_frag, clean(inv_no), clean(po), clean(asin))
        eff_max = max_depth_override if max_depth_override is not None else self.MAX_DEPTH
        if state in visited or depth >= eff_max: return [], 0.0
        visited = visited | {state}
        if state in self.loop_cache and depth > 0: return self.loop_cache[state]
        rows, actionable, rq, acc, n_rem, ex = self._build_level_logic(barcode, inv_no, sid, po, asin, iqty, rem_pqv, depth, is_claiming, is_manual=is_manual, initial_cp=initial_cp)
        if row_callback:
            for r in rows: row_callback(r)
        if depth == 0 and not is_manual:
            for c in self.detect_cross_po(sid_frag, po, asin):
                cp_row = self._make_row('[CROSS PO]', '—', c['sid'], c['po'], c['asin'], c['inv_qty'], c['rec_qty'], '', '', f"Phase 4: Auto-detected Cross PO | {c['cross_type']} | Overage={fmt_qty(c['rec_qty'])} units | Note-only for Auto Mode", c['date'], depth, rtype='crosspo')
                rows.append(cp_row)
                if row_callback: row_callback(cp_row)
        total_acc = min(branch_budget, max(0.0, acc)); cur_budget = branch_budget - total_acc
        if cur_budget <= 0 or not actionable or 'REBNI' in rows[0].get('remarks', '') or rows[0].get('remarks', '') == 'SR':
            if rows and total_acc > 0 and _remark_overwritable(rows[0].get('remarks', '')):
                rows[0]['remarks'] = f"Accounted for {int(total_acc)} units at this level — Budget Explained"
            if depth > 0: self.loop_cache[state] = (list(rows), total_acc)
            return rows, total_acc
        for match in actionable:
            if self.stop_requested: break
            n_inv, n_po, n_asin = match['mtc_inv'], match['mtc_po'], match['mtc_asin']
            n_budget = safe_num(match['mtc_qty']) if safe_num(match['mtc_qty']) > 0 else cur_budget
            n_iqty = self._resolve_inv_qty(n_inv, n_asin, match['inv_qty'])
            n_sid = self.cache_sid.get(n_inv) or self._find_sid(n_po, n_asin, n_inv)
            if not n_sid and self.sid_cb:
                n_sid = self.sid_cb(n_inv, n_po, n_asin)
                if n_sid: self.cache_sid[n_inv] = n_sid
            if not n_sid:
                rows.append(self._make_row("[DICES]", n_inv, "[ENTER SID FROM DICES]", n_po, n_asin, n_iqty, "", "", "", "Phase 2: SID not found — validate in DICES", "", depth + 1))
                continue
            child_rows, child_acc = self.run_auto(self.cache_bc.get(n_inv, "[DICES]"), n_inv, n_sid, n_po, n_asin, n_iqty, pqv, depth+1, visited, rem_pqv-total_acc, False, n_budget, max_depth_override, is_manual, row_callback)
            rows.extend(child_rows); contribution = min(cur_budget, child_acc); total_acc += contribution; cur_budget -= contribution
        if rows and total_acc > 0 and _remark_overwritable(rows[0].get('remarks', '')):
            rows[0]['remarks'] = f"Accounted for {int(total_acc)} units of budget {int(branch_budget)} — Branch {'explained' if total_acc >= branch_budget else 'partially explained'}"
        if depth > 0: self.loop_cache[state] = (list(rows), total_acc)
        return rows, total_acc

    def run_mismatch_investigation(self, data, budget, depth=0):
        rows, match, rq, shortage, n_rem, ex = self._build_level_logic("[MISMATCH]", "", extract_sid(data.get('sid', '')), clean(data.get('po', '')), clean(data.get('asin', '')), safe_num(data.get('inv_qty', 0)), budget, depth, False)
        return rows, match, rq, shortage, n_rem

    def build_one_level(self, b, i, s, p, a, iq, rem, depth=0, is_claiming=True, is_manual=False, initial_cp=0.0):
        rows, matches, rq, shortage, n_rem, ex = self._build_level_logic(b, i, s, p, a, iq, rem, depth, is_claiming, cross_po_indicator_only=is_manual, initial_cp=initial_cp)
        return rows, [m for m in matches if m['mtc_inv'] != clean(i)], rq, n_rem

    def run_cross_po_investigation(self, c, case_type, budget, depth=0, visited=None):
        if visited is None: visited = set()
        c_sid, c_po, c_asin = c['sid'], c['po'], c['asin']; c_iq = safe_num(c.get('inv_qty', 0))
        raw = self.inv_p.get((extract_sid(c_sid), clean(c_po), clean(c_asin)), [])
        unique = sorted(raw, key=lambda x: safe_num(x['mtc_qty']), reverse=True)
        rebni = self.rebni_p.get((extract_sid(c_sid), clean(c_po), clean(c_asin)), [])
        rq = sum(safe_num(r.get('quantity_unpacked', 0)) for r in rebni) if rebni else 0.0
        ra = sum(safe_num(r.get('rebni_available', 0)) for r in rebni) if rebni else 0.0
        ex = sum(safe_num(r.get('quantity_adjusted', 0)) for r in rebni) if rebni else 0.0
        rd = clean(rebni[0].get('received_datetime', '')) if rebni else ''
        shortage = max(0.0, c_iq - rq); acc = shortage + ra + ex
        m_inv = unique[0]['mtc_inv'] if unique else "Short Received"; m_qty = fmt_qty(unique[0]['mtc_qty']) if unique else ""
        rem = f"Phase 4 Cross PO ({case_type}): Accounted for {int(acc)} units"
        if not unique and shortage > 0: rem += " — Target met via Direct Shortage"
        elif ra > 0: rem += " — Suggest TSP to utilize REBNI"
        rows = [self._make_row('[CROSS PO]', '—', c_sid, c_po, c_asin, fmt_qty(c_iq), rq, m_qty, m_inv, rem, rd, depth, mtc_asin=unique[0].get('mtc_asin','') if unique else '', mtc_po=unique[0].get('mtc_po','') if unique else '')]
        for m in unique[1:]: rows.append(self._make_row("", "", "", "", "", "", "", fmt_qty(m['mtc_qty']), m['mtc_inv'], "", "", depth, 'subrow', mtc_asin=m.get('mtc_asin',''), mtc_po=m.get('mtc_po','')))
        total_acc = min(budget, max(0.0, acc)); cur_rem = budget - total_acc
        if ra > 0 or not unique or cur_rem <= 0: return rows, total_acc
        for match in unique:
            if cur_rem <= 0: break
            n_inv, n_po, n_asin = match['mtc_inv'], match['mtc_po'], match['mtc_asin']
            n_bud = safe_num(match['mtc_qty']) if safe_num(match['mtc_qty']) > 0 else cur_rem
            n_iq = self._resolve_inv_qty(n_inv, n_asin, match['inv_qty'])
            state = (extract_sid(c_sid), clean(n_inv), clean(n_po), clean(n_asin))
            if state in visited: continue
            visited = visited | {state}
            n_sid = self.cache_sid.get(n_inv) or self._find_sid(n_po, n_asin, n_inv)
            if not n_sid and self.sid_cb:
                n_sid = self.sid_cb(n_inv, n_po, n_asin)
                if n_sid: self.cache_sid[n_inv] = n_sid
            if not n_sid:
                rows.append(self._make_row("[DICES]", n_inv, "[ENTER SID]", n_po, n_asin, n_iq, "", "", "", "Phase 4: SID not found — validate in DICES", "", depth + 1))
                continue
            child_rows, child_acc = self.run_auto(self.cache_bc.get(n_inv, "[DICES]"), n_inv, n_sid, n_po, n_asin, n_iq, cur_rem, depth+1, visited, cur_rem, False, min(n_bud, cur_rem))
            rows.extend(child_rows); contri = min(cur_rem, child_acc); total_acc += contri; cur_rem -= contri
        return rows, total_acc


# ═══════════════════════════════════════════════════════════
#  EXCEL WRITER
# ═══════════════════════════════════════════════════════════

def write_excel(all_blocks, path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Investigation"
    headers = ["Barcode", "Inv no", "SID", "PO", "ASIN", "Inv Qty", "Rec Qty", "Mtc Qty", "Mtc Inv", "Mtc ASIN", "Mtc PO", "Remarks", "Date", "CP"]
    H_FILL, DOM_F, SUB_F, ROOT_F, DICES_F, SR_F, INVLD_F, REBNI_F, CROSS_F, MIS_F = [PatternFill("solid", fgColor=c) for c in ["203864", "E2EFDA", "EBF3FB", "FFE0E0", "FFF2CC", "FFD7D7", "FFD0D0", "D0F0FF", "FFF0C0", "D0E8FF"]]
    H_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=10); N_FONT = Font(name="Calibri", size=10); ROOT_FT = Font(bold=True, color="9C0006", name="Calibri", size=10); SR_FT = Font(bold=True, color="CC0000", name="Calibri", size=10)
    INVLD_FT = Font(bold=True, color="880000", name="Calibri", size=10, italic=True); REBNI_FT = Font(bold=True, color="005580", name="Calibri", size=10); CROSS_FT = Font(bold=True, color="7a5c00", name="Calibri", size=10)
    SHORT_FILL = PatternFill("solid", fgColor="FFCCCC"); SHORT_FONT = Font(bold=True, color="9C0006", name="Calibri", size=10); BDR = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),  bottom=Side(style='thin'))
    KM = {'Barcode':'barcode', 'Inv no':'invoice', 'SID':'sid', 'PO':'po', 'ASIN':'asin', 'Inv Qty':'inv_qty', 'Rec Qty':'rec_qty', 'Mtc Qty':'mtc_qty', 'Mtc Inv':'mtc_inv', 'Mtc ASIN':'mtc_asin', 'Mtc PO':'mtc_po', 'Remarks':'remarks', 'Date':'date', 'CP':'cp_status'}
    curr = 1
    for block in all_blocks:
        if not block: continue
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=curr, column=c, value=h); cell.fill, cell.font, cell.border = H_FILL, H_FONT, BDR
        curr += 1
        for rd in block:
            rem, rtyp, dep = str(rd.get('remarks', '')), rd.get('type', 'dominant'), rd.get('depth', 0)
            iq, rq = safe_num(rd.get('inv_qty')), safe_num(rd.get('rec_qty')); is_short = (iq > 0 and rq < iq)
            if is_short: fill, fnt = SHORT_FILL, SHORT_FONT
            elif 'invalid invoice' in rem.lower(): fill, fnt = INVLD_F, INVLD_FT
            elif 'REBNI Available' in rem or 'Shipment-level REBNI' in rem or rtyp == 'rebni_shipment': fill, fnt = REBNI_F, REBNI_FT
            elif 'Cross PO' in rem or rtyp == 'crosspo': fill, fnt = CROSS_F, CROSS_FT
            elif 'short received directly' in rem.lower() or 'Direct Shortage' in rem or 'Root cause' in rem or ('Found' in rem and 'short' in rem.lower()): fill, fnt = ROOT_F, ROOT_FT
            elif rem == 'SR': fill, fnt = SR_F, SR_FT
            elif rtyp == 'subrow': fill, fnt = SUB_F, N_FONT
            elif dep > 0: fill, fnt = DOM_F, N_FONT
            else: fill, fnt = None, N_FONT
            for c, h in enumerate(headers, 1):
                val = rd.get(KM[h], ""); final_val = val; str_val = str(val).strip()
                if val not in (None, '') and str_val.replace('.','',1).isdigit(): final_val = safe_num(str_val)
                cell = ws.cell(row=curr, column=c, value=final_val if final_val not in (None, '') else None); cell.border, cell.font = BDR, fnt
                if is_short and h == "Remarks": cell.fill, cell.font = SHORT_FILL, SHORT_FONT
                elif fill: cell.fill = fill
            curr += 1
        curr += 1
    for i, w in enumerate([18, 22, 18, 12, 14, 9, 9, 9, 26, 18, 18, 42, 22, 36], 1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws2 = wb.create_sheet(title="Unique Summary")
    SH_FILL, SH_FONT = PatternFill("solid", fgColor="1F4E79"), Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    SN_FONT, SRED, SGRN = Font(name="Calibri", size=10), PatternFill("solid", fgColor="FFE2E2"), PatternFill("solid", fgColor="E2F0D9")
    CTR, LFT, SBDR = Alignment(horizontal="center", vertical="center"), Alignment(horizontal="left", vertical="center"), Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),  bottom=Side(style='thin'))
    sh = ["Barcode", "Inv no", "SID", "PO", "ASIN", "Inv Qty", "Rec Qty", "Missing QTY", "Cost Price", "Shortage Amount"]
    sw = [20, 14, 18, 12, 14, 10, 10, 14, 12, 18]
    for ci, (h, w) in enumerate(zip(sh, sw), 1):
        cell = ws2.cell(row=1, column=ci, value=h); cell.fill, cell.font, cell.border, cell.alignment = SH_FILL, SH_FONT, SBDR, CTR
        ws2.column_dimensions[get_column_letter(ci)].width = w
    seen, sr = set(), 2
    for block in all_blocks:
        for rd in block:
            b = str(rd.get('barcode', '')).strip()
            if not b or (b.startswith('[') and b != '[INVALID]'): continue
            i, s, p, a, iq, rq = str(rd.get('invoice','')).strip(), str(rd.get('sid','')).strip(), str(rd.get('po','')).strip(), str(rd.get('asin','')).strip(), rd.get('inv_qty', 0), rd.get('rec_qty', 0)
            key = (b, i, s, p, a, str(iq), str(rq))
            if key in seen: continue
            seen.add(key); iq_n, rq_n = safe_num(iq), safe_num(rq) if str(rq).strip() != '' else 0.0
            rf = SRED if iq_n > rq_n else SGRN
            v = [b, i, s, p, a, int(iq_n) if iq_n == int(iq_n) else iq_n, int(rq_n) if rq_n == int(rq_n) else rq_n, None, safe_num(rd.get('cp_status', 0)), None]
            for ci, val in enumerate(v, 1):
                if ci == 8: cell = ws2.cell(row=sr, column=ci, value=f"=F{sr}-G{sr}")
                elif ci == 10: cell = ws2.cell(row=sr, column=ci, value=f"=H{sr}*I{sr}")
                else: cell = ws2.cell(row=sr, column=ci, value=val)
                cell.font, cell.border, cell.fill, cell.alignment = SN_FONT, SBDR, rf, (CTR if ci >= 6 else LFT)
            sr += 1
    ws2.freeze_panes = "A2"; wb.save(path)


# ═══════════════════════════════════════════════════════════
#  MAIN GUI
# ═══════════════════════════════════════════════════════════

class MFIToolApp:
    def __init__(self):
        self.root = tk.Tk(); self.root.title("MFI Investigation Tool  v5.9.2  |  ROW IB")
        try: self.root.state('zoomed')
        except: self.root.attributes('-zoomed', True)
        self.root.minsize(900, 620); self.root.configure(bg="#0f0f1a")
        self.claims_path, self.rebni_path, self.inv_path, self.ticket_id, self.mode_var, self.ticket_type_var = tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar(value="auto"), tk.StringVar(value="PDTT")
        self.all_blocks, self.preview = [], None; self._build_ui()

    def _build_ui(self):
        t = tk.Frame(self.root, bg="#16213e", height=62); t.pack(fill="x")
        tk.Label(t, text="  MFI Investigation Tool", fg="#e94560", bg="#16213e", font=("Segoe UI", 20, "bold")).pack(side="left", padx=16, pady=12)
        tk.Label(t, text="Developed by Mukesh", fg="#4a9eff", bg="#16213e", font=("Segoe UI", 10, "italic")).pack(side="right", padx=6)
        tk.Label(t, text="v5.9.2  |  ROW IB", fg="#8888aa", bg="#16213e", font=("Segoe UI", 10)).pack(side="right", padx=16)
        leg = tk.Frame(self.root, bg="#1a1a2e", height=30); leg.pack(fill="x")
        for tx, f, b in [("Claiming","white","#0f0f1a"),("Dominant","black","#E2EFDA"),("Sub-rows","black","#EBF3FB"),("Root/Short","#9C0006","#FFE0E0"),("DICES","black","#FFF2CC"),("SR","#CC0000","#FFD7D7"),("Invalid inv","#333","#FFD0D0"),("REBNI","#333","#D0F0FF"),("Cross PO","#7a5c00","#FFF0C0"),("Mismatch","#333","#D0E8FF")]:
            tk.Label(leg, text=f"  {tx}  ", fg=f, bg=b, font=("Segoe UI", 8, "bold"), padx=8).pack(side="left", padx=3, pady=3)
        body = tk.Frame(self.root, bg="#0d0d1a", padx=24, pady=12); body.pack(fill="both", expand=True)
        inp = tk.LabelFrame(body, text="  Input Files  (Excel .xlsx or CSV .csv supported)  ", fg="#4a9eff", bg="#0d0d1a", font=("Segoe UI", 10, "bold"), padx=12, pady=8); inp.pack(fill="x", pady=6)
        self._f_row(inp, "Claims Sheet:", self.claims_path, 0); self._f_row(inp, "REBNI Result:", self.rebni_path, 1); self._f_row(inp, "Invoice Search:", self.inv_path, 2)
        tf = tk.Frame(body, bg="#0f0f1a"); tf.pack(anchor="w", pady=4); tk.Label(tf, text="Ticket ID:", fg="white", bg="#0f0f1a", font=("Segoe UI", 11)).pack(side="left"); tk.Entry(tf, textvariable=self.ticket_id, width=28, font=("Segoe UI", 11), bg="#1e1e3a", fg="white", insertbackground="white", relief="flat").pack(side="left", padx=8)
        m = tk.LabelFrame(body, text="Investigation Mode", fg="white", bg="#0f0f1a", padx=10, pady=5); m.pack(fill="x", pady=8); tk.Radiobutton(m, text="AUTO  —  Automatic. SID popup when not found in REBNI.", variable=self.mode_var, value="auto", fg="white", bg="#0f0f1a", selectcolor="#16213e", font=("Segoe UI", 10)).pack(anchor="w", padx=10); tk.Radiobutton(m, text="MANUAL  —  One level at a time. Live preview. Parallel interaction enabled.", variable=self.mode_var, value="manual", fg="white", bg="#0f0f1a", selectcolor="#16213e", font=("Segoe UI", 10)).pack(anchor="w", padx=10)
        tt = tk.LabelFrame(body, text="Ticket Type", fg="white", bg="#0f0f1a", padx=10, pady=5); tt.pack(fill="x", pady=6); tk.Radiobutton(tt, text="PDTT  —  Full chain investigation across all shipments (default).", variable=self.ticket_type_var, value="PDTT", fg="white", bg="#0f0f1a", selectcolor="#16213e", font=("Segoe UI", 10)).pack(anchor="w", padx=10); tk.Radiobutton(tt, text="REMASH TT  —  Claiming shipment level investigation ONLY. No sub-shipment chain tracing.", variable=self.ticket_type_var, value="REMASH", fg="#f0c060", bg="#0f0f1a", selectcolor="#1a1500", font=("Segoe UI", 10)).pack(anchor="w", padx=10)
        self.status = tk.Label(body, text="Ready", fg="#4a9eff", bg="#0f0f1a", font=("Segoe UI", 10)); self.status.pack(pady=(10, 0)); self.pb = ttk.Progressbar(body, mode='determinate'); self.pb.pack(fill="x", pady=4)
        bf = tk.Frame(body, bg="#0f0f1a"); bf.pack(pady=10)
        self.run_btn = tk.Button(bf, text="▶  RUN INVESTIGATION", bg="#e94560", fg="white", font=("Segoe UI", 15, "bold"), padx=36, pady=14, relief="flat", cursor="hand2", command=self.start_run); self.run_btn.pack(side="left", padx=10)
        self.stop_inv_btn = tk.Button(bf, text="⏸  STOP INVESTIGATION", bg="#4a2020", fg="white", font=("Segoe UI", 11, "bold"), padx=16, pady=14, relief="flat", state="disabled", cursor="hand2", command=self.request_stop_investigation); self.stop_inv_btn.pack(side="left", padx=6)
        self.stop_sess_btn = tk.Button(bf, text="⏹  STOP SESSION", bg="#3a0000", fg="white", font=("Segoe UI", 11, "bold"), padx=16, pady=14, relief="flat", state="disabled", cursor="hand2", command=self.request_stop_session); self.stop_sess_btn.pack(side="left", padx=6)
        self.save_btn = tk.Button(bf, text="💾  SAVE OUTPUT", bg="#2d6a4f", fg="white", font=("Segoe UI", 13, "bold"), padx=28, pady=14, relief="flat", state="normal", cursor="hand2", command=self.save_output); self.save_btn.pack(side="left", padx=10)
        
        # --- v5.9.2 INTEGRATION: UNIQUE SUMMARY PORTAL ---
        self.portal_btn = tk.Button(bf, text="📑  UNIQUE SUMMARY PORTAL", bg="#1c2c42", fg="#4a9eff", font=("Segoe UI", 11, "bold"), padx=16, pady=14, relief="flat", cursor="hand2", command=self.open_summary_portal)
        self.portal_btn.pack(side="left", padx=6)

    def _f_row(self, p, l, v, r):
        tk.Label(p, text=l, fg="#cccccc", bg="#131320", width=18, anchor="w", font=("Segoe UI", 10)).grid(row=r, column=0, sticky="w", pady=3); tk.Entry(p, textvariable=v, width=62, font=("Segoe UI", 10), bg="#1e1e3a", fg="white", insertbackground="white", relief="flat").grid(row=r, column=1, padx=6); tk.Button(p, text="Browse", command=lambda: v.set(filedialog.askopenfilename(filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv")])), bg="#2d2d5e", fg="white", relief="flat", cursor="hand2", padx=8).grid(row=r, column=2)

    def _set_status(self, msg, pct=None):
        self.root.after(0, lambda: (self.status.config(text=msg), (self.pb.__setitem__('value', pct) if pct is not None else None)))

    def start_run(self):
        if not all([self.claims_path.get(), self.rebni_path.get(), self.inv_path.get()]):
            messagebox.showerror("Error", "Please select all 3 input files."); return
        self.run_btn.config(state="disabled"); self.save_btn.config(state="disabled"); self.portal_btn.config(state="disabled"); self.stop_inv_btn.config(state="normal"); self.stop_sess_btn.config(state="normal"); self.all_blocks = []
        if self.mode_var.get() == "manual":
            if not self.preview or not self.preview.winfo_exists(): self.preview = PreviewPanel(self.root); self.preview._app = self
            else: self.preview.clear_all()
        threading.Thread(target=self._process, daemon=True).start()

    def request_stop_investigation(self):
        if hasattr(self, 'engine'): self.engine.stop_requested = True
        self._set_status("Investigation paused — current results preserved. Save or resume."); self.root.after(0, lambda: (self.save_btn.config(state="normal"), self.portal_btn.config(state="normal")))

    def request_stop_session(self):
        if hasattr(self, 'engine'): self.engine.stop_requested = True
        self._set_status("Session ended — saving current results."); self._finish()

    def _process(self):
        try:
            self._set_status("Loading Claims file…", 5); df_c = load_claims(self.claims_path.get()); mc, corr = detect_claim_cols(df_c)
            if corr or [f for f in COLUMN_ALIASES if f not in mc]:
                conf, done = [None], threading.Event(); self.root.after(0, lambda: HeaderCorrectionDialog(self.root, corr, mc, list(df_c.columns), lambda res: (conf.__setitem__(0, res['mapping']) if res['action'] == 'proceed' else None, done.set()))); done.wait()
                if conf[0] is None: self.root.after(0, lambda: (self.run_btn.config(state="normal"), self.stop_inv_btn.config(state="disabled"), self.stop_sess_btn.config(state="disabled"))); return
                mc = conf[0]
            self._set_status("Loading REBNI…", 12); rp, rs, rfb = build_rebni_index(load_rebni(self.rebni_path.get())); self._set_status("Loading Invoice Search…", 30); ip, ifb, iam = build_invoice_index(load_invoice_search(self.inv_path.get())); self.engine = InvestigationEngine(rp, rs, rfb, ip, ifb, iam, self._req_sid); tot = len(df_c)
            if self.mode_var.get() == "auto":
                self.preview = PreviewPanel(self.root); self.preview._app = self
                for i, (_, r) in enumerate(df_c.iterrows()):
                    if self.engine.stop_requested: break
                    self._set_status(f"Auto: {i+1}/{tot}  ASIN: {clean(r.get(mc.get('ASIN',''),''))}", 60 + int((i / max(tot, 1)) * 35)); self.preview.add_header_row(f"{i+1}/{tot}: {clean(r.get(mc.get('ASIN',''),''))}")
                    rows, _ = self.engine.run_auto(clean(r.get(mc.get('Barcode', ''), '')), clean(r.get(mc.get('Invoice', ''), '')), extract_sid(clean(r.get(mc.get('SID', ''), ''))), clean(r.get(mc.get('PO', ''), '')), clean(r.get(mc.get('ASIN', ''), '')), safe_num(r.get(mc.get('InvQty', ''), 0)), safe_num(r.get(mc.get('PQV', ''), 0)), initial_cp=safe_num(r.get(mc.get('CP', ''), 0)), row_callback=lambda row: (self.preview.add_row(row), self.root.update()))
                    self.all_blocks.append(rows)
                self._finish()
            else: self.manual_q, self.map_cols = df_c.to_dict('records'), mc; self._next_man()
        except Exception as e:
            import traceback; tb = traceback.format_exc(); self.root.after(0, lambda: messagebox.showerror("Error", f"{e}\n\n{tb}")); self._finish()

    def _req_sid(self, inv, po, asin):
        if inv in self.engine.cache_sid: return self.engine.cache_sid[inv]
        res, done = [None], threading.Event(); self.root.after(0, lambda: SIDRequestDialog(self.root, inv, po, asin, lambda s: (self.engine.cache_sid.__setitem__(inv, s) if s else None, res.__setitem__(0, s), done.set()))); done.wait(); return res[0]

    def _next_man(self, force_next=False):
        if hasattr(self, '_pending_cross_po') and self._pending_cross_po:
            p = self._pending_cross_po.pop(0)
            def inv_p():
                cr, fn = self.engine.run_cross_po_investigation(p['candidate'], p['case'], p['budget'], depth=self.curr_m['depth']+1)
                for r in cr:
                    self.curr_m['block'].append(r)
                    if self.preview and self.preview.winfo_exists(): self.root.after(0, lambda row=r: self.preview.add_row(row))
                self._man_step()
            threading.Thread(target=inv_p, daemon=True).start(); return
        if hasattr(self, 'curr_m') and not force_next:
            if self._collect_all_pending(): self.root.after(0, self._show_pending_gateway); return
        if not self.manual_q or self.engine.stop_requested: self._finish(); return
        r, mc = self.manual_q.pop(0), self.map_cols
        self.curr_m = {'b':clean(r.get(mc.get('Barcode',''),'')), 'i':clean(r.get(mc.get('Invoice',''),'')), 's':extract_sid(clean(r.get(mc.get('SID',''),''))), 'p':clean(r.get(mc.get('PO',''),'')), 'a':clean(r.get(mc.get('ASIN',''),'')), 'iq':safe_num(r.get(mc.get('InvQty',''),0)), 'pqv':safe_num(r.get(mc.get('PQV',''),0)), 'rem':safe_num(r.get(mc.get('PQV',''),0)), 'budget':safe_num(r.get(mc.get('PQV',''),0)), 'depth':0, 'block':[], 'processed':set(), 'asin_rendered_levels':set(), 'is_new_block':True, 'rendered':False, 'siblings_stack':[], 'all_seen_matches':[], 'initial_cp':safe_num(r.get(mc.get('CP',''),0))}
        self.preview.add_header_row(self.curr_m['a']); threading.Thread(target=self._man_step, daemon=True).start()

    def _man_step(self):
        if self.engine.stop_requested: self._finish(); return
        m = self.curr_m; rows, matches, rq, n_rem = self.engine.build_one_level(m['b'], m['i'], m['s'], m['p'], m['a'], m['iq'], m['rem'], m['depth'], is_claiming=(m['depth']==0), is_manual=True, initial_cp=m.get('initial_cp',0.0))
        k = (m['depth'], clean(m['s']), clean(m['p']), clean(m['a']), clean(m['i']))
        if not m['rendered'] and k not in m['asin_rendered_levels']:
            if m.get('is_new_block') and rows: rows[0]['is_new_block']=True; m['is_new_block']=False
            m['block'].extend(rows); [self.preview.add_row(r) for r in rows]; self.root.update(); m['asin_rendered_levels'].add(k); m['rendered']=True
        elif not m['rendered']: [self.preview.add_row(r) for r in rows[1:]]; m['rendered']=True
        if not m.get('cross_po_checked'):
            m['cross_po_checked']=True; cands=self.engine.detect_cross_po(m['s'], m['p'], m['a'])
            if cands: m['_awaiting_cross_po']=True; self.root.after(0, lambda: CrossPODialog(self.root, cands, m['i'], m['s'], lambda res: self._handle_cross_po_and_finish(res))); return
        for mt in matches:
            aug = dict(mt); aug['_depth'] = m.get('depth',0); m.setdefault('all_seen_matches',[]).append(aug)
        m['rem']=n_rem; matches=[x for x in matches if self._get_loop_key(x) not in m['processed']]; rem_s=rows[0].get('remarks','') if rows else ''
        if not matches: rem_s="No unprocessed matches remaining"
        if not matches or any(kw in rem_s for kw in ["Root cause", "REBNI", "SR", "short received directly", "Direct Shortage", "Phase 1", "No Invoice Search"]):
            if m.get('siblings_stack'):
                ctx=m['siblings_stack'].pop(0); m.update({'depth':ctx['depth'], 'rem':ctx['rem'], 'budget':ctx['budget'], 'b':ctx['b'], 'i':ctx['i'], 'iq':ctx['iq'], 's':ctx['s'], 'p':ctx['p'], 'a':ctx['a'], 'processed':ctx['processed'], 'rendered':False}); self.root.after(0, lambda: self._show_dlg(ctx['siblings'])); return
            if hasattr(self, '_pending_cross_po') and self._pending_cross_po: self.root.after(0, self._next_man); return
            self.all_blocks.append(m['block']); self._next_man(); return
        self.root.after(0, lambda: self._show_dlg(matches))

    def _show_dlg(self, matches):
        m, f = self.curr_m, matches[0]
        if self.engine.user_overrides:
            pm = []
            for mt in matches:
                ov = self.engine.user_overrides.get(clean(mt.get('mtc_inv','')), {})
                if ov:
                    mt=dict(mt)
                    if 'mtc_qty' in ov: mt['mtc_qty']=ov['mtc_qty']
                    if 'inv_qty' in ov: mt['inv_qty']=ov['inv_qty']
                pm.append(mt)
            matches, f = pm, pm[0]
        if f['mtc_inv'] in self.engine.cache_sid: self._handle_res({'action':'valid','chosen_match':f,'sid':self.engine.cache_sid[f['mtc_inv']],'barcode':self.engine.cache_bc.get(f['mtc_inv'],"[DICES]")}, matches); return
        if getattr(self, 'active_manual_dlg', None) and self.active_manual_dlg.winfo_exists(): self.active_manual_dlg.destroy()
        self.active_manual_dlg = ManualLevelDialog(self.root, matches, m['rem'], m['budget'], lambda rs: self._handle_res(rs, matches), pending_cb=self._show_pending_invoices_from_dialog, engine=self.engine)

    def _handle_res(self, res, matches):
        if not res or res['action'] == 'stop': self.all_blocks.append(self.curr_m['block']); self._next_man(); return
        mc = res.get('chosen_match')
        if mc: self.curr_m['processed'].add(self._get_loop_key(mc))
        if res['action'] == 'invalid':
            ex = res['invalid_qty']; ro = self.engine._make_row('[INVALID]', mc['mtc_inv'], '—', mc['mtc_po'], mc['mtc_asin'], mc['inv_qty'], '', '', '', f"{int(ex)} units matched to invalid invoice {mc['mtc_inv']} — excluded from PQV", '', self.curr_m['depth'], 'subrow')
            self.curr_m['block'].append(ro); self.preview.add_row(ro); self.curr_m['rem']=max(0, self.curr_m['rem']-ex)
            if self.curr_m['rem']<=0: self.all_blocks.append(self.curr_m['block']); self._next_man()
            else:
                rm = [x for x in matches if x['mtc_inv'] != mc['mtc_inv']]
                if rm: self.root.after(0, lambda: self._show_dlg(rm))
                else: self.all_blocks.append(self.curr_m['block']); self._next_man()
        elif res['action'] == 'cross_po':
            cands = self.engine.detect_cross_po(self.curr_m['s'], self.curr_m['p'], self.curr_m['a'])
            if cands: CrossPODialog(self.root, cands, self.curr_m['i'], self.curr_m['s'], lambda r: self._handle_cross_po(r))
            else: messagebox.showinfo("No Cross PO", "No Cross PO candidates found."); self.root.after(0, lambda: self._show_dlg(matches))
        elif res['action'] == 'mismatch':
            d = res['mismatch_data']; bug = safe_num(d.get('ovg_qty', 0)) or self.curr_m['rem']
            rows, sm, rq, sh, nr = self.engine.run_mismatch_investigation(d, bug, depth=self.curr_m['depth'])
            for r in rows: self.curr_m['block'].append(r) or self.preview.add_row(r)
            if sm: self.root.after(0, lambda: self._show_dlg(sm))
            else:
                rm = [x for x in matches if x != res.get('chosen_match')]
                if rm: self.root.after(0, lambda: self._show_dlg(rm))
                else: self.all_blocks.append(self.curr_m['block']); self._next_man()
        else:
            self.engine.cache_sid[mc['mtc_inv']], self.engine.cache_bc[mc['mtc_inv']] = res['sid'], res['barcode']
            rem_s = [x for x in matches if x['mtc_inv'] != mc['mtc_inv'] and self._get_loop_key(x) not in self.curr_m['processed']]
            ps = list(self.curr_m['siblings_stack'])
            if rem_s: ps.append({'siblings':rem_s, 'depth':self.curr_m['depth'], 'rem':self.curr_m['rem'], 'budget':self.curr_m['budget'], 'b':self.curr_m.get('b',''), 'i':self.curr_m.get('i',''), 'iq':self.curr_m.get('iq',0), 's':self.curr_m['s'], 'p':self.curr_m['p'], 'a':self.curr_m['a'], 'processed':self.curr_m['processed'], 'cross_po_checked':self.curr_m.get('cross_po_checked',False)})
            bb = safe_num(mc['mtc_qty']) or self.curr_m['rem']
            self.curr_m.update({'b':res['barcode'], 'i':mc['mtc_inv'], 's':res['sid'], 'p':mc['mtc_po'], 'a':mc['mtc_asin'], 'iq':mc['inv_qty'], 'rem':bb, 'budget':bb, 'depth':self.curr_m['depth']+1, 'rendered':False, 'processed':self.curr_m['processed'], 'siblings_stack':ps, 'pending_siblings':[], 'cross_po_checked':False, '_awaiting_cross_po':False})
            threading.Thread(target=self._man_step, daemon=True).start()

    def _collect_all_pending(self):
        m, proc, seen, ded = self.curr_m, self.curr_m.get('processed', set()), set(), []
        for inv in m.get('all_seen_matches', []):
            k = self._get_loop_key(inv)
            if k and k not in seen and k not in proc:
                seen.add(k)
                if hasattr(self, 'engine') and self.engine.user_overrides:
                    ov = self.engine.user_overrides.get(clean(k), {})
                    if ov:
                        inv=dict(inv)
                        if 'mtc_qty' in ov: inv['mtc_qty']=ov['mtc_qty']
                        if 'inv_qty' in ov: inv['inv_qty']=ov['inv_qty']
                ded.append(inv)
        return ded

    def _show_pending_invoices_from_dialog(self):
        ap = self._collect_all_pending()
        if not ap: messagebox.showinfo("No Pending Invoices", "All matched invoices have already been investigated or no matches were found yet.", parent=self.root); return
        if getattr(self, 'active_pending_dlg', None) and self.active_pending_dlg.winfo_exists(): self.active_pending_dlg.destroy()
        self.active_pending_dlg = PendingInvoicesDialog(self.root, ap, self.curr_m.get('a', ''), lambda res: (self.curr_m['processed'].add(self._get_loop_key(res['match'])), self.root.after(0, lambda: self._show_dlg([res['match']]))) if res['action']=='investigate' else None)

    def _show_pending_gateway(self):
        ap = self._collect_all_pending()
        if not ap: self._next_man(force_next=True); return
        if getattr(self, 'active_pending_dlg', None) and self.active_pending_dlg.winfo_exists(): self.active_pending_dlg.destroy()
        self.active_pending_dlg = PendingInvoicesDialog(self.root, ap, f"{self.curr_m.get('a', '')} (Conclusion Review)", lambda res: (self.curr_m['processed'].add(self._get_loop_key(res['match'])), self.root.after(0, lambda m=res['match']: self._show_dlg([m]))) if res['action']=='investigate' else (self.all_blocks.append(self.curr_m['block']) if self.curr_m.get('block') else None, self._next_man(force_next=True)))

    def _handle_cross_po(self, res):
        if res['action'] == 'skip': threading.Thread(target=self._man_step, daemon=True).start(); return
        if not hasattr(self, '_pending_cross_po'): self._pending_cross_po = []
        self._pending_cross_po.append({'candidate':res['candidate'], 'case':res.get('case','Case 1'), 'budget':safe_num(res['candidate']['rec_qty'])})
        self._set_status(f"Cross PO stored ({res['candidate']['po']}) — continuing normal investigation.", None); threading.Thread(target=self._man_step, daemon=True).start()

    def _handle_cross_po_and_finish(self, res):
        if res['action'] == 'skip': self.all_blocks.append(self.curr_m['block']); self._next_man(); return
        c, bud = res['candidate'], safe_num(res['candidate']['rec_qty'])
        self._set_status(f"Cross PO confirmed ({c['po']}) — starting manual investigation of {int(bud)} units…", None); self.curr_m.pop('_awaiting_cross_po', None)
        self.curr_m.update({'b':'', 'i':'', 's':c['sid'], 'p':c['po'], 'a':c['asin'], 'iq':0, 'rem':bud, 'budget':bud, 'depth':self.curr_m['depth']+1, 'rendered':False, 'processed':set(), 'cross_po_checked':True, 'asin_rendered_levels':set()}); threading.Thread(target=self._man_step, daemon=True).start()

    def _finish(self):
        msg = "Investigation complete!" if not (hasattr(self, 'engine') and self.engine.stop_requested) else "Investigation stopped by user."
        self._set_status("Complete. Click SAVE.", 100); self.root.after(0, lambda: (self.run_btn.config(state="normal"), self.save_btn.config(state="normal"), self.portal_btn.config(state="normal"), self.stop_inv_btn.config(state="disabled"), self.stop_sess_btn.config(state="disabled"), messagebox.showinfo("Done", msg)))

    def _get_loop_key(self, mt): return (clean(mt.get('mtc_inv','')), clean(mt.get('mtc_asin','')), clean(mt.get('mtc_po','')), fmt_qty(mt.get('mtc_qty',0)))

    def save_output(self):
        t, ts = self.ticket_id.get().strip().replace(' ','_'), datetime.now().strftime('%Y%m%d_%H%M%S'); o = f"MFI_{t}_{ts}.xlsx" if t else f"MFI_Investigation_{ts}.xlsx"; p = os.path.join(os.path.dirname(self.claims_path.get()) or os.getcwd(), o)
        try:
            bl = self.all_blocks
            if self.mode_var.get() == "manual" and self.preview and self.preview.winfo_exists():
                ar, fb, cur = self.preview.get_all_rows(), [], []
                for r in ar:
                    if r.get('is_new_block') and cur: fb.append(cur); cur=[]
                    cur.append(r)
                if cur: fb.append(cur)
                bl = fb
            write_excel(bl, p); messagebox.showinfo("Saved", f"Saved to:\n{p}")
        except Exception as e: messagebox.showerror("Save Error", str(e))

    def open_summary_portal(self):
        """v5.9.2 Integration: Automatically launch the embedded HTML portal."""
        # ── v5.9.2 BUNDLING LOGIC ───────────────────────────────────────────
        # This allows the tool to find the HTML file whether running as a 
        # script or as a bundled EXE (via sys._MEIPASS).
        if hasattr(sys, '_MEIPASS'):
            # Running as a PyInstaller bundle
            base_dir = sys._MEIPASS
        else:
            # Running as a normal script
            base_dir = os.path.dirname(os.path.abspath(__file__))
            
        portal_path = os.path.join(base_dir, "MFI_unique_summary_upload_export.html")
        
        if os.path.exists(portal_path):
            self._set_status("Launching Unique Summary Portal...")
            # Use as_uri() to handle spaces and special characters safely in all browsers
            from pathlib import Path
            webbrowser.open_new_tab(Path(portal_path).as_uri())
        else:
            messagebox.showerror("File Not Found", 
                                f"Could not find the portal file:\n{portal_path}\n\n"
                                f"If you are using the EXE version, ensure it was compiled with the portal data.")

    def run(self):
        try: self.root.mainloop()
        except KeyboardInterrupt: pass

if __name__ == '__main__': MFIToolApp().run()
