"""
MFI Investigation Tool  v5.0.4  |  Roy B Workflow Version
========================================================
ROW IB  |  Amazon
Developed by Mukesh

CHANGES IN v5.0.4 (Roy B Workflow):
  ✔ PHASE 1: Implemented Direct Shortage gateway (Shortage >= PQV).
  ✔ PHASE 1-3: Implemented accounting formula: Total Accounted = Shortage + REBNI + EX.
  ✔ PHASE 2: Position-matching rule reinforced for (SID, PO, ASIN) matching in Invoice Search.
  ✔ PHASE 2: Sequential Investigation Rule (siblings_stack) reinforced for complete accounting.
  ✔ PHASE 4: Added Phase 4 flagging for Cross PO and Mismatch ASIN overages.
  ✔ REMARKS: Detailed accounting breakdown added to remarks.
  ✔ EXCEL: Maintained single-sheet output (Phase 6 multi-sheet excluded per request).
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, re, threading
from datetime import datetime


# ═══════════════════════════════════════════════════════════
#  UTILITIES
# ═══════════════════════════════════════════════════════════

def extract_sid(val):
    s = str(val).strip()
    parts = re.findall(r'\d{10,}', s)
    return max(parts, key=len) if parts else s

def strip_scr(inv_no):
    import re
    return re.sub(r'(?:SCR)+$', '', str(inv_no).strip(), flags=re.IGNORECASE)

def safe_num(val):
    try:
        if pd.isna(val): return 0.0
    except: pass
    try:
        return float(str(val).replace(',', '').strip())
    except: return 0.0

def clean(val):
    try:
        if pd.isna(val): return ""
    except: pass
    return str(val).strip()

def split_comma(val):
    if not val: return []
    try:
        if pd.isna(val): return []
    except: pass
    return [s.strip() for s in str(val).split(',') if s.strip()]

def fmt_qty(val):
    n = safe_num(val)
    if n == 0: return ""
    return str(int(n)) if n == int(n) else str(n)


# ═══════════════════════════════════════════════════════════
#  DIALOGS
# ═══════════════════════════════════════════════════════════


class HeaderCorrectionDialog(tk.Toplevel):
    def __init__(self, parent, corrections, mapping, df_columns, callback):
        super().__init__(parent)
        self.callback    = callback
        self.corrections = corrections
        self.mapping     = mapping
        self.df_columns  = df_columns
        self.title("Column Header Mismatch Detected — v5.0.4")
        self.geometry("700x480")
        self.configure(bg="#0f0f1a")
        self.resizable(True, True)
        self.lift(); self.focus_force()

        tk.Label(self, text="⚠  Non-standard column headers detected in Claims file",
                 bg="#16213e", fg="#f0a500", font=("Segoe UI",12,"bold"), height=2).pack(fill="x")
        tk.Label(self, text="The tool has automatically matched the columns below.",
                 bg="#0f0f1a", fg="#cccccc", font=("Segoe UI",9)).pack(pady=4)

        outer = tk.Frame(self, bg="#0f0f1a"); outer.pack(fill="both", expand=True, padx=16, pady=6)
        hdrs = ["Field", "Expected", "Found in file", "Status"]
        w = [14, 20, 28, 12]
        for ci, (h, ww) in enumerate(zip(hdrs, w)):
            tk.Label(outer, text=h, bg="#203864", fg="white", font=("Calibri",10,"bold"), width=ww, anchor="w", padx=4).grid(row=0, column=ci, padx=1, pady=1, sticky="w")

        all_fields = list(COLUMN_ALIASES.keys())
        corrected_fields = {c[0] for c in corrections}
        self._override_vars = {}

        for ri, field in enumerate(all_fields, 1):
            canonical = COLUMN_ALIASES[field][0]
            found_col = mapping.get(field, "")
            is_corrected = field in corrected_fields
            is_missing   = field not in mapping
            if is_missing: status_txt, status_fg, row_bg = "MISSING", "#ff4444", "#2a0000"
            elif is_corrected: status_txt, status_fg, row_bg = "Auto-fixed", "#f0a500", "#1a1500"
            else: status_txt, status_fg, row_bg = "✔ OK", "#44ff88", "#001a00"

            tk.Label(outer, text=field, bg=row_bg, fg="#e0e0e0", font=("Calibri",10,"bold"), width=14, anchor="w", padx=4).grid(row=ri, column=0, padx=1, pady=1, sticky="w")
            tk.Label(outer, text=canonical, bg=row_bg, fg="#aaaacc", font=("Calibri",10), width=20, anchor="w", padx=4).grid(row=ri, column=1, padx=1, pady=1, sticky="w")
            v = tk.StringVar(value=found_col or "— not found —"); self._override_vars[field] = v
            cb = ttk.Combobox(outer, textvariable=v, values=["— not found —"] + list(df_columns), state="readonly", width=26, font=("Calibri",9)); cb.grid(row=ri, column=2, padx=1, pady=1, sticky="w")
            tk.Label(outer, text=status_txt, bg=row_bg, fg=status_fg, font=("Calibri",10,"bold"), width=12, anchor="w", padx=4).grid(row=ri, column=3, padx=1, pady=1, sticky="w")

        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=10)
        tk.Button(bf, text="✔  Auto-correct & Proceed", command=self._proceed, bg="#2d6a4f", fg="white", font=("Segoe UI",12,"bold"), padx=20, pady=8, relief="flat", cursor="hand2").pack(side="left", padx=10)
        tk.Button(bf, text="✖  Cancel", command=self._cancel, bg="#4a2020", fg="white", font=("Segoe UI",11), padx=16, pady=8, relief="flat", cursor="hand2").pack(side="left", padx=10)
        self.protocol("WM_DELETE_WINDOW", self._cancel)
        self.update_idletasks()
        px, py = parent.winfo_x() + (parent.winfo_width() - self.winfo_width()) // 2, parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _proceed(self):
        final = {f: v.get() for f, v in self._override_vars.items() if v.get() and v.get() != "— not found —"}
        self.callback({'action': 'proceed', 'mapping': final}); self.destroy()

    def _cancel(self): self.callback({'action': 'cancel'}); self.destroy()

class SIDRequestDialog(tk.Toplevel):
    def __init__(self, parent, invoice, po, asin, callback):
        super().__init__(parent); self.callback = callback
        self.title("SID Required — DICES Validation"); self.geometry("540x210"); self.resizable(True, True); self.configure(bg="#16213e"); self.lift(); self.focus_force()
        tk.Label(self, text="⚠  SID Not Found in REBNI", bg="#16213e", fg="#e94560", font=("Segoe UI",13,"bold")).pack(pady=(14,4))
        tk.Label(self, text=f"Invoice: {invoice}   PO: {po}   ASIN: {asin}", bg="#16213e", fg="#e0e0e0", font=("Segoe UI",9)).pack(pady=2)
        tk.Label(self, text="Validate this invoice in DICES and enter the SID below:", bg="#16213e", fg="#aaaacc", font=("Segoe UI",9)).pack(pady=6)
        ef = tk.Frame(self, bg="#16213e"); ef.pack()
        tk.Label(ef, text="SID from DICES:", bg="#16213e", fg="#e0e0e0", font=("Segoe UI",10)).pack(side="left", padx=8)
        self._sid = tk.StringVar(); self._entry = tk.Entry(ef, textvariable=self._sid, width=30, font=("Segoe UI",10), bg="#1e1e3a", fg="#e0e0e0", insertbackground="white", relief="flat"); self._entry.pack(side="left", padx=4); self._entry.focus_set()
        bf = tk.Frame(self, bg="#16213e"); bf.pack(pady=12)
        tk.Button(bf, text="✔  Continue", command=self._ok, bg="#2d6a4f", fg="white", font=("Segoe UI",11,"bold"), padx=16, pady=7, relief="flat", cursor="hand2").pack(side="left", padx=8)
        tk.Button(bf, text="✖  Skip", command=self._skip, bg="#6b2737", fg="white", font=("Segoe UI",10), padx=16, pady=7, relief="flat", cursor="hand2").pack(side="left", padx=8)
        self.bind('<Return>', lambda e: self._ok()); self.protocol("WM_DELETE_WINDOW", self._skip)
        self.update_idletasks()
        px, py = parent.winfo_x() + (parent.winfo_width() - self.winfo_width()) // 2, parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _ok(self):
        sid = extract_sid(self._sid.get().strip())
        if sid: self.callback(sid); self.destroy()
        else: self._entry.config(bg="#3a1e1e")

    def _skip(self): self.callback(None); self.destroy()


class CrossPODialog(tk.Toplevel):
    CASE_DESCRIPTIONS = {
        "Case 1": ("Case 1 — No PO, but ASIN received", "Rec=0 at claiming PO. Same ASIN received in different PO within same SID.\nThose units are overage under a different PO."),
        "Case 2": ("Case 2 — PO exists but ASIN not ordered there", "This PO exists in the claiming SID, but the ASIN was never invoiced there.\nInv Qty = 0, but units were received. This is a Cross PO overage."),
        "Case 3": ("Case 3 — PO and ASIN exist but Rec > Inv", "Both PO and ASIN are present. Invoiced qty = X, but received more than X.\nExcess units are Cross PO overage."),
    }

    def __init__(self, parent, candidates, current_inv, sid, callback):
        super().__init__(parent); self.callback = callback; self.candidates = candidates
        self.title("Cross PO Overage — Phase 4"); self.geometry("740x540"); self.resizable(True, True); self.configure(bg="#0f0f1a"); self.lift(); self.focus_force()
        tk.Label(self, text="🔄  Cross PO Overage Detected (Phase 4)", bg="#16213e", fg="#f0a500", font=("Segoe UI",13,"bold"), height=2).pack(fill="x")
        tk.Label(self, text=f"SID: {sid}   |   Investigation Invoice: {current_inv}", bg="#0f0f1a", fg="#cccccc", font=("Segoe UI",9)).pack(pady=2)
        
        tf = tk.LabelFrame(self, text="  Detected Cross PO Candidates  ", bg="#0f0f1a", fg="#e0e0e0", font=("Segoe UI",9,"bold"), padx=10, pady=6); tf.pack(fill="x", padx=16, pady=6)
        for ci, h in enumerate(["Cross PO","ASIN","Inv Qty","Rec Qty","Overage","Type"]): tk.Label(tf, text=h, bg="#203864", fg="white", font=("Calibri",10,"bold"), width=14, anchor="w", padx=3).grid(row=0, column=ci, padx=1, pady=1)
        for ri, c in enumerate(candidates, 1):
            inv_n, rec_n = safe_num(c.get('inv_qty', 0)), safe_num(c['rec_qty']); ovg = max(0.0, rec_n - inv_n)
            for ci, v in enumerate([c['po'], c['asin'], fmt_qty(inv_n), fmt_qty(rec_n), fmt_qty(ovg) or "—", c['cross_type'].split("—")[0].strip()]):
                tk.Label(tf, text=str(v), bg="#1e1e3a", fg="#e0e0e0", font=("Calibri",10), width=14, anchor="w", padx=3).grid(row=ri, column=ci, padx=1, pady=1)

        sf = tk.Frame(self, bg="#0f0f1a"); sf.pack(fill="x", padx=16, pady=4)
        tk.Label(sf, text="Select Cross PO to investigate:", bg="#0f0f1a", fg="#e0e0e0", font=("Segoe UI",10), width=30, anchor="w").pack(side="left")
        opts = [f"PO={c['po']}  Rec={fmt_qty(c['rec_qty'])}  {c['cross_type'].split(chr(8212))[0].strip()}" for c in candidates] + ["None — Skip"]
        self._sel_cb = ttk.Combobox(sf, values=opts, state="readonly", width=50, font=("Segoe UI",9)); self._sel_cb.current(0); self._sel_cb.pack(side="left", padx=6)

        cf = tk.LabelFrame(self, text="  Confirm Cross PO Case  ", bg="#0f0f1a", fg="#e0e0e0", font=("Segoe UI",9,"bold"), padx=12, pady=8); cf.pack(fill="x", padx=16, pady=4)
        self._case_var = tk.StringVar(value="Case 1"); self._case_desc_lbl = tk.Label(cf, text="", bg="#0f0f1a", fg="#aaaacc", font=("Segoe UI",9), justify="left", wraplength=640, anchor="w")
        for case_key, (case_label, _) in self.CASE_DESCRIPTIONS.items(): tk.Radiobutton(cf, text=case_label, variable=self._case_var, value=case_key, bg="#0f0f1a", fg="#f0c060", selectcolor="#1a1500", font=("Segoe UI",10,"bold"), command=self._on_case_change).pack(anchor="w", pady=2)
        self._case_desc_lbl.pack(anchor="w", pady=4, padx=8); self._on_case_change()

        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=10)
        tk.Button(bf, text="✔  Confirm & Investigate", command=self._confirm, bg="#2d6a4f", fg="white", font=("Segoe UI",12,"bold"), padx=20, pady=9, relief="flat", cursor="hand2").pack(side="left", padx=10)
        tk.Button(bf, text="✖  Skip", command=self._skip, bg="#4a2020", fg="white", font=("Segoe UI",11), padx=16, pady=9, relief="flat", cursor="hand2").pack(side="left", padx=10)
        self.protocol("WM_DELETE_WINDOW", self._skip); self.update_idletasks()
        px, py = parent.winfo_x() + (parent.winfo_width() - self.winfo_width()) // 2, parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _on_case_change(self): self._case_desc_lbl.config(text=self.CASE_DESCRIPTIONS.get(self._case_var.get(), ("", ""))[1])
    def _confirm(self):
        idx = self._sel_cb.current()
        if idx >= len(self.candidates): self.callback({'action': 'skip'}); self.destroy(); return
        self.callback({'action':'confirmed', 'candidate':self.candidates[idx], 'case':self._case_var.get()}); self.destroy()
    def _skip(self): self.callback({'action': 'skip'}); self.destroy()


class ManualLevelDialog(tk.Toplevel):
    def __init__(self, parent, matches, remaining_pqv, branch_budget, callback):
        super().__init__(parent); self.callback = callback; self.matches = matches; self.rem_pqv = remaining_pqv; self.branch_budget = branch_budget
        self.title("Manual Investigation — Phase 2/3 Step"); self.geometry("660x500"); self.configure(bg="#0f0f1a"); self.resizable(True, True); self.lift(); self.focus_force()
        tk.Label(self, text="  Manual Investigation — Sequential Rule (Phase 2)", bg="#16213e", fg="#4a9eff", font=("Segoe UI",12,"bold"), height=2).pack(fill="x")
        tk.Label(self, text=f"Remaining PQV: {int(remaining_pqv)}    Branch budget: {int(branch_budget)}", bg="#0f0f1a", fg="#cccccc", font=("Segoe UI",9)).pack(pady=2)
        inv_f = tk.LabelFrame(self, text="  Select Invoice to Continue  ", font=("Segoe UI",9,"bold"), bg="#0f0f1a", fg="#e0e0e0", padx=10, pady=6); inv_f.pack(fill="x", padx=16, pady=4)
        opts = [f"Qty={fmt_qty(m['mtc_qty'])}  |  Inv={m['mtc_inv']}  |  PO={m['mtc_po']}" for m in matches]
        self._branch_cb = ttk.Combobox(inv_f, values=opts, state="readonly", width=70, font=("Segoe UI",9)); if opts: self._branch_cb.current(0); self._branch_cb.pack()
        
        ibc_f = tk.LabelFrame(self, text="  IBC = PBC Validation  ", font=("Segoe UI",9,"bold"), bg="#0f0f1a", fg="#e0e0e0", padx=10, pady=6); ibc_f.pack(fill="x", padx=16, pady=4)
        self._validity = tk.StringVar(value="valid"); rf = tk.Frame(ibc_f, bg="#0f0f1a"); rf.pack(fill="x")
        tk.Radiobutton(rf, text="✔  IBC = PBC  VALID — Continue", variable=self._validity, value="valid", bg="#0f0f1a", fg="#90ee90", selectcolor="#1e3a28", font=("Segoe UI",10,"bold"), command=self._toggle).pack(side="left", padx=6)
        tk.Radiobutton(rf, text="✗  IBC ≠ PBC  INVALID — Exclude", variable=self._validity, value="invalid", bg="#0f0f1a", fg="#ff8888", selectcolor="#3a1e1e", font=("Segoe UI",10,"bold"), command=self._toggle).pack(side="left", padx=14)
        self._invalid_frame = tk.Frame(ibc_f, bg="#0f0f1a"); tk.Label(self._invalid_frame, text="Units matched to invalid invoice:", bg="#0f0f1a", fg="#ff8888", font=("Segoe UI",9)).pack(side="left", padx=4); self._inv_qty_var = tk.StringVar(); tk.Entry(self._invalid_frame, textvariable=self._inv_qty_var, width=10, font=("Segoe UI",10), bg="#1e1e3a", fg="#ff8888", insertbackground="white", relief="flat").pack(side="left", padx=4)

        self._dices_frame = tk.LabelFrame(self, text="  DICES Details  ", font=("Segoe UI",9,"bold"), bg="#0f0f1a", fg="#e0e0e0", padx=10, pady=6); self._dices_frame.pack(fill="x", padx=16, pady=4)
        r1 = tk.Frame(self._dices_frame, bg="#0f0f1a"); r1.pack(fill="x", pady=2); tk.Label(r1, text="SID from DICES:", bg="#0f0f1a", fg="#e0e0e0", font=("Segoe UI",9), width=20, anchor="w").pack(side="left"); self._sid_var = tk.StringVar(); tk.Entry(r1, textvariable=self._sid_var, width=28, font=("Segoe UI",9), bg="#1e1e3a", fg="#e0e0e0", insertbackground="white", relief="flat").pack(side="left", padx=4)
        r2 = tk.Frame(self._dices_frame, bg="#0f0f1a"); r2.pack(fill="x", pady=2); tk.Label(r2, text="Barcode from DICES:", bg="#0f0f1a", fg="#e0e0e0", font=("Segoe UI",9), width=20, anchor="w").pack(side="left"); self._bc_var = tk.StringVar(); tk.Entry(r2, textvariable=self._bc_var, width=28, font=("Segoe UI",9), bg="#1e1e3a", fg="#e0e0e0", insertbackground="white", relief="flat").pack(side="left", padx=4)
        self._toggle()

        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=10)
        tk.Button(bf, text="▶  CONTINUE", command=self._ok, bg="#2d6a4f", fg="white", font=("Segoe UI",12,"bold"), padx=16, pady=8, relief="flat", cursor="hand2").pack(side="left", padx=6)
        tk.Button(bf, text="🔄  CROSS PO", command=self._cross_po, bg="#7a5c00", fg="white", font=("Segoe UI",10,"bold"), padx=12, pady=8, relief="flat", cursor="hand2").pack(side="left", padx=6)
        tk.Button(bf, text="⚖  MISMATCH", command=self._mismatch, bg="#2d4a7a", fg="white", font=("Segoe UI",10,"bold"), padx=12, pady=8, relief="flat", cursor="hand2").pack(side="left", padx=6)
        tk.Button(bf, text="⬛  STOP ASIN", command=self._stop, bg="#4a2020", fg="white", font=("Segoe UI",10), padx=12, pady=8, relief="flat", cursor="hand2").pack(side="left", padx=6)
        self.update_idletasks(); px, py = parent.winfo_x() + (parent.winfo_width() - self.winfo_width()) // 2, parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2; self.geometry(f"+{px}+{py}")

    def _toggle(self):
        if self._validity.get() == "valid": self._invalid_frame.pack_forget(); self._dices_frame.pack(fill="x", padx=16, pady=4)
        else: self._dices_frame.pack_forget(); self._invalid_frame.pack(fill="x", pady=3)

    def _ok(self):
        sel = self._branch_cb.current()
        if sel < 0: messagebox.showwarning("Select Invoice", "Please select an invoice.", parent=self); return
        match = self.matches[sel]
        if self._validity.get() == "valid":
            sid = extract_sid(self._sid_var.get().strip())
            if not sid: messagebox.showwarning("SID Required", "Please enter SID from DICES.", parent=self); return
            self.callback({'action':'valid', 'chosen_match':match, 'sid':sid, 'barcode':self._bc_var.get().strip() or "[DICES]"})
        else:
            try: qty = float(self._inv_qty_var.get().strip())
            except: messagebox.showwarning("Qty Required", "Enter units matched to invalid invoice.", parent=self); return
            self.callback({'action':'invalid', 'chosen_match':match, 'invalid_qty':qty})
        self.destroy()

    def _cross_po(self): self.callback({'action':'cross_po', 'chosen_match': self.matches[self._branch_cb.current()] if self.matches else None}); self.destroy()
    def _mismatch(self):
        dlg = tk.Toplevel(self); dlg.title("Mismatch / Overage Details — Phase 4"); dlg.geometry("460x260"); dlg.configure(bg="#0f0f1a"); dlg.lift(); dlg.focus_force()
        fields = [("ASIN received:", "asin"), ("SID:", "sid"), ("PO:", "po"), ("Inv Qty:", "inv_qty"), ("Overage Qty:", "ovg_qty")]; vars_ = {}
        for i, (lbl, key) in enumerate(fields):
            tk.Label(dlg, text=lbl, bg="#0f0f1a", fg="#e0e0e0", font=("Segoe UI",10), width=22, anchor="w").grid(row=i, column=0, padx=12, pady=5)
            v = tk.StringVar(); tk.Entry(dlg, textvariable=v, width=26, font=("Segoe UI",10), bg="#1e1e3a", fg="#e0e0e0", insertbackground="white", relief="flat").grid(row=i, column=1, padx=8, pady=5); vars_[key] = v
        tk.Button(dlg, text="✔  Submit Mismatch (Phase 4)", command=lambda: (self.callback({'action':'mismatch', 'mismatch_data': {k: v.get().strip() for k, v in vars_.items()}}), dlg.destroy(), self.destroy()), bg="#2d6a4f", fg="white", font=("Segoe UI",11,"bold"), padx=14, pady=7, relief="flat", cursor="hand2").grid(row=len(fields), column=0, columnspan=2, pady=12)

    def _stop(self): self.callback({'action':'stop'}); self.destroy()


class PreviewPanel(tk.Toplevel):
    COLS = ['Barcode','Inv no','SID','PO','ASIN','Inv Qty','Rec Qty','Mtc Qty','Mtc Inv','Remarks','Date']
    COL_W_PX = [130,160,130,90,110,60,60,60,160,240,150]

    def __init__(self, parent):
        super().__init__(parent); self.title("Investigation Preview — Manual Mode Workflow"); self.geometry("1280x520"); self.configure(bg="#0f0f1a"); self.resizable(True, True)
        tk.Label(self, text="  Live Investigation Preview — Double-click to edit", bg="#16213e", fg="#4a9eff", font=("Segoe UI",10,"bold"), height=2).pack(fill="x")
        frame = tk.Frame(self, bg="#0f0f1a"); frame.pack(fill="both", expand=True, padx=8, pady=6)
        vsb, hsb = ttk.Scrollbar(frame, orient="vertical"), ttk.Scrollbar(frame, orient="horizontal")
        self.tree = ttk.Treeview(frame, columns=self.COLS, show='headings', yscrollcommand=vsb.set, xscrollcommand=hsb.set, height=22)
        vsb.config(command=self.tree.yview); hsb.config(command=self.tree.xview)
        for col, w in zip(self.COLS, self.COL_W_PX): self.tree.heading(col, text=col); self.tree.column(col, width=w, minwidth=40, anchor='w')
        self.tree.grid(row=0, column=0, sticky="nsew"); vsb.grid(row=0, column=1, sticky="ns"); hsb.grid(row=1, column=0, sticky="ew")
        frame.grid_rowconfigure(0, weight=1); frame.grid_columnconfigure(0, weight=1); self.tree.bind('<Double-1>', self._on_double_click); self._row_data = {}
        bb = tk.Frame(self, bg="#0f0f1a"); bb.pack(fill="x", padx=8, pady=4); tk.Label(bb, text="Double-click to edit", bg="#0f0f1a", fg="#8888aa", font=("Segoe UI",8)).pack(side="left"); tk.Button(bb, text="Clear All", command=self.clear_all, bg="#2d2d5e", fg="white", font=("Segoe UI",9), relief="flat", padx=10, pady=4, cursor="hand2").pack(side="right")
        s = ttk.Style(); s.configure("Treeview", font=("Calibri",10), rowheight=22, background="#1e1e3a", foreground="#e0e0e0", fieldbackground="#1e1e3a"); s.configure("Treeview.Heading", font=("Calibri",10,"bold"), background="#203864", foreground="white")
        for tag, bg, fg in [('header', '#203864', 'white'), ('d0', '#1e1e3a', '#e0e0e0'), ('d1', '#1e3a28', '#e0e0e0'), ('sub', '#1a1a35', '#e0e0e0'), ('root', '#3a1e1e', '#ff8888'), ('sr', '#3a1a1a', '#ff6666'), ('dices', '#3a3010', '#e0e0e0'), ('invalid', '#3a1010', '#ff9999'), ('rebni', '#0d2535', '#88ddff'), ('crosspo', '#2a1a00', '#f0c060'), ('mismatch', '#002040', '#66aaff')]: self.tree.tag_configure(tag, background=bg, foreground=fg)

    def add_header_row(self, label=""): vals = list(self.COLS); if label: vals[4] = f"── {label} ──"; iid = self.tree.insert('', 'end', values=vals, tags=('header',)); self._row_data[iid] = dict(zip(self.COLS, vals))
    def add_row(self, rd):
        vals = [rd.get(k, '') for k in ['barcode','invoice','sid','po','asin','inv_qty','rec_qty','mtc_qty','mtc_inv','remarks','date']]; rem = rd.get('remarks','').lower(); tag = ('sub' if rd.get('type') == 'subrow' else 'root' if 'accounted' in rem or 'short' in rem else 'sr' if rem == 'sr' else 'invalid' if 'invalid' in rem else 'rebni' if 'rebni available' in rem else 'crosspo' if 'cross po' in rem or rd.get('barcode','') == '[CROSS PO]' else 'mismatch' if 'mismatch' in rem else 'dices' if '[dices]' in str(rd.get('barcode','')).lower() else f"d{min(rd.get('depth',0),1)}")
        iid = self.tree.insert('', 'end', values=vals, tags=(tag,)); self._row_data[iid] = dict(zip(self.COLS, vals)); self._row_data[iid]['_rd'] = rd; self.tree.see(iid)
    def get_all_rows(self):
        KEY = {'Barcode':'barcode','Inv no':'invoice','SID':'sid','PO':'po','ASIN':'asin','Inv Qty':'inv_qty','Rec Qty':'rec_qty','Mtc Qty':'mtc_qty','Mtc Inv':'mtc_inv','Remarks':'remarks','Date':'date'}; rows = []
        for iid in self.tree.get_children():
            d = self._row_data.get(iid, {}); if d.get(self.COLS[0]) == self.COLS[0]: continue
            rd = d.get('_rd', {}).copy(); [rd.__setitem__(KEY[col], d.get(col, '')) for col in self.COLS]; rows.append(rd)
        return rows
    def clear_all(self): [self.tree.delete(iid) for iid in self.tree.get_children()]; self._row_data.clear()
    def _on_double_click(self, event):
        reg = self.tree.identify_region(event.x, event.y); iid = self.tree.identify_row(event.y); col = self.tree.identify_column(event.x)
        if reg != 'cell' or not iid or not col: return
        ci = int(col.replace('#', '')) - 1; cn = self.COLS[ci]; bbox = self.tree.bbox(iid, col); if not bbox: return
        x, y, w, h = bbox; ev = tk.StringVar(value=str(self._row_data.get(iid, {}).get(cn, ''))); e = tk.Entry(self.tree, textvariable=ev, font=("Calibri", 10), bg="#2d2d5e", fg="white", insertbackground="white", relief="flat", bd=1); e.place(x=x, y=y, width=w, height=h); e.focus_force(); e.select_range(0, 'end')
        def save(ev_=None): nv = ev.get(); if iid in self._row_data: self._row_data[iid][cn] = nv; v = list(self.tree.item(iid, 'values')); v[ci] = nv; self.tree.item(iid, values=v); e.destroy()
        e.bind('<Return>', save); e.bind('<Tab>', save); e.bind('<Escape>', lambda _: e.destroy()); e.bind('<FocusOut>', save)


# ═══════════════════════════════════════════════════════════
#  DATA LOADERS & INDICES
# ═══════════════════════════════════════════════════════════

def _load_file(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == '.csv':
        try: return pd.read_csv(path, dtype=str, encoding='utf-8')
        except: return pd.read_csv(path, dtype=str, encoding='latin-1')
    return pd.read_excel(path, header=0, dtype=str)

COLUMN_ALIASES = {
    'Barcode': ['barcode','upc','ean','item code'], 'Invoice': ['inv no','invoice_no','invoice number'],
    'SID': ['sid','shipment id','shipment_id'], 'PO': ['po','po no','po number','purchase order'],
    'ASIN': ['asin','po_asin','amazon asin'], 'InvQty': ['inv qty','invoice qty','invoiced quantity'],
    'PQV': ['pqv','pqv qty','missing qty','shortage'],
}

def detect_claim_cols(df):
    ac, mapping, corr = list(df.columns), {}, []
    for f, al in COLUMN_ALIASES.items():
        found = None
        for a in al:
            for c in ac:
                if a == c.lower().strip() or a in c.lower() or c.lower() in a: found = c; break
            if found: break
        if found:
            if found.lower().strip() != al[0]: corr.append((f, found, al[0]))
            mapping[f] = found
    return mapping, corr

def load_rebni(path):
    df = _load_file(path); n = ['vendor_code','po','asin','shipment_id','received_datetime','wh_id','item_cost','quantity_unpacked','quantity_adjusted','qty_postadj','quantity_matched','rebni_available','cnt_mtc','mtc_invs']
    df.columns = n[:len(df.columns)]; return df

def load_invoice_search(path):
    df = _load_file(path); n = ['vendor_code','po_id','asin','inv_no','inv_date','status','qty_inv','qty_mtc','no_ship','shipment_id','shipmentwise_matched_qty','matched_po','matched_asin']
    df.columns = n[:len(df.columns)]; return df

def build_rebni_index(df):
    p, s, fb = {}, {}, {}
    for _, r in df.iterrows():
        sid, po, asin = extract_sid(clean(r.get('shipment_id',''))), clean(r.get('po','')), clean(r.get('asin',''))
        if not sid or not asin: continue
        p.setdefault((sid, po, asin), []).append(r.to_dict())
        s.setdefault((po, asin), []).append(r.to_dict())
        for inv in split_comma(r.get('mtc_invs','')):
            if inv: fb.setdefault((sid, po, inv), []).append(r.to_dict())
    return p, s, fb

def build_invoice_index(df):
    idx, fb = {}, {}
    for _, r in df.iterrows():
        sids, pos, asins, qtys = split_comma(r.get('shipment_id','')), split_comma(r.get('matched_po','')), split_comma(r.get('matched_asin','')), split_comma(r.get('shipmentwise_matched_qty',''))
        for i in range(max(len(sids), len(pos), len(asins), len(qtys))):
            sf, pv, av = extract_sid(sids[i] if i < len(sids) else ""), pos[i] if i < len(pos) else "", asins[i] if i < len(asins) else ""
            qv, inv, mpo, mas = safe_num(qtys[i] if i < len(qtys) else "0"), clean(r.get('inv_no','')), clean(r.get('po_id','')), clean(r.get('asin',''))
            if not sf or not pv or not av: continue
            e = {'mtc_inv':inv, 'mtc_po':mpo, 'mtc_asin':mas, 'inv_qty':safe_num(r.get('qty_inv','0')), 'mtc_qty':qv, 'date':clean(r.get('inv_date',''))}
            idx.setdefault((sf, pv, av), []).append(e)
            if inv: fb.setdefault((sf, pv, inv), []).append(e)
    return idx, fb


# ═══════════════════════════════════════════════════════════
#  INVESTIGATION ENGINE — ROY B WORKFLOW (v5.0.4)
# ═══════════════════════════════════════════════════════════

class InvestigationEngine:
    MAX_DEPTH = 10

    def __init__(self, rp, rs, rfb, ip, ifb, sid_cb=None):
        self.rebni_p, self.rebni_s, self.rebni_fb = rp, rs, rfb; self.inv_p, self.inv_fb = ip, ifb
        self.sid_cb, self.stop_requested = sid_cb, False; self.cache_sid, self.cache_bc, self.loop_cache = {}, {}, {}

    def _resolve_inv_qty(self, inv_no, asin, fallback):
        base = strip_scr(inv_no); if base == clean(inv_no): return fallback
        for (s, p, a), entries in self.inv_p.items():
            if a == clean(asin):
                for e in entries:
                    if clean(e.get('mtc_inv', '')) == base: return safe_num(e.get('inv_qty', 0))
        return fallback

    def detect_cross_po(self, sid, po, asin):
        cands = []; seen = set(); cur_rows = self.rebni_p.get((sid, po, asin), [])
        rec_at_cur = sum(safe_num(r.get('quantity_unpacked', 0)) for r in cur_rows)
        for (s, p, a), rows in self.rebni_p.items():
            if s != sid or a != asin or p == po or p in seen: continue
            for r in rows:
                rec = safe_num(r.get('quantity_unpacked', 0)); if rec <= 0: continue
                seen.add(p); mtcs = self.inv_p.get((sid, p, asin), []); iq = safe_num(mtcs[0].get('inv_qty', 0)) if mtcs else 0.0
                ct = "Case 2 — Not Invoiced" if rec_at_cur == 0 and iq == 0 else "Case 3 — Rec > Inv" if rec > iq and iq > 0 else "Case 1 — Rec=0 at Cur"
                cands.append({'po': p, 'asin': asin, 'sid': sid, 'inv_qty': fmt_qty(iq), 'rec_qty': rec, 'cross_type': ct, 'date': clean(r.get('received_datetime', ''))})
        return cands

    def _build_level_logic(self, barcode, inv_no, sid, po, asin, inv_qty, rem_pqv, depth, is_claiming, indicator_only=False):
        sf = extract_sid(sid); rows = self.rebni_p.get((sf, clean(po), clean(asin)), [])
        rec_qty = sum(safe_num(r.get('quantity_unpacked', 0)) for r in rows)
        rebni_avail = sum(safe_num(r.get('rebni_available', 0)) for r in rows)
        ex_adj = sum(safe_num(r.get('quantity_adjusted', 0)) for r in rows)
        rec_date = clean(rows[0].get('received_datetime', '')) if rows else ''
        remarks = "SR" if depth > 0 and not rows else ""
        shortage = max(0.0, safe_num(inv_qty) - rec_qty)
        
        # Accounting formula: Shortage + Available + Adjusted
        found_at_level = shortage + rebni_avail + ex_adj

        # PHASE 1: Direct Shortage Gateway
        if shortage >= rem_pqv > 0 and not remarks:
            rem = f"Accounted for {int(found_at_level)} units (Shortage={int(shortage)}, REBNI={int(rebni_avail)}, EX={int(ex_adj)}) — Phase 1 Direct Shortage"
            main = self._make_row(barcode, inv_no, sf, po, asin, inv_qty, rec_qty, shortage, "Short Received", rem, rec_date, depth)
            res = [main]
            if indicator_only:
                [res.append({'barcode': '[CROSS PO?]', 'invoice': '—', 'sid': c['sid'], 'po': c['po'], 'asin': c['asin'], 'inv_qty': c.get('inv_qty', ''), 'rec_qty': fmt_qty(c['rec_qty']), 'mtc_qty': '', 'mtc_inv': '', 'remarks': f"Phase 4: Cross PO Candidate Detected | {c['cross_type']}", 'date': c['date'], 'depth': depth, 'type': 'crosspo'}) for c in self.detect_cross_po(sf, clean(po), clean(asin))]
            else: res.extend(self._build_cross_po_rows(sf, clean(po), clean(asin), depth))
            return res, [], rec_qty, found_at_level, 0.0, ex_adj

        if rebni_avail > 0: remarks = f"REBNI Available = {int(rebni_avail)} — Accounting towards shortage"
        if remarks == "SR": return [self._make_row(barcode, inv_no, sf, po, asin, inv_qty, rec_qty, "", "", remarks, rec_date, depth)], [], rec_qty, found_at_level, max(0.0, rem_pqv - found_at_level), ex_adj

        # PHASE 2: Matching Investigation
        raw = self.inv_p.get((sf, clean(po), clean(asin)), []); seen, unique = set(), []
        for m in raw:
            cb = (m['mtc_inv'], m['mtc_po'], m['mtc_asin']); if cb not in seen: seen.add(cb); unique.append(m)
        sm = sorted(unique, key=lambda x: safe_num(x['mtc_qty']), reverse=True); mi, mq = "", ""
        if sm:
            top = sm[0]
            if top['mtc_inv'] == clean(inv_no): mi, mq, sm = "Self Matching", fmt_qty(rec_qty), []
            else: mi, mq = top['mtc_inv'], fmt_qty(top['mtc_qty'])
        elif found_at_level > 0: mi, mq, remarks = "Short Received", fmt_qty(found_at_level), f"Accounted for {int(found_at_level)} units (Shortage={int(shortage)}, REBNI={int(rebni_avail)}, EX={int(ex_adj)})"

        main = self._make_row(barcode, inv_no, sf, po, asin, inv_qty, rec_qty, mq, mi, remarks, rec_date, depth); r = [main]
        start = 1 if (sm and mi not in ("Self Matching", "Short Received")) else 0
        for m in sm[start:]: r.append(self._make_row("", "", "", "", "", "", "", fmt_qty(m['mtc_qty']), m['mtc_inv'], "", "", depth, 'subrow'))
        
        act = []
        for m in sm:
            if m['mtc_inv'] != clean(inv_no): act.append({**m, 'inv_qty': self._resolve_inv_qty(m['mtc_inv'], m['mtc_asin'], m['inv_qty'])})
        
        if not indicator_only: r.extend(self._build_cross_po_rows(sf, clean(po), clean(asin), depth))
        return r, act, rec_qty, found_at_level, max(0.0, rem_pqv - found_at_level), ex_adj

    def _build_cross_po_rows(self, sid, po, asin, depth):
        cands, rows = self.detect_cross_po(sid, po, asin), []
        for c in cands:
            rows.append({'barcode': '[CROSS PO]', 'invoice': '—', 'sid': c['sid'], 'po': c['po'], 'asin': c['asin'], 'inv_qty': c.get('inv_qty', ''), 'rec_qty': fmt_qty(c['rec_qty']), 'mtc_qty': '', 'mtc_inv': '', 'remarks': f"Phase 4: Cross PO — {c['cross_type']} | Investigating chain", 'date': c['date'], 'depth': depth, 'type': 'crosspo'})
            r, _ = self.run_cross_po_investigation(c, c['cross_type'].split("\u2014")[0], safe_num(c['rec_qty']), depth+1); rows.extend(r)
        return rows

    def _make_row(self, b, i, s, p, a, iq, rq, mq, mi, rem, d, depth, rtype='dominant'):
        return {'barcode': b, 'invoice': i, 'sid': extract_sid(s) if s else '', 'po': p, 'asin': a, 'inv_qty': fmt_qty(iq), 'rec_qty': fmt_qty(rq), 'mtc_qty': fmt_qty(mq), 'mtc_inv': mi, 'remarks': rem, 'date': d, 'depth': depth, 'type': rtype}

    def run_auto(self, b, i, s, p, a, iq, pqv, depth=0, visited=None, rem_pqv=None, is_claiming=True, budget=None):
        if self.stop_requested: return [], 0.0
        if visited is None: visited = set()
        if rem_pqv is None: rem_pqv = safe_num(pqv)
        if budget is None: budget = rem_pqv
        sf = extract_sid(s); st = (sf, clean(i), clean(p), clean(a))
        if st in visited or depth >= self.MAX_DEPTH: return [], 0.0
        visited = visited | {st}
        if st in self.loop_cache and depth > 0: return list(self.loop_cache[st]), sum(safe_num(r.get('mtc_qty',0)) for r in self.loop_cache[st] if r.get('mtc_inv') == 'Short Received')

        rows, act, rq, accounted, n_rem, ex = self._build_level_logic(b, i, s, p, a, iq, rem_pqv, depth, is_claiming)
        total_accounted = min(budget, max(0.0, accounted)); cur_rem = budget - total_accounted

        if cur_rem > 0 and act:
            for m in act:
                if self.stop_requested or cur_rem <= 0: break
                ns, nb = self.cache_sid.get(m['mtc_inv']) or self._find_sid(m['mtc_po'], m['mtc_asin'], m['mtc_inv']), safe_num(m['mtc_qty'])
                if not ns and self.sid_cb: ns = self.sid_cb(m['mtc_inv'], m['mtc_po'], m['mtc_asin']); if ns: self.cache_sid[m['mtc_inv']] = ns
                if not ns: rows.append(self._make_row("[DICES]", m['mtc_inv'], "[ENTER SID]", m['mtc_po'], m['mtc_asin'], m['inv_qty'], "", "", "", "Phase 2: Missing SID", "", depth+1)); continue
                cr, fnd = self.run_auto(self.cache_bc.get(m['mtc_inv'], "[DICES]"), m['mtc_inv'], ns, m['mtc_po'], m['mtc_asin'], m['inv_qty'], pqv, depth+1, visited, rem_pqv - total_accounted, False, nb)
                rows.extend(cr); contrib = min(cur_rem, fnd); total_accounted += contrib; cur_rem -= contrib

        if rows and total_accounted > 0 and 'Direct' not in rows[0]['remarks']:
            rows[0]['remarks'] = f"Accounted for {int(total_accounted)} units (Budget: {int(budget)}) — Phase 3 complete" if cur_rem <= 0 else f"Accounted for {int(total_accounted)} units (Remaining: {int(cur_rem)})"
        if depth > 0: self.loop_cache[st] = list(rows)
        return rows, total_accounted

    def build_one_level(self, b, i, s, p, a, iq, rem, d=0, is_claiming=True): return self._build_level_logic(b, i, s, p, a, iq, rem, d, is_claiming, True)

    def run_cross_po_investigation(self, c, ct, bgt, depth=0, visited=None):
        if visited is None: visited = set()
        cs, cp, ca, ciq = c['sid'], c['po'], c['asin'], c.get('inv_qty', bgt); raw = self.inv_p.get((extract_sid(cs), clean(cp), clean(ca)), []); seen, unique = set(), []
        for m in raw:
            cb = (m['mtc_inv'], m['mtc_po'], m['mtc_asin']); if cb not in seen: seen.add(cb); unique.append(m)
        rrows = self.rebni_p.get((extract_sid(cs), clean(cp), clean(ca)), []); rq = sum(safe_num(r.get('quantity_unpacked', 0)) for r in rrows); ra = sum(safe_num(r.get('rebni_available', 0)) for r in rrows); ex = sum(safe_num(r.get('quantity_adjusted', 0)) for r in rrows)
        shortage = max(0.0, safe_num(ciq) - rq); mi = unique[0]['mtc_inv'] if unique else "Short Received"; mq = fmt_qty(unique[0]['mtc_qty']) if unique else ""
        rem = f"Phase 4: Accounted for {int(shortage+ra+ex)} units of Cross PO overage"
        rows = [self._make_row('[CROSS PO]', '—', cs, cp, ca, fmt_qty(ciq), rq, mq, mi, rem, clean(rrows[0].get('received_datetime','')) if rrows else '', depth)]
        total = min(shortage + ra + ex, bgt); cur = bgt - total
        for m in unique:
            if cur <= 0: break
            ni, np, na, nb = m['mtc_inv'], m['mtc_po'], m['mtc_asin'], safe_num(m['mtc_qty']); ns = self.cache_sid.get(ni) or self._find_sid(np, na, ni)
            if ns: cr, fnd = self.run_auto("[DICES]", ni, ns, np, na, m['inv_qty'], cur, depth+1, visited, cur, False, min(nb, cur)); rows.extend(cr); ctri = min(cur, fnd); total += ctri; cur -= ctri
        return rows, total

    def _find_sid(self, po, asin, inv):
        rows = self.rebni_s.get((clean(po), clean(asin)), [])
        for r in rows:
            if inv in split_comma(r.get('mtc_invs','')): return extract_sid(r['shipment_id'])
        return extract_sid(rows[0]['shipment_id']) if rows else None


# ═══════════════════════════════════════════════════════════
#  MAIN APP — UI PRESERVED FROM v5.0.3
# ═══════════════════════════════════════════════════════════

def write_excel(all_blocks, path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Investigation"
    h = ["Barcode","Inv no","SID","PO","ASIN","Inv Qty","Rec Qty","Mtc Qty","Mtc Inv","Remarks","Date"]
    H_FILL, DOM_F, SUB_F, ROOT_F, DICES_F, SR_F, INVLD_F, REBNI_F, CROSS_F = [PatternFill("solid", fgColor=c) for c in ["203864","E2EFDA","EBF3FB","FFE0E0","FFF2CC","FFD7D7","FFD0D0","D0F0FF","FFF0C0"]]
    H_FT, N_FT, ROOT_FT, SR_FT, CROSS_FT = Font(color="FFFFFF", bold=True, size=10), Font(size=10), Font(bold=True, color="9C0006", size=10), Font(bold=True, color="CC0000", size=10), Font(bold=True, color="7a5c00", size=10)
    BDR = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    K = {'Barcode':'barcode','Inv no':'invoice','SID':'sid','PO':'po','ASIN':'asin','Inv Qty':'inv_qty','Rec Qty':'rec_qty','Mtc Qty':'mtc_qty','Mtc Inv':'mtc_inv','Remarks':'remarks','Date':'date'}
    curr = 1
    for block in all_blocks:
        if not block: continue
        for c, hv in enumerate(h, 1): cell = ws.cell(row=curr, column=c, value=hv); cell.fill, cell.font, cell.border = H_FILL, H_FT, BDR
        curr += 1
        for rd in block:
            rem, rtyp, dep = str(rd.get('remarks','')), rd.get('type','dominant'), rd.get('depth',0)
            fill, fnt = None, N_FT
            if 'invalid' in rem.lower(): fill = INVLD_F
            elif 'REBNI' in rem: fill = REBNI_F
            elif 'Cross PO' in rem or rtyp == 'crosspo': fill, fnt = CROSS_F, CROSS_FT
            elif 'Accounted' in rem or 'short' in rem.lower(): fill, fnt = ROOT_F, ROOT_FT
            elif rem == 'SR': fill, fnt = SR_F, SR_FT
            elif rtyp == 'subrow': fill = SUB_F
            elif dep > 0: fill = DOM_F
            for c, hv in enumerate(h, 1):
                cell = ws.cell(row=curr, column=c, value=rd.get(K[hv], "")); cell.border, cell.font = BDR, fnt
                if fill: cell.fill = fill
            curr += 1
        curr += 1
    for i, w in enumerate([18,22,18,12,14,9,9,9,26,42,22], 1): ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(path)

class MFIToolApp:
    def __init__(self):
        self.root = tk.Tk(); self.root.title("MFI Investigation Tool  v5.0.4  |  Roy B Workflow")
        try: self.root.state('zoomed')
        except: self.root.attributes('-zoomed', True)
        self.root.minsize(900, 620); self.root.configure(bg="#0f0f1a")
        self.claims_path, self.rebni_path, self.inv_path, self.ticket_id, self.mode_var = tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar(value="auto")
        self.all_blocks, self.preview = [], None; self._build_ui()

    def _build_ui(self):
        t = tk.Frame(self.root, bg="#16213e", height=62); t.pack(fill="x")
        tk.Label(t, text="  MFI Investigation Tool", fg="#e94560", bg="#16213e", font=("Segoe UI",20,"bold")).pack(side="left", padx=16, pady=12)
        tk.Label(t, text="v5.0.4 | Roy B Workflow", fg="#8888aa", bg="#16213e", font=("Segoe UI",10)).pack(side="right", padx=16)
        body = tk.Frame(self.root, bg="#0d0d1a", padx=24, pady=12); body.pack(fill="both", expand=True)
        inp = tk.LabelFrame(body, text="  Input Files  ", fg="#4a9eff", bg="#0d0d1a", font=("Segoe UI",10,"bold"), padx=12, pady=8); inp.pack(fill="x", pady=6)
        for lbl, var, row in [("Claims Sheet:", self.claims_path, 0), ("REBNI Result:", self.rebni_path, 1), ("Invoice Search:", self.inv_path, 2)]:
            tk.Label(inp, text=lbl, fg="#cccccc", bg="#131320", width=18, anchor="w").grid(row=row, column=0, pady=3)
            tk.Entry(inp, textvariable=var, width=62, bg="#1e1e3a", fg="white").grid(row=row, column=1, padx=6)
            tk.Button(inp, text="Browse", command=lambda v=var: v.set(filedialog.askopenfilename())).grid(row=row, column=2)
        tf = tk.Frame(body, bg="#0f0f1a"); tf.pack(anchor="w", pady=4); tk.Label(tf, text="Ticket ID:", fg="white", bg="#0f0f1a").pack(side="left"); tk.Entry(tf, textvariable=self.ticket_id, width=28, bg="#1e1e3a", fg="white").pack(side="left", padx=8)
        m = tk.LabelFrame(body, text="Investigation Mode", fg="white", bg="#0f0f1a", padx=10, pady=5); m.pack(fill="x", pady=8); [tk.Radiobutton(m, text=t, variable=self.mode_var, value=v, fg="white", bg="#0f0f1a", selectcolor="#16213e").pack(anchor="w", padx=10) for t, v in [("AUTO", "auto"), ("MANUAL", "manual")]]
        self.status = tk.Label(body, text="Ready", fg="#4a9eff", bg="#0f0f1a"); self.status.pack(pady=(10,0))
        self.pb = ttk.Progressbar(body, mode='determinate'); self.pb.pack(fill="x", pady=4)
        bf = tk.Frame(body, bg="#0f0f1a"); bf.pack(pady=10)
        self.run_btn = tk.Button(bf, text="▶  RUN INVESTIGATION", bg="#e94560", fg="white", font=("Segoe UI",15,"bold"), padx=36, pady=14, command=self.start_run); self.run_btn.pack(side="left", padx=10)
        self.save_btn = tk.Button(bf, text="💾  SAVE OUTPUT", bg="#2d6a4f", fg="white", font=("Segoe UI",13,"bold"), padx=28, pady=14, state="disabled", command=self.save_output); self.save_btn.pack(side="left", padx=10)

    def start_run(self):
        if not all([self.claims_path.get(), self.rebni_path.get(), self.inv_path.get()]): messagebox.showerror("Error", "Select all files."); return
        self.run_btn.config(state="disabled"); self.all_blocks = []
        if self.mode_var.get() == "manual":
            if not self.preview or not self.preview.winfo_exists(): self.preview = PreviewPanel(self.root)
            else: self.preview.clear_all()
        threading.Thread(target=self._process, daemon=True).start()

    def _process(self):
        try:
            df_c = load_claims(self.claims_path.get()); mc, corrections = detect_claim_cols(df_c)
            if corrections or [f for f in COLUMN_ALIASES if f not in mc]:
                conf, done = [None], threading.Event(); self.root.after(0, lambda: HeaderCorrectionDialog(self.root, corrections, mc, list(df_c.columns), lambda res: (conf.__setitem__(0, res['mapping']) if res['action']=='proceed' else None, done.set()))); done.wait()
                if conf[0] is None: self.root.after(0, lambda: self.run_btn.config(state="normal")); return
                mc = conf[0]
            rp, rs, rfb = build_rebni_index(load_rebni(self.rebni_path.get())); ip, ifb = build_invoice_index(load_invoice_search(self.inv_path.get()))
            self.engine = InvestigationEngine(rp, rs, rfb, ip, ifb, self._req_sid); tot = len(df_c)
            if self.mode_var.get() == "auto":
                for i, (_, r) in enumerate(df_c.iterrows()):
                    rows, _ = self.engine.run_auto(clean(r.get(mc.get('Barcode',''),'')), clean(r.get(mc.get('Invoice',''),'')), extract_sid(clean(r.get(mc.get('SID',''),''))), clean(r.get(mc.get('PO',''),'')), clean(r.get(mc.get('ASIN',''),'')), safe_num(r.get(mc.get('InvQty',''),0)), safe_num(r.get(mc.get('PQV',''),0))); self.all_blocks.append(rows)
                self._finish()
            else: self.manual_q, self.map_cols = df_c.to_dict('records'), mc; self._next_man()
        except Exception as e: messagebox.showerror("Error", str(e)); self._finish()

    def _req_sid(self, inv, po, asin):
        if inv in self.engine.cache_sid: return self.engine.cache_sid[inv]
        res, done = [None], threading.Event(); self.root.after(0, lambda: SIDRequestDialog(self.root, inv, po, asin, lambda s: (self.engine.cache_sid.__setitem__(inv, s) if s else None, res.__setitem__(0, s), done.set()))); done.wait(); return res[0]

    def _next_man(self):
        if not self.manual_q: self._finish(); return
        r, mc = self.manual_q.pop(0), self.map_cols; self.curr_m = {'b': clean(r.get(mc.get('Barcode',''), '')), 'i': clean(r.get(mc.get('Invoice',''), '')), 's': extract_sid(clean(r.get(mc.get('SID',''), ''))), 'p': clean(r.get(mc.get('PO',''), '')), 'a': clean(r.get(mc.get('ASIN',''), '')), 'iq': safe_num(r.get(mc.get('InvQty',''), 0)), 'pqv': safe_num(r.get(mc.get('PQV',''), 0)), 'rem': safe_num(r.get(mc.get('PQV',''), 0)), 'budget': safe_num(r.get(mc.get('PQV',''), 0)), 'depth': 0, 'block': [], 'processed': set(), 'rendered': False, 'siblings_stack': []}; self.preview.add_header_row(self.curr_m['a']); threading.Thread(target=self._man_step, daemon=True).start()

    def _man_step(self):
        m = self.curr_m; rows, matches, rq, n_rem = self.engine.build_one_level(m['b'], m['i'], m['s'], m['p'], m['a'], m['iq'], m['rem'], m['depth'], m['depth']==0)
        if not m['rendered']: m['block'].extend(rows); [self.preview.add_row(r) for r in rows]; m['rendered'] = True
        m['rem'] = n_rem; matches = [x for x in matches if x['mtc_inv'] not in m['processed']]; rem_str = rows[0].get('remarks', '')
        if not matches or any(x in rem_str for x in ["Direct Shortage", "REBNI", "SR"]):
            if m['siblings_stack']:
                ctx = m['siblings_stack'].pop(); m.update(ctx); self.root.after(0, lambda sibs=ctx['siblings']: self._show_dlg(sibs)); return
            self.all_blocks.append(m['block']); self._next_man(); return
        self.root.after(0, lambda: self._show_dlg(matches))

    def _show_dlg(self, matches):
        m, f = self.curr_m, matches[0]
        if f['mtc_inv'] in self.engine.cache_sid: self._handle_res({'action':'valid','chosen_match':f,'sid':self.engine.cache_sid[f['mtc_inv']],'barcode':self.engine.cache_bc.get(f['mtc_inv'],"[DICES]")}, matches); return
        ManualLevelDialog(self.root, matches, m['rem'], m['budget'], lambda res: self._handle_res(res, matches))

    def _handle_res(self, res, matches):
        if not res or res['action'] == 'stop': self.all_blocks.append(self.curr_m['block']); self._next_man(); return
        match = res.get('chosen_match'); if match: self.curr_m['processed'].add(match['mtc_inv'])
        if res['action'] == 'invalid':
            excl = res['invalid_qty']; r = self.curr_m; row = {'barcode':'[INVALID]','invoice':match['mtc_inv'],'sid':'—','po':match['mtc_po'],'asin':match['mtc_asin'],'inv_qty':fmt_qty(match['inv_qty']),'remarks':f"Excluded {int(excl)} units — Phase 2 validation",'depth':r['depth'],'type':'subrow'}; r['block'].append(row); self.preview.add_row(row); r['rem'] = max(0, r['rem'] - excl); (self.root.after(0, lambda: self._show_dlg([x for x in matches if x['mtc_inv'] != match['mtc_inv']])) if r['rem'] > 0 else (self.all_blocks.append(r['block']), self._next_man()))
        elif res['action'] == 'cross_po':
            cands = self.engine.detect_cross_po(self.curr_m['s'], self.curr_m['p'], self.curr_m['a']); CrossPODialog(self.root, cands, self.curr_m['i'], self.curr_m['s'], lambda r: self._handle_cross_po(r))
        elif res['action'] == 'mismatch':
            data = res['mismatch_data']; row = {'barcode':'[PHASE 4]','invoice':'','sid':data.get('sid',''),'po':data.get('po',''),'asin':data.get('asin',''),'remarks':f"Phase 4 Target: Mismatch {data.get('ovg_qty','')} units",'depth':self.curr_m['depth'],'type':'subrow'}; self.curr_m['block'].append(row); self.preview.add_row(row); self.root.after(0, lambda: self._show_dlg([x for x in matches if x != res.get('chosen_match')]))
        else:
            self.engine.cache_sid[match['mtc_inv']], self.engine.cache_bc[match['mtc_inv']] = res['sid'], res['barcode']
            rem_sibs = [x for x in matches if x['mtc_inv'] != match['mtc_inv'] and x['mtc_inv'] not in self.curr_m['processed']]
            if rem_sibs: self.curr_m['siblings_stack'].append({'siblings': rem_sibs, 'depth': self.curr_m['depth'], 'rem': self.curr_m['rem'], 'budget': self.curr_m['budget'], 's': self.curr_m['s'], 'p': self.curr_m['p'], 'a': self.curr_m['a'], 'processed': set(self.curr_m['processed'])})
            bb = safe_num(match['mtc_qty']); self.curr_m.update({'b': res['barcode'], 'i': match['mtc_inv'], 's': res['sid'], 'p': match['mtc_po'], 'a': match['mtc_asin'], 'iq': match['inv_qty'], 'rem': bb, 'budget': bb, 'depth': self.curr_m['depth'] + 1, 'rendered': False}); threading.Thread(target=self._man_step, daemon=True).start()

    def _handle_cross_po(self, res):
        if res['action'] == 'skip': threading.Thread(target=self._man_step, daemon=True).start(); return
        r, budget = res['candidate'], safe_num(res['candidate']['rec_qty']); rrows, fnd = self.engine.run_cross_po_investigation(r, res.get('case','Case 1'), budget, self.curr_m['depth']+1); self.curr_m['block'].extend(rrows); [self.preview.add_row(rr) for rr in rrows]; threading.Thread(target=self._man_step, daemon=True).start()

    def _finish(self): self.status.config(text="Complete. Click SAVE."); self.root.after(0, lambda: (self.save_btn.config(state="normal"), self.run_btn.config(state="normal"), messagebox.showinfo("Done", "Investigation Complete!")))
    def save_output(self):
        t, ts = self.ticket_id.get().strip().replace(' ','_'), datetime.now().strftime('%Y%m%d_%H%M%S'); out = f"RoyB_Investigation_{ts}.xlsx"; p = os.path.join(os.path.dirname(self.claims_path.get()) or os.getcwd(), out)
        try: write_excel(self.all_blocks, p); messagebox.showinfo("Saved", f"Saved to:\n{p}")
        except Exception as e: messagebox.showerror("Save Error", str(e))

if __name__ == '__main__': MFIToolApp().run()
