"""
MFI Investigation Tool  v4.9.3
==============================
ROW IB  |  Amazon
Developed by Mukesh

CHANGES IN v4.9.3:
  ✔ Removed transient property from all dialogs to restore Minimize/Maximize buttons.
  ✔ Removed grab_set() (modality) to allow parallel interaction with PreviewPanel and main window.
  ✔ Added lift() and focus_force() for reliable window surfacing.
  ✔ Synced version labels across all UI components.
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, re, threading
from datetime import datetime


# ═══════════════════════════════════════════════════════════
#  UTILITIES
# ═══════════════════════════════════════════════════════════

def extract_sid(val):
    s = str(val).strip()
    parts = re.findall(r'\d{10,}', s)
    return max(parts, key=len) if parts else s

def safe_num(val):
    try:
        if pd.isna(val): return 0.0
    except: pass
    try:
        return float(str(val).replace(',', '').strip())
    except: return 0.0

def clean(val):
    try:
        if pd.isna(val): return ""
    except: pass
    return str(val).strip()

def split_comma(val):
    if not val: return []
    try:
        if pd.isna(val): return []
    except: pass
    return [s.strip() for s in str(val).split(',') if s.strip()]

def fmt_qty(val):
    n = safe_num(val)
    if n == 0: return ""
    return str(int(n)) if n == int(n) else str(n)


# ═══════════════════════════════════════════════════════════
#  DIALOGS
# ═══════════════════════════════════════════════════════════


class HeaderCorrectionDialog(tk.Toplevel):
    """
    Shown when claims file has non-standard column headers.
    Lists what was detected vs canonical name.
    User can confirm auto-correct or cancel.
    """
    def __init__(self, parent, corrections, mapping, df_columns, callback):
        super().__init__(parent)
        self.callback    = callback
        self.corrections = corrections
        self.mapping     = mapping
        self.df_columns  = df_columns

        self.title("Column Header Mismatch Detected — v4.9.3")
        self.geometry("700x480")
        self.configure(bg="#0f0f1a")
        self.resizable(True, True)
        # self.transient(parent)  # Removed to ensure minimize/maximize buttons
        self.lift(); self.focus_force()

        # Header
        tk.Label(self,
                 text="⚠  Non-standard column headers detected in Claims file",
                 bg="#16213e", fg="#f0a500",
                 font=("Segoe UI",12,"bold"), height=2).pack(fill="x")
        tk.Label(self,
                 text="The tool has automatically matched the columns below. "
                      "Please confirm or correct before proceeding.",
                 bg="#0f0f1a", fg="#cccccc",
                 font=("Segoe UI",9)).pack(pady=4)

        # Table showing all field mappings
        outer = tk.Frame(self, bg="#0f0f1a"); outer.pack(fill="both", expand=True, padx=16, pady=6)

        # Header row
        hdrs = ["Field", "Expected", "Found in file", "Status"]
        w = [14, 20, 28, 12]
        for ci, (h, ww) in enumerate(zip(hdrs, w)):
            tk.Label(outer, text=h, bg="#203864", fg="white",
                     font=("Calibri",10,"bold"),
                     width=ww, anchor="w", padx=4).grid(row=0, column=ci, padx=1, pady=1, sticky="w")

        # All fields — show green (exact) or yellow (corrected) or red (missing)
        all_fields = list(COLUMN_ALIASES.keys())
        corrected_fields = {c[0] for c in corrections}
        missing_fields   = [f for f in all_fields if f not in mapping]

        self._override_vars = {}

        for ri, field in enumerate(all_fields, 1):
            canonical = COLUMN_ALIASES[field][0]
            found_col = mapping.get(field, "")
            is_corrected = field in corrected_fields
            is_missing   = field not in mapping

            if is_missing:
                status_txt, status_fg, row_bg = "MISSING", "#ff4444", "#2a0000"
            elif is_corrected:
                status_txt, status_fg, row_bg = "Auto-fixed", "#f0a500", "#1a1500"
            else:
                status_txt, status_fg, row_bg = "✔ OK", "#44ff88", "#001a00"

            tk.Label(outer, text=field, bg=row_bg, fg="#e0e0e0",
                     font=("Calibri",10,"bold"), width=14, anchor="w", padx=4
                     ).grid(row=ri, column=0, padx=1, pady=1, sticky="w")
            tk.Label(outer, text=canonical, bg=row_bg, fg="#aaaacc",
                     font=("Calibri",10), width=20, anchor="w", padx=4
                     ).grid(row=ri, column=1, padx=1, pady=1, sticky="w")

            # Dropdown for found column — allows user to override
            v = tk.StringVar(value=found_col or "— not found —")
            self._override_vars[field] = v
            opts = ["— not found —"] + list(df_columns)
            cb = ttk.Combobox(outer, textvariable=v, values=opts,
                              state="readonly", width=26, font=("Calibri",9))
            cb.grid(row=ri, column=2, padx=1, pady=1, sticky="w")

            tk.Label(outer, text=status_txt, bg=row_bg, fg=status_fg,
                     font=("Calibri",10,"bold"), width=12, anchor="w", padx=4
                     ).grid(row=ri, column=3, padx=1, pady=1, sticky="w")

        # Bottom note
        tk.Label(self,
                 text="You can change any column assignment using the dropdowns above.",
                 bg="#0f0f1a", fg="#888899", font=("Segoe UI",8)).pack(pady=2)

        # Buttons
        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=10)
        tk.Button(bf, text="✔  Auto-correct & Proceed",
                  command=self._proceed,
                  bg="#2d6a4f", fg="white",
                  font=("Segoe UI",12,"bold"),
                  padx=20, pady=8, relief="flat",
                  cursor="hand2").pack(side="left", padx=10)
        tk.Button(bf, text="✖  Cancel",
                  command=self._cancel,
                  bg="#4a2020", fg="white",
                  font=("Segoe UI",11),
                  padx=16, pady=8, relief="flat",
                  cursor="hand2").pack(side="left", padx=10)

        self.protocol("WM_DELETE_WINDOW", self._cancel)
        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _proceed(self):
        # Build final mapping from user selections
        final = {}
        for field, var in self._override_vars.items():
            v = var.get()
            if v and v != "— not found —":
                final[field] = v
        self.callback({'action': 'proceed', 'mapping': final})
        self.destroy()

    def _cancel(self):
        self.callback({'action': 'cancel'})
        self.destroy()

class SIDRequestDialog(tk.Toplevel):
    """Auto mode: SID not found in REBNI → ask user."""
    def __init__(self, parent, invoice, po, asin, callback):
        super().__init__(parent)
        self.callback = callback
        self.title("SID Required — DICES Validation")
        self.geometry("540x210")
        self.resizable(True, True)
        self.configure(bg="#16213e")
        # self.transient(parent)  # Removed to ensure minimize/maximize buttons
        self.lift(); self.focus_force()

        tk.Label(self, text="⚠  SID Not Found in REBNI",
                 bg="#16213e", fg="#e94560",
                 font=("Segoe UI",13,"bold")).pack(pady=(14,4))
        tk.Label(self, text=f"Invoice: {invoice}   PO: {po}   ASIN: {asin}",
                 bg="#16213e", fg="#e0e0e0", font=("Segoe UI",9)).pack(pady=2)
        tk.Label(self, text="Validate this invoice in DICES and enter the SID below:",
                 bg="#16213e", fg="#aaaacc", font=("Segoe UI",9)).pack(pady=6)

        ef = tk.Frame(self, bg="#16213e"); ef.pack()
        tk.Label(ef, text="SID from DICES:", bg="#16213e", fg="#e0e0e0",
                 font=("Segoe UI",10)).pack(side="left", padx=8)
        self._sid = tk.StringVar()
        self._entry = tk.Entry(ef, textvariable=self._sid, width=30,
                               font=("Segoe UI",10), bg="#1e1e3a", fg="#e0e0e0",
                               insertbackground="white", relief="flat")
        self._entry.pack(side="left", padx=4)
        self._entry.focus_set()

        bf = tk.Frame(self, bg="#16213e"); bf.pack(pady=12)
        tk.Button(bf, text="✔  Continue", command=self._ok,
                  bg="#2d6a4f", fg="white", font=("Segoe UI",11,"bold"),
                  padx=16, pady=7, relief="flat", cursor="hand2").pack(side="left", padx=8)
        tk.Button(bf, text="✖  Skip", command=self._skip,
                  bg="#6b2737", fg="white", font=("Segoe UI",10),
                  padx=16, pady=7, relief="flat", cursor="hand2").pack(side="left", padx=8)

        self.bind('<Return>', lambda e: self._ok())
        self.protocol("WM_DELETE_WINDOW", self._skip)
        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _ok(self):
        sid = extract_sid(self._sid.get().strip())
        if sid:
            self.callback(sid); self.destroy()
        else: self._entry.config(bg="#3a1e1e")

    def _skip(self):
        self.callback(None); self.destroy()


class CrossPODialog(tk.Toplevel):
    """
    Cross PO Confirmation Dialog — v4.9.3
    
    - Resizable and minimizable
    - Non-modal (allows parallel check of PreviewPanel)
    """
    CASE_DESCRIPTIONS = {
        "Case 1": (
            "Case 1 — No PO, but ASIN received",
            "Rec=0 at claiming PO. Same ASIN received in different PO within same SID.\n"
            "Those units are overage under a different PO."
        ),
        "Case 2": (
            "Case 2 — PO exists but ASIN not ordered there",
            "This PO exists in the claiming SID, but the ASIN was never invoiced there.\n"
            "Inv Qty = 0, but units were received. This is a Cross PO overage."
        ),
        "Case 3": (
            "Case 3 — PO and ASIN exist but Rec > Inv",
            "Both PO and ASIN are present. Invoiced qty = X, but received more than X.\n"
            "Excess units are Cross PO overage."
        ),
    }

    def __init__(self, parent, candidates, current_inv, sid, callback):
        super().__init__(parent)
        self.callback   = callback
        self.candidates = candidates

        self.title("Cross PO Overage — Confirm & Investigate")
        self.geometry("740x540")
        self.resizable(True, True)
        self.configure(bg="#0f0f1a")
        # self.transient(parent)  # Removed to ensure minimize/maximize buttons
        self.lift(); self.focus_force()

        # Header
        tk.Label(self, text="🔄  Cross PO Overage Detected",
                 bg="#16213e", fg="#f0a500",
                 font=("Segoe UI",13,"bold"), height=2).pack(fill="x")
        tk.Label(self,
                 text=f"SID: {sid}   |   Investigation Invoice: {current_inv}",
                 bg="#0f0f1a", fg="#cccccc",
                 font=("Segoe UI",9)).pack(pady=2)
        tk.Label(self,
                 text="On confirming, the tool will investigate the Cross PO chain "
                      "to find equivalent shortage.",
                 bg="#0f0f1a", fg="#4a9eff",
                 font=("Segoe UI",9)).pack(pady=2)

        # Candidate table
        tf = tk.LabelFrame(self, text="  Detected Cross PO Candidates  ",
                           bg="#0f0f1a", fg="#e0e0e0",
                           font=("Segoe UI",9,"bold"), padx=10, pady=6)
        tf.pack(fill="x", padx=16, pady=6)
        for ci, h in enumerate(["Cross PO","ASIN","Inv Qty","Rec Qty","Overage","Type"]):
            tk.Label(tf, text=h, bg="#203864", fg="white",
                     font=("Calibri",10,"bold"), width=14, anchor="w", padx=3
                     ).grid(row=0, column=ci, padx=1, pady=1)
        for ri, c in enumerate(candidates, 1):
            inv_n = safe_num(c.get('inv_qty', 0))
            rec_n = safe_num(c['rec_qty'])
            ovg   = max(0.0, rec_n - inv_n)
            for ci, v in enumerate([c['po'], c['asin'],
                                     fmt_qty(inv_n), fmt_qty(rec_n),
                                     fmt_qty(ovg) or "—",
                                     c['cross_type'].split("—")[0].strip()]):
                tk.Label(tf, text=str(v), bg="#1e1e3a", fg="#e0e0e0",
                         font=("Calibri",10), width=14, anchor="w", padx=3
                         ).grid(row=ri, column=ci, padx=1, pady=1)

        # Candidate selection
        sf = tk.Frame(self, bg="#0f0f1a"); sf.pack(fill="x", padx=16, pady=4)
        tk.Label(sf, text="Select Cross PO to investigate:",
                 bg="#0f0f1a", fg="#e0e0e0",
                 font=("Segoe UI",10), width=30, anchor="w").pack(side="left")
        opts = [f"PO={c['po']}  Rec={fmt_qty(c['rec_qty'])}  {c['cross_type'].split(chr(8212))[0].strip()}"
                for c in candidates] + ["None — Skip"]
        self._sel_var = tk.StringVar()
        self._sel_cb  = ttk.Combobox(sf, textvariable=self._sel_var,
                                      values=opts, state="readonly", width=50,
                                      font=("Segoe UI",9))
        self._sel_cb.current(0)
        self._sel_cb.pack(side="left", padx=6)
        self._sel_cb.bind("<<ComboboxSelected>>", self._on_candidate_change)

        # Case selection
        cf = tk.LabelFrame(self, text="  Confirm Cross PO Case  ",
                           bg="#0f0f1a", fg="#e0e0e0",
                           font=("Segoe UI",9,"bold"), padx=12, pady=8)
        cf.pack(fill="x", padx=16, pady=4)

        self._case_var = tk.StringVar(value="Case 1")
        self._case_desc_lbl = tk.Label(cf, text="",
                                        bg="#0f0f1a", fg="#aaaacc",
                                        font=("Segoe UI",9), justify="left",
                                        wraplength=640, anchor="w")

        for case_key, (case_label, case_desc) in self.CASE_DESCRIPTIONS.items():
            tk.Radiobutton(cf, text=case_label,
                           variable=self._case_var, value=case_key,
                           bg="#0f0f1a", fg="#f0c060",
                           selectcolor="#1a1500",
                           font=("Segoe UI",10,"bold"),
                           command=self._on_case_change
                           ).pack(anchor="w", pady=2)

        self._case_desc_lbl.pack(anchor="w", pady=4, padx=8)
        self._on_case_change()  # init description

        # Investigation note
        tk.Label(self,
                 text="⚡  On confirming: tool will investigate Cross PO chain "
                      "until full Cross PO rec_qty is explained as shortage.",
                 bg="#0f0f1a", fg="#88ccff",
                 font=("Segoe UI",9,"italic")).pack(pady=4, padx=16, anchor="w")

        # Buttons
        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=10)
        tk.Button(bf, text="✔  Confirm & Investigate",
                  command=self._confirm,
                  bg="#2d6a4f", fg="white",
                  font=("Segoe UI",12,"bold"),
                  padx=20, pady=9, relief="flat",
                  cursor="hand2").pack(side="left", padx=10)
        tk.Button(bf, text="✖  Skip",
                  command=self._skip,
                  bg="#4a2020", fg="white",
                  font=("Segoe UI",11),
                  padx=16, pady=9, relief="flat",
                  cursor="hand2").pack(side="left", padx=10)

        self.protocol("WM_DELETE_WINDOW", self._skip)
        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _on_candidate_change(self, event=None):
        pass

    def _on_case_change(self):
        case_key = self._case_var.get()
        _, desc  = self.CASE_DESCRIPTIONS.get(case_key, ("", ""))
        self._case_desc_lbl.config(text=desc)

    def _confirm(self):
        idx = self._sel_cb.current()
        if idx >= len(self.candidates):
            self.callback({'action': 'skip'}); self.destroy(); return
        self.callback({
            'action'   : 'confirmed',
            'candidate': self.candidates[idx],
            'case'     : self._case_var.get(),
        })
        self.destroy()

    def _skip(self):
        self.callback({'action': 'skip'}); self.destroy()


class ManualLevelDialog(tk.Toplevel):
    """Manual mode: per-level dialog with IBC/PBC + Cross PO + Mismatches."""
    def __init__(self, parent, matches, remaining_pqv, branch_budget, callback):
        super().__init__(parent)
        self.callback = callback
        self.matches = matches
        self.rem_pqv = remaining_pqv
        self.branch_budget = branch_budget

        self.title("Manual Investigation — Next Step")
        self.geometry("660x500")
        self.configure(bg="#0f0f1a")
        self.resizable(True, True)
        # self.transient(parent)  # Removed to ensure minimize/maximize buttons
        self.lift(); self.focus_force()

        # Header
        tk.Label(self, text="  Manual Investigation — Continue",
                 bg="#16213e", fg="#4a9eff",
                 font=("Segoe UI",12,"bold"), height=2).pack(fill="x")

        info = f"Remaining PQV: {int(remaining_pqv)}    Branch budget: {int(branch_budget)}"
        tk.Label(self, text=info, bg="#0f0f1a", fg="#cccccc",
                 font=("Segoe UI",9)).pack(pady=2)

        # Invoice selection
        inv_f = tk.LabelFrame(self, text="  Select Invoice to Continue  ",
                              font=("Segoe UI",9,"bold"),
                              bg="#0f0f1a", fg="#e0e0e0", padx=10, pady=6)
        inv_f.pack(fill="x", padx=16, pady=4)
        opts = [f"Qty={fmt_qty(m['mtc_qty'])}  |  Inv={m['mtc_inv']}  |  PO={m['mtc_po']}  |  ASIN={m['mtc_asin']}"
                for m in matches]
        self._branch_var = tk.StringVar()
        self._branch_cb = ttk.Combobox(inv_f, textvariable=self._branch_var,
                                        values=opts, state="readonly", width=70,
                                        font=("Segoe UI",9))
        if opts: self._branch_cb.current(0)
        self._branch_cb.pack()

        # IBC/PBC section
        ibc_f = tk.LabelFrame(self, text="  IBC = PBC Validation  ",
                               font=("Segoe UI",9,"bold"),
                               bg="#0f0f1a", fg="#e0e0e0", padx=10, pady=6)
        ibc_f.pack(fill="x", padx=16, pady=4)

        self._validity = tk.StringVar(value="valid")
        rf = tk.Frame(ibc_f, bg="#0f0f1a"); rf.pack(fill="x")
        tk.Radiobutton(rf, text="✔  IBC = PBC  VALID — Continue investigation",
                       variable=self._validity, value="valid",
                       bg="#0f0f1a", fg="#90ee90", selectcolor="#1e3a28",
                       font=("Segoe UI",10,"bold"),
                       command=self._toggle).pack(side="left", padx=6)
        tk.Radiobutton(rf, text="✗  IBC ≠ PBC  INVALID — Exclude units",
                       variable=self._validity, value="invalid",
                       bg="#0f0f1a", fg="#ff8888", selectcolor="#3a1e1e",
                       font=("Segoe UI",10,"bold"),
                       command=self._toggle).pack(side="left", padx=14)

        self._invalid_frame = tk.Frame(ibc_f, bg="#0f0f1a"); self._invalid_frame.pack(fill="x", pady=3)
        tk.Label(self._invalid_frame, text="Units matched to invalid invoice:",
                 bg="#0f0f1a", fg="#ff8888", font=("Segoe UI",9)).pack(side="left", padx=4)
        self._inv_qty_var = tk.StringVar()
        tk.Entry(self._invalid_frame, textvariable=self._inv_qty_var, width=10,
                 font=("Segoe UI",10), bg="#1e1e3a", fg="#ff8888",
                 insertbackground="white", relief="flat").pack(side="left", padx=4)

        # DICES details (for valid)
        self._dices_frame = tk.LabelFrame(self, text="  DICES Details  ",
                                           font=("Segoe UI",9,"bold"),
                                           bg="#0f0f1a", fg="#e0e0e0", padx=10, pady=6)
        self._dices_frame.pack(fill="x", padx=16, pady=4)
        r1 = tk.Frame(self._dices_frame, bg="#0f0f1a"); r1.pack(fill="x", pady=2)
        tk.Label(r1, text="SID from DICES:", bg="#0f0f1a", fg="#e0e0e0",
                 font=("Segoe UI",9), width=20, anchor="w").pack(side="left")
        self._sid_var = tk.StringVar()
        tk.Entry(r1, textvariable=self._sid_var, width=28, font=("Segoe UI",9),
                 bg="#1e1e3a", fg="#e0e0e0", insertbackground="white", relief="flat").pack(side="left", padx=4)

        r2 = tk.Frame(self._dices_frame, bg="#0f0f1a"); r2.pack(fill="x", pady=2)
        tk.Label(r2, text="Barcode from DICES:", bg="#0f0f1a", fg="#e0e0e0",
                 font=("Segoe UI",9), width=20, anchor="w").pack(side="left")
        self._bc_var = tk.StringVar()
        tk.Entry(r2, textvariable=self._bc_var, width=28, font=("Segoe UI",9),
                 bg="#1e1e3a", fg="#e0e0e0", insertbackground="white", relief="flat").pack(side="left", padx=4)

        self._toggle()

        # Buttons
        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=10)
        tk.Button(bf, text="▶  CONTINUE",
                  command=self._ok, bg="#2d6a4f", fg="white",
                  font=("Segoe UI",12,"bold"), padx=16, pady=8,
                  relief="flat", cursor="hand2").pack(side="left", padx=6)
        tk.Button(bf, text="🔄  CROSS PO",
                  command=self._cross_po, bg="#7a5c00", fg="white",
                  font=("Segoe UI",10,"bold"), padx=12, pady=8,
                  relief="flat", cursor="hand2").pack(side="left", padx=6)
        tk.Button(bf, text="⚖  MISMATCH / OVERAGE",
                  command=self._mismatch, bg="#2d4a7a", fg="white",
                  font=("Segoe UI",10,"bold"), padx=12, pady=8,
                  relief="flat", cursor="hand2").pack(side="left", padx=6)
        tk.Button(bf, text="⬛  STOP THIS ASIN",
                  command=self._stop, bg="#4a2020", fg="white",
                  font=("Segoe UI",10), padx=12, pady=8,
                  relief="flat", cursor="hand2").pack(side="left", padx=6)

        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _toggle(self):
        if self._validity.get() == "valid":
            self._invalid_frame.pack_forget()
            self._dices_frame.pack(fill="x", padx=16, pady=4)
        else:
            self._dices_frame.pack_forget()
            self._invalid_frame.pack(fill="x", pady=3)

    def _ok(self):
        sel = self._branch_cb.current()
        if sel < 0 or sel >= len(self.matches):
            messagebox.showwarning("Select Invoice", "Please select an invoice.", parent=self); return
        match = self.matches[sel]
        if self._validity.get() == "valid":
            sid = extract_sid(self._sid_var.get().strip())
            if not sid:
                messagebox.showwarning("SID Required", "Please enter SID from DICES.", parent=self); return
            self.callback({'action':'valid', 'chosen_match':match,
                           'sid':sid, 'barcode':self._bc_var.get().strip() or "[DICES]"})
        else:
            qty_str = self._inv_qty_var.get().strip()
            try: qty = float(qty_str)
            except:
                messagebox.showwarning("Qty Required", "Enter units matched to invalid invoice.", parent=self); return
            self.callback({'action':'invalid', 'chosen_match':match, 'invalid_qty':qty})
        self.destroy()

    def _cross_po(self):
        self.callback({'action':'cross_po',
                       'chosen_match': self.matches[self._branch_cb.current()] if self.matches else None})
        self.destroy()

    def _mismatch(self):
        # Mini dialog for mismatch/overage details
        dlg = tk.Toplevel(self)
        dlg.title("Mismatch / Overage Details")
        dlg.geometry("460x260")
        dlg.configure(bg="#0f0f1a")
        # self.grab_set()  # Removed to allow parallel interaction
        dlg.lift(); dlg.focus_force()

        fields = [("ASIN received:", "asin"), ("SID:", "sid"), ("PO:", "po"),
                  ("Inv Qty (invoiced):", "inv_qty"), ("Overage Qty received:", "ovg_qty")]
        vars_ = {}
        for i, (lbl, key) in enumerate(fields):
            tk.Label(dlg, text=lbl, bg="#0f0f1a", fg="#e0e0e0",
                     font=("Segoe UI",10), width=22, anchor="w").grid(row=i, column=0, padx=12, pady=5)
            v = tk.StringVar()
            tk.Entry(dlg, textvariable=v, width=26, font=("Segoe UI",10),
                     bg="#1e1e3a", fg="#e0e0e0", insertbackground="white",
                     relief="flat").grid(row=i, column=1, padx=8, pady=5)
            vars_[key] = v

        def submit():
            data = {k: v.get().strip() for k, v in vars_.items()}
            dlg.destroy()
            self.callback({'action':'mismatch', 'mismatch_data': data})
            self.destroy()

        tk.Button(dlg, text="✔  Submit Mismatch", command=submit,
                  bg="#2d6a4f", fg="white", font=("Segoe UI",11,"bold"),
                  padx=14, pady=7, relief="flat", cursor="hand2"
                  ).grid(row=len(fields), column=0, columnspan=2, pady=12)

    def _stop(self):
        self.callback({'action':'stop'}); self.destroy()


class PreviewPanel(tk.Toplevel):
    COLS = ['Barcode','Inv no','SID','PO','ASIN','Inv Qty','Rec Qty','Mtc Qty','Mtc Inv','Remarks','Date']
    COL_W_PX = [130,160,130,90,110,60,60,60,160,240,150]

    def __init__(self, parent):
        super().__init__(parent)
        self.title("Investigation Preview — Manual Mode (editable)")
        self.geometry("1280x520")
        self.configure(bg="#0f0f1a")
        self.resizable(True, True)

        tk.Label(self, text="  Live Investigation Preview — double-click any cell to edit",
                 bg="#16213e", fg="#4a9eff",
                 font=("Segoe UI",10,"bold"), height=2).pack(fill="x")

        frame = tk.Frame(self, bg="#0f0f1a")
        frame.pack(fill="both", expand=True, padx=8, pady=6)
        vsb = ttk.Scrollbar(frame, orient="vertical")
        hsb = ttk.Scrollbar(frame, orient="horizontal")
        self.tree = ttk.Treeview(frame, columns=self.COLS, show='headings',
                                  yscrollcommand=vsb.set, xscrollcommand=hsb.set, height=22)
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        for col, w in zip(self.COLS, self.COL_W_PX):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, minwidth=40, anchor='w')
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.grid_rowconfigure(0, weight=1); frame.grid_columnconfigure(0, weight=1)
        self.tree.bind('<Double-1>', self._on_double_click)
        self._row_data = {}

        bb = tk.Frame(self, bg="#0f0f1a"); bb.pack(fill="x", padx=8, pady=4)
        tk.Label(bb, text="Double-click any cell to edit",
                 bg="#0f0f1a", fg="#8888aa", font=("Segoe UI",8)).pack(side="left")
        tk.Button(bb, text="Clear All", command=self.clear_all,
                  bg="#2d2d5e", fg="white", font=("Segoe UI",9),
                  relief="flat", padx=10, pady=4, cursor="hand2").pack(side="right")

        s = ttk.Style()
        s.configure("Treeview", font=("Calibri",10), rowheight=22,
                     background="#1e1e3a", foreground="#e0e0e0", fieldbackground="#1e1e3a")
        s.configure("Treeview.Heading", font=("Calibri",10,"bold"),
                     background="#203864", foreground="white")
        for tag, bg, fg in [
            ('header', '#203864', 'white'), ('d0', '#1e1e3a', '#e0e0e0'),
            ('d1', '#1e3a28', '#e0e0e0'), ('sub', '#1a1a35', '#e0e0e0'),
            ('root', '#3a1e1e', '#ff8888'), ('sr', '#3a1a1a', '#ff6666'),
            ('dices', '#3a3010', '#e0e0e0'), ('invalid', '#3a1010', '#ff9999'),
            ('rebni', '#0d2535', '#88ddff'), ('crosspo', '#2a1a00', '#f0c060'),
            ('mismatch', '#002040', '#66aaff'),
        ]:
            self.tree.tag_configure(tag, background=bg, foreground=fg)

    def add_header_row(self, label=""):
        vals = list(self.COLS)
        if label: vals[4] = f"── {label} ──"
        iid = self.tree.insert('', 'end', values=vals, tags=('header',))
        self._row_data[iid] = dict(zip(self.COLS, vals))

    def add_row(self, rd):
        vals = [rd.get('barcode',''), rd.get('invoice',''), rd.get('sid',''),
                rd.get('po',''), rd.get('asin',''), rd.get('inv_qty',''),
                rd.get('rec_qty',''), rd.get('mtc_qty',''), rd.get('mtc_inv',''),
                rd.get('remarks',''), rd.get('date','')]
        remarks = rd.get('remarks','').lower()
        tag = ('sub'     if rd.get('type') == 'subrow' else
               'root'    if 'root cause' in remarks or 'short' in remarks else
               'sr'      if remarks == 'sr' else
               'invalid' if 'invalid invoice' in remarks else
               'rebni'   if 'rebni available' in remarks else
               'crosspo' if 'cross po' in remarks or rd.get('barcode','') == '[CROSS PO]' else
               'mismatch' if 'mismatch' in remarks else
               'dices'   if '[dices]' in str(rd.get('barcode','')).lower() else
               f"d{min(rd.get('depth',0),1)}")
        iid = self.tree.insert('', 'end', values=vals, tags=(tag,))
        self._row_data[iid] = dict(zip(self.COLS, vals))
        self._row_data[iid]['_rd'] = rd
        self.tree.see(iid)

    def get_all_rows(self):
        KEY = {'Barcode':'barcode','Inv no':'invoice','SID':'sid','PO':'po','ASIN':'asin',
               'Inv Qty':'inv_qty','Rec Qty':'rec_qty','Mtc Qty':'mtc_qty',
               'Mtc Inv':'mtc_inv','Remarks':'remarks','Date':'date'}
        rows = []
        for iid in self.tree.get_children():
            d = self._row_data.get(iid, {})
            if d.get(self.COLS[0]) == self.COLS[0]: continue
            rd = d.get('_rd', {}).copy()
            for col in self.COLS: rd[KEY[col]] = d.get(col, '')
            rows.append(rd)
        return rows

    def clear_all(self):
        for iid in self.tree.get_children(): self.tree.delete(iid)
        self._row_data.clear()

    def _on_double_click(self, event):
        region = self.tree.identify_region(event.x, event.y)
        if region != 'cell':
            return
        iid = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if not iid or not col:
            return

        col_idx  = int(col.replace('#', '')) - 1
        col_name = self.COLS[col_idx]

        bbox = self.tree.bbox(iid, col)
        if not bbox:
            return
        x, y, w, h = bbox

        current  = self._row_data.get(iid, {}).get(col_name, '')
        ev       = tk.StringVar(value=str(current))

        e = tk.Entry(self.tree, textvariable=ev,
                     font=("Calibri", 10),
                     bg="#2d2d5e", fg="white",
                     insertbackground="white",
                     relief="flat", bd=1)
        e.place(x=x, y=y, width=w, height=h)
        e.focus_force()
        e.select_range(0, 'end')

        def save(ev_=None):
            nv = ev.get()
            if iid in self._row_data:
                self._row_data[iid][col_name] = nv
            vals        = list(self.tree.item(iid, 'values'))
            vals[col_idx] = nv
            self.tree.item(iid, values=vals)
            try: e.destroy()
            except: pass

        e.bind('<Return>',   save)
        e.bind('<Tab>',      save)
        e.bind('<Escape>',   lambda _: e.destroy())
        e.bind('<FocusOut>', save)


# ═══════════════════════════════════════════════════════════
#  DATA LOADERS
# ═══════════════════════════════════════════════════════════

def _load_file(path):
    """Load Excel or CSV file into DataFrame."""
    ext = os.path.splitext(path)[1].lower()
    if ext == '.csv':
        try: return pd.read_csv(path, dtype=str, encoding='utf-8')
        except: return pd.read_csv(path, dtype=str, encoding='latin-1')
    else:
        return pd.read_excel(path, header=0, dtype=str)

def load_claims(path): return _load_file(path)

# ── Column aliases for smart detection ──────────────────────
COLUMN_ALIASES = {
    'Barcode': [
        'barcode', 'bar code', 'bar_code', 'upc', 'ean',
        'scan code', 'item code', 'carton barcode', 'pkg barcode',
    ],
    'Invoice': [
        'inv no', 'inv_no', 'invoice_no', 'invoice no', 'invoice number',
        'invoice_number', 'invoice', 'inv num', 'inv number',
        'inv#', 'invoice#', 'invc no', 'invc number', 'invc',
        'bill no', 'bill number',
    ],
    'SID': [
        'sid', 'shipment id', 'shipment_id', 'shipment no',
        'shipment number', 'ship id', 'fba shipment id',
        'inbound shipment', 'shipment', 'inbound sid',
    ],
    'PO': [
        'po', 'po_no', 'p.o.', 'po no', 'po number', 'po#',
        'purchase order', 'purchase_order', 'purchase order no',
        'purchase order number', 'po id', 'order no', 'order number',
        'purch order', 'header po', 'line po', 'po num',
        'invoice_lineitem_po',
    ],
    'ASIN': [
        'asin', 'po_asin', 'amazon asin', 'amazon product id',
        'product id', 'asin no', 'item asin', 'amazon id',
    ],
    'InvQty': [
        'inv qty', 'inv_qty', 'invoice qty', 'invoice quantity',
        'invoiced qty', 'invoiced quantity', 'quantity invoiced',
        'quantity_invoiced', 'billed qty', 'billed quantity',
        'total qty', 'total quantity', 'item qty',
        'po_ordered_quantity',
    ],
    'PQV': [
        'pqv', 'pqv qty', 'missing qty', 'missing quantity',
        'missing_qty', 'shortage', 'short qty', 'short quantity',
        'claim qty', 'claimed qty', 'dispute qty',
        'pending qty', 'outstanding qty', 'difference qty',
    ],
}

def detect_claim_cols(df):
    """
    Smart column mapper.
    Returns (mapping_dict, corrections_list).
    corrections_list = list of (field, original_col, matched_col) for any fuzzy matches.
    """
    actual_cols = list(df.columns)
    lower_map   = {c.lower().strip(): c for c in actual_cols}
    mapping      = {}
    corrections  = []   # (field_name, found_header, canonical_name)

    for field, aliases in COLUMN_ALIASES.items():
        found = None
        for alias in aliases:
            if alias in lower_map:
                found = lower_map[alias]; break
        
        if not found:
            # Partial match
            for alias in aliases:
                for col in actual_cols:
                    if alias in col.lower() or col.lower() in alias:
                        found = col; break
                if found: break

        if found:
            canonical = aliases[0]
            if found.lower().strip() != canonical:
                corrections.append((field, found, canonical))
            mapping[field] = found

    return mapping, corrections

def load_rebni(path):
    df = _load_file(path)
    names = ['vendor_code','po','asin','shipment_id','received_datetime',
             'warehouse_id','item_cost','quantity_unpacked','quantity_adjusted',
             'qty_received_postadj','quantity_matched','rebni_available',
             'cnt_invoice_matched','matched_invoice_numbers']
    df.columns = names[:len(df.columns)]
    return df

def load_invoice_search(path):
    df = _load_file(path)
    names = ['vendor_code','purchase_order_id','asin','invoice_number','invoice_date',
             'invoice_item_status','quantity_invoiced','quantity_matched_total',
             'no_of_shipments','shipment_id','shipmentwise_matched_qty','matched_po','matched_asin']
    df.columns = names[:len(df.columns)]
    return df


# ═══════════════════════════════════════════════════════════
#  INDEX BUILDERS
# ═══════════════════════════════════════════════════════════

def build_rebni_index(df):
    p, s, fb = {}, {}, {}
    for _, row in df.iterrows():
        sid  = extract_sid(clean(row.get('shipment_id','')))
        po   = clean(row.get('po',''))
        asin = clean(row.get('asin',''))
        if not sid or not asin: continue
        p.setdefault((sid, po, asin), []).append(row.to_dict())
        s.setdefault((po, asin), []).append(row.to_dict())
        for inv in split_comma(row.get('matched_invoice_numbers','')):
            if inv: fb.setdefault((sid, po, inv), []).append(row.to_dict())
    return p, s, fb

def build_invoice_index(df):
    idx, fb = {}, {}
    for _, row in df.iterrows():
        sids  = split_comma(row.get('shipment_id',''))
        pos   = split_comma(row.get('matched_po',''))
        asins = split_comma(row.get('matched_asin',''))
        qtys  = split_comma(row.get('shipmentwise_matched_qty',''))
        for i in range(max(len(sids), len(pos), len(asins), len(qtys))):
            s_frag = extract_sid(sids[i] if i < len(sids) else "")
            p_val  = pos[i]   if i < len(pos)   else ""
            a_val  = asins[i] if i < len(asins) else ""
            q_val  = safe_num(qtys[i] if i < len(qtys) else "0")
            inv_no   = clean(row.get('invoice_number',''))
            mtc_po   = clean(row.get('purchase_order_id',''))
            mtc_asin = clean(row.get('asin',''))
            if not s_frag or not p_val or not a_val: continue
            entry = {'mtc_inv':inv_no, 'mtc_po':mtc_po, 'mtc_asin':mtc_asin,
                     'inv_qty':safe_num(row.get('quantity_invoiced','0')),
                     'mtc_qty':q_val, 'date':clean(row.get('invoice_date',''))}
            idx.setdefault((s_frag, p_val, a_val), []).append(entry)
            if inv_no: fb.setdefault((s_frag, p_val, inv_no), []).append(entry)
    return idx, fb


# ═══════════════════════════════════════════════════════════
#  INVESTIGATION ENGINE
# ═══════════════════════════════════════════════════════════

class InvestigationEngine:
    MAX_DEPTH = 10

    def __init__(self, rp, rs, rfb, ip, ifb, sid_cb=None):
        self.rebni_p  = rp
        self.rebni_s  = rs
        self.rebni_fb = rfb
        self.inv_p    = ip
        self.inv_fb   = ifb
        self.sid_cb   = sid_cb
        self.stop_requested = False
        self.cache_sid = {}
        self.cache_bc  = {}
        self.loop_cache = {}
        

    def _rebni_lookup(self, sid, po, asin, inv_no=None):
        rows = self.rebni_p.get((sid, po, asin), [])
        if not rows and inv_no:
            rows = self.rebni_fb.get((sid, po, inv_no), [])
        return rows

    def _inv_lookup(self, sid, po, asin):
        # STRICT ASIN MATCHING — FIXED LOGIC v4.9.4.1
        return self.inv_p.get((sid, po, asin), [])

    def _find_sid(self, po, asin, inv_no):
        rows = self.rebni_s.get((clean(po), clean(asin)), [])
        for r in rows:
            if inv_no in split_comma(r.get('matched_invoice_numbers','')):
                return extract_sid(r['shipment_id'])
        return None

    def detect_cross_po(self, sid, current_po, asin):
        candidates = []
        seen_po = set()
        rec_at_current = 0.0
        current_rows = self.rebni_p.get((sid, current_po, asin), [])
        if current_rows:
            rec_at_current = safe_num(current_rows[0].get('quantity_unpacked', 0))

        for (s, p, a), rebni_rows in self.rebni_p.items():
            if s != sid or a != asin or p == current_po or p in seen_po:
                continue
            for r in rebni_rows:
                rec = safe_num(r.get('quantity_unpacked', 0))
                if rec <= 0:
                    continue
                seen_po.add(p)
                inv_matches = self.inv_p.get((sid, p, asin), [])
                inv_qty_cross = safe_num(inv_matches[0].get('inv_qty', 0)) if inv_matches else 0.0
                if rec_at_current == 0 and inv_qty_cross == 0:
                    cross_type = "Case 2 - ASIN not invoiced at this PO, but received"
                elif rec > inv_qty_cross and inv_qty_cross > 0:
                    cross_type = "Case 3 - Rec qty > Inv qty (overage in cross PO)"
                else:
                    cross_type = "Case 1 - Rec=0 at current PO, units received here"
                candidates.append({
                    'po': p,
                    'asin': asin,
                    'sid': sid,
                    'inv_qty': fmt_qty(inv_qty_cross),
                    'rec_qty': rec,
                    'cross_type': cross_type,
                    'date': clean(r.get('received_datetime', '')),
                })
        return candidates

    def _build_level_logic(self, barcode, inv_no, sid, po, asin,
                            inv_qty, rem_pqv, depth, is_claiming):
        sid_frag    = extract_sid(sid)
        rebni_rows  = self._rebni_lookup(sid_frag, clean(po), clean(asin), clean(inv_no))
        rec_qty, remarks, rec_date, r_avail = 0.0, "", "", 0.0

        if rebni_rows:
            r = rebni_rows[0]
            rec_qty, rec_date, r_avail = safe_num(r.get('quantity_unpacked', 0)), clean(r.get('received_datetime', '')), safe_num(r.get('rebni_available', 0))
            if r_avail > 0:
                remarks = f"REBNI Available = {int(r_avail)} units at {'claiming' if is_claiming else 'matching'} level"
        else:
            remarks = "SID not found in REBNI — validate in DICES" if depth == 0 else "SR"

        shortage = max(0.0, safe_num(inv_qty) - rec_qty) if rebni_rows or depth > 0 else 0.0
        
        # STRICT ASIN MATCHING - FIXED v4.9.4.1
        raw = self._inv_lookup(sid_frag, clean(po), clean(asin))
        seen, unique = set(), []
        for m in raw:
            if m['mtc_inv'] not in seen:
                seen.add(m['mtc_inv']); unique.append(m)
        sorted_m = sorted(unique, key=lambda x: safe_num(x['mtc_qty']), reverse=True)

        main_mtc_inv, main_mtc_qty = "", ""
        if sorted_m:
            if sorted_m[0]['mtc_inv'] == clean(inv_no):
                main_mtc_inv, main_mtc_qty = "Self Matching", fmt_qty(sorted_m[0]['mtc_qty'])
                sorted_m = sorted_m[1:]
            else:
                main_mtc_inv, main_mtc_qty = sorted_m[0]['mtc_inv'], fmt_qty(sorted_m[0]['mtc_qty'])
        elif shortage > 0 and not remarks:
            main_mtc_inv, remarks = "Short Received", f"Found {int(shortage)} short as loop started from {int(rem_pqv)} matched qty, no remaining pqv"
            sorted_m = []

        rows = [self._make_row(barcode, inv_no, sid, po, asin, inv_qty, rec_qty, main_mtc_qty, main_mtc_inv, remarks, rec_date, depth)]
        if main_mtc_inv not in ("Short Received", "Self Matching"):
            for m in sorted_m[1:]:
                rows.append(self._make_row("","","","","","","",fmt_qty(m['mtc_qty']), m['mtc_inv'],"","", depth, 'subrow'))

        new_rem = max(0.0, rem_pqv - min(rem_pqv, shortage))
        rows.extend(self._build_cross_po_rows(sid_frag, clean(po), clean(asin), depth))
        return rows, [m for m in sorted_m if m['mtc_inv'] != clean(inv_no)], rec_qty, shortage, new_rem

    def _build_cross_po_rows(self, sid, po, asin, depth):
        rows = []
        for c in self.detect_cross_po(sid, po, asin):
            rows.append({
                'barcode': '[CROSS PO]',
                'invoice': '-',
                'sid': sid,
                'po': c['po'],
                'asin': asin,
                'inv_qty': c.get('inv_qty', ''),
                'rec_qty': fmt_qty(c['rec_qty']),
                'mtc_qty': '',
                'mtc_inv': '',
                'remarks': f"Cross PO - {c['cross_type']}",
                'date': c['date'],
                'depth': depth,
                'type': 'crosspo',
            })
            case_label = c['cross_type'].split("-")[0].strip()
            sub, _ = self.run_cross_po_investigation(c, case_label, c['rec_qty'], depth + 1)
            rows.extend(sub)
        return rows

    def _make_row(self, b, i, s, p, a, iq, rq, mq, mi, rem, d, depth, rtype='dominant'):
        return {'barcode': b, 'invoice': i, 'sid': extract_sid(s) if s else '',
                'po': p, 'asin': a, 'inv_qty': fmt_qty(iq), 'rec_qty': fmt_qty(rq),
                'mtc_qty': fmt_qty(mq), 'mtc_inv': mi, 'remarks': rem, 'date': d,
                'depth': depth, 'type': rtype}

    def run_auto(self, barcode, inv_no, sid, po, asin, inv_qty, pqv,
                 depth=0, visited=None, rem_pqv=None, is_claiming=True,
                 branch_budget=None):
        if self.stop_requested: return [], 0.0
        visited = visited or set()
        rem_pqv = rem_pqv if rem_pqv is not None else safe_num(pqv)
        budget = branch_budget if branch_budget is not None else rem_pqv
        
        state = (extract_sid(sid), clean(inv_no), clean(po), clean(asin))
        if state in visited or depth >= self.MAX_DEPTH: return [], 0.0
        visited.add(state)

        rows, actionable, rec_qty, shortage, n_rem = self._build_level_logic(barcode, inv_no, sid, po, asin, inv_qty, rem_pqv, depth, is_claiming)
        
        found = min(budget, shortage)
        if budget - found <= 0 or not actionable or 'REBNI' in rows[0].get('remarks','') or rows[0].get('remarks','') == 'SR':
            if rows and shortage > 0:
                rows[0]['remarks'] = f"Found {int(shortage)} units short, no remaining pqv"
            return rows, found

        cur_b = budget - found
        for m in actionable:
            if self.stop_requested or cur_b <= 0: break
            nsid = self.cache_sid.get(m['mtc_inv']) or self._find_sid(m['mtc_po'], m['mtc_asin'], m['mtc_inv'])
            if not nsid and self.sid_cb: nsid = self.sid_cb(m['mtc_inv'], m['mtc_po'], m['mtc_asin'])
            if not nsid: continue
            child, f_b = self.run_auto(self.cache_bc.get(m['mtc_inv'], "[DICES]"), m['mtc_inv'], nsid, m['mtc_po'], m['mtc_asin'], m['inv_qty'], pqv, depth+1, visited, rem_pqv - found, False, m['mtc_qty'])
            rows.extend(child); c = min(cur_b, f_b); found += c; cur_b -= c
        return rows, found

    def build_one_level(self, b, i, s, p, a, iq, rem, depth=0, is_clm=True):
        return self._build_level_logic(b, i, s, p, a, iq, rem, depth, is_clm)

    def run_cross_po_investigation(self, cand, c_type, budget, depth=0, visited=None):
        raw = self._inv_lookup(cand['sid'], cand['po'], cand['asin'])
        m_inv = raw[0]['mtc_inv'] if raw else "Short Received"
        rows = [
            self._make_row(
                '[CROSS PO]',
                '-',
                cand['sid'],
                cand['po'],
                cand['asin'],
                cand.get('inv_qty', '0'),
                cand['rec_qty'],
                '',
                m_inv,
                f"Cross PO {c_type} - investigating overage",
                cand['date'],
                depth,
            )
        ]
        return rows, 0.0

# ═══════════════════════════════════════════════════════════
#  EXCEL WRITER
# ═══════════════════════════════════════════════════════════

def write_excel(all_blocks, path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Investigation"
    headers = ["Barcode","Inv no","SID","PO","ASIN","Inv Qty","Rec Qty","Mtc Qty","Mtc Inv","Remarks","Date"]
    H_FILL, DOM_F, SUB_F, ROOT_F, REBNI_F, CROSS_F = [PatternFill("solid", fgColor=c) for c in ["203864","E2EFDA","EBF3FB","FFE0E0","D0F0FF","FFF0C0"]]
    H_FONT, N_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=10), Font(name="Calibri", size=10)
    BDR = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    KM  = {'Barcode':'barcode','Inv no':'invoice','SID':'sid','PO':'po','ASIN':'asin','Inv Qty':'inv_qty','Rec Qty':'rec_qty','Mtc Qty':'mtc_qty','Mtc Inv':'mtc_inv','Remarks':'remarks','Date':'date'}
    curr = 1
    for block in all_blocks:
        if not block: continue
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=curr, column=c, value=h); cell.fill, cell.font, cell.border = H_FILL, H_FONT, BDR
        curr += 1
        for rd in block:
            rem = str(rd.get('remarks','')).lower()
            fill = REBNI_F if 'rebni' in rem else CROSS_F if 'cross' in rem or rd.get('barcode') == '[CROSS PO]' else ROOT_F if 'short' in rem else SUB_F if rd.get('type')=='subrow' else DOM_F if rd.get('depth',0)>0 else None
            for c, h in enumerate(headers, 1):
                val = rd.get(KM[h], ""); cell = ws.cell(row=curr, column=c, value=val if val not in (None, '') else None); cell.border = BDR; cell.font = N_FONT
                if fill: cell.fill = fill
            curr += 1
        curr += 1
    wb.save(path)

# ═══════════════════════════════════════════════════════════
#  MAIN GUI
# ═══════════════════════════════════════════════════════════

class MFIToolApp:
    def __init__(self):
        self.root = tk.Tk(); self.root.title("MFI Investigation Tool  v4.9.4.1  |  ROW IB")
        try: self.root.state('zoomed')
        except: self.root.attributes('-zoomed', True)
        self.root.minsize(1000, 700); self.root.configure(bg="#0f0f1a")
        self.claims_path, self.rebni_path, self.inv_path, self.ticket_id, self.mode_var = tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar(value="auto")
        self.all_blocks, self.preview = [], None; self._build_ui()

    def _build_ui(self):
        t = tk.Frame(self.root, bg="#16213e", height=65); t.pack(fill="x")
        tk.Label(t, text="  MFI Investigation Tool", fg="#e94560", bg="#16213e", font=("Segoe UI",24,"bold")).pack(side="left", padx=20, pady=10)
        tk.Label(t, text="v4.9.4.1  |  ROW IB", fg="#8888aa", bg="#16213e", font=("Segoe UI",12)).pack(side="right", padx=20)
        leg = tk.Frame(self.root, bg="#1a1a2e", height=35); leg.pack(fill="x")
        for txt, fg, bg in [("Claiming","white","#0f0f1a"), ("Dominant","black","#E2EFDA"), ("Sub-rows","black","#EBF3FB"), ("Short/Root","#9C0006","#FFE0E0"), ("REBNI","#333","#D0F0FF"), ("Cross PO","#7a5c00","#FFF0C0")]:
            tk.Label(leg, text=f"  {txt}  ", fg=fg, bg=bg, font=("Segoe UI",9,"bold"), padx=10).pack(side="left", padx=5, pady=5)
        body = tk.Frame(self.root, bg="#0d0d1a", padx=30, pady=20); body.pack(fill="both", expand=True)
        inp = tk.LabelFrame(body, text="  Configuration & Input  ", fg="#4a9eff", bg="#0d0d1a", font=("Segoe UI",12,"bold"), padx=15, pady=15); inp.pack(fill="x", pady=10)
        self._f_row(inp, "Claims Sheet:", self.claims_path, 0); self._f_row(inp, "REBNI Result:", self.rebni_path, 1); self._f_row(inp, "Invoice Search:", self.inv_path, 2)
        tf = tk.Frame(body, bg="#0f0f1a"); tf.pack(anchor="w", pady=5); tk.Label(tf, text="Ticket ID:", fg="white", bg="#0f0f1a", font=("Segoe UI",12)).pack(side="left"); tk.Entry(tf, textvariable=self.ticket_id, width=35, font=("Segoe UI",12), bg="#1e1e3a", fg="white").pack(side="left", padx=10)
        m = tk.LabelFrame(body, text=" Investigation Mode ", fg="white", bg="#0f0f1a", font=("Segoe UI",11), padx=15, pady=10); m.pack(fill="x", pady=15)
        tk.Radiobutton(m, text="AUTO Investigation", variable=self.mode_var, value="auto", fg="white", bg="#0f0f1a", font=("Segoe UI",11)).pack(anchor="w", pady=2)
        tk.Radiobutton(m, text="MANUAL Investigation (Full Features: IBC/PBC, Mismatch, Overages)", variable=self.mode_var, value="manual", fg="white", bg="#0f0f1a", font=("Segoe UI",11)).pack(anchor="w", pady=2)
        self.status = tk.Label(body, text="System Ready", fg="#4a9eff", bg="#0f0f1a", font=("Segoe UI",12)); self.status.pack(pady=10); self.pb = ttk.Progressbar(body, mode='determinate'); self.pb.pack(fill="x", pady=10)
        bf = tk.Frame(body, bg="#0f0f1a"); bf.pack(pady=20)
        self.run_btn = tk.Button(bf, text="▶  START INVESTIGATION", bg="#e94560", fg="white", font=("Segoe UI",18,"bold"), padx=50, pady=15, relief="flat", command=self.start_run); self.run_btn.pack(side="left", padx=15)
        self.stop_btn = tk.Button(bf, text="🛑  STOP", bg="#4a2020", fg="white", font=("Segoe UI",14), padx=20, pady=15, state="disabled", command=self.request_stop); self.stop_btn.pack(side="left", padx=10)
        self.save_btn = tk.Button(bf, text="💾  SAVE REPORT", bg="#2d6a4f", fg="white", font=("Segoe UI",14), padx=30, pady=15, state="disabled", command=self.save_output); self.save_btn.pack(side="left", padx=15)

    def _f_row(self, p, l, v, r):
        tk.Label(p, text=l, fg="#cccccc", bg="#131320", width=20, anchor="w", font=("Segoe UI",11)).grid(row=r, column=0, pady=5); tk.Entry(p, textvariable=v, width=80, font=("Segoe UI",10), bg="#1e1e3a", fg="white").grid(row=r, column=1, padx=10); tk.Button(p, text="Browse", command=lambda: v.set(filedialog.askopenfilename())).grid(row=r, column=2)

    def _set_status(self, msg, pct=None):
        if threading.current_thread() is not threading.main_thread():
            self.root.after(0, lambda m=msg, p=pct: self._set_status(m, p))
            return
        self.status.config(text=msg)
        if pct is not None:
            self.pb['value'] = pct
        self.root.update_idletasks()

    def start_run(self):
        if not all([self.claims_path.get(), self.rebni_path.get(), self.inv_path.get()]):
            messagebox.showerror("Error", "Please select all 3 input files."); return
        self.run_btn.config(state="disabled"); self.save_btn.config(state="disabled"); self.stop_btn.config(state="normal")
        self.all_blocks = []
        if self.mode_var.get() == "manual":
            if not self.preview or not self.preview.winfo_exists(): self.preview = PreviewPanel(self.root)
            else: self.preview.clear_all()
        threading.Thread(target=self._process, daemon=True).start()

    def request_stop(self):
        if hasattr(self, 'engine'): self.engine.stop_requested = True

    def _process(self):
        try:
            df_c = load_claims(self.claims_path.get()); mc, corrections = detect_claim_cols(df_c)
            rp, rs, rfb = build_rebni_index(load_rebni(self.rebni_path.get())); ip, ifb = build_invoice_index(load_invoice_search(self.inv_path.get()))
            self.engine = InvestigationEngine(rp, rs, rfb, ip, ifb, self._req_sid); tot = len(df_c)
            if self.mode_var.get() == "auto":
                for i, (_, r) in enumerate(df_c.iterrows()):
                    if self.engine.stop_requested: break
                    sid = extract_sid(clean(r.get(mc.get('SID',''), '')))
                    rows, _ = self.engine.run_auto(clean(r.get(mc.get('Barcode',''),'')), clean(r.get(mc.get('Invoice',''),'')), sid, clean(r.get(mc.get('PO',''),'')), clean(r.get(mc.get('ASIN',''),'')), safe_num(r.get(mc.get('InvQty',''),0)), safe_num(r.get(mc.get('PQV',''),0)))
                    self.all_blocks.append(rows); self._set_status(f"Processing ASIN {i+1}/{tot}", (i/tot)*100)
                self._finish()
            else: self.manual_q, self.map_cols = df_c.to_dict('records'), mc; self._next_man()
        except Exception as e:
            self._finish()
            self.root.after(0, lambda err=str(e): messagebox.showerror("Engine Error", err))

    def _req_sid(self, inv, po, asin):
        if inv in self.engine.cache_sid: return self.engine.cache_sid[inv]
        res, done = [None], threading.Event()
        self.root.after(0, lambda: SIDRequestDialog(self.root, inv, po, asin, lambda s: (self.engine.cache_sid.__setitem__(inv, s) if s else None, res.__setitem__(0, s), done.set())))
        done.wait(); return res[0]

    def _next_man(self):
        if threading.current_thread() is not threading.main_thread():
            self.root.after(0, self._next_man)
            return
        if not self.manual_q or self.engine.stop_requested:
            self._finish()
            return
        r, mc = self.manual_q.pop(0), self.map_cols
        self.curr_m = {'b': clean(r.get(mc.get('Barcode',''), '')), 'i': clean(r.get(mc.get('Invoice',''), '')), 's': extract_sid(clean(r.get(mc.get('SID',''), ''))), 'p': clean(r.get(mc.get('PO',''), '')), 'a': clean(r.get(mc.get('ASIN',''), '')), 'iq': safe_num(r.get(mc.get('InvQty',''), 0)), 'pqv': safe_num(r.get(mc.get('PQV',''), 0)), 'rem': safe_num(r.get(mc.get('PQV',''), 0)), 'budget': safe_num(r.get(mc.get('PQV',''), 0)), 'depth': 0, 'block': [], 'processed': set(), 'rendered': False}
        self.preview.add_header_row(self.curr_m['a']); threading.Thread(target=self._man_step, daemon=True).start()

    def _man_step(self):
        if self.engine.stop_requested: self._finish(); return
        m = self.curr_m; rows, matches, rq, n_rem = self.engine.build_one_level(m['b'], m['i'], m['s'], m['p'], m['a'], m['iq'], m['rem'], m['depth'], m['depth']==0)
        if not m['rendered']:
            m['block'].extend(rows)
            if self.preview and self.preview.winfo_exists():
                self.root.after(0, lambda rs=list(rows): [self.preview.add_row(r) for r in rs])
            m['rendered'] = True
        m['rem'] = n_rem; matches = [x for x in matches if x['mtc_inv'] not in m['processed']]; rem_str = rows[0].get('remarks','') if rows else ''
        if not matches or any(x in rem_str for x in ["Root cause","REBNI","SR","Found"]): self.all_blocks.append(m['block']); self._next_man(); return
        self.root.after(0, lambda: self._show_dlg(matches))

    def _show_dlg(self, matches):
        m, first = self.curr_m, matches[0]
        if first['mtc_inv'] in self.engine.cache_sid: self._handle_res({'action':'valid','chosen_match':first,'sid':self.engine.cache_sid[first['mtc_inv']],'barcode':self.engine.cache_bc.get(first['mtc_inv'],"[DICES]")}, matches); return
        ManualLevelDialog(self.root, matches, m['rem'], m['budget'], lambda res: self._handle_res(res, matches))

    def _handle_res(self, res, matches):
        if not res or res['action'] == 'stop': self.all_blocks.append(self.curr_m['block']); self._next_man(); return
        match = res.get('chosen_match'); 
        if match: self.curr_m['processed'].add(match['mtc_inv'])
        if res['action'] == 'invalid':
            excl = res['invalid_qty']; row = {'barcode':'[INVALID]','invoice':match['mtc_inv'],'sid':'—','po':match['mtc_po'],'asin':match['mtc_asin'],'inv_qty':fmt_qty(match['inv_qty']),'rec_qty':'','mtc_qty':'','mtc_inv':'','remarks':f"{int(excl)} units excluded",'date':'','depth':self.curr_m['depth'],'type':'subrow'}
            self.curr_m['block'].append(row); self.preview.add_row(row); self.curr_m['rem'] = max(0, self.curr_m['rem'] - excl)
            if self.curr_m['rem'] <= 0: self.all_blocks.append(self.curr_m['block']); self._next_man()
            else: rem_m = [x for x in matches if x['mtc_inv'] != match['mtc_inv']]; (self.root.after(0, lambda: self._show_dlg(rem_m)) if rem_m else (self.all_blocks.append(self.curr_m['block']), self._next_man()))
        elif res['action'] == 'cross_po':
            cands = self.engine.detect_cross_po(self.curr_m['s'], self.curr_m['p'], self.curr_m['a'])
            if cands: CrossPODialog(self.root, cands, self.curr_m['i'], self.curr_m['s'], lambda r: self._handle_cross_po(r))
            else: self.root.after(0, lambda: self._show_dlg(matches))
        elif res['action'] == 'mismatch':
            data = res['mismatch_data']; row = {'barcode':'[MISMATCH]','invoice':'','sid':data.get('sid',''),'po':data.get('po',''),'asin':data.get('asin',''),'inv_qty':data.get('inv_qty',''),'rec_qty':data.get('ovg_qty',''),'mtc_qty':'','mtc_inv':'','remarks':f"Mismatch: {data['asin']} Rec={data['ovg_qty']}",'date':'','depth':self.curr_m['depth'],'type':'subrow'}
            self.curr_m['block'].append(row)
            self.preview.add_row(row)
            rem_m = [x for x in matches if not match or x['mtc_inv'] != match['mtc_inv']]
            (self.root.after(0, lambda: self._show_dlg(rem_m)) if rem_m else (self.all_blocks.append(self.curr_m['block']), self._next_man()))
        else:
            self.engine.cache_sid[match['mtc_inv']] = res['sid']; self.engine.cache_bc[match['mtc_inv']] = res['barcode']
            self.curr_m.update({'b': res['barcode'], 'i': match['mtc_inv'], 's': res['sid'], 'p': match['mtc_po'], 'a': match['mtc_asin'], 'iq': match['inv_qty'], 'budget': safe_num(match['mtc_qty']), 'depth': self.curr_m['depth'] + 1, 'rendered': False, 'processed': set()}); threading.Thread(target=self._man_step, daemon=True).start()

    def _handle_cross_po(self, res):
        if res['action'] == 'skip':
            return
        cand = res['candidate']
        c_type = res.get('case', 'Case 1')
        budget = safe_num(cand.get('rec_qty', 0))

        def investigate():
            cross_rows, found = self.engine.run_cross_po_investigation(cand, c_type, budget, depth=self.curr_m['depth'] + 1)
            for row in cross_rows:
                self.curr_m['block'].append(row)
                if self.preview and self.preview.winfo_exists():
                    self.root.after(0, lambda r=row: self.preview.add_row(r))
            self.curr_m['rem'] = max(0.0, self.curr_m['rem'] - safe_num(found))
            if self.curr_m['rem'] <= 0:
                self.root.after(100, lambda: (self.all_blocks.append(self.curr_m['block']), self._next_man()))
            else:
                self.root.after(100, self._man_step)

        threading.Thread(target=investigate, daemon=True).start()

    def _finish(self):
        self._set_status("Investigation Done", 100); self.root.after(0, lambda: (self.save_btn.config(state="normal"), self.run_btn.config(state="normal"), self.stop_btn.config(state="disabled")))

    def save_output(self):
        p = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if p: write_excel(self.all_blocks, p); messagebox.showinfo("Saved", p)

    def run(self): self.root.mainloop()

# ═══════════════════════════════════════════════════════════

if __name__ == '__main__': MFIToolApp().run()
