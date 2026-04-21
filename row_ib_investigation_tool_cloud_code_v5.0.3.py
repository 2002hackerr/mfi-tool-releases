"""
MFI Investigation Tool  v5.0.3
==============================
ROW IB  |  Amazon
Developed by Mukesh

CHANGES IN v5.0.3 (over v5.0.2):
  ✔ _handle_res: replaced flat pending_siblings with siblings_stack (per-depth context stack)
    — siblings from each depth level are now tracked independently, never overwritten
  ✔ _handle_res: 'rem' now correctly set to match mtc_qty when diving into sub-invoice
  ✔ _man_step: siblings_stack used to resume correct sibling at correct depth after branch done
  ✔ _man_step: removed 'Found' from should_stop keywords (caused premature stop at sub-levels)
  ✔ run_cross_po_investigation: uses inv_p primary-only lookup (no _inv_lookup fallback)
  ✔ run_cross_po_investigation: uses _resolve_inv_qty for correct SCR invoice quantities
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, re, threading
from datetime import datetime


# ═══════════════════════════════════════════════════════════
#  UTILITIES
# ═══════════════════════════════════════════════════════════

def extract_sid(val):
    s = str(val).strip()
    parts = re.findall(r'\d{10,}', s)
    return max(parts, key=len) if parts else s

def strip_scr(inv_no):
    """
    Remove trailing SCR suffix(es) from an invoice number.
    e.g., '9527185027SCRSCRSCR' → '9527185027'
         '9527194660SCR'        → '9527194660'
         '9527200210'           → '9527200210'  (no change)
    """
    import re
    return re.sub(r'(?:SCR)+$', '', str(inv_no).strip(), flags=re.IGNORECASE)

def safe_num(val):
    try:
        if pd.isna(val): return 0.0
    except: pass
    try:
        return float(str(val).replace(',', '').strip())
    except: return 0.0

def clean(val):
    try:
        if pd.isna(val): return ""
    except: pass
    return str(val).strip()

def split_comma(val):
    if not val: return []
    try:
        if pd.isna(val): return []
    except: pass
    return [s.strip() for s in str(val).split(',') if s.strip()]

def fmt_qty(val):
    n = safe_num(val)
    if n == 0: return ""
    return str(int(n)) if n == int(n) else str(n)


# ═══════════════════════════════════════════════════════════
#  DIALOGS
# ═══════════════════════════════════════════════════════════


class HeaderCorrectionDialog(tk.Toplevel):
    """
    Shown when claims file has non-standard column headers.
    Lists what was detected vs canonical name.
    User can confirm auto-correct or cancel.
    """
    def __init__(self, parent, corrections, mapping, df_columns, callback):
        super().__init__(parent)
        self.callback    = callback
        self.corrections = corrections
        self.mapping     = mapping
        self.df_columns  = df_columns

        self.title("Column Header Mismatch Detected — v5.0.3")
        self.geometry("700x480")
        self.configure(bg="#0f0f1a")
        self.resizable(True, True)
        # self.transient(parent)  # Removed to ensure minimize/maximize buttons
        self.lift(); self.focus_force()

        # Header
        tk.Label(self,
                 text="⚠  Non-standard column headers detected in Claims file",
                 bg="#16213e", fg="#f0a500",
                 font=("Segoe UI",12,"bold"), height=2).pack(fill="x")
        tk.Label(self,
                 text="The tool has automatically matched the columns below. "
                      "Please confirm or correct before proceeding.",
                 bg="#0f0f1a", fg="#cccccc",
                 font=("Segoe UI",9)).pack(pady=4)

        # Table showing all field mappings
        outer = tk.Frame(self, bg="#0f0f1a"); outer.pack(fill="both", expand=True, padx=16, pady=6)

        # Header row
        hdrs = ["Field", "Expected", "Found in file", "Status"]
        w = [14, 20, 28, 12]
        for ci, (h, ww) in enumerate(zip(hdrs, w)):
            tk.Label(outer, text=h, bg="#203864", fg="white",
                     font=("Calibri",10,"bold"),
                     width=ww, anchor="w", padx=4).grid(row=0, column=ci, padx=1, pady=1, sticky="w")

        # All fields — show green (exact) or yellow (corrected) or red (missing)
        all_fields = list(COLUMN_ALIASES.keys())
        corrected_fields = {c[0] for c in corrections}
        missing_fields   = [f for f in all_fields if f not in mapping]

        self._override_vars = {}

        for ri, field in enumerate(all_fields, 1):
            canonical = COLUMN_ALIASES[field][0]
            found_col = mapping.get(field, "")
            is_corrected = field in corrected_fields
            is_missing   = field not in mapping

            if is_missing:
                status_txt, status_fg, row_bg = "MISSING", "#ff4444", "#2a0000"
            elif is_corrected:
                status_txt, status_fg, row_bg = "Auto-fixed", "#f0a500", "#1a1500"
            else:
                status_txt, status_fg, row_bg = "✔ OK", "#44ff88", "#001a00"

            tk.Label(outer, text=field, bg=row_bg, fg="#e0e0e0",
                     font=("Calibri",10,"bold"), width=14, anchor="w", padx=4
                     ).grid(row=ri, column=0, padx=1, pady=1, sticky="w")
            tk.Label(outer, text=canonical, bg=row_bg, fg="#aaaacc",
                     font=("Calibri",10), width=20, anchor="w", padx=4
                     ).grid(row=ri, column=1, padx=1, pady=1, sticky="w")

            # Dropdown for found column — allows user to override
            v = tk.StringVar(value=found_col or "— not found —")
            self._override_vars[field] = v
            opts = ["— not found —"] + list(df_columns)
            cb = ttk.Combobox(outer, textvariable=v, values=opts,
                              state="readonly", width=26, font=("Calibri",9))
            cb.grid(row=ri, column=2, padx=1, pady=1, sticky="w")

            tk.Label(outer, text=status_txt, bg=row_bg, fg=status_fg,
                     font=("Calibri",10,"bold"), width=12, anchor="w", padx=4
                     ).grid(row=ri, column=3, padx=1, pady=1, sticky="w")

        # Bottom note
        tk.Label(self,
                 text="You can change any column assignment using the dropdowns above.",
                 bg="#0f0f1a", fg="#888899", font=("Segoe UI",8)).pack(pady=2)

        # Buttons
        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=10)
        tk.Button(bf, text="✔  Auto-correct & Proceed",
                  command=self._proceed,
                  bg="#2d6a4f", fg="white",
                  font=("Segoe UI",12,"bold"),
                  padx=20, pady=8, relief="flat",
                  cursor="hand2").pack(side="left", padx=10)
        tk.Button(bf, text="✖  Cancel",
                  command=self._cancel,
                  bg="#4a2020", fg="white",
                  font=("Segoe UI",11),
                  padx=16, pady=8, relief="flat",
                  cursor="hand2").pack(side="left", padx=10)

        self.protocol("WM_DELETE_WINDOW", self._cancel)
        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _proceed(self):
        # Build final mapping from user selections
        final = {}
        for field, var in self._override_vars.items():
            v = var.get()
            if v and v != "— not found —":
                final[field] = v
        self.callback({'action': 'proceed', 'mapping': final})
        self.destroy()

    def _cancel(self):
        self.callback({'action': 'cancel'})
        self.destroy()

class SIDRequestDialog(tk.Toplevel):
    """Auto mode: SID not found in REBNI → ask user."""
    def __init__(self, parent, invoice, po, asin, callback):
        super().__init__(parent)
        self.callback = callback
        self.title("SID Required — DICES Validation")
        self.geometry("540x210")
        self.resizable(True, True)
        self.configure(bg="#16213e")
        # self.transient(parent)  # Removed to ensure minimize/maximize buttons
        self.lift(); self.focus_force()

        tk.Label(self, text="⚠  SID Not Found in REBNI",
                 bg="#16213e", fg="#e94560",
                 font=("Segoe UI",13,"bold")).pack(pady=(14,4))
        tk.Label(self, text=f"Invoice: {invoice}   PO: {po}   ASIN: {asin}",
                 bg="#16213e", fg="#e0e0e0", font=("Segoe UI",9)).pack(pady=2)
        tk.Label(self, text="Validate this invoice in DICES and enter the SID below:",
                 bg="#16213e", fg="#aaaacc", font=("Segoe UI",9)).pack(pady=6)

        ef = tk.Frame(self, bg="#16213e"); ef.pack()
        tk.Label(ef, text="SID from DICES:", bg="#16213e", fg="#e0e0e0",
                 font=("Segoe UI",10)).pack(side="left", padx=8)
        self._sid = tk.StringVar()
        self._entry = tk.Entry(ef, textvariable=self._sid, width=30,
                               font=("Segoe UI",10), bg="#1e1e3a", fg="#e0e0e0",
                               insertbackground="white", relief="flat")
        self._entry.pack(side="left", padx=4)
        self._entry.focus_set()

        bf = tk.Frame(self, bg="#16213e"); bf.pack(pady=12)
        tk.Button(bf, text="✔  Continue", command=self._ok,
                  bg="#2d6a4f", fg="white", font=("Segoe UI",11,"bold"),
                  padx=16, pady=7, relief="flat", cursor="hand2").pack(side="left", padx=8)
        tk.Button(bf, text="✖  Skip", command=self._skip,
                  bg="#6b2737", fg="white", font=("Segoe UI",10),
                  padx=16, pady=7, relief="flat", cursor="hand2").pack(side="left", padx=8)

        self.bind('<Return>', lambda e: self._ok())
        self.protocol("WM_DELETE_WINDOW", self._skip)
        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _ok(self):
        sid = extract_sid(self._sid.get().strip())
        if sid:
            self.callback(sid); self.destroy()
        else: self._entry.config(bg="#3a1e1e")

    def _skip(self):
        self.callback(None); self.destroy()


class CrossPODialog(tk.Toplevel):
    """
    Cross PO Confirmation Dialog — v5.0.2
    
    - Resizable and minimizable
    - Non-modal (allows parallel check of PreviewPanel)
    """
    CASE_DESCRIPTIONS = {
        "Case 1": (
            "Case 1 — No PO, but ASIN received",
            "Rec=0 at claiming PO. Same ASIN received in different PO within same SID.\n"
            "Those units are overage under a different PO."
        ),
        "Case 2": (
            "Case 2 — PO exists but ASIN not ordered there",
            "This PO exists in the claiming SID, but the ASIN was never invoiced there.\n"
            "Inv Qty = 0, but units were received. This is a Cross PO overage."
        ),
        "Case 3": (
            "Case 3 — PO and ASIN exist but Rec > Inv",
            "Both PO and ASIN are present. Invoiced qty = X, but received more than X.\n"
            "Excess units are Cross PO overage."
        ),
    }

    def __init__(self, parent, candidates, current_inv, sid, callback):
        super().__init__(parent)
        self.callback   = callback
        self.candidates = candidates

        self.title("Cross PO Overage — Confirm & Investigate")
        self.geometry("740x540")
        self.resizable(True, True)
        self.configure(bg="#0f0f1a")
        # self.transient(parent)  # Removed to ensure minimize/maximize buttons
        self.lift(); self.focus_force()

        # Header
        tk.Label(self, text="🔄  Cross PO Overage Detected",
                 bg="#16213e", fg="#f0a500",
                 font=("Segoe UI",13,"bold"), height=2).pack(fill="x")
        tk.Label(self,
                 text=f"SID: {sid}   |   Investigation Invoice: {current_inv}",
                 bg="#0f0f1a", fg="#cccccc",
                 font=("Segoe UI",9)).pack(pady=2)
        tk.Label(self,
                 text="On confirming, the tool will investigate the Cross PO chain "
                      "to find equivalent shortage.",
                 bg="#0f0f1a", fg="#4a9eff",
                 font=("Segoe UI",9)).pack(pady=2)

        # Candidate table
        tf = tk.LabelFrame(self, text="  Detected Cross PO Candidates  ",
                           bg="#0f0f1a", fg="#e0e0e0",
                           font=("Segoe UI",9,"bold"), padx=10, pady=6)
        tf.pack(fill="x", padx=16, pady=6)
        for ci, h in enumerate(["Cross PO","ASIN","Inv Qty","Rec Qty","Overage","Type"]):
            tk.Label(tf, text=h, bg="#203864", fg="white",
                     font=("Calibri",10,"bold"), width=14, anchor="w", padx=3
                     ).grid(row=0, column=ci, padx=1, pady=1)
        for ri, c in enumerate(candidates, 1):
            inv_n = safe_num(c.get('inv_qty', 0))
            rec_n = safe_num(c['rec_qty'])
            ovg   = max(0.0, rec_n - inv_n)
            for ci, v in enumerate([c['po'], c['asin'],
                                     fmt_qty(inv_n), fmt_qty(rec_n),
                                     fmt_qty(ovg) or "—",
                                     c['cross_type'].split("—")[0].strip()]):
                tk.Label(tf, text=str(v), bg="#1e1e3a", fg="#e0e0e0",
                         font=("Calibri",10), width=14, anchor="w", padx=3
                         ).grid(row=ri, column=ci, padx=1, pady=1)

        # Candidate selection
        sf = tk.Frame(self, bg="#0f0f1a"); sf.pack(fill="x", padx=16, pady=4)
        tk.Label(sf, text="Select Cross PO to investigate:",
                 bg="#0f0f1a", fg="#e0e0e0",
                 font=("Segoe UI",10), width=30, anchor="w").pack(side="left")
        opts = [f"PO={c['po']}  Rec={fmt_qty(c['rec_qty'])}  {c['cross_type'].split(chr(8212))[0].strip()}"
                for c in candidates] + ["None — Skip"]
        self._sel_var = tk.StringVar()
        self._sel_cb  = ttk.Combobox(sf, textvariable=self._sel_var,
                                      values=opts, state="readonly", width=50,
                                      font=("Segoe UI",9))
        self._sel_cb.current(0)
        self._sel_cb.pack(side="left", padx=6)
        self._sel_cb.bind("<<ComboboxSelected>>", self._on_candidate_change)

        # Case selection
        cf = tk.LabelFrame(self, text="  Confirm Cross PO Case  ",
                           bg="#0f0f1a", fg="#e0e0e0",
                           font=("Segoe UI",9,"bold"), padx=12, pady=8)
        cf.pack(fill="x", padx=16, pady=4)

        self._case_var = tk.StringVar(value="Case 1")
        self._case_desc_lbl = tk.Label(cf, text="",
                                        bg="#0f0f1a", fg="#aaaacc",
                                        font=("Segoe UI",9), justify="left",
                                        wraplength=640, anchor="w")

        for case_key, (case_label, case_desc) in self.CASE_DESCRIPTIONS.items():
            tk.Radiobutton(cf, text=case_label,
                           variable=self._case_var, value=case_key,
                           bg="#0f0f1a", fg="#f0c060",
                           selectcolor="#1a1500",
                           font=("Segoe UI",10,"bold"),
                           command=self._on_case_change
                           ).pack(anchor="w", pady=2)

        self._case_desc_lbl.pack(anchor="w", pady=4, padx=8)
        self._on_case_change()  # init description

        # Investigation note
        tk.Label(self,
                 text="⚡  On confirming: tool will investigate Cross PO chain "
                      "until full Cross PO rec_qty is explained as shortage.",
                 bg="#0f0f1a", fg="#88ccff",
                 font=("Segoe UI",9,"italic")).pack(pady=4, padx=16, anchor="w")

        # Buttons
        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=10)
        tk.Button(bf, text="✔  Confirm & Investigate",
                  command=self._confirm,
                  bg="#2d6a4f", fg="white",
                  font=("Segoe UI",12,"bold"),
                  padx=20, pady=9, relief="flat",
                  cursor="hand2").pack(side="left", padx=10)
        tk.Button(bf, text="✖  Skip",
                  command=self._skip,
                  bg="#4a2020", fg="white",
                  font=("Segoe UI",11),
                  padx=16, pady=9, relief="flat",
                  cursor="hand2").pack(side="left", padx=10)

        self.protocol("WM_DELETE_WINDOW", self._skip)
        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _on_candidate_change(self, event=None):
        pass

    def _on_case_change(self):
        case_key = self._case_var.get()
        _, desc  = self.CASE_DESCRIPTIONS.get(case_key, ("", ""))
        self._case_desc_lbl.config(text=desc)

    def _confirm(self):
        idx = self._sel_cb.current()
        if idx >= len(self.candidates):
            self.callback({'action': 'skip'}); self.destroy(); return
        self.callback({
            'action'   : 'confirmed',
            'candidate': self.candidates[idx],
            'case'     : self._case_var.get(),
        })
        self.destroy()

    def _skip(self):
        self.callback({'action': 'skip'}); self.destroy()


class ManualLevelDialog(tk.Toplevel):
    """Manual mode: per-level dialog with IBC/PBC + Cross PO + Mismatches."""
    def __init__(self, parent, matches, remaining_pqv, branch_budget, callback):
        super().__init__(parent)
        self.callback = callback
        self.matches = matches
        self.rem_pqv = remaining_pqv
        self.branch_budget = branch_budget

        self.title("Manual Investigation — Next Step")
        self.geometry("660x500")
        self.configure(bg="#0f0f1a")
        self.resizable(True, True)
        # self.transient(parent)  # Removed to ensure minimize/maximize buttons
        self.lift(); self.focus_force()

        # Header
        tk.Label(self, text="  Manual Investigation — Continue",
                 bg="#16213e", fg="#4a9eff",
                 font=("Segoe UI",12,"bold"), height=2).pack(fill="x")

        info = f"Remaining PQV: {int(remaining_pqv)}    Branch budget: {int(branch_budget)}"
        tk.Label(self, text=info, bg="#0f0f1a", fg="#cccccc",
                 font=("Segoe UI",9)).pack(pady=2)

        # Invoice selection
        inv_f = tk.LabelFrame(self, text="  Select Invoice to Continue  ",
                              font=("Segoe UI",9,"bold"),
                              bg="#0f0f1a", fg="#e0e0e0", padx=10, pady=6)
        inv_f.pack(fill="x", padx=16, pady=4)
        opts = [f"Qty={fmt_qty(m['mtc_qty'])}  |  Inv={m['mtc_inv']}  |  PO={m['mtc_po']}  |  ASIN={m['mtc_asin']}"
                for m in matches]
        self._branch_var = tk.StringVar()
        self._branch_cb = ttk.Combobox(inv_f, textvariable=self._branch_var,
                                        values=opts, state="readonly", width=70,
                                        font=("Segoe UI",9))
        if opts: self._branch_cb.current(0)
        self._branch_cb.pack()

        # IBC/PBC section
        ibc_f = tk.LabelFrame(self, text="  IBC = PBC Validation  ",
                               font=("Segoe UI",9,"bold"),
                               bg="#0f0f1a", fg="#e0e0e0", padx=10, pady=6)
        ibc_f.pack(fill="x", padx=16, pady=4)

        self._validity = tk.StringVar(value="valid")
        rf = tk.Frame(ibc_f, bg="#0f0f1a"); rf.pack(fill="x")
        tk.Radiobutton(rf, text="✔  IBC = PBC  VALID — Continue investigation",
                       variable=self._validity, value="valid",
                       bg="#0f0f1a", fg="#90ee90", selectcolor="#1e3a28",
                       font=("Segoe UI",10,"bold"),
                       command=self._toggle).pack(side="left", padx=6)
        tk.Radiobutton(rf, text="✗  IBC ≠ PBC  INVALID — Exclude units",
                       variable=self._validity, value="invalid",
                       bg="#0f0f1a", fg="#ff8888", selectcolor="#3a1e1e",
                       font=("Segoe UI",10,"bold"),
                       command=self._toggle).pack(side="left", padx=14)

        self._invalid_frame = tk.Frame(ibc_f, bg="#0f0f1a"); self._invalid_frame.pack(fill="x", pady=3)
        tk.Label(self._invalid_frame, text="Units matched to invalid invoice:",
                 bg="#0f0f1a", fg="#ff8888", font=("Segoe UI",9)).pack(side="left", padx=4)
        self._inv_qty_var = tk.StringVar()
        tk.Entry(self._invalid_frame, textvariable=self._inv_qty_var, width=10,
                 font=("Segoe UI",10), bg="#1e1e3a", fg="#ff8888",
                 insertbackground="white", relief="flat").pack(side="left", padx=4)

        # DICES details (for valid)
        self._dices_frame = tk.LabelFrame(self, text="  DICES Details  ",
                                           font=("Segoe UI",9,"bold"),
                                           bg="#0f0f1a", fg="#e0e0e0", padx=10, pady=6)
        self._dices_frame.pack(fill="x", padx=16, pady=4)
        r1 = tk.Frame(self._dices_frame, bg="#0f0f1a"); r1.pack(fill="x", pady=2)
        tk.Label(r1, text="SID from DICES:", bg="#0f0f1a", fg="#e0e0e0",
                 font=("Segoe UI",9), width=20, anchor="w").pack(side="left")
        self._sid_var = tk.StringVar()
        tk.Entry(r1, textvariable=self._sid_var, width=28, font=("Segoe UI",9),
                 bg="#1e1e3a", fg="#e0e0e0", insertbackground="white", relief="flat").pack(side="left", padx=4)

        r2 = tk.Frame(self._dices_frame, bg="#0f0f1a"); r2.pack(fill="x", pady=2)
        tk.Label(r2, text="Barcode from DICES:", bg="#0f0f1a", fg="#e0e0e0",
                 font=("Segoe UI",9), width=20, anchor="w").pack(side="left")
        self._bc_var = tk.StringVar()
        tk.Entry(r2, textvariable=self._bc_var, width=28, font=("Segoe UI",9),
                 bg="#1e1e3a", fg="#e0e0e0", insertbackground="white", relief="flat").pack(side="left", padx=4)

        self._toggle()

        # Buttons
        bf = tk.Frame(self, bg="#0f0f1a"); bf.pack(pady=10)
        tk.Button(bf, text="▶  CONTINUE",
                  command=self._ok, bg="#2d6a4f", fg="white",
                  font=("Segoe UI",12,"bold"), padx=16, pady=8,
                  relief="flat", cursor="hand2").pack(side="left", padx=6)
        tk.Button(bf, text="🔄  CROSS PO",
                  command=self._cross_po, bg="#7a5c00", fg="white",
                  font=("Segoe UI",10,"bold"), padx=12, pady=8,
                  relief="flat", cursor="hand2").pack(side="left", padx=6)
        tk.Button(bf, text="⚖  MISMATCH / OVERAGE",
                  command=self._mismatch, bg="#2d4a7a", fg="white",
                  font=("Segoe UI",10,"bold"), padx=12, pady=8,
                  relief="flat", cursor="hand2").pack(side="left", padx=6)
        tk.Button(bf, text="⬛  STOP THIS ASIN",
                  command=self._stop, bg="#4a2020", fg="white",
                  font=("Segoe UI",10), padx=12, pady=8,
                  relief="flat", cursor="hand2").pack(side="left", padx=6)

        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{px}+{py}")

    def _toggle(self):
        if self._validity.get() == "valid":
            self._invalid_frame.pack_forget()
            self._dices_frame.pack(fill="x", padx=16, pady=4)
        else:
            self._dices_frame.pack_forget()
            self._invalid_frame.pack(fill="x", pady=3)

    def _ok(self):
        sel = self._branch_cb.current()
        if sel < 0 or sel >= len(self.matches):
            messagebox.showwarning("Select Invoice", "Please select an invoice.", parent=self); return
        match = self.matches[sel]
        if self._validity.get() == "valid":
            sid = extract_sid(self._sid_var.get().strip())
            if not sid:
                messagebox.showwarning("SID Required", "Please enter SID from DICES.", parent=self); return
            self.callback({'action':'valid', 'chosen_match':match,
                           'sid':sid, 'barcode':self._bc_var.get().strip() or "[DICES]"})
        else:
            qty_str = self._inv_qty_var.get().strip()
            try: qty = float(qty_str)
            except:
                messagebox.showwarning("Qty Required", "Enter units matched to invalid invoice.", parent=self); return
            self.callback({'action':'invalid', 'chosen_match':match, 'invalid_qty':qty})
        self.destroy()

    def _cross_po(self):
        self.callback({'action':'cross_po',
                       'chosen_match': self.matches[self._branch_cb.current()] if self.matches else None})
        self.destroy()

    def _mismatch(self):
        # Mini dialog for mismatch/overage details
        dlg = tk.Toplevel(self)
        dlg.title("Mismatch / Overage Details")
        dlg.geometry("460x260")
        dlg.configure(bg="#0f0f1a")
        # self.grab_set()  # Removed to allow parallel interaction
        dlg.lift(); dlg.focus_force()

        fields = [("ASIN received:", "asin"), ("SID:", "sid"), ("PO:", "po"),
                  ("Inv Qty (invoiced):", "inv_qty"), ("Overage Qty received:", "ovg_qty")]
        vars_ = {}
        for i, (lbl, key) in enumerate(fields):
            tk.Label(dlg, text=lbl, bg="#0f0f1a", fg="#e0e0e0",
                     font=("Segoe UI",10), width=22, anchor="w").grid(row=i, column=0, padx=12, pady=5)
            v = tk.StringVar()
            tk.Entry(dlg, textvariable=v, width=26, font=("Segoe UI",10),
                     bg="#1e1e3a", fg="#e0e0e0", insertbackground="white",
                     relief="flat").grid(row=i, column=1, padx=8, pady=5)
            vars_[key] = v

        def submit():
            data = {k: v.get().strip() for k, v in vars_.items()}
            dlg.destroy()
            self.callback({'action':'mismatch', 'mismatch_data': data})
            self.destroy()

        tk.Button(dlg, text="✔  Submit Mismatch", command=submit,
                  bg="#2d6a4f", fg="white", font=("Segoe UI",11,"bold"),
                  padx=14, pady=7, relief="flat", cursor="hand2"
                  ).grid(row=len(fields), column=0, columnspan=2, pady=12)

    def _stop(self):
        self.callback({'action':'stop'}); self.destroy()


class PreviewPanel(tk.Toplevel):
    COLS = ['Barcode','Inv no','SID','PO','ASIN','Inv Qty','Rec Qty','Mtc Qty','Mtc Inv','Remarks','Date']
    COL_W_PX = [130,160,130,90,110,60,60,60,160,240,150]

    def __init__(self, parent):
        super().__init__(parent)
        self.title("Investigation Preview — Manual Mode (editable)")
        self.geometry("1280x520")
        self.configure(bg="#0f0f1a")
        self.resizable(True, True)

        tk.Label(self, text="  Live Investigation Preview — double-click any cell to edit",
                 bg="#16213e", fg="#4a9eff",
                 font=("Segoe UI",10,"bold"), height=2).pack(fill="x")

        frame = tk.Frame(self, bg="#0f0f1a")
        frame.pack(fill="both", expand=True, padx=8, pady=6)
        vsb = ttk.Scrollbar(frame, orient="vertical")
        hsb = ttk.Scrollbar(frame, orient="horizontal")
        self.tree = ttk.Treeview(frame, columns=self.COLS, show='headings',
                                  yscrollcommand=vsb.set, xscrollcommand=hsb.set, height=22)
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        for col, w in zip(self.COLS, self.COL_W_PX):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, minwidth=40, anchor='w')
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.grid_rowconfigure(0, weight=1); frame.grid_columnconfigure(0, weight=1)
        self.tree.bind('<Double-1>', self._on_double_click)
        self._row_data = {}

        bb = tk.Frame(self, bg="#0f0f1a"); bb.pack(fill="x", padx=8, pady=4)
        tk.Label(bb, text="Double-click any cell to edit",
                 bg="#0f0f1a", fg="#8888aa", font=("Segoe UI",8)).pack(side="left")
        tk.Button(bb, text="Clear All", command=self.clear_all,
                  bg="#2d2d5e", fg="white", font=("Segoe UI",9),
                  relief="flat", padx=10, pady=4, cursor="hand2").pack(side="right")

        s = ttk.Style()
        s.configure("Treeview", font=("Calibri",10), rowheight=22,
                     background="#1e1e3a", foreground="#e0e0e0", fieldbackground="#1e1e3a")
        s.configure("Treeview.Heading", font=("Calibri",10,"bold"),
                     background="#203864", foreground="white")
        for tag, bg, fg in [
            ('header', '#203864', 'white'), ('d0', '#1e1e3a', '#e0e0e0'),
            ('d1', '#1e3a28', '#e0e0e0'), ('sub', '#1a1a35', '#e0e0e0'),
            ('root', '#3a1e1e', '#ff8888'), ('sr', '#3a1a1a', '#ff6666'),
            ('dices', '#3a3010', '#e0e0e0'), ('invalid', '#3a1010', '#ff9999'),
            ('rebni', '#0d2535', '#88ddff'), ('crosspo', '#2a1a00', '#f0c060'),
            ('mismatch', '#002040', '#66aaff'),
        ]:
            self.tree.tag_configure(tag, background=bg, foreground=fg)

    def add_header_row(self, label=""):
        vals = list(self.COLS)
        if label: vals[4] = f"── {label} ──"
        iid = self.tree.insert('', 'end', values=vals, tags=('header',))
        self._row_data[iid] = dict(zip(self.COLS, vals))

    def add_row(self, rd):
        vals = [rd.get('barcode',''), rd.get('invoice',''), rd.get('sid',''),
                rd.get('po',''), rd.get('asin',''), rd.get('inv_qty',''),
                rd.get('rec_qty',''), rd.get('mtc_qty',''), rd.get('mtc_inv',''),
                rd.get('remarks',''), rd.get('date','')]
        remarks = rd.get('remarks','').lower()
        tag = ('sub'     if rd.get('type') == 'subrow' else
               'root'    if 'root cause' in remarks or 'short' in remarks else
               'sr'      if remarks == 'sr' else
               'invalid' if 'invalid invoice' in remarks else
               'rebni'   if 'rebni available' in remarks else
               'crosspo' if 'cross po' in remarks or rd.get('barcode','') == '[CROSS PO]' else
               'mismatch' if 'mismatch' in remarks else
               'dices'   if '[dices]' in str(rd.get('barcode','')).lower() else
               f"d{min(rd.get('depth',0),1)}")
        iid = self.tree.insert('', 'end', values=vals, tags=(tag,))
        self._row_data[iid] = dict(zip(self.COLS, vals))
        self._row_data[iid]['_rd'] = rd
        self.tree.see(iid)

    def get_all_rows(self):
        KEY = {'Barcode':'barcode','Inv no':'invoice','SID':'sid','PO':'po','ASIN':'asin',
               'Inv Qty':'inv_qty','Rec Qty':'rec_qty','Mtc Qty':'mtc_qty',
               'Mtc Inv':'mtc_inv','Remarks':'remarks','Date':'date'}
        rows = []
        for iid in self.tree.get_children():
            d = self._row_data.get(iid, {})
            if d.get(self.COLS[0]) == self.COLS[0]: continue
            rd = d.get('_rd', {}).copy()
            for col in self.COLS: rd[KEY[col]] = d.get(col, '')
            rows.append(rd)
        return rows

    def clear_all(self):
        for iid in self.tree.get_children(): self.tree.delete(iid)
        self._row_data.clear()

    def _on_double_click(self, event):
        region = self.tree.identify_region(event.x, event.y)
        if region != 'cell':
            return
        iid = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if not iid or not col:
            return

        col_idx  = int(col.replace('#', '')) - 1
        col_name = self.COLS[col_idx]

        bbox = self.tree.bbox(iid, col)
        if not bbox:
            return
        x, y, w, h = bbox

        current  = self._row_data.get(iid, {}).get(col_name, '')
        ev       = tk.StringVar(value=str(current))

        e = tk.Entry(self.tree, textvariable=ev,
                     font=("Calibri", 10),
                     bg="#2d2d5e", fg="white",
                     insertbackground="white",
                     relief="flat", bd=1)
        e.place(x=x, y=y, width=w, height=h)
        e.focus_force()
        e.select_range(0, 'end')

        def save(ev_=None):
            nv = ev.get()
            if iid in self._row_data:
                self._row_data[iid][col_name] = nv
            vals        = list(self.tree.item(iid, 'values'))
            vals[col_idx] = nv
            self.tree.item(iid, values=vals)
            try: e.destroy()
            except: pass

        e.bind('<Return>',   save)
        e.bind('<Tab>',      save)
        e.bind('<Escape>',   lambda _: e.destroy())
        e.bind('<FocusOut>', save)


# ═══════════════════════════════════════════════════════════
#  DATA LOADERS
# ═══════════════════════════════════════════════════════════

def _load_file(path):
    """Load Excel or CSV file into DataFrame."""
    ext = os.path.splitext(path)[1].lower()
    if ext == '.csv':
        try: return pd.read_csv(path, dtype=str, encoding='utf-8')
        except: return pd.read_csv(path, dtype=str, encoding='latin-1')
    else:
        return pd.read_excel(path, header=0, dtype=str)

def load_claims(path): return _load_file(path)

COLUMN_ALIASES = {
    'Barcode': [
        'barcode', 'bar code', 'bar_code', 'upc', 'ean',
        'scan code', 'item code', 'carton barcode', 'pkg barcode',
    ],
    'Invoice': [
        'inv no', 'inv_no', 'invoice_no', 'invoice no', 'invoice number',
        'invoice_number', 'invoice', 'inv num', 'inv number',
        'inv#', 'invoice#', 'invc no', 'invc number', 'invc',
        'bill no', 'bill number',
    ],
    'SID': [
        'sid', 'shipment id', 'shipment_id', 'shipment no',
        'shipment number', 'ship id', 'fba shipment id',
        'inbound shipment', 'shipment', 'inbound sid',
    ],
    'PO': [
        'po', 'po_no', 'p.o.', 'po no', 'po number', 'po#',
        'purchase order', 'purchase_order', 'purchase order no',
        'purchase order number', 'po id', 'order no', 'order number',
        'purch order', 'header po', 'line po', 'po num',
        'invoice_lineitem_po',
    ],
    'ASIN': [
        'asin', 'po_asin', 'amazon asin', 'amazon product id',
        'product id', 'asin no', 'item asin', 'amazon id',
    ],
    'InvQty': [
        'inv qty', 'inv_qty', 'invoice qty', 'invoice quantity',
        'invoiced qty', 'invoiced quantity', 'quantity invoiced',
        'quantity_invoiced', 'billed qty', 'billed quantity',
        'total qty', 'total quantity', 'item qty',
        'po_ordered_quantity',
    ],
    'PQV': [
        'pqv', 'pqv qty', 'missing qty', 'missing quantity',
        'missing_qty', 'shortage', 'short qty', 'short quantity',
        'claim qty', 'claimed qty', 'dispute qty',
        'pending qty', 'outstanding qty', 'difference qty',
    ],
}

def detect_claim_cols(df):
    actual_cols = list(df.columns)
    lower_map   = {c.lower().strip(): c for c in actual_cols}
    mapping      = {}
    corrections  = []

    for field, aliases in COLUMN_ALIASES.items():
        found = None
        for alias in aliases:
            if alias in lower_map:
                found = lower_map[alias]; break
        if not found:
            for alias in aliases:
                for col in actual_cols:
                    if alias in col.lower() or col.lower() in alias:
                        found = col; break
                if found: break
        if found:
            canonical = aliases[0]
            if found.lower().strip() != canonical:
                corrections.append((field, found, canonical))
            mapping[field] = found
    return mapping, corrections

def load_rebni(path):
    df = _load_file(path)
    names = ['vendor_code','po','asin','shipment_id','received_datetime',
             'warehouse_id','item_cost','quantity_unpacked','quantity_adjusted',
             'qty_received_postadj','quantity_matched','rebni_available',
             'cnt_invoice_matched','matched_invoice_numbers']
    df.columns = names[:len(df.columns)]
    return df

def load_invoice_search(path):
    df = _load_file(path)
    names = ['vendor_code','purchase_order_id','asin','invoice_number','invoice_date',
             'invoice_item_status','quantity_invoiced','quantity_matched_total',
             'no_of_shipments','shipment_id','shipmentwise_matched_qty','matched_po','matched_asin']
    df.columns = names[:len(df.columns)]
    return df


# ═══════════════════════════════════════════════════════════
#  INDEX BUILDERS
# ═══════════════════════════════════════════════════════════

def build_rebni_index(df):
    p, s, fb = {}, {}, {}
    for _, row in df.iterrows():
        sid  = extract_sid(clean(row.get('shipment_id','')))
        po   = clean(row.get('po',''))
        asin = clean(row.get('asin',''))
        if not sid or not asin: continue
        p.setdefault((sid, po, asin), []).append(row.to_dict())
        s.setdefault((po, asin), []).append(row.to_dict())
        for inv in split_comma(row.get('matched_invoice_numbers','')):
            if inv: fb.setdefault((sid, po, inv), []).append(row.to_dict())
    return p, s, fb

def build_invoice_index(df):
    idx, fb = {}, {}
    for _, row in df.iterrows():
        sids  = split_comma(row.get('shipment_id',''))
        pos   = split_comma(row.get('matched_po',''))
        asins = split_comma(row.get('matched_asin',''))
        qtys  = split_comma(row.get('shipmentwise_matched_qty',''))
        for i in range(max(len(sids), len(pos), len(asins), len(qtys))):
            s_frag = extract_sid(sids[i] if i < len(sids) else "")
            p_val  = pos[i]   if i < len(pos)   else ""
            a_val  = asins[i] if i < len(asins) else ""
            q_val  = safe_num(qtys[i] if i < len(qtys) else "0")
            inv_no   = clean(row.get('invoice_number',''))
            mtc_po   = clean(row.get('purchase_order_id',''))
            mtc_asin = clean(row.get('asin',''))
            if not s_frag or not p_val or not a_val: continue
            entry = {'mtc_inv':inv_no, 'mtc_po':mtc_po, 'mtc_asin':mtc_asin,
                     'inv_qty':safe_num(row.get('quantity_invoiced','0')),
                     'mtc_qty':q_val, 'date':clean(row.get('invoice_date',''))}
            idx.setdefault((s_frag, p_val, a_val), []).append(entry)
            if inv_no: fb.setdefault((s_frag, p_val, inv_no), []).append(entry)
    return idx, fb


# ═══════════════════════════════════════════════════════════
#  INVESTIGATION ENGINE
# ═══════════════════════════════════════════════════════════

class InvestigationEngine:
    MAX_DEPTH = 10

    def __init__(self, rp, rs, rfb, ip, ifb, sid_cb=None):
        self.rebni_p  = rp
        self.rebni_s  = rs
        self.rebni_fb = rfb
        self.inv_p    = ip
        self.inv_fb   = ifb
        self.sid_cb   = sid_cb
        self.stop_requested = False
        self.cache_sid = {}
        self.cache_bc  = {}
        self.loop_cache = {}
        

    def _rebni_lookup(self, sid, po, asin, inv_no=None):
        rows = self.rebni_p.get((sid, po, asin), [])
        if not rows and inv_no:
            rows = self.rebni_fb.get((sid, po, inv_no), [])
        return rows

    def _inv_lookup(self, sid, po, asin, inv_no=None):
        m = self.inv_p.get((sid, po, asin), [])
        if not m and inv_no:
            m = self.inv_fb.get((sid, po, inv_no), [])
        return m

    def _find_sid(self, po, asin, inv_no):
        rows = self.rebni_s.get((clean(po), clean(asin)), [])
        for r in rows:
            if inv_no in split_comma(r.get('matched_invoice_numbers','')):
                return extract_sid(r['shipment_id'])
        return extract_sid(rows[0]['shipment_id']) if rows else None

    def _resolve_inv_qty(self, inv_no, asin, fallback_qty):
        """
        Return the correct invoice quantity for a given invoice number + ASIN.

        For SCR-suffixed invoices (e.g., 9527185027SCRSCRSCR):
          - Strip the SCR suffix → base_inv = 9527185027
          - Search invoice_search data for invoice_number=base_inv AND asin=asin
          - Return that row's quantity_invoiced

        For normal invoices:
          - Return fallback_qty (the value already stored in the match entry)

        This is necessary because SCR-suffixed invoice entries in the index may
        carry the SCR row's own quantity, not the base invoice's quantity.
        """
        base = strip_scr(inv_no)
        if base == clean(inv_no):
            # No SCR suffix — use the stored value directly
            return fallback_qty

        # Look up the base invoice number in the invoice primary index
        # inv_p is keyed by (sid_frag, po, asin) — we need to search by invoice number
        # Use invoice_search dataframe directly via a scan of inv_p entries
        for (s, p, a), entries in self.inv_p.items():
            if a != clean(asin):
                continue
            for entry in entries:
                if clean(entry.get('mtc_inv', '')) == base:
                    qty = safe_num(entry.get('inv_qty', 0))
                    if qty > 0:
                        return qty

        # Fallback: if base not found, return original
        return fallback_qty

    def detect_cross_po(self, sid, current_po, asin):
        candidates = []
        seen_po = set()
        rec_at_current = 0.0
        current_rows = self.rebni_p.get((sid, current_po, asin), [])
        if current_rows:
            rec_at_current = sum(safe_num(r.get('quantity_unpacked', 0)) for r in current_rows)

        for (s, p, a), rebni_rows in self.rebni_p.items():
            if s != sid or a != asin or p == current_po or p in seen_po:
                continue
            for r in rebni_rows:
                rec = safe_num(r.get('quantity_unpacked', 0))
                if rec <= 0: continue
                seen_po.add(p)
                inv_matches = self.inv_p.get((sid, p, asin), [])
                inv_qty_cross = safe_num(inv_matches[0].get('inv_qty', 0)) if inv_matches else 0.0
                if rec_at_current == 0 and inv_qty_cross == 0:
                    cross_type = "Case 2 — ASIN not invoiced at this PO, but received"
                elif rec > inv_qty_cross and inv_qty_cross > 0:
                    cross_type = "Case 3 — Rec qty > Inv qty (overage in cross PO)"
                else:
                    cross_type = "Case 1 — Rec=0 at current PO, units received here"
                candidates.append({
                    'po': p, 'asin': asin, 'sid': sid,
                    'inv_qty': fmt_qty(inv_qty_cross), 'rec_qty': rec,
                    'cross_type': cross_type, 'date': clean(r.get('received_datetime', '')),
                })
        return candidates

    def _build_level_logic(self, barcode, inv_no, sid, po, asin,
                            inv_qty, rem_pqv, depth, is_claiming,
                            cross_po_indicator_only=False):
        sid_frag = extract_sid(sid)

        # ── REBNI: PRIMARY-ONLY, MULTI-ROW AGGREGATED ─────────────────────────
        # Use rebni_p DIRECTLY. NOT self._rebni_lookup() — its fallback returns
        # rows for different ASINs that share the same invoice number (cross-ASIN contamination).
        # Sum ALL rows for same (SID, PO, ASIN) to get true received quantity.
        rebni_rows  = self.rebni_p.get((sid_frag, clean(po), clean(asin)), [])
        rec_qty     = 0.0
        rebni_avail = 0.0
        ex_adj      = 0.0
        remarks     = ""
        rec_date    = ""

        if rebni_rows:
            # Aggregate multiple warehouse receipt rows for same (SID, PO, ASIN)
            rec_qty     = sum(safe_num(r.get('quantity_unpacked', 0)) for r in rebni_rows)
            rebni_avail = sum(safe_num(r.get('rebni_available', 0))   for r in rebni_rows)
            ex_adj      = sum(safe_num(r.get('quantity_adjusted', 0)) for r in rebni_rows)
            rec_date    = clean(rebni_rows[0].get('received_datetime', ''))
            if rebni_avail > 0:
                lvl     = 'claiming shipment' if is_claiming else 'matching shipment'
                remarks = f"REBNI Available = {int(rebni_avail)} units at {lvl} level — Suggest TSP to utilize"
        else:
            # No REBNI row for this exact (SID, PO, ASIN).
            # depth=0: rec_qty=0, remarks="" — direct-short logic handles it below.
            # depth>0: mark SR.
            if depth > 0:
                remarks = "SR"

        shortage = max(0.0, safe_num(inv_qty) - rec_qty)

        # ── DIRECT SHORT EARLY RETURN ──────────────────────────────────────────
        # Condition: physical shortage >= remaining PQV budget AND no blocking remark.
        # Covers: (a) ASIN not in REBNI (rec_qty=0, shortage=inv_qty),
        #         (b) ASIN in REBNI but shortage >= PQV.
        if shortage >= rem_pqv > 0 and not remarks:
            # Rec Qty: show "0" explicitly when nothing received (not empty string)
            rec_qty_display = "0" if rec_qty == 0.0 else fmt_qty(rec_qty)
            # Mtc Qty: if nothing received, nothing was matched → show "0"
            # Rule: Matched Qty = Received Qty (rec_qty)
            mtc_qty_display = "0" if rec_qty == 0.0 else fmt_qty(rec_qty)

            main_row = {
                'barcode': barcode, 'invoice': inv_no,
                'sid': extract_sid(sid) if sid else '',
                'po': po, 'asin': asin,
                'inv_qty': fmt_qty(inv_qty),
                'rec_qty': rec_qty_display,
                'mtc_qty': mtc_qty_display,
                'mtc_inv': "Short Received",
                'remarks': f"{int(shortage)} units short received directly",
                'date': rec_date, 'depth': depth, 'type': 'dominant'
            }

            result_rows = [main_row]

            # ── SHIPMENT-LEVEL REBNI CHECK ─────────────────────────────────────
            shipment_rebni = self._get_shipment_rebni(sid_frag, clean(po))
            if shipment_rebni > 0:
                result_rows.append({
                    'barcode': '[REBNI-SHP]', 'invoice': inv_no,
                    'sid': sid_frag, 'po': po, 'asin': asin,
                    'inv_qty': '', 'rec_qty': '',
                    'mtc_qty': fmt_qty(shipment_rebni), 'mtc_inv': '',
                    'remarks': (f"Shipment-level REBNI = {int(shipment_rebni)} units available "
                                f"across all ASINs in this shipment — Suggest TSP to utilize"),
                    'date': rec_date, 'depth': depth, 'type': 'rebni_shipment'
                })

            # ── CROSS PO CHECK ────────────────────────────────────────────────────
            # AUTO mode (cross_po_indicator_only=False):
            #   → Full chain investigation via _build_cross_po_rows.
            # MANUAL mode (cross_po_indicator_only=True):
            #   → Indicator rows only; _man_step will show CrossPODialog for user confirmation.
            if cross_po_indicator_only:
                candidates = self.detect_cross_po(sid_frag, clean(po), clean(asin))
                for c in candidates:
                    result_rows.append({
                        'barcode': '[CROSS PO?]', 'invoice': '—',
                        'sid': c['sid'], 'po': c['po'], 'asin': c['asin'],
                        'inv_qty': c.get('inv_qty', ''),
                        'rec_qty': fmt_qty(c['rec_qty']),
                        'mtc_qty': '', 'mtc_inv': '',
                        'remarks': (f"Cross PO candidate detected | {c['cross_type']} "
                                    f"| Rec={fmt_qty(c['rec_qty'])} units "
                                    f"| PENDING USER CONFIRMATION — verify before investigating"),
                        'date': c['date'], 'depth': depth, 'type': 'crosspo'
                    })
            else:
                # Auto mode: detect and immediately investigate the full chain
                result_rows.extend(
                    self._build_cross_po_rows(sid_frag, clean(po), clean(asin), depth))

            return result_rows, [], rec_qty, shortage, 0.0, ex_adj
        # ───────────────────────────────────────────────────────────────────────

        # ── STOP CONDITIONS ────────────────────────────────────────────────────
        if 'REBNI Available' in remarks or remarks == 'SR':
            rows = [self._make_row(barcode, inv_no, sid, po, asin,
                                    inv_qty, rec_qty, "", "", remarks, rec_date, depth)]
            return rows, [], rec_qty, shortage, max(0.0, rem_pqv - shortage), ex_adj

        # ── MATCHING INVESTIGATION ─────────────────────────────────────────────
        # shortage < rem_pqv: physical short alone doesn't cover all PQV.
        # OR shortage == 0: Rec = Inv but PQV still exists — trace matching chain.
        # PRIMARY-ONLY Invoice Search lookup.
        # Use inv_p directly — do NOT use self._inv_lookup() which falls back to
        # inv_fb[(sid, po, inv_no)]. That fallback returns entries for ALL ASINs
        # at that invoice number, contaminating results with wrong mtc_qty values.
        # The primary index inv_p is already ASIN-scoped (keyed by matched_asin),
        # so no additional mtc_asin filter is needed or correct.
        # NOTE: the previous mtc_asin filter was WRONG — it filtered by the invoice's
        # own product ASIN, which for cross-PO cases differs from the current ASIN.
        raw  = self.inv_p.get((sid_frag, clean(po), clean(asin)), [])
        seen = set()
        unique = []
        for m in raw:
            combo = (m['mtc_inv'], m['mtc_po'], m['mtc_asin'])
            if combo not in seen:
                seen.add(combo)
                unique.append(m)
        sorted_m = sorted(unique, key=lambda x: safe_num(x['mtc_qty']), reverse=True)

        main_mtc_inv = ""
        main_mtc_qty = ""

        if sorted_m:
            top = sorted_m[0]
            if top['mtc_inv'] == clean(inv_no):
                # Self Matching: invoice matched itself.
                # Mtc Qty = rec_qty (this ASIN's received qty), NOT top['mtc_qty'] (invoice total).
                # sorted_m MUST be cleared — remaining entries are invoice-system distributions,
                # not investigation branches for this specific ASIN.
                main_mtc_inv = "Self Matching"
                main_mtc_qty = fmt_qty(rec_qty)
                sorted_m     = []
            else:
                # Non-self match: different invoice — legitimate investigation branch.
                main_mtc_inv = top['mtc_inv']
                main_mtc_qty = fmt_qty(top['mtc_qty'])
        else:
            if not remarks and shortage > 0:
                main_mtc_inv = "Short Received"
                main_mtc_qty = fmt_qty(shortage)
                remarks = f"{int(shortage)} units short received directly"

        # ── MAIN ROW ───────────────────────────────────────────────────────────
        rows = [self._make_row(barcode, inv_no, sid, po, asin,
                                inv_qty, rec_qty, main_mtc_qty, main_mtc_inv,
                                remarks, rec_date, depth)]

        # ── SUB-ROWS ───────────────────────────────────────────────────────────
        # For non-self-matching: sorted_m[0] is the main row → write from [1:] as sub-rows.
        # For Self Matching: sorted_m=[] → no sub-rows written.
        sub_start = 1 if (sorted_m and main_mtc_inv not in ("Self Matching", "Short Received")) else 0
        for m in sorted_m[sub_start:]:
            rows.append(self._make_row("", "", "", "", "", "", "",
                                        fmt_qty(m['mtc_qty']), m['mtc_inv'],
                                        "", "", depth, 'subrow'))

        # ── ACTIONABLE MATCHES FOR RECURSION ──────────────────────────────────
        # Resolve inv_qty for SCR-suffixed invoices to get the correct base quantity.
        actionable = []
        for m in sorted_m:
            if m['mtc_inv'] != clean(inv_no):
                resolved_iqty = self._resolve_inv_qty(m['mtc_inv'], m['mtc_asin'], m['inv_qty'])
                actionable.append({**m, 'inv_qty': resolved_iqty})

        new_rem = max(0.0, rem_pqv - min(rem_pqv, shortage))

        # ── SHIPMENT-LEVEL REBNI CHECK (for matching nodes too) ───────────────
        if shortage > 0:
            shipment_rebni = self._get_shipment_rebni(sid_frag, clean(po))
            if shipment_rebni > 0:
                rows.append({
                    'barcode': '[REBNI-SHP]', 'invoice': inv_no,
                    'sid': sid_frag, 'po': po, 'asin': asin,
                    'inv_qty': '', 'rec_qty': '',
                    'mtc_qty': fmt_qty(shipment_rebni), 'mtc_inv': '',
                    'remarks': (f"Shipment-level REBNI = {int(shipment_rebni)} units available "
                                f"across all ASINs in this shipment — Suggest TSP to utilize"),
                    'date': rec_date, 'depth': depth, 'type': 'rebni_shipment'
                })

        # ── CROSS PO CHECK ──────────────────────────────────────────────────────
        # AUTO mode: full chain investigation via _build_cross_po_rows.
        # MANUAL mode (cross_po_indicator_only=True): suppress auto-detection here.
        # In manual mode, Cross PO at matching level is handled by the user clicking
        # the "Cross PO" button in ManualLevelDialog. Auto-detection here would cause
        # circular detection (cross PO's PO → detects original claiming PO as a new
        # Cross PO) and duplicate rows.
        if not cross_po_indicator_only:
            rows.extend(self._build_cross_po_rows(sid_frag, clean(po), clean(asin), depth))

        return rows, actionable, rec_qty, shortage, new_rem, ex_adj

    def _build_cross_po_rows(self, sid, po, asin, depth):
        candidates = self.detect_cross_po(sid, po, asin)
        rows = []
        for c in candidates:
            budget = safe_num(c['rec_qty'])
            rows.append({
                'barcode' : '[CROSS PO]', 'invoice' : '—', 'sid' : c['sid'],
                'po' : c['po'], 'asin' : c['asin'], 'inv_qty' : c.get('inv_qty', ''),
                'rec_qty' : fmt_qty(c['rec_qty']), 'mtc_qty' : '', 'mtc_inv' : '',
                'remarks' : f"Cross PO — {c['cross_type']} | Overage = {fmt_qty(c['rec_qty'])} units — investigating chain",
                'date' : c['date'], 'depth' : depth, 'type' : 'crosspo',
            })
            if budget > 0:
                case_label = c['cross_type'].split("\u2014")[0].strip()
                child_rows, _ = self.run_cross_po_investigation(c, case_label, budget, depth=depth+1)
                rows.extend(child_rows)
        return rows

    def _get_shipment_rebni(self, sid, po):
        """
        Return total rebni_available summed across ALL ASINs for a given SID + PO.
        Used to detect shipment-level REBNI that may offset claiming shortages.
        """
        total = 0.0
        for (s, p, a), rows in self.rebni_p.items():
            if s == sid and p == po:
                for r in rows:
                    total += safe_num(r.get('rebni_available', 0))
        return total

    def _make_row(self, b, i, s, p, a, iq, rq, mq, mi, rem, d, depth, rtype='dominant'):
        return {'barcode': b, 'invoice': i, 'sid': extract_sid(s) if s else '',
                'po': p, 'asin': a, 'inv_qty': fmt_qty(iq), 'rec_qty': fmt_qty(rq),
                'mtc_qty': fmt_qty(mq), 'mtc_inv': mi, 'remarks': rem, 'date': d,
                'depth': depth, 'type': rtype}

    def run_auto(self, barcode, inv_no, sid, po, asin, inv_qty, pqv,
                 depth=0, visited=None, rem_pqv=None, is_claiming=True,
                 branch_budget=None):
        if self.stop_requested: return [], 0.0
        if visited is None: visited = set()
        if rem_pqv is None: rem_pqv = safe_num(pqv)
        if branch_budget is None: branch_budget = rem_pqv
        sid_frag = extract_sid(sid); state = (sid_frag, clean(inv_no), clean(po), clean(asin))
        if state in visited or depth >= self.MAX_DEPTH: return [], 0.0
        visited = visited | {state}
        if state in self.loop_cache and depth > 0:
            cached = self.loop_cache[state]
            found = sum(safe_num(r.get('rec_qty',0)) for r in cached if 'short' in r.get('remarks','').lower() or r.get('mtc_inv','') == 'Short Received')
            return list(cached), found

        rows, actionable, rec_qty, shortage, new_rem, ex_adj = self._build_level_logic(barcode, inv_no, sid, po, asin, inv_qty, rem_pqv, depth, is_claiming)
        total_found = min(branch_budget, max(0.0, shortage))
        rem_budget = branch_budget - total_found
        if rem_budget <= 0 or not actionable or 'REBNI' in rows[0].get('remarks','') or rows[0].get('remarks','') == 'SR':
            if rows and shortage > 0 and 'short received directly' not in rows[0].get('remarks', '').lower():
                c_str = "now continue for remaining pqv" if rem_pqv - total_found > 0 else "no remaining pqv"
                rows[0]['remarks'] = f"Found {int(shortage)} units short as loop started from {int(branch_budget)} matched qty, {c_str}"
            return rows, total_found

        cur_budget = rem_budget
        for match in actionable:
            if self.stop_requested or cur_budget <= 0: break
            n_inv = match['mtc_inv']; n_po = match['mtc_po']; n_asin = match['mtc_asin']
            n_budget = safe_num(match['mtc_qty'])
            # For SCR-suffixed invoices, look up the real invoice qty from the base number.
            # The stored inv_qty may reflect the SCR row's own quantity, not the base invoice.
            n_iqty = self._resolve_inv_qty(n_inv, n_asin, match['inv_qty'])
            n_sid = self.cache_sid.get(n_inv) or self._find_sid(n_po, n_asin, n_inv)
            if not n_sid and self.sid_cb:
                n_sid = self.sid_cb(n_inv, n_po, n_asin)
                if n_sid: self.cache_sid[n_inv] = n_sid
            if not n_sid:
                rows.append(self._make_row("[DICES]", n_inv, "[ENTER SID FROM DICES]", n_po, n_asin, n_iqty, "", "", "", "SID not found — validate in DICES", "", depth+1))
                continue
            child_rows, found_in_b = self.run_auto(self.cache_bc.get(n_inv, "[DICES]"), n_inv, n_sid, n_po, n_asin, n_iqty, pqv, depth+1, visited, rem_pqv - total_found, False, n_budget)
            rows.extend(child_rows); contribution = min(cur_budget, found_in_b); total_found += contribution; cur_budget -= contribution

        if (rows and total_found > 0
                and 'Root' not in rows[0].get('remarks', '')
                and 'REBNI' not in rows[0].get('remarks', '')
                and 'short received directly' not in rows[0].get('remarks', '').lower()):
            c_str = "now continue for remaining pqv" if rem_pqv - total_found > 0 else "no remaining pqv"
            rows[0]['remarks'] = f"Found {int(total_found)} units short as loop started from {int(branch_budget)} matched qty, {c_str}"
        if depth > 0: self.loop_cache[state] = list(rows)
        return rows, total_found

    def build_one_level(self, b, i, s, p, a, iq, rem, depth=0, is_claiming=True):
        # cross_po_indicator_only=True: in MANUAL mode, write [CROSS PO?] indicator rows
        # instead of auto-investigating the chain. _man_step will detect candidates and
        # show CrossPODialog for user confirmation before any investigation begins.
        rows, matches, rq, shortage, new_rem, ex_adj = self._build_level_logic(
            b, i, s, p, a, iq, rem, depth, is_claiming, cross_po_indicator_only=True)
        return rows, [m for m in matches if m['mtc_inv'] != clean(i)], rq, new_rem

    def run_cross_po_investigation(self, candidate, case_type, budget, depth=0, visited=None):
        if visited is None: visited = set()
        c_sid = candidate['sid']; c_po = candidate['po']; c_asin = candidate['asin']; c_iqty = candidate.get('inv_qty', budget)
        raw = self.inv_p.get((extract_sid(c_sid), clean(c_po), clean(c_asin)), []); seen = set(); unique = []
        for m in raw:
            combo = (m['mtc_inv'], m['mtc_po'], m['mtc_asin'])
            if combo not in seen: seen.add(combo); unique.append(m)
        unique.sort(key=lambda x: safe_num(x['mtc_qty']), reverse=True)
        rebni_rows = self.rebni_p.get((extract_sid(c_sid), clean(c_po), clean(c_asin)), [])
        rec_qty  = sum(safe_num(r.get('quantity_unpacked', 0)) for r in rebni_rows) if rebni_rows else 0.0
        r_avail  = sum(safe_num(r.get('rebni_available', 0))   for r in rebni_rows) if rebni_rows else 0.0
        rec_date = clean(rebni_rows[0].get('received_datetime', '')) if rebni_rows else ''
        shortage = max(0.0, safe_num(c_iqty) - rec_qty)
        m_inv = unique[0]['mtc_inv'] if unique else "Short Received"
        m_qty = fmt_qty(unique[0]['mtc_qty']) if unique else ""
        if not unique and shortage > 0: rem = f"Found {int(min(shortage, budget))} units short as loop started from {int(budget)} matched qty, no remaining pqv"
        elif r_avail > 0: rem = f"REBNI Available = {int(r_avail)} units at matching shipment level — Suggest TSP to utilize"
        else: rem = f"Cross PO {case_type} — investigating {int(budget)} overage units"
        rows = [self._make_row('[CROSS PO]', '—', c_sid, c_po, c_asin, fmt_qty(c_iqty), rec_qty, m_qty, m_inv, rem, rec_date, depth)]
        for m in unique[1:]: rows.append(self._make_row("","","","","","","",fmt_qty(m['mtc_qty']), m['mtc_inv'],"","", depth, 'subrow'))
        if r_avail > 0 or not unique: return rows, min(shortage, budget)
        total_found = min(shortage, budget); current_rem = budget - total_found
        for match in unique:
            if current_rem <= 0: break
            n_inv = match['mtc_inv']; n_po = match['mtc_po']; n_asin = match['mtc_asin']
            n_budget = safe_num(match['mtc_qty'])
            n_iqty = self._resolve_inv_qty(n_inv, n_asin, match['inv_qty'])
            state = (c_sid, clean(n_inv), clean(n_po), clean(n_asin))
            if state in visited: continue
            visited = visited | {state}
            n_sid = self.cache_sid.get(n_inv) or self._find_sid(n_po, n_asin, n_inv)
            if not n_sid and self.sid_cb:
                n_sid = self.sid_cb(n_inv, n_po, n_asin)
                if n_sid: self.cache_sid[n_inv] = n_sid
            if not n_sid:
                rows.append(self._make_row("[DICES]", n_inv, "[ENTER SID FROM DICES]", n_po, n_asin, n_iqty, "", "", "", "SID not found — validate in DICES", "", depth+1))
                continue
            child_rows, found = self.run_auto(self.cache_bc.get(n_inv, "[DICES]"), n_inv, n_sid, n_po, n_asin, n_iqty, current_rem, depth+1, visited, current_rem, False, min(n_budget, current_rem))
            rows.extend(child_rows); contribution = min(current_rem, found); total_found += contribution; current_rem -= contribution
        return rows, total_found


# ═══════════════════════════════════════════════════════════
#  EXCEL WRITER
# ═══════════════════════════════════════════════════════════

def write_excel(all_blocks, path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Investigation"
    headers = ["Barcode","Inv no","SID","PO","ASIN","Inv Qty","Rec Qty","Mtc Qty","Mtc Inv","Remarks","Date"]
    H_FILL, DOM_F, SUB_F, ROOT_F, DICES_F, SR_F, INVLD_F, REBNI_F, CROSS_F, MIS_F = [PatternFill("solid", fgColor=c) for c in ["203864","E2EFDA","EBF3FB","FFE0E0","FFF2CC","FFD7D7","FFD0D0","D0F0FF","FFF0C0","D0E8FF"]]
    H_FONT, N_FONT, ROOT_FT, SR_FT, INVLD_FT, REBNI_FT, CROSS_FT = Font(color="FFFFFF", bold=True, name="Calibri", size=10), Font(name="Calibri", size=10), Font(bold=True, color="9C0006", name="Calibri", size=10), Font(bold=True, color="CC0000", name="Calibri", size=10), Font(bold=True, color="880000", name="Calibri", size=10, italic=True), Font(bold=True, color="005580", name="Calibri", size=10), Font(bold=True, color="7a5c00", name="Calibri", size=10)
    BDR = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    KM  = {'Barcode':'barcode','Inv no':'invoice','SID':'sid','PO':'po','ASIN':'asin','Inv Qty':'inv_qty','Rec Qty':'rec_qty','Mtc Qty':'mtc_qty','Mtc Inv':'mtc_inv','Remarks':'remarks','Date':'date'}
    curr = 1
    for block in all_blocks:
        if not block: continue
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=curr, column=c, value=h); cell.fill, cell.font, cell.border = H_FILL, H_FONT, BDR
        curr += 1
        for rd in block:
            rem, mti, rtyp, dep = str(rd.get('remarks','')), str(rd.get('mtc_inv','')), rd.get('type','dominant'), rd.get('depth',0)
            if 'invalid invoice' in rem.lower():
                fill, fnt = INVLD_F, INVLD_FT
            elif 'REBNI Available' in rem or 'Shipment-level REBNI' in rem or rtyp == 'rebni_shipment':
                fill, fnt = REBNI_F, REBNI_FT
            elif 'Cross PO' in rem or rtyp == 'crosspo':
                fill, fnt = CROSS_F, CROSS_FT
            elif ('short received directly' in rem.lower()
                  or 'Root cause' in rem
                  or ('Found' in rem and 'short' in rem.lower())):
                fill, fnt = ROOT_F, ROOT_FT
            elif rem == 'SR':
                fill, fnt = SR_F, SR_FT
            elif rtyp == 'subrow':
                fill, fnt = SUB_F, N_FONT
            elif dep > 0:
                fill, fnt = DOM_F, N_FONT
            else:
                fill, fnt = None, N_FONT
            for c, h in enumerate(headers, 1):
                val = rd.get(KM[h], ""); cell = ws.cell(row=curr, column=c, value=val if val not in (None, '') else None); cell.border = BDR; cell.font = fnt
                if fill: cell.fill = fill
            curr += 1
        curr += 1
    for i, w in enumerate([18,22,18,12,14,9,9,9,26,42,22], 1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"; wb.save(path)


# ═══════════════════════════════════════════════════════════
#  MAIN GUI
# ═══════════════════════════════════════════════════════════

class MFIToolApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("MFI Investigation Tool  v5.0.3  |  ROW IB")
        try: self.root.state('zoomed')
        except: self.root.attributes('-zoomed', True)
        self.root.minsize(900, 620); self.root.configure(bg="#0f0f1a")
        self.claims_path, self.rebni_path, self.inv_path, self.ticket_id, self.mode_var = tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar(value="auto")
        self.all_blocks, self.preview = [], None; self._build_ui()

    def _build_ui(self):
        t = tk.Frame(self.root, bg="#16213e", height=62); t.pack(fill="x")
        tk.Label(t, text="  MFI Investigation Tool", fg="#e94560", bg="#16213e", font=("Segoe UI",20,"bold")).pack(side="left", padx=16, pady=12)
        tk.Label(t, text="Developed by Mukesh", fg="#4a9eff", bg="#16213e", font=("Segoe UI",10,"italic")).pack(side="right", padx=6)
        tk.Label(t, text="v5.0.3  |  ROW IB", fg="#8888aa", bg="#16213e", font=("Segoe UI",10)).pack(side="right", padx=16)
        leg = tk.Frame(self.root, bg="#1a1a2e", height=30); leg.pack(fill="x")
        for txt, fg, bg in [("Claiming","white","#0f0f1a"), ("Dominant","black","#E2EFDA"), ("Sub-rows","black","#EBF3FB"), ("Root/Short","#9C0006","#FFE0E0"), ("DICES","black","#FFF2CC"), ("SR","#CC0000","#FFD7D7"), ("Invalid inv","#333","#FFD0D0"), ("REBNI","#333","#D0F0FF"), ("Cross PO","#7a5c00","#FFF0C0"), ("Mismatch","#333","#D0E8FF")]:
            tk.Label(leg, text=f"  {txt}  ", fg=fg, bg=bg, font=("Segoe UI",8,"bold"), padx=8).pack(side="left", padx=3, pady=3)
        body = tk.Frame(self.root, bg="#0d0d1a", padx=24, pady=12); body.pack(fill="both", expand=True)
        inp = tk.LabelFrame(body, text="  Input Files  (Excel .xlsx or CSV .csv supported)  ", fg="#4a9eff", bg="#0d0d1a", font=("Segoe UI",10,"bold"), padx=12, pady=8); inp.pack(fill="x", pady=6)
        self._f_row(inp, "Claims Sheet:", self.claims_path, 0); self._f_row(inp, "REBNI Result:", self.rebni_path, 1); self._f_row(inp, "Invoice Search:", self.inv_path, 2)
        tf = tk.Frame(body, bg="#0f0f1a"); tf.pack(anchor="w", pady=4)
        tk.Label(tf, text="Ticket ID:", fg="white", bg="#0f0f1a", font=("Segoe UI",11)).pack(side="left")
        tk.Entry(tf, textvariable=self.ticket_id, width=28, font=("Segoe UI",11), bg="#1e1e3a", fg="white", insertbackground="white", relief="flat").pack(side="left", padx=8)
        m = tk.LabelFrame(body, text="Investigation Mode", fg="white", bg="#0f0f1a", padx=10, pady=5); m.pack(fill="x", pady=8)
        tk.Radiobutton(m, text="AUTO  —  Automatic. SID popup when not found in REBNI.", variable=self.mode_var, value="auto", fg="white", bg="#0f0f1a", selectcolor="#16213e", font=("Segoe UI",10)).pack(anchor="w", padx=10)
        tk.Radiobutton(m, text="MANUAL  —  One level at a time. Live preview. Parallel interaction enabled.", variable=self.mode_var, value="manual", fg="white", bg="#0f0f1a", selectcolor="#16213e", font=("Segoe UI",10)).pack(anchor="w", padx=10)
        self.status = tk.Label(body, text="Ready", fg="#4a9eff", bg="#0f0f1a", font=("Segoe UI",10)); self.status.pack(pady=(10,0))
        self.pb = ttk.Progressbar(body, mode='determinate'); self.pb.pack(fill="x", pady=4)
        bf = tk.Frame(body, bg="#0f0f1a"); bf.pack(pady=10)
        self.run_btn = tk.Button(bf, text="▶  RUN INVESTIGATION", bg="#e94560", fg="white", font=("Segoe UI",15,"bold"), padx=36, pady=14, relief="flat", cursor="hand2", command=self.start_run); self.run_btn.pack(side="left", padx=10)
        
        self.stop_inv_btn = tk.Button(bf, text="⏸  STOP INVESTIGATION",
                                      bg="#4a2020", fg="white",
                                      font=("Segoe UI",11,"bold"),
                                      padx=16, pady=14, relief="flat",
                                      state="disabled", cursor="hand2",
                                      command=self.request_stop_investigation)
        self.stop_inv_btn.pack(side="left", padx=6)

        self.stop_sess_btn = tk.Button(bf, text="⏹  STOP SESSION",
                                       bg="#3a0000", fg="white",
                                       font=("Segoe UI",11,"bold"),
                                       padx=16, pady=14, relief="flat",
                                       state="disabled", cursor="hand2",
                                       command=self.request_stop_session)
        self.stop_sess_btn.pack(side="left", padx=6)

        self.save_btn = tk.Button(bf, text="💾  SAVE OUTPUT", bg="#2d6a4f", fg="white", font=("Segoe UI",13,"bold"), padx=28, pady=14, relief="flat", state="disabled", cursor="hand2", command=self.save_output); self.save_btn.pack(side="left", padx=10)

    def _f_row(self, p, l, v, r):
        tk.Label(p, text=l, fg="#cccccc", bg="#131320", width=18, anchor="w", font=("Segoe UI",10)).grid(row=r, column=0, sticky="w", pady=3)
        tk.Entry(p, textvariable=v, width=62, font=("Segoe UI",10), bg="#1e1e3a", fg="white", insertbackground="white", relief="flat").grid(row=r, column=1, padx=6)
        tk.Button(p, text="Browse", command=lambda: v.set(filedialog.askopenfilename(filetypes=[("Excel/CSV","*.xlsx *.xls *.csv")])), bg="#2d2d5e", fg="white", relief="flat", cursor="hand2", padx=8).grid(row=r, column=2)

    def _set_status(self, msg, pct=None):
        self.status.config(text=msg)
        if pct is not None: self.pb['value'] = pct
        self.root.update_idletasks()

    def start_run(self):
        if not all([self.claims_path.get(), self.rebni_path.get(), self.inv_path.get()]):
            messagebox.showerror("Error", "Please select all 3 input files."); return
        self.run_btn.config(state="disabled"); self.save_btn.config(state="disabled")
        self.stop_inv_btn.config(state="normal")
        self.stop_sess_btn.config(state="normal")
        self.all_blocks = []
        if self.mode_var.get() == "manual":
            if not self.preview or not self.preview.winfo_exists(): self.preview = PreviewPanel(self.root)
            else: self.preview.clear_all()
        threading.Thread(target=self._process, daemon=True).start()

    def request_stop_investigation(self):
        """
        STOP INVESTIGATION: Pause the current ASIN investigation.
        Stay on the same page. Engine marks stop but does not advance.
        User can resume or save what is completed so far.
        """
        if hasattr(self, 'engine'):
            self.engine.stop_requested = True
        self._set_status("Investigation paused — current results preserved. Save or resume.")
        # Do NOT call _finish() — stay on same page
        self.root.after(0, lambda: self.save_btn.config(state="normal"))

    def request_stop_session(self):
        """
        STOP SESSION: End the current session entirely.
        Saves current results and closes/resets for next claim.
        """
        if hasattr(self, 'engine'):
            self.engine.stop_requested = True
        self._set_status("Session ended — saving current results.")
        self._finish()

    def request_stop(self):
        # Kept for backward compatibility if needed, but redirects to session stop
        self.request_stop_session()

    def _process(self):
        try:
            self._set_status("Loading Claims file…", 5); df_c = load_claims(self.claims_path.get()); mc, corrections = detect_claim_cols(df_c)
            missing = [f for f in COLUMN_ALIASES if f not in mc]
            if corrections or missing:
                confirmed, done = [None], threading.Event()
                self.root.after(0, lambda: HeaderCorrectionDialog(self.root, corrections, mc, list(df_c.columns), lambda res: (confirmed.__setitem__(0, res['mapping']) if res['action']=='proceed' else None, done.set())))
                done.wait()
                if confirmed[0] is None: self.root.after(0, lambda: (self.run_btn.config(state="normal"), self.stop_btn.config(state="disabled"))); return
                mc = confirmed[0]
            self._set_status("Loading REBNI…", 12); rp, rs, rfb = build_rebni_index(load_rebni(self.rebni_path.get()))
            self._set_status("Loading Invoice Search…", 30); ip, ifb = build_invoice_index(load_invoice_search(self.inv_path.get()))
            self.engine = InvestigationEngine(rp, rs, rfb, ip, ifb, self._req_sid); tot = len(df_c)
            if self.mode_var.get() == "auto":
                for i, (_, r) in enumerate(df_c.iterrows()):
                    if self.engine.stop_requested: break
                    self._set_status(f"Auto: {i+1}/{tot}  ASIN: {clean(r.get(mc.get('ASIN',''),''))}", 60 + int((i/max(tot,1))*35))
                    rows, _ = self.engine.run_auto(clean(r.get(mc.get('Barcode',''),'')), clean(r.get(mc.get('Invoice',''),'')), extract_sid(clean(r.get(mc.get('SID',''), ''))), clean(r.get(mc.get('PO',''),'')), clean(r.get(mc.get('ASIN',''),'')), safe_num(r.get(mc.get('InvQty',''),0)), safe_num(r.get(mc.get('PQV',''),0)))
                    self.all_blocks.append(rows)
                self._finish()
            else: self.manual_q, self.map_cols = df_c.to_dict('records'), mc; self._next_man()
        except Exception as e:
            import traceback; tb = traceback.format_exc(); self.root.after(0, lambda: messagebox.showerror("Error", f"{e}\n\n{tb}")); self._finish()

    def _req_sid(self, inv, po, asin):
        if inv in self.engine.cache_sid: return self.engine.cache_sid[inv]
        res, done = [None], threading.Event()
        self.root.after(0, lambda: SIDRequestDialog(self.root, inv, po, asin, lambda s: (self.engine.cache_sid.__setitem__(inv, s) if s else None, res.__setitem__(0, s), done.set())))
        done.wait(); return res[0]

    def _next_man(self):
        # Investigate any pending Cross PO before moving to next ASIN
        if hasattr(self, '_pending_cross_po') and self._pending_cross_po:
            pending = self._pending_cross_po.pop(0)
            def investigate_pending():
                cross_rows, found = self.engine.run_cross_po_investigation(
                    pending['candidate'], pending['case'], pending['budget'],
                    depth=self.curr_m['depth'] + 1)
                for r in cross_rows:
                    self.curr_m['block'].append(r)
                    if self.preview and self.preview.winfo_exists():
                        self.root.after(0, lambda row=r: self.preview.add_row(row))
                # Important: _next_man will recursively call itself until _pending_cross_po is empty
                self._next_man()
            threading.Thread(target=investigate_pending, daemon=True).start()
            return

        if not self.manual_q or self.engine.stop_requested: self._finish(); return
        r, mc = self.manual_q.pop(0), self.map_cols
        self.curr_m = {'b': clean(r.get(mc.get('Barcode',''), '')), 'i': clean(r.get(mc.get('Invoice',''), '')), 's': extract_sid(clean(r.get(mc.get('SID',''), ''))), 'p': clean(r.get(mc.get('PO',''), '')), 'a': clean(r.get(mc.get('ASIN',''), '')), 'iq': safe_num(r.get(mc.get('InvQty',''), 0)), 'pqv': safe_num(r.get(mc.get('PQV',''), 0)), 'rem': safe_num(r.get(mc.get('PQV',''), 0)), 'budget': safe_num(r.get(mc.get('PQV',''), 0)), 'depth': 0, 'block': [], 'processed': set(), 'rendered': False}
        self.preview.add_header_row(self.curr_m['a']); threading.Thread(target=self._man_step, daemon=True).start()

    def _man_step(self):
        if self.engine.stop_requested: self._finish(); return
        m = self.curr_m
        rows, matches, rq, n_rem = self.engine.build_one_level(
            m['b'], m['i'], m['s'], m['p'], m['a'], m['iq'], m['rem'], m['depth'], m['depth']==0)
        if not m['rendered']:
            m['block'].extend(rows)
            [self.preview.add_row(r) for r in rows]
            m['rendered'] = True
        m['rem'] = n_rem
        matches = [x for x in matches if x['mtc_inv'] not in m['processed']]
        rem_str = rows[0].get('remarks', '') if rows else ''

        should_stop = (not matches or
                       any(x in rem_str for x in
                           ["Root cause", "REBNI", "SR", "short received directly"]))

        if should_stop:
            # ── CROSS PO CHECK before finalizing ──────────────────────────────
            # Only check once per ASIN (guard with 'cross_po_checked' flag).
            if not m.get('cross_po_checked'):
                m['cross_po_checked'] = True
                cross_cands = self.engine.detect_cross_po(m['s'], m['p'], m['a'])
                if cross_cands:
                    # Set guard flag: prevents ANY subsequent _man_step call from
                    # finalizing the block while the CrossPODialog is still open.
                    # Cleared by _handle_cross_po_and_finish when user responds.
                    m['_awaiting_cross_po'] = True
                    self.root.after(0, lambda cands=cross_cands: CrossPODialog(
                        self.root, cands, m['i'], m['s'],
                        lambda r: self._handle_cross_po_and_finish(r)))
                    return  # Wait for user response

            # ── Guard: if CrossPODialog is still open, do NOT finalize ─────────
            if m.get('_awaiting_cross_po'):
                return  # Dialog pending — block must not be finalized yet

            # ── Check siblings_stack for unfinished sibling branches ──────────────
            # siblings_stack is a per-depth stack. Each entry = siblings waiting at
            # that parent depth. Pop from the stack, restore parent context, show dialog.
            stack = m.get('siblings_stack', [])
            if stack:
                ctx = stack.pop()           # restore parent-level context
                siblings = ctx['siblings']  # siblings that were waiting at this depth
                # Filter out any already-processed siblings
                siblings = [x for x in siblings if x['mtc_inv'] not in ctx.get('processed', set())]
                if siblings:
                    # Restore parent-level context so _show_dlg runs at correct depth
                    m.update({
                        'depth'           : ctx['depth'],
                        'rem'             : ctx['rem'],
                        'budget'          : ctx['budget'],
                        's'               : ctx['s'],
                        'p'               : ctx['p'],
                        'a'               : ctx['a'],
                        'processed'       : ctx.get('processed', set()),
                        'cross_po_checked': ctx.get('cross_po_checked', False),
                        'rendered'        : True,   # already rendered parent row
                        'siblings_stack'  : stack,
                        'pending_siblings': [],
                    })
                    self.root.after(0, lambda sibs=siblings: self._show_dlg(sibs))
                    return
                elif stack:
                    # All siblings at this level done, check higher levels
                    m['siblings_stack'] = stack
                    # Re-enter finalization with updated stack
                    self.root.after(0, self._man_step)
                    return

            # ── Stack empty → finalize block ─────────────────────────────────────
            self.all_blocks.append(m['block'])
            self._next_man()
            return

        self.root.after(0, lambda: self._show_dlg(matches))

    def _show_dlg(self, matches):
        m, first = self.curr_m, matches[0]
        if first['mtc_inv'] in self.engine.cache_sid: self._handle_res({'action':'valid','chosen_match':first,'sid':self.engine.cache_sid[first['mtc_inv']],'barcode':self.engine.cache_bc.get(first['mtc_inv'],"[DICES]")}, matches); return
        ManualLevelDialog(self.root, matches, m['rem'], m['budget'], lambda res: self._handle_res(res, matches))

    def _handle_res(self, res, matches):
        if not res or res['action'] == 'stop': self.all_blocks.append(self.curr_m['block']); self._next_man(); return
        match = res.get('chosen_match')
        if match: self.curr_m['processed'].add(match['mtc_inv'])
        if res['action'] == 'invalid':
            excl = res['invalid_qty']; row = {'barcode':'[INVALID]','invoice':match['mtc_inv'],'sid':'—','po':match['mtc_po'],'asin':match['mtc_asin'],'inv_qty':fmt_qty(match['inv_qty']),'rec_qty':'','mtc_qty':'','mtc_inv':'','remarks':f"{int(excl)} units matched to invalid invoice {match['mtc_inv']} — excluded from PQV",'date':'','depth':self.curr_m['depth'],'type':'subrow'}
            self.curr_m['block'].append(row); self.preview.add_row(row); self.curr_m['rem'] = max(0, self.curr_m['rem'] - excl)
            if self.curr_m['rem'] <= 0: self.all_blocks.append(self.curr_m['block']); self._next_man()
            else: rem_m = [x for x in matches if x['mtc_inv'] != match['mtc_inv']]; (self.root.after(0, lambda: self._show_dlg(rem_m)) if rem_m else (self.all_blocks.append(self.curr_m['block']), self._next_man()))
        elif res['action'] == 'cross_po':
            cands = self.engine.detect_cross_po(self.curr_m['s'], self.curr_m['p'], self.curr_m['a'])
            if cands: CrossPODialog(self.root, cands, self.curr_m['i'], self.curr_m['s'], lambda r: self._handle_cross_po(r))
            else: messagebox.showinfo("No Cross PO", "No Cross PO candidates found."); self.root.after(0, lambda: self._show_dlg(matches))
        elif res['action'] == 'mismatch':
            data = res['mismatch_data']; row = {'barcode':'[MISMATCH]','invoice':'','sid':data.get('sid',''),'po':data.get('po',''),'asin':data.get('asin',''),'inv_qty':data.get('inv_qty',''),'rec_qty':data.get('ovg_qty',''),'mtc_qty':'','mtc_inv':'','remarks':f"Mismatch/Overage: ASIN {data.get('asin','')} Inv={data.get('inv_qty','')} Rec={data.get('ovg_qty','')}",'date':'','depth':self.curr_m['depth'],'type':'subrow'}
            self.curr_m['block'].append(row); self.preview.add_row(row); rem_m = [x for x in matches if x != res.get('chosen_match')]; (self.root.after(0, lambda: self._show_dlg(rem_m)) if rem_m else (self.all_blocks.append(self.curr_m['block']), self._next_man()))
        else:
            self.engine.cache_sid[match['mtc_inv']] = res['sid']
            self.engine.cache_bc[match['mtc_inv']]  = res['barcode']
            # Build the siblings_stack — a per-depth stack of unfinished sibling branches.
            # Each entry captures the full context (depth, rem, budget, siblings list) so
            # that after a branch chain finishes, we can pop back up and resume the correct
            # siblings at the correct parent depth.
            remaining_siblings = [x for x in matches
                                   if x['mtc_inv'] != match['mtc_inv']
                                   and x['mtc_inv'] not in self.curr_m.get('processed', set())]
            parent_stack = list(self.curr_m.get('siblings_stack', []))
            if remaining_siblings:
                parent_stack.append({
                    'siblings' : remaining_siblings,
                    'depth'    : self.curr_m['depth'],
                    'rem'      : self.curr_m['rem'],
                    'budget'   : self.curr_m['budget'],
                    's'        : self.curr_m['s'],
                    'p'        : self.curr_m['p'],
                    'a'        : self.curr_m['a'],
                    'processed': set(self.curr_m.get('processed', set())),
                    'cross_po_checked': self.curr_m.get('cross_po_checked', False),
                })
            branch_budget = safe_num(match['mtc_qty'])
            self.curr_m.update({
                'b'              : res['barcode'],
                'i'              : match['mtc_inv'],
                's'              : res['sid'],
                'p'              : match['mtc_po'],
                'a'              : match['mtc_asin'],
                'iq'             : match['inv_qty'],
                'rem'            : branch_budget,   # branch budget = mtc_qty of this match
                'budget'         : branch_budget,
                'depth'          : self.curr_m['depth'] + 1,
                'rendered'       : False,
                'processed'      : set(),
                'siblings_stack' : parent_stack,
                'pending_siblings': [],  # clear legacy field
            })
            threading.Thread(target=self._man_step, daemon=True).start()

    def _handle_cross_po(self, res):
        if res['action'] == 'skip':
            # User skipped Cross PO — resume normal matching investigation.
            # Without this, investigation freezes after skip (no dialog, no continuation).
            threading.Thread(target=self._man_step, daemon=True).start()
            return
        # Store candidate for later investigation — user may continue normal flow first
        if not hasattr(self, '_pending_cross_po'):
            self._pending_cross_po = []
        self._pending_cross_po.append({
            'candidate': res['candidate'],
            'case': res.get('case', 'Case 1'),
            'budget': safe_num(res['candidate']['rec_qty'])
        })
        self._set_status(
            f"Cross PO stored ({res['candidate']['po']}) — continuing normal investigation. "
            f"Cross PO will be investigated after current ASIN.", None)
        # Continue normal manual investigation step
        threading.Thread(target=self._man_step, daemon=True).start()

    def _handle_cross_po_and_finish(self, res):
        """
        Handles CrossPODialog result for DIRECT-SHORT claiming ASINs.

        SKIP: Finalize the current ASIN block and advance to next ASIN.

        CONFIRMED: Start MANUAL investigation of the Cross PO chain by
        updating curr_m with the Cross PO candidate's details and calling
        _man_step. This is the same manual flow used for normal matching:
          - _man_step builds the Cross PO level row (REBNI data + matched invoices)
          - ManualLevelDialog (IBC/PBC popup) appears for EACH matched invoice
          - User validates each invoice (barcode, SID, IBC/PBC)
          - Investigation recurses through all sub-invoices manually
          - When total shortage ≥ budget, _man_step finalizes block + calls _next_man()

        KEY: curr_m['block'] is a Python list reference — all rows added by _man_step
        during the Cross PO chain automatically appear in the claiming ASIN's block.
        We do NOT call self.all_blocks.append() here — _man_step does it when done.
        """
        if res['action'] == 'skip':
            # User skipped — finalize block as-is, no Cross PO investigation
            self.all_blocks.append(self.curr_m['block'])
            self._next_man()
            return

        # User confirmed a Cross PO candidate
        candidate = res['candidate']
        budget    = safe_num(candidate['rec_qty'])

        self._set_status(
            f"Cross PO confirmed ({candidate['po']}) — starting manual investigation "
            f"of {int(budget)} units…", None)

        # Clear the guard flag — the user has responded to the dialog.
        # Without this, _man_step would keep returning early thinking dialog is still open.
        self.curr_m.pop('_awaiting_cross_po', None)

        # Transition curr_m to investigate the Cross PO chain manually.
        #
        # After this update, _man_step will call build_one_level with these details,
        # which calls _build_level_logic to:
        #   1. Look up REBNI for (SID, PO, ASIN) → get rec_qty for cross PO
        #   2. Look up Invoice Search for (SID, PO, ASIN) → get matched invoices
        #   3. Build the Cross PO level row showing REBNI data + first matched invoice
        #   4. Return matches for ManualLevelDialog to present to the user
        #
        # IMPORTANT: curr_m['block'] is NOT replaced — it remains the same Python list.
        # All cross PO investigation rows appended by _man_step automatically become
        # part of the claiming ASIN's output block.
        self.curr_m.update({
            'b'              : '',        # Cross PO: no barcode yet — user provides via ManualLevelDialog
            'i'              : '',        # Cross PO: no invoice number — lookup by SID+PO+ASIN (primary)
            's'              : candidate['sid'],
            'p'              : candidate['po'],
            'a'              : candidate['asin'],
            'iq'             : 0,         # MUST be 0. Using candidate's inv_qty causes:
                                          #   shortage = max(0, inv_qty - rec_qty) >= budget
                                          #   → direct-short fires → matching investigation skipped!
                                          # With iq=0: shortage=0 < budget → correct matching flow.
            'rem'            : budget,    # Budget = overage units (rec_qty at cross PO)
            'budget'         : budget,
            'depth'          : self.curr_m['depth'] + 1,
            'rendered'       : False,     # Let _man_step build the cross PO level row
            'processed'      : set(),
            'cross_po_checked': True,    # Prevent recursive Cross PO dialogs in this chain
        })
        # NOTE: 'block' is intentionally NOT replaced — same list reference.
        # All rows added by _man_step during cross PO investigation automatically
        # appear in the claiming ASIN's output block.

        # Start manual investigation. _man_step builds the cross PO level row,
        # shows ManualLevelDialog (IBC/PBC) for each matched invoice, and recurses.
        # When budget is exhausted, _man_step finalizes: all_blocks.append + _next_man().
        threading.Thread(target=self._man_step, daemon=True).start()

    def _finish(self):
        msg = "Investigation complete!" if not (hasattr(self,'engine') and self.engine.stop_requested) else "Investigation stopped by user."
        self._set_status("Complete. Click SAVE.", 100)
        self.root.after(0, lambda: (
            self.save_btn.config(state="normal"),
            self.run_btn.config(state="normal"),
            self.stop_inv_btn.config(state="disabled"),
            self.stop_sess_btn.config(state="disabled"),
            messagebox.showinfo("Done", msg)
        ))

    def save_output(self):
        t, ts = self.ticket_id.get().strip().replace(' ','_'), datetime.now().strftime('%Y%m%d_%H%M%S')
        out = f"MFI_{t}_{ts}.xlsx" if t else f"MFI_Investigation_{ts}.xlsx"; p = os.path.join(os.path.dirname(self.claims_path.get()) or os.getcwd(), out)
        try:
            blocks = self.all_blocks
            if self.mode_var.get() == "manual" and self.preview and self.preview.winfo_exists():
                all_rows, fb, cur = self.tree.get_all_rows() if hasattr(self,'tree') else self.preview.get_all_rows(), [], []
                for r in all_rows: (fb.append(cur) if r.get('depth') == 0 and cur else None, (cur := []) if r.get('depth') == 0 and cur else None, cur.append(r))
                if cur: fb.append(cur)
                blocks = fb
            write_excel(blocks, p); messagebox.showinfo("Saved", f"Saved to:\n{p}")
        except Exception as e: messagebox.showerror("Save Error", str(e))

    def run(self): self.root.mainloop()

if __name__ == '__main__': MFIToolApp().run()
